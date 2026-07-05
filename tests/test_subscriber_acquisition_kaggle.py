from __future__ import annotations

import base64
import importlib.util
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
import pytest


def load_runtime():
    path = Path("kaggle/SubscriberAcquisitionDiscovery/subscriber_acquisition_discovery.py")
    spec = importlib.util.spec_from_file_location("acq_discovery_runtime", path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


def load_retrieval():
    path = Path("kaggle/SubscriberAcquisitionDiscovery/comment_semantic_retrieval.py")
    spec = importlib.util.spec_from_file_location("comment_semantic_retrieval", path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


def _auth_bundle(session: str, **extra):
    payload = {"session": session, **extra}
    raw = json.dumps(payload).encode("utf-8")
    return base64.urlsafe_b64encode(raw).decode("ascii")


def test_extract_candidate_surfaces_from_public_links():
    runtime = load_runtime()
    surfaces = runtime.extract_candidate_surfaces("Вот t.me/some_kgd_chat, https://vk.com/wall-12345_9 и https://vk.com/wall12345_9")
    by_external = {s["external_id"]: s for s in surfaces}
    assert by_external["tg:some_kgd_chat"]["source"] == "discovered"
    assert by_external["vk:club12345"]["platform"] == "vk"
    assert by_external["vk:club12345"]["url"] == "https://vk.com/club12345"
    assert by_external["vk:id12345"]["surface_type"] == "profile"


def test_extract_candidate_surfaces_skips_telegram_bots_and_service_links():
    runtime = load_runtime()

    surfaces = runtime.extract_candidate_surfaces(
        "https://t.me/somehelperbot https://t.me/addstickers/Foo "
        "https://t.me/share/url?url=x https://t.me/real_kgd_chat"
    )

    assert [surface["external_id"] for surface in surfaces] == ["tg:real_kgd_chat"]
    assert runtime._is_tg_discovery_bot_or_service_handle("weatherbot") is True
    assert runtime._is_tg_discovery_bot_or_service_handle("real_kgd_chat") is False


def test_vk_surface_extractor_skips_non_wall_links_and_keeps_profiles():
    runtime = load_runtime()

    surfaces = runtime.extract_candidate_surfaces(
        "https://vk.com/album-123_456 https://vk.com/app7070938_-1 https://vk.com/market-1 "
        "https://vk.com/away.php https://vk.com/id6786438 https://vk.com/club123"
    )

    assert [surface["external_id"] for surface in surfaces] == ["vk:id6786438", "vk:club123"]
    assert surfaces[0]["surface_type"] == "profile"
    assert runtime._is_vk_scan_domain_candidate("album-123_456") is False
    assert runtime._is_vk_scan_domain_candidate("club123") is True
    assert runtime._is_vk_scan_domain_candidate("id6786438") is True




def test_out_of_region_telegram_surface_is_rejected_not_queued():
    runtime = load_runtime()

    surfaces = runtime.extract_candidate_surfaces("Смотрите ещё https://t.me/visitNavahrudak")

    assert len(surfaces) == 1
    surface = surfaces[0]
    assert surface["external_id"] == "tg:visitNavahrudak"
    assert surface["status"] == "rejected_out_of_region"
    assert surface["topic_cluster"] == "out_of_region"
    assert runtime._is_surface_scan_candidate(surface) is False


def test_rejected_no_comments_surface_is_not_scan_candidate():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://t.me/no_comments_channel", platform="tg")
    surface["status"] = "rejected_no_comments"

    assert runtime._is_surface_scan_candidate(surface) is False


def test_telegram_channel_resolution_status_helpers():
    runtime = load_runtime()
    channel = runtime._seed_surface("https://t.me/some_channel", platform="tg")
    linked = runtime._seed_surface("https://t.me/some_channel_chat", platform="tg")
    linked.update({"surface_type": "linked_discussion", "status": "candidate"})

    assert channel["status"] == "needs_comment_resolve"

    resolved = runtime._mark_tg_channel_resolved_with_linked_discussion(channel, linked_surface=linked)
    assert resolved["status"] == "resolved_has_linked_discussion"
    assert resolved["risk"]["reply_policy"] == "use_linked_discussion"
    assert resolved["risk"]["linked_discussion_url"] == "https://t.me/some_channel_chat"
    assert runtime._is_surface_scan_candidate(resolved) is False

    rejected = runtime._mark_tg_channel_rejected_no_comments(channel)
    assert rejected["status"] == "rejected_no_comments"
    assert rejected["risk"]["reply_policy"] == "no_reply_surface"




def test_trip_recommendation_requirement_is_acquisition_topic():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://t.me/example", platform="tg")

    opp = runtime.build_opportunity_from_message(
        surface,
        SimpleNamespace(id=31, message="Куда съездить на один день из Калининграда на электричке в выходные?"),
        default_target_url="https://t.me/kenigevents",
    )

    assert opp is not None
    assert opp["matched_intent"] == "trip_route_recommendation_context"
    assert opp["topic_cluster"] == "trip_route_recommendation"
    assert opp["link_target"]["kind"] == "other"
    assert opp["link_target"]["url"] is None
    assert "Конкретный маршрут" in opp["link_target"]["label"]
    assert opp["fallback_link_target"]["url"] is None


def test_trip_recommendation_prefilter_covers_looser_route_questions():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://t.me/example", platform="tg")

    examples = [
        "Что посмотреть за день в области с детьми?",
        "Куда поехать на выходных из Калининграда?",
        "Посоветуйте маршрут по области: замки и побережье",
    ]

    for index, text in enumerate(examples, start=40):
        opp = runtime.build_opportunity_from_message(
            surface,
            SimpleNamespace(id=index, message=text),
            default_target_url="https://t.me/kenigevents",
        )
        assert opp is not None, text
        assert opp["topic_cluster"] == "trip_route_recommendation"


def test_acquisition_intent_covers_site_filters_and_partnership(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_STATIC_SITE_BASE_URL", "https://kenigevents.ru")
    surface = runtime._seed_surface("https://t.me/example", platform="tg")

    partner = runtime.build_opportunity_from_message(
        surface,
        SimpleNamespace(id=21, message="Куда прислать афишу и как добавить мероприятие?"),
        default_target_url="https://t.me/kenigevents",
    )
    badge = runtime.build_opportunity_from_message(
        surface,
        SimpleNamespace(id=22, message="Есть поиск мероприятий по Пушкинской карте и для детей?"),
        default_target_url="https://t.me/kenigevents",
    )
    exhibitions = runtime.build_opportunity_from_message(
        surface,
        SimpleNamespace(id=23, message="Где посмотреть все выставки в Калининграде?"),
        default_target_url="https://t.me/kenigevents",
    )

    assert partner["topic_cluster"] == "organizer_partnership"
    assert partner["link_target"]["url"] == "https://kenigevents.ru/partnerstvo/"
    assert badge["topic_cluster"] == "event_badges_filters"
    assert badge["matched_intent"] == "event_badge_or_filter_request"
    assert exhibitions["topic_cluster"] == "event_site_search"
    assert exhibitions["link_target"]["url"] == "https://kenigevents.ru/vystavki/"


def test_build_opportunity_from_message_is_review_only_with_sticker_observation():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://t.me/example", platform="tg")
    message = SimpleNamespace(id=12, message="Где послушать концерт на выходных? 😂", date=datetime(2026, 7, 1, tzinfo=timezone.utc))
    opp = runtime.build_opportunity_from_message(surface, message, default_target_url="https://t.me/kenigevents")
    assert opp is not None
    assert opp["context_url"] == "https://t.me/example/12"
    assert opp["link_target"]["kind"] == "pka_channel"
    assert opp["scores"]["source"] == "deterministic_shadow_prefilter"
    assert opp["sticker_observation"]["fit"] == "possible"






def test_badge_word_in_advice_is_not_filter_request():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://vk.com/example", platform="vk")

    opp = runtime.build_vk_opportunity(
        surface,
        owner_id=-1,
        post_id=2,
        comment={"id": 3, "text": "Вот вам совет: добавьте игры с картинками растений. Детям нравятся такие занятия."},
        default_target_url="https://t.me/kenigevents",
    )

    assert opp is None


def test_organizer_thanks_is_not_partnership_opportunity():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://vk.com/example", platform="vk")

    opp = runtime.build_vk_opportunity(
        surface,
        owner_id=-1,
        post_id=2,
        comment={"id": 3, "text": "Спасибо организаторам за прекрасно проведенное время❤"},
        default_target_url="https://t.me/kenigevents",
    )

    assert opp is None


def test_existing_event_logistics_comment_is_not_acquisition_opportunity():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://vk.com/example", platform="vk")

    assert runtime.build_vk_opportunity(
        surface,
        owner_id=-1,
        post_id=2,
        comment={"id": 3, "text": "Подскажите, до скольки мероприятие?"},
        default_target_url="https://t.me/kenigevents",
    ) is None


def test_venue_policy_question_is_not_badge_filter_opportunity():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://vk.com/yunost_park", platform="vk")

    assert runtime.build_vk_opportunity(
        surface,
        owner_id=-1,
        post_id=2,
        comment={"id": 3, "text": "Здравствуйте, есть ли у вас какие-то льготы для детей-инвалидов????"},
        default_target_url="https://t.me/kenigevents",
    ) is None


def test_citywide_accessible_event_search_is_still_badge_filter_opportunity():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://vk.com/example", platform="vk")

    opp = runtime.build_vk_opportunity(
        surface,
        owner_id=-1,
        post_id=2,
        comment={"id": 4, "text": "Где найти мероприятия с льготами и доступностью для детей-инвалидов в Калининграде?"},
        default_target_url="https://t.me/kenigevents",
    )

    assert opp is not None
    assert opp["matched_intent"] == "event_badge_or_filter_request"
    assert opp["topic_cluster"] == "event_badges_filters"


def test_generic_where_comment_is_not_event_opportunity():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://t.me/example", platform="tg")
    message = SimpleNamespace(id=12, message="А где 34 автобус?!", date=datetime(2026, 7, 1, tzinfo=timezone.utc))

    assert runtime.build_opportunity_from_message(surface, message, default_target_url="https://t.me/kenigevents") is None


def test_telegram_opportunity_filter_requires_comments_not_channel_posts():
    runtime = load_runtime()

    channel_post = SimpleNamespace(post=True, fwd_from=None, reply_to=None)
    copied_discussion_post = SimpleNamespace(post=False, fwd_from=object(), reply_to=object())
    linked_comment = SimpleNamespace(post=False, fwd_from=None, reply_to=object())
    group_message = SimpleNamespace(post=False, fwd_from=None, reply_to=None)

    assert runtime.is_comment_opportunity_message(channel_post, surface_type="channel", relation=None) is False
    assert runtime.is_comment_opportunity_message(copied_discussion_post, surface_type="linked_discussion", relation="linked_discussion") is False
    assert runtime.is_comment_opportunity_message(linked_comment, surface_type="linked_discussion", relation="linked_discussion") is True
    assert runtime.is_comment_opportunity_message(group_message, surface_type="group", relation=None) is True


def test_shadow_payload_keeps_vk_seed_candidate_without_allowlist(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_TG_SEEDS_JSON", '["https://t.me/a_public"]')
    monkeypatch.setenv("ACQ_VK_SEEDS_JSON", '["https://vk.com/test_public"]')
    monkeypatch.delenv("ACQ_VK_ALLOWLIST_JSON", raising=False)
    payload = runtime.build_shadow_payload()
    by_external = {s["external_id"]: s for s in payload["surfaces"]}
    assert by_external["tg:a_public"]["status"] == "needs_comment_resolve"
    assert by_external["vk:test_public"]["status"] == "candidate"
    assert payload["stats"]["external_sends"] == 0
    assert payload["stats"]["comments_posted"] == 0
    assert payload["stats"]["stickers_sent"] == 0
    assert "tg_scan" in payload["stats"]


def test_shadow_payload_preserves_scanned_vk_surface_metadata(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_VK_SEEDS_JSON", '["https://vk.com/vagonka39"]')
    monkeypatch.setenv("ACQ_VK_ALLOWLIST_JSON", '["https://vk.com/vagonka39"]')
    monkeypatch.setenv("ACQ_VK_SEED_SURFACES_JSON", json.dumps([
        {
            "url": "https://vk.com/vagonka39",
            "handle": "vagonka39",
            "external_id": "vk:vagonka39",
            "title": "Вагонка",
            "source": "vk_source",
            "topic_hint": "existing VK monitoring source",
            "reach": {"members": 1234, "basis": "vk_source_seed"},
        }
    ], ensure_ascii=False))

    scanned = runtime._seed_surface("https://vk.com/vagonka39", platform="vk")
    scanned.update({"source": "allowlist", "status": "approved", "reach": {"basis": "vk_wall", "confidence": "low"}})
    payload = runtime.build_shadow_payload(scanned_surfaces=[scanned], scanned_opportunities=[])

    by_external = {s["external_id"]: s for s in payload["surfaces"]}
    assert by_external["vk:vagonka39"]["source"] == "allowlist"
    assert by_external["vk:vagonka39"]["reach"]["basis"] == "vk_wall"


def test_shadow_payload_applies_vk_seed_metadata_without_scan(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_VK_SEEDS_JSON", '["https://vk.com/vagonka39"]')
    monkeypatch.delenv("ACQ_VK_ALLOWLIST_JSON", raising=False)
    monkeypatch.setenv("ACQ_VK_SEED_SURFACES_JSON", json.dumps([
        {
            "url": "https://vk.com/vagonka39",
            "handle": "vagonka39",
            "external_id": "vk:vagonka39",
            "title": "Вагонка",
            "source": "vk_source",
            "topic_hint": "existing VK monitoring source",
            "reach": {"members": 4321, "basis": "vk_source_seed"},
        }
    ], ensure_ascii=False))

    payload = runtime.build_shadow_payload()

    by_external = {s["external_id"]: s for s in payload["surfaces"]}
    assert by_external["vk:vagonka39"]["title"] == "Вагонка"
    assert by_external["vk:vagonka39"]["source"] == "vk_source"
    assert by_external["vk:vagonka39"]["reach"]["members"] == 4321


def test_comment_retrieval_surface_inventory_includes_unscanned_event_sources(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_TG_SEEDS_JSON", json.dumps([
        "https://t.me/scanned_chat",
        "https://t.me/event_source_waiting",
    ], ensure_ascii=False))
    monkeypatch.setenv("ACQ_TG_SEED_SURFACES_JSON", json.dumps([
        {
            "url": "https://t.me/event_source_waiting",
            "handle": "event_source_waiting",
            "external_id": "tg:event_source_waiting",
            "surface_type": "unknown_public",
            "source": "tg_monitoring_canonical",
            "title": "Event Source Waiting",
            "topic_hint": "canonical Telegram Monitoring source",
        }
    ], ensure_ascii=False))

    inventory = runtime._surface_inventory_by_external([
        {
            "platform": "tg",
            "surface_type": "group",
            "url": "https://t.me/scanned_chat",
            "handle": "scanned_chat",
            "external_id": "tg:scanned_chat",
            "status": "candidate",
            "source": "linked_discussion",
        }
    ])

    assert inventory["tg:scanned_chat"]["status"] == "candidate"
    assert inventory["tg:event_source_waiting"]["source"] == "tg_monitoring_canonical"
    assert inventory["tg:event_source_waiting"]["status"] == "needs_comment_resolve"
    assert inventory["tg:event_source_waiting"]["title"] == "Event Source Waiting"


def test_shadow_payload_exposes_opportunity_screening_counters(monkeypatch):
    runtime = load_runtime()
    for key in runtime.OPPORTUNITY_SCREENING_STATS:
        runtime.OPPORTUNITY_SCREENING_STATS[key] = 0

    surface = runtime._seed_surface("https://vk.com/club123", platform="vk")
    comment = {"id": 7, "date": 1782900000, "text": "Подскажите, куда сходить на выходных?"}

    assert runtime.build_vk_opportunity(surface, owner_id=-123, post_id=55, comment=comment, default_target_url="https://t.me/kenigevents")
    payload = runtime.build_shadow_payload(scanned_surfaces=[surface], scanned_opportunities=[])

    screening = payload["stats"]["opportunity_screening"]
    assert screening["texts_screened"] == 1
    assert screening["matched_event_question"] == 1
    assert screening["no_intent"] == 0


def test_comment_semantic_retrieval_writes_profiles_candidates_and_xlsx(monkeypatch, tmp_path):
    retrieval = load_retrieval()
    monkeypatch.setenv("ACQ_COMMENT_RETRIEVAL_MODELS_JSON", '["intfloat/multilingual-e5-base", "BAAI/bge-m3"]')
    monkeypatch.setenv("ACQ_COMMENT_RETRIEVAL_MAX_LLM_CANDIDATES", "6")
    monkeypatch.setenv("ACQ_COMMENT_RETRIEVAL_MANUAL_SAMPLE_ROWS", "20")
    records = [
        {
            "surface_key": "tg:vKalinigrad_recomendations",
            "surface_external_id": "tg:vKalinigrad_recomendations",
            "surface_url": "https://t.me/vKalinigrad_recomendations",
            "platform": "tg",
            "surface_type": "group",
            "context_url": "https://t.me/vKalinigrad_recomendations/11",
            "comment_id": "11",
            "created_at": "2026-07-01T10:00:00+00:00",
            "author_id": "101",
            "relation": "group_message",
            "text": "Куда съездить из Калининграда на один день на электричке, Светлогорск или Зеленоградск?",
        },
        {
            "surface_key": "vk:vagonka39",
            "surface_external_id": "vk:vagonka39",
            "surface_url": "https://vk.com/vagonka39",
            "platform": "vk",
            "surface_type": "community",
            "context_url": "https://vk.com/wall-1_2?reply=3",
            "comment_id": "3",
            "created_at": "2026-07-02T10:00:00+00:00",
            "author_id": "202",
            "relation": "vk_comment",
            "text": "Будет ли мероприятие для детей, нужна регистрация и есть ли билеты?",
        },
        {
            "surface_key": "vk:vagonka39",
            "surface_external_id": "vk:vagonka39",
            "surface_url": "https://vk.com/vagonka39",
            "platform": "vk",
            "surface_type": "community",
            "context_url": "https://vk.com/wall-1_2",
            "comment_id": "",
            "post_id": "2",
            "created_at": "2026-07-02T09:00:00+00:00",
            "author_id": "-1",
            "relation": "vk_social_wall_post",
            "is_post": True,
            "text": "Подскажите, куда сходить на выходных?",
        },
    ]

    result = retrieval.run_comment_semantic_retrieval(
        records,
        surfaces_by_external={
            "tg:vKalinigrad_recomendations": {"title": "Калининград рекомендации", "url": "https://t.me/vKalinigrad_recomendations"},
            "vk:vagonka39": {"title": "Вагонка", "url": "https://vk.com/vagonka39", "reach": {"members": 1234}},
        },
        output_dir=tmp_path,
        backend=retrieval.HashingEmbeddingBackend(),
    )

    summary = result["summary"]
    assert summary["stage"] == retrieval.STAGE_NAME
    assert summary["comments_embedded"] == 3
    assert summary["surface_profiles_count"] == 2
    assert summary["candidate_count"] > 0
    assert result["llm_gate_candidates"]
    assert Path(summary["artifacts"]["manual_review_xlsx"]).exists()
    assert Path(summary["artifacts"]["candidates_csv"]).exists()
    assert Path(summary["artifacts"]["question_patterns_csv"]).exists()
    from openpyxl import load_workbook

    workbook = load_workbook(summary["artifacts"]["manual_review_xlsx"], read_only=True)
    assert "summary_ru" in workbook.sheetnames
    assert "decision_deltas" in workbook.sheetnames
    assert "processed_comments_last_run" in workbook.sheetnames
    assert "rejected_noise_examples" in workbook.sheetnames
    assert "summary_counts" in workbook.sheetnames
    assert "intent_catalog" in workbook.sheetnames
    assert "region_catalog" in workbook.sheetnames
    assert "full_surface_list" in workbook.sheetnames
    assert "question_patterns" in workbook.sheetnames
    assert "canonical_questions" in workbook.sheetnames
    assert "monitoring_targets" in workbook.sheetnames
    assert "surface_backlog" in workbook.sheetnames
    assert "semantic_union_candidates" in workbook.sheetnames
    assert "reply_event_candidates" in workbook.sheetnames
    assert "reply_route_candidates" in workbook.sheetnames
    assert "ask_organizer_candidates" in workbook.sheetnames
    assert "llm_gate_results" in workbook.sheetnames
    assert "model_comparison" in workbook.sheetnames
    assert "surface_scan_funnel" in workbook.sheetnames
    assert "goal_ask_event_details" in workbook.sheetnames
    assert "goal_reply_events" in workbook.sheetnames
    assert "goal_reply_routes" in workbook.sheetnames
    assert "answerable_e5_base" in workbook.sheetnames
    assert "answerable_bge_m3" in workbook.sheetnames
    assert workbook.sheetnames[:10] == [
        "summary_ru",
        "run_delta_sources",
        "monitoring_targets",
        "surface_backlog",
        "semantic_union_candidates",
        "reply_event_candidates",
        "reply_route_candidates",
        "ask_organizer_candidates",
        "llm_gate_results",
        "model_comparison",
    ]
    summary_values = [row[1] for row in workbook["summary_ru"].iter_rows(min_row=3, max_col=2, values_only=True)]
    assert "Площадок с анализом комментариев" in summary_values
    route_candidates = [c for c in result["candidates"] if c["candidate_action_type"] == "trip_route_poi_recommendation"]
    assert route_candidates
    assert any(c["target_hint"]["route_target_status"] == "route_needed" for c in route_candidates)
    assert {p["surface_key"] for p in result["surface_profiles"]} == {"tg:vKalinigrad_recomendations", "vk:vagonka39"}
    vk_profile = next(p for p in result["surface_profiles"] if p["surface_key"] == "vk:vagonka39")
    assert vk_profile["members_or_subscribers"] == 1234
    assert vk_profile["unique_commenters"] == 2
    assert vk_profile["source_post_records"] == 1
    assert vk_profile["period_days"] is not None
    assert Path(summary["artifacts"]["surface_decision_summary_csv"]).exists()
    assert result["surface_decision_summaries"]
    assert "latest_route_recommendation_at" in result["surface_decision_summaries"][0]
    assert result["surface_decision_summaries"][0].get("comments_per_30d") is not None
    assert result["summary"]["artifacts"]["question_patterns_csv"].endswith("comment_retrieval_question_patterns.csv")
    assert result["summary"]["artifacts"]["canonical_questions_csv"].endswith("comment_retrieval_canonical_questions.csv")
    assert result["canonical_questions"]
    assert result["summary"]["region_gate_required"] is True
    assert any(p["region_confidence"] == "confirmed" for p in result["surface_profiles"])


def test_comment_retrieval_xlsx_highlights_selected_and_increment(tmp_path):
    retrieval = load_retrieval()
    path = tmp_path / "review.xlsx"

    retrieval._write_xlsx(
        path,
        [],
        surface_summaries=[
            {
                "recommendation": "monitor_for_reply_opportunities",
                "selection_status": "selected",
                "surface_key": "tg:selected",
                "surface_url": "https://t.me/selected",
                "increment_status": "analyzed_comments_this_run",
            }
        ],
        surface_inventory=[
            {
                "surface_key": "vk:id123",
                "platform": "vk",
                "surface_type": "profile",
                "surface_url": "https://vk.com/id123",
                "selection_status": "candidate",
                "increment_status": "newly_discovered_this_run",
            }
        ],
        summary_counts=[
            {
                "platform": "ALL",
                "surface_type": "ALL",
                "total_surfaces": 1,
                "selected_surfaces": 1,
                "candidate_surfaces": 0,
                "rejected_surfaces": 0,
                "newly_discovered_this_run": 1,
                "removed_surfaces_this_run": 0,
                "increment_touched_this_run": 1,
                "analyzed_comments_this_run": 1,
                "selected_in_delta_this_run": 1,
                "candidate_in_delta_this_run": 0,
                "rejected_in_delta_this_run": 0,
                "comments_embedded_delta_this_run": 3,
                "answerable_questions_delta_this_run": 1,
            }
        ],
        decision_delta_rows=[
            {
                "delta_type": "newly_discovered_this_run",
                "decision_change_ru": "новая площадка",
                "surface_key": "vk:id123",
                "surface_url": "https://vk.com/id123",
                "selection_status": "candidate",
                "increment_status": "newly_discovered_this_run",
            }
        ],
        monitoring_target_rows=[
            {
                "selection_status": "selected",
                "recommendation": "monitor_for_reply_opportunities",
                "surface_title": "Selected",
                "surface_url": "https://t.me/selected",
                "platform": "tg",
                "surface_type_ru": "Telegram-группа/чат",
                "answerable_question_candidates": 1,
            }
        ],
        goal_candidate_rows={
            "goal_reply_routes": [
                {
                    "future_goal_ru": "отвечать на вопросы рекомендацией маршрута",
                    "context_url": "https://t.me/selected/1",
                    "text_snapshot": "Куда съездить на один день?",
                    "source_post_text_snapshot": "Про поездки по области",
                    "candidate_action_type": "trip_route_poi_recommendation",
                    "models_matched": "intfloat/multilingual-e5-base, BAAI/bge-m3",
                    "model_count": 2,
                    "best_score": 0.5,
                }
            ],
            "goal_ask_event_details": [],
            "goal_reply_events": [],
            "goal_other_acq": [],
        },
        processed_comment_rows=[
            {
                "criteria_status_ru": "соответствует: реальный вопрос",
                "criteria_status": "accepted_reply_candidate",
                "analysis_kind": "user_comment",
                "context_url": "https://vk.com/wall-1_2?reply=3",
                "text_snapshot": "Куда сходить?",
                "source_post_text_snapshot": "Анонс концерта",
            },
            {
                "criteria_status_ru": "это исходный пост/контекст",
                "criteria_status": "rejected_source_post_context",
                "analysis_kind": "post_context",
                "context_url": "https://vk.com/wall-1_2",
                "relation": "vk_social_wall_post",
                "is_post": True,
                "text_snapshot": "Анонс события, по которому можно спросить детали",
            },
        ],
        rejected_noise_rows=[
            {
                "criteria_status_ru": "отфильтровано как шум",
                "criteria_status": "rejected_noise:intent_without_text_support",
                "analysis_kind": "user_comment",
                "context_url": "https://vk.com/wall-1_2?reply=4",
                "text_snapshot": "Потерял телефон",
            }
        ],
    )

    from openpyxl import load_workbook

    workbook = load_workbook(path)
    assert workbook.sheetnames[:4] == ["run_delta_sources", "monitoring_targets", "surface_backlog", "goal_ask_event_details"]
    assert workbook["surface_summary"]["A3"].fill.fgColor.rgb == "00C6EFCE"
    assert workbook["full_surface_list"]["A3"].fill.fgColor.rgb == "00FFF2CC"
    assert workbook["summary_counts"]["G3"].fill.fgColor.rgb == "00FFF2CC"
    assert workbook["monitoring_targets"]["A3"].fill.fgColor.rgb == "00C6EFCE"
    assert "surface_backlog" in workbook.sheetnames
    assert workbook["goal_reply_routes"]["A3"].fill.fgColor.rgb == "00E2F0D9"
    processed = workbook["processed_comments_last_run"]
    headers = [cell.value for cell in processed[2]]
    assert "Текущий комментарий" in headers
    assert "Текущий пост / анонс" in headers
    assert "Комментарий" not in headers
    comment_col = headers.index("Текущий комментарий") + 1
    post_col = headers.index("Текущий пост / анонс") + 1
    source_post_col = headers.index("Пост, под которым написан комментарий") + 1
    assert processed.cell(3, comment_col).value == "Куда сходить?"
    assert processed.cell(3, post_col).value in (None, "")
    assert processed.cell(3, source_post_col).value == "Анонс концерта"
    assert processed.cell(3, 1).fill.fgColor.rgb == "00C6EFCE"
    assert processed.cell(4, comment_col).value in (None, "")
    assert processed.cell(4, post_col).value == "Анонс события, по которому можно спросить детали"
    assert processed.cell(4, source_post_col).value in (None, "")
    assert processed.cell(4, 1).fill.fgColor.rgb == "00F4CCCC"


def test_comment_retrieval_increment_requires_verified_run_id(monkeypatch):
    retrieval = load_retrieval()
    monkeypatch.delenv("KAGGLE_RUN_ID", raising=False)
    monkeypatch.delenv("ACQ_SOURCE_RUN_ID", raising=False)

    inventory = retrieval._surface_inventory_rows(
        {
            "tg:test": {
                "platform": "tg",
                "surface_type": "group",
                "url": "https://t.me/test",
                "external_id": "tg:test",
                "scan_state": "scanned",
                "scan_run_id": "old-run",
            }
        },
        [
            {
                "surface_key": "tg:test",
                "platform": "tg",
                "surface_type": "group",
                "comments_embedded": 3,
                "selection_status": "candidate",
            }
        ],
    )

    assert inventory[0]["increment_status"] == "no_verified_run_id"
    counts = retrieval._summary_count_rows(inventory)
    assert counts[0]["increment_touched_this_run"] == 0
    assert counts[0]["analyzed_comments_this_run"] == 0

    monkeypatch.setenv("KAGGLE_RUN_ID", "run-1")
    inventory = retrieval._surface_inventory_rows(
        {
            "tg:test": {
                "platform": "tg",
                "surface_type": "group",
                "url": "https://t.me/test",
                "external_id": "tg:test",
                "scan_state": "scanned",
                "scan_run_id": "run-1",
            }
        },
        [
            {
                "surface_key": "tg:test",
                "platform": "tg",
                "surface_type": "group",
                "comments_embedded": 3,
                "selection_status": "candidate",
                "last_analyzed_run_id": "run-1",
            }
        ],
    )

    assert inventory[0]["increment_status"] == "analyzed_comments_this_run"
    counts = retrieval._summary_count_rows(inventory)
    assert counts[0]["increment_touched_this_run"] == 1
    assert counts[0]["analyzed_comments_this_run"] == 1

    inventory = retrieval._surface_inventory_rows(
        {
            "vk:id123": {
                "platform": "vk",
                "surface_type": "profile",
                "url": "https://vk.com/id123",
                "external_id": "vk:id123",
                "source": "discovered_vk_author",
                "discovered_in_run_id": "run-1",
            },
            "tg:seed": {
                "platform": "tg",
                "surface_type": "channel",
                "url": "https://t.me/seed",
                "external_id": "tg:seed",
                "source": "tg_monitoring_canonical",
                "discovered_in_run_id": "run-1",
            },
        },
        [],
    )
    by_key = {row["surface_key"]: row for row in inventory}
    assert by_key["vk:id123"]["increment_status"] == "queued_discovered_backlog_this_run"
    assert by_key["tg:seed"]["increment_status"] == "seed_backlog_visible_this_run"
    counts = retrieval._summary_count_rows(inventory)
    assert counts[0]["queued_discovered_backlog_this_run"] == 1
    assert counts[0]["seed_backlog_visible_this_run"] == 1
    assert counts[0]["increment_touched_this_run"] == 0
    delta_rows = retrieval._decision_delta_rows(inventory, scope_rows=[{"metric": "source_run_provenance", "value": "kaggle_run_id"}])
    assert {row["delta_type"] for row in delta_rows} == {"queued_discovered_backlog_this_run", "seed_backlog_visible_this_run"}


def test_comment_semantic_retrieval_vector_scan_does_not_prefilter_before_llm(monkeypatch):
    retrieval = load_retrieval()
    monkeypatch.delenv("ACQ_COMMENT_RETRIEVAL_DETERMINISTIC_PREFILTER", raising=False)

    question = retrieval._text_quality_features("Подскажите, куда сходить в Светлогорске с детьми?")
    offer = retrieval._text_quality_features("Маршрут одного дня в Калининграде. Сохраняйте и записывайтесь на экскурсию по ссылке https://example.com")

    assert question["question_signal"] is True
    assert question["hard_noise"] is False
    assert offer["question_signal"] is False
    assert offer["hard_noise"] is True
    assert offer["noise_type"] == "explicit_offer_or_ad"

    row = {"text_snapshot": "Сохраняйте подборку и бронируйте тур", "positive_negative_margin": 0.5}
    retrieval._apply_text_quality_to_candidate(row, scoring_method="positive_negative_margin")
    assert row["pre_llm_candidate_eligible"] is True
    assert row["score_for_rank"] == 0.5
    assert row["llm_gate_selection_basis"] == "semantic_top_n_no_deterministic_prefilter"

    unsupported = {"text_snapshot": "Как зовут детишек?", "positive_negative_margin": 0.5, "intent_set": "organizer_comment_fit"}
    retrieval._apply_text_quality_to_candidate(unsupported, scoring_method="positive_negative_margin")
    assert unsupported["pre_llm_candidate_eligible"] is True
    assert unsupported["candidate_noise_type"] == "intent_without_text_support"

    row2 = {"text_snapshot": "Куда съездить из Калининграда на один день?", "positive_negative_margin": 0.5, "intent_set": "route_poi_close_actionable"}
    retrieval._apply_text_quality_to_candidate(row2, scoring_method="positive_negative_margin")
    assert row2["pre_llm_candidate_eligible"] is True
    assert row2["score_for_rank"] == 0.5

    source_post = {
        "text_snapshot": "Куда съездить из Калининграда на один день?",
        "positive_negative_margin": 0.5,
        "intent_set": "route_poi_close_actionable",
        "relation": "vk_social_wall_post",
    }
    retrieval._apply_text_quality_to_candidate(source_post, scoring_method="positive_negative_margin")
    assert source_post["pre_llm_candidate_eligible"] is True
    assert source_post["candidate_noise_type"] == "source_post_context"

    real_estate = {"text_snapshot": "Где снять квартиру на длительный срок?", "positive_negative_margin": 0.5, "intent_set": "event_site_search_or_listing"}
    retrieval._apply_text_quality_to_candidate(real_estate, scoring_method="positive_negative_margin")
    assert real_estate["pre_llm_candidate_eligible"] is True
    assert real_estate["candidate_noise_type"] == "out_of_scope_real_estate"

    medicine = {"text_snapshot": "Посоветуйте стоматолога и клинику", "positive_negative_margin": 0.5, "intent_set": "badge_filter_need"}
    retrieval._apply_text_quality_to_candidate(medicine, scoring_method="positive_negative_margin")
    assert medicine["pre_llm_candidate_eligible"] is True
    assert medicine["candidate_noise_type"] == "out_of_scope_medicine"

    route_with_flat_context = {"text_snapshot": "Снимаем квартиру в Калининграде, куда съездить на один день?", "positive_negative_margin": 0.5, "intent_set": "route_poi_close_actionable"}
    retrieval._apply_text_quality_to_candidate(route_with_flat_context, scoring_method="positive_negative_margin")
    assert route_with_flat_context["pre_llm_candidate_eligible"] is True


def _semantic_goal_row(retrieval, *, key: str, goal: str, model: str, margin: float, channel: str | None = None, created_at: str = "2026-07-01T10:00:00+00:00", relation: str = "group_message", is_post: bool = False, text: str = "Нужен совет по событию без вопросительного знака"):
    return {
        "surface_key": "tg:test",
        "platform": "tg",
        "surface_type": "linked_discussion",
        "context_url": f"https://t.me/test/{key}",
        "comment_id": key,
        "created_at": created_at,
        "relation": relation,
        "is_post": is_post,
        "text": text,
        "text_snapshot": text,
        "goal": goal,
        "intent_set": goal,
        "candidate_action_type": retrieval.SEMANTIC_GOAL_ACTION_TYPES[goal],
        "model_name": model,
        "retrieval_channel": channel or retrieval._model_dense_channel(model),
        "positive_score": margin + 0.5,
        "negative_score": 0.5,
        "semantic_margin": margin,
        "production_bucket": retrieval._production_bucket_for_goal({"created_at": created_at}, goal),
        "region_confidence": "confirmed",
        "region_gate_status": "semantic_region_confirmed",
    }


def test_semantic_union_allows_bge_only_and_e5_only_candidates(monkeypatch):
    retrieval = load_retrieval()
    monkeypatch.setenv("ACQ_COMMENT_RETRIEVAL_MAX_LLM_CANDIDATES", "10")
    rows = [
        _semantic_goal_row(retrieval, key="e5", goal="reply_to_user_event_question", model="intfloat/multilingual-e5-base", margin=0.9),
        _semantic_goal_row(retrieval, key="bge", goal="reply_to_user_event_question", model="BAAI/bge-m3", margin=0.95),
    ]

    union = retrieval._build_semantic_union_candidates(rows)
    by_url = {row["context_url"]: row for row in union}

    assert by_url["https://t.me/test/e5"]["hit_by_e5_dense"] is True
    assert by_url["https://t.me/test/e5"]["hit_by_bge_m3_dense"] is False
    assert by_url["https://t.me/test/bge"]["hit_by_bge_m3_dense"] is True
    assert by_url["https://t.me/test/bge"]["hit_by_e5_dense"] is False
    selected = retrieval._select_semantic_llm_gate_candidates(union, 10)
    assert {row["context_url"] for row in selected} == {"https://t.me/test/e5", "https://t.me/test/bge"}


def test_semantic_union_does_not_drop_source_posts_or_comments_without_question_mark(monkeypatch):
    retrieval = load_retrieval()
    monkeypatch.delenv("ACQ_INCLUDE_HISTORICAL_IN_LLM_QUEUE", raising=False)
    source_post = _semantic_goal_row(
        retrieval,
        key="post",
        goal="ask_organizer_event_details",
        model="BAAI/bge-m3",
        margin=0.9,
        relation="vk_social_wall_post",
        is_post=True,
        text="В субботу состоится мастер-класс для детей",
    )
    no_question_comment = _semantic_goal_row(
        retrieval,
        key="comment",
        goal="reply_to_user_route_recommendation",
        model="intfloat/multilingual-e5-base",
        margin=0.91,
        text="Посоветуйте маршрут по области на выходные",
    )

    union = retrieval._build_semantic_union_candidates([source_post, no_question_comment])
    selected = retrieval._select_semantic_llm_gate_candidates(union, 10)

    assert {row["context_url"] for row in selected} == {"https://t.me/test/post", "https://t.me/test/comment"}
    assert any(row["is_post"] is True and row["goal"] == "ask_organizer_event_details" for row in selected)


def test_semantic_union_separates_historical_rows_from_production_queue(monkeypatch):
    retrieval = load_retrieval()
    monkeypatch.delenv("ACQ_INCLUDE_HISTORICAL_IN_LLM_QUEUE", raising=False)
    old_row = _semantic_goal_row(
        retrieval,
        key="old",
        goal="reply_to_user_event_question",
        model="BAAI/bge-m3",
        margin=0.99,
        created_at="2025-01-01T10:00:00+00:00",
    )
    new_row = _semantic_goal_row(
        retrieval,
        key="new",
        goal="reply_to_user_event_question",
        model="intfloat/multilingual-e5-base",
        margin=0.8,
    )

    union = retrieval._build_semantic_union_candidates([old_row, new_row])
    selected = retrieval._select_semantic_llm_gate_candidates(union, 10)

    assert any(row["production_bucket"] == "historical_calibration_candidates" for row in union)
    assert {row["context_url"] for row in selected} == {"https://t.me/test/new"}


def test_semantic_goals_are_split_between_event_route_and_organizer():
    retrieval = load_retrieval()

    assert "reply_to_user_event_or_route" not in retrieval.SEMANTIC_GOAL_SETS
    assert retrieval.SEMANTIC_GOAL_ACTION_TYPES["reply_to_user_event_question"] == "event_recommendation_reply"
    assert retrieval.SEMANTIC_GOAL_ACTION_TYPES["reply_to_user_route_recommendation"] == "trip_route_poi_recommendation"
    assert retrieval.SEMANTIC_GOAL_ACTION_TYPES["ask_organizer_event_details"] == "organizer_visibility_clarification"


def test_semantic_topk_is_adaptive_for_tiny_samples(monkeypatch):
    retrieval = load_retrieval()
    monkeypatch.setenv("ACQ_E5_DENSE_TOPK_PER_GOAL", "80")
    monkeypatch.setenv("ACQ_SEMANTIC_TOPK_FRACTION_PER_GOAL", "0.20")
    monkeypatch.setenv("ACQ_SEMANTIC_TOPK_MIN_PER_GOAL", "5")
    rows = [
        _semantic_goal_row(
            retrieval,
            key=str(i),
            goal="reply_to_user_event_question",
            model="intfloat/multilingual-e5-base",
            margin=1.0 - (i / 100.0),
        )
        for i in range(50)
    ]

    top = retrieval._topk_per_goal_channel(rows)

    assert len(top) == 10
    assert {row["adaptive_topk_effective"] for row in top} == {10}


def test_semantic_llm_queue_is_goal_balanced(monkeypatch):
    retrieval = load_retrieval()
    monkeypatch.setenv("ACQ_LLM_FINAL_SCOPE_MAX_TOTAL", "6")
    monkeypatch.setenv("ACQ_LLM_FINAL_SCOPE_MAX_PER_GOAL", "2")
    rows = []
    for goal in ["reply_to_user_event_question", "reply_to_user_route_recommendation", "ask_organizer_event_details"]:
        for idx in range(4):
            rows.append(_semantic_goal_row(
                retrieval,
                key=f"{goal}-{idx}",
                goal=goal,
                model="intfloat/multilingual-e5-base",
                margin=0.95 - (idx / 100.0),
                relation="vk_social_wall_post" if goal == "ask_organizer_event_details" else "group_message",
                is_post=goal == "ask_organizer_event_details",
            ))
    union = retrieval._build_semantic_union_candidates(rows)
    selected = retrieval._select_semantic_llm_gate_candidates(union, 99)
    counts = Counter(row["goal"] for row in selected)

    assert len(selected) == 6
    assert counts["reply_to_user_event_question"] == 2
    assert counts["reply_to_user_route_recommendation"] == 2
    assert counts["ask_organizer_event_details"] == 2


def test_comment_semantic_retrieval_requires_two_models_even_if_env_single(monkeypatch):
    retrieval = load_retrieval()
    monkeypatch.setenv("ACQ_COMMENT_RETRIEVAL_MODELS_JSON", '["intfloat/multilingual-e5-base"]')

    assert retrieval._load_models_from_env() == ["intfloat/multilingual-e5-base", "BAAI/bge-m3"]


def test_comment_semantic_region_gate_uses_surface_and_context_evidence():
    retrieval = load_retrieval()

    by_surface = retrieval._assess_record_region({
        "surface_title": "Калининград рекомендации",
        "surface_url": "https://t.me/vKalinigrad_recomendations",
        "text": "Куда съездить на один день?",
    })
    by_comment = retrieval._assess_record_region({
        "surface_title": "Городские советы",
        "text": "Куда съездить из Калининграда в Светлогорск на электричке?",
    })
    unknown = retrieval._assess_record_region({
        "surface_title": "Городские советы",
        "text": "Куда съездить на один день?",
    })
    out = retrieval._assess_record_region({
        "surface_title": "Visit Navahrudak",
        "surface_url": "https://t.me/visitNavahrudak",
        "text": "Что посмотреть за день?",
    })

    assert by_surface["region_confidence"] == "confirmed"
    assert by_surface["region_signal_source"] == "surface_metadata_keyword"
    assert by_comment["region_confidence"] == "confirmed"
    assert by_comment["region_signal_source"] == "comment_context_keyword"
    assert unknown["region_confidence"] == "unknown"
    assert out["region_confidence"] == "out_of_region"


def test_comment_semantic_region_gate_blocks_report_and_surface_selection():
    retrieval = load_retrieval()
    row = {
        "surface_key": "tg:unknown",
        "platform": "tg",
        "surface_type": "group",
        "candidate_action_type": "trip_route_poi_recommendation",
        "intent_set": "route_poi_close_actionable",
        "text_snapshot": "Куда съездить на один день?",
        "context_url": "https://t.me/unknown/1",
        "candidate_usage_scope": "monitoring_candidate",
        "question_signal": True,
        "intent_text_supported": True,
        "candidate_noise_type": "",
        "score": 0.5,
        "region_confidence": "unknown",
        "region_evidence_ru": "нет Калининградской зацепки",
    }

    assert retrieval._report_candidate_eligible(row) is False
    status, status_ru = retrieval._criteria_status(row)
    assert status == "rejected_region:unknown"
    assert "Калининградской области" in status_ru

    profile = {
        "surface_key": "tg:unknown",
        "platform": "tg",
        "surface_type": "group",
        "surface_title": "Городские советы",
        "surface_url": "https://t.me/unknown",
        "monitoring_decision_hint": "monitor",
        "region_confidence": "unknown",
        "region_evidence_ru": "нет Калининградской зацепки",
        "comments_embedded": 1,
    }
    summary = retrieval._surface_decision_summaries(
        [profile],
        eligible_rows_by_surface={"tg:unknown": [row]},
        all_gate_rows_by_surface={"tg:unknown": [row]},
    )[0]
    assert summary["recommendation"] == "reject_region_unknown"
    assert summary["selection_status"] == "rejected"


def test_comment_semantic_hard_filters_past_events_and_gasoline(monkeypatch):
    retrieval = load_retrieval()
    monkeypatch.setenv("ACQ_COMMENT_RETRIEVAL_NOW_ISO", "2026-07-05T12:00:00+00:00")

    past_event = {
        "surface_key": "vk:vagonka39",
        "platform": "vk",
        "surface_type": "community",
        "candidate_action_type": "organizer_visibility_clarification",
        "intent_set": "organizer_event_post_context",
        "text_snapshot": "В субботу 4 июля прошёл концерт, спасибо организаторам!",
        "context_url": "https://vk.com/wall-1_2",
        "created_at": "2026-07-04T08:00:00+00:00",
        "candidate_usage_scope": "monitoring_candidate",
        "region_confidence": "confirmed",
        "positive_negative_margin": 0.5,
    }
    retrieval._apply_text_quality_to_candidate(past_event, scoring_method="positive_negative_margin")

    assert past_event["semantic_candidate_rejected"] is True
    assert past_event["semantic_exclusion_type"] == "past_event"
    assert past_event["pre_llm_candidate_eligible"] is False
    assert retrieval._report_candidate_eligible(past_event) is False
    status, status_ru = retrieval._criteria_status(past_event)
    assert status == "rejected_temporal:past_event"
    assert "будущими событиями" in status_ru

    future_event = {
        **past_event,
        "text_snapshot": "8 июля состоится концерт, нужна ли регистрация?",
        "created_at": "2026-07-04T08:00:00+00:00",
        "candidate_action_type": "event_recommendation_reply",
        "intent_set": "event_close_question",
    }
    retrieval._apply_text_quality_to_candidate(future_event, scoring_method="positive_negative_margin")
    assert future_event["semantic_candidate_rejected"] is False
    assert future_event["event_temporal_status"] == "future_or_today"
    assert future_event["pre_llm_candidate_eligible"] is True

    mixed_past_today = {
        **past_event,
        "text_snapshot": "На стадионе прошёл большой концерт, сегодня снова будет шоу.",
        "created_at": "2026-07-05T08:00:00+00:00",
        "candidate_action_type": "organizer_visibility_clarification",
        "intent_set": "organizer_event_post_context",
    }
    retrieval._apply_text_quality_to_candidate(mixed_past_today, scoring_method="positive_negative_margin")
    assert mixed_past_today["semantic_exclusion_type"] == "past_event_signal"
    assert retrieval._goal_candidate_rows([mixed_past_today], actions={"organizer_visibility_clarification"}, source_posts=True) == []

    already_started = {
        **past_event,
        "text_snapshot": "На стадионе стартовал вечерний концерт, который открыл Родион Газманов.",
        "created_at": "2026-07-05T08:00:00+00:00",
        "candidate_action_type": "organizer_visibility_clarification",
        "intent_set": "organizer_event_post_context",
    }
    retrieval._apply_text_quality_to_candidate(already_started, scoring_method="positive_negative_margin")
    assert already_started["semantic_exclusion_type"] == "past_event_signal"

    gasoline = {
        "surface_key": "vk:club42481124",
        "platform": "vk",
        "surface_type": "community",
        "candidate_action_type": "trip_route_poi_recommendation",
        "intent_set": "route_poi_close_actionable",
        "text_snapshot": "Подскажите где сейчас есть бензин 95 на заправках?",
        "context_url": "https://vk.com/topic-1_2?post=3",
        "candidate_usage_scope": "monitoring_candidate",
        "region_confidence": "confirmed",
        "positive_negative_margin": 0.5,
    }
    retrieval._apply_text_quality_to_candidate(gasoline, scoring_method="positive_negative_margin")
    assert gasoline["candidate_noise_type"] == "out_of_scope_gasoline_availability"
    assert gasoline["semantic_candidate_rejected"] is True
    assert gasoline["pre_llm_candidate_eligible"] is False
    assert retrieval._report_candidate_eligible(gasoline) is False
    status, status_ru = retrieval._criteria_status(gasoline)
    assert status == "rejected_noise:out_of_scope_gasoline_availability"
    assert "бензина" in status_ru


def test_comment_semantic_legacy_deterministic_prefilter_is_explicit_opt_in(monkeypatch):
    retrieval = load_retrieval()
    monkeypatch.setenv("ACQ_COMMENT_RETRIEVAL_DETERMINISTIC_PREFILTER", "1")

    row = {"text_snapshot": "Сохраняйте подборку и бронируйте тур", "positive_negative_margin": 0.5}
    retrieval._apply_text_quality_to_candidate(row, scoring_method="positive_negative_margin")

    assert row["pre_llm_candidate_eligible"] is False
    assert row["score_for_rank"] < 0.5
    assert row["llm_gate_selection_basis"] == "legacy_deterministic_prefilter"


def test_question_pattern_label_covers_route_event_and_badges():
    retrieval = load_retrieval()

    assert retrieval._question_pattern_label("Куда съездить с детьми на один день?", action_type="trip_route_poi_recommendation") == "route_with_children"
    assert retrieval._question_pattern_label("Сколько стоит билет и нужен ли вход?", action_type="event_recommendation_reply") == "event_ticket_or_price"
    assert retrieval._question_pattern_label("Есть по Пушкинской карте?", action_type="badge_filter_need") == "event_badge_pushkin_card"
    rows = retrieval._build_question_patterns([
        {
            "surface_key": "tg:test",
            "platform": "tg",
            "surface_type": "group",
            "candidate_action_type": "trip_route_poi_recommendation",
            "intent_set": "route_poi_close_actionable",
            "text_snapshot": "Куда съездить с детьми на один день?",
            "context_url": "https://t.me/test/1",
            "model_name": "intfloat/multilingual-e5-base",
            "pre_llm_candidate_eligible": True,
            "candidate_usage_scope": "monitoring_candidate",
            "question_signal": True,
            "intent_text_supported": True,
            "candidate_noise_type": "",
            "score": 0.05,
            "region_confidence": "confirmed",
        },
        {
            "surface_key": "tg:test",
            "platform": "tg",
            "surface_type": "group",
            "candidate_action_type": "trip_route_poi_recommendation",
            "intent_set": "route_poi_close_actionable",
            "text_snapshot": "Ездили с детьми в парк",
            "context_url": "https://t.me/test/2",
            "model_name": "intfloat/multilingual-e5-base",
            "pre_llm_candidate_eligible": True,
            "candidate_usage_scope": "monitoring_candidate",
            "question_signal": False,
            "intent_text_supported": True,
            "candidate_noise_type": "non_question_statement",
            "score": 0.05,
        },
    ])
    assert rows[0]["pattern"] == "route_with_children"
    assert rows[0]["canonical_question_ru"].endswith("?")
    assert "Ездили" not in rows[0]["example_questions"]


def test_report_quality_filter_does_not_promote_vector_noise_to_canonical_questions():
    retrieval = load_retrieval()
    garbage = {
        "surface_key": "vk:club42481124",
        "platform": "vk",
        "surface_type": "community",
        "candidate_action_type": "trip_route_poi_recommendation",
        "intent_set": "route_poi_close_actionable",
        "text_snapshot": "Че там по роснефти за 159 дизель?",
        "context_url": "https://vk.com/wall-42481124_1?reply=2",
        "model_name": "intfloat/multilingual-e5-base",
        "candidate_usage_scope": "monitoring_candidate",
        "question_signal": True,
        "intent_text_supported": False,
        "candidate_noise_type": "intent_without_text_support",
        "score": 0.044,
        # Important regression contract: default vector/LLM gate still sees top
        # semantic rows; only report/canonical evidence is stricter.
        "pre_llm_candidate_eligible": True,
    }

    assert garbage["pre_llm_candidate_eligible"] is True
    assert retrieval._report_candidate_eligible(garbage) is False
    assert retrieval._build_question_patterns([garbage]) == []
    assert retrieval._canonical_question_catalog_rows([garbage]) == []


def test_report_quality_keeps_source_posts_as_ask_context_but_not_question_catalog():
    retrieval = load_retrieval()
    source_post = {
        "surface_key": "vk:organizer",
        "platform": "vk",
        "surface_type": "community",
        "candidate_action_type": "organizer_visibility_clarification",
        "intent_set": "organizer_comment_fit",
        "text_snapshot": "Приходите на открытие выставки 10 июля, начало в 18:00.",
        "context_url": "https://vk.com/wall-1_2",
        "model_name": "intfloat/multilingual-e5-base",
        "candidate_usage_scope": "monitoring_candidate",
        "question_signal": False,
        "intent_text_supported": True,
        "candidate_noise_type": "source_post_context",
        "score": 0.05,
        "relation": "vk_social_wall_post",
        "is_post": True,
        "region_confidence": "confirmed",
        "pre_llm_candidate_eligible": True,
    }

    assert retrieval._report_candidate_eligible(source_post) is True
    assert retrieval._report_real_question_row(source_post) is False
    assert retrieval._model_example_rows([source_post], "intfloat/multilingual-e5-base") == []
    assert retrieval._model_ask_context_rows([source_post], "intfloat/multilingual-e5-base") == [source_post]
    assert retrieval._build_question_patterns([source_post]) == []
    assert retrieval._canonical_question_catalog_rows([source_post]) == []

    user_question_wall_post = {
        **source_post,
        "text_snapshot": "У кого есть два лишних билета на концерт?",
        "question_signal": True,
        "context_url": "https://vk.com/wall-1_3",
    }
    assert retrieval._model_ask_context_rows([user_question_wall_post], "intfloat/multilingual-e5-base") == []
    assert retrieval._goal_candidate_rows(
        [user_question_wall_post],
        actions={"organizer_visibility_clarification"},
        source_posts=True,
    ) == []


def test_report_quality_allows_historical_real_questions_only_for_canonical_catalog():
    retrieval = load_retrieval()
    historical = {
        "surface_key": "tg:golden",
        "platform": "tg",
        "surface_type": "group",
        "candidate_action_type": "trip_route_poi_recommendation",
        "intent_set": "route_poi_close_actionable",
        "text_snapshot": "Куда съездить из Калининграда на один день с детьми?",
        "context_url": "https://t.me/golden/1",
        "model_name": "intfloat/multilingual-e5-base",
        "candidate_usage_scope": "historical_calibration",
        "question_signal": True,
        "intent_text_supported": True,
        "candidate_noise_type": "",
        "score": 0.05,
        "region_confidence": "confirmed",
        "pre_llm_candidate_eligible": True,
    }

    assert retrieval._report_candidate_eligible(historical) is False
    assert retrieval._report_candidate_eligible(historical, allow_historical=True) is True
    catalog = retrieval._canonical_question_catalog_rows([historical])
    assert catalog
    assert catalog[0]["historical_calibration_examples"] == 1


def test_model_answerable_examples_are_only_fresh_reply_questions():
    retrieval = load_retrieval()
    answerable = {
        "surface_key": "tg:golden",
        "platform": "tg",
        "surface_type": "group",
        "candidate_action_type": "trip_route_poi_recommendation",
        "intent_set": "route_poi_close_actionable",
        "text_snapshot": "Куда съездить из Калининграда на один день с детьми?",
        "context_url": "https://t.me/golden/1",
        "model_name": "intfloat/multilingual-e5-base",
        "candidate_usage_scope": "monitoring_candidate",
        "question_signal": True,
        "intent_text_supported": True,
        "candidate_noise_type": "",
        "score": 0.05,
        "region_confidence": "confirmed",
        "pre_llm_candidate_eligible": True,
    }
    source_post = {
        **answerable,
        "text_snapshot": "Приходите на выставку 10 июля, начало в 18:00.",
        "context_url": "https://t.me/golden/2",
        "relation": "tg_channel_post_context",
        "is_post": True,
        "intent_set": "organizer_event_post_context",
        "candidate_action_type": "organizer_visibility_clarification",
        "question_signal": False,
        "candidate_noise_type": "source_post_context",
    }

    assert retrieval._model_example_rows([answerable, source_post], "intfloat/multilingual-e5-base") == [answerable]
    assert retrieval._model_ask_context_rows([answerable, source_post], "intfloat/multilingual-e5-base") == [source_post]


def test_freshness_policy_keeps_old_records_for_calibration_but_rejects_stale_surfaces(monkeypatch):
    retrieval = load_retrieval()
    monkeypatch.setenv("ACQ_COMMENT_RETRIEVAL_MAX_COMMENT_AGE_DAYS", "365")
    monkeypatch.setenv("ACQ_COMMENT_RETRIEVAL_STALE_ACTIVITY_DAYS", "92")
    old = {
        "surface_key": "tg:old",
        "created_at": "2024-01-01T00:00:00+00:00",
        "text": "Куда сходить?",
    }
    fresh_but_stale_surface = {
        "surface_key": "tg:stale",
        "created_at": "2026-02-01T00:00:00+00:00",
        "text": "Куда сходить?",
    }

    deduped = retrieval._dedupe_records([old, fresh_but_stale_surface])

    assert [r["surface_key"] for r in deduped] == ["tg:old", "tg:stale"]
    assert deduped[0]["candidate_usage_scope"] == "historical_calibration"
    assert deduped[1]["candidate_usage_scope"] == "monitoring_candidate"
    profile = retrieval._surface_profile(
        "tg:stale",
        [fresh_but_stale_surface],
        [
            {
                "intent_set": "event_close_question",
                "score": 1.0,
                "funnel_bucket": "top_1pct",
                "pre_llm_candidate_eligible": True,
            }
        ],
        scoring_method="positive_negative_margin",
    )
    assert profile["monitoring_decision_hint"] == "reject_stale_inactive"
    summaries = retrieval._surface_decision_summaries(
        [profile],
        eligible_rows_by_surface={"tg:stale": []},
        all_gate_rows_by_surface={"tg:stale": []},
    )
    assert summaries[0]["selection_status"] == "rejected"


def test_summary_counts_groups_surface_inventory_statuses():
    retrieval = load_retrieval()
    rows = retrieval._summary_count_rows([
        {"platform": "tg", "surface_type": "group", "selection_status": "selected"},
        {"platform": "tg", "surface_type": "group", "selection_status": "candidate"},
        {"platform": "vk", "surface_type": "community", "selection_status": "rejected"},
    ])

    assert rows[0]["platform"] == "ALL"
    assert rows[0]["total_surfaces"] == 3
    assert rows[0]["selected_surfaces"] == 1
    assert rows[0]["candidate_surfaces"] == 1
    assert rows[0]["rejected_surfaces"] == 1


def test_surface_decision_summary_separates_reply_and_question_asking_modes():
    retrieval = load_retrieval()
    profile = {
        "surface_key": "tg:example",
        "platform": "tg",
        "surface_type": "group",
        "comments_embedded": 10,
        "dominant_detected_interests": ["route_poi", "event_questions"],
        "monitoring_decision_hint": "monitor",
        "monitoring_reason": "test",
        "region_confidence": "confirmed",
    }
    rows = [
        {
            "context_url": "https://t.me/example/1",
            "rank_global": 1,
            "candidate_action_type": "trip_route_poi_recommendation",
            "intent_set": "route_poi_close_actionable",
            "text_snapshot": "Куда съездить?",
            "pre_llm_candidate_eligible": True,
            "relation": "group_message",
        },
        {
            "context_url": "https://t.me/example/2",
            "rank_global": 2,
            "candidate_action_type": "event_recommendation_reply",
            "intent_set": "event_close_question",
            "text_snapshot": "А где купить билеты?",
            "pre_llm_candidate_eligible": True,
            "relation": "group_message",
        },
        {
            "context_url": "https://t.me/example/3",
            "rank_global": 3,
            "candidate_action_type": "event_recommendation_reply",
            "intent_set": "event_close_question",
            "text_snapshot": "Будет ли регистрация?",
            "pre_llm_candidate_eligible": True,
            "relation": "group_message",
        },
        {
            "context_url": "https://t.me/example/source",
            "rank_global": 4,
            "candidate_action_type": "trip_route_poi_recommendation",
            "intent_set": "route_poi_close_actionable",
            "text_snapshot": "Куда съездить?",
            "pre_llm_candidate_eligible": False,
            "relation": "vk_social_wall_post",
            "candidate_noise_type": "source_post_not_comment",
        },
    ]

    summaries = retrieval._surface_decision_summaries(
        [profile],
        eligible_rows_by_surface={"tg:example": rows[:3]},
        all_gate_rows_by_surface={"tg:example": [*rows, {"context_url": "x", "candidate_noise_type": "explicit_offer_or_ad"}]},
    )

    assert summaries[0]["recommendation"] in {"both_monitor_replies_and_ask_clarifications", "monitor_for_reply_opportunities"}
    assert summaries[0]["answerable_question_candidates"] == 3
    assert summaries[0]["ask_clarification_contexts"] == 2
    assert summaries[0]["source_post_contexts"] == 1
    assert summaries[0]["filtered_noise_contexts"] == 2


def test_shadow_payload_exposes_comment_semantic_retrieval_result(monkeypatch, tmp_path):
    runtime = load_runtime()
    surface = runtime._seed_surface("https://t.me/vKalinigrad_recomendations", platform="tg")
    profile = {
        "surface_key": surface["external_id"],
        "monitoring_decision_hint": "monitor",
        "monitoring_reason": "route questions",
        "dominant_detected_interests": ["route_poi"],
        "llm_budget_recommendation": {"send_top_comments_to_llm": 1},
    }
    result = {
        "summary": {"stage": "acq_comment_semantic_retrieval.v1", "comments_embedded": 10},
        "surface_profiles": [profile],
        "llm_gate_candidates": [{"context_url": "https://t.me/vKalinigrad_recomendations/11"}],
        "artifacts": {"manual_review_xlsx": str(tmp_path / "manual.xlsx")},
    }

    runtime._attach_comment_semantic_profiles([surface], result)
    payload = runtime.build_shadow_payload(scanned_surfaces=[surface], scanned_opportunities=[], comment_retrieval_result=result)

    by_external = {s["external_id"]: s for s in payload["surfaces"]}
    enriched = by_external[surface["external_id"]]
    assert enriched["monitoring_decision_hint"] == "monitor"
    assert enriched["risk"]["comment_semantic_retrieval"]["dominant_detected_interests"] == ["route_poi"]
    assert payload["comment_semantic_retrieval"]["artifacts"]["manual_review_xlsx"].endswith("manual.xlsx")
    assert payload["stats"]["comment_semantic_retrieval"]["summary"]["comments_embedded"] == 10


def test_runtime_builds_route_opportunity_from_retrieval_candidate():
    runtime = load_runtime()
    candidate = {
        "platform": "tg",
        "surface_key": "tg:vKalinigrad_recomendations",
        "context_url": "https://t.me/vKalinigrad_recomendations/11",
        "comment_id": "11",
        "text_snapshot": "Куда съездить из Калининграда на один день?",
        "candidate_action_type": "trip_route_poi_recommendation",
        "intent_set": "route_poi_close_actionable",
        "score": 0.42,
        "positive_score": 0.61,
        "negative_score": 0.19,
        "top_intent_phrase": "посоветуйте маршрут на один день из Калининграда",
        "region_confidence": "confirmed",
        "region_gate_status": "region_ok",
        "region_evidence_ru": "есть явная зацепка Калининградской области",
        "rank_global": 1,
        "rank_within_surface": 1,
        "target_hint": {"route_target_status": "route_needed", "destination_hint": "", "event_ids": []},
    }

    opp = runtime._build_opportunity_from_retrieval_candidate(candidate, default_target_url="https://t.me/kenigevents")

    assert opp["action_type"] == "trip_route_poi_recommendation"
    assert opp["target_hint"]["route_target_status"] == "route_needed"
    assert opp["link_target"]["kind"] == "route_needed"
    assert opp["evidence"]["semantic_retrieval"]["rank_global"] == 1
    assert opp["scores"]["source"] == "comment_semantic_retrieval"


def test_comment_retrieval_llm_gate_reviews_only_retrieved_candidates(monkeypatch):
    runtime = load_runtime()
    surface = runtime._seed_surface("https://t.me/vKalinigrad_recomendations", platform="tg")
    calls = []

    def fake_review(opp, _surface, _diagnostics):
        calls.append(opp["context_url"])
        return opp

    monkeypatch.setattr(runtime, "_llm_review_opportunity_sync", fake_review)
    scanned_opportunities = []
    result = {
        "llm_gate_candidates": [
            {
                "platform": "tg",
                "surface_key": surface["external_id"],
                "context_url": "https://t.me/vKalinigrad_recomendations/11",
                "comment_id": "11",
                "text_snapshot": "Куда съездить за день?",
                "candidate_action_type": "trip_route_poi_recommendation",
                "intent_set": "route_poi_close_actionable",
                "score": 0.7,
                "region_confidence": "confirmed",
                "target_hint": {"route_target_status": "route_needed"},
            },
            {
                # Same context must be skipped before spending a second LLM call.
                "platform": "tg",
                "surface_key": surface["external_id"],
                "context_url": "https://t.me/vKalinigrad_recomendations/11",
                "comment_id": "11",
                "text_snapshot": "Куда съездить за день?",
                "candidate_action_type": "event_recommendation_reply",
                "intent_set": "event_far_context",
                "score": 0.6,
                "region_confidence": "confirmed",
            },
        ],
    }

    runtime._run_comment_retrieval_llm_gate(
        result,
        surfaces_by_external={surface["external_id"]: surface},
        scanned_opportunities=scanned_opportunities,
        diagnostics=[],
    )

    assert calls == ["https://t.me/vKalinigrad_recomendations/11"]
    assert len(scanned_opportunities) == 1


def test_vk_scan_with_semantic_retrieval_collects_comments_without_per_comment_llm(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_ENABLE_COMMENT_SEMANTIC_RETRIEVAL", "1")
    monkeypatch.setattr(runtime, "_semantic_retrieval_enabled_from_module", lambda: True)
    monkeypatch.setattr(runtime, "_vk_token_lanes", lambda: [("TEST_TOKEN", "token")])
    monkeypatch.setattr(runtime, "_human_pause_sync", lambda *args, **kwargs: None)
    monkeypatch.setenv("ACQ_MAX_VK_BOARD_TOPICS_PER_SURFACE", "0")

    def fake_vk(method, *, token_lanes, params):
        if method == "wall.get":
            return {
                "items": [
                    {
                        "owner_id": -42481124,
                        "id": 5,
                        "date": 1782900000,
                        "text": "Подскажите, куда съездить из Калининграда на день?",
                        "comments": {"count": 1},
                    }
                ]
            }, "TEST_TOKEN"
        if method == "wall.getComments":
            return {
                "items": [
                    {"id": 9, "from_id": 101, "date": 1782900000, "text": "Куда поехать на электричке в область?"}
                ]
            }, "TEST_TOKEN"
        if method == "groups.getById":
            return {"groups": [{"id": 42481124, "name": "Подслушано", "screen_name": "club42481124", "members_count": 999}]}, "TEST_TOKEN"
        raise AssertionError(method)

    monkeypatch.setattr(runtime, "_vk_api_with_fallback", fake_vk)
    monkeypatch.setattr(runtime, "_llm_review_opportunity_sync", lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("per-comment LLM must be skipped")))
    records = []

    surfaces, opportunities, _diagnostics = runtime.scan_vk_shadow_surfaces(
        ["https://vk.com/club42481124"],
        ["https://vk.com/club42481124"],
        comment_records=records,
    )

    assert opportunities == []
    assert surfaces[0]["status"] == "comments_available"
    assert len(records) == 2
    assert {r["relation"] for r in records} == {"vk_social_wall_post", "vk_comment"}
    assert {r["author_id"] for r in records} == {"-42481124", "101"}
    assert surfaces[0]["title"] == "Подслушано"
    assert surfaces[0]["reach"]["members"] == 999


def test_vk_scan_can_read_active_personal_profile_wall(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_ENABLE_COMMENT_SEMANTIC_RETRIEVAL", "1")
    monkeypatch.setattr(runtime, "_semantic_retrieval_enabled_from_module", lambda: True)
    monkeypatch.setattr(runtime, "_vk_token_lanes", lambda: [("TEST_TOKEN", "token")])
    monkeypatch.setattr(runtime, "_human_pause_sync", lambda *args, **kwargs: None)
    monkeypatch.setenv("ACQ_MAX_VK_BOARD_TOPICS_PER_SURFACE", "1")

    calls = []

    def fake_vk(method, *, token_lanes, params):
        calls.append(method)
        if method == "users.get":
            return [{"id": 6786438, "first_name": "Иван", "last_name": "Иванов", "followers_count": 321}], "TEST_TOKEN"
        if method == "wall.get":
            return {
                "items": [
                    {
                        "owner_id": 6786438,
                        "from_id": 6786438,
                        "id": 5,
                        "date": 1782900000,
                        "text": "Пост про выходные",
                        "comments": {"count": 1},
                    }
                ]
            }, "TEST_TOKEN"
        if method == "wall.getComments":
            return {
                "items": [
                    {"id": 9, "from_id": 101, "date": 1782900000, "text": "Куда сходить рядом?"}
                ]
            }, "TEST_TOKEN"
        raise AssertionError(method)

    monkeypatch.setattr(runtime, "_vk_api_with_fallback", fake_vk)
    records = []

    surfaces, opportunities, diagnostics = runtime.scan_vk_shadow_surfaces(
        ["https://vk.com/id6786438"],
        ["https://vk.com/id6786438"],
        comment_records=records,
    )

    assert opportunities == []
    assert surfaces[0]["surface_type"] == "profile"
    assert surfaces[0]["title"] == "Иван Иванов"
    assert surfaces[0]["reach"]["followers"] == 321
    assert surfaces[0]["status"] == "comments_available"
    assert {r["relation"] for r in records} == {"vk_comment", "vk_social_wall_post"}
    assert any(r["is_post"] for r in records)
    assert "board.getTopics" not in calls
    assert any("vk id6786438: wall.get ok" in item for item in diagnostics)


def test_build_vk_opportunity_is_read_only_review_payload():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://vk.com/test_public", platform="vk")
    comment = {"id": 7, "date": 1782900000, "text": "Где найти афишу на выходные?", "attachments": [{"type": "sticker"}]}
    opp = runtime.build_vk_opportunity(surface, owner_id=-123, post_id=55, comment=comment, default_target_url="https://t.me/kenigevents")
    assert opp is not None
    assert opp["platform"] == "vk"
    assert opp["context_url"] == "https://vk.com/wall-123_55?reply=7"
    assert opp["sticker_observation"]["fit"] == "possible"
    assert opp["scores"]["source"] == "deterministic_shadow_prefilter"


def test_vk_social_wall_post_can_be_reply_opportunity_but_official_post_cannot():
    runtime = load_runtime()
    social_surface = runtime._seed_surface("https://vk.com/club42481124", platform="vk")
    official_surface = runtime._seed_surface("https://vk.com/shaman_kaliningrad", platform="vk")
    post = {
        "owner_id": -42481124,
        "id": 101,
        "date": 1782900000,
        "text": "Подскажите, куда съездить на один день из Калининграда с детьми?",
        "comments": {"count": 3},
        "views": {"count": 1200},
    }

    opp = runtime.build_vk_wall_post_opportunity(social_surface, post=post, default_target_url="https://t.me/kenigevents")

    assert opp is not None
    assert opp["context_url"] == "https://vk.com/wall-42481124_101"
    assert opp["topic_cluster"] == "trip_route_recommendation"
    assert opp["evidence"]["relation"] == "vk_social_wall_post"
    assert runtime.build_vk_wall_post_opportunity(official_surface, post=post, default_target_url="https://t.me/kenigevents") is None


def test_vk_board_comment_can_be_reply_opportunity():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://vk.com/kidsreview_kaliningrad", platform="vk")
    topic = {"id": 55, "title": "Вопросы родителей", "comments": 12}
    comment = {
        "id": 77,
        "date": 1782900000,
        "text": "Подскажите, что посмотреть с детьми за день в Калининградской области?",
    }

    opp = runtime.build_vk_board_opportunity(
        surface,
        group_id=56154842,
        topic=topic,
        comment=comment,
        default_target_url="https://t.me/kenigevents",
    )

    assert opp is not None
    assert opp["context_url"] == "https://vk.com/topic-56154842_55?post=77"
    assert opp["topic_cluster"] == "trip_route_recommendation"
    assert opp["evidence"]["relation"] == "vk_board_comment"


def test_vk_api_guard_rejects_write_method():
    runtime = load_runtime()
    try:
        runtime._vk_api("wall.createComment", token="x", params={})
    except RuntimeError as exc:
        assert "forbidden VK method" in str(exc)
    else:
        raise AssertionError("VK write method was not rejected")


def test_vk_allowlist_without_token_is_safe_seed_only(monkeypatch):
    runtime = load_runtime()
    monkeypatch.delenv("VK_ACCESS_TOKEN", raising=False)
    monkeypatch.delenv("VK_ACCESS_TOKEN4", raising=False)
    surfaces, opportunities, diagnostics = runtime.scan_vk_shadow_surfaces(["https://vk.com/test_public"], ["https://vk.com/test_public"])
    assert surfaces == []
    assert opportunities == []
    assert diagnostics and "token is not configured" in diagnostics[0]


def test_vk_scan_marks_comments_available_surface(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setattr(runtime, "_vk_token_lanes", lambda: [("TEST_TOKEN", "token")])
    monkeypatch.setattr(runtime, "_human_pause_sync", lambda *args, **kwargs: None)
    monkeypatch.setenv("ACQ_MAX_VK_BOARD_TOPICS_PER_SURFACE", "1")
    monkeypatch.setenv("ACQ_MAX_VK_BOARD_COMMENTS_PER_TOPIC", "2")

    def fake_vk(method, *, token_lanes, params):
        if method == "wall.get":
            return {"items": [{"owner_id": -123, "id": 5, "text": "Отчёт", "comments": {"count": 1}}]}, "TEST_TOKEN"
        if method == "wall.getComments":
            return {"items": [{"id": 9, "date": 1782900000, "text": "Спасибо"}]}, "TEST_TOKEN"
        if method == "board.getTopics":
            return {"items": []}, "TEST_TOKEN"
        raise AssertionError(method)

    monkeypatch.setattr(runtime, "_vk_api_with_fallback", fake_vk)

    surfaces, opportunities, diagnostics = runtime.scan_vk_shadow_surfaces(
        ["https://vk.com/club123"],
        ["https://vk.com/club123"],
    )

    assert opportunities == []
    assert diagnostics[0] == "vk club123: wall.get ok via TEST_TOKEN"
    assert surfaces[0]["status"] == "comments_available"
    assert surfaces[0]["scan_state"] == "comments_available"
    assert surfaces[0]["reach"]["wall_comments_seen"] == 1
    assert surfaces[0]["risk"]["reply_policy"] == "candidate_comment_surface"


def test_vk_scan_marks_inaccessible_surface(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setattr(runtime, "_vk_token_lanes", lambda: [("TEST_TOKEN", "token")])
    monkeypatch.setattr(runtime, "_human_pause_sync", lambda *args, **kwargs: None)

    def fake_vk(method, *, token_lanes, params):
        raise RuntimeError(f"{method} access denied")

    monkeypatch.setattr(runtime, "_vk_api_with_fallback", fake_vk)

    surfaces, opportunities, diagnostics = runtime.scan_vk_shadow_surfaces(
        ["https://vk.com/club123"],
        ["https://vk.com/club123"],
    )

    assert opportunities == []
    assert any("wall.get failed" in item for item in diagnostics)
    assert surfaces[0]["status"] == "rejected_inaccessible"
    assert surfaces[0]["scan_state"] == "checked_inaccessible"
    assert surfaces[0]["risk"]["reply_policy"] == "no_reply_surface"


def test_vk_scan_marks_no_comments_surface(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setattr(runtime, "_vk_token_lanes", lambda: [("TEST_TOKEN", "token")])
    monkeypatch.setattr(runtime, "_human_pause_sync", lambda *args, **kwargs: None)
    monkeypatch.setenv("ACQ_MAX_VK_BOARD_TOPICS_PER_SURFACE", "1")

    def fake_vk(method, *, token_lanes, params):
        if method == "wall.get":
            return {"items": [{"owner_id": -123, "id": 5, "text": "Пост", "comments": {"count": 0}}]}, "TEST_TOKEN"
        if method == "board.getTopics":
            return {"items": []}, "TEST_TOKEN"
        raise AssertionError(method)

    monkeypatch.setattr(runtime, "_vk_api_with_fallback", fake_vk)

    surfaces, opportunities, diagnostics = runtime.scan_vk_shadow_surfaces(
        ["https://vk.com/club123"],
        ["https://vk.com/club123"],
    )

    assert opportunities == []
    assert surfaces[0]["status"] == "rejected_no_comments"
    assert surfaces[0]["scan_state"] == "checked_no_public_comments_or_boards"
    assert surfaces[0]["risk"]["reply_policy"] == "no_reply_surface"


def test_vk_board_comments_count_is_capped_to_vk_api_limit(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setattr(runtime, "_vk_token_lanes", lambda: [("TEST_TOKEN", "token")])
    monkeypatch.setattr(runtime, "_human_pause_sync", lambda *args, **kwargs: None)
    monkeypatch.setenv("ACQ_MAX_VK_POSTS_PER_SURFACE", "0")
    monkeypatch.setenv("ACQ_MAX_VK_BOARD_TOPICS_PER_SURFACE", "1")
    monkeypatch.setenv("ACQ_MAX_VK_BOARD_COMMENTS_PER_TOPIC", "200")
    counts: list[int] = []

    def fake_vk(method, *, token_lanes, params):
        if method == "wall.get":
            return {"items": []}, "TEST_TOKEN"
        if method == "board.getTopics":
            return {"items": [{"id": 55, "title": "Вопросы"}]}, "TEST_TOKEN"
        if method == "board.getComments":
            counts.append(params["count"])
            return {"items": []}, "TEST_TOKEN"
        raise AssertionError(method)

    monkeypatch.setattr(runtime, "_vk_api_with_fallback", fake_vk)

    runtime.scan_vk_shadow_surfaces(["https://vk.com/club123"], ["https://vk.com/club123"])

    assert counts == [100]



def test_vk_token_lanes_prefer_monitoring_token_without_leaking_values(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("VK_ACCESS_TOKEN", "generic-token")
    monkeypatch.setenv("VK_ACCESS_TOKEN4", "monitoring-token")

    lanes = runtime._vk_token_lanes()

    assert [name for name, _token in lanes] == ["VK_ACCESS_TOKEN4", "VK_ACCESS_TOKEN"]
    assert [token for _name, token in lanes] == ["monitoring-token", "generic-token"]


def test_tg_frontier_queue_dedupes_and_respects_limit():
    runtime = load_runtime()
    queue = []
    queued = set()

    assert runtime._enqueue_tg_url(queue, queued, "https://t.me/first", limit=2) is True
    assert runtime._enqueue_tg_url(queue, queued, "https://t.me/first/", limit=2) is False
    assert runtime._enqueue_tg_url(queue, queued, "https://t.me/second", limit=2) is True
    assert runtime._enqueue_tg_url(queue, queued, "https://t.me/third", limit=2) is False

    assert queue == ["https://t.me/first", "https://t.me/second"]


def test_runtime_deadline_helpers_stop_after_budget(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_RUNTIME_DEADLINE_SECONDS", "1")

    deadline = runtime._deadline_after_seconds()

    assert isinstance(deadline, float)
    assert runtime._deadline_reached(deadline - 10) is True
    assert runtime._deadline_reached(None) is False


def test_seen_context_is_skipped_before_spending_gemma(monkeypatch):
    runtime = load_runtime()
    runtime.LLM_GATE_STATS["skipped_seen_context"] = 0
    opp = {"platform": "tg", "context_url": "https://t.me/example/1", "context_text_snippet": "Куда сходить?"}
    keys = set()
    diagnostics = []

    assert runtime._should_skip_opportunity_before_llm(
        opp,
        seen_contexts={"https://t.me/example/1"},
        opportunity_keys=keys,
        diagnostics=diagnostics,
    ) is True

    assert runtime.LLM_GATE_STATS["skipped_seen_context"] == 1
    assert not keys
    assert "before Gemma" in diagnostics[0]


def test_llm_gate_has_visible_per_run_budget(monkeypatch):
    runtime = load_runtime()
    runtime.LLM_GATE_STATS["calls"] = 0
    runtime.LLM_GATE_STATS["reserved"] = 0
    runtime.LLM_GATE_STATS["blocked_rate_limit"] = 0
    runtime.LLM_GATE_STATS["estimated_input_tokens"] = 0
    monkeypatch.setenv("ACQ_MAX_LLM_CALLS_PER_RUN", "1")

    diagnostics = []
    assert runtime._reserve_llm_gate_call("Куда сходить?", diagnostics) is True
    runtime.LLM_GATE_STATS["calls"] = 1
    assert runtime._reserve_llm_gate_call("Куда сходить?", diagnostics) is False

    snapshot = runtime._llm_limit_snapshot()
    assert snapshot["max_calls_per_run"] == 1
    assert snapshot["calls_used_this_run"] == 1
    assert snapshot["blocked_rate_limit"] == 1
    assert snapshot["estimated_input_tokens_this_run"] > 0


def test_kaggle_config_overrides_stale_acq_env(monkeypatch, tmp_path):
    runtime = load_runtime()
    config_path = tmp_path / "config.json"
    config_path.write_text(json.dumps({"ACQ_ENABLE_LIVE_TG_SCAN": "1"}), encoding="utf-8")

    monkeypatch.setenv("ACQ_ENABLE_LIVE_TG_SCAN", "0")
    monkeypatch.setattr(runtime, "_find_input_file", lambda name: config_path if name == "config.json" else None)

    loaded = runtime._load_kaggle_env()
    assert loaded["config_loaded"] is True
    assert runtime.os.environ["ACQ_ENABLE_LIVE_TG_SCAN"] == "1"


def test_telega_in_kaliningrad_seeds_are_tagged():
    runtime = load_runtime()
    assert "https://t.me/anons39" in runtime.DEFAULT_TG_SEEDS
    assert "https://t.me/vKalinigrad_recomendations" in runtime.DEFAULT_TG_SEEDS
    surface = runtime._seed_surface("https://t.me/anons39", platform="tg")
    assert surface["source"] == "telega_in"
    assert surface["status"] == "needs_comment_resolve"
    assert "Telega.in" in surface["topic_hint"]


def test_llm_gate_rejects_low_confidence_acceptance(monkeypatch):
    runtime = load_runtime()
    surface = runtime._seed_surface("https://t.me/example", platform="tg")
    opp = {
        "platform": "tg",
        "surface_external_id": surface["external_id"],
        "context_url": "https://t.me/example/44",
        "context_text_snippet": "Куда сходить?",
        "matched_intent": "event_recommendation_question",
        "topic_cluster": "local_events",
        "link_target": {"kind": "pka_channel", "label": "ПКА", "url": "https://t.me/kenigevents"},
        "scores": {"relevance": 0.5, "spam_risk": "low", "safety_risk": "low"},
    }
    monkeypatch.setattr(runtime, "_google_api_key", lambda: ("test-key", "GOOGLE_API_KEY3"))
    monkeypatch.setattr(runtime, "_call_acq_llm_gate_sync", lambda _opp, _surface: {
        "is_candidate": True,
        "matched_intent": "event_recommendation_question",
        "topic_cluster": "local_events",
        "best_reply_strategy": "unclear",
        "target_kind": "pka_channel",
        "target_label": "ПКА",
        "target_url": "https://t.me/kenigevents",
        "reason": "weak",
        "relevance": 0.4,
        "spam_risk": "low",
        "safety_risk": "low",
        "checklist": [{"id": "need", "question": "Есть потребность?", "answer": True, "note": "да"}],
    })

    diagnostics = []
    assert runtime._llm_review_opportunity_sync(opp, surface, diagnostics) is None
    assert runtime.LLM_GATE_STATS["rejected_low_confidence"] == 1
    assert "relevance" in diagnostics[0]


def test_llm_gate_json_salvage_and_fallback_needs_human_review(monkeypatch):
    runtime = load_runtime()
    parsed = runtime._parse_llm_json('prefix {"is_candidate": false, "reason": "нет"} suffix')
    assert parsed["is_candidate"] is False
    assert runtime.LLM_GATE_STATS["json_salvaged"] >= 1

    monkeypatch.setattr(runtime, "_google_api_key", lambda: ("test-key", "GOOGLE_API_KEY_TEST"))
    monkeypatch.setattr(runtime, "_reserve_llm_gate_call", lambda _prompt, _diagnostics: True)
    monkeypatch.setattr(runtime, "_call_acq_llm_gate_sync", lambda _opp, _surface: (_ for _ in ()).throw(RuntimeError("boom")))
    diagnostics = []
    reviewed = runtime._llm_review_opportunity_sync(
        {
            "platform": "tg",
            "context_url": "https://t.me/test/1",
            "context_text_snippet": "Посоветуйте куда сходить",
            "matched_intent": "event_or_route_recommendation_reply",
            "topic_cluster": "reply_to_user_event_or_route",
            "scores": {"relevance": 0.9},
            "evidence": {"semantic_retrieval": {"selection_source": "bge_m3_dense"}},
        },
        {"title": "Тест", "url": "https://t.me/test"},
        diagnostics,
    )

    assert reviewed["review_status"] == "needs_human_review"
    assert reviewed["evidence"]["llm_gate"]["status"] == "needs_human_review"


def test_llm_gate_prompt_rejects_event_local_schedule_logistics():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://vk.com/vagonka39", platform="vk")
    prompt = runtime._llm_gate_prompt({"platform": "vk", "context_text_snippet": "А где афиша 1 дня ?", "matched_intent": "event_recommendation_question", "topic_cluster": "local_events", "link_target": {}}, surface)
    assert "афиша/программа/расписание 1 дня" in prompt
    assert "локальная логистика текущего события" in prompt
    assert "общий канал" in prompt
    assert "льготы/скидки/билеты/условия/доступность" in prompt
    assert "venue policy" in prompt


def test_llm_gate_rejects_post_event_praise(monkeypatch):
    runtime = load_runtime()
    surface = runtime._seed_surface("https://vk.com/example", platform="vk")
    opp = {
        "platform": "vk",
        "surface_external_id": surface["external_id"],
        "context_url": "https://vk.com/wall-1_2?reply=3",
        "context_text_snippet": "Погода была отличная, компания веселая. Спасибо организаторам за прекрасно проведенное время❤",
        "matched_intent": "organizer_partnership",
        "topic_cluster": "organizer_partnership",
        "link_target": {"kind": "topic_landing", "label": "Добавить событие", "url": "https://kenigevents.ru/partnerstvo/"},
        "scores": {"relevance": 0.6, "spam_risk": "low", "safety_risk": "low"},
    }
    monkeypatch.setattr(runtime, "_google_api_key", lambda: ("test-key", "GOOGLE_API_KEY3"))
    monkeypatch.setattr(runtime, "_call_acq_llm_gate_sync", lambda _opp, _surface: {
        "is_candidate": False,
        "matched_intent": "none",
        "topic_cluster": "none",
        "best_reply_strategy": "no_reply",
        "target_kind": "none",
        "target_label": "",
        "target_url": "",
        "reason": "Постфактум-отзыв и благодарность, нет вопроса или полезного acquisition-ответа.",
        "relevance": 0.0,
        "spam_risk": "low",
        "safety_risk": "low",
        "checklist": [{"id": "need", "question": "Есть потребность?", "answer": False, "note": "только благодарность"}],
    })

    diagnostics = []
    assert runtime._llm_review_opportunity_sync(opp, surface, diagnostics) is None
    assert runtime.LLM_GATE_STATS["rejected"] == 1
    assert "rejected" in diagnostics[0]


def test_llm_gate_accepts_and_persists_checklist(monkeypatch):
    runtime = load_runtime()
    surface = runtime._seed_surface("https://t.me/example", platform="tg")
    opp = runtime.build_opportunity_from_message(
        surface,
        SimpleNamespace(id=44, message="Куда сходить с ребёнком на выходных в Калининграде?", date=datetime(2026, 7, 1, tzinfo=timezone.utc)),
        default_target_url="https://t.me/kenigevents",
    )
    assert opp is not None
    monkeypatch.setattr(runtime, "_google_api_key", lambda: ("test-key", "GOOGLE_API_KEY3"))
    monkeypatch.setattr(runtime, "_call_acq_llm_gate_sync", lambda _opp, _surface: {
        "is_candidate": True,
        "matched_intent": "event_recommendation_request",
        "topic_cluster": "kids_weekend_events",
        "best_reply_strategy": "short_native_reply_with_one_relevant_link",
        "target_kind": "pka_channel",
        "target_label": "Полюбить Калининград Анонсы",
        "target_url": "https://t.me/kenigevents",
        "reason": "Есть явный запрос куда сходить с ребёнком на выходных.",
        "relevance": 0.92,
        "spam_risk": "low",
        "safety_risk": "low",
        "checklist": [
            {"id": "question", "question": "Есть вопрос?", "answer": True, "note": "куда сходить"},
            {"id": "future_need", "question": "Будущая потребность?", "answer": True, "note": "на выходных"},
        ],
    })

    accepted = runtime._llm_review_opportunity_sync(opp, surface, [])

    assert accepted is not None
    assert accepted["scores"]["source"] == "gemma4_acquisition_gate"
    assert accepted["scores"]["relevance"] == 0.92
    assert accepted["evidence"]["llm_gate"]["model"].startswith("models/gemma-4")
    assert accepted["evidence"]["llm_gate"]["checklist"][0]["answer"] is True


def test_kaggle_runtime_static_no_external_write_calls():
    source = Path("kaggle/SubscriberAcquisitionDiscovery/subscriber_acquisition_discovery.py").read_text(encoding="utf-8")
    forbidden_snippets = [
        ".send_message(",
        ".send_file(",
        ".send_reaction(",
        "JoinChannelRequest",
        "ImportChatInviteRequest",
        "messages.send",
        "wall.post",
        "wall.createComment",
        "stories.",
    ]
    for snippet in forbidden_snippets:
        assert snippet not in source


def test_review_module_only_sends_to_configured_review_chat_static():
    source = Path("subscriber_acquisition/review.py").read_text(encoding="utf-8")
    assert "bot.send_message" in source
    assert "cfg.review_chat_id" in source
    assert "ensure_review_chat(sent_chat_id" in source


def test_kaggle_runtime_env_allowlists_vk_monitoring_seeds_for_discovery():
    from subscriber_acquisition.config import AcqConfig
    from subscriber_acquisition.kaggle_runner import _runtime_env_from_config

    payload = {"surfaces": [
        {"platform": "vk", "url": "https://vk.com/club1", "external_id": "vk:club1", "status": "candidate", "source": "vk_source"},
        {"platform": "vk", "url": "https://vk.com/club2", "external_id": "vk:club2", "status": "candidate", "source": "vk_source"},
    ]}
    env = _runtime_env_from_config(AcqConfig(), payload)
    assert env["ACQ_VK_SEEDS_JSON"] == '["https://vk.com/club1", "https://vk.com/club2"]'
    assert env["ACQ_VK_ALLOWLIST_JSON"] == env["ACQ_VK_SEEDS_JSON"]
    assert env["ACQ_ENABLE_LLM_GATE"] == "1"
    assert env["ACQ_GOOGLE_KEY_ENV"] == "GOOGLE_API_KEY3"
    assert env["ACQ_LLM_MODEL"].startswith("models/gemma-4")
    assert env["ACQ_LLM_GATE_MIN_RELEVANCE"]
    assert env["ACQ_RUNTIME_DEADLINE_SECONDS"]
    assert env["ACQ_MAX_LLM_CALLS_PER_RUN"]
    assert "ACQ_TG_SEARCH_MESSAGES_PER_QUERY" in env
    assert "ACQ_MAX_VK_SURFACES_PER_RUN" in env
    assert "ACQ_MAX_VK_POSTS_PER_SURFACE" in env
    assert "ACQ_MAX_VK_COMMENTS_PER_POST" in env
    assert "ACQ_ENABLE_COMMENT_SEMANTIC_RETRIEVAL" in env
    assert "ACQ_COMMENT_RETRIEVAL_MODELS_JSON" in env
    assert json.loads(env["ACQ_COMMENT_RETRIEVAL_MODELS_JSON"]) == ["intfloat/multilingual-e5-base", "BAAI/bge-m3"]
    assert env["ACQ_COMMENT_RETRIEVAL_MAX_LLM_CANDIDATES"] == "24"
    assert env["ACQ_ENABLE_VK_MEMBER_PROFILE_DISCOVERY"] == "1"


def test_kaggle_runtime_env_forces_two_semantic_models(monkeypatch):
    from subscriber_acquisition.config import AcqConfig
    from subscriber_acquisition.kaggle_runner import _runtime_env_from_config

    monkeypatch.setenv("ACQ_COMMENT_RETRIEVAL_MODELS_JSON", '["intfloat/multilingual-e5-base"]')

    env = _runtime_env_from_config(AcqConfig(), {"surfaces": []})

    assert json.loads(env["ACQ_COMMENT_RETRIEVAL_MODELS_JSON"]) == ["intfloat/multilingual-e5-base", "BAAI/bge-m3"]


def test_kaggle_runtime_env_allows_explicit_vk_profile_wall_seed():
    from subscriber_acquisition.config import AcqConfig
    from subscriber_acquisition.kaggle_runner import _runtime_env_from_config

    payload = {"surfaces": [
        {"platform": "vk", "url": "https://vk.com/id123", "external_id": "vk:id123", "surface_type": "profile", "status": "candidate", "source": "discovered_vk_author"},
    ]}

    env = _runtime_env_from_config(AcqConfig(), payload)

    assert json.loads(env["ACQ_VK_SEEDS_JSON"]) == ["https://vk.com/id123"]
    assert json.loads(env["ACQ_VK_ALLOWLIST_JSON"]) == ["https://vk.com/id123"]


def test_vk_member_profile_discovery_adds_bounded_profile_candidates(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("KAGGLE_RUN_ID", "member-run")
    monkeypatch.setenv("ACQ_ENABLE_VK_MEMBER_PROFILE_DISCOVERY", "1")
    monkeypatch.setenv("ACQ_MAX_VK_MEMBER_PROFILES_DISCOVERED_PER_RUN", "2")
    monkeypatch.setenv("ACQ_MAX_VK_MEMBER_PROFILES_PER_GROUP", "3")
    runtime.VK_SCAN_STATS["profile_surfaces_discovered_from_members"] = 0

    def fake_vk(method, *, token_lanes, params):
        assert method == "groups.getMembers"
        assert params["group_id"] == "club123"
        return {"items": [{"id": 10}, {"id": 11}, {"id": 12}]}, "TEST_TOKEN"

    monkeypatch.setattr(runtime, "_vk_api_with_fallback", fake_vk)
    surfaces = {}
    diagnostics = []

    runtime._discover_vk_member_profile_surfaces(
        surfaces,
        domain="club123",
        token_lanes=[("TEST_TOKEN", "token")],
        diagnostics=diagnostics,
    )

    assert list(surfaces) == ["vk:id10", "vk:id11"]
    assert surfaces["vk:id10"]["surface_type"] == "profile"
    assert surfaces["vk:id10"]["discovered_in_run_id"] == "member-run"
    assert runtime.VK_SCAN_STATS["profile_surfaces_discovered_from_members"] == 2
    assert "discovered 2 member profile-wall candidates" in diagnostics[0]


def test_vk_profile_metadata_enrichment_names_discovered_profiles(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_MAX_VK_PROFILE_NAME_ENRICH_PER_RUN", "10")
    runtime.VK_SCAN_STATS["profile_surfaces_metadata_enriched"] = 0
    surface = runtime._profile_surface_from_user_id(10, source="discovered_vk_author", context="comment author")
    assert surface is not None
    surfaces = {str(surface["external_id"]): surface}

    def fake_vk(method, *, token_lanes, params):
        assert method == "users.get"
        assert params["user_ids"] == "id10"
        return [{"id": 10, "first_name": "Иван", "last_name": "Иванов", "screen_name": "ivanov", "followers_count": 42}], "TEST_TOKEN"

    monkeypatch.setattr(runtime, "_vk_api_with_fallback", fake_vk)
    diagnostics = []

    runtime._enrich_vk_discovered_profile_surfaces(
        surfaces,
        token_lanes=[("TEST_TOKEN", "token")],
        diagnostics=diagnostics,
    )

    assert surfaces["vk:id10"]["title"] == "Иван Иванов"
    assert surfaces["vk:id10"]["handle"] == "ivanov"
    assert surfaces["vk:id10"]["url"] == "https://vk.com/ivanov"
    assert surfaces["vk:id10"]["reach"]["followers"] == 42
    assert runtime.VK_SCAN_STATS["profile_surfaces_metadata_enriched"] == 1


def test_runtime_env_passes_tg_channel_resolve_budget():
    from subscriber_acquisition.config import AcqConfig
    from subscriber_acquisition.kaggle_runner import _runtime_env_from_config

    env = _runtime_env_from_config(AcqConfig(max_surfaces_per_run=10), {"surfaces": [
        {"platform": "tg", "url": "https://t.me/channel", "external_id": "tg:channel", "status": "needs_comment_resolve"},
    ]})

    assert env["ACQ_TG_SEEDS_JSON"] == '["https://t.me/channel"]'
    assert "ACQ_MAX_TG_CHANNEL_RESOLVES_PER_RUN" in env
    assert "ACQ_MAX_TG_CHANNEL_POSTS_FOR_LINKS" in env
    assert "ACQ_TG_SEED_SURFACES_JSON" in env


def test_runtime_env_passes_private_telegram_access_metadata():
    from subscriber_acquisition.config import AcqConfig
    from subscriber_acquisition.kaggle_runner import _runtime_env_from_config

    env = _runtime_env_from_config(AcqConfig(), {"surfaces": [
        {
            "platform": "tg",
            "url": "https://t.me/c/1481648829",
            "handle": "1481648829",
            "external_id": "tg:1481648829",
            "surface_type": "linked_discussion",
            "source": "linked_discussion",
            "status": "candidate",
            "risk_json": {"_telegram_access": {"id": "1481648829", "access_hash": "123456789"}},
        },
    ]})

    seed_meta = json.loads(env["ACQ_TG_SEED_SURFACES_JSON"])
    assert seed_meta[0]["telegram_access"]["access_hash"] == "123456789"


def test_smartik_vk_seeds_are_configured():
    from subscriber_acquisition.kaggle_runner import SMARTIK_KALININGRAD_VK_SEEDS

    handles = {handle for handle, _title, _url in SMARTIK_KALININGRAD_VK_SEEDS}

    assert "club42481124" in handles
    assert "club31556867" in handles
    assert all(url.startswith("https://smartik.ru/kaliningrad/group/") for _handle, _title, url in SMARTIK_KALININGRAD_VK_SEEDS)


def test_vk_scanner_prefers_live_comment_threads_static():
    source = Path("kaggle/SubscriberAcquisitionDiscovery/subscriber_acquisition_discovery.py").read_text(encoding="utf-8")
    assert '"filter": "all"' in source
    assert '"sort": "desc"' in source
    assert "posts_without_comments_skipped" in source
    assert "rate_limit_backoffs" in source
    assert "board.getTopics" in source
    assert "board.getComments" in source


def test_kaggle_secrets_use_isolated_gemma_key_lane(monkeypatch):
    from subscriber_acquisition.kaggle_runner import _build_secrets_payload

    monkeypatch.setenv("GOOGLE_API_KEY3", "key3")
    monkeypatch.setenv("GOOGLE_API_KEY", "generic-key")
    monkeypatch.delenv("ACQ_ALLOW_GOOGLE_KEY_FALLBACKS", raising=False)

    payload = json.loads(_build_secrets_payload())

    assert payload["GOOGLE_API_KEY3"] == "key3"
    assert "GOOGLE_API_KEY" not in payload


def test_acquisition_discovery_auth_bundle_is_preferred_over_s22(monkeypatch):
    from subscriber_acquisition import kaggle_runner

    monkeypatch.setenv("ACQ_ENABLE_LIVE_TG_SCAN", "1")
    monkeypatch.setenv("TELEGRAM_AUTH_BUNDLE_DISCOVERY", "discovery-bundle")
    monkeypatch.setenv("TELEGRAM_AUTH_BUNDLE_S22", "s22-bundle")
    monkeypatch.setenv("TG_API_ID", "1")
    monkeypatch.setenv("TG_API_HASH", "hash")

    payload = json.loads(kaggle_runner._build_secrets_payload())

    assert kaggle_runner.discovery_remote_auth_scope() == "TELEGRAM_AUTH_BUNDLE_DISCOVERY"
    assert kaggle_runner._discovery_resource_lease_key() == "telegram_session:env:TELEGRAM_AUTH_BUNDLE_DISCOVERY"
    assert payload["TELEGRAM_AUTH_BUNDLE_DISCOVERY"] == "discovery-bundle"
    assert "TELEGRAM_AUTH_BUNDLE_S22" not in payload


def test_runtime_decode_tg_auth_prefers_discovery_bundle(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_TELEGRAM_AUTH_BUNDLE_ENV", "TELEGRAM_AUTH_BUNDLE_DISCOVERY")
    monkeypatch.setenv("TELEGRAM_AUTH_BUNDLE_DISCOVERY", _auth_bundle("discovery-session", device_model="Discovery phone"))
    monkeypatch.setenv("TELEGRAM_AUTH_BUNDLE_S22", _auth_bundle("s22-session", device_model="S22"))

    session, device = runtime._decode_tg_auth()

    assert session == "discovery-session"
    assert device["device_model"] == "Discovery phone"
    assert "auth_bundle_env" not in device


class _FakeSupabaseRpc:
    def __init__(self, data):
        self.data = data

    def execute(self):
        return SimpleNamespace(data=self.data)


class _FakeSupabaseLeaseClient:
    def __init__(self, acquire_data):
        self.acquire_data = acquire_data
        self.calls = []

    def rpc(self, name, payload):
        self.calls.append({"name": name, "payload": payload})
        if name == "runtime_resource_acquire":
            return _FakeSupabaseRpc(self.acquire_data)
        if name == "runtime_resource_release":
            return _FakeSupabaseRpc({"ok": True, "released": 1})
        raise AssertionError(name)


@pytest.mark.asyncio
async def test_discovery_supabase_session_lease_acquire_and_release(monkeypatch):
    from subscriber_acquisition import kaggle_runner

    client = _FakeSupabaseLeaseClient({"ok": True, "expires_at": "2026-07-05T12:00:00Z"})
    monkeypatch.setenv("ACQ_ENABLE_LIVE_TG_SCAN", "1")
    monkeypatch.setenv("TELEGRAM_AUTH_BUNDLE_DISCOVERY", "bundle")
    monkeypatch.setattr(kaggle_runner, "_get_supabase_client_for_shared_limits", lambda: client)

    lease = await kaggle_runner._acquire_supabase_session_lease(run_id="20260705-120000")
    await kaggle_runner._release_supabase_session_lease(lease, status="released")

    assert lease is not None
    assert lease.resource_key == "telegram_session:env:TELEGRAM_AUTH_BUNDLE_DISCOVERY"
    assert [call["name"] for call in client.calls] == ["runtime_resource_acquire", "runtime_resource_release"]
    assert client.calls[0]["payload"]["p_holder_id"] == "acq_discovery:20260705-120000"


@pytest.mark.asyncio
async def test_discovery_supabase_session_lease_blocks_busy_holder(monkeypatch):
    from subscriber_acquisition import kaggle_runner

    client = _FakeSupabaseLeaseClient({
        "ok": False,
        "holder_id": "acq_discovery:other",
        "expires_at": "2026-07-05T12:30:00Z",
    })
    monkeypatch.setenv("ACQ_ENABLE_LIVE_TG_SCAN", "1")
    monkeypatch.setenv("TELEGRAM_AUTH_BUNDLE_DISCOVERY", "bundle")
    monkeypatch.setattr(kaggle_runner, "_get_supabase_client_for_shared_limits", lambda: client)

    with pytest.raises(RuntimeError, match="Telegram discovery session is busy"):
        await kaggle_runner._acquire_supabase_session_lease(run_id="20260705-120000")


def test_runtime_env_passes_seen_context_urls():
    from subscriber_acquisition.config import AcqConfig
    from subscriber_acquisition.kaggle_runner import _runtime_env_from_config

    env = _runtime_env_from_config(AcqConfig(), {
        "surfaces": [],
        "seen_opportunities": [{"context_url": "https://t.me/example/1"}],
        "known_terminal_tg_handles": ["Anons39"],
    })

    assert env["ACQ_SEEN_CONTEXT_URLS_JSON"] == '["https://t.me/example/1"]'
    assert env["ACQ_KNOWN_TERMINAL_TG_HANDLES_JSON"] == '["anons39"]'


def test_remote_session_marker_cooldown_blocks_fast_reuse(monkeypatch, tmp_path):
    from subscriber_acquisition import kaggle_runner

    marker = tmp_path / "remote-session.json"
    monkeypatch.setenv("ACQ_ENABLE_LIVE_TG_SCAN", "1")
    monkeypatch.setenv("ACQ_REMOTE_SESSION_MARKER_PATH", str(marker))
    monkeypatch.setenv("ACQ_REMOTE_SESSION_COOLDOWN_SECONDS", "600")

    kaggle_runner._write_remote_session_marker(state="complete", run_id="run-1", kernel_ref="owner/kernel")
    remaining, payload = kaggle_runner._remote_session_cooldown_remaining()

    assert remaining > 0
    assert payload["run_id"] == "run-1"
    assert payload["state"] == "complete"


def test_discovery_auth_scope_none_when_tg_scan_disabled(monkeypatch):
    from subscriber_acquisition import kaggle_runner

    monkeypatch.setenv("ACQ_ENABLE_LIVE_TG_SCAN", "0")

    assert kaggle_runner.discovery_remote_auth_scope() == "none"


def test_seen_context_urls_env_parser(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_SEEN_CONTEXT_URLS_JSON", '["https://t.me/example/1"]')

    assert runtime._seen_context_urls() == {"https://t.me/example/1"}


def test_known_terminal_tg_handles_env_parser(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_KNOWN_TERMINAL_TG_HANDLES_JSON", '["Anons39", "kpkld"]')

    assert runtime._known_terminal_tg_handles() == {"anons39", "kpkld"}


def test_tg_seed_metadata_env_parser(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv(
        "ACQ_TG_SEED_SURFACES_JSON",
        '[{"url":"https://t.me/c/1481648829","handle":"1481648829","external_id":"tg:1481648829","surface_type":"linked_discussion","source":"linked_discussion"}]',
    )

    meta = runtime._tg_seed_metadata()

    assert meta["1481648829"]["surface_type"] == "linked_discussion"
    assert meta["https://t.me/c/1481648829"]["source"] == "linked_discussion"
    assert runtime._metadata_for_tg_seed("https://t.me/c/1481648829", "1481648829", meta)["external_id"] == "tg:1481648829"


def test_telegram_entity_ref_from_seed_uses_injected_input_peer_class():
    runtime = load_runtime()

    class DummyInputPeerChannel:
        def __init__(self, channel_id, access_hash):
            self.channel_id = channel_id
            self.access_hash = access_hash

    ref = runtime._telegram_entity_ref_from_seed(
        "1481648829",
        {"telegram_access": {"id": "1481648829", "access_hash": "5526881181816195856"}},
        input_peer_channel_cls=DummyInputPeerChannel,
    )

    assert isinstance(ref, DummyInputPeerChannel)
    assert ref.channel_id == 1481648829
    assert ref.access_hash == 5526881181816195856


def test_tg_comment_record_and_llm_prompt_preserve_context_chain():
    runtime = load_runtime()
    surface = runtime._seed_surface("https://t.me/example_chat", platform="tg")
    surface["surface_type"] = "linked_discussion"
    records = []

    runtime._collect_tg_comment_record(
        records,
        surface,
        SimpleNamespace(id=42, message="А детям можно?", date=datetime(2026, 7, 1, tzinfo=timezone.utc), reply_to=SimpleNamespace(reply_to_msg_id=10)),
        relation="linked_discussion",
        retrieval="latest",
        reply_parent_text="Можно ли с коляской?",
        source_post_text="Концерт в парке, начало в 18:00",
        source_post_id=7,
    )
    candidate = {
        **records[0],
        "model_name": "intfloat/multilingual-e5-base",
        "intent_set": "event_close_question",
        "score": 0.9,
        "candidate_action_type": "event_recommendation_reply",
        "region_confidence": "confirmed",
    }
    opp = runtime._build_opportunity_from_retrieval_candidate(candidate, default_target_url="https://t.me/kenigevents")
    prompt = runtime._llm_gate_prompt(opp, surface)

    assert records[0]["source_post_text_snapshot"].startswith("Концерт в парке")
    assert records[0]["reply_parent_text_snapshot"].startswith("Можно ли")
    assert opp["context_chain"]["source_post_text"].startswith("Концерт в парке")
    assert "source_post_text" in prompt
    assert "parent_comment_text" in prompt
    assert "current_comment_text" in prompt


def test_shadow_payload_applies_tg_seed_metadata_to_linked_discussion_rows(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_TG_SEEDS_JSON", '["https://t.me/c/1481648829"]')
    monkeypatch.setenv(
        "ACQ_TG_SEED_SURFACES_JSON",
        '[{"url":"https://t.me/c/1481648829","handle":"1481648829","external_id":"tg:1481648829","surface_type":"linked_discussion","source":"linked_discussion","title":"КП Chat","topic_hint":"linked discussion for kpkld"}]',
    )

    payload = runtime.build_shadow_payload(scanned_surfaces=[], scanned_opportunities=[])

    surface = payload["surfaces"][0]
    assert surface["external_id"] == "tg:1481648829"
    assert surface["surface_type"] == "linked_discussion"
    assert surface["source"] == "linked_discussion"
    assert surface["status"] == "candidate"
    assert surface["scan_state"] == "queued_waiting_replyable_budget"
    assert surface["title"] == "КП Chat"


def test_shadow_payload_applies_tg_seed_title_and_reach_without_scan(monkeypatch):
    runtime = load_runtime()
    monkeypatch.setenv("ACQ_TG_SEEDS_JSON", '["https://t.me/monitoring_public"]')
    monkeypatch.setenv(
        "ACQ_TG_SEED_SURFACES_JSON",
        json.dumps([
            {
                "url": "https://t.me/monitoring_public",
                "handle": "monitoring_public",
                "external_id": "tg:monitoring_public",
                "surface_type": "unknown_public",
                "source": "tg_monitoring",
                "title": "Городские анонсы",
                "topic_hint": "existing Telegram monitoring source",
                "reach": {"basis": "telegram_monitoring_source", "confidence": "low"},
            }
        ], ensure_ascii=False),
    )

    payload = runtime.build_shadow_payload(scanned_surfaces=[], scanned_opportunities=[])

    surface = payload["surfaces"][0]
    assert surface["title"] == "Городские анонсы"
    assert surface["source"] == "tg_monitoring"
    assert surface["reach"]["basis"] == "telegram_monitoring_source"


def test_tg_seed_metadata_is_initialized_inside_telegram_scan():
    source = Path("kaggle/SubscriberAcquisitionDiscovery/subscriber_acquisition_discovery.py").read_text(encoding="utf-8")
    telegram_body = source.split("async def scan_telegram_shadow_surfaces", 1)[1].split("def _seen_context_urls", 1)[0]
    vk_body = source.split("def scan_vk_shadow_surfaces", 1)[1].split("def _load_status_loader", 1)[0]

    assert "tg_seed_meta = _tg_seed_metadata()" in telegram_body
    assert "tg_seed_meta = _tg_seed_metadata()" not in vk_body


def test_kaggle_dataset_slug_stays_within_current_api_limit():
    from subscriber_acquisition.kaggle_runner import (
        CONFIG_DATASET_CIPHER,
        CONFIG_DATASET_KEY,
        _build_dataset_slug,
    )

    run_id = "20260701T010203Z-abcdef1234567890"
    for prefix in [CONFIG_DATASET_CIPHER, CONFIG_DATASET_KEY, "subscriber-acquisition-discovery-extra-long-prefix"]:
        slug = _build_dataset_slug(prefix, run_id)
        assert 6 <= len(slug) <= 50
        assert "20260701t010203z" in slug

    assert _build_dataset_slug(CONFIG_DATASET_CIPHER, run_id) != _build_dataset_slug(CONFIG_DATASET_KEY, run_id)


def test_kaggle_status_parser_handles_sdk_enum_values():
    from subscriber_acquisition.kaggle_runner import _status_to_text

    class StatusEnum:
        name = "COMPLETE"

    class Response:
        status = StatusEnum()

    assert _status_to_text(Response()) == "COMPLETE"
    assert _status_to_text({"status": "KernelWorkerStatus.COMPLETE"}) == "COMPLETE"
