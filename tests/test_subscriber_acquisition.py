from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace

import pytest
import pytest_asyncio
from sqlalchemy import select, text

from db import Database
from models import AcqDiscoveryRun, AcqLinkTarget, AcqOpportunity, AcqReviewFeedback, AcqSurface
from subscriber_acquisition.config import AcqConfig
from subscriber_acquisition.importer import import_discovery_result
from subscriber_acquisition.link_targets import select_link_target
from subscriber_acquisition.review import build_surface_keyboard, find_opportunity_by_review_message, publish_frontier_summary, publish_review_cards, record_feedback, record_surface_feedback
from subscriber_acquisition.safety import AcquisitionSafetyError, ensure_review_chat, ensure_vk_read_only
from subscriber_acquisition.scoring import conservative_reach_low, priority_score
from subscriber_acquisition.service import add_surface_seed, export_surface_map_xlsx, list_surfaces, run_acq_discovery_shadow


@pytest_asyncio.fixture
async def db(tmp_path, monkeypatch):
    monkeypatch.setenv("DB_INIT_SKIP_TG_SOURCES_SEED", "1")
    monkeypatch.setenv("DB_INIT_SKIP_GUIDE_SOURCES_SEED", "1")
    monkeypatch.setenv("DB_INIT_MINIMAL", "1")
    database = Database(str(tmp_path / "acq.sqlite"))
    await database.init()
    try:
        yield database
    finally:
        await database.close()


@pytest.fixture
def sample_payload():
    return json.loads(Path("tests/fixtures/acq_discovery_result.sample.json").read_text(encoding="utf-8"))


@pytest.mark.asyncio
async def test_import_discovery_result_creates_schema_rows(db, sample_payload):
    result = await import_discovery_result(db, sample_payload)
    assert result.run.id
    assert len(result.surfaces) == 1
    assert len(result.opportunities) == 1
    async with db.get_session() as session:
        surfaces = (await session.execute(select(AcqSurface))).scalars().all()
        opps = (await session.execute(select(AcqOpportunity))).scalars().all()
        targets = (await session.execute(select(AcqLinkTarget))).scalars().all()
        runs = (await session.execute(select(AcqDiscoveryRun))).scalars().all()
    assert surfaces[0].external_id == "tg:example"
    assert opps[0].link_target_kind == "topic_landing"
    assert opps[0].sticker_fit == "possible"
    assert targets[0].kind == "topic_landing"
    assert runs[0].status == "done"




@pytest.mark.asyncio
async def test_import_marks_only_scanned_surfaces_as_scanned(db, sample_payload):
    payload = json.loads(json.dumps(sample_payload))
    payload["surfaces"] = [
        {
            "platform": "vk",
            "surface_type": "community",
            "url": "https://vk.com/scanned",
            "handle": "scanned",
            "external_id": "vk:scanned",
            "status": "approved",
            "source": "allowlist",
            "reach": {"basis": "vk_wall", "confidence": "low"},
            "risk": {},
        },
        {
            "platform": "vk",
            "surface_type": "community",
            "url": "https://vk.com/waiting",
            "handle": "waiting",
            "external_id": "vk:waiting",
            "status": "approved",
            "source": "seed",
            "reach": {"basis": "seed_only", "confidence": "low"},
            "risk": {},
        },
    ]
    payload["opportunities"] = []

    await import_discovery_result(db, payload)

    async with db.get_session() as session:
        rows = (await session.execute(select(AcqSurface))).scalars().all()
    by_external = {row.external_id: row for row in rows}
    assert by_external["vk:scanned"].last_scan_at is not None
    assert by_external["vk:scanned"].next_scan_after is not None
    assert by_external["vk:waiting"].last_scan_at is None
    assert by_external["vk:waiting"].next_scan_after is None


@pytest.mark.asyncio
async def test_import_updates_existing_surface_to_out_of_region_rejected(db, sample_payload):
    first = json.loads(json.dumps(sample_payload))
    first["surfaces"][0]["external_id"] = "tg:visitNavahrudak"
    first["surfaces"][0]["url"] = "https://t.me/visitNavahrudak"
    first["surfaces"][0]["handle"] = "visitNavahrudak"
    first["opportunities"] = []
    await import_discovery_result(db, first)

    rejected = json.loads(json.dumps(first))
    rejected["run_id"] = "second-run"
    rejected["surfaces"][0]["status"] = "rejected_out_of_region"
    rejected["surfaces"][0]["topic_cluster"] = "out_of_region"
    rejected["surfaces"][0]["risk"] = {"level": "rejected", "reason": "out-of-region"}
    await import_discovery_result(db, rejected)

    async with db.get_session() as session:
        surface = (await session.execute(select(AcqSurface).where(AcqSurface.external_id == "tg:visitNavahrudak"))).scalar_one()
    assert surface.status == "rejected_out_of_region"
    assert surface.topic_cluster == "out_of_region"


@pytest.mark.asyncio
async def test_import_rejects_telegram_bot_surfaces(db, sample_payload):
    payload = json.loads(json.dumps(sample_payload))
    payload["surfaces"][0]["external_id"] = "tg:RK39_bot"
    payload["surfaces"][0]["url"] = "https://t.me/RK39_bot"
    payload["surfaces"][0]["handle"] = "RK39_bot"
    payload["opportunities"] = []

    await import_discovery_result(db, payload)

    async with db.get_session() as session:
        surface = (await session.execute(select(AcqSurface).where(AcqSurface.external_id == "tg:RK39_bot"))).scalar_one()
    assert surface.status == "rejected_bot_or_service"


@pytest.mark.asyncio
async def test_import_rejects_vk_non_community_surfaces(db, sample_payload):
    payload = json.loads(json.dumps(sample_payload))
    payload["surfaces"][0]["platform"] = "vk"
    payload["surfaces"][0]["external_id"] = "vk:album-1_2"
    payload["surfaces"][0]["url"] = "https://vk.com/album-1_2"
    payload["surfaces"][0]["handle"] = "album-1_2"
    payload["opportunities"] = []

    await import_discovery_result(db, payload)

    async with db.get_session() as session:
        surface = (await session.execute(select(AcqSurface).where(AcqSurface.external_id == "vk:album-1_2"))).scalar_one()
    assert surface.status == "rejected_non_community"


@pytest.mark.asyncio
async def test_import_dedupes_context(db, sample_payload):
    first = await import_discovery_result(db, sample_payload)
    second = await import_discovery_result(db, sample_payload)
    assert len(first.opportunities) == 1
    assert len(second.opportunities) == 0
    assert second.skipped_duplicate_contexts == 1


def test_conservative_reach_and_priority_do_not_override_high_risk():
    cfg = AcqConfig(tg_comment_readthrough_factor=0.02, reach_unknown_group_low=5)
    reach = conservative_reach_low(platform="tg", surface_type="channel", recent_views_p10=1000, config=cfg)
    assert reach["low"] == 20
    assert priority_score(relevance=0.95, reach_low=10000, spam_risk="high", safety_risk="low") == 0.0


def test_link_target_selection_topic_and_event():
    cfg = AcqConfig(default_link_target_url="https://t.me/kenigevents", pka_channel_url="https://t.me/kenigevents")
    topic = select_link_target(topic_cluster="organ_concerts", candidate_events=[], config=cfg)
    assert topic.kind == "topic_landing"
    event = SimpleNamespace(ticket_link="https://tickets.example/1", telegraph_url="https://telegra.ph/x")
    chosen = select_link_target(topic_cluster="organ_concerts", candidate_events=[event], config=cfg)
    assert chosen.kind == "event_site"
    assert chosen.url == "https://tickets.example/1"


class FakeBot:
    def __init__(self):
        self.calls = []

    async def send_message(self, chat_id, text, **kwargs):
        assert chat_id == 777  # no external TG send: review chat only
        self.calls.append((chat_id, text, kwargs))
        return SimpleNamespace(chat=SimpleNamespace(id=chat_id), message_id=100 + len(self.calls))


@pytest.mark.asyncio
async def test_review_cards_and_feedback_capture(db, sample_payload):
    result = await import_discovery_result(db, sample_payload)
    cfg = AcqConfig(review_chat_id=777, review_group_max_cards_per_run=20)
    bot = FakeBot()
    posted = await publish_review_cards(db, bot, result.opportunities, config=cfg)
    assert posted == 1
    assert "✅ Да" in repr(bot.calls[0][2]["reply_markup"])
    async with db.get_session() as session:
        opp = (await session.execute(select(AcqOpportunity))).scalar_one()
    assert opp.review_message_chat_id == 777
    fb = await record_feedback(db, opportunity_id=opp.id, reviewer_id=42, action="approve", review_message_chat_id=777, review_message_id=101)
    assert fb.action == "approve"
    found = await find_opportunity_by_review_message(db, chat_id=777, message_id=101)
    assert found and found.id == opp.id
    comment = await record_feedback(db, opportunity_id=opp.id, reviewer_id=42, action="comment", note="хороший кейс", review_message_chat_id=777, review_message_id=101)
    assert comment.note == "хороший кейс"
    async with db.get_session() as session:
        fbs = (await session.execute(select(AcqReviewFeedback))).scalars().all()
        opp2 = await session.get(AcqOpportunity, opp.id)
    assert [fb.action for fb in fbs] == ["shown", "approve", "comment"]
    assert opp2.status == "approved"


def test_review_card_renders_gemma_checklist():
    from subscriber_acquisition.review import format_review_card

    opp = AcqOpportunity(
        id=5,
        platform="vk",
        context_url="https://vk.com/wall-1_2?reply=3",
        context_text_snippet="А где афиша на выходные?",
        evidence_json={"llm_gate": {"checklist": [
            {"id": "question", "question": "Есть явный вопрос?", "answer": True, "note": "спрашивает афишу"},
            {"id": "post_factum", "question": "Не постфактум-отзыв?", "answer": True, "note": "запрос будущей афиши"},
        ]}},
        matched_intent="event_recommendation_request",
        topic_cluster="local_events",
        link_target_kind="pka_channel",
        link_target_label="Полюбить Калининград Анонсы",
        reach_low=3,
        spam_risk="low",
        safety_risk="low",
    )

    card = format_review_card(opp)

    assert "Контрольный список Gemma" in card
    assert "✅ Есть явный вопрос" in card
    assert "спрашивает афишу" in card


def test_review_card_hides_non_russian_gemma_notes():
    from subscriber_acquisition.review import format_review_card

    opp = AcqOpportunity(
        id=6,
        platform="tg",
        context_url="https://t.me/example/1",
        context_text_snippet="А где афиша 1 дня?",
        evidence_json={"llm_gate": {"checklist": [
            {"id": "question", "question": "Есть явный вопрос?", "answer": True, "note": "User asks where the schedule is"},
        ]}},
        matched_intent="event_recommendation_question",
        topic_cluster="local_events",
        link_target_kind="pka_channel",
        reach_low=3,
        spam_risk="low",
        safety_risk="low",
    )

    card = format_review_card(opp)

    assert "✅ Есть явный вопрос?" in card
    assert "User asks" not in card


@pytest.mark.asyncio
async def test_frontier_summary_publishes_new_surfaces_without_approval_buttons(db):
    async with db.get_session() as session:
        new_surface = AcqSurface(
            platform="tg",
            surface_type="linked_discussion",
            url="https://t.me/c/123",
            external_id="tg:123",
            title="New linked chat",
            status="candidate",
            source="linked_discussion",
        )
        old_seed = AcqSurface(
            platform="tg",
            surface_type="channel",
            url="https://t.me/old",
            external_id="tg:old",
            title="Old seed",
            status="candidate",
            source="seed",
        )
        session.add(new_surface)
        session.add(old_seed)
        await session.commit()
        await session.refresh(new_surface)
        await session.refresh(old_seed)

    bot = FakeBot()
    shown = await publish_frontier_summary(db, bot, [old_seed, new_surface], config=AcqConfig(review_chat_id=777))

    assert shown == 1
    assert len(bot.calls) == 1
    assert "New linked chat" in bot.calls[0][1]
    assert "Согласование не требуется" in bot.calls[0][1]
    assert bot.calls[0][2].get("reply_markup") is None
    async with db.get_session() as session:
        feedback = (await session.execute(select(AcqReviewFeedback))).scalars().all()
    assert [row.action for row in feedback] == ["frontier_summary_shown"]


@pytest.mark.asyncio
async def test_surface_seed_add_and_review_feedback(db):
    surface = await add_surface_seed(db, "https://t.me/new_public", reviewer_id=42)
    assert surface.external_id == "tg:new_public"
    assert surface.status == "candidate"
    keyboard = build_surface_keyboard(surface)
    texts = [button.text for row in keyboard.inline_keyboard for button in row]
    assert texts == ["✅ Да", "❌ Нет + причина", "🕒 Потом", "💬 Оставить причину", "🔗 Открыть"]

    feedback = await record_surface_feedback(db, surface_id=surface.id, reviewer_id=42, action="approve")
    surfaces = await list_surfaces(db, limit=5)

    assert feedback.action == "surface_approve"
    assert surfaces[0].status == "approved"
    async with db.get_session() as session:
        stored_feedback = (await session.execute(select(AcqReviewFeedback))).scalars().all()
    assert stored_feedback[0].surface_id == surface.id


@pytest.mark.asyncio
async def test_review_cards_are_hard_capped_at_20_and_have_required_buttons(db, sample_payload):
    bulk = json.loads(json.dumps(sample_payload))
    base = bulk["opportunities"][0]
    bulk["opportunities"] = []
    for i in range(25):
        item = json.loads(json.dumps(base))
        item["context_url"] = f"https://t.me/example/{1000 + i}"
        item["context_external_id"] = str(1000 + i)
        item["context_text_snippet"] = f"Где событие #{i}?"
        bulk["opportunities"].append(item)
    result = await import_discovery_result(db, bulk)
    cfg = AcqConfig(review_chat_id=777, review_group_max_cards_per_run=50)
    bot = FakeBot()

    posted = await publish_review_cards(db, bot, result.opportunities, config=cfg)

    assert posted == 20
    assert len(bot.calls) == 20
    keyboard = bot.calls[0][2]["reply_markup"].inline_keyboard
    texts = [button.text for row in keyboard for button in row]
    assert texts == ["✅ Да", "❌ Нет + причина", "🕒 Потом", "💬 Оставить причину", "🔗 Контекст", "🎯 Куда"]


@pytest.mark.asyncio
async def test_shadow_run_with_configured_fixture_path_creates_review_artifacts(db, sample_payload, tmp_path, monkeypatch):
    async def fake_report(run, surfaces, opportunities):
        return "https://telegra.ph/acq-report"

    fixture = tmp_path / "acq_result.json"
    fixture.write_text(json.dumps(sample_payload, ensure_ascii=False), encoding="utf-8")
    monkeypatch.setattr("subscriber_acquisition.service.publish_telegraph_report", fake_report)
    cfg = AcqConfig(review_chat_id=777, fixture_path=str(fixture))
    bot = FakeBot()

    result = await run_acq_discovery_shadow(db, bot=bot, config=cfg)
    opp_id = result.opportunities[0].id
    fb = await record_feedback(db, opportunity_id=opp_id, reviewer_id=42, action="approve", review_message_chat_id=777, review_message_id=101)

    async with db.get_session() as session:
        runs = (await session.execute(select(AcqDiscoveryRun))).scalars().all()
        surfaces = (await session.execute(select(AcqSurface))).scalars().all()
        opportunities = (await session.execute(select(AcqOpportunity))).scalars().all()
        targets = (await session.execute(select(AcqLinkTarget))).scalars().all()
        feedback = (await session.execute(select(AcqReviewFeedback))).scalars().all()

    assert result.run.telegraph_url == "https://telegra.ph/acq-report"
    assert result.run.stats_json["review_cards_posted"] == 1
    assert len(bot.calls) == 1
    assert runs and surfaces and opportunities and targets and feedback
    assert {row.action for row in feedback} == {"shown", "approve"}
    assert fb.action == "approve"


class FakeWrongChatBot:
    async def send_message(self, chat_id, text, **kwargs):
        return SimpleNamespace(chat=SimpleNamespace(id=123), message_id=999)


@pytest.mark.asyncio
async def test_review_card_send_verifies_returned_review_chat(db, sample_payload):
    result = await import_discovery_result(db, sample_payload)
    cfg = AcqConfig(review_chat_id=777, review_group_max_cards_per_run=20)

    with pytest.raises(AcquisitionSafetyError):
        await publish_review_cards(db, FakeWrongChatBot(), result.opportunities, config=cfg)


def test_no_send_guard_blocks_external_targets():
    ensure_review_chat(777, review_chat_id=777)
    with pytest.raises(AcquisitionSafetyError):
        ensure_review_chat(123, review_chat_id=777)
    with pytest.raises(AcquisitionSafetyError):
        ensure_vk_read_only("wall.createComment")


@pytest.mark.asyncio
async def test_surface_map_xlsx_has_clickable_group_links(db):
    surface = await add_surface_seed(db, "https://t.me/map_test", reviewer_id=42)
    await record_surface_feedback(db, surface_id=surface.id, reviewer_id=42, action="approve")

    path = await export_surface_map_xlsx(db)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["groups"]
    headers = [cell.value for cell in ws[1]]
    assert "reply_policy" in headers
    assert "scan_state" in headers
    rows = list(ws.iter_rows(min_row=2, values_only=False))
    row = next(row for row in rows if row[4].value == "https://t.me/map_test")
    assert row[4].hyperlink.target == "https://t.me/map_test"
    assert row[8].value == "confirmed_can_reply_after_human_review"
    assert "summary" in wb.sheetnames


@pytest.mark.asyncio
async def test_runtime_seed_payload_includes_telega_in_kaliningrad_seeds(db):
    from subscriber_acquisition.kaggle_runner import collect_runtime_seed_payload

    payload = await collect_runtime_seed_payload(db)

    by_external = {item["external_id"]: item for item in payload["surfaces"]}
    assert by_external["tg:anons39"]["source"] == "telega_in"
    assert "telega.in/channels/anons39" in by_external["tg:anons39"]["topic_hint"]


@pytest.mark.asyncio
async def test_runtime_seed_payload_includes_existing_vk_monitoring_groups(db):
    from subscriber_acquisition.kaggle_runner import collect_runtime_seed_payload

    async with db.get_session() as session:
        await session.execute(text(
            "CREATE TABLE vk_source(group_id INTEGER, screen_name TEXT, name TEXT, owner_type TEXT)"
        ))
        await session.execute(text(
            "INSERT INTO vk_source(group_id, screen_name, name, owner_type) VALUES (123, 'club123', 'VK Club', 'group'), (456, 'person456', 'VK Person', 'user')"
        ))
        await session.commit()

    payload = await collect_runtime_seed_payload(db)
    by_external = {item["external_id"]: item for item in payload["surfaces"]}
    assert by_external["vk:club123"]["source"] == "vk_source"
    assert "vk:person456" not in by_external


@pytest.mark.asyncio
async def test_runtime_seed_payload_rejects_existing_telegram_bots(db):
    from subscriber_acquisition.kaggle_runner import collect_runtime_seed_payload

    async with db.get_session() as session:
        session.add(AcqSurface(
            platform="tg",
            surface_type="unknown_public",
            url="https://t.me/RK39_bot",
            handle="RK39_bot",
            external_id="tg:RK39_bot",
            status="candidate",
            source="discovered",
        ))
        await session.commit()

    payload = await collect_runtime_seed_payload(db)

    assert "tg:RK39_bot" not in {item["external_id"] for item in payload["surfaces"]}
    async with db.get_session() as session:
        surface = (await session.execute(select(AcqSurface).where(AcqSurface.external_id == "tg:RK39_bot"))).scalar_one()
    assert surface.status == "rejected_bot_or_service"


@pytest.mark.asyncio
async def test_runtime_seed_payload_rejects_existing_vk_non_communities(db):
    from subscriber_acquisition.kaggle_runner import collect_runtime_seed_payload

    async with db.get_session() as session:
        session.add(AcqSurface(
            platform="vk",
            surface_type="community",
            url="https://vk.com/album-123_456",
            handle="album-123_456",
            external_id="vk:album-123_456",
            status="candidate",
            source="discovered",
        ))
        await session.commit()

    payload = await collect_runtime_seed_payload(db)

    assert "vk:album-123_456" not in {item["external_id"] for item in payload["surfaces"]}
    async with db.get_session() as session:
        surface = (await session.execute(select(AcqSurface).where(AcqSurface.external_id == "vk:album-123_456"))).scalar_one()
    assert surface.status == "rejected_non_community"


@pytest.mark.asyncio
async def test_runtime_seed_payload_prioritizes_new_frontier_surfaces(db):
    from subscriber_acquisition.kaggle_runner import collect_runtime_seed_payload

    async with db.get_session() as session:
        session.add(AcqSurface(
            platform="tg",
            surface_type="channel",
            url="https://t.me/old_seed",
            external_id="tg:old_seed",
            status="candidate",
            source="seed",
        ))
        session.add(AcqSurface(
            platform="tg",
            surface_type="linked_discussion",
            url="https://t.me/c/123",
            external_id="tg:123",
            status="candidate",
            source="linked_discussion",
        ))
        await session.commit()

    payload = await collect_runtime_seed_payload(db)

    assert payload["surfaces"][0]["external_id"] == "tg:123"
    assert any(item["external_id"] == "tg:anons39" for item in payload["surfaces"][:14])
    assert payload["surfaces"].index(next(item for item in payload["surfaces"] if item["external_id"] == "tg:old_seed")) > 1


@pytest.mark.asyncio
async def test_runtime_seed_payload_prioritizes_smartik_communities(db):
    from subscriber_acquisition.kaggle_runner import collect_runtime_seed_payload, _runtime_env_from_config

    async with db.get_session() as session:
        session.add(AcqSurface(
            platform="vk",
            surface_type="community",
            url="https://vk.com/club42481124",
            handle="club42481124",
            external_id="vk:club42481124",
            status="approved",
            source="seed",
        ))
        for idx in range(8):
            session.add(AcqSurface(
                platform="vk",
                surface_type="community",
                url=f"https://vk.com/club9{idx}",
                handle=f"club9{idx}",
                external_id=f"vk:club9{idx}",
                status="candidate",
                source="discovered",
            ))
        await session.commit()

    payload = await collect_runtime_seed_payload(db)
    env = _runtime_env_from_config(AcqConfig(max_surfaces_per_run=10), payload)
    vk_seeds = json.loads(env["ACQ_VK_SEEDS_JSON"])

    assert vk_seeds[0] == "https://vk.com/club42481124"
    by_external = {item["external_id"]: item for item in payload["surfaces"]}
    assert by_external["vk:club42481124"]["source"] == "smartik_kaliningrad_catalog"


@pytest.mark.asyncio
async def test_shadow_run_without_payload_uses_kaggle_runner_by_default(db, sample_payload, monkeypatch):
    async def fake_report(run, surfaces, opportunities):
        return "https://telegra.ph/acq-report"

    async def fake_kaggle(db_arg, *, config, seed_payload=None):
        assert db_arg is db
        assert config.runner == "kaggle"
        assert seed_payload and any(item.get("source") == "telega_in" for item in seed_payload.get("surfaces", []))
        return SimpleNamespace(payload=sample_payload, output_path=Path("/tmp/acq.json"), runner="kaggle", kernel_ref="user/subscriber-acquisition-discovery", run_id="test-run", status="complete")

    monkeypatch.setattr("subscriber_acquisition.service.publish_telegraph_report", fake_report)
    monkeypatch.setattr("subscriber_acquisition.service.run_kaggle_discovery_runtime", fake_kaggle)
    cfg = AcqConfig(review_chat_id=777)
    bot = FakeBot()

    result = await run_acq_discovery_shadow(db, bot=bot, config=cfg)

    assert result.run.telegraph_url == "https://telegra.ph/acq-report"
    assert len(result.surfaces) == 1
    assert len(result.opportunities) == 1
    async with db.get_session() as session:
        feedback = (await session.execute(select(AcqReviewFeedback))).scalars().all()
    assert [row.action for row in feedback] == ["shown"]


@pytest.mark.asyncio
async def test_shadow_run_refuses_live_scan_when_remote_session_busy(db, monkeypatch):
    class Busy(RuntimeError):
        pass

    async def fake_busy():
        raise Busy("remote session busy")

    def should_not_run(*, config, seed_payload=None):
        raise AssertionError("runtime must not start while remote session is busy")

    monkeypatch.setattr("subscriber_acquisition.service.ensure_remote_telegram_session_available_for_discovery", fake_busy)
    monkeypatch.setattr("subscriber_acquisition.service.run_local_discovery_runtime", should_not_run)

    with pytest.raises(Busy):
        await run_acq_discovery_shadow(db, bot=None, config=AcqConfig(runner="local"))


@pytest.mark.asyncio
async def test_shadow_run_without_payload_uses_local_runtime(db, sample_payload, monkeypatch):
    async def fake_report(run, surfaces, opportunities):
        return "https://telegra.ph/acq-report"

    def fake_runtime(*, config, seed_payload=None):
        assert config.runner == "local"
        assert seed_payload and any(item.get("source") == "telega_in" for item in seed_payload.get("surfaces", []))
        return SimpleNamespace(payload=sample_payload, output_path=Path("/tmp/acq.json"), runner="local_shadow_runtime")

    monkeypatch.setattr("subscriber_acquisition.service.publish_telegraph_report", fake_report)
    monkeypatch.setattr("subscriber_acquisition.service.run_local_discovery_runtime", fake_runtime)
    cfg = AcqConfig(review_chat_id=777, runner="local")
    bot = FakeBot()

    result = await run_acq_discovery_shadow(db, bot=bot, config=cfg)

    assert result.run.telegraph_url == "https://telegra.ph/acq-report"
    assert len(result.surfaces) == 1
    assert len(result.opportunities) == 1
    async with db.get_session() as session:
        targets = (await session.execute(select(AcqLinkTarget))).scalars().all()
        feedback = (await session.execute(select(AcqReviewFeedback))).scalars().all()
    assert targets and targets[0].kind == "topic_landing"
    assert [row.action for row in feedback] == ["shown"]


@pytest.mark.asyncio
async def test_shadow_run_publishes_report_and_review_cards(db, sample_payload, monkeypatch):
    async def fake_report(run, surfaces, opportunities):
        return "https://telegra.ph/acq-report"

    monkeypatch.setattr("subscriber_acquisition.service.publish_telegraph_report", fake_report)
    cfg = AcqConfig(review_chat_id=777, review_group_max_cards_per_run=20)
    bot = FakeBot()
    result = await run_acq_discovery_shadow(db, bot=bot, payload=sample_payload, config=cfg)
    assert result.run.telegraph_url == "https://telegra.ph/acq-report"
    assert result.run.stats_json["review_cards_posted"] == 1
    assert len(bot.calls) == 1
    async with db.get_session() as session:
        feedback = (await session.execute(select(AcqReviewFeedback))).scalars().all()
    assert [row.action for row in feedback] == ["shown"]


def test_ydb_stats_disabled_by_default(monkeypatch):
    from subscriber_acquisition.ydb_stats import ydb_stats_enabled

    monkeypatch.delenv("ACQ_YDB_STATS_ENABLED", raising=False)
    assert ydb_stats_enabled() is False


@pytest.mark.asyncio
async def test_import_records_ydb_export_when_enabled(db, sample_payload, monkeypatch):
    async def fake_export(payload, *, run_db_id=None):
        return {"run_uid": payload["run_id"], "surfaces": len(payload["surfaces"]), "opportunities": len(payload["opportunities"])}

    monkeypatch.setenv("ACQ_YDB_STATS_ENABLED", "1")
    monkeypatch.setattr("subscriber_acquisition.ydb_stats.export_discovery_payload_to_ydb", fake_export)

    result = await import_discovery_result(db, sample_payload)

    assert result.run.stats_json["ydb_export"]["run_uid"] == sample_payload["run_id"]
    assert result.run.stats_json["ydb_export"]["surfaces"] == 1
