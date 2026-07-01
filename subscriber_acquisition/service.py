from __future__ import annotations

import json
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import select

from models import AcqDiscoveryRun, AcqOpportunity, AcqReviewFeedback, AcqSurface
from .config import AcqConfig, load_config
from .importer import ImportResult, import_discovery_result
from .kaggle_runner import collect_runtime_seed_payload, ensure_remote_telegram_session_available_for_discovery, run_kaggle_discovery_runtime, run_local_discovery_runtime
from .report import publish_telegraph_report
from .review import publish_frontier_summary, publish_review_cards

_SAMPLE_FIXTURE = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "acq_discovery_result.sample.json"


async def run_acq_discovery_shadow(db, bot: Any | None = None, *, payload: dict[str, Any] | None = None, config: AcqConfig | None = None) -> ImportResult:
    cfg = config or load_config()
    if payload is None:
        candidate_paths = [cfg.discovery_results_path, cfg.fixture_path]
        for raw_path in candidate_paths:
            if not raw_path:
                continue
            path = Path(raw_path)
            if path.exists():
                payload = json.loads(path.read_text(encoding="utf-8"))
                break
        if payload is None and cfg.use_sample_fixture and _SAMPLE_FIXTURE.exists():
            payload = json.loads(_SAMPLE_FIXTURE.read_text(encoding="utf-8"))
        if payload is None and cfg.runner in {"kaggle", "kaggle_shadow"}:
            seed_payload = await collect_runtime_seed_payload(db)
            runtime_result = await run_kaggle_discovery_runtime(db, config=cfg, seed_payload=seed_payload)
            payload = runtime_result.payload
        if payload is None and cfg.runner in {"local", "local_shadow", "local_shadow_runtime"}:
            await ensure_remote_telegram_session_available_for_discovery()
            seed_payload = await collect_runtime_seed_payload(db)
            runtime_result = run_local_discovery_runtime(config=cfg, seed_payload=seed_payload)
            payload = runtime_result.payload
        if payload is None:
            payload = {"run_id": "empty", "surfaces": [], "opportunities": [], "diagnostics": ["ACQ_DISCOVERY_RUNNER disabled or unsupported"]}
    result = await import_discovery_result(db, payload)
    telegraph_url = None
    review_cards_posted = 0
    frontier_surfaces_shown = 0
    if result.opportunities:
        try:
            telegraph_url = await publish_telegraph_report(result.run, result.surfaces, result.opportunities)
        except Exception:
            telegraph_url = None
    if bot is not None and cfg.review_chat_id:
        if result.opportunities:
            review_cards_posted = await publish_review_cards(db, bot, result.opportunities, config=cfg)
        frontier_surfaces_shown = await publish_frontier_summary(db, bot, result.surfaces, config=cfg)
    if telegraph_url or review_cards_posted or frontier_surfaces_shown:
        async with db.get_session() as session:
            run = await session.get(AcqDiscoveryRun, result.run.id)
            if run is not None:
                stats = dict(run.stats_json or {})
                if review_cards_posted:
                    stats["review_cards_posted"] = review_cards_posted
                if frontier_surfaces_shown:
                    stats["frontier_surfaces_shown"] = frontier_surfaces_shown
                    stats["frontier_summary_posted"] = 1
                run.stats_json = stats
                run.telegraph_url = telegraph_url
                session.add(run)
                await session.commit()
                await session.refresh(run)
                result.run = run
    return result


async def queue_counts(db) -> dict[str, int]:
    async with db.get_session() as session:
        rows = (await session.execute(select(AcqOpportunity.status))).all()
    counts: dict[str, int] = {}
    for (status,) in rows:
        counts[str(status or "pending")] = counts.get(str(status or "pending"), 0) + 1
    return counts


async def surface_counts(db) -> dict[str, int]:
    async with db.get_session() as session:
        rows = (await session.execute(select(AcqSurface.status))).all()
    counts: dict[str, int] = {}
    for (status,) in rows:
        counts[str(status or "candidate")] = counts.get(str(status or "candidate"), 0) + 1
    return counts


async def latest_report_url(db) -> str | None:
    async with db.get_session() as session:
        run = (await session.execute(
            select(AcqDiscoveryRun).where(AcqDiscoveryRun.telegraph_url.is_not(None)).order_by(AcqDiscoveryRun.id.desc()).limit(1)
        )).scalar_one_or_none()
    return run.telegraph_url if run else None


async def export_feedback_jsonl(db) -> str:
    async with db.get_session() as session:
        rows = (await session.execute(select(AcqReviewFeedback).order_by(AcqReviewFeedback.created_at.asc()))).scalars().all()
    lines = []
    for fb in rows:
        lines.append(json.dumps({
            "id": fb.id,
            "opportunity_id": fb.opportunity_id,
            "surface_id": fb.surface_id,
            "reviewer_id": fb.reviewer_id,
            "action": fb.action,
            "note": fb.note,
            "created_at": fb.created_at.isoformat() if fb.created_at else None,
        }, ensure_ascii=False))
    return "\n".join(lines) + ("\n" if lines else "")


def _surface_scan_state(surface: AcqSurface, *, now: datetime | None = None) -> str:
    now = now or datetime.now(timezone.utc)
    if surface.last_scan_at is None:
        return "pending_first_scan"
    next_scan = surface.next_scan_after
    if next_scan is not None:
        if next_scan.tzinfo is None:
            next_scan = next_scan.replace(tzinfo=timezone.utc)
        if next_scan > now:
            return "scanned_waiting_rescan"
    return "ready_for_rescan"


def _surface_reply_policy(surface: AcqSurface) -> str:
    status = str(surface.status or "candidate").strip().lower()
    if status == "approved":
        return "confirmed_can_reply_after_human_review"
    if status == "rejected":
        return "rejected_do_not_reply"
    if status == "paused":
        return "paused_recheck_later"
    return "pending_discovery_analysis"


def _dt_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        return value.isoformat(sep=" ", timespec="seconds")
    except Exception:
        return str(value)


async def export_surface_map_xlsx(db) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    async with db.get_session() as session:
        surfaces = list((await session.execute(
            select(AcqSurface).order_by(AcqSurface.platform.asc(), AcqSurface.status.asc(), AcqSurface.updated_at.desc(), AcqSurface.id.asc())
        )).scalars().all())
        opportunities = list((await session.execute(select(AcqOpportunity))).scalars().all())

    opp_counts: dict[int, dict[str, int]] = {}
    for opp in opportunities:
        if not opp.surface_id:
            continue
        counts = opp_counts.setdefault(int(opp.surface_id), {})
        status = str(opp.status or "pending")
        counts[status] = counts.get(status, 0) + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "groups"
    headers = [
        "id", "platform", "type", "title", "url", "handle", "status",
        "scan_state", "reply_policy", "source", "topic_hint", "reach", "risk",
        "last_scan_at", "next_scan_after", "opportunities_total", "opportunities_by_status",
        "review_note", "created_at", "updated_at",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for surface in surfaces:
        counts = opp_counts.get(int(surface.id or 0), {})
        total = sum(counts.values())
        row = [
            surface.id,
            surface.platform,
            surface.surface_type,
            surface.title or surface.handle or surface.external_id or surface.url,
            surface.url,
            surface.handle,
            surface.status,
            _surface_scan_state(surface),
            _surface_reply_policy(surface),
            surface.source,
            surface.topic_hint,
            json.dumps(surface.reach_json or {}, ensure_ascii=False),
            json.dumps(surface.risk_json or {}, ensure_ascii=False),
            _dt_text(surface.last_scan_at),
            _dt_text(surface.next_scan_after),
            total,
            json.dumps(counts, ensure_ascii=False, sort_keys=True),
            surface.review_note,
            _dt_text(surface.created_at),
            _dt_text(surface.updated_at),
        ]
        ws.append(row)
        url_cell = ws.cell(row=ws.max_row, column=5)
        if surface.url:
            url_cell.hyperlink = surface.url
            url_cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {1: 8, 2: 12, 3: 20, 4: 38, 5: 44, 7: 14, 8: 24, 9: 34, 10: 18, 11: 38, 16: 18, 17: 24}
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(idx, 18)

    summary = wb.create_sheet("summary")
    summary.append(["metric", "value"])
    summary["A1"].font = summary["B1"].font = Font(bold=True)
    by_status: dict[str, int] = {}
    by_platform: dict[str, int] = {}
    by_policy: dict[str, int] = {}
    by_scan: dict[str, int] = {}
    for surface in surfaces:
        by_status[surface.status or "candidate"] = by_status.get(surface.status or "candidate", 0) + 1
        by_platform[surface.platform or "unknown"] = by_platform.get(surface.platform or "unknown", 0) + 1
        by_policy[_surface_reply_policy(surface)] = by_policy.get(_surface_reply_policy(surface), 0) + 1
        by_scan[_surface_scan_state(surface)] = by_scan.get(_surface_scan_state(surface), 0) + 1
    rows = [("total_surfaces", len(surfaces)), ("total_opportunities", len(opportunities))]
    rows += [(f"status:{k}", v) for k, v in sorted(by_status.items())]
    rows += [(f"platform:{k}", v) for k, v in sorted(by_platform.items())]
    rows += [(f"reply_policy:{k}", v) for k, v in sorted(by_policy.items())]
    rows += [(f"scan_state:{k}", v) for k, v in sorted(by_scan.items())]
    for row in rows:
        summary.append(list(row))
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 18

    out = Path(tempfile.gettempdir()) / f"acq_surface_map_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out)
    return out


async def list_surfaces(db, *, status: str | None = None, limit: int = 10) -> list[AcqSurface]:
    async with db.get_session() as session:
        stmt = select(AcqSurface).order_by(AcqSurface.updated_at.desc(), AcqSurface.id.desc()).limit(max(1, min(int(limit), 25)))
        if status:
            stmt = select(AcqSurface).where(AcqSurface.status == status).order_by(AcqSurface.updated_at.desc(), AcqSurface.id.desc()).limit(max(1, min(int(limit), 25)))
        return list((await session.execute(stmt)).scalars().all())


def normalize_surface_seed(url: str) -> dict[str, str]:
    raw = str(url or "").strip()
    if not raw:
        raise ValueError("surface url is required")
    lowered = raw.lower()
    platform = "vk" if "vk.com/" in lowered else "tg"
    if platform == "tg" and raw.startswith("@"):
        raw = "https://t.me/" + raw.lstrip("@")
    if platform == "tg" and "t.me/" not in raw.lower():
        raise ValueError("Telegram seed must be @username or https://t.me/... URL")
    handle = raw.rstrip("/").split("/")[-1].split("?")[0].lstrip("@")
    if not handle:
        raise ValueError("could not extract surface handle")
    return {
        "platform": platform,
        "surface_type": "community" if platform == "vk" else "unknown_public",
        "url": raw,
        "handle": handle,
        "external_id": f"{platform}:{handle}",
    }


async def add_surface_seed(db, url: str, *, reviewer_id: int | None = None, note: str | None = None) -> AcqSurface:
    seed = normalize_surface_seed(url)
    async with db.get_session() as session:
        existing = (await session.execute(
            select(AcqSurface).where(
                AcqSurface.platform == seed["platform"],
                AcqSurface.external_id == seed["external_id"],
            )
        )).scalar_one_or_none()
        if existing is None:
            surface = AcqSurface(
                platform=seed["platform"],
                surface_type=seed["surface_type"],
                url=seed["url"],
                handle=seed["handle"],
                external_id=seed["external_id"],
                status="candidate",
                source="manual",
                review_note=note,
            )
        else:
            surface = existing
            surface.url = seed["url"]
            surface.handle = seed["handle"]
            surface.source = surface.source or "manual"
            surface.review_note = note or surface.review_note
        if reviewer_id:
            surface.reviewed_by = reviewer_id
        session.add(surface)
        await session.commit()
        await session.refresh(surface)
        return surface
