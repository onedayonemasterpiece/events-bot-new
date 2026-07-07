#!/usr/bin/env python3
from __future__ import annotations

import asyncio
import atexit
import base64
import csv
import gc
import hashlib
import html
import json
import os
import random
import re
import subprocess
import sys
import time
import zipfile
import urllib.parse
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from xml.sax.saxutils import escape

RUN_STARTED_AT = datetime.now(timezone.utc)
DEFAULT_ANCHORS = ["Калининград", "Калининградская область", "Куршская коса", "Зеленоградск", "Светлогорск", "Балтийское море", "Кёнигсберг", "Краснолесье", "Виштынец", "Роминтенская пуща", "Балтийская коса", "Янтарный", "Балтийск", "Советск", "Неман", "Правдинск", "Черняховск"]
TELEGRAM_KEYWORD_DISCOVERY_TERMS = [
    "Калининград", "Кёнигсберг", "Зеленоградск", "Светлогорск", "Янтарный", "Балтийск", "Пионерский",
    "Светлый", "Приморск", "Советск", "Черняховск", "Гусев", "Неман", "Полесск", "Правдинск",
    "Багратионовск", "Мамоново", "Нестеров", "Гвардейск", "Гурьевск", "Славск", "Озёрск",
    "Краснознаменск", "Ладушкин", "Краснолесье", "Донское", "Куликово", "Отрадное", "Морское",
    "Рыбачий", "Лесной", "Куршская коса", "Балтийская коса", "Виштынецкое озеро", "Роминтенская пуща",
    "Роминта", "Танцующий лес", "Высота Эфа", "Королевский бор", "Балтийское море", "Куршский залив",
    "Вислинский залив", "Филинская бухта", "Янтарный карьер", "Амалиенау", "Ратсхоф", "Рыбная деревня",
    "Остров Канта", "Кафедральный собор", "Музей Мирового океана", "Форт №5", "Фридрихсбургские ворота",
    "Закхаймские ворота", "Королевские ворота", "Верхнее озеро", "Нижнее озеро", "кирха Арнау",
    "Дом Советов", "Готическое кольцо", "замок Нессельбек", "замок Шаакен", "замок Тапиау",
    "Тильзит", "Инстербург", "Рагнит", "маяк Заливино", "маяк Балтийск",
]
NEWS_WORDS = ["происшеств", "дтп", "авар", "полици", "суд", "задерж", "штраф", "войн", "полит", "скандал", "убий", "пожар"]
AD_WORDS = ["скидк", "промокод", "купить", "заказать", "реклама", "партнёр", "партнер", "оплат", "бронь", "регистрация", "анонс", "конкурс", "диктант", "географический диктант", "билеты", "забронировать"]
TRASH_WORDS = ["жесть", "треш", "трэш", "шок", "кошмар"]
POSITIVE_WORDS = ["красив", "атмосфер", "море", "дюны", "архитект", "истори", "маршрут", "музей", "пляж", "курорт", "прогул", "путешеств"]

_REGION_TALK_SUPABASE_CLIENT: Any | None = None
_REGION_TALK_GOOGLE_CLIENT: Any | None = None
_REGION_TALK_CLIP_MODEL: Any | None = None
_REGION_TALK_CLIP_PROCESSOR: Any | None = None
_REGION_TALK_CLIP_DEVICE: str | None = None
_REGION_TALK_TEXT_MODELS: dict[str, Any] = {}
_REGION_TALK_TEXT_PROTOTYPE_CACHE: dict[str, Any] = {}
_REGION_TALK_TELEGRAM_RUNTIME: dict[str, Any] = {}



def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def getenv_bool(name: str, default: bool) -> bool:
    raw = (os.getenv(name) or "").strip().lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "on"}


def getenv_int(name: str, default: int) -> int:
    raw = (os.getenv(name) or "").strip()
    try:
        return int(raw) if raw else default
    except Exception:
        return default


def stable_hash(*parts: Any, length: int = 16) -> str:
    h = hashlib.sha256()
    for part in parts:
        h.update(str(part or "").strip().lower().encode("utf-8"))
        h.update(b"\0")
    return h.hexdigest()[:length]


def canonical_handle(value: str) -> str:
    raw = (value or "").strip()
    raw = re.sub(r"^https?://t\.me/", "@", raw, flags=re.I)
    raw = re.sub(r"^https?://(?:www\.)?vk\.com/", "@", raw, flags=re.I)
    if raw and not raw.startswith("@") and re.fullmatch(r"[A-Za-z0-9_\.]+", raw):
        raw = "@" + raw
    return raw


def seed_sort_number(value: Any) -> int:
    m = re.search(r"\d+", str(value or ""))
    return int(m.group(0)) if m else 999999


def canonical_source_url(platform: str, handle: str, url: str) -> str:
    if url:
        raw = url.strip()
        if platform == "telegram":
            m = re.search(r"(?:t\.me|telegram\.me)/(@?[A-Za-z0-9_]{4,})(?:/)?(?:\?.*)?$", raw, re.I)
            if m:
                return "https://t.me/" + m.group(1).lstrip("@")
        if platform in {"vk", "vkvideo"}:
            m = re.search(r"vk\.com/(?:club|public)?([A-Za-z0-9_.-]+)", raw, re.I)
            if m:
                ident = m.group(1)
                if ident.isdigit():
                    return "https://vk.com/club" + ident
                return "https://vk.com/" + ident
        return raw.rstrip("/")
    h = canonical_handle(handle).lstrip("@")
    if not h:
        return ""
    if platform == "telegram":
        return f"https://t.me/{h}"
    if platform in {"vk", "vkvideo"}:
        return f"https://vk.com/{h}"
    return h




VK_NON_SOURCE_PATHS = {
    "video", "videos", "clip", "clips", "feed", "search", "places", "market",
    "audio", "audios", "albums", "album", "photos", "photo", "app", "apps",
    "im", "groups", "groups_create", "mail", "away", "share", "login", "join",
}


def _url_host_path(url: str) -> tuple[str, str]:
    raw = (url or "").strip()
    if raw.startswith("@"):
        return "", raw
    try:
        parsed = urllib.parse.urlparse(raw if re.match(r"^[a-z][a-z0-9+.-]*://", raw, re.I) else "https://" + raw.lstrip("/"))
        return (parsed.netloc or "").lower(), (parsed.path or "").strip("/")
    except Exception:
        return "", ""


def target_source_url_reason(platform: str, url: str) -> str:
    """Guardrail for the durable source queue: only real TG channels and VK walls/communities.

    This is URL-shape validation, not content semantics. It prevents search/result/media
    pages (tgstat search, vk video, web pages) from becoming durable scan sources.
    """
    p = (platform or "").strip().lower()
    raw = (url or "").strip()
    host, path = _url_host_path(raw)
    first = (path.split("/", 1)[0] if path else "").lower()
    if p == "telegram":
        if host not in {"t.me", "telegram.me"}:
            return "telegram_source_must_be_t_me_channel"
        if "/" in path:
            return "telegram_source_not_public_channel_url"
        if not first or first in {"s", "c", "joinchat", "+", "share", "addstickers", "proxy", "iv"}:
            return "telegram_source_not_public_channel_url"
        if not re.fullmatch(r"[a-z0-9_]{4,}", first, re.I):
            return "telegram_source_handle_invalid"
        return "ok"
    if p == "vk":
        if host not in {"vk.com", "www.vk.com", "m.vk.com"}:
            return "vk_source_must_be_vk_com_wall_or_community"
        if not first or first in VK_NON_SOURCE_PATHS or re.match(r"(?:video|clip|photo|album|doc|story|market)-?", first, re.I):
            return "vk_source_not_wall_or_community_url"
        if first.startswith("wall") and "_" in first:
            return "vk_source_not_wall_or_community_url"
        if re.fullmatch(r"(?:club|public)\d+", first, re.I) or re.fullmatch(r"wall-?\d+", first, re.I) or re.fullmatch(r"[a-z0-9_.-]{3,}", first, re.I):
            return "ok"
        return "vk_source_identifier_invalid"
    return "unsupported_platform_for_source_queue"


def is_target_source_url(platform: str, url: str) -> bool:
    return target_source_url_reason(platform, url) == "ok"

def canonical_source_key(platform: str, handle: str = "", url: str = "") -> str:
    p = (platform or "").strip().lower()
    cu = canonical_source_url(p, handle, url).lower().rstrip("/")
    if p == "telegram":
        h = canonical_handle(handle or cu).lstrip("@").lower()
        if h:
            return "telegram:" + h
    if p.startswith("vk"):
        return "vk:" + re.sub(r"^https://vk\.com/", "", cu)
    return (p or "unknown") + ":" + (cu or canonical_handle(handle).lower())




def normalize_source_platform(value: str, url: str = "") -> str:
    raw = (value or "").strip().lower().replace("_", "-")
    low_url = (url or "").strip().lower()
    if raw in {"telegram", "tg", "telegram-channel", "telegram channel", "telegram_post", "telegram-post"} or "t.me/" in low_url or "telegram.me/" in low_url:
        return "telegram"
    if raw in {"vk", "vkontakte", "vk-public", "vk public", "vk-community", "vk community", "vk group", "vk_group"} or "vk.com/" in low_url:
        if "video" in raw:
            return "vkvideo"
        return "vk"
    if raw in {"youtube", "youtube-channel", "youtube channel"} or "youtube.com" in low_url or "youtu.be" in low_url:
        return "youtube"
    if raw in {"dzen", "zen", "dzen-channel", "dzen channel"} or "dzen.ru" in low_url:
        return "dzen"
    return raw or "unknown"

def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        if not line.strip() or line.lstrip().startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


def load_split_runtime_from_kaggle_input() -> dict[str, Any]:
    roots = [Path("/kaggle/input"), Path.cwd()]
    config: dict[str, Any] = {}
    secret_files: list[Path] = []
    key_files: list[Path] = []
    for root in roots:
        if not root.exists():
            continue
        for p in root.rglob("region_talk_run_config.json"):
            try:
                config.update(json.loads(p.read_text(encoding="utf-8")))
            except Exception:
                pass
        secret_files.extend(root.rglob("region_talk_secrets.enc"))
        key_files.extend(root.rglob("region_talk_fernet.key"))
    if secret_files and key_files:
        try:
            from cryptography.fernet import Fernet

            def _slug(path: Path, prefix: str) -> str:
                name = path.parent.name
                return name[len(prefix):] if name.startswith(prefix) else name

            key_by_slug = {_slug(k, "region-talk-key-"): k for k in key_files}
            pairs: list[tuple[Path, Path]] = []
            for secret_file in secret_files:
                slug = _slug(secret_file, "region-talk-secrets-")
                key_file = key_by_slug.get(slug)
                if key_file:
                    pairs.append((secret_file, key_file))
            # Backward compatible case: older runner may mount both files in one dataset folder.
            key_by_parent = {k.parent.resolve(): k for k in key_files}
            for secret_file in secret_files:
                key_file = key_by_parent.get(secret_file.parent.resolve())
                if key_file and (secret_file, key_file) not in pairs:
                    pairs.append((secret_file, key_file))
            for secret_file in secret_files:
                for key_file in key_files:
                    if (secret_file, key_file) not in pairs:
                        pairs.append((secret_file, key_file))
            failures: list[str] = []
            loaded = False
            for secret_file, key_file in pairs:
                try:
                    secrets = json.loads(Fernet(key_file.read_bytes().strip()).decrypt(secret_file.read_bytes()).decode("utf-8"))
                    for k, v in secrets.items():
                        if v is not None and str(v).strip():
                            os.environ.setdefault(str(k), str(v))
                    loaded = True
                    break
                except Exception as pair_exc:
                    failures.append(f"{secret_file.parent.name}+{key_file.parent.name}:{type(pair_exc).__name__}")
                    continue
            if not loaded:
                raise RuntimeError("no matching region_talk_secrets.enc/region_talk_fernet.key pair; " + "; ".join(failures[:5]))
        except Exception as exc:
            raise RuntimeError(f"failed to load encrypted Region Talk secrets: {type(exc).__name__}") from exc
    return config


class Status:
    def __init__(self) -> None:
        self.client = None
        self.seq = 0
        try:
            from kaggle_status_client import KaggleStatusClient  # type: ignore
            self.client = KaggleStatusClient.discover()
        except Exception:
            self.client = None
        self.events: list[dict[str, Any]] = []

    def event(self, name: str, **payload: Any) -> None:
        self.seq += 1
        clean = {k: v for k, v in payload.items() if v is not None}
        clean.setdefault("event_name", name)
        clean.setdefault("run_id", os.getenv("REGION_TALK_RUN_ID") or "")
        clean.setdefault("event_seq", self.seq)
        clean.setdefault("created_at", utc_now_iso())
        self.events.append(clean)
        print(f"[region-talk] {name} {json.dumps({k:v for k,v in clean.items() if k not in {'token'}}, ensure_ascii=False)[:800]}", flush=True)
        try:
            write_region_talk_business_heartbeat(clean)
        except Exception as exc:
            print(f"[region-talk] business_heartbeat_ydb_failed {type(exc).__name__}: {str(exc)[:160]}", flush=True)
        if self.client:
            try:
                self.client.event(name, phase=clean.get("phase"), status=clean.get("status"), progress=clean)
            except Exception:
                pass


@dataclass
class Seed:
    source_seed_id: str
    platform: str
    source_title: str
    handle: str
    url: str
    source_kind: str
    source_scope_guess: str
    priority: int
    discovered_from: str
    discovered_from_url: str
    why_seeded: str
    expected_value: str
    known_risks: str
    initial_status: str
    monitoring_enabled: bool
    rights_policy: str
    notes: str

    @property
    def source_id(self) -> str:
        return "src_" + stable_hash(self.platform, canonical_source_url(self.platform, self.handle, self.url))

    @property
    def canonical_url(self) -> str:
        return canonical_source_url(self.platform, self.handle, self.url)


def find_seed_file(config: dict[str, Any]) -> Path:
    candidates = []
    for raw in [os.getenv("REGION_TALK_SEED_FILE"), config.get("seed_file"), "seed-sources-v2.csv", "docs/features/region-talk-channel/seed-sources-v2.csv", "seed-sources-v1.csv", "docs/features/region-talk-channel/seed-sources-v1.csv"]:
        if raw:
            candidates.append(Path(str(raw)))
    for root in [Path.cwd(), Path(__file__).resolve().parent, Path("/kaggle/input")]:
        if root.exists():
            candidates.extend(root.rglob("seed-sources-v2.csv"))
            candidates.extend(root.rglob("seed-sources-v1.csv"))
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError("seed-sources-v2.csv/seed-sources-v1.csv not found")


def load_seeds(path: Path) -> list[Seed]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    seeds: list[Seed] = []
    for row in rows:
        seeds.append(Seed(
            source_seed_id=str(row.get("source_seed_id") or "").strip(),
            platform=str(row.get("platform") or "").strip().lower(),
            source_title=str(row.get("source_title") or "").strip(),
            handle=canonical_handle(str(row.get("handle") or "").strip()),
            url=str(row.get("url") or "").strip(),
            source_kind=str(row.get("source_kind") or "").strip(),
            source_scope_guess=str(row.get("source_scope_guess") or "").strip(),
            priority=getenv_int("_NO_SUCH_ENV", int(str(row.get("priority") or "999").strip() or "999")),
            discovered_from=str(row.get("discovered_from") or "").strip(),
            discovered_from_url=str(row.get("discovered_from_url") or "").strip(),
            why_seeded=str(row.get("why_seeded") or "").strip(),
            expected_value=str(row.get("expected_value") or "").strip(),
            known_risks=str(row.get("known_risks") or "").strip(),
            initial_status=str(row.get("initial_status") or "").strip(),
            monitoring_enabled=str(row.get("monitoring_enabled") or "").strip().lower() in {"1","true","yes","on"},
            rights_policy=str(row.get("rights_policy") or "unknown").strip() or "unknown",
            notes=str(row.get("notes") or "").strip(),
        ))
    return seeds



PLACE_LEXICON_FIELDS = [
    "place_id", "canonical_name", "place_type", "municipality", "district_or_okrug", "is_city", "is_settlement",
    "is_tourist_place", "is_nature_place", "is_historical_name", "aliases", "old_names", "latin_aliases",
    "common_misspellings", "geo_scope", "priority_tier", "ambiguity_level", "allowed_for_kaliningrad_scope",
    "requires_context", "reject_if_external_context", "source_url", "source_note",
]
EXTERNAL_REGION_TERMS = [
    "карелия", "алтай", "дагестан", "байкал", "камчатка", "сахалин", "кавказ", "крым", "сочи",
    "санкт-петербург", "петербург", "ленинградская область", "москва", "московская область", "казань",
    "татарстан", "владимир", "суздаль", "ярославль", "кострома", "нижний новгород", "псков",
    "новгород", "мурманск", "териберка", "архангельск", "вологда", "урал", "сибирь", "приморье",
    "владивосток", "краснодарский край", "адыгея", "эльбрус", "чечня", "ингушетия", "осетия",
    "башкирия", "пермский край", "самара", "саратов", "волгоград", "астрахань", "тюмень",
    "челябинск", "челябинская область", "якутия", "саха",
]
EXTERNAL_COUNTRY_TERMS = [
    "польша", "литва", "латвия", "эстония", "германия", "беларусь", "грузия", "армения", "турция",
    "италия", "франция", "испания", "португалия", "норвегия", "финляндия", "швеция", "китай",
]
AD_PROMO_WORDS = [
    "реклама", "промокод", "скидк", "акция", "конкурс", "розыгрыш", "партнёр", "партнер", "спонсор",
    "купить", "заказать", "забронировать", "бронь", "билеты", "билет", "стоимость", "цена", "руб",
    "регистрация", "зарегистр", "анонс", "приглашаем", "приходите", "участвуйте", "географический диктант",
    "диктант", "тур ", "туры", "экскурсия", "экскурсии", "программа мероприятия", "мероприятие состоится",
]
AD_PROMO_HARD_PATTERNS = [
    ("ad_label", r"(?<![а-яa-z])(?:реклама|на правах рекламы|партнерск(?:ий|ая|ое|ие)|партн[её]рский материал)(?![а-яa-z])"),
    ("promo_code", r"(?<![а-яa-z])(?:промокод|скидк[аиуой]?|акци[ияю]|розыгрыш|конкурс)(?![а-яa-z])"),
    ("price_rub", r"(?:\b\d[\d\s]*(?:₽|руб(?:\.|лей|ля|ль)?\b)|(?:стоимость|цена|оплата|оплатить)\b)"),
    ("booking_tickets", r"(?<![а-яa-z])(?:купить|заказать|забронировать|бронь|билеты?|регистраци[яию]|зарегистр(?:ироваться|ируйтесь|ируйся)|приходите|участвуйте)(?![а-яa-z])"),
    ("paid_tour_service", r"(?<![а-яa-z])(?:туры?|экскурси[яию]|гид|туроператор|пут[её]вк[аиу]|программа мероприятия|мероприятие состоится|географический диктант|диктант)(?![а-яa-z])"),
    ("app_download", r"(?:скачайте|установите|приложени[ея]|app store|google play)"),
]
AD_PROMO_POSSIBLE_PATTERNS = [
    ("announcement_tone", r"(?<![а-яa-z])(?:анонс|приглашаем|афиша|состоится)(?![а-яa-z])"),
    ("service_context", r"(?<![а-яa-z])(?:маршрут от|проект|сервис|официальный портал)(?![а-яa-z])"),
]
SUBSTANCE_WORDS = ["маршрут", "дорога", "путь", "добраться", "совет", "полезн", "истори", "место", "что посмотреть", "где", "когда", "почему"]
VISIT_WORDS = ["побывал", "побывали", "посетил", "посетили", "ездили", "поехали", "приехали", "гуляли", "увидели", "запомнил", "запомнилось", "впечатлен", "впечатления"]
EMOTION_WORDS = ["красив", "впечатля", "атмосфер", "удивител", "люблю", "очар", "запомни", "магия", "спокойн", "вдохнов", "вау", "эмоци"]
MEMORABLE_WORDS = ["больше всего", "особенно", "запомни", "неожиданно", "удивило", "главное", "лучшее", "самое", "деталь", "история"]
URL_RE = re.compile(r"(?i)\b(?:https?://|www\.)\S+|\bt\.me/[A-Za-z0-9_+/.-]+|\bvk\.com/[A-Za-z0-9_./-]+")
HANDLE_RE = re.compile(r"(?<!\w)@[A-Za-z0-9_]{4,}")
STATE_FILE_NAME = "region-talk-state.json"

def selected_sources_for_run(seeds: list[Seed], max_sources: int) -> list[Seed]:
    enabled = [s for s in seeds if s.monitoring_enabled]
    fallback = [s for s in seeds if not s.monitoring_enabled]
    ordered = sorted(enabled, key=lambda s: (s.priority, seed_sort_number(s.source_seed_id)))
    seen = {s.source_seed_id for s in ordered}
    for seed in sorted(fallback, key=lambda s: (s.priority, seed_sort_number(s.source_seed_id))):
        if seed.source_seed_id not in seen:
            ordered.append(seed)
            seen.add(seed.source_seed_id)
    return ordered[:max(0, max_sources)]


def seed_from_frontier_row(row: dict[str, Any], idx: int) -> Seed | None:
    platform = str(row.get("platform") or row.get("platform_guess") or "").strip().lower()
    if not platform.startswith("telegram"):
        return None
    url = str(row.get("canonical_url") or row.get("normalized_url") or row.get("recommended_canonical_url") or "").strip()
    handle = canonical_handle(str(row.get("username_or_handle") or row.get("recommended_username") or url).strip())
    if not handle or "/" in handle.lstrip("@"):
        return None
    title = str(row.get("title_guess") or row.get("recommended_title") or row.get("to_source_title") or handle).strip()
    return Seed(
        source_seed_id=f"frontier_dynamic_{idx}_{stable_hash(url or handle, length=8)}",
        platform="telegram",
        source_title=title,
        handle=handle,
        url=url or ("https://t.me/" + handle.lstrip("@")),
        source_kind="frontier_telegram",
        source_scope_guess="travel_frontier",
        priority=2,
        discovered_from=str(row.get("best_discovered_from_source") or row.get("discovered_from_source") or "source_frontier"),
        discovered_from_url=str(row.get("best_discovered_from_post_url") or row.get("telegram_similar_seed_channel") or ""),
        why_seeded="promoted from persistent source frontier queue",
        expected_value="shallow probe / delta history scan",
        known_risks=str(row.get("rejection_or_skip_reason") or ""),
        initial_status="frontier_promoted",
        monitoring_enabled=True,
        rights_policy="unknown",
        notes="dynamic seed from previous Region Talk state; no auto-join",
    )


def frontier_dynamic_seeds(previous_state: dict[str, Any], max_items: int) -> list[Seed]:
    rows: list[dict[str, Any]] = []
    queue = previous_state.get("source_frontier_queue_next")
    if isinstance(queue, dict):
        rows.extend([v for v in queue.values() if isinstance(v, dict)])
    unique = previous_state.get("source_frontier_unique")
    if isinstance(unique, dict):
        rows.extend([v for v in unique.values() if isinstance(v, dict)])
    selected: list[Seed] = []
    seen: set[str] = set()
    rows = sorted(rows, key=lambda r: (
        str(r.get("frontier_stage") or "") not in {"history_due", "probe_due", "unresolved"},
        -float(r.get("frontier_priority") or r.get("source_candidate_score") or 0),
        str(r.get("canonical_url") or r.get("normalized_url") or ""),
    ))
    for i, row in enumerate(rows, start=1):
        seed = seed_from_frontier_row(row, i)
        if not seed:
            continue
        key = seed.canonical_url
        if key in seen:
            continue
        seen.add(key)
        selected.append(seed)
        if len(selected) >= max_items:
            break
    return selected


def seed_from_unified_queue_row(row: dict[str, Any], idx: int) -> Seed | None:
    seed = _source_queue_seed_from_row(row)
    if not seed:
        return None
    platform = str(seed.get("platform") or "")
    handle = canonical_handle(str(seed.get("handle") or seed.get("source_url") or ""))
    url = str(seed.get("source_url") or seed.get("canonical_url") or "")
    title = str(seed.get("source_title") or handle or url).strip()
    if platform == "telegram" and (not handle or "/" in handle.lstrip("@")):
        return None
    return Seed(
        source_seed_id=f"unified_queue_{idx}_{stable_hash(seed.get('canonical_source_key'), length=8)}",
        platform=platform,
        source_title=title,
        handle=handle,
        url=url,
        source_kind="unified_source_queue",
        source_scope_guess="canonical_queue",
        priority=0,
        discovered_from=str(row.get("added_from") or "unified_source_queue"),
        discovered_from_url=str(row.get("source_url") or ""),
        why_seeded="selected from canonical single Telegram/VK source queue after cursor",
        expected_value="queue-ordered history scan",
        known_risks=str(row.get("last_scan_status") or ""),
        initial_status=str(row.get("source_queue_status") or "pending_scan"),
        monitoring_enabled=True,
        rights_policy="unknown",
        notes=f"queue_order={row.get('queue_order')}; cursor_marker={row.get('cursor_marker')}",
    )


def unified_queue_dynamic_seeds(previous_state: dict[str, Any], max_items: int) -> list[Seed]:
    rows = _previous_rows_dict(previous_state.get("unified_source_queue") or previous_state.get("canonical_source_queue"))
    if not rows:
        return []
    cursor = int(previous_state.get("unified_source_queue_cursor_position") or previous_state.get("canonical_source_cursor_position") or 0)
    rows = [
        r for r in rows
        if normalize_source_platform(str(r.get("platform") or ""), str(r.get("source_url") or r.get("canonical_url") or "")) in {"telegram", "vk"}
    ]
    rows = sorted(rows, key=lambda r: (
        int(r.get("queue_order") or 999999999) <= cursor,
        str(r.get("source_queue_status") or "") not in {"pending_scan", "needs_rescan_or_retry"},
        int(r.get("queue_order") or 999999999),
    ))
    selected: list[Seed] = []
    seen: set[str] = set()
    for i, row in enumerate(rows, start=1):
        seed = seed_from_unified_queue_row(row, i)
        if not seed:
            continue
        key = seed.canonical_url
        if key in seen:
            continue
        seen.add(key)
        selected.append(seed)
        if len(selected) >= max_items:
            break
    return selected


def source_status_row(seed: Seed, status_value: str, **extra: Any) -> dict[str, Any]:
    return {
        **asdict(seed),
        "source_id": seed.source_id,
        "canonical_url": seed.canonical_url,
        "canonical_source_key": canonical_source_key(seed.platform, seed.handle, seed.canonical_url),
        "fetch_status": status_value,
        "posts_scanned": 0,
        "source_selected_this_run": "true",
        "source_cursor_strategy": "fresh_first_min_post_date_then_anchor_search",
        **extra,
    }


def state_file_candidates(output_dir: Path) -> list[Path]:
    out: list[Path] = []
    if os.getenv("REGION_TALK_STATE_FILE"):
        out.append(Path(str(os.getenv("REGION_TALK_STATE_FILE"))))
    cwd = Path.cwd()
    out.extend([
        output_dir.parent.parent / "state" / STATE_FILE_NAME if len(output_dir.parents) >= 2 else output_dir.parent / STATE_FILE_NAME,
        output_dir.parent / STATE_FILE_NAME,
    ])
    try:
        if output_dir.resolve().is_relative_to(cwd.resolve()):
            out.append(cwd / "artifacts" / "region-talk" / "state" / STATE_FILE_NAME)
    except Exception:
        pass
    inp = Path("/kaggle/input")
    if inp.exists():
        out.extend(inp.rglob(STATE_FILE_NAME))
    seen: set[str] = set()
    unique: list[Path] = []
    for p in out:
        key = str(p)
        if key not in seen:
            seen.add(key)
            unique.append(p)
    return unique


def parse_iso_datetime(value: Any) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def iso_after_seconds(seconds: int) -> str:
    return datetime.fromtimestamp(time.time() + max(0, seconds), timezone.utc).isoformat()




def region_talk_state_backend_requested() -> str:
    return (os.getenv("REGION_TALK_STATE_BACKEND") or "json").strip().lower() or "json"


def ydb_config_status() -> dict[str, str]:
    endpoint = (os.getenv("REGION_TALK_YDB_ENDPOINT") or "").strip()
    database = (os.getenv("REGION_TALK_YDB_DATABASE") or "").strip()
    if "?database=" in endpoint:
        endpoint_part, database_part = endpoint.split("?database=", 1)
        endpoint = endpoint_part
        if not database:
            database = database_part
    endpoint = endpoint.rstrip("/")
    namespace = (os.getenv("REGION_TALK_YDB_NAMESPACE") or "region_talk").strip() or "region_talk"
    missing = [k for k, v in {"REGION_TALK_YDB_ENDPOINT": endpoint, "REGION_TALK_YDB_DATABASE": database}.items() if not v]
    return {"endpoint": endpoint, "database": database, "namespace": namespace, "missing": ",".join(missing)}


def ydb_table_plan() -> str:
    return ";".join([
        "region_talk_state_kv(compact)", "region_talk_sources(compact)", "region_talk_processed_posts(compact)",
        "region_talk_image_metrics(compact)", "region_talk_run_metrics(compact)",
    ])


def ydb_table_name(suffix: str = "state_kv") -> str:
    namespace = re.sub(r"[^A-Za-z0-9_]+", "_", (os.getenv("REGION_TALK_YDB_NAMESPACE") or "region_talk").strip() or "region_talk").strip("_") or "region_talk"
    return f"{namespace}_{suffix}"


def compact_scalar(value: Any, *, max_len: int = 500) -> Any:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    text = str(value)
    return text if len(text) <= max_len else text[:max_len]


def compact_record(row: dict[str, Any], fields: Iterable[str], *, max_len: int = 500) -> dict[str, Any]:
    return {field: compact_scalar(row.get(field), max_len=max_len) for field in fields if row.get(field) not in (None, "", [], {})}


def build_queue_cursor_state(state: dict[str, Any], source_queue: dict[str, Any], image_queue: dict[str, Any]) -> dict[str, Any]:
    source_rows = [v for v in source_queue.values() if isinstance(v, dict)]
    image_rows = [v for v in image_queue.values() if isinstance(v, dict)]
    updated_at = state.get("updated_at") or datetime.now(timezone.utc).isoformat()
    return {
        "source": {
            "queue_name": "unified_source_queue",
            "cursor_position": int(state.get("unified_source_queue_cursor_position") or 0),
            "cursor_key": state.get("unified_source_queue_cursor_key") or "",
            "total": len(source_rows),
            "pending_total": sum(1 for r in source_rows if r.get("source_queue_status") == "pending_scan"),
            "processed_total": sum(1 for r in source_rows if str(r.get("source_queue_status") or "").startswith("processed")),
            "retry_total": sum(1 for r in source_rows if r.get("source_queue_status") == "needs_rescan_or_retry"),
            "updated_at": updated_at,
        },
        "image": {
            "queue_name": "image_candidate_queue",
            "cursor_position": int(state.get("image_candidate_queue_cursor_position") or 0),
            "cursor_key": state.get("image_candidate_queue_cursor_key") or "",
            "total": len(image_rows),
            "needs_actual_fetch_total": sum(1 for r in image_rows if r.get("image_queue_status") == "needs_actual_image_fetch"),
            "in_progress_total": sum(1 for r in image_rows if r.get("image_queue_status") == "image_analysis_in_progress"),
            "actual_scored_total": sum(1 for r in image_rows if r.get("image_queue_status") == "actual_scored"),
            "updated_at": updated_at,
        },
    }


SOURCE_STATE_FIELDS = [
    "source_id", "source_seed_id", "canonical_source_key",
    "platform", "handle", "username_or_handle", "source_title", "canonical_url", "normalized_url",
    "source_url", "fetch_status", "vk_wall_probe_status",
    "posts_scanned", "last_seen_post_date", "monitor_priority_score", "source_quality_score",
    "next_action", "updated_at", "last_run_id",
]
SOURCE_QUEUE_STATE_FIELDS = [
    "source_queue_id", "queue_order", "display_order", "source_queue_status", "previous_source_queue_status",
    "status_changed_this_run", "last_status_changed_at", "status_color_hint", "row_fill_color",
    "cursor_marker", "is_after_cursor", "added_at", "added_from", "first_seen_run_id",
    "last_processed_at", "last_scan_run_id", "last_scan_status", "platform",
    "source_url", "canonical_url", "canonical_source_key", "source_title", "handle",
    "posts_scanned", "ko_posts_found", "candidate_posts_found",
    "actual_images_scored_count", "avg_actual_image_score", "low_actual_image_count",
    "source_image_quality_status", "source_image_quality_min_actual_scored",
    "source_image_quality_min_avg_score", "monitoring_exclusion_reason",
    "next_action", "queue_item_updated_at", "source_visual_rollup_updated_at", "source_visual_rollup_run_id",
]
IMAGE_QUEUE_STATE_FIELDS = [
    "image_queue_id", "image_queue_order", "display_order", "image_queue_status", "previous_image_queue_status",
    "status_changed_this_run", "last_status_changed_at", "status_color_hint", "row_fill_color",
    "cursor_marker", "is_after_cursor", "selected_for_next_image_batch", "image_queue_batch_target",
    "added_at", "added_from", "last_attempt_run_id", "last_attempt_at", "lease_run_id", "lease_at",
    "last_image_diag_run_id", "last_image_diag_stage", "last_image_diag_at", "post_id",
    "post_url", "platform_post_key", "source_id", "source_title", "source_url", "post_date",
    "text_region_confirmation_status", "kaliningrad_oblast_only_scope", "kaliningrad_mention_role",
    "matched_place_names", "external_geo_mentions", "mentioned_external_regions",
    "is_ad_or_promo", "current_stage", "current_lifecycle_status", "vector_gate_status",
    "vector_content_type", "candidate_score", "overall_media_score", "postcardness_score", "aesthetic_score",
    "technical_quality_score", "publication_safety_score", "cv_overall_media_score", "clip_postcardness_score",
    "laion_aesthetic_score", "nima_quality_score", "final_visual_score", "final_visual_status",
    "model_disagreement_score", "image_width", "image_height",
    "image_model_input_type", "image_model_type", "media_acquisition_status", "media_fetch_status",
    "media_acquisition_error_type", "media_fetch_error", "images_scored_actual_count", "next_action",
]
POST_STATE_FIELDS = [
    "post_id", "source_id", "source_title", "platform", "platform_post_key", "post_url", "post_date",
    "current_stage", "fresh_enough", "kaliningrad_oblast_only_scope", "matched_place_names",
    "candidate_score", "media_score", "overall_media_score", "image_quality_bucket",
    "image_model_input_type", "image_download_status", "image_reviewable", "image_publication_ready",
    "vector_gate_status", "llm_gate_status", "llm_decision", "llm_reason", "content_type",
    "is_forwarded_or_repost", "forwarded_from_post_url",
]
POST_LIVE_STATE_FIELDS = POST_STATE_FIELDS + [
    "run_id", "updated_at", "last_seen_run_id", "online_update_stage", "has_media",
    "media_count", "text_hash", "text_excerpt_hash", "source_url", "handle",
    "fetch_status", "post_observation_status",
]
CANDIDATE_MEMORY_STATE_FIELDS = [
    "candidate_memory_id", "post_id", "source_id", "source_title", "platform", "post_url", "post_date",
    "current_stage", "current_lifecycle_status", "best_candidate_score_ever", "best_media_score_ever",
    "publication_story_score", "nonlocal_value_score", "manual_decision", "last_seen_run_id",
    "kaliningrad_oblast_only_scope", "kaliningrad_mention_role", "matched_place_names",
    "external_geo_mentions", "mentioned_external_regions", "is_ad_or_promo",
    "vector_gate_status", "vector_content_type", "first_seen_run_id", "seen_run_count",
    "short_summary", "why_selected", "final_verifier_status", "final_verifier_decision",
    "final_verifier_reason", "final_verifier_model", "llm_gate_status", "llm_decision",
    "image_model_input_type", "image_queue_status", "overall_media_score", "postcardness_score",
    "aesthetic_score", "image_publication_ready",
]
PUBLICATION_CANDIDATE_STATE_FIELDS = [
    "publication_candidate_id", "publication_goal_id", "publication_rank", "publication_candidate_status",
    "post_id", "post_url", "source_id", "source_title", "source_geo_class", "source_topic_class",
    "post_date", "short_summary", "why_selected", "why_not_selected", "matched_place_names",
    "candidate_score", "publication_score", "visual_score", "text_story_score", "diversity_penalty",
    "overall_media_score", "postcardness_score", "aesthetic_score", "image_model_input_type",
    "image_queue_status", "vector_gate_status", "vector_content_type", "vector_positive_score",
    "publication_llm_status", "publication_llm_decision", "publication_llm_reason",
    "publication_llm_model", "visual_confirmation_source", "goal_stop_candidate", "sent_to_chat",
    "sent_message_id", "created_at", "last_seen_run_id", "last_confirmed_run_id",
]
PUBLICATION_CONFIRMED_STATUSES = {"llm_confirmed", "sent_to_chat", "accepted_for_publication"}


def compact_region_talk_state_for_ydb(state: dict[str, Any]) -> dict[str, Any]:
    """Keep only durable, valuable state in YDB; no raw post text or media bytes.

    Telegram/VK posts can be re-fetched by URL/key, so the YDB sidecar stores
    cursors, source/post links, candidate lifecycle, run metrics and image
    scoring metrics. The full JSON backup remains a local/Kaggle artifact only.
    """
    posts = state.get("posts") if isinstance(state.get("posts"), dict) else {}
    if not posts:
        posts = state.get("processed_posts") if isinstance(state.get("processed_posts"), dict) else {}
    sources = state.get("region_talk_sources") if isinstance(state.get("region_talk_sources"), dict) else {}
    if not sources:
        sources = state.get("sources") if isinstance(state.get("sources"), dict) else {}
    if not sources:
        sources = state.get("source_frontier_unique") if isinstance(state.get("source_frontier_unique"), dict) else {}
    candidate_memory = state.get("candidate_memory") if isinstance(state.get("candidate_memory"), dict) else {}
    publication_candidates = state.get("publication_candidate_queue") if isinstance(state.get("publication_candidate_queue"), dict) else {}
    max_posts = getenv_int("REGION_TALK_YDB_MAX_POST_ROWS", 20000)
    max_sources = getenv_int("REGION_TALK_YDB_MAX_SOURCE_ROWS", 5000)
    max_candidates = getenv_int("REGION_TALK_YDB_MAX_CANDIDATE_ROWS", 5000)
    compact_posts = {
        str(k): compact_record(v, POST_STATE_FIELDS, max_len=700)
        for k, v in list(posts.items())[-max_posts:]
        if isinstance(v, dict)
    }
    compact_sources = {
        str(k): compact_record(v, SOURCE_STATE_FIELDS, max_len=700)
        for k, v in list(sources.items())[-max_sources:]
        if isinstance(v, dict)
    }
    compact_candidates = {
        str(k): compact_record(v, CANDIDATE_MEMORY_STATE_FIELDS, max_len=900)
        for k, v in list(candidate_memory.items())[-max_candidates:]
        if isinstance(v, dict)
    }
    source_queue = state.get("unified_source_queue") if isinstance(state.get("unified_source_queue"), dict) else {}
    if not source_queue and isinstance(state.get("canonical_source_queue"), dict):
        source_queue = state.get("canonical_source_queue") or {}
    if not source_queue:
        synthetic: dict[str, dict[str, Any]] = {}
        synthetic_rows: list[dict[str, Any]] = []
        for value in [sources, state.get("source_frontier_queue_next"), state.get("source_frontier_unique")]:
            if isinstance(value, dict):
                synthetic_rows.extend([v for v in value.values() if isinstance(v, dict)])
            elif isinstance(value, list):
                synthetic_rows.extend([v for v in value if isinstance(v, dict)])
        for row in synthetic_rows:
            seed = _source_queue_seed_from_row(row) if "_source_queue_seed_from_row" in globals() else None
            if not seed:
                continue
            key = str(seed.get("canonical_source_key") or "")
            if not key or key in synthetic:
                continue
            order = len(synthetic) + 1
            synthetic[key] = {
                **seed,
                "source_queue_id": "srcq_" + stable_hash(key),
                "queue_order": order,
                "source_queue_status": row.get("source_queue_status") or row.get("scan_status") or row.get("frontier_status") or "pending_scan",
                "added_at": row.get("added_at") or row.get("updated_at") or state.get("updated_at") or "",
                "added_from": row.get("added_from") or row.get("discovery_types") or row.get("discovery_type") or "legacy_ydb_queue_migration",
                "last_scan_status": row.get("last_scan_status") or row.get("fetch_status") or row.get("frontier_status") or "",
                "posts_scanned": row.get("posts_scanned") or 0,
                "next_action": row.get("next_action") or "scan_when_cursor_reaches_source",
            }
        source_queue = synthetic
    image_queue = state.get("image_candidate_queue") if isinstance(state.get("image_candidate_queue"), dict) else {}
    return {
        "run_id": state.get("run_id"),
        "state_schema_version": "region-talk-ydb-compact-v3",
        "queue_contract_version": "ydb_row_level_unified_source_queue_v2_and_image_candidate_queue_v2",
        "full_state_schema_version": state.get("state_schema_version"),
        "updated_at": state.get("updated_at"),
        "all_time_metrics": state.get("all_time_metrics") or {},
        "source_cursors": state.get("source_cursors") or {},
        "telegram_entity_cache": state.get("telegram_entity_cache") or {},
        "telegram_cooldowns": state.get("telegram_cooldowns") or {},
        "sources": compact_sources,
        "processed_posts": compact_posts,
        "candidate_memory": compact_candidates,
        "publication_goal": state.get("publication_goal") or {},
        "publication_candidate_queue": {
            str(k): compact_record(v, PUBLICATION_CANDIDATE_STATE_FIELDS, max_len=900)
            for k, v in list(publication_candidates.items())[:max_candidates]
            if isinstance(v, dict)
        },
        "unified_source_queue_cursor_position": state.get("unified_source_queue_cursor_position", 0),
        "unified_source_queue_cursor_key": state.get("unified_source_queue_cursor_key", ""),
        "unified_source_queue": {
            str(k): compact_record(v, SOURCE_QUEUE_STATE_FIELDS, max_len=700)
            for k, v in list(source_queue.items())[:max_sources]
            if isinstance(v, dict)
        },
        "image_candidate_queue_cursor_position": state.get("image_candidate_queue_cursor_position", 0),
        "image_candidate_queue_cursor_key": state.get("image_candidate_queue_cursor_key", ""),
        "image_candidate_queue": {
            str(k): compact_record(v, IMAGE_QUEUE_STATE_FIELDS, max_len=700)
            for k, v in list(image_queue.items())[:max_candidates]
            if isinstance(v, dict)
        },
        "queue_cursors": build_queue_cursor_state(state, source_queue, image_queue),
    }


def ensure_ydb_module() -> Any:
    try:
        import ydb  # type: ignore
        return ydb
    except Exception:
        if getenv_bool("REGION_TALK_AUTO_INSTALL", True):
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "ydb[yc]"])
            import ydb  # type: ignore
            return ydb
        raise


def ydb_credentials(ydb: Any) -> Any:
    token = (os.getenv("REGION_TALK_YDB_IAM_TOKEN") or os.getenv("YC_IAM_TOKEN") or os.getenv("YDB_ACCESS_TOKEN") or "").strip()
    if token:
        return ydb.AccessTokenCredentials(token)
    key_json = (os.getenv("REGION_TALK_YDB_SERVICE_ACCOUNT_KEY_JSON") or "").strip()
    if key_json:
        import tempfile
        import ydb.iam  # type: ignore
        fd, path = tempfile.mkstemp(prefix="region-talk-ydb-sa-", suffix=".json")
        os.close(fd)
        Path(path).write_text(key_json, encoding="utf-8")
        return ydb.iam.ServiceAccountCredentials.from_file(path)
    if os.getenv("YDB_USER"):
        return ydb.StaticCredentials.from_user_password(os.getenv("YDB_USER"), os.getenv("YDB_PASSWORD", ""))
    return ydb.credentials_from_env_variables()


def ydb_connect() -> tuple[Any, Any, dict[str, str]]:
    cfg = ydb_config_status()
    ydb = ensure_ydb_module()
    driver = ydb.Driver(endpoint=cfg["endpoint"], database=cfg["database"], credentials=ydb_credentials(ydb))
    driver.wait(timeout=getenv_int("REGION_TALK_YDB_CONNECT_TIMEOUT_SECONDS", 20), fail_fast=True)
    return ydb, driver, cfg


def ydb_kv_table_path(cfg: dict[str, str]) -> str:
    return cfg["database"].rstrip("/") + "/" + ydb_table_name("state_kv")


def ensure_ydb_kv_table(ydb: Any, session: Any, table_path: str) -> None:
    try:
        session.describe_table(table_path)
        return
    except Exception:
        pass
    desc = (
        ydb.TableDescription()
        .with_column(ydb.Column("pk", ydb.OptionalType(ydb.PrimitiveType.Utf8)))
        .with_column(ydb.Column("kind", ydb.OptionalType(ydb.PrimitiveType.Utf8)))
        .with_column(ydb.Column("payload_json", ydb.OptionalType(ydb.PrimitiveType.Json)))
        .with_column(ydb.Column("updated_at", ydb.OptionalType(ydb.PrimitiveType.Utf8)))
        .with_primary_key("pk")
    )
    session.create_table(table_path, desc)



_YDB_BUSINESS_HEARTBEAT_CACHE: dict[str, Any] = {"ydb": None, "driver": None, "pool": None, "table_path": "", "last_alive_at": 0.0}


def _compact_business_payload(payload: dict[str, Any]) -> dict[str, Any]:
    allowed = [
        "run_id", "event_seq", "event_name", "created_at", "phase", "status", "progress_label",
        "sources_done", "sources_total", "source_index", "source_id", "current_source_title",
        "current_source_url", "current_source_handle", "current_source_platform", "fetch_status",
        "telegram_resolve_status", "posts_scanned", "last_seen_post_date", "history_fetch_runtime_seconds",
        "posts_fetched", "posts_scanned", "posts_for_scoring", "posts_to_score", "posts_scored",
        "posts_deferred", "sources_scanned", "candidates_created", "favorites_created",
        "current_run_reviewable_candidates", "state_backend", "ydb_read_status", "ydb_write_status",
        "ydb_state_mode", "vk_wall_probe_status", "image_queue_total", "image_queue_pruned_non_region_previous",
        "image_queue_rejected_non_region_inputs", "image_queue_text_region_confirmed_total",
        "fetch_error_code", "fetch_error_message", "xlsx",
    ]
    out = {k: compact_scalar(payload.get(k), max_len=900) for k in allowed if payload.get(k) not in (None, "", [], {})}
    out.setdefault("run_id", os.getenv("REGION_TALK_RUN_ID") or "")
    out.setdefault("updated_at", utc_now_iso())
    return out


def _business_heartbeat_enabled() -> bool:
    if os.getenv("REGION_TALK_YDB_BUSINESS_HEARTBEAT") is not None:
        return getenv_bool("REGION_TALK_YDB_BUSINESS_HEARTBEAT", True)
    return (os.getenv("REGION_TALK_STATE_BACKEND") or "").strip().lower() == "ydb"


def _get_business_heartbeat_pool() -> tuple[Any, Any, Any, str]:
    cache = _YDB_BUSINESS_HEARTBEAT_CACHE
    if cache.get("pool") is not None and cache.get("ydb") is not None and cache.get("table_path"):
        return cache["ydb"], cache["driver"], cache["pool"], str(cache["table_path"])
    ydb, driver, cfg = ydb_connect()
    table_path = ydb_kv_table_path(cfg)
    pool = ydb.SessionPool(driver)
    cache.update({"ydb": ydb, "driver": driver, "pool": pool, "table_path": table_path})
    return ydb, driver, pool, table_path


def write_region_talk_business_heartbeat(payload: dict[str, Any]) -> None:
    """Persist online business-stage progress to YDB while Kaggle is still running.

    This is deliberately separate from the final compact state snapshot: it makes
    a running notebook observable even when Kaggle live logs are empty.
    """
    if not _business_heartbeat_enabled():
        return
    event_name = str(payload.get("event_name") or "")
    now_monotonic = time.monotonic()
    min_interval = max(0, getenv_int("REGION_TALK_YDB_HEARTBEAT_MIN_INTERVAL_SECONDS", 5))
    if event_name == "alive" and now_monotonic - float(_YDB_BUSINESS_HEARTBEAT_CACHE.get("last_alive_at") or 0) < min_interval:
        return
    if event_name == "alive":
        _YDB_BUSINESS_HEARTBEAT_CACHE["last_alive_at"] = now_monotonic
    compact = _compact_business_payload(payload)
    run_id = str(compact.get("run_id") or os.getenv("REGION_TALK_RUN_ID") or "unknown-run")
    seq = int(compact.get("event_seq") or 0)
    updated_at = str(compact.get("created_at") or utc_now_iso())
    ydb, _driver, pool, table_path = _get_business_heartbeat_pool()

    def op(session: Any) -> None:
        ensure_ydb_kv_table(ydb, session, table_path)
        ydb_upsert_json(session, ydb, table_path, "business_heartbeat:" + run_id, "business_heartbeat", compact, updated_at)
        ydb_upsert_json(session, ydb, table_path, "latest_business_heartbeat", "business_heartbeat", compact, updated_at)
        if getenv_bool("REGION_TALK_YDB_BUSINESS_EVENT_LOG", True):
            ydb_upsert_json(session, ydb, table_path, f"business_event:{run_id}:{seq:06d}", "business_event", compact, updated_at)

    pool.retry_operation_sync(op)


def _online_source_key(row: dict[str, Any]) -> str:
    platform = str(row.get("platform") or row.get("platform_guess") or "telegram")
    handle = str(row.get("handle") or row.get("recommended_username") or row.get("username_or_handle") or "")
    url = str(row.get("canonical_url") or row.get("source_url") or row.get("normalized_url") or row.get("recommended_canonical_url") or row.get("raw_url") or "")
    key = str(row.get("canonical_source_key") or "")
    return key or canonical_source_key(platform, handle, url) or stable_hash(platform, handle, url, row.get("source_id") or row.get("source_candidate_id") or "")


def _online_source_payload(row: dict[str, Any], *, run_id: str, stage: str, status: str | None = None) -> dict[str, Any]:
    now = utc_now_iso()
    platform = str(row.get("platform") or row.get("platform_guess") or "telegram")
    url = str(row.get("canonical_url") or row.get("source_url") or row.get("normalized_url") or row.get("recommended_canonical_url") or row.get("raw_url") or "")
    handle = str(row.get("handle") or row.get("recommended_username") or row.get("username_or_handle") or "")
    title = str(row.get("source_title") or row.get("resolved_title") or row.get("recommended_title") or row.get("discovered_from_source") or handle or url)
    fetch_status = str(status or row.get("fetch_status") or row.get("candidate_source_status") or row.get("frontier_status") or row.get("method_status") or "observed")
    payload = {
        "run_id": run_id,
        "updated_at": now,
        "last_seen_run_id": run_id,
        "online_update_stage": stage,
        "canonical_source_key": _online_source_key(row),
        "source_queue_id": "srcq_" + stable_hash(_online_source_key(row)),
        "source_id": row.get("source_id") or row.get("source_candidate_id") or ("src_" + stable_hash(platform, handle or url)),
        "source_candidate_id": row.get("source_candidate_id") or "",
        "platform": platform,
        "handle": handle,
        "source_title": title,
        "resolved_title": row.get("resolved_title") or title,
        "canonical_url": url,
        "source_url": url,
        "fetch_status": fetch_status,
        "queue_status": "candidate" if row.get("source_candidate_id") else ("scanned" if fetch_status == "ok" else "skipped_or_rejected"),
        "frontier_status": row.get("frontier_status") or "",
        "frontier_action": row.get("frontier_action") or "",
        "frontier_reason": row.get("frontier_reason") or row.get("source_probe_reason") or "",
        "discovery_type": row.get("discovery_type") or "",
        "edge_type": row.get("edge_type") or "",
        "posts_scanned": row.get("posts_scanned") or 0,
        "ko_posts_found": row.get("ko_posts_found") or 0,
        "candidate_posts_found": row.get("candidate_posts_found") or 0,
        "last_seen_post_date": row.get("last_seen_post_date") or "",
        "telegram_resolve_status": row.get("telegram_resolve_status") or "",
        "fetch_error_code": row.get("fetch_error_code") or row.get("last_resolve_error_code") or row.get("method_error_code") or "",
        "fetch_error_message": row.get("fetch_error_message") or row.get("last_resolve_error_message_short") or row.get("method_error_message_short") or "",
        "confidence": row.get("confidence") or row.get("similarity_edge_confidence") or "",
        "source_probe_reason": row.get("source_probe_reason") or row.get("frontier_reason") or "",
    }
    return compact_record(payload, SOURCE_QUEUE_STATE_FIELDS + [
        "run_id", "updated_at", "last_seen_run_id", "online_update_stage", "queue_status",
        "source_candidate_id", "platform", "handle", "resolved_title", "discovery_type",
        "edge_type", "frontier_status", "frontier_action", "frontier_reason", "confidence",
    ], max_len=700)


def write_region_talk_online_source_item(row: dict[str, Any], *, run_id: str, stage: str, status: str | None = None) -> None:
    """Upsert source/frontier progress while discovery is still running.

    Final state snapshots are still written at report end, but the product
    discovery loop needs live YDB visibility for newly seen / skipped channels.
    """
    if (os.getenv("REGION_TALK_STATE_BACKEND") or "").strip().lower() != "ydb":
        return
    if not getenv_bool("REGION_TALK_YDB_ONLINE_SOURCE_WRITES", True):
        return
    try:
        payload = _online_source_payload(row, run_id=run_id, stage=stage, status=status)
        key = str(payload.get("canonical_source_key") or "")
        if not key:
            return
        updated_at = str(payload.get("updated_at") or utc_now_iso())
        ydb, _driver, pool, table_path = _get_business_heartbeat_pool()

        def op(session: Any) -> None:
            ensure_ydb_kv_table(ydb, session, table_path)
            ydb_upsert_json(session, ydb, table_path, "source_queue_item:" + key, "source_queue_item", payload, updated_at)
            ydb_upsert_json(session, ydb, table_path, "source_status_item:" + key, "source_status_item", payload, updated_at)
            ydb_upsert_json(session, ydb, table_path, "online_source_item:" + key, "online_source_item", payload, updated_at)

        pool.retry_operation_sync(op)
    except Exception as exc:
        print(f"[region-talk] online_source_ydb_failed {type(exc).__name__}: {str(exc)[:160]}", flush=True)


def _online_post_payload(row: dict[str, Any], *, run_id: str, stage: str, status: str | None = None) -> dict[str, Any]:
    now = utc_now_iso()
    text = str(row.get("text") or row.get("text_excerpt") or "")
    payload = {
        **row,
        "run_id": run_id,
        "updated_at": now,
        "last_seen_run_id": run_id,
        "online_update_stage": stage,
        "post_observation_status": status or row.get("current_stage") or "fetched",
        "text_hash": stable_hash(text) if text else row.get("text_hash", ""),
        "text_excerpt_hash": stable_hash(str(row.get("text_excerpt") or "")) if row.get("text_excerpt") else "",
    }
    payload.pop("text", None)
    payload.pop("raw", None)
    return compact_record(payload, POST_LIVE_STATE_FIELDS, max_len=900)


def write_region_talk_online_post_item(row: dict[str, Any], *, run_id: str, stage: str, status: str | None = None) -> None:
    if (os.getenv("REGION_TALK_STATE_BACKEND") or "").strip().lower() != "ydb":
        return
    if not getenv_bool("REGION_TALK_YDB_ONLINE_POST_WRITES", True):
        return
    try:
        payload = _online_post_payload(row, run_id=run_id, stage=stage, status=status)
        key = str(payload.get("post_id") or payload.get("platform_post_key") or payload.get("post_url") or "")
        if not key:
            return
        updated_at = str(payload.get("updated_at") or utc_now_iso())
        ydb, _driver, pool, table_path = _get_business_heartbeat_pool()

        def op(session: Any) -> None:
            ensure_ydb_kv_table(ydb, session, table_path)
            ydb_upsert_json(session, ydb, table_path, "post_live_item:" + key, "post_live_item", payload, updated_at)
            ydb_upsert_json(session, ydb, table_path, "processed_post_item:" + key, "processed_post_item", payload, updated_at)

        pool.retry_operation_sync(op)
    except Exception as exc:
        print(f"[region-talk] online_post_ydb_failed {type(exc).__name__}: {str(exc)[:160]}", flush=True)


def write_region_talk_online_candidate_items(rows: list[dict[str, Any]], *, run_id: str, stage: str) -> int:
    if (os.getenv("REGION_TALK_STATE_BACKEND") or "").strip().lower() != "ydb":
        return 0
    if not rows or not getenv_bool("REGION_TALK_YDB_ONLINE_CANDIDATE_WRITES", True):
        return 0
    try:
        now = utc_now_iso()
        items: list[tuple[str, str, dict[str, Any]]] = []
        for row in rows:
            payload = compact_record({**row, "run_id": run_id, "updated_at": now, "last_seen_run_id": run_id, "online_update_stage": stage}, CANDIDATE_MEMORY_STATE_FIELDS + ["run_id", "updated_at", "online_update_stage"], max_len=900)
            key = str(payload.get("candidate_memory_id") or payload.get("post_id") or payload.get("post_url") or "")
            if key:
                items.append(("candidate_memory_item:" + key.replace("candidate_memory_item:", ""), "candidate_memory_item", payload))
        if not items:
            return 0
        ydb, _driver, pool, table_path = _get_business_heartbeat_pool()

        def op(session: Any) -> int:
            ensure_ydb_kv_table(ydb, session, table_path)
            return ydb_upsert_json_many(session, ydb, table_path, items, now, chunk_size=getenv_int("REGION_TALK_YDB_ROW_UPSERT_CHUNK_SIZE", 100))

        return int(pool.retry_operation_sync(op) or 0)
    except Exception as exc:
        print(f"[region-talk] online_candidate_ydb_failed {type(exc).__name__}: {str(exc)[:160]}", flush=True)
        return 0


def write_region_talk_online_queue_items(rows: list[dict[str, Any]], *, kind: str, id_fields: list[str], fields: list[str], run_id: str, stage: str) -> int:
    if (os.getenv("REGION_TALK_STATE_BACKEND") or "").strip().lower() != "ydb":
        return 0
    if not rows or not getenv_bool("REGION_TALK_YDB_ONLINE_QUEUE_WRITES", True):
        return 0
    try:
        now = utc_now_iso()
        items: list[tuple[str, str, dict[str, Any]]] = []
        for row in rows:
            payload = compact_record({**row, "run_id": run_id, "updated_at": now, "last_seen_run_id": run_id, "online_update_stage": stage}, fields + ["run_id", "updated_at", "last_seen_run_id", "online_update_stage"], max_len=900)
            key = ""
            for f in id_fields:
                if payload.get(f):
                    key = str(payload.get(f))
                    break
            if key:
                items.append((kind + ":" + key.replace(kind + ":", ""), kind, payload))
        if not items:
            return 0
        ydb, _driver, pool, table_path = _get_business_heartbeat_pool()

        def op(session: Any) -> int:
            ensure_ydb_kv_table(ydb, session, table_path)
            return ydb_upsert_json_many(session, ydb, table_path, items, now, chunk_size=getenv_int("REGION_TALK_YDB_ROW_UPSERT_CHUNK_SIZE", 100))

        return int(pool.retry_operation_sync(op) or 0)
    except Exception as exc:
        print(f"[region-talk] online_queue_ydb_failed kind={kind} {type(exc).__name__}: {str(exc)[:160]}", flush=True)
        return 0


def write_region_talk_online_stats(payload: dict[str, Any]) -> None:
    if (os.getenv("REGION_TALK_STATE_BACKEND") or "").strip().lower() != "ydb":
        return
    if not getenv_bool("REGION_TALK_YDB_ONLINE_SOURCE_WRITES", True):
        return
    try:
        compact = compact_record({**payload, "updated_at": utc_now_iso()}, [
            "run_id", "updated_at", "phase", "status", "progress_label",
            "sources_seen_this_run", "sources_done", "sources_total",
            "sources_rejected_or_skipped_this_run", "sources_with_ko_posts_this_run",
            "ko_candidate_posts_this_run", "image_strong_this_run",
        ], max_len=700)
        run_id = str(compact.get("run_id") or os.getenv("REGION_TALK_RUN_ID") or "unknown-run")
        ydb, _driver, pool, table_path = _get_business_heartbeat_pool()
        updated_at = str(compact.get("updated_at") or utc_now_iso())

        def op(session: Any) -> None:
            ensure_ydb_kv_table(ydb, session, table_path)
            ydb_upsert_json(session, ydb, table_path, "online_stats:" + run_id, "online_stats", compact, updated_at)
            ydb_upsert_json(session, ydb, table_path, "online_stats:latest", "online_stats", compact, updated_at)

        pool.retry_operation_sync(op)
    except Exception as exc:
        print(f"[region-talk] online_stats_ydb_failed {type(exc).__name__}: {str(exc)[:160]}", flush=True)


def append_source_row_online(source_rows: list[dict[str, Any]], row: dict[str, Any], *, run_id: str, stage: str, sources_total: int | None = None) -> None:
    source_rows.append(row)
    write_region_talk_online_source_item(row, run_id=run_id, stage=stage)
    skipped = sum(1 for r in source_rows if str(r.get("fetch_status") or r.get("method_status") or "").startswith(("skipped", "error", "debug_self_loop_rejected")))
    with_ko = sum(1 for r in source_rows if int(r.get("ko_posts_found") or 0) > 0)
    ko_posts = sum(int(r.get("ko_posts_found") or 0) for r in source_rows)
    write_region_talk_online_stats({
        "run_id": run_id,
        "phase": stage,
        "status": "running",
        "sources_seen_this_run": len(source_rows),
        "sources_done": len(source_rows),
        "sources_total": sources_total or "",
        "sources_rejected_or_skipped_this_run": skipped,
        "sources_with_ko_posts_this_run": with_ko,
        "ko_candidate_posts_this_run": ko_posts,
        "progress_label": f"online discovery sources {len(source_rows)}" + (f"/{sources_total}" if sources_total else ""),
    })


def append_post_online(posts: list[dict[str, Any]], row: dict[str, Any], *, run_id: str, stage: str) -> None:
    posts.append(row)
    write_region_talk_online_post_item(row, run_id=run_id, stage=stage)


def close_region_talk_business_heartbeat() -> None:
    driver = _YDB_BUSINESS_HEARTBEAT_CACHE.get("driver")
    if driver is not None:
        try:
            driver.stop(timeout=5)
        except Exception:
            pass
    _YDB_BUSINESS_HEARTBEAT_CACHE.update({"ydb": None, "driver": None, "pool": None, "table_path": "", "last_alive_at": 0.0})


atexit.register(close_region_talk_business_heartbeat)


def ydb_upsert_json(session: Any, ydb: Any, table_path: str, pk: str, kind: str, payload: dict[str, Any], updated_at: str) -> None:
    query = session.prepare(f"""
DECLARE $pk AS Utf8;
DECLARE $kind AS Utf8;
DECLARE $payload_json AS Json;
DECLARE $updated_at AS Utf8;
UPSERT INTO `{table_path}` (pk, kind, payload_json, updated_at)
VALUES ($pk, $kind, $payload_json, $updated_at);
""")
    session.transaction(ydb.SerializableReadWrite()).execute(
        query,
        {"$pk": pk, "$kind": kind, "$payload_json": json.dumps(payload, ensure_ascii=False), "$updated_at": updated_at},
        commit_tx=True,
    )


def ydb_upsert_json_many(session: Any, ydb: Any, table_path: str, rows: list[tuple[str, str, dict[str, Any]]], updated_at: str, *, chunk_size: int = 100) -> int:
    if not rows:
        return 0
    query = session.prepare(f"""
DECLARE $pk AS Utf8;
DECLARE $kind AS Utf8;
DECLARE $payload_json AS Json;
DECLARE $updated_at AS Utf8;
UPSERT INTO `{table_path}` (pk, kind, payload_json, updated_at)
VALUES ($pk, $kind, $payload_json, $updated_at);
""")
    written = 0
    chunk_size = max(1, int(chunk_size or 100))
    for start in range(0, len(rows), chunk_size):
        tx = session.transaction(ydb.SerializableReadWrite())
        for pk, kind, payload in rows[start:start + chunk_size]:
            tx.execute(query, {"$pk": pk, "$kind": kind, "$payload_json": json.dumps(payload, ensure_ascii=False), "$updated_at": updated_at}, commit_tx=False)
            written += 1
        tx.commit()
    return written


def ydb_select_kind_items(session: Any, ydb: Any, table_path: str, kind: str, *, limit: int = 10000) -> dict[str, dict[str, Any]]:
    max_items = max(1, int(limit))
    page_size = max(1, min(max_items, getenv_int("REGION_TALK_YDB_SELECT_PAGE_SIZE", 200)))
    out: dict[str, dict[str, Any]] = {}
    after = ""
    while len(out) < max_items:
        query = session.prepare(f"""
DECLARE $kind AS Utf8;
DECLARE $after AS Utf8;
SELECT pk, payload_json FROM `{table_path}`
WHERE kind = $kind AND pk > $after
ORDER BY pk
LIMIT {min(page_size, max_items - len(out))};
""")
        result_sets = session.transaction(ydb.StaleReadOnly()).execute(query, {"$kind": kind, "$after": after}, commit_tx=True)
        rows = result_sets[0].rows if result_sets else []
        if not rows:
            break
        for row in rows:
            pk = str(row.pk)
            payload = row.payload_json
            data = json.loads(payload) if isinstance(payload, str) else dict(payload or {})
            if isinstance(data, dict):
                out[pk] = data
            after = pk
        if len(rows) < page_size:
            break
    return out


def ydb_select_latest_state(session: Any, ydb: Any, table_path: str) -> dict[str, Any]:
    result_sets = session.transaction(ydb.StaleReadOnly()).execute(
        f"SELECT payload_json FROM `{table_path}` WHERE pk = 'latest_state';",
        commit_tx=True,
    )
    rows = result_sets[0].rows if result_sets else []
    if not rows:
        return {}
    payload = rows[0].payload_json
    return json.loads(payload) if isinstance(payload, str) else dict(payload or {})


def ydb_prune_legacy_queue_payloads(session: Any, ydb: Any, table_path: str) -> int:
    """Rewrite compact state rows to the current queue contract without deleting data."""
    if not getenv_bool("REGION_TALK_YDB_PRUNE_LEGACY_QUEUE_PAYLOADS", True):
        return 0
    max_rows = getenv_int("REGION_TALK_YDB_PRUNE_MAX_ROWS", 200)
    result_sets = session.transaction(ydb.StaleReadOnly()).execute(
        f"SELECT pk, kind, payload_json, updated_at FROM `{table_path}` "
        f"WHERE kind IN ('state_snapshot', 'run_state_snapshot') LIMIT {max(1, max_rows)};",
        commit_tx=True,
    )
    rows = result_sets[0].rows if result_sets else []
    changed = 0
    for row in rows:
        pk = str(row.pk)
        kind = str(row.kind)
        payload = row.payload_json
        data = json.loads(payload) if isinstance(payload, str) else dict(payload or {})
        if not isinstance(data, dict):
            continue
        pruned = compact_region_talk_state_for_ydb(data)
        if data.get("source_frontier_queue_next") or data.get("similar_seed_queue") or data.get("canonical_source_queue") or data.get("state_schema_version") != pruned.get("state_schema_version"):
            ydb_upsert_json(session, ydb, table_path, pk, kind, pruned, str(row.updated_at or pruned.get("updated_at") or utc_now_iso()))
            changed += 1
    return changed


def load_region_talk_ydb_state() -> tuple[dict[str, Any], dict[str, Any]]:
    cfg = ydb_config_status()
    meta = {"state_backend_requested": "ydb", "state_backend": "ydb", "state_fallback_used": "false", "ydb_namespace": cfg["namespace"], "ydb_tables_expected": ydb_table_plan(), "ydb_read_status": "not_started", "ydb_write_status": "not_started"}
    if cfg["missing"]:
        meta.update({"ydb_read_status": "missing_config", "ydb_error": "missing " + cfg["missing"]})
        return {}, meta
    try:
        ydb, driver, cfg = ydb_connect()
    except Exception as exc:
        meta.update({"ydb_read_status": "connect_error", "ydb_error": f"{type(exc).__name__}: {str(exc)[:220]}"})
        return {}, meta
    snapshot = (os.getenv("REGION_TALK_YDB_STATE_SNAPSHOT_FILE") or "").strip()
    if not snapshot:
        table_path = ydb_kv_table_path(cfg)
        try:
            pool = ydb.SessionPool(driver)
            def op(session: Any) -> dict[str, Any]:
                ensure_ydb_kv_table(ydb, session, table_path)
                data0 = ydb_select_latest_state(session, ydb, table_path)
                # Merge row-level queue items written by parallel notebooks (e.g. ImageDiagnostic).
                image_items = ydb_select_kind_items(session, ydb, table_path, "image_queue_item", limit=getenv_int("REGION_TALK_YDB_MAX_CANDIDATE_ROWS", 5000))
                source_items = ydb_select_kind_items(session, ydb, table_path, "source_queue_item", limit=getenv_int("REGION_TALK_YDB_MAX_SOURCE_ROWS", 5000))
                publication_items = ydb_select_kind_items(session, ydb, table_path, "publication_candidate_item", limit=getenv_int("REGION_TALK_YDB_MAX_CANDIDATE_ROWS", 5000))
                post_items = ydb_select_kind_items(session, ydb, table_path, "processed_post_item", limit=getenv_int("REGION_TALK_YDB_MAX_POST_ROWS", 20000))
                candidate_items = ydb_select_kind_items(session, ydb, table_path, "candidate_memory_item", limit=getenv_int("REGION_TALK_YDB_MAX_CANDIDATE_ROWS", 5000))
                queue_cursors = ydb_select_kind_items(session, ydb, table_path, "queue_cursor", limit=20)
                if isinstance(data0, dict):
                    if post_items:
                        p = data0.get("processed_posts") if isinstance(data0.get("processed_posts"), dict) else {}
                        p = dict(p)
                        for _pk, item in post_items.items():
                            key = str(item.get("post_id") or item.get("platform_post_key") or item.get("post_url") or _pk.replace("processed_post_item:", ""))
                            if key:
                                p[key] = {**p.get(key, {}), **item}
                        data0["processed_posts"] = p
                        data0["ydb_row_level_processed_post_items_loaded"] = len(post_items)
                    if candidate_items:
                        c = data0.get("candidate_memory") if isinstance(data0.get("candidate_memory"), dict) else {}
                        c = dict(c)
                        for _pk, item in candidate_items.items():
                            key = str(item.get("candidate_memory_id") or item.get("post_id") or item.get("post_url") or _pk.replace("candidate_memory_item:", ""))
                            if key:
                                c[key] = {**c.get(key, {}), **item}
                        data0["candidate_memory"] = c
                        data0["ydb_row_level_candidate_memory_items_loaded"] = len(candidate_items)
                    if image_items:
                        q = data0.get("image_candidate_queue") if isinstance(data0.get("image_candidate_queue"), dict) else {}
                        q = dict(q)
                        for _pk, item in image_items.items():
                            key = str(item.get("image_queue_id") or item.get("post_url") or _pk.replace("image_queue_item:", ""))
                            if key:
                                q[key] = {**q.get(key, {}), **item}
                        data0["image_candidate_queue"] = q
                        data0["ydb_row_level_image_queue_items_loaded"] = len(image_items)
                    if source_items:
                        q = data0.get("unified_source_queue") if isinstance(data0.get("unified_source_queue"), dict) else {}
                        q = dict(q)
                        for _pk, item in source_items.items():
                            key = str(item.get("canonical_source_key") or _pk.replace("source_queue_item:", ""))
                            if key:
                                q[key] = {**q.get(key, {}), **item}
                        data0["unified_source_queue"] = q
                        data0["ydb_row_level_source_queue_items_loaded"] = len(source_items)
                    if publication_items:
                        q = data0.get("publication_candidate_queue") if isinstance(data0.get("publication_candidate_queue"), dict) else {}
                        q = dict(q)
                        for _pk, item in publication_items.items():
                            key = str(item.get("publication_candidate_id") or item.get("post_url") or _pk.replace("publication_candidate_item:", ""))
                            if key:
                                q[key] = {**q.get(key, {}), **item}
                        data0["publication_candidate_queue"] = q
                        data0["ydb_row_level_publication_candidate_items_loaded"] = len(publication_items)
                    if queue_cursors:
                        cursors = data0.get("queue_cursors") if isinstance(data0.get("queue_cursors"), dict) else {}
                        cursors = dict(cursors)
                        for _pk, item in queue_cursors.items():
                            name = str(item.get("queue_name") or item.get("name") or _pk.replace("queue_cursor:", ""))
                            short = "source" if "source" in name else ("image" if "image" in name else name)
                            cursors[short] = {**(cursors.get(short) or {}), **item}
                            if short == "source":
                                data0["unified_source_queue_cursor_position"] = item.get("cursor_position", data0.get("unified_source_queue_cursor_position", 0))
                                data0["unified_source_queue_cursor_key"] = item.get("cursor_key", data0.get("unified_source_queue_cursor_key", ""))
                            elif short == "image":
                                data0["image_candidate_queue_cursor_position"] = item.get("cursor_position", data0.get("image_candidate_queue_cursor_position", 0))
                                data0["image_candidate_queue_cursor_key"] = item.get("cursor_key", data0.get("image_candidate_queue_cursor_key", ""))
                        data0["queue_cursors"] = cursors
                        data0["ydb_queue_cursor_items_loaded"] = len(queue_cursors)
                return data0
            data = pool.retry_operation_sync(op)
            driver.stop(timeout=5)
            if not isinstance(data, dict) or not data:
                meta.update({"ydb_read_status": "empty", "previous_state_loaded": "false", "increment_state_loaded": "false", "previous_state_source": "ydb:" + table_path})
                return {}, meta
            if "posts" not in data and isinstance(data.get("processed_posts"), dict):
                data["posts"] = data.get("processed_posts") or {}
            if "region_talk_sources" not in data and isinstance(data.get("sources"), dict):
                data["region_talk_sources"] = data.get("sources") or {}
            raw = json.dumps(data, ensure_ascii=False, sort_keys=True).encode("utf-8")
            state_hash = hashlib.sha256(raw).hexdigest()
            data["_loaded_from_path"] = "ydb:" + table_path + "#latest_state"
            meta.update({"ydb_read_status": "ok", "previous_state_loaded": "true", "increment_state_loaded": "true", "previous_state_source": "ydb:" + table_path, "increment_state_path": "ydb:" + table_path, "previous_state_run_id": data.get("run_id", ""), "previous_run_id": data.get("run_id", ""), "previous_state_hash": state_hash, "state_schema_version_previous": data.get("state_schema_version", ""), "previous_seen_post_count": len(data.get("processed_posts") or data.get("posts") or {}), "previous_candidate_memory_count": len(data.get("candidate_memory") or {}), "ydb_table_path": table_path, "ydb_state_mode": "compact_kv"})
            return data, meta
        except Exception as exc:
            try:
                driver.stop(timeout=5)
            except Exception:
                pass
            meta.update({"ydb_read_status": "error", "ydb_error": f"{type(exc).__name__}: {str(exc)[:220]}"})
            return {}, meta
    path = Path(snapshot)
    try:
        raw = path.read_bytes()
        data = json.loads(raw.decode("utf-8"))
        if not isinstance(data, dict):
            raise ValueError("snapshot is not JSON object")
        state_hash = hashlib.sha256(raw).hexdigest()
        data["_loaded_from_path"] = str(path)
        meta.update({"ydb_read_status": "ok", "previous_state_loaded": "true", "increment_state_loaded": "true", "previous_state_source": "ydb:" + str(path), "increment_state_path": "ydb:" + str(path), "previous_state_run_id": data.get("run_id", ""), "previous_run_id": data.get("run_id", ""), "previous_state_hash": state_hash, "state_schema_version_previous": data.get("state_schema_version", ""), "previous_seen_post_count": len(data.get("posts") or {}), "previous_candidate_memory_count": len(data.get("candidate_memory") or {})})
        return data, meta
    except Exception as exc:
        meta.update({"ydb_read_status": "error", "ydb_error": f"{type(exc).__name__}: {str(exc)[:220]}"})
        return {}, meta


def save_region_talk_ydb_state(output_dir: Path, state: dict[str, Any]) -> dict[str, Any]:
    cfg = ydb_config_status()
    meta = {"state_backend_requested": "ydb", "state_backend": "ydb", "state_fallback_used": "false", "ydb_namespace": cfg["namespace"], "ydb_tables_expected": ydb_table_plan(), "ydb_write_status": "not_started"}
    if cfg["missing"]:
        return {**meta, "ydb_write_status": "missing_config", "state_write_status": "error", "state_write_error": "missing " + cfg["missing"]}
    snapshot = (os.getenv("REGION_TALK_YDB_STATE_SNAPSHOT_FILE") or "").strip()
    if not snapshot:
        try:
            ydb, driver, cfg = ydb_connect()
            table_path = ydb_kv_table_path(cfg)
            compact = compact_region_talk_state_for_ydb(state)
            updated_at = str(compact.get("updated_at") or utc_now_iso())
            payload = json.dumps(compact, ensure_ascii=False, sort_keys=True)
            state_hash = hashlib.sha256(payload.encode("utf-8")).hexdigest()
            pool = ydb.SessionPool(driver)
            def op(session: Any) -> None:
                ensure_ydb_kv_table(ydb, session, table_path)
                ydb_upsert_json(session, ydb, table_path, "latest_state", "state_snapshot", compact, updated_at)
                ydb_upsert_json(session, ydb, table_path, "run:" + str(compact.get("run_id") or updated_at), "run_state_snapshot", compact, updated_at)
                ydb_upsert_json(session, ydb, table_path, "metrics:" + str(compact.get("run_id") or updated_at), "run_metrics", {"run_id": compact.get("run_id"), "updated_at": updated_at, "all_time_metrics": compact.get("all_time_metrics") or {}}, updated_at)
                row_items: list[tuple[str, str, dict[str, Any]]] = []
                row_write_mode = (os.getenv("REGION_TALK_YDB_ROW_WRITE_MODE") or "changed").strip().lower()
                skip_row_rewrite = getenv_bool("REGION_TALK_YDB_SKIP_ROW_LEVEL_REWRITE", False)
                current_run_id = str(compact.get("run_id") or "")
                def should_write_queue_item(item: dict[str, Any], *, item_type: str) -> bool:
                    if skip_row_rewrite:
                        return False
                    if row_write_mode == "full":
                        return True
                    if str(item.get("status_changed_this_run") or "").lower() == "true":
                        return True
                    if current_run_id and current_run_id in {str(item.get("first_seen_run_id") or ""), str(item.get("last_scan_run_id") or ""), str(item.get("last_attempt_run_id") or ""), str(item.get("last_image_diag_run_id") or ""), str(item.get("source_visual_rollup_run_id") or "")} :
                        return True
                    if item_type == "image" and str(item.get("image_queue_status") or "") in {"image_analysis_in_progress", "actual_scored"}:
                        return True
                    return False
                for item in (compact.get("image_candidate_queue") or {}).values():
                    if isinstance(item, dict) and should_write_queue_item(item, item_type="image"):
                        key = str(item.get("image_queue_id") or item.get("post_url") or "")
                        if key:
                            row_items.append(("image_queue_item:" + key, "image_queue_item", item))
                for item in (compact.get("processed_posts") or {}).values():
                    if isinstance(item, dict):
                        key = str(item.get("post_id") or item.get("platform_post_key") or item.get("post_url") or "")
                        if key:
                            row_items.append(("processed_post_item:" + key, "processed_post_item", item))
                for item in (compact.get("candidate_memory") or {}).values():
                    if isinstance(item, dict):
                        key = str(item.get("candidate_memory_id") or item.get("post_id") or item.get("post_url") or "")
                        if key:
                            row_items.append(("candidate_memory_item:" + key, "candidate_memory_item", item))
                for item in (compact.get("publication_candidate_queue") or {}).values():
                    if isinstance(item, dict):
                        key = str(item.get("publication_candidate_id") or item.get("post_url") or "")
                        if key:
                            row_items.append(("publication_candidate_item:" + key, "publication_candidate_item", item))
                for item in (compact.get("unified_source_queue") or {}).values():
                    if isinstance(item, dict) and should_write_queue_item(item, item_type="source"):
                        key = str(item.get("canonical_source_key") or item.get("source_queue_id") or "")
                        if key:
                            row_items.append(("source_queue_item:" + key, "source_queue_item", item))
                if not skip_row_rewrite:
                    for name, item in (compact.get("queue_cursors") or {}).items():
                        if isinstance(item, dict):
                            row_items.append(("queue_cursor:" + str(name), "queue_cursor", item))
                    row_items.append(("queue_metrics:latest", "queue_metrics", {"run_id": compact.get("run_id"), "updated_at": updated_at, "queue_cursors": compact.get("queue_cursors") or {}, "source_queue_total": len(compact.get("unified_source_queue") or {}), "image_queue_total": len(compact.get("image_candidate_queue") or {})}))
                compact["_ydb_row_level_publication_candidate_items_written"] = len(compact.get("publication_candidate_queue") or {})
                compact["_ydb_row_level_items_written"] = ydb_upsert_json_many(session, ydb, table_path, row_items, updated_at, chunk_size=getenv_int("REGION_TALK_YDB_ROW_UPSERT_CHUNK_SIZE", 100))
                compact["_ydb_pruned_legacy_queue_payload_rows"] = ydb_prune_legacy_queue_payloads(session, ydb, table_path)
            pool.retry_operation_sync(op)
            driver.stop(timeout=5)
            return {**meta, "ydb_write_status": "ok", "state_write_status": "ok", "state_write_path": "ydb:" + table_path, "latest_state_run_id": state.get("run_id", ""), "latest_state_uri": "ydb:" + table_path + "#latest_state", "latest_state_hash": state_hash, "latest_state_pointer_path": "ydb:" + table_path + "#latest_state", "ydb_table_path": table_path, "ydb_state_mode": "compact_kv", "ydb_compact_schema_version": compact.get("state_schema_version", ""), "ydb_compact_queue_contract_version": compact.get("queue_contract_version", ""), "ydb_pruned_legacy_queue_payload_rows": compact.get("_ydb_pruned_legacy_queue_payload_rows", 0), "ydb_compact_sources": len(compact.get("sources") or {}), "ydb_compact_processed_posts": len(compact.get("processed_posts") or {}), "ydb_compact_candidate_memory": len(compact.get("candidate_memory") or {}), "ydb_compact_publication_candidate_queue": len(compact.get("publication_candidate_queue") or {}), "ydb_compact_unified_source_queue": len(compact.get("unified_source_queue") or {}), "ydb_compact_image_candidate_queue": len(compact.get("image_candidate_queue") or {}), "ydb_row_level_processed_post_items_written": len(compact.get("processed_posts") or {}), "ydb_row_level_candidate_memory_items_written": len(compact.get("candidate_memory") or {}), "ydb_row_level_source_queue_items_written": len(compact.get("unified_source_queue") or {}), "ydb_row_level_image_queue_items_written": len(compact.get("image_candidate_queue") or {}), "ydb_row_level_publication_candidate_items_written": compact.get("_ydb_row_level_publication_candidate_items_written", 0), "ydb_queue_cursor_items_written": len(compact.get("queue_cursors") or {}), "ydb_row_level_items_written": compact.get("_ydb_row_level_items_written", 0)}
        except Exception as exc:
            return {**meta, "ydb_write_status": "error", "state_write_status": "error", "state_write_error": f"{type(exc).__name__}: {str(exc)[:220]}"}
    try:
        target = Path(snapshot)
        target.parent.mkdir(parents=True, exist_ok=True)
        payload = json.dumps(state, ensure_ascii=False, indent=2)
        target.write_text(payload, encoding="utf-8")
        state_hash = hashlib.sha256(payload.encode("utf-8")).hexdigest()
        return {**meta, "ydb_write_status": "ok", "state_write_status": "ok", "state_write_path": "ydb:" + str(target), "latest_state_run_id": state.get("run_id", ""), "latest_state_uri": "ydb:" + str(target), "latest_state_hash": state_hash, "latest_state_pointer_path": "ydb:" + str(target)}
    except Exception as exc:
        return {**meta, "ydb_write_status": "error", "state_write_status": "error", "state_write_error": f"{type(exc).__name__}: {str(exc)[:220]}"}

def load_region_talk_state(output_dir: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    requested_backend = region_talk_state_backend_requested()
    if requested_backend == "ydb":
        ydb_state, ydb_meta = load_region_talk_ydb_state()
        if ydb_state or ydb_meta.get("ydb_read_status") == "empty":
            return ydb_state, ydb_meta
        if getenv_bool("REGION_TALK_REQUIRE_YDB_STATE", False):
            raise RuntimeError("REGION_TALK_REQUIRE_YDB_STATE=1 but YDB state is unavailable: " + str(ydb_meta.get("ydb_error") or ydb_meta.get("ydb_read_status")))
        fallback_state, fallback_meta = load_region_talk_json_state(output_dir)
        fallback_meta.update({**ydb_meta, "state_backend": "json_fallback", "state_fallback_used": "true", "state_fallback_reason": ydb_meta.get("ydb_error") or ydb_meta.get("ydb_read_status") or "ydb_unavailable", "previous_state_loaded": fallback_meta.get("previous_state_loaded", "false"), "previous_state_hash": fallback_meta.get("previous_state_hash", "")})
        return fallback_state, fallback_meta
    return load_region_talk_json_state(output_dir)


def load_region_talk_json_state(output_dir: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    for path in state_file_candidates(output_dir):
        try:
            if path.exists():
                raw = path.read_bytes()
                data = json.loads(raw.decode("utf-8"))
                if isinstance(data, dict):
                    data["_loaded_from_path"] = str(path)
                    state_hash = hashlib.sha256(raw).hexdigest()
                    return data, {
                        "increment_state_loaded": "true",
                        "increment_state_path": str(path),
                        "previous_state_loaded": "true",
                        "previous_state_source": str(path),
                        "previous_state_run_id": data.get("run_id", ""),
                        "previous_state_hash": state_hash,
                        "previous_run_id": data.get("run_id", ""),
                        "state_schema_version_previous": data.get("state_schema_version", ""),
                        "previous_seen_post_count": len(data.get("posts") or {}),
                        "previous_candidate_memory_count": len(data.get("candidate_memory") or {}),
                    }
        except Exception as exc:
            return {}, {"increment_state_loaded": "false", "increment_state_path": str(path), "previous_state_loaded": "false", "previous_state_source": str(path), "previous_state_hash": "", "increment_state_error": f"{type(exc).__name__}: {str(exc)[:180]}", "previous_run_id": "", "previous_seen_post_count": 0}
    if getenv_bool("REGION_TALK_REQUIRE_PREVIOUS_STATE", False):
        raise RuntimeError("REGION_TALK_REQUIRE_PREVIOUS_STATE=1 but no previous Region Talk state was found")
    return {}, {"increment_state_loaded": "false", "increment_state_path": "", "previous_state_loaded": "false", "previous_state_source": "", "previous_state_hash": "", "increment_state_note": "baseline run, not real increment: no previous dry-run state file found", "previous_run_id": "", "previous_seen_post_count": 0, "previous_candidate_memory_count": 0}


def load_public_blogger_links(config: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    candidates = []
    for raw in [
        os.getenv("REGION_TALK_PUBLIC_BLOGGER_LINKS_FILE"),
        (config or {}).get("public_blogger_links_file") if config else None,
        "public_travel_blogger_channel_links.xlsx",
        "artifacts/public_travel_blogger_channel_links.xlsx",
        "/home/dev/projects/events-bot-new/artifacts/public_travel_blogger_channel_links.xlsx",
    ]:
        if raw:
            candidates.append(Path(str(raw)))
    for root in [Path.cwd(), Path(__file__).resolve().parent, Path("/kaggle/input")]:
        if root.exists():
            candidates.extend(root.rglob("public_travel_blogger_channel_links.xlsx"))
    path = next((p for p in candidates if p.exists()), None)
    if not path:
        _REGION_TALK_TELEGRAM_RUNTIME.update({
            "catalog_import_file_status": "missing",
            "catalog_import_file_path": "",
            "catalog_import_rows_total": 0,
            "catalog_import_telegram_unique": 0,
            "catalog_import_vk_unique": 0,
            "catalog_import_external_quarantine": 0,
        })
        return []
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        if getenv_bool("REGION_TALK_AUTO_INSTALL", True):
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
            from openpyxl import load_workbook  # type: ignore
        else:
            return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Links" not in wb.sheetnames:
        return []
    ws = wb["Links"]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(x or "").strip() for x in next(rows_iter, [])]
    out: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    rows_total = 0
    external_count = 0
    for row in rows_iter:
        rows_total += 1
        rec = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
        url = str(rec.get("URL") or "").strip()
        platform = normalize_source_platform(str(rec.get("Platform") or ""), url)
        handle = canonical_handle(str(rec.get("Handle") or "").strip())
        if not platform or not url:
            continue
        canonical_key = canonical_source_key(platform, handle, url)
        key = (platform, canonical_key)
        if key in seen:
            continue
        seen.add(key)
        canonical = canonical_source_url(platform, handle, url)
        cand_id = "src_cand_" + stable_hash(platform, canonical_key)
        is_target = platform in {"telegram", "vk"}
        if not is_target:
            external_count += 1
        out.append({
            "source_candidate_id": cand_id,
            "frontier_source_id": cand_id,
            "canonical_source_key": canonical_key,
            "platform": platform,
            "platform_guess": platform,
            "canonical_url": canonical,
            "normalized_url": canonical,
            "username_or_handle": handle,
            "title_guess": str(rec.get("Handle") or handle or canonical).strip(),
            "source_title": str(rec.get("Handle") or handle or canonical).strip(),
            "category": str(rec.get("Category") or ""),
            "source_catalog": str(rec.get("Source") or ""),
            "source_page": str(rec.get("Source page") or ""),
            "discovery_type": "public_travel_blogger_catalog",
            "edge_type": "public_travel_blogger_catalog",
            "candidate_source_status": "source_frontier",
            "frontier_status": "queued_unresolved" if is_target else "external_quarantine",
            "scan_status": "pending_primary_scan" if is_target else "external_quarantine",
            "confidence": 0.45 if platform == "telegram" else 0.35,
            "frontier_priority": 0.62 if platform == "telegram" else 0.40,
            "frontier_reason": "public travel/blogger catalog import; target Telegram/VK active, external platforms quarantined",
            "discovered_from_source": str(rec.get("Source") or "public_travel_blogger_channel_links.xlsx"),
            "discovered_from_post_url": "",
            "catalog_import_row": rows_total,
            "catalog_import_source": f"catalog_import:{path.name}:{rows_total}",
            "external_quarantine_reason": "" if is_target else "non_target_platform_for_region_talk_active_scan",
            "raw_url": url,
        })
    _REGION_TALK_TELEGRAM_RUNTIME.update({
        "catalog_import_file_status": "found",
        "catalog_import_file_path": str(path),
        "catalog_import_rows_total": rows_total,
        "catalog_import_telegram_unique": len({r.get("canonical_source_key") for r in out if r.get("platform") == "telegram"}),
        "catalog_import_vk_unique": len({r.get("canonical_source_key") for r in out if r.get("platform") == "vk"}),
        "catalog_import_external_quarantine": external_count,
    })
    return out


def save_region_talk_state(output_dir: Path, state: dict[str, Any]) -> dict[str, Any]:
    requested_backend = region_talk_state_backend_requested()
    if requested_backend == "ydb":
        ydb_meta = save_region_talk_ydb_state(output_dir, state)
        json_meta = save_region_talk_json_state(output_dir, state)
        if ydb_meta.get("state_write_status") == "ok":
            return {**json_meta, **ydb_meta, "json_backup_write_status": json_meta.get("state_write_status", "")}
        if getenv_bool("REGION_TALK_REQUIRE_YDB_STATE", False):
            raise RuntimeError("REGION_TALK_REQUIRE_YDB_STATE=1 but YDB state write failed: " + str(ydb_meta.get("state_write_error") or ydb_meta.get("ydb_write_status")))
        return {**json_meta, **ydb_meta, "state_backend": "json_fallback", "state_fallback_used": "true", "state_fallback_reason": ydb_meta.get("state_write_error") or ydb_meta.get("ydb_write_status") or "ydb_write_failed"}
    return save_region_talk_json_state(output_dir, state)


def save_region_talk_json_state(output_dir: Path, state: dict[str, Any]) -> dict[str, Any]:
    target = state_file_candidates(output_dir)[0]
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        payload = json.dumps(state, ensure_ascii=False, indent=2)
        state_hash = hashlib.sha256(payload.encode("utf-8")).hexdigest()
        target.write_text(payload, encoding="utf-8")
        run_copy = output_dir / STATE_FILE_NAME
        if run_copy.resolve() != target.resolve():
            run_copy.write_text(payload, encoding="utf-8")
        latest_pointer = target.parent / "latest-region-talk-state.json"
        latest_pointer.write_text(json.dumps({"latest_state_run_id": state.get("run_id"), "latest_state_uri": str(target), "latest_state_hash": state_hash, "updated_at": state.get("updated_at")}, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"state_write_status": "ok", "state_write_path": str(target), "latest_state_run_id": state.get("run_id", ""), "latest_state_uri": str(target), "latest_state_hash": state_hash, "latest_state_pointer_path": str(latest_pointer)}
    except Exception as exc:
        return {"state_write_status": "error", "state_write_error": f"{type(exc).__name__}: {str(exc)[:240]}", "state_write_path": str(target)}


def classify_pre_llm_reject_reason(fresh: dict[str, Any], scope: dict[str, Any], ad_gate: dict[str, Any], substance: dict[str, Any], ts: dict[str, Any]) -> str:
    if not fresh.get("fresh_enough"):
        return "reject_old_post"
    if not scope.get("kaliningrad_oblast_only_scope"):
        return "reject_not_kaliningrad_oblast_only"
    if ad_gate.get("is_ad_or_promo"):
        return "reject_ad_or_promo"
    if float(ts.get("newsiness_score") or 0) >= 0.70 or float(ts.get("trash_score") or 0) >= 0.70:
        return "reject_news_or_trash"
    if float(substance.get("text_substance_score") or 0) < 0.18:
        return "reject_low_substance"
    return "reject_source_boilerplate"


def git_provenance() -> dict[str, str]:
    out = {"git_sha": os.getenv("GIT_SHA", ""), "git_sha_short": "", "branch": os.getenv("GIT_BRANCH", "")}
    try:
        sha = subprocess.check_output(["git", "rev-parse", "HEAD"], text=True, stderr=subprocess.DEVNULL).strip()
        out["git_sha"] = out["git_sha"] or sha
        out["git_sha_short"] = sha[:8]
    except Exception:
        out["git_sha_short"] = (out.get("git_sha") or "")[:8]
    try:
        branch = subprocess.check_output(["git", "rev-parse", "--abbrev-ref", "HEAD"], text=True, stderr=subprocess.DEVNULL).strip()
        out["branch"] = out["branch"] or branch
    except Exception:
        pass
    return out


class TelegramRequestGovernor:
    def __init__(self, run_id: str, output_dir: Path, state: dict[str, Any]) -> None:
        self.run_id = run_id
        self.output_dir = output_dir
        self.state = state if isinstance(state, dict) else {}
        self.entity_cache: dict[str, Any] = dict(self.state.get("telegram_entity_cache") or {})
        self.entity_cache_loaded_from_path = str((self.state.get("_loaded_from_path") or self.state.get("state_path") or ""))
        self.cooldowns: dict[str, Any] = dict(self.state.get("telegram_cooldowns") or {})
        self.ledger: list[dict[str, Any]] = []
        self.requests_by_method: dict[str, int] = {}
        self.total_attempted = 0
        self.total_ok = 0
        self.total_error = 0
        self.resolve_cache_hits = 0
        self.resolve_network_attempts = 0
        self.resolve_network_ok = 0
        self.resolve_network_floodwait = 0
        self.resolve_skipped_by_cooldown = 0
        self.recommendation_calls_attempted = 0
        self.recommendation_calls_ok = 0
        self.recommendation_channels_returned = 0
        self.recommendation_channels_added_to_frontier = 0
        self.history_sources_attempted = 0
        self.history_sources_ok = 0
        self.history_posts_fetched = 0
        self.media_downloads_attempted = 0
        self.media_downloads_ok = 0
        self.max_floodwait_seconds = 0
        self.floodwait_method = ""
        self.floodwait_cooldown_until = ""
        self.telegram_phase_status = "ok"
        self.degraded = False
        self.max_total_requests = getenv_int("REGION_TALK_TG_MAX_TOTAL_REQUESTS_PER_RUN", 800)
        self.max_network_resolves = getenv_int("REGION_TALK_TG_MAX_NETWORK_RESOLVES_PER_RUN", 8)
        self.max_history_sources = getenv_int("REGION_TALK_HISTORY_SOURCES_TARGET", getenv_int("REGION_TALK_TG_MAX_HISTORY_SOURCES_PER_RUN", 100))
        self.max_media_downloads = getenv_int("REGION_TALK_TG_MAX_MEDIA_DOWNLOADS_PER_RUN", 120)
        self.max_recommendation_calls = getenv_int("REGION_TALK_MAX_SIMILAR_SEEDS_PER_RUN", getenv_int("REGION_TALK_TG_MAX_RECOMMENDATION_CALLS_PER_RUN", 100))
        self.floodwait_abort_threshold = getenv_int("REGION_TALK_TG_FLOODWAIT_ABORT_THRESHOLD_SECONDS", 300)
        self.floodwait_margin = getenv_int("REGION_TALK_TG_FLOODWAIT_COOLDOWN_MARGIN_SECONDS", 1800)

    def cache_key(self, handle_or_url: str) -> str:
        raw = canonical_handle(str(handle_or_url or "").strip()).lstrip("@").lower()
        raw = re.sub(r"^https?://t\.me/", "", raw, flags=re.I).split("/", 1)[0].lower()
        return "telegram:username:" + raw

    def cooldown_active(self, key: str) -> tuple[bool, str]:
        rec = self.cooldowns.get(key) or {}
        until = str(rec.get("cooldown_until") or "")
        dt = parse_iso_datetime(until)
        if dt and dt > datetime.now(timezone.utc):
            return True, until
        return False, until

    def log(self, method_name: str, method_class: str, source_id: str, canonical_url: str, decision: str, **extra: Any) -> None:
        rec = {
            "ts": utc_now_iso(), "run_id": self.run_id, "method_name": method_name, "method_class": method_class,
            "source_id": source_id, "canonical_url": canonical_url, "decision": decision, **extra,
        }
        self.ledger.append(rec)

    def mark_floodwait(self, method_name: str, source_id: str, canonical_url: str, seconds: int) -> str:
        cooldown_until = iso_after_seconds(seconds + self.floodwait_margin)
        self.max_floodwait_seconds = max(self.max_floodwait_seconds, seconds)
        self.floodwait_method = method_name
        self.floodwait_cooldown_until = cooldown_until
        self.telegram_phase_status = "degraded_floodwait"
        self.degraded = True
        self.cooldowns[f"method:{method_name}"] = {"cooldown_until": cooldown_until, "last_error_seconds": seconds, "last_error_run_id": self.run_id}
        if canonical_url:
            self.cooldowns[f"source:{canonical_url}"] = {"cooldown_until": cooldown_until, "reason": "telegram_floodwait", "last_error_seconds": seconds}
        self.log(method_name, "entity_resolve_expensive", source_id, canonical_url, "error_floodwait", ok=False, floodwait_seconds=seconds, cooldown_until=cooldown_until)
        return cooldown_until

    def has_total_request_budget(self, method_name: str, source_id: str = "", canonical_url: str = "") -> bool:
        if self.total_attempted >= self.max_total_requests:
            self.telegram_phase_status = "degraded_request_budget_exhausted"
            self.degraded = True
            self.log(method_name, "rate_budget", source_id, canonical_url, "skipped_total_request_budget", ok=False, max_total_requests=self.max_total_requests)
            return False
        return True

    async def resolve_entity(self, client: Any, seed: Seed) -> tuple[Any | None, dict[str, Any]]:
        canonical_url = seed.canonical_url
        source_id = seed.source_id
        handle = seed.handle.lstrip("@")
        key = self.cache_key(seed.handle or seed.url)
        source_cd, source_until = self.cooldown_active(f"source:{canonical_url}")
        method_cd, method_until = self.cooldown_active("method:ResolveUsernameRequest")
        if self.degraded or source_cd or method_cd:
            self.resolve_skipped_by_cooldown += 1
            status = "skipped_telegram_global_cooldown" if self.degraded or method_cd else "skipped_telegram_resolve_cooldown"
            self.log("ResolveUsernameRequest", "entity_resolve_expensive", source_id, canonical_url, status, ok=False, cooldown_until=method_until or source_until)
            return None, {"fetch_status": status, "telegram_resolve_status": status, "next_allowed_resolve_at": method_until or source_until}
        cached = self.entity_cache.get(key) or {}
        if cached.get("channel_id_private") and cached.get("access_hash_private"):
            try:
                from telethon.tl.types import InputPeerChannel  # type: ignore
                entity = InputPeerChannel(int(cached["channel_id_private"]), int(cached["access_hash_private"]))
                self.resolve_cache_hits += 1
                self.log("ResolveUsernameRequest", "entity_resolve_expensive", source_id, canonical_url, "skipped_cache_hit", cache_hit=True, network_call=False, ok=True)
                return entity, {"telegram_resolve_status": "resolved_from_private_cache", "resolved_title": cached.get("title_last_seen", "")}
            except Exception:
                pass
        if self.resolve_network_attempts >= self.max_network_resolves or not self.has_total_request_budget("ResolveUsernameRequest", source_id, canonical_url):
            self.log("ResolveUsernameRequest", "entity_resolve_expensive", source_id, canonical_url, "skipped_budget", ok=False)
            return None, {"fetch_status": "skipped_telegram_budget_exhausted", "telegram_resolve_status": "skipped_budget", "resolve_attempt_count": self.resolve_network_attempts}
        self.resolve_network_attempts += 1
        self.total_attempted += 1
        self.requests_by_method["ResolveUsernameRequest"] = self.requests_by_method.get("ResolveUsernameRequest", 0) + 1
        try:
            entity = await client.get_entity(handle)
            self.total_ok += 1
            self.resolve_network_ok += 1
            self.entity_cache[key] = {
                "canonical_url": canonical_url,
                "username": handle,
                "title_last_seen": str(getattr(entity, "title", None) or seed.source_title),
                "entity_kind": "channel",
                "channel_id_private": str(getattr(entity, "id", "") or ""),
                "access_hash_private": str(getattr(entity, "access_hash", "") or ""),
                "resolved_at": utc_now_iso(),
                "last_used_at": utc_now_iso(),
                "last_success_at": utc_now_iso(),
                "last_error_at": None,
                "last_error_code": None,
                "resolve_attempt_count": int((cached or {}).get("resolve_attempt_count") or 0) + 1,
                "resolve_network_count": int((cached or {}).get("resolve_network_count") or 0) + 1,
                "resolve_cache_hit_count": int((cached or {}).get("resolve_cache_hit_count") or 0),
            }
            self.log("ResolveUsernameRequest", "entity_resolve_expensive", source_id, canonical_url, "allowed", cache_hit=False, network_call=True, ok=True)
            return entity, {"telegram_resolve_status": "resolved_network", "resolved_title": str(getattr(entity, "title", None) or seed.source_title)}
        except Exception as exc:
            self.total_error += 1
            seconds = int(getattr(exc, "seconds", 0) or 0)
            if type(exc).__name__ == "FloodWaitError" or seconds:
                self.resolve_network_floodwait += 1
                cooldown_until = self.mark_floodwait("ResolveUsernameRequest", source_id, canonical_url, seconds)
                return None, {"fetch_status": "error_floodwait_resolve", "telegram_resolve_status": "error_floodwait", "fetch_error_code": "FloodWaitError", "fetch_error_message": str(exc)[:180], "next_allowed_resolve_at": cooldown_until}
            self.log("ResolveUsernameRequest", "entity_resolve_expensive", source_id, canonical_url, "error", ok=False, error_code=type(exc).__name__, error_message=str(exc)[:180])
            return None, {"fetch_status": "error_telegram_rpc", "telegram_resolve_status": "error", "fetch_error_code": type(exc).__name__, "fetch_error_message": str(exc)[:180]}

    def write_ledger(self) -> None:
        try:
            path = self.output_dir.parent.parent / "logs" / "telegram-request-ledger.jsonl"
            path.parent.mkdir(parents=True, exist_ok=True)
            with path.open("a", encoding="utf-8") as f:
                for rec in self.ledger:
                    f.write(json.dumps(rec, ensure_ascii=False) + "\n")
        except Exception:
            pass

    def observability_row(self, started_at: str, finished_at: str) -> dict[str, Any]:
        return {
            "run_id": self.run_id, "started_at": started_at, "finished_at": finished_at,
            "telegram_phase_status": self.telegram_phase_status,
            "telethon_version": _REGION_TALK_TELEGRAM_RUNTIME.get("telethon_version", ""),
            "session_cache_loaded": "true",
            "entity_cache_loaded": str(bool(self.state.get("telegram_entity_cache"))).lower(),
            "entity_cache_loaded_from_path": self.entity_cache_loaded_from_path,
            "entity_cache_write_path": str((self.output_dir.parent.parent / "state" / "region-talk-state.json")),
            "entity_cache_entries": len(self.entity_cache),
            "entity_cache_hit_rate": round(self.resolve_cache_hits / max(1, self.resolve_cache_hits + self.resolve_network_attempts), 3),
            "resolved_sources_available_for_history_fetch": sum(1 for v in self.entity_cache.values() if isinstance(v, dict) and v.get("channel_id_private") and v.get("access_hash_private")),
            "resolved_sources_used_without_network_resolve": self.resolve_cache_hits,
            "telegram_total_requests_attempted": self.total_attempted,
            "telegram_total_requests_ok": self.total_ok,
            "telegram_total_requests_error": self.total_error,
            "telegram_requests_by_method_json": json.dumps(self.requests_by_method, ensure_ascii=False),
            "resolve_cache_hits": self.resolve_cache_hits,
            "resolve_network_attempts": self.resolve_network_attempts,
            "resolve_network_ok": self.resolve_network_ok,
            "resolve_network_floodwait": self.resolve_network_floodwait,
            "resolve_skipped_by_cache": self.resolve_cache_hits,
            "resolve_skipped_by_cooldown": self.resolve_skipped_by_cooldown,
            "max_floodwait_seconds": self.max_floodwait_seconds,
            "floodwait_method": self.floodwait_method,
            "floodwait_cooldown_until": self.floodwait_cooldown_until,
            "recommendation_calls_attempted": self.recommendation_calls_attempted,
            "recommendation_calls_ok": self.recommendation_calls_ok,
            "recommendation_channels_returned": self.recommendation_channels_returned,
            "recommendation_channels_added_to_frontier": self.recommendation_channels_added_to_frontier,
            "history_sources_attempted": self.history_sources_attempted,
            "history_sources_ok": self.history_sources_ok,
            "history_sources_target": self.max_history_sources,
            "history_posts_fetched": self.history_posts_fetched,
            "media_downloads_attempted": self.media_downloads_attempted,
            "media_downloads_ok": self.media_downloads_ok,
            "telegram_governor_decisions_json": json.dumps({"max_network_resolves": self.max_network_resolves, "max_history_sources": self.max_history_sources, "max_recommendation_calls": self.max_recommendation_calls}, ensure_ascii=False),
        }


def normalize_geo_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").lower().replace("ё", "е")).strip()


def text_main_content_for_geo(text: str) -> str:
    kept: list[str] = []
    for raw_line in (text or "").splitlines():
        line = raw_line.strip()
        low = normalize_geo_text(line)
        if not line:
            continue
        if URL_RE.search(line) and len(line) < 140:
            continue
        if low.startswith(("подпис", "источник:", "фото:", "подробнее", "читать", "наш сайт", "реклама")):
            continue
        if line.startswith("#") or ("#" in line and len(line) < 80):
            continue
        kept.append(line)
    return "\n".join(kept) or (text or "")


def split_semicolon(value: str) -> list[str]:
    return [x.strip() for x in re.split(r"\s*;\s*", value or "") if x.strip()]


def find_place_lexicon_file(config: dict[str, Any] | None = None) -> Path | None:
    raw_candidates = [
        os.getenv("REGION_TALK_PLACE_LEXICON_FILE"),
        (config or {}).get("place_lexicon_file") if config else None,
        "kaliningrad-place-lexicon-v1.csv",
        "docs/features/region-talk-channel/kaliningrad-place-lexicon-v1.csv",
    ]
    candidates = [Path(str(x)) for x in raw_candidates if x]
    for root in [Path.cwd(), Path(__file__).resolve().parent, Path("/kaggle/input")]:
        if root.exists():
            candidates.extend(root.rglob("kaliningrad-place-lexicon-v1.csv"))
    for path in candidates:
        if path.exists():
            return path
    return None


def load_place_lexicon(path: Path | None) -> list[dict[str, Any]]:
    if not path or not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    out: list[dict[str, Any]] = []
    for row in rows:
        clean = {field: str(row.get(field) or "").strip() for field in PLACE_LEXICON_FIELDS}
        terms: list[tuple[str, str]] = [(clean["canonical_name"], "canonical_name")]
        for col in ("aliases", "old_names", "latin_aliases", "common_misspellings"):
            for value in split_semicolon(clean.get(col, "")):
                terms.append((value, col))
        clean["match_terms"] = [(term, kind, normalize_geo_text(term)) for term, kind in terms if term]
        out.append(clean)
    return out


def term_in_text(norm_term: str, norm_text: str) -> bool:
    if not norm_term:
        return False
    escaped = re.escape(norm_term)
    return re.search(rf"(?<![0-9a-zа-я]){escaped}(?![0-9a-zа-я])", norm_text, flags=re.I) is not None


def match_kaliningrad_places(text: str, lexicon: list[dict[str, Any]]) -> list[dict[str, Any]]:
    norm = normalize_geo_text(text)
    matches: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    strong_context_terms = [normalize_geo_text(x) for x in DEFAULT_ANCHORS[:8]] + [
        "калининградская область", "калининградской области", "зеленоградский", "светлогорский",
        "балтийский район", "куршская коса", "балтийская коса", "пос.", "поселок", "поселке",
        "маршрут", "поездка", "путешествие",
    ]
    for place in lexicon:
        if str(place.get("allowed_for_kaliningrad_scope") or "").lower() != "true":
            continue
        for term, kind, norm_term in place.get("match_terms") or []:
            if len(norm_term) < 3:
                continue
            if term_in_text(norm_term, norm):
                key = (place.get("place_id") or place.get("canonical_name") or "", term)
                if key in seen:
                    continue
                seen.add(key)
                idx = norm.find(norm_term)
                raw_text = text or ""
                requires_context = str(place.get("requires_context") or "").lower() == "true"
                context_norm = normalize_geo_text(raw_text[max(0, idx-120):idx+len(term)+120]) if idx >= 0 else norm
                accepted_context = (not requires_context) or any(t and t in context_norm for t in strong_context_terms)
                matches.append({
                    "place_id": place.get("place_id", ""),
                    "matched_place_name": term,
                    "canonical_name": place.get("canonical_name", ""),
                    "place_type": place.get("place_type", ""),
                    "municipality": place.get("municipality", ""),
                    "priority_tier": place.get("priority_tier", ""),
                    "alias_used": term if kind != "canonical_name" else "",
                    "match_context": re.sub(r"\s+", " ", raw_text[max(0, idx-80):idx+len(term)+80])[:240] if idx >= 0 else "",
                    "requires_context": str(requires_context).lower(),
                    "accepted_as_region_evidence": str(bool(accepted_context)).lower(),
                    "match_context_short": re.sub(r"\s+", " ", raw_text[max(0, idx-80):idx+len(term)+80])[:180] if idx >= 0 else "",
                })
    return matches


def external_geo_mentions(text: str) -> tuple[list[str], list[str]]:
    norm = normalize_geo_text(text_main_content_for_geo(text))
    regions = [term for term in EXTERNAL_REGION_TERMS if term_in_text(normalize_geo_text(term), norm)]
    countries = [term for term in EXTERNAL_COUNTRY_TERMS if term_in_text(normalize_geo_text(term), norm)]
    return sorted(set(regions)), sorted(set(countries))


def kaliningrad_oblast_only_scope_gate(text: str, lexicon: list[dict[str, Any]]) -> dict[str, Any]:
    main_text = text_main_content_for_geo(text)
    matches = match_kaliningrad_places(main_text, lexicon)
    strong_matches = [m for m in matches if m.get("accepted_as_region_evidence") == "true"]
    ambiguous_matches = [m for m in matches if m.get("accepted_as_region_evidence") != "true"]
    external_regions, external_countries = external_geo_mentions(text)
    ok = bool(strong_matches) and not external_regions and not external_countries
    reason = ""
    if not matches:
        reason = "reject: no Kaliningrad Oblast place evidence in main content"
    elif external_regions or external_countries:
        reason = "reject: external destinations in main content: " + "; ".join(external_regions + external_countries)
    else:
        reason = "accepted: matched Kaliningrad Oblast places only: " + "; ".join(sorted({m['canonical_name'] for m in matches})[:10])
    return {
        "kaliningrad_oblast_only_scope": ok,
        "matched_place_names": "; ".join(sorted({m["canonical_name"] for m in matches})),
        "matched_place_types": "; ".join(sorted({m["place_type"] for m in matches if m.get("place_type")})),
        "matched_place_priority_tiers": "; ".join(sorted({m["priority_tier"] for m in matches if m.get("priority_tier")})),
        "matched_place_aliases": "; ".join(sorted({m["alias_used"] for m in matches if m.get("alias_used")})),
        "ambiguous_place_names": "; ".join(sorted({m["canonical_name"] for m in ambiguous_matches})),
        "requires_context_place_names": "; ".join(sorted({m["canonical_name"] for m in ambiguous_matches})),
        "matched_place_requires_context": "; ".join(sorted({m["canonical_name"] for m in matches if m.get("requires_context") == "true"})),
        "matched_place_accepted_as_region_evidence": "; ".join(sorted({m["canonical_name"] for m in strong_matches})),
        "matched_place_rejected_as_ambiguous": "; ".join(sorted({m["canonical_name"] for m in ambiguous_matches})),
        "match_context_short": "; ".join([str(m.get("match_context_short") or "") for m in matches[:3] if m.get("match_context_short")]),
        "mentioned_kaliningrad_places": "; ".join(sorted({m["canonical_name"] for m in matches})),
        "mentioned_external_regions": "; ".join(external_regions),
        "mentioned_external_countries": "; ".join(external_countries),
        "external_region_count": len(external_regions),
        "external_country_count": len(external_countries),
        "external_geo_mentions": "; ".join(external_regions + external_countries),
        "region_scope_decision": "accept_kaliningrad_oblast_only" if ok else "reject_not_kaliningrad_oblast_only",
        "region_scope_reason": reason,
        "place_matches": matches,
    }


def parse_post_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def freshness_gate(post_date: Any) -> dict[str, Any]:
    min_raw = os.getenv("REGION_TALK_MIN_POST_DATE", "2026-01-01")
    min_dt = datetime.fromisoformat(min_raw).replace(tzinfo=timezone.utc)
    now = RUN_STARTED_AT
    dt = parse_post_datetime(post_date)
    if dt is None:
        return {"fresh_enough": False, "post_age_days": "", "freshness_score": 0.0, "freshness_reason": "reject: missing post date", "min_post_date": min_raw}
    age_days = max(0, (now - dt).days)
    half_life = max(1, getenv_int("REGION_TALK_FRESHNESS_HALF_LIFE_DAYS", 30))
    score = round(1 / (1 + age_days / half_life), 3)
    ok = dt >= min_dt and dt.year >= min_dt.year
    return {"fresh_enough": ok, "post_age_days": age_days, "freshness_score": score if ok else 0.0, "freshness_reason": "accepted" if ok else f"reject: post_date before {min_raw}", "min_post_date": min_raw}


def score_substance(text: str) -> dict[str, Any]:
    low = normalize_geo_text(text)
    length_score = min(0.35, len(text or "") / 1200)
    useful = min(1.0, 0.18 * sum(1 for w in SUBSTANCE_WORDS if w in low))
    visit = min(1.0, 0.22 * sum(1 for w in VISIT_WORDS if w in low))
    emotion = min(1.0, 0.16 * sum(1 for w in EMOTION_WORDS if w in low))
    memorable = min(1.0, 0.22 * sum(1 for w in MEMORABLE_WORDS if w in low))
    substance = round(min(1.0, length_score + 0.35*useful + 0.30*visit + 0.20*emotion + 0.25*memorable), 3)
    return {
        "text_substance_score": substance,
        "visit_impression_score": round(visit, 3),
        "useful_route_score": round(useful, 3),
        "emotion_observation_score": round(emotion, 3),
        "memorable_details_score": round(memorable, 3),
        "substance_reason": "accepted substantive/experience text" if substance >= 0.25 else "reject: thin text, visual dump, SEO list, or low-context mention",
    }


def ad_promo_gate(text: str) -> dict[str, Any]:
    low = normalize_geo_text(text)
    hard_hits: list[str] = []
    possible_hits: list[str] = []
    for label, pattern in AD_PROMO_HARD_PATTERNS:
        if re.search(pattern, low, flags=re.I):
            hard_hits.append(label)
    for label, pattern in AD_PROMO_POSSIBLE_PATTERNS:
        if re.search(pattern, low, flags=re.I):
            possible_hits.append(label)
    is_hard = bool(hard_hits)
    is_possible = bool(possible_hits)
    reason = (
        "reject: hard ad/promo/announcement cues: " + "; ".join(sorted(set(hard_hits)))
        if is_hard else (
            "possible promo cues, send to LLM if content is strong: " + "; ".join(sorted(set(possible_hits)))
            if is_possible else "accepted: no hard ad/promo cues"
        )
    )
    return {
        "is_ad_or_promo": is_hard,
        "is_ad_or_promo_hard": str(is_hard).lower(),
        "is_ad_or_promo_possible": str(is_possible).lower(),
        "ad_promo_hits": "; ".join(sorted(set(hard_hits))),
        "ad_promo_possible_hits": "; ".join(sorted(set(possible_hits))),
        "ad_promo_reason": reason,
    }


def extract_urls_and_handles(text: str) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for m in URL_RE.finditer(text or ""):
        raw = m.group(0).rstrip(').,;!?:»”')
        norm = raw if raw.startswith(("http://", "https://")) else "https://" + raw.lstrip("/")
        if norm not in seen:
            seen.add(norm); out.append({"raw_url": raw, "normalized_url": norm, "extracted_handle": ""})
    for m in HANDLE_RE.finditer(text or ""):
        handle = m.group(0)
        norm = "https://t.me/" + handle.lstrip("@")
        if norm not in seen:
            seen.add(norm); out.append({"raw_url": handle, "normalized_url": norm, "extracted_handle": handle})
    return out


def classify_platform_url(url: str) -> tuple[str, str]:
    low = (url or "").lower()
    if "t.me/" in low:
        return ("telegram_post" if re.search(r"t\.me/[a-z0-9_]+/\d+", low) else "telegram_channel", "telegram")
    if "vk.com/wall" in low:
        return "vk_wall_post", "vk"
    if "vk.com/" in low:
        return "vk_public", "vk"
    if "dzen.ru" in low:
        return "dzen_article", "dzen"
    if "youtube.com" in low or "youtu.be" in low:
        return "youtube_channel", "youtube"
    if "rutube.ru" in low:
        return "rutube_channel", "rutube"
    return "website", "web"


def discover_links_for_post(post: dict[str, Any], run_id: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    edges: list[dict[str, Any]] = []
    from_source = str(post.get("source_id") or "")
    source_title = str(post.get("source_title") or "")
    post_id = str(post.get("post_id") or "")
    post_url = str(post.get("post_url") or "")
    items = extract_urls_and_handles(str(post.get("text") or ""))
    if post.get("forwarded_from_url"):
        items.append({"raw_url": str(post.get("forwarded_from_url")), "normalized_url": str(post.get("forwarded_from_url")), "extracted_handle": str(post.get("forwarded_from_handle") or "")})
    for i, item in enumerate(items, start=1):
        normalized = item["normalized_url"]
        link_type, platform = classify_platform_url(normalized)
        edge_type = "forward_origin" if normalized == post.get("forwarded_from_url") else "post_text_link"
        cand_id = "src_cand_" + stable_hash(platform, normalized)
        edge_id = "edge_" + stable_hash(from_source, post_id, normalized, edge_type)
        rows.append({
            "source_candidate_id": cand_id, "discovered_from_source": source_title, "discovered_from_post_url": post_url,
            "discovered_from_comment_hash": "", "discovery_type": "forwarded_from" if edge_type == "forward_origin" else "post_text",
            "edge_type": edge_type, "raw_url": item["raw_url"], "normalized_url": normalized, "platform_guess": link_type,
            "candidate_source_status": "source_frontier", "confidence": 0.85 if edge_type == "forward_origin" else 0.55,
        })
        edges.append({
            "edge_id": edge_id, "from_source_id": from_source, "from_source_title": source_title, "from_post_id": post_id,
            "to_source_candidate_id": cand_id, "to_source_title": item.get("extracted_handle") or normalized, "edge_type": edge_type,
            "evidence_url": normalized, "evidence_context_short": make_summary(str(post.get("text_excerpt") or post.get("text") or ""))[:240],
            "confidence": 0.85 if edge_type == "forward_origin" else 0.55, "discovery_depth": 1, "run_id": run_id,
        })
    return rows, edges


def source_candidate_score(row: dict[str, Any]) -> float:
    title = normalize_geo_text(str(row.get("recommended_title") or row.get("title_guess") or row.get("to_source_title") or row.get("normalized_url") or ""))
    positive = ["путешеств", "travel", "места", "росси", "маршрут", "город", "прогул", "балтик", "калининград", "блог"]
    negative = ["скидк", "авиабил", "турфирм", "турагент", "казино", "crypto", "крипт", "ваканси", "работа", "новости", "полит"]
    score = float(row.get("confidence") or 0.45)
    score += min(0.20, 0.04 * sum(1 for w in positive if w in title))
    score -= min(0.35, 0.08 * sum(1 for w in negative if w in title))
    if row.get("edge_type") == "telegram_similar_channel":
        score += 0.12
    if row.get("platform_guess") in {"telegram", "telegram_channel"}:
        score += 0.05
    return round(max(0.0, min(0.95, score)), 3)


def build_source_frontier_unique(rows: list[dict[str, Any]], previous_discovered: dict[str, Any], run_id: str) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    prev_by_key: dict[str, dict[str, Any]] = {}
    for pv in previous_discovered.values() if isinstance(previous_discovered, dict) else []:
        if isinstance(pv, dict):
            pkey = str(pv.get("canonical_source_key") or canonical_source_key(str(pv.get("platform") or pv.get("platform_guess") or ""), str(pv.get("username_or_handle") or ""), str(pv.get("canonical_url") or pv.get("normalized_url") or "")))
            if pkey:
                prev_by_key[pkey] = pv
    for row in rows:
        platform = normalize_source_platform(str(row.get("platform") or row.get("platform_guess") or ""), str(row.get("canonical_url") or row.get("normalized_url") or row.get("raw_url") or ""))
        canonical_url = canonical_source_url(platform, str(row.get("username_or_handle") or row.get("recommended_username") or ""), str(row.get("canonical_url") or row.get("normalized_url") or ""))
        ckey = str(row.get("canonical_source_key") or canonical_source_key(platform, str(row.get("username_or_handle") or row.get("recommended_username") or ""), canonical_url))
        cid = str(row.get("source_candidate_id") or "src_cand_" + stable_hash(ckey or platform, canonical_url))
        prev = previous_discovered.get(cid) or prev_by_key.get(ckey) or {}
        group_key = ckey or cid
        g = grouped.setdefault(group_key, {
            "frontier_source_id": cid,
            "source_candidate_id": cid,
            "platform": platform or prev.get("platform_guess", ""),
            "platform_guess": platform or prev.get("platform_guess", ""),
            "canonical_source_key": ckey,
            "canonical_url": canonical_url or prev.get("normalized_url", ""),
            "normalized_url": canonical_url or prev.get("normalized_url", ""),
            "username_or_handle": row.get("recommended_username") or row.get("username_or_handle") or "",
            "title_guess": row.get("recommended_title") or row.get("title_guess") or row.get("to_source_title") or "",
            "discovery_first_run_id": prev.get("first_seen_run_id") or run_id,
            "discovery_last_run_id": run_id,
            "discovery_count": int(prev.get("seen_run_count") or 0),
            "edge_types_all": set(),
            "discovery_types": set(),
            "seed_sources": set(),
            "best_confidence": 0.0,
            "frontier_priority": 0.0,
            "frontier_status": row.get("frontier_status") or row.get("candidate_source_status") or "source_frontier",
            "resolve_status": row.get("resolve_status") or "",
            "resolve_attempt_count": row.get("resolve_attempt_count") or "",
            "last_resolve_error_code": row.get("last_resolve_error_code") or "",
            "last_resolve_error_message_short": row.get("last_resolve_error_message_short") or "",
            "next_allowed_resolve_at": row.get("next_allowed_resolve_at") or "",
            "probe_status": row.get("probe_status") or "probe_later",
            "monitor_decision": row.get("monitor_decision") or "",
            "monitor_decision_reason": row.get("monitor_decision_reason") or "",
            "private_state_key": row.get("private_state_key") or "",
        })
        g["discovery_count"] = int(g.get("discovery_count") or 0) + 1
        if row.get("edge_type"): g["edge_types_all"].add(str(row.get("edge_type")))
        if row.get("discovery_type"): g["discovery_types"].add(str(row.get("discovery_type")))
        if row.get("discovered_from_source"): g["seed_sources"].add(str(row.get("discovered_from_source")))
        confidence = float(row.get("confidence") or 0)
        if confidence >= float(g.get("best_confidence") or 0):
            g["best_confidence"] = confidence
            g["best_edge_type"] = row.get("edge_type") or ""
            g["best_discovered_from_source"] = row.get("discovered_from_source") or row.get("recommendation_source_channel_title") or ""
            g["best_discovered_from_post_url"] = row.get("discovered_from_post_url") or ""
            g["telegram_similar_seed_channel"] = row.get("recommendation_source_channel_url") or ""
            g["telegram_similar_rank"] = row.get("recommendation_rank") or ""
        g["frontier_priority"] = max(float(g.get("frontier_priority") or 0), source_candidate_score(row))
    out: list[dict[str, Any]] = []
    for g in grouped.values():
        edge_types = sorted(g.pop("edge_types_all") or [])
        discovery_types = sorted(g.pop("discovery_types") or [])
        seed_sources = sorted(g.pop("seed_sources") or [])
        priority = float(g.get("frontier_priority") or 0)
        platform = str(g.get("platform") or g.get("platform_guess") or "").lower()
        target_reason = target_source_url_reason(platform, str(g.get("canonical_url") or g.get("normalized_url") or ""))
        resolve_status = str(g.get("resolve_status") or "")
        prev_status = str(g.get("frontier_status") or "")
        if target_reason != "ok":
            stage = "unsupported"
            status = "unsupported_source_url"
        elif platform == "vk":
            stage = "vk_not_configured"
            status = "vk_wall_setup_required"
        elif platform and not platform.startswith("telegram"):
            stage = "unsupported"
            status = "unsupported_probe_backlog"
        elif resolve_status in {"resolved_network", "resolved_from_private_cache"} or prev_status in {"history_due", "resolved", "source_frontier"} and priority >= 0.55:
            stage = "history_due"
            status = "history_due"
        elif priority >= 0.75:
            stage = "probe_due"
            status = "probe_due"
        elif priority >= 0.35:
            stage = "unresolved"
            status = "unresolved_frontier"
        else:
            stage = "inactive_low_quality"
            status = "inactive_low_quality"
        g.update({
            "edge_types_all": "; ".join(edge_types),
            "discovery_types": "; ".join(discovery_types),
            "seed_sources_evidence_count": len(seed_sources),
            "candidate_source_status": "source_frontier" if stage != "inactive_low_quality" else "inactive_low_quality",
            "frontier_stage": stage,
            "frontier_status": status,
            "source_candidate_score": round(priority, 3),
            "target_source_url_status": target_reason,
            "active_scan_allowed": str(target_reason == "ok").lower(),
            "next_action": "probe_later" if stage in {"unresolved", "probe_due"} else ("fetch_history_when_selected" if stage == "history_due" else ("configure_vk_wall_reader" if stage == "vk_not_configured" else "keep_in_low_quality_pool")),
            "rejection_or_skip_reason": target_reason if target_reason != "ok" else ("" if priority >= 0.35 else "low source candidate score; retained as inactive low-quality pool, not a blocking status"),
        })
        g.update(classify_source_profile({
            "source_title": g.get("title_guess"),
            "resolved_title": g.get("title_guess"),
            "source_kind": g.get("edge_types_all"),
            "source_type": g.get("discovery_types"),
            "canonical_url": g.get("canonical_url"),
        }, []))
        out.append(g)
    return sorted(out, key=lambda r: (-float(r.get("source_candidate_score") or 0), str(r.get("normalized_url") or "")))



CANDIDATE_MEMORY_STAGES = {
    "favorite", "semantic_candidate", "needs_image_review", "image_fetch_retry_needed",
    "good_text_weak_media", "low_substance_but_region_relevant", "manual_keep",
}
ACTIVE_CANDIDATE_MEMORY_STATUSES = {
    "text_candidate_pending_image", "image_fetch_retry_needed", "image_reviewable",
    "good_text_weak_media", "publication_ready_candidate", "manual_keep", "source_not_refetched_this_run",
}
VISIT_EVIDENCE_PATTERN = re.compile(r"\b(были|ездили|поехали|вернулись|посетили|гуляли|увидели|понравил|впечатлил|запомнил|ощущени|эмоци|отзыв|фотоотч[её]т)\b", re.I)
EMOTION_PATTERN = re.compile(r"\b(понравил|впечатлил|запомнил|атмосфер|красив|удивил|эмоци|ощущени|восторг|любов|спокойн|магия)\b", re.I)
ORIGINAL_PHOTO_PATTERN = re.compile(r"\b(мои фото|наши фото|фотоотч[её]т|снял[аи]?|снимали|кадры|фотографи[ия]|подписчик|читател)\b", re.I)


def infer_visit_semantic_fields(text: str, llm_gate: dict[str, Any], substance: dict[str, Any], post: dict[str, Any] | None = None) -> dict[str, Any]:
    low = (text or "").lower()
    llm_type = str(llm_gate.get("content_type") or llm_gate.get("llm_content_type") or "").strip()
    positive_markers = [
        "мы приехали", "я был", "я была", "мы были", "нам понравилось", "советую", "делюсь маршрутом",
        "нашли место", "зашли", "прогулялись", "поехали на выходные", "мой отзыв", "впечатлило",
        "запомнилось", "вот что посмотреть", "наш маршрут", "первым делом", "вернулись из калининграда",
    ]
    marker_hits = [m for m in positive_markers if m in low]
    has_firsthand = bool(marker_hits) or bool(VISIT_EVIDENCE_PATTERN.search(text or "")) or llm_type == "visit_impression_candidate" or float(substance.get("visit_impression_score") or 0) >= 0.25
    has_emotion = bool(EMOTION_PATTERN.search(text or "")) or float(substance.get("emotion_observation_score") or 0) >= 0.20
    has_original_photo = bool(ORIGINAL_PHOTO_PATTERN.search(text or "")) or bool((post or {}).get("is_forwarded_or_repost"))
    if has_firsthand:
        visit_type = "firsthand_author_visit"
    elif "подписчик" in low or "читател" in low or "фотоотчет" in low or "фотоотчёт" in low:
        visit_type = "subscriber_photo_report"
    elif llm_type == "route_useful_candidate" or float(substance.get("useful_route_score") or 0) >= 0.18:
        visit_type = "route_guide"
    elif llm_type in {"single_location_photo_card", "encyclopedic_card_candidate"} or (post or {}).get("has_media"):
        visit_type = "single_location_photo_card"
    elif llm_type == "news_or_event":
        visit_type = "news_or_event"
    elif "рго" in str((post or {}).get("source_title") or "").lower() or "проект" in low:
        visit_type = "official_project_material"
    else:
        visit_type = "unknown"
    text_bucket = "publication_candidate_text" if has_firsthand and has_emotion else ("research_candidate" if visit_type in {"route_guide", "single_location_photo_card", "subscriber_photo_report"} else "low_priority_research")
    return {
        "has_firsthand_visit_evidence": str(has_firsthand).lower(),
        "visit_evidence_type": visit_type,
        "first_person_markers": "; ".join(marker_hits),
        "emotion_or_impression_evidence": str(has_emotion).lower(),
        "review_or_opinion_evidence": str(has_firsthand or has_emotion).lower(),
        "useful_route_evidence": str(float(substance.get("useful_route_score") or 0) >= 0.18 or "маршрут" in low or "что посмотреть" in low).lower(),
        "memorable_detail_evidence": str(float(substance.get("memorable_details_score") or 0) >= 0.20 or "запом" in low).lower(),
        "original_photo_evidence": str(has_original_photo or bool((post or {}).get("primary_media_path"))).lower(),
        "nonlocal_blogger_visit_score": round((0.45 if has_firsthand else 0.0) + (0.20 if has_emotion else 0.0) + (0.15 if marker_hits else 0.0) + (0.10 if (post or {}).get("has_media") else 0.0), 3),
        "publication_story_score": round((0.35 if has_firsthand else 0.0) + (0.25 if has_emotion else 0.0) + (0.20 if float(substance.get("useful_route_score") or 0) >= 0.18 else 0.0) + (0.10 if has_original_photo else 0.0), 3),
        "text_bucket": text_bucket,
    }


def classify_source_profile(srow: dict[str, Any], sampled: list[dict[str, Any]]) -> dict[str, Any]:
    title = str(srow.get("source_title") or srow.get("resolved_title") or "").lower()
    kind = str(srow.get("source_kind") or srow.get("source_type") or "").lower()
    joined = " ".join([title, kind, str(srow.get("canonical_url") or "").lower()])
    n = max(1, len(sampled))
    ko_hits = sum(1 for r in sampled if r.get("kaliningrad_oblast_only_scope"))
    ad_hits = sum(1 for r in sampled if r.get("is_ad_or_promo"))
    news_hits = sum(1 for r in sampled if float(r.get("newsiness_score") or 0) >= 0.45)
    personal_hits = sum(1 for r in sampled if str(r.get("has_firsthand_visit_evidence") or "").lower() == "true" or str(r.get("first_person_markers") or ""))
    ko_ratio = round(ko_hits / n, 3)
    if any(w in joined for w in ["калининград", "кёнигсберг", "kenig", "kgd", "39"]):
        geo = "kaliningrad_local"
    elif sampled and 0 < ko_ratio < 0.7:
        geo = "nonlocal_russia"
    elif ko_ratio >= 0.7:
        geo = "kaliningrad_local"
    else:
        geo = "unknown"
    if any(w in joined for w in ["новост", "сми", "инфо", "афиша"]):
        topic = "local_news" if geo == "kaliningrad_local" else "federal_media"
    elif any(w in joined for w in ["тур", "экскурс", "бронь"]):
        topic = "ads_tours"
    elif any(w in joined for w in ["travel", "trip", "путеше", "маршрут", "around", "vokrug"]):
        topic = "travel_blogger" if personal_hits else "travel_media"
    elif personal_hits:
        topic = "personal_blog"
    elif news_hits / n > 0.5:
        topic = "local_news" if geo == "kaliningrad_local" else "federal_media"
    else:
        topic = "unknown"
    travel_score = round(min(1.0, (0.25 if "travel" in joined or "путеше" in joined else 0.0) + (personal_hits / n) * 0.35 + max(0, 1 - news_hits / n) * 0.20 + max(0, 1 - ad_hits / n) * 0.20), 3)
    personal_score = round(min(1.0, personal_hits / n + (0.2 if topic == "personal_blog" else 0.0)), 3)
    nonlocal_value = round(min(1.0, (0.45 if geo == "nonlocal_russia" else 0.10 if geo == "unknown" else 0.0) + travel_score * 0.35 + (0.20 if 0 < ko_ratio < 0.5 else 0.0)), 3)
    return {
        "source_geo_class": geo,
        "source_topic_class": topic,
        "ko_mention_ratio_recent": ko_ratio,
        "travel_blogger_score": travel_score,
        "personal_voice_score": personal_score,
        "nonlocal_value_score": nonlocal_value,
        "source_priority_reason": f"geo={geo}; topic={topic}; ko_ratio={ko_ratio}; personal_hits={personal_hits}/{len(sampled)}",
    }


def normalize_image_status(stage: str, ms: dict[str, Any]) -> dict[str, Any]:
    input_type = str(ms.get("image_model_input_type") or "")
    has_media_fallback = input_type != "actual_image" and str(ms.get("image_download_status") or "") in {"actual_image_missing_or_fallback", "download_failed_or_missing", "queued_for_region_talk_image_diagnostic"}
    if has_media_fallback:
        return {
            "current_stage": "image_fetch_retry_needed",
            "image_status": "needs_actual_image_fetch",
            "image_bucket": "text_accepted_metadata_only",
            "visual_decision": "pending",
            "next_action": "retry_media_download_or_manual_open",
            "why_not_publication_ready": "metadata-only image fallback; actual image bytes were not scored",
            "is_selected_for_publication": False,
            "image_publication_ready": "false",
            "image_reviewable": "false",
            "image_quality_bucket": "metadata_pending_actual_image",
            "failure_reason": "needs_actual_image_fetch",
        }
    if input_type != "actual_image" and str(ms.get("image_publication_ready")) == "true":
        return {
            "current_stage": "image_fetch_retry_needed",
            "image_status": "needs_actual_image_fetch",
            "image_bucket": "text_accepted_metadata_only",
            "visual_decision": "pending",
            "next_action": "retry_media_download_or_manual_open",
            "why_not_publication_ready": "non-actual image scoring cannot mark a row publication-ready",
            "is_selected_for_publication": False,
            "image_publication_ready": "false",
            "image_reviewable": "false",
            "image_quality_bucket": "metadata_pending_actual_image",
            "failure_reason": "needs_actual_image_fetch",
        }
    if str(ms.get("image_publication_ready")) == "true":
        return {"image_status": "publication_ready", "image_bucket": "text_accepted_image_publication_ready", "visual_decision": "accept", "next_action": "human_final_check", "why_not_publication_ready": ""}
    if str(ms.get("image_reviewable")) == "true":
        return {"image_status": "image_reviewable", "image_bucket": "text_accepted_image_reviewable", "visual_decision": "review", "next_action": "human_image_review", "why_not_publication_ready": "image reviewable but below publication-ready threshold"}
    if input_type == "actual_image":
        return {"image_status": "actual_image_weak", "image_bucket": "text_accepted_image_weak", "visual_decision": "reject_visual", "next_action": "manual_open_if_text_is_strong", "why_not_publication_ready": str(ms.get("failure_reason") or "actual image below threshold")}
    return {"image_status": "text_rejected_no_image_spend" if stage in {"dropped_text_gate", "debug_reject"} else "not_scored", "image_bucket": "text_rejected_no_image_spend", "visual_decision": "not_run", "next_action": "", "why_not_publication_ready": str(ms.get("failure_reason") or "")}


def candidate_lifecycle_status(row: dict[str, Any]) -> str:
    manual = str(row.get("manual_decision") or "").strip().lower()
    if manual == "reject":
        return "manual_reject"
    if manual in {"keep", "manual_keep", "favorite", "approve_for_preview", "approve_for_queue"}:
        return "manual_keep"
    if str(row.get("image_publication_ready")) == "true" or row.get("current_stage") in {"favorite", "semantic_candidate"}:
        return "publication_ready_candidate" if str(row.get("image_publication_ready")) == "true" else "image_reviewable"
    if row.get("current_stage") == "image_fetch_retry_needed" or row.get("image_status") == "needs_actual_image_fetch":
        return "image_fetch_retry_needed"
    if row.get("current_stage") == "needs_image_review" or str(row.get("image_reviewable")) == "true":
        return "image_reviewable"
    if row.get("current_stage") == "good_text_weak_media":
        return "good_text_weak_media"
    if row.get("llm_decision") == "accept" or row.get("text_bucket") in {"research_candidate", "publication_candidate_text"}:
        return "text_candidate_pending_image"
    if row.get("current_stage") == "source_not_refetched_this_run":
        return "source_not_refetched_this_run"
    return "research_candidate"


def is_candidate_memory_row(row: dict[str, Any]) -> bool:
    vector_status = str(row.get("vector_gate_status") or "")
    vector_region_accept = vector_status == "vector_accept_candidate" and str(row.get("vector_content_type") or "") in {"visit_impression_candidate", "route_useful_candidate", "single_location_photo_card"}
    if not row.get("kaliningrad_oblast_only_scope") and not vector_region_accept:
        return row.get("manual_decision") in {"manual_keep", "keep", "favorite"} and str(row.get("manual_region_override") or "").lower() == "true"
    if str(row.get("kaliningrad_mention_role") or "main_subject") not in {"", "main_subject", "unclear"} and not vector_region_accept:
        return False
    return row.get("current_stage") in CANDIDATE_MEMORY_STAGES or row.get("llm_decision") == "accept" or row.get("manual_decision") in {"manual_keep", "keep", "favorite"}


def build_candidate_memory(previous_state: dict[str, Any], current_rows: list[dict[str, Any]], source_rows: list[dict[str, Any]], run_id: str, run_now: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    previous_memory = previous_state.get("candidate_memory") if isinstance(previous_state.get("candidate_memory"), dict) else {}
    previous_posts = previous_state.get("posts") if isinstance(previous_state.get("posts"), dict) else {}
    memory: dict[str, dict[str, Any]] = {str(k): dict(v) for k, v in previous_memory.items() if isinstance(v, dict)}
    # Bootstrap legacy candidate memory from pre-z4 post state.
    for pid, prev in previous_posts.items():
        if not isinstance(prev, dict):
            continue
        if str(prev.get("current_stage") or "") in CANDIDATE_MEMORY_STAGES:
            mid = "cmem_" + stable_hash(str(pid))
            memory.setdefault(mid, {
                "candidate_memory_id": mid, "post_id": pid, "post_url": prev.get("post_url", ""),
                "platform_post_key": prev.get("platform_post_key", ""), "source_id": prev.get("source_id", ""),
                "source_title": prev.get("source_title", ""), "post_date": prev.get("post_date", ""),
                "first_seen_run_id": prev.get("first_seen_run_id", ""), "first_candidate_run_id": prev.get("first_seen_run_id", ""),
                "last_seen_run_id": prev.get("last_seen_run_id", ""), "last_refetched_run_id": prev.get("last_seen_run_id", ""),
                "seen_run_count": prev.get("seen_run_count", 1), "current_lifecycle_status": candidate_lifecycle_status(prev),
                "current_stage": prev.get("current_stage", ""), "best_stage_ever": prev.get("current_stage", ""),
                "best_candidate_score_ever": prev.get("candidate_score_current", ""), "candidate_score_current": prev.get("candidate_score_current", ""),
                "best_media_score_ever": prev.get("media_score_current", ""), "media_score_current": prev.get("media_score_current", ""),
                "manual_decision": prev.get("manual_decision", ""), "next_action": "keep_in_memory_or_refetch_source_later",
                "why_keep_in_memory": "legacy candidate from previous post state", "why_not_publication_ready": "not refetched in this run yet",
            })
    current_by_id = {str(r.get("post_id")): r for r in current_rows if r.get("post_id")}
    for row in current_rows:
        if not is_candidate_memory_row(row):
            continue
        pid = str(row.get("post_id") or stable_hash(row.get("post_url")))
        mid = "cmem_" + stable_hash(pid)
        prev = memory.get(mid) or {}
        status = candidate_lifecycle_status(row)
        prev_score = prev.get("candidate_score_current", "")
        prev_media = prev.get("media_score_current", "")
        best_score = max([float(x) for x in [prev.get("best_candidate_score_ever"), row.get("candidate_score")] if str(x) not in {"", "None"}], default=float(row.get("candidate_score") or 0))
        best_media = max([float(x) for x in [prev.get("best_media_score_ever"), row.get("overall_media_score")] if str(x) not in {"", "None"}], default=float(row.get("overall_media_score") or 0))
        try:
            score_delta = round(float(row.get("candidate_score") or 0) - float(prev_score), 3) if prev_score != "" else ""
        except Exception:
            score_delta = ""
        try:
            media_delta = round(float(row.get("overall_media_score") or 0) - float(prev_media), 3) if prev_media != "" else ""
        except Exception:
            media_delta = ""
        memory[mid] = {
            **prev,
            "candidate_memory_id": mid, "post_id": pid, "post_url": row.get("post_url", prev.get("post_url", "")),
            "platform_post_key": row.get("platform_post_key", prev.get("platform_post_key", "")), "source_id": row.get("source_id", prev.get("source_id", "")),
            "source_title": row.get("source_title", prev.get("source_title", "")), "post_date": row.get("post_date", prev.get("post_date", "")),
            "first_seen_run_id": prev.get("first_seen_run_id") or row.get("first_seen_run_id") or run_id,
            "first_candidate_run_id": prev.get("first_candidate_run_id") or run_id,
            "last_seen_run_id": run_id, "last_refetched_run_id": run_id, "not_refetched_this_run": "false",
            "seen_run_count": int(prev.get("seen_run_count") or 0) + (0 if prev.get("last_refetched_run_id") == run_id else 1),
            "previous_lifecycle_status": prev.get("current_lifecycle_status", ""), "current_lifecycle_status": status,
            "best_lifecycle_status_ever": prev.get("best_lifecycle_status_ever") or status,
            "previous_stage": prev.get("current_stage", ""), "current_stage": row.get("current_stage", ""),
            "best_stage_ever": prev.get("best_stage_ever") or row.get("current_stage", ""),
            "best_candidate_score_ever": round(best_score, 3), "candidate_score_current": row.get("candidate_score", ""), "candidate_score_delta": score_delta,
            "best_media_score_ever": round(best_media, 3), "media_score_current": row.get("overall_media_score", ""), "media_score_delta": media_delta,
            "short_summary": row.get("short_summary", prev.get("short_summary", "")),
            "kaliningrad_oblast_only_scope": row.get("kaliningrad_oblast_only_scope", prev.get("kaliningrad_oblast_only_scope", "")),
            "matched_place_names": row.get("matched_place_names", prev.get("matched_place_names", "")),
            "matched_place_accepted_as_region_evidence": row.get("matched_place_accepted_as_region_evidence", prev.get("matched_place_accepted_as_region_evidence", "")),
            "kaliningrad_mention_role": row.get("kaliningrad_mention_role", prev.get("kaliningrad_mention_role", "")),
            "region_scope_reason": row.get("region_scope_reason", prev.get("region_scope_reason", "")),
            "content_type": row.get("content_type", prev.get("content_type", "")),
            "llm_content_type": row.get("llm_content_type", prev.get("llm_content_type", "")),
            "vector_content_type": row.get("vector_content_type", prev.get("vector_content_type", "")),
            "source_geo_class": row.get("source_geo_class", prev.get("source_geo_class", "")),
            "source_topic_class": row.get("source_topic_class", prev.get("source_topic_class", "")),
            "ko_mention_ratio_recent": row.get("ko_mention_ratio_recent", prev.get("ko_mention_ratio_recent", "")),
            "travel_blogger_score": row.get("travel_blogger_score", prev.get("travel_blogger_score", "")),
            "personal_voice_score": row.get("personal_voice_score", prev.get("personal_voice_score", "")),
            "nonlocal_value_score": row.get("nonlocal_value_score", prev.get("nonlocal_value_score", "")),
            "source_priority_reason": row.get("source_priority_reason", prev.get("source_priority_reason", "")),
            "text_bucket": row.get("text_bucket", ""), "image_bucket": row.get("image_bucket", ""), "image_status": row.get("image_status", ""),
            "image_model_input_type": row.get("image_model_input_type", ""), "image_model_type": row.get("image_model_type", ""),
            "image_publication_ready": row.get("image_publication_ready", "false"), "image_reviewable": row.get("image_reviewable", "false"),
            "visual_decision": row.get("visual_decision", ""),
            "llm_status": row.get("llm_status", ""), "llm_stage": row.get("llm_stage", ""), "needs_llm_final_verify": row.get("needs_llm_final_verify", ""),
            "llm_gate_status": row.get("llm_gate_status", ""), "llm_decision": row.get("llm_decision", ""), "llm_reason": row.get("llm_reason", ""), "llm_model": row.get("llm_model", ""),
            "final_verifier_status": row.get("final_verifier_status", ""), "final_verifier_decision": row.get("final_verifier_decision", ""), "final_verifier_reason": row.get("final_verifier_reason", ""), "final_verifier_model": row.get("final_verifier_model", ""),
            "vector_gate_status": row.get("vector_gate_status", ""), "vector_positive_score": row.get("vector_positive_score", ""),
            "has_firsthand_visit_evidence": row.get("has_firsthand_visit_evidence", ""), "visit_evidence_type": row.get("visit_evidence_type", ""),
            "first_person_markers": row.get("first_person_markers", prev.get("first_person_markers", "")),
            "emotion_or_impression_evidence": row.get("emotion_or_impression_evidence", ""), "review_or_opinion_evidence": row.get("review_or_opinion_evidence", ""),
            "useful_route_evidence": row.get("useful_route_evidence", prev.get("useful_route_evidence", "")),
            "nonlocal_blogger_visit_score": row.get("nonlocal_blogger_visit_score", prev.get("nonlocal_blogger_visit_score", "")),
            "publication_story_score": row.get("publication_story_score", prev.get("publication_story_score", "")),
            "original_photo_evidence": row.get("original_photo_evidence", ""), "manual_decision": row.get("manual_decision") or prev.get("manual_decision", ""),
            "manual_decision_at": prev.get("manual_decision_at", ""), "next_action": row.get("next_action") or row.get("suggested_action") or "manual_review",
            "why_keep_in_memory": "LLM/text/image gate reached research candidate status", "why_not_publication_ready": row.get("why_not_publication_ready") or row.get("rejection_reason") or "",
            "expires_at": prev.get("expires_at", ""),
        }
    source_status = {str(s.get("source_id") or ""): s for s in source_rows}
    for mid, rec in list(memory.items()):
        if rec.get("post_id") in current_by_id:
            continue
        sid = str(rec.get("source_id") or "")
        srow = source_status.get(sid) or {}
        rec["not_refetched_this_run"] = "true"
        rec["current_lifecycle_status"] = "source_not_refetched_this_run" if rec.get("current_lifecycle_status") not in {"manual_reject", "expired"} else rec.get("current_lifecycle_status")
        rec["visibility_status"] = "not_refetched_this_run"
        rec["source_fetch_status_this_run"] = srow.get("fetch_status", "source_not_selected_this_run")
        rec["next_action"] = "keep_in_memory_or_refetch_source_later"
        rec["why_keep_in_memory"] = rec.get("why_keep_in_memory") or "previous candidate retained despite source not being refetched"
        rec["why_not_publication_ready"] = rec.get("why_not_publication_ready") or "source not refetched this run"
    region_excluded: list[dict[str, Any]] = []
    for mid, rec in list(memory.items()):
        if rec.get("manual_decision") in {"manual_keep", "keep", "favorite"}:
            continue
        vector_region_accept = str(rec.get("vector_gate_status") or "") == "vector_accept_candidate" and str(rec.get("vector_content_type") or "") in {"visit_impression_candidate", "route_useful_candidate", "single_location_photo_card"}
        if (str(rec.get("kaliningrad_oblast_only_scope") or "").lower() not in {"true", "1", "yes"} and not vector_region_accept) or (str(rec.get("kaliningrad_mention_role") or "main_subject") not in {"", "main_subject", "unclear"} and not vector_region_accept):
            rec["current_lifecycle_status"] = "region_evidence_missing_needs_refetch"
            rec["visibility_status"] = "excluded_from_candidate_memory_by_region_vector_gate"
            rec["next_action"] = "refetch_source_or_manual_region_override"
            rec["why_not_publication_ready"] = "region/vector evidence missing in candidate memory"
            region_excluded.append({**rec, "candidate_memory_id": mid})
            memory.pop(mid, None)
    rows = sorted(memory.values(), key=lambda r: (str(r.get("manual_decision") or "") == "reject", str(r.get("not_refetched_this_run") or "") == "true", -float(r.get("best_candidate_score_ever") or 0), str(r.get("post_date") or "")))
    not_refetched = [r for r in rows if str(r.get("not_refetched_this_run")) == "true" and r.get("current_lifecycle_status") != "manual_reject"]
    deltas = []
    for r in rows:
        prev_stage = str(r.get("previous_stage") or "")
        cur_stage = str(r.get("current_stage") or "")
        if str(r.get("not_refetched_this_run")) == "true":
            bucket = "not_refetched_this_run"
        elif not prev_stage:
            bucket = "new_to_system"
        elif prev_stage == cur_stage:
            bucket = "stage_unchanged"
        elif cur_stage in {"favorite", "semantic_candidate", "needs_image_review", "image_fetch_retry_needed", "good_text_weak_media"} and prev_stage not in CANDIDATE_MEMORY_STAGES:
            bucket = "became_candidate"
        elif float(r.get("candidate_score_delta") or 0) > 0.02:
            bucket = "stage_upgraded"
        elif float(r.get("candidate_score_delta") or 0) < -0.02:
            bucket = "stage_downgraded"
        else:
            bucket = "re_seen"
        deltas.append({**r, "delta_bucket": bucket, "current_run_id": run_id})
    for r in region_excluded:
        deltas.append({**r, "delta_bucket": "excluded_by_hard_region_gate", "current_run_id": run_id})
    return rows, not_refetched, deltas


def build_source_frontier_queue_next(frontier_rows: list[dict[str, Any]], source_rows: list[dict[str, Any]], candidate_memory: list[dict[str, Any]], run_id: str) -> list[dict[str, Any]]:
    candidate_source_ids = {str(r.get("source_id") or "") for r in candidate_memory if r.get("source_id")}
    resolved = {str(s.get("canonical_url") or "") for s in source_rows if s.get("telegram_resolve_status") in {"resolved_network", "resolved_from_private_cache"} or s.get("fetch_status") == "ok"}
    out = []
    for row in frontier_rows:
        url = str(row.get("canonical_url") or row.get("normalized_url") or "")
        platform = str(row.get("platform") or row.get("platform_guess") or "")
        edge_types = str(row.get("edge_types_all") or row.get("best_edge_type") or "")
        discovery = str(row.get("discovery_types") or "")
        score = float(row.get("source_candidate_score") or row.get("frontier_priority") or 0)
        pclass = "P3"
        why = "low priority/generic backlog"
        if str(row.get("source_candidate_id") or "") in candidate_source_ids or url in resolved:
            pclass, why, score = "P0", "previous candidate or already resolved source", score + 0.35
        elif "telegram_similar_channel" in edge_types:
            pclass, why, score = "P1", "Telegram similar-channel edge", score + 0.20
        elif any(w in (url + " " + str(row.get("title_guess") or "")).lower() for w in ["travel", "trip", "photo", "путеше", "блог", "фото", "route", "туризм"]):
            pclass, why, score = "P1", "author/travel/photo/catalog signal", score + 0.12
        elif "public_travel_blogger_catalog" in discovery or "public_travel_blogger_catalog" in edge_types:
            pclass, why, score = "P2", "sample public travel blogger catalog", score + 0.05
        frontier_stage = str(row.get("frontier_stage") or ("vk_not_configured" if platform == "vk" else ("unsupported" if platform and not platform.startswith("telegram") else "unresolved")))
        planned = "fetch_cached_entity_history" if frontier_stage == "history_due" else ("probe_telegram_entity_then_history" if platform.startswith("telegram") else ("probe_vk_wall_when_configured" if platform == "vk" else "keep_frontier"))
        blocked_reason = "vk_wall_not_configured" if platform == "vk" and not (os.getenv("VK_ACCESS_TOKEN") or os.getenv("VK_SERVICE_TOKEN") or os.getenv("VK_TOKEN")) else ""
        out.append({
            "queue_rank": 0, "canonical_url": url, "platform": platform, "title_guess": row.get("title_guess") or row.get("username_or_handle") or url,
            "frontier_priority": round(score, 3), "frontier_stage": frontier_stage, "frontier_status": row.get("frontier_status", ""), "why_next": why,
            "discovery_types": discovery, "best_edge_type": row.get("best_edge_type") or edge_types,
            "source_quality_score": round(min(1.0, score), 3), "kaliningrad_prior_score": row.get("best_confidence", ""),
            "authorial_voice_score": 0.65 if pclass in {"P0", "P1"} else 0.35,
            "original_photo_prior_score": 0.60 if pclass in {"P0", "P1"} else 0.30,
            "blogger_source_score": 0.70 if any(w in (url + " " + str(row.get("title_guess") or "")).lower() for w in ["travel", "путеше", "блог", "trip"]) else 0.35,
            "ad_risk_score": 0.20 if pclass in {"P0", "P1", "P2"} else 0.45,
            "resolve_status": row.get("resolve_status", ""), "next_allowed_resolve_at": row.get("next_allowed_resolve_at", ""),
            "probe_status": row.get("probe_status", "probe_later"), "planned_action": planned, "planned_after_run_id": run_id,
            "blocked_reason": blocked_reason, "promotion_class": pclass,
        })
    selected = []
    for pclass, limit in [("P0", 25), ("P1", 40), ("P2", 25), ("P3", 10)]:
        selected.extend(sorted([r for r in out if r["promotion_class"] == pclass], key=lambda r: (-float(r.get("frontier_priority") or 0), r.get("canonical_url") or ""))[:limit])
    seen = set(); final = []
    for row in selected:
        key = row.get("canonical_url")
        if key in seen:
            continue
        seen.add(key); final.append(row)
    for i, row in enumerate(final[:100], start=1):
        row["queue_rank"] = i
    return final[:100]


def _source_queue_seed_rejection_reason(row: dict[str, Any]) -> str:
    platform = normalize_source_platform(str(row.get("platform") or row.get("platform_guess") or ""), str(row.get("canonical_url") or row.get("normalized_url") or row.get("source_url") or row.get("url") or row.get("keyword_hit_source_url") or row.get("recommended_canonical_url") or row.get("raw_url") or ""))
    if platform not in {"telegram", "vk"}:
        return "unsupported_platform_for_source_queue"
    url = canonical_source_url(platform, str(row.get("handle") or row.get("username_or_handle") or row.get("recommended_username") or ""), str(row.get("canonical_url") or row.get("normalized_url") or row.get("source_url") or row.get("url") or row.get("keyword_hit_source_url") or row.get("recommended_canonical_url") or row.get("raw_url") or ""))
    if not url:
        return "missing_source_url"
    return target_source_url_reason(platform, url)


def _source_queue_seed_from_row(row: dict[str, Any], *, default_added_from: str = "") -> dict[str, Any] | None:
    platform = normalize_source_platform(str(row.get("platform") or row.get("platform_guess") or ""), str(row.get("canonical_url") or row.get("normalized_url") or row.get("source_url") or row.get("url") or row.get("keyword_hit_source_url") or row.get("recommended_canonical_url") or row.get("raw_url") or ""))
    if platform not in {"telegram", "vk"}:
        return None
    url = canonical_source_url(platform, str(row.get("handle") or row.get("username_or_handle") or row.get("recommended_username") or ""), str(row.get("canonical_url") or row.get("normalized_url") or row.get("source_url") or row.get("url") or row.get("keyword_hit_source_url") or row.get("recommended_canonical_url") or row.get("raw_url") or ""))
    ckey = str(row.get("canonical_source_key") or canonical_source_key(platform, str(row.get("handle") or row.get("username_or_handle") or row.get("recommended_username") or ""), url))
    if not url or not ckey:
        return None
    target_reason = target_source_url_reason(platform, url)
    if target_reason != "ok":
        return None
    added_from = default_added_from or str(row.get("discovery_type") or row.get("edge_type") or row.get("source_kind") or row.get("source_type") or "unknown")
    return {
        "canonical_source_key": ckey,
        "platform": platform,
        "source_url": url,
        "canonical_url": url,
        "handle": canonical_handle(str(row.get("handle") or row.get("username_or_handle") or row.get("recommended_username") or url)),
        "source_title": row.get("source_title") or row.get("title_guess") or row.get("recommended_title") or row.get("resolved_title") or row.get("seed_channel_title") or row.get("source_seed_id") or url,
        "added_from": added_from,
        "source_candidate_id": row.get("source_candidate_id") or row.get("frontier_source_id") or "",
        "source_id": row.get("source_id") or "",
        "discovery_types": row.get("discovery_types") or row.get("discovery_type") or "",
        "edge_types_all": row.get("edge_types_all") or row.get("edge_type") or "",
        "source_priority_score": row.get("source_candidate_score") or row.get("frontier_priority") or row.get("monitor_priority_score") or "",
    }


def _previous_rows_dict(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, dict):
        return [dict(v) for v in value.values() if isinstance(v, dict)]
    if isinstance(value, list):
        return [dict(v) for v in value if isinstance(v, dict)]
    return []


def build_unified_source_queue(
    previous_state: dict[str, Any],
    seeds: list[Seed],
    source_rows: list[dict[str, Any]],
    source_frontier_unique: list[dict[str, Any]],
    public_blogger_rows: list[dict[str, Any]],
    keyword_discovery_rows: list[dict[str, Any]],
    keyword_post_hit_rows: list[dict[str, Any]],
    posts_by_source: dict[str, list[dict[str, Any]]],
    run_id: str,
    run_now: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Build the single human-visible source URL queue.

    Contract: only Telegram/VK sources, deduped by canonical_source_key. Rows
    discovered by Telegram keyword search are inserted immediately after the
    persisted cursor so they become the next backlog; all other new sources are
    appended to the tail. Existing order is preserved.
    """
    previous_queue_rows = _previous_rows_dict(previous_state.get("unified_source_queue") or previous_state.get("canonical_source_queue"))
    previous_image_queue_rows = _previous_rows_dict(previous_state.get("image_candidate_queue"))
    prev_cursor = int(previous_state.get("unified_source_queue_cursor_position") or previous_state.get("canonical_source_cursor_position") or 0)
    entries: dict[str, dict[str, Any]] = {}
    skipped_non_target_queue_rows = 0

    def merge_existing(row: dict[str, Any]) -> None:
        nonlocal skipped_non_target_queue_rows
        seed = _source_queue_seed_from_row(row)
        if not seed:
            if _source_queue_seed_rejection_reason(row) != "missing_source_url":
                skipped_non_target_queue_rows += 1
            return
        key = seed["canonical_source_key"]
        order = int(row.get("queue_order") or 0)
        entries[key] = {**row, **{k: v for k, v in seed.items() if v not in ("", None)}, "queue_order": order}

    for row in sorted(previous_queue_rows, key=lambda r: int(r.get("queue_order") or 999999999)):
        merge_existing(row)

    def append_or_update(row: dict[str, Any], *, added_from: str) -> bool:
        nonlocal skipped_non_target_queue_rows
        seed = _source_queue_seed_from_row(row, default_added_from=added_from)
        if not seed:
            if _source_queue_seed_rejection_reason(row) != "missing_source_url":
                skipped_non_target_queue_rows += 1
            return False
        key = seed["canonical_source_key"]
        if key in entries:
            entries[key].update({k: v for k, v in seed.items() if v not in ("", None)})
            return False
        next_order = max([int(v.get("queue_order") or 0) for v in entries.values()] or [0]) + 1
        entries[key] = {
            **seed,
            "source_queue_id": "srcq_" + stable_hash(key),
            "queue_order": next_order,
            "added_at": run_now,
            "added_from": added_from,
            "first_seen_run_id": run_id,
        }
        return True

    for s in seeds:
        append_or_update({**asdict(s), "canonical_url": s.canonical_url, "source_id": s.source_id}, added_from="seed_file")
    for collection, origin in [
        (source_rows, "monitored_source"),
        (source_frontier_unique, "frontier"),
        (public_blogger_rows, "public_travel_blogger_catalog"),
    ]:
        for row in collection:
            append_or_update(row, added_from=origin)

    keyword_inserted = 0
    keyword_candidates = list(keyword_discovery_rows or []) + list(keyword_post_hit_rows or [])
    missing_keyword: list[dict[str, Any]] = []
    seen_keyword: set[str] = set()
    for row in keyword_candidates:
        seed = _source_queue_seed_from_row(row, default_added_from="telegram_keyword_search")
        if not seed or seed["platform"] != "telegram" or seed["canonical_source_key"] in entries or seed["canonical_source_key"] in seen_keyword:
            continue
        seen_keyword.add(seed["canonical_source_key"])
        missing_keyword.append(seed)
    if missing_keyword:
        for rec in entries.values():
            if int(rec.get("queue_order") or 0) > prev_cursor:
                rec["queue_order"] = int(rec.get("queue_order") or 0) + len(missing_keyword)
        for offset, seed in enumerate(missing_keyword, start=1):
            key = seed["canonical_source_key"]
            entries[key] = {
                **seed,
                "source_queue_id": "srcq_" + stable_hash(key),
                "queue_order": prev_cursor + offset,
                "added_at": run_now,
                "added_from": "telegram_keyword_search",
                "first_seen_run_id": run_id,
                "insertion_policy": "insert_after_cursor",
            }
            keyword_inserted += 1

    source_by_key: dict[str, dict[str, Any]] = {}
    for srow in source_rows:
        seed = _source_queue_seed_from_row(srow)
        if seed:
            source_by_key[seed["canonical_source_key"]] = srow
    processed_orders: list[int] = []
    out: list[dict[str, Any]] = []
    for key, rec in entries.items():
        srow = source_by_key.get(key) or {}
        sid = str(srow.get("source_id") or rec.get("source_id") or "")
        sampled = posts_by_source.get(sid, []) if sid else []
        ko_posts = sum(1 for r in sampled if r.get("kaliningrad_oblast_only_scope"))
        candidate_posts = sum(1 for r in sampled if r.get("current_stage") in CANDIDATE_MEMORY_STAGES)
        image_quality_min_n = getenv_int("REGION_TALK_SOURCE_IMAGE_MIN_ACTUAL_SCORED", 3)
        try:
            image_quality_min_score = float(os.getenv("REGION_TALK_SOURCE_IMAGE_MIN_AVG_SCORE", "0.55"))
        except Exception:
            image_quality_min_score = 0.55
        visual_rows = [
            r for r in sampled
            if r.get("kaliningrad_oblast_only_scope") and not r.get("is_ad_or_promo") and r.get("current_stage") in CANDIDATE_MEMORY_STAGES
        ]
        source_urls = {str(rec.get("source_url") or ""), str(rec.get("canonical_url") or ""), str(srow.get("source_url") or ""), str(srow.get("canonical_url") or "")}
        source_urls = {u.rstrip("/") for u in source_urls if u}
        previous_visual_rows = []
        for ir in previous_image_queue_rows:
            ir_source_url = str(ir.get("source_url") or "").rstrip("/")
            if sid and str(ir.get("source_id") or "") == sid:
                previous_visual_rows.append(ir)
            elif ir_source_url and ir_source_url in source_urls:
                previous_visual_rows.append(ir)
        visual_rows_all = visual_rows + [r for r in previous_visual_rows if image_queue_text_region_confirmed(r)]
        actual_scores: list[float] = []
        seen_score_rows: set[str] = set()
        for vr in visual_rows_all:
            if str(vr.get("image_model_input_type") or "") != "actual_image" and str(vr.get("image_queue_status") or "") != "actual_scored":
                continue
            score_key = str(vr.get("post_url") or vr.get("post_id") or id(vr))
            if score_key in seen_score_rows:
                continue
            seen_score_rows.add(score_key)
            try:
                actual_scores.append(float(vr.get("overall_media_score") or vr.get("final_visual_score") or 0))
            except Exception:
                pass
        ko_posts = max(ko_posts, int(rec.get("ko_posts_found") or 0), sum(1 for r in previous_visual_rows if image_queue_text_region_confirmed(r)))
        candidate_posts = max(candidate_posts, int(rec.get("candidate_posts_found") or 0), sum(1 for r in previous_visual_rows if image_queue_text_region_confirmed(r) and str(r.get("current_stage") or "") in CANDIDATE_MEMORY_STAGES))
        actual_n = len(actual_scores) if actual_scores else int(rec.get("actual_images_scored_count") or 0)
        avg_score = round(sum(actual_scores) / len(actual_scores), 3) if actual_scores else rec.get("avg_actual_image_score", "")
        low_count = sum(1 for x in actual_scores if x < image_quality_min_score) if actual_scores else int(rec.get("low_actual_image_count") or 0)
        if candidate_posts > 0 and actual_n >= image_quality_min_n and avg_score != "" and float(avg_score) < image_quality_min_score:
            image_quality_source_status = "exclude_low_image_quality"
            monitoring_exclusion_reason = "kaliningrad_posts_found_but_actual_images_systematically_low_score"
        elif candidate_posts > 0 and actual_n > 0:
            image_quality_source_status = "monitor_candidate_image_quality_ok"
            monitoring_exclusion_reason = ""
        elif candidate_posts > 0:
            image_quality_source_status = "needs_more_actual_image_evidence"
            monitoring_exclusion_reason = ""
        elif ko_posts > 0:
            image_quality_source_status = "ko_posts_no_candidate_images_yet"
            monitoring_exclusion_reason = ""
        else:
            image_quality_source_status = "no_ko_posts_yet"
            monitoring_exclusion_reason = ""
        fetch_status = str(srow.get("fetch_status") or rec.get("last_scan_status") or "")
        previous_queue_status = str(rec.get("source_queue_status") or "")
        scanned = bool(fetch_status) or previous_queue_status.startswith("processed") or bool(rec.get("last_processed_at")) or actual_n > 0 or ko_posts > 0 or candidate_posts > 0
        if scanned:
            processed_orders.append(int(rec.get("queue_order") or 0))
        if scanned and image_quality_source_status == "exclude_low_image_quality":
            qstatus, color, next_action = "processed_found_ko_low_image_quality", "yellow_retry", "exclude_from_monitoring_candidates_but_keep_posts_for_review"
        elif scanned and (ko_posts > 0 or candidate_posts > 0):
            qstatus, color, next_action = "processed_found_ko_candidate", "green_found_ko", "prioritize_delta_rescan_or_manual_review_posts"
        elif scanned and fetch_status == "ok":
            qstatus, color, next_action = "processed_no_ko", "red_no_ko", "rescan_later_or_deprioritize"
        elif scanned:
            qstatus, color, next_action = "needs_rescan_or_retry", "yellow_retry", "retry_or_fix_source_access"
        else:
            qstatus, color, next_action = "pending_scan", "white_pending", "scan_when_cursor_reaches_source"
        previous_status = str(rec.get("source_queue_status") or "")
        status_changed = bool(previous_status and previous_status != qstatus)
        out.append({
            **rec,
            "source_queue_id": rec.get("source_queue_id") or "srcq_" + stable_hash(key),
            "source_queue_status": qstatus,
            "previous_source_queue_status": previous_status if status_changed else rec.get("previous_source_queue_status", ""),
            "status_changed_this_run": str(status_changed).lower(),
            "last_status_changed_at": run_now if status_changed or not rec.get("last_status_changed_at") else rec.get("last_status_changed_at"),
            "status_color_hint": color,
            "row_fill_color": color,
            "last_processed_at": run_now if scanned else rec.get("last_processed_at", ""),
            "last_scan_run_id": run_id if scanned else rec.get("last_scan_run_id", ""),
            "last_scan_status": fetch_status,
            "posts_scanned": int(srow.get("posts_scanned") or len(sampled) or rec.get("posts_scanned") or 0),
            "ko_posts_found": ko_posts,
            "candidate_posts_found": candidate_posts,
            "actual_images_scored_count": actual_n,
            "avg_actual_image_score": avg_score,
            "low_actual_image_count": low_count,
            "source_image_quality_status": image_quality_source_status,
            "source_image_quality_min_actual_scored": image_quality_min_n,
            "source_image_quality_min_avg_score": image_quality_min_score,
            "monitoring_exclusion_reason": monitoring_exclusion_reason,
            "next_action": next_action,
            "queue_item_updated_at": run_now,
        })
    cursor_position = max([prev_cursor] + processed_orders) if out else 0
    cursor_key = next((str(r.get("canonical_source_key") or "") for r in out if int(r.get("queue_order") or 0) == cursor_position), "")
    display_rows = sorted(out, key=lambda r: (0 if str(r.get("source_queue_status") or "").startswith("processed") or str(r.get("source_queue_status")) == "needs_rescan_or_retry" else 1, int(r.get("queue_order") or 0)))
    next_pending_marked = False
    for idx, row in enumerate(display_rows, start=1):
        order = int(row.get("queue_order") or 0)
        row["display_order"] = idx
        row["cursor_marker"] = "cursor_here" if order == cursor_position else ""
        row["is_after_cursor"] = str(order > cursor_position).lower()
        if not next_pending_marked and order > cursor_position:
            row["cursor_marker"] = "next_after_cursor"
            next_pending_marked = True
    metrics = {
        "source_queue_total": len(out),
        "source_queue_pending_total": sum(1 for r in out if r.get("source_queue_status") == "pending_scan"),
        "source_queue_processed_total": sum(1 for r in out if str(r.get("source_queue_status") or "").startswith("processed")),
        "source_queue_retry_total": sum(1 for r in out if r.get("source_queue_status") == "needs_rescan_or_retry"),
        "source_queue_cursor_position": cursor_position,
        "source_queue_cursor_key": cursor_key,
        "source_queue_keyword_inserted_this_run": keyword_inserted,
        "source_queue_catalog_sources_total": sum(1 for r in out if "public_travel_blogger_catalog" in str(r.get("added_from") or r.get("discovery_types") or "")),
        "source_queue_telegram_total": sum(1 for r in out if r.get("platform") == "telegram"),
        "source_queue_vk_total": sum(1 for r in out if r.get("platform") == "vk"),
        "source_queue_pending_telegram_total": sum(1 for r in out if r.get("platform") == "telegram" and r.get("source_queue_status") == "pending_scan"),
        "source_queue_pending_vk_total": sum(1 for r in out if r.get("platform") == "vk" and r.get("source_queue_status") == "pending_scan"),
        "source_queue_non_target_skipped_this_run": skipped_non_target_queue_rows,
        "source_queue_low_image_quality_excluded_total": sum(1 for r in out if r.get("source_image_quality_status") == "exclude_low_image_quality"),
        "source_queue_only_telegram_vk": str(all(r.get("platform") in {"telegram", "vk"} for r in out)).lower(),
        "source_queue_only_target_source_urls": str(all(is_target_source_url(str(r.get("platform") or ""), str(r.get("canonical_url") or r.get("source_url") or "")) for r in out)).lower(),
    }
    return display_rows, metrics


def image_queue_text_region_confirmed(row: dict[str, Any]) -> bool:
    """Only text-confirmed Kaliningrad Oblast candidate posts may enter image analysis.

    The image queue is downstream of semantic/vector text gates. Rows that are
    merely image/media rows, old queue carry-over, ads, other-region posts, or
    vector rejects must not be sent to image diagnostics.
    """
    if str(row.get("is_ad_or_promo") or "").lower() in {"true", "1", "yes"}:
        return False
    vector_status = str(row.get("vector_gate_status") or row.get("memory_vector_recheck_status") or "")
    if vector_status.startswith("vector_reject"):
        return False
    if str(row.get("kaliningrad_oblast_only_scope") or "").lower() not in {"true", "1", "yes"}:
        return False
    mention_role = str(row.get("kaliningrad_mention_role") or "main_subject")
    if mention_role not in {"", "main_subject", "unclear"}:
        return False
    if str(row.get("external_geo_mentions") or row.get("mentioned_external_regions") or row.get("mentioned_external_countries") or "").strip():
        # Multi-region/travel roundups and other-region official posts are text failures for image analysis.
        return False
    stage = str(row.get("current_stage") or row.get("current_lifecycle_status") or "")
    if stage and stage not in CANDIDATE_MEMORY_STAGES and stage not in ACTIVE_CANDIDATE_MEMORY_STATUSES and stage not in {"actual_scored", "needs_actual_image_fetch"}:
        return False
    return True


def build_image_candidate_queue(
    previous_state: dict[str, Any],
    new_posts: list[dict[str, Any]],
    candidate_memory_rows: list[dict[str, Any]],
    media_rows: list[dict[str, Any]],
    run_id: str,
    run_now: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    previous_queue_rows = _previous_rows_dict(previous_state.get("image_candidate_queue"))
    entries: dict[str, dict[str, Any]] = {}
    prev_cursor = int(previous_state.get("image_candidate_queue_cursor_position") or 0)
    image_queue_pruned_non_region_previous = 0
    image_queue_rejected_non_region_inputs = 0
    for row in sorted(previous_queue_rows, key=lambda r: int(r.get("image_queue_order") or 999999999)):
        key = str(row.get("post_url") or row.get("post_id") or row.get("image_queue_id") or "")
        if not key:
            continue
        if not image_queue_text_region_confirmed(row):
            image_queue_pruned_non_region_previous += 1
            continue
        entries[key] = dict(row)
    media_by_url = {str(r.get("post_url") or ""): r for r in media_rows if r.get("post_url")}

    def add_or_update(row: dict[str, Any], *, added_from: str) -> None:
        nonlocal image_queue_rejected_non_region_inputs
        post_url = str(row.get("post_url") or "")
        key = post_url or str(row.get("post_id") or "")
        if not key:
            return
        if added_from != "media_scoring" and not image_queue_text_region_confirmed(row):
            image_queue_rejected_non_region_inputs += 1
            return
        if added_from == "media_scoring" and key not in entries and not image_queue_text_region_confirmed(row):
            image_queue_rejected_non_region_inputs += 1
            return
        media = media_by_url.get(post_url) or row
        actual = str(media.get("image_model_input_type") or row.get("image_model_input_type") or "") == "actual_image"
        metadata = str(media.get("image_model_input_type") or row.get("image_model_input_type") or "") == "metadata_only"
        if key not in entries:
            order = max([int(v.get("image_queue_order") or 0) for v in entries.values()] or [0]) + 1
            entries[key] = {"image_queue_id": "imgq_" + stable_hash(key), "image_queue_order": order, "added_at": run_now, "added_from": added_from}
        previous_status = str(entries[key].get("image_queue_status") or "")
        if actual or previous_status == "actual_scored":
            status = "actual_scored"
        elif previous_status == "image_analysis_in_progress":
            status = "image_analysis_in_progress"
        else:
            status = "needs_actual_image_fetch" if metadata or row.get("has_media") else "not_reviewable_no_media"
        color = "green_found_ko" if status == "actual_scored" else ("blue_cursor" if status == "image_analysis_in_progress" else ("yellow_retry" if status == "needs_actual_image_fetch" else "red_no_ko"))
        status_changed = bool(previous_status and previous_status != status)
        entries[key].update({
            "previous_image_queue_status": previous_status if status_changed else entries[key].get("previous_image_queue_status", ""),
            "status_changed_this_run": str(status_changed).lower(),
            "last_status_changed_at": run_now if status_changed or not entries[key].get("last_status_changed_at") else entries[key].get("last_status_changed_at"),
            "post_id": row.get("post_id", entries[key].get("post_id", "")),
            "post_url": post_url,
            "platform_post_key": row.get("platform_post_key", entries[key].get("platform_post_key", "")),
            "source_id": row.get("source_id", entries[key].get("source_id", "")),
            "source_title": row.get("source_title", entries[key].get("source_title", "")),
            "source_url": row.get("source_url", entries[key].get("source_url", "")),
            "post_date": row.get("post_date", entries[key].get("post_date", "")),
            "kaliningrad_oblast_only_scope": row.get("kaliningrad_oblast_only_scope", entries[key].get("kaliningrad_oblast_only_scope", "")),
            "kaliningrad_mention_role": row.get("kaliningrad_mention_role", entries[key].get("kaliningrad_mention_role", "")),
            "matched_place_names": row.get("matched_place_names", entries[key].get("matched_place_names", "")),
            "external_geo_mentions": row.get("external_geo_mentions", entries[key].get("external_geo_mentions", "")),
            "mentioned_external_regions": row.get("mentioned_external_regions", entries[key].get("mentioned_external_regions", "")),
            "is_ad_or_promo": row.get("is_ad_or_promo", entries[key].get("is_ad_or_promo", "")),
            "current_stage": row.get("current_stage", entries[key].get("current_stage", "")),
            "current_lifecycle_status": row.get("current_lifecycle_status", entries[key].get("current_lifecycle_status", "")),
            "vector_gate_status": row.get("vector_gate_status", entries[key].get("vector_gate_status", "")),
            "vector_content_type": row.get("vector_content_type", entries[key].get("vector_content_type", "")),
            "text_region_confirmation_status": "text_confirmed_ko_only_for_image_analysis",
            "candidate_score": row.get("candidate_score") or row.get("candidate_score_current") or row.get("best_candidate_score_ever") or entries[key].get("candidate_score", ""),
            "overall_media_score": media.get("overall_media_score", row.get("overall_media_score", entries[key].get("overall_media_score", ""))),
            "postcardness_score": media.get("postcardness_score", row.get("postcardness_score", entries[key].get("postcardness_score", ""))),
            "aesthetic_score": media.get("aesthetic_score", row.get("aesthetic_score", entries[key].get("aesthetic_score", ""))),
            "technical_quality_score": media.get("technical_quality_score", row.get("technical_quality_score", entries[key].get("technical_quality_score", ""))),
            "publication_safety_score": media.get("publication_safety_score", row.get("publication_safety_score", entries[key].get("publication_safety_score", ""))),
            "image_model_input_type": media.get("image_model_input_type", row.get("image_model_input_type", entries[key].get("image_model_input_type", ""))),
            "image_model_type": media.get("image_model_type", row.get("image_model_type", entries[key].get("image_model_type", ""))),
            "image_url_or_local_path": media.get("image_url_or_local_path", row.get("primary_media_path", "")),
            "image_queue_status": status,
            "status_color_hint": color,
            "row_fill_color": color,
            "media_acquisition_status": "actual_image_downloaded_and_scored" if actual else ("needs_actual_image_fetch" if status == "needs_actual_image_fetch" else "no_media_or_not_supported"),
            "media_acquisition_error_type": "" if actual else (media.get("failure_reason") or row.get("failure_reason") or status),
            "images_scored_actual_count": 1 if actual else 0,
            "last_attempt_run_id": run_id if media.get("post_url") else entries[key].get("last_attempt_run_id", ""),
            "last_attempt_at": run_now if media.get("post_url") else entries[key].get("last_attempt_at", ""),
            "next_action": "human_review_best_image" if status == "actual_scored" else ("wait_for_image_diagnostic" if status == "image_analysis_in_progress" else ("download_actual_image_bytes_next" if status == "needs_actual_image_fetch" else "skip_or_manual_open")),
            "queue_item_updated_at": run_now,
        })

    for row in new_posts:
        if row.get("has_media"):
            add_or_update(row, added_from="current_run_text_gate")
    for row in candidate_memory_rows:
        if row.get("post_url") and (row.get("current_lifecycle_status") in ACTIVE_CANDIDATE_MEMORY_STATUSES or row.get("image_status") == "needs_actual_image_fetch"):
            add_or_update(row, added_from="candidate_memory")
    for media in media_rows:
        if media.get("post_url"):
            add_or_update(media, added_from="media_scoring")

    processed_orders = [int(r.get("image_queue_order") or 0) for r in entries.values() if r.get("last_attempt_run_id") == run_id or r.get("image_queue_status") == "actual_scored"]
    cursor_position = max([prev_cursor] + processed_orders) if entries else 0
    target = getenv_int("REGION_TALK_IMAGE_QUEUE_TARGET_PER_RUN", 30)
    ordered = sorted(entries.values(), key=lambda r: (int(r.get("image_queue_order") or 0)))
    pending_after_cursor = [r for r in ordered if int(r.get("image_queue_order") or 0) > cursor_position and r.get("image_queue_status") != "actual_scored"]
    target_ids = {str(r.get("image_queue_id")) for r in pending_after_cursor[:target]}
    display = sorted(ordered, key=lambda r: (0 if r.get("image_queue_status") == "actual_scored" else 1, int(r.get("image_queue_order") or 0)))
    for idx, row in enumerate(display, start=1):
        order = int(row.get("image_queue_order") or 0)
        row["display_order"] = idx
        row["cursor_marker"] = "cursor_here" if order == cursor_position else ("next_after_cursor" if order > cursor_position and str(row.get("image_queue_id")) in target_ids else "")
        row["is_after_cursor"] = str(order > cursor_position).lower()
        row["selected_for_next_image_batch"] = str(str(row.get("image_queue_id")) in target_ids).lower()
        row["image_queue_batch_target"] = target
    top = [
        {"image_rank": i, **r}
        for i, r in enumerate(sorted([r for r in display if r.get("image_queue_status") == "actual_scored"], key=lambda r: (-float(r.get("overall_media_score") or 0), -float(r.get("postcardness_score") or 0), -float(r.get("aesthetic_score") or 0))), start=1)
    ]
    metrics = {
        "image_queue_total": len(display),
        "image_queue_cursor_position": cursor_position,
        "image_queue_target_this_run": target,
        "image_queue_selected_next_batch": len(target_ids),
        "image_queue_actual_scored_total": sum(1 for r in display if r.get("image_queue_status") == "actual_scored"),
        "image_queue_needs_actual_fetch_total": sum(1 for r in display if r.get("image_queue_status") == "needs_actual_image_fetch"),
        "image_queue_pruned_non_region_previous": image_queue_pruned_non_region_previous,
        "image_queue_rejected_non_region_inputs": image_queue_rejected_non_region_inputs,
        "image_queue_text_region_confirmed_total": sum(1 for r in display if image_queue_text_region_confirmed(r)),
    }
    return display, top, metrics


def compact_shortlist_row(r: dict[str, Any]) -> dict[str, Any]:
    return {
        "rank": r.get("human_shortlist_rank") or r.get("rank"),
        "decision_bucket": r.get("decision_bucket"),
        "next_action": r.get("next_action"),
        "post_url": r.get("post_url"),
        "source_title": r.get("source_title"),
        "platform": r.get("platform"),
        "post_date": r.get("post_date"),
        "short_summary": r.get("short_summary"),
        "why_relevant": r.get("why_this_is_about_kaliningrad"),
        "matched_place_names": r.get("matched_place_names"),
        "content_type": r.get("content_type") or r.get("llm_content_type") or r.get("vector_content_type"),
        "vector_gate_status": r.get("vector_gate_status"),
        "vector_negative_class": r.get("vector_negative_class"),
        "vector_positive_score": r.get("vector_positive_score"),
        "vector_negative_score": r.get("vector_negative_score"),
        "candidate_score": r.get("candidate_score"),
        "text_substance_score": r.get("text_substance_score"),
        "source_geo_class": r.get("source_geo_class"),
        "source_topic_class": r.get("source_topic_class"),
        "nonlocal_value_score": r.get("nonlocal_value_score"),
        "has_firsthand_visit_evidence": r.get("has_firsthand_visit_evidence"),
        "visit_evidence_type": r.get("visit_evidence_type"),
        "first_person_markers": r.get("first_person_markers"),
        "emotion_or_impression_evidence": r.get("emotion_or_impression_evidence"),
        "review_or_opinion_evidence": r.get("review_or_opinion_evidence"),
        "useful_route_evidence": r.get("useful_route_evidence"),
        "publication_story_score": r.get("publication_story_score"),
        "visit_impression_score": r.get("visit_impression_score"),
        "useful_route_score": r.get("useful_route_score"),
        "emotion_observation_score": r.get("emotion_observation_score"),
        "llm_gate_status": r.get("llm_gate_status"),
        "llm_model": r.get("llm_model"),
        "llm_decision": r.get("llm_decision"),
        "llm_reason_short": str(r.get("llm_reason") or "")[:260],
        "image_model_id": r.get("model_id"),
        "image_model_type": r.get("image_model_type"),
        "image_model_runtime": r.get("image_model_runtime"),
        "image_model_input_type": r.get("image_model_input_type"),
        "image_model_device": r.get("image_model_device"),
        "postcardness_score": r.get("postcardness_score"),
        "aesthetic_score": r.get("aesthetic_score"),
        "region_visual_relevance_score": r.get("region_visual_relevance_score"),
        "publication_safety_score": r.get("publication_safety_score"),
        "overall_media_score": r.get("overall_media_score"),
        "image_reviewable": r.get("image_reviewable"),
        "image_publication_ready": r.get("image_publication_ready"),
        "image_model_explanation": r.get("model_short_explanation"),
        "risk_flags": r.get("risk_flags"),
    }


def _rt_float(value: Any, default: float = 0.0) -> float:
    try:
        if value in ("", None):
            return default
        return float(value)
    except Exception:
        return default


def _rt_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "on", "accept", "accepted"}


def publication_goal_defaults(previous_state: dict[str, Any], *, run_now: str) -> dict[str, Any]:
    prev = previous_state.get("publication_goal") if isinstance(previous_state.get("publication_goal"), dict) else {}
    goal_id = (os.getenv("REGION_TALK_PUBLICATION_GOAL_ID") or str(prev.get("publication_goal_id") or "region-talk-product-goal-v1")).strip()
    target = getenv_int("REGION_TALK_PUBLICATION_GOAL_TARGET", int(prev.get("target_confirmed") or 20))
    budget = getenv_int("REGION_TALK_PUBLICATION_LLM_BUDGET_MAX", int(prev.get("llm_budget_max") or 100))
    return {
        **prev,
        "publication_goal_id": goal_id,
        "target_confirmed": target,
        "llm_budget_max": budget,
        "started_at": prev.get("started_at") or run_now,
        "updated_at": run_now,
    }


def _publication_candidate_base_ok(row: dict[str, Any]) -> tuple[bool, str]:
    if not str(row.get("post_url") or "").strip():
        return False, "missing_post_url"
    if str(row.get("source_geo_class") or "") == "kaliningrad_local":
        return False, "local_kaliningrad_source_for_separate_monitoring"
    topic = str(row.get("source_topic_class") or "")
    if topic in {"local_news", "federal_media", "ads_tours"}:
        return False, "source_topic_not_nonlocal_blogger_travel"
    if not _rt_bool(row.get("kaliningrad_oblast_only_scope")):
        return False, "not_confirmed_kaliningrad_oblast_scope"
    if str(row.get("kaliningrad_mention_role") or "main_subject") not in {"", "main_subject", "unclear"}:
        return False, "kaliningrad_not_main_subject"
    if _rt_bool(row.get("is_ad_or_promo")):
        return False, "ad_or_promo"
    if str(row.get("vector_gate_status") or "").startswith("vector_reject"):
        return False, str(row.get("vector_gate_status") or "vector_reject")
    if str(row.get("image_model_input_type") or "") != "actual_image":
        return False, "actual_image_required"
    if str(row.get("image_queue_status") or "") not in {"", "actual_scored"}:
        return False, "image_queue_not_actual_scored"
    if _rt_float(row.get("overall_media_score")) < _rt_float(os.getenv("REGION_TALK_PUBLICATION_MIN_OVERALL_MEDIA_SCORE"), 0.66):
        return False, "overall_media_score_below_threshold"
    if _rt_float(row.get("postcardness_score")) < _rt_float(os.getenv("REGION_TALK_PUBLICATION_MIN_POSTCARDNESS_SCORE"), 0.55):
        return False, "postcardness_score_below_threshold"
    return True, ""


def _publication_text_story_score(row: dict[str, Any]) -> float:
    bool_bonus = 0.0
    for key, weight in [
        ("has_firsthand_visit_evidence", 0.22),
        ("emotion_or_impression_evidence", 0.18),
        ("review_or_opinion_evidence", 0.16),
        ("memorable_detail_evidence", 0.16),
        ("useful_route_evidence", 0.12),
    ]:
        if _rt_bool(row.get(key)):
            bool_bonus += weight
    numeric = max(
        _rt_float(row.get("publication_story_score")),
        _rt_float(row.get("nonlocal_blogger_visit_score")),
        _rt_float(row.get("vector_visit_impression_score")),
        _rt_float(row.get("vector_route_useful_score")),
        _rt_float(row.get("vector_emotion_observation_score")),
    )
    return round(min(1.0, max(numeric, bool_bonus, _rt_float(row.get("candidate_score_current")), _rt_float(row.get("candidate_score")))), 3)


def _publication_diversity_penalty(row: dict[str, Any], accepted_rows: list[dict[str, Any]]) -> float:
    """Small anti-overlap penalty; vector-nearest-neighbour diversity can replace this after first accepted set exists."""
    if not accepted_rows:
        return 0.0
    source = str(row.get("source_id") or row.get("source_title") or "")
    places = {p.strip().lower() for p in str(row.get("matched_place_names") or "").replace(",", ";").split(";") if p.strip()}
    ctype = str(row.get("content_type") or row.get("vector_content_type") or "")
    penalty = 0.0
    for prev in accepted_rows:
        if source and source == str(prev.get("source_id") or prev.get("source_title") or ""):
            penalty = max(penalty, 0.18)
        prev_places = {p.strip().lower() for p in str(prev.get("matched_place_names") or "").replace(",", ";").split(";") if p.strip()}
        if places and prev_places and places & prev_places:
            penalty = max(penalty, 0.12)
        if ctype and ctype == str(prev.get("content_type") or prev.get("vector_content_type") or ""):
            penalty = max(penalty, 0.05)
    return round(penalty, 3)


def _merge_candidate_and_image_rows(candidate_rows: list[dict[str, Any]], image_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_url = {str(r.get("post_url") or ""): r for r in image_rows if str(r.get("post_url") or "")}
    merged = []
    for row in candidate_rows:
        img = by_url.get(str(row.get("post_url") or "")) or {}
        merged.append({**row, **{k: v for k, v in img.items() if v not in ("", None)}})
    return merged


def build_publication_candidate_queue(
    candidate_rows: list[dict[str, Any]],
    image_rows: list[dict[str, Any]],
    previous_publication_rows: list[dict[str, Any]],
    previous_goal: dict[str, Any],
    *,
    run_id: str,
    run_now: str,
    llm_model: str,
    llm_default_env_var_name: str,
    current_run_preverified_llm_calls: int = 0,
    report_event: Any | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    goal = publication_goal_defaults({"publication_goal": previous_goal}, run_now=run_now)
    goal_id = str(goal.get("publication_goal_id") or "region-talk-product-goal-v1")
    budget_max = int(goal.get("llm_budget_max") or 100)
    target = int(goal.get("target_confirmed") or 20)
    previous_by_url = {str(r.get("post_url") or ""): dict(r) for r in previous_publication_rows if str(r.get("post_url") or "")}
    confirmed_urls = {u for u, r in previous_by_url.items() if str(r.get("publication_candidate_status") or "") in PUBLICATION_CONFIRMED_STATUSES}
    sent_urls = {u for u, r in previous_by_url.items() if _rt_bool(r.get("sent_to_chat"))}
    previous_llm_used = int(goal.get("llm_calls_used_total") or 0)
    preverified_calls = max(0, int(current_run_preverified_llm_calls or 0))
    calls_this_run = 0
    per_run_cap = getenv_int("REGION_TALK_PUBLICATION_MAX_LLM_VERIFY_PER_RUN", 20)
    rows: list[dict[str, Any]] = []
    accepted_for_diversity: list[dict[str, Any]] = [previous_by_url[u] for u in confirmed_urls if u in previous_by_url]
    pool = _merge_candidate_and_image_rows(candidate_rows, image_rows)
    pre_ranked: list[dict[str, Any]] = []
    for row in pool:
        ok, reason = _publication_candidate_base_ok(row)
        visual = max(_rt_float(row.get("overall_media_score")), _rt_float(row.get("final_visual_score")))
        postcard = _rt_float(row.get("postcardness_score") or row.get("clip_postcardness_score") or row.get("cv_postcardness_score"))
        text_story = _publication_text_story_score(row)
        nonlocal_score = _rt_float(row.get("nonlocal_value_score"))
        base_score = round(min(1.0, 0.42 * visual + 0.18 * postcard + 0.30 * text_story + 0.10 * nonlocal_score), 3)
        pre_ranked.append({**row, "_publication_base_ok": ok, "_publication_base_reject_reason": reason, "_publication_base_score": base_score, "_publication_visual_score": visual, "_publication_text_story_score": text_story})
    pre_ranked.sort(key=lambda r: (-_rt_float(r.get("_publication_base_score")), -_rt_float(r.get("overall_media_score")), -_rt_float(r.get("postcardness_score")), str(r.get("post_date") or "")))
    for row in pre_ranked:
        post_url = str(row.get("post_url") or "")
        prev = previous_by_url.get(post_url, {})
        candidate_id = "pubcand_" + stable_hash(goal_id, post_url or row.get("candidate_memory_id") or row.get("post_id"))
        diversity_penalty = _publication_diversity_penalty(row, accepted_for_diversity)
        publication_score = round(max(0.0, _rt_float(row.get("_publication_base_score")) - diversity_penalty), 3)
        base_ok = bool(row.get("_publication_base_ok"))
        status = str(prev.get("publication_candidate_status") or ("ready_for_llm" if base_ok else "filtered_before_llm"))
        llm_status = str(prev.get("publication_llm_status") or "")
        llm_decision = str(prev.get("publication_llm_decision") or "")
        llm_reason = str(prev.get("publication_llm_reason") or "")
        llm_model_used = str(prev.get("publication_llm_model") or "")
        if base_ok and post_url not in confirmed_urls:
            final_status = str(row.get("final_verifier_status") or row.get("llm_gate_status") or "")
            final_decision = str(row.get("final_verifier_decision") or row.get("llm_decision") or "")
            if final_status == "ok" and final_decision == "accept":
                llm_status, llm_decision = "confirmed_from_final_verifier", "accept"
                llm_reason = str(row.get("final_verifier_reason") or row.get("llm_reason") or "Gemini final verifier accepted text criteria")
                llm_model_used = str(row.get("final_verifier_model") or row.get("llm_model") or llm_model)
                status = "llm_confirmed"
            elif previous_llm_used + preverified_calls + calls_this_run < budget_max and calls_this_run < per_run_cap:
                evidence = {
                    "stage": "publication_queue_final_verifier",
                    "visual_score": row.get("_publication_visual_score"),
                    "overall_media_score": row.get("overall_media_score"),
                    "postcardness_score": row.get("postcardness_score"),
                    "aesthetic_score": row.get("aesthetic_score"),
                    "image_model_input_type": row.get("image_model_input_type"),
                    "image_model_type": row.get("image_model_type"),
                    "visual_confirmation_source": "RegionTalkImageDiagnostic actual-image scoring",
                    "vector_gate_status": row.get("vector_gate_status"),
                    "source_geo_class": row.get("source_geo_class"),
                    "source_topic_class": row.get("source_topic_class"),
                    "publication_text_story_score": row.get("_publication_text_story_score"),
                }
                result = call_region_talk_semantic_llm(row, evidence, model=llm_model, default_env_var_name=llm_default_env_var_name)
                if result.get("llm_gate_status") in {"ok", "error", "rate_limited"}:
                    calls_this_run += 1
                llm_status = str(result.get("llm_gate_status") or "")
                llm_decision = str(result.get("llm_decision") or "")
                llm_reason = str(result.get("llm_reason") or "")
                llm_model_used = str(result.get("llm_model") or llm_model)
                status = "llm_confirmed" if llm_status == "ok" and llm_decision == "accept" else ("llm_rejected" if llm_status == "ok" and llm_decision == "reject" else "llm_needs_review")
            else:
                status = "llm_budget_deferred"
                llm_status = "budget_deferred"
                llm_reason = f"Goal LLM budget/cap reached: used={previous_llm_used + preverified_calls + calls_this_run}/{budget_max}, per_run={calls_this_run}/{per_run_cap}, preverified={preverified_calls}"
        elif not base_ok and status not in PUBLICATION_CONFIRMED_STATUSES:
            status = "filtered_before_llm"
        if status in PUBLICATION_CONFIRMED_STATUSES:
            accepted_for_diversity.append(row)
        is_confirmed = status in PUBLICATION_CONFIRMED_STATUSES
        why_selected = (
            f"визуал={_rt_float(row.get('_publication_visual_score')):.2f}, открытка={_rt_float(row.get('postcardness_score')):.2f}, "
            f"текст/эмоция={_rt_float(row.get('_publication_text_story_score')):.2f}, источник={row.get('source_topic_class') or 'unknown'}, "
            f"diversity_penalty={diversity_penalty:.2f}"
        )
        rows.append({
            "publication_candidate_id": candidate_id,
            "publication_goal_id": goal_id,
            "publication_rank": 0,
            "publication_candidate_status": status,
            "post_id": row.get("post_id"),
            "post_url": post_url,
            "source_id": row.get("source_id"),
            "source_title": row.get("source_title"),
            "source_geo_class": row.get("source_geo_class"),
            "source_topic_class": row.get("source_topic_class"),
            "post_date": row.get("post_date"),
            "short_summary": row.get("short_summary"),
            "matched_place_names": row.get("matched_place_names"),
            "candidate_score": row.get("candidate_score") or row.get("candidate_score_current") or row.get("best_candidate_score_ever"),
            "publication_score": publication_score,
            "visual_score": round(_rt_float(row.get("_publication_visual_score")), 3),
            "text_story_score": row.get("_publication_text_story_score"),
            "diversity_penalty": diversity_penalty,
            "overall_media_score": row.get("overall_media_score") or row.get("final_visual_score"),
            "postcardness_score": row.get("postcardness_score"),
            "aesthetic_score": row.get("aesthetic_score"),
            "image_model_input_type": row.get("image_model_input_type"),
            "image_queue_status": row.get("image_queue_status"),
            "vector_gate_status": row.get("vector_gate_status"),
            "vector_content_type": row.get("vector_content_type"),
            "vector_positive_score": row.get("vector_positive_score"),
            "publication_llm_status": llm_status,
            "publication_llm_decision": llm_decision,
            "publication_llm_reason": llm_reason[:500],
            "publication_llm_model": llm_model_used,
            "visual_confirmation_source": "RegionTalkImageDiagnostic actual-image scoring" if base_ok else "",
            "why_selected": why_selected if is_confirmed else "",
            "why_not_selected": "" if is_confirmed else (row.get("_publication_base_reject_reason") or llm_reason),
            "goal_stop_candidate": "false",
            "sent_to_chat": str(post_url in sent_urls).lower(),
            "sent_message_id": prev.get("sent_message_id", ""),
            "created_at": prev.get("created_at") or run_now,
            "last_seen_run_id": run_id,
            "last_confirmed_run_id": run_id if is_confirmed else prev.get("last_confirmed_run_id", ""),
        })
    rows.sort(key=lambda r: (
        0 if str(r.get("publication_candidate_status")) in PUBLICATION_CONFIRMED_STATUSES else 1 if str(r.get("publication_candidate_status")) in {"ready_for_llm", "llm_needs_review"} else 2,
        -_rt_float(r.get("publication_score")),
        -_rt_float(r.get("visual_score")),
        str(r.get("post_date") or ""),
    ))
    for i, row in enumerate(rows, start=1):
        row["publication_rank"] = i
        if i <= target and str(row.get("publication_candidate_status") or "") in PUBLICATION_CONFIRMED_STATUSES:
            row["goal_stop_candidate"] = "true"
    confirmed_count = sum(1 for r in rows if str(r.get("publication_candidate_status")) in PUBLICATION_CONFIRMED_STATUSES)
    total_llm_calls = previous_llm_used + preverified_calls + calls_this_run
    goal.update({
        "updated_at": run_now,
        "last_run_id": run_id,
        "confirmed_count": confirmed_count,
        "sent_count": sum(1 for r in rows if _rt_bool(r.get("sent_to_chat"))),
        "llm_calls_used_total": total_llm_calls,
        "llm_calls_used_this_run": preverified_calls + calls_this_run,
        "llm_calls_used_publication_queue_this_run": calls_this_run,
        "llm_calls_used_preverified_this_run": preverified_calls,
        "llm_budget_remaining": max(0, budget_max - total_llm_calls),
        "goal_status": "complete" if confirmed_count >= target else ("llm_budget_exhausted" if total_llm_calls >= budget_max else "running"),
        "top_confirmed_post_urls": [r.get("post_url") for r in rows if str(r.get("publication_candidate_status")) in PUBLICATION_CONFIRMED_STATUSES][:target],
    })
    if callable(report_event):
        try:
            report_event("publication_queue_built", phase="publication_queue", status=goal.get("goal_status"), publication_goal_id=goal_id, confirmed_count=confirmed_count, target_confirmed=target, llm_calls_used_total=goal.get("llm_calls_used_total"), llm_budget_max=budget_max)
        except Exception:
            pass
    return rows, goal


def candidate_memory_shortlist_row(r: dict[str, Any], rank: int) -> dict[str, Any]:
    input_type = str(r.get("image_model_input_type") or "")
    stage = str(r.get("current_stage") or "")
    if str(r.get("image_publication_ready") or "") == "true" and input_type == "actual_image":
        bucket = "publication_ready_actual_image"
        action = "human_final_check"
    elif stage == "image_fetch_retry_needed" or r.get("image_status") == "needs_actual_image_fetch":
        bucket = "needs_actual_image_retry"
        action = "retry_media_download_or_manual_open"
    elif stage == "good_text_weak_media":
        bucket = "good_text_weak_actual_image" if input_type == "actual_image" else "manual_open_promising_text"
        action = "manual_open_if_text_is_strong"
    else:
        bucket = "manual_open_promising_text"
        action = r.get("next_action") or "human_review_without_llm"
    return {
        "rank": rank,
        "decision_bucket": bucket,
        "next_action": action,
        "post_url": r.get("post_url"),
        "source_title": r.get("source_title"),
        "platform": r.get("platform"),
        "post_date": r.get("post_date"),
        "short_summary": r.get("short_summary"),
        "why_relevant": r.get("why_keep_in_memory"),
        "matched_place_names": r.get("matched_place_names"),
        "content_type": r.get("content_type") or r.get("llm_content_type") or r.get("vector_content_type"),
        "vector_gate_status": r.get("vector_gate_status"),
        "memory_vector_recheck_status": r.get("memory_vector_recheck_status"),
        "vector_negative_class": r.get("vector_negative_class") or r.get("memory_vector_recheck_negative_class"),
        "vector_positive_score": r.get("vector_positive_score"),
        "memory_product_exclusion_reason": r.get("memory_product_exclusion_reason"),
        "source_geo_class": r.get("source_geo_class"),
        "source_topic_class": r.get("source_topic_class"),
        "nonlocal_value_score": r.get("nonlocal_value_score"),
        "has_firsthand_visit_evidence": r.get("has_firsthand_visit_evidence"),
        "visit_evidence_type": r.get("visit_evidence_type"),
        "first_person_markers": r.get("first_person_markers"),
        "emotion_or_impression_evidence": r.get("emotion_or_impression_evidence"),
        "review_or_opinion_evidence": r.get("review_or_opinion_evidence"),
        "useful_route_evidence": r.get("useful_route_evidence"),
        "publication_story_score": r.get("publication_story_score"),
        "candidate_score": r.get("candidate_score_current") or r.get("best_candidate_score_ever"),
        "overall_media_score": r.get("media_score_current") or r.get("best_media_score_ever"),
        "image_status": r.get("image_status"),
        "image_model_input_type": r.get("image_model_input_type"),
        "image_reviewable": r.get("image_reviewable"),
        "image_publication_ready": r.get("image_publication_ready"),
        "llm_status": r.get("llm_status") or "not_called_or_previous",
        "current_lifecycle_status": r.get("current_lifecycle_status"),
        "not_refetched_this_run": r.get("not_refetched_this_run"),
    }


async def discover_telegram_similar_channels(client: Any, source_rows: list[dict[str, Any]], entity_by_source: dict[str, Any], governor: TelegramRequestGovernor, run_id: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not getenv_bool("REGION_TALK_TG_SIMILAR_ENABLED", getenv_bool("REGION_TALK_DISCOVERY_ENABLE_TELEGRAM_SIMILAR", True)):
        _REGION_TALK_TELEGRAM_RUNTIME.update({"telegram_similar_channels_status": "disabled"})
        return [], []
    max_seed_channels = getenv_int("REGION_TALK_MAX_SIMILAR_SEEDS_PER_RUN", getenv_int("REGION_TALK_TG_SIMILAR_MAX_SEED_CHANNELS_PER_RUN", getenv_int("REGION_TALK_DISCOVERY_MAX_SIMILAR_SEED_CHANNELS", 100)))
    max_per_seed = getenv_int("REGION_TALK_TG_SIMILAR_MAX_RECOMMENDATIONS_PER_SEED", getenv_int("REGION_TALK_DISCOVERY_MAX_SIMILAR_PER_SEED", 10))
    max_frontier = getenv_int("REGION_TALK_MAX_NEW_FRONTIER_PER_RUN", getenv_int("REGION_TALK_TG_SIMILAR_MAX_NEW_FRONTIER_PER_RUN", getenv_int("REGION_TALK_DISCOVERY_MAX_NEW_FRONTIER_PER_RUN", 1000)))
    rows: list[dict[str, Any]] = []
    edges: list[dict[str, Any]] = []
    try:
        from telethon.tl import functions  # type: ignore
    except Exception as exc:
        _REGION_TALK_TELEGRAM_RUNTIME.update({"telegram_similar_channels_status": "not_supported_by_telethon_version", "telegram_similar_channels_error": type(exc).__name__})
        return rows, edges
    request_cls = getattr(getattr(functions, "channels", None), "GetChannelRecommendationsRequest", None)
    if request_cls is None:
        _REGION_TALK_TELEGRAM_RUNTIME.update({"telegram_similar_channels_status": "not_supported_by_telethon_version"})
        return rows, edges
    seeds = [r for r in source_rows if (r.get("fetch_status") == "ok" or str(r.get("fetch_status") or "").startswith("profile_resolved")) and r.get("source_id") in entity_by_source]
    seed_pool = sorted(seeds, key=lambda r: (
        bool(str(r.get("similar_next_allowed_at") or "") and str(r.get("similar_next_allowed_at")) > utc_now_iso()),
        str(r.get("similar_last_scanned_at") or ""),
        -float(r.get("monitor_priority_score") or 0),
        int(r.get("priority") or 999),
        str(r.get("canonical_url") or ""),
    ))[:max(max_seed_channels * 4, max_seed_channels)]
    seeds = seed_pool[:max_seed_channels]
    raw_count = 0
    errors = 0
    floodwait = 0
    seen: set[str] = set()
    self_loop_count = 0
    duplicate_count = 0
    seed_updates: dict[str, dict[str, Any]] = {}
    for seed_idx, srow in enumerate(seeds, start=1):
        if governor.recommendation_calls_attempted >= governor.max_recommendation_calls or not governor.has_total_request_budget("channels.getChannelRecommendations", str(srow.get("source_id") or ""), str(srow.get("canonical_url") or "")):
            break
        source_id = str(srow.get("source_id") or "")
        canonical = str(srow.get("canonical_url") or "")
        similar_seed_id = "similar_seed_" + stable_hash(canonical)
        entity = entity_by_source.get(source_id)
        governor.recommendation_calls_attempted += 1
        governor.total_attempted += 1
        governor.requests_by_method["channels.getChannelRecommendations"] = governor.requests_by_method.get("channels.getChannelRecommendations", 0) + 1
        try:
            result = await client(request_cls(channel=entity))
            governor.total_ok += 1
            governor.recommendation_calls_ok += 1
            chats = list(getattr(result, "chats", None) or [])
            raw_count += len(chats)
            unique_for_seed = 0
            governor.recommendation_channels_returned += len(chats)
            for rank, ch in enumerate(chats[:max_per_seed], start=1):
                username = str(getattr(ch, "username", None) or "").strip()
                title = str(getattr(ch, "title", None) or "").strip()
                if not username:
                    continue
                rec_url = "https://t.me/" + username
                cand_id = "src_cand_" + stable_hash("telegram", rec_url)
                if canonical_source_key("telegram", username, rec_url) == canonical_source_key("telegram", srow.get("handle", ""), canonical):
                    self_loop_count += 1
                    debug_row = {"run_id": run_id, "seed_source_id": source_id, "seed_channel_url": canonical, "recommended_canonical_url": rec_url, "method_status": "debug_self_loop_rejected", "edge_type": "telegram_similar_channel", "frontier_action": "none", "frontier_reason": "self-loop rejected"}
                    rows.append(debug_row)
                    write_region_talk_online_source_item(debug_row, run_id=run_id, stage="telegram_similar_discovery", status="debug_self_loop_rejected")
                    continue
                if cand_id in seen:
                    duplicate_count += 1
                    continue
                seen.add(cand_id)
                unique_for_seed += 1
                confidence = round(max(0.35, 0.85 - 0.025 * (rank - 1)), 3)
                row = {
                    "run_id": run_id,
                    "source_candidate_id": cand_id,
                    "discovered_from_source": srow.get("source_title") or srow.get("resolved_title"),
                    "discovered_from_post_url": "",
                    "discovery_type": "telegram_similar_channels",
                    "edge_type": "telegram_similar_channel",
                    "raw_url": rec_url,
                    "normalized_url": rec_url,
                    "platform_guess": "telegram",
                    "candidate_source_status": "source_frontier",
                    "frontier_status": "queued_unresolved",
                    "confidence": confidence,
                    "recommendation_rank": rank,
                    "recommendation_count_for_seed": len(chats),
                    "recommendation_source_channel_url": canonical,
                    "recommendation_source_channel_title": srow.get("source_title") or srow.get("resolved_title"),
                    "recommended_title": title,
                    "recommended_username": username,
                    "recommended_canonical_url": rec_url,
                    "recommended_is_channel": str(bool(getattr(ch, "broadcast", False))).lower(),
                    "recommended_is_megagroup": str(bool(getattr(ch, "megagroup", False))).lower(),
                    "recommended_verified": str(bool(getattr(ch, "verified", False))).lower(),
                    "similarity_seed_source_id": source_id,
                    "similarity_edge_confidence": confidence,
                    "frontier_action": "queued_resolve_later",
                    "frontier_reason": "Telegram channels.getChannelRecommendations; stored as frontier only, no auto-join/no auto-monitor",
                    "private_state_key": "telegram:username:" + username.lower(),
                }
                row["frontier_priority"] = source_candidate_score(row)
                rows.append(row)
                write_region_talk_online_source_item(row, run_id=run_id, stage="telegram_similar_discovery", status="source_frontier")
                edge_id = "edge_" + stable_hash(source_id, cand_id, "telegram_similar_channel")
                edges.append({
                    "edge_id": edge_id, "from_source_id": source_id, "from_source_title": srow.get("source_title") or srow.get("resolved_title"),
                    "from_post_id": "", "to_source_candidate_id": cand_id, "to_source_title": title or username,
                    "edge_type": "telegram_similar_channel", "evidence_url": rec_url,
                    "evidence_context_short": f"Similar channel recommendation rank {rank} from {srow.get('source_title')}",
                    "confidence": confidence, "discovery_depth": 1, "run_id": run_id,
                })
                governor.entity_cache["telegram:username:" + username.lower()] = {
                    "canonical_url": rec_url, "username": username, "title_last_seen": title, "entity_kind": "channel",
                    "channel_id_private": str(getattr(ch, "id", "") or ""),
                    "access_hash_private": str(getattr(ch, "access_hash", "") or ""),
                    "resolved_at": utc_now_iso(), "last_used_at": "", "last_success_at": "", "source": "telegram_similar_channels",
                }
                if len(rows) >= max_frontier:
                    break
            seed_updates[similar_seed_id] = {
                "similar_seed_id": similar_seed_id,
                "canonical_url": canonical,
                "similar_last_scanned_at": utc_now_iso(),
                "similar_last_used_at": utc_now_iso(),
                "similar_use_count_increment": 1,
                "similar_last_result_count": len(chats),
                "similar_last_unique_count": unique_for_seed,
                "similar_next_allowed_at": iso_after_seconds(getenv_int("REGION_TALK_SIMILAR_SEED_TTL_SECONDS", 7 * 86400)),
                "similar_error_count_increment": 0,
            }
            if len(rows) >= max_frontier:
                break
        except Exception as exc:
            errors += 1
            seconds = int(getattr(exc, "seconds", 0) or 0)
            if type(exc).__name__ == "FloodWaitError" or seconds:
                floodwait += 1
                governor.mark_floodwait("channels.getChannelRecommendations", source_id, canonical, seconds)
                break
            governor.total_error += 1
            seed_updates[similar_seed_id] = {
                "similar_seed_id": similar_seed_id,
                "canonical_url": canonical,
                "similar_last_scanned_at": utc_now_iso(),
                "similar_last_used_at": utc_now_iso(),
                "similar_use_count_increment": 1,
                "similar_last_result_count": 0,
                "similar_last_unique_count": 0,
                "similar_next_allowed_at": iso_after_seconds(getenv_int("REGION_TALK_SIMILAR_ERROR_BACKOFF_SECONDS", 86400)),
                "similar_error_count_increment": 1,
            }
            error_row = {
                "run_id": run_id, "seed_source_id": source_id, "seed_channel_url": canonical,
                "seed_channel_title": srow.get("source_title"), "method_status": "error",
                "method_error_code": type(exc).__name__, "method_error_message_short": str(exc)[:180],
                "edge_type": "telegram_similar_channel", "frontier_action": "none", "frontier_reason": "recommendation call failed",
            }
            rows.append(error_row)
            write_region_talk_online_source_item(error_row, run_id=run_id, stage="telegram_similar_discovery", status="error")
    governor.recommendation_channels_added_to_frontier = len([r for r in rows if r.get("candidate_source_status") == "source_frontier"])
    status = "ok" if rows or not errors else "error"
    _REGION_TALK_TELEGRAM_RUNTIME.update({
        "telegram_similar_channels_status": status,
        "telegram_similar_channels_seed_count": len(seeds),
        "telegram_similar_channels_raw_count": raw_count,
        "telegram_similar_channels_unique_count": len({r.get("source_candidate_id") for r in rows if r.get("source_candidate_id")}),
        "telegram_similar_channels_added_to_frontier": governor.recommendation_channels_added_to_frontier,
        "telegram_similar_channels_fetch_errors": errors,
        "telegram_similar_channels_floodwait_count": floodwait,
        "telegram_similar_self_loop_count": self_loop_count,
        "telegram_similar_duplicate_canonical_count": duplicate_count,
        "similar_seed_cursor_advanced": str(bool(seed_updates)).lower(),
        "similar_seed_updates": seed_updates,
    })
    return rows, edges


async def discover_telegram_keyword_sources(client: Any, governor: TelegramRequestGovernor, run_id: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    mode = (os.getenv("REGION_TALK_DISCOVERY_MODE") or "mixed").strip().lower()
    if mode in {"off", "similar_only"} or not getenv_bool("REGION_TALK_ENABLE_TELEGRAM_KEYWORD_DISCOVERY", mode in {"mixed", "keyword_only"}):
        _REGION_TALK_TELEGRAM_RUNTIME["telegram_keyword_discovery_status"] = "disabled"
        return [], []
    terms_raw = os.getenv("REGION_TALK_TELEGRAM_KEYWORD_QUERIES") or ""
    terms = [x.strip() for x in re.split(r"[;\n|]+", terms_raw) if x.strip()] or TELEGRAM_KEYWORD_DISCOVERY_TERMS
    max_queries = getenv_int("REGION_TALK_MAX_TELEGRAM_KEYWORD_QUERIES", 30)
    per_query = getenv_int("REGION_TALK_TELEGRAM_KEYWORD_RESULTS_PER_QUERY", 20)
    max_frontier = getenv_int("REGION_TALK_MAX_KEYWORD_DISCOVERED_SOURCES_PER_RUN", 300)
    rows: list[dict[str, Any]] = []
    edges: list[dict[str, Any]] = []
    post_hits: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    processed = 0
    errors = 0
    raw_hits = 0
    try:
        for query in terms[:max(0, max_queries)]:
            if not governor.has_total_request_budget("messages.searchGlobal", "keyword_discovery", query):
                break
            processed += 1
            governor.total_attempted += 1
            governor.requests_by_method["messages.searchGlobal"] = governor.requests_by_method.get("messages.searchGlobal", 0) + 1
            try:
                async for msg in client.iter_messages(None, search=query, limit=per_query):
                    raw_hits += 1
                    chat = None
                    try:
                        chat = await msg.get_chat()
                    except Exception:
                        chat = getattr(msg, "chat", None)
                    username = str(getattr(chat, "username", None) or "").strip()
                    title = str(getattr(chat, "title", None) or getattr(chat, "first_name", None) or "").strip()
                    if not username:
                        continue
                    url = "https://t.me/" + username
                    key = canonical_source_key("telegram", username, url)
                    post_url = f"https://t.me/{username}/{getattr(msg, 'id', '')}"
                    text_excerpt = re.sub(r"\s+", " ", str(getattr(msg, "message", "") or ""))[:500]
                    post_hit_row = {
                        "run_id": run_id,
                        "matched_query": query,
                        "keyword_hit_post_url": post_url,
                        "keyword_hit_source_url": url,
                        "keyword_hit_text_excerpt": text_excerpt,
                        "source_candidate_id": "src_cand_" + stable_hash("telegram", url),
                        "source_title": title,
                        "username_or_handle": username,
                        "canonical_source_key": key,
                        "frontier_action": "source_seen" if key in seen_keys else "source_queued",
                        "sent_to_pipeline": "true",
                        "acceptance_note": "keyword hit is discovery/context only, never auto-accepts publication candidate",
                    }
                    post_hits.append(post_hit_row)
                    write_region_talk_online_source_item(post_hit_row, run_id=run_id, stage="keyword_hit_source_context", status=post_hit_row.get("frontier_action"))
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)
                    cand_id = "src_cand_" + stable_hash("telegram", url)
                    row = {
                        "run_id": run_id,
                        "source_candidate_id": cand_id,
                        "discovered_from_source": "telegram_keyword_search",
                        "discovered_from_post_url": post_url,
                        "keyword_hit_post_url": post_url,
                        "keyword_hit_text_excerpt": text_excerpt,
                        "discovery_type": "telegram_keyword_search",
                        "edge_type": "telegram_keyword_search",
                        "matched_query": query,
                        "raw_url": url,
                        "normalized_url": url,
                        "platform_guess": "telegram",
                        "canonical_source_key": key,
                        "candidate_source_status": "source_frontier",
                        "frontier_status": "queued_unresolved",
                        "confidence": 0.62,
                        "recommended_title": title,
                        "recommended_username": username,
                        "recommended_canonical_url": url,
                        "frontier_action": "queued_resolve_later",
                        "frontier_reason": "Telegram keyword search by Kaliningrad toponym; source candidate only, post is not accepted here",
                        "keyword_evidence_excerpt": text_excerpt,
                        "private_state_key": "telegram:username:" + username.lower(),
                    }
                    row["frontier_priority"] = source_candidate_score(row)
                    rows.append(row)
                    write_region_talk_online_source_item(row, run_id=run_id, stage="telegram_keyword_discovery", status="source_frontier")
                    edges.append({
                        "edge_id": "edge_" + stable_hash("keyword", query, cand_id),
                        "from_source_id": "telegram_keyword_search",
                        "from_source_title": query,
                        "from_post_id": "",
                        "to_source_candidate_id": cand_id,
                        "to_source_title": title or username,
                        "edge_type": "telegram_keyword_search",
                        "evidence_url": row["discovered_from_post_url"],
                        "evidence_context_short": f"Keyword '{query}' matched a public Telegram result; source only",
                        "confidence": row["confidence"],
                        "discovery_depth": 0,
                        "run_id": run_id,
                    })
                    if len(rows) >= max_frontier:
                        break
                governor.total_ok += 1
            except Exception as exc:
                errors += 1
                seconds = int(getattr(exc, "seconds", 0) or 0)
                if type(exc).__name__ == "FloodWaitError" or seconds:
                    governor.mark_floodwait("messages.searchGlobal", "keyword_discovery", query, seconds)
                    break
                governor.total_error += 1
                error_row = {"run_id": run_id, "discovery_type": "telegram_keyword_search", "edge_type": "telegram_keyword_search", "matched_query": query, "method_status": "error", "method_error_code": type(exc).__name__, "method_error_message_short": str(exc)[:180], "frontier_action": "none"}
                rows.append(error_row)
                write_region_talk_online_source_item(error_row, run_id=run_id, stage="telegram_keyword_discovery", status="error")
            if len(rows) >= max_frontier:
                break
    finally:
        _REGION_TALK_TELEGRAM_RUNTIME.update({
            "telegram_keyword_discovery_status": "ok" if rows or not errors else "error",
            "keyword_search_queries_processed": processed,
            "keyword_discovered_sources_unique": len({r.get("source_candidate_id") for r in rows if r.get("source_candidate_id")}),
            "keyword_post_hits_raw": raw_hits,
            "keyword_unique_channels": len({r.get("canonical_source_key") for r in post_hits if r.get("canonical_source_key")}),
            "keyword_hit_posts_sent_to_pipeline": len(post_hits),
            "keyword_post_hit_rows": post_hits,
            "keyword_discovery_errors": errors,
        })
    return rows, edges


def score_text(text: str) -> dict[str, Any]:
    low = (text or "").lower()
    anchor_hits = [a for a in DEFAULT_ANCHORS if a.lower() in low]
    positive_hits = [w for w in POSITIVE_WORDS if w in low]
    news_hits = [w for w in NEWS_WORDS if w in low]
    ad_hits = [w for w in AD_WORDS if w in low]
    trash_hits = [w for w in TRASH_WORDS if w in low]
    region = min(1.0, 0.25 * len(anchor_hits) + (0.25 if len(text) > 240 and anchor_hits else 0))
    news = min(1.0, 0.28 * len(news_hits))
    ad = min(1.0, 0.20 * len(ad_hits))
    trash = min(1.0, 0.35 * len(trash_hits))
    value = min(1.0, 0.25 + 0.08 * len(positive_hits) + min(0.35, len(text) / 1200)) if anchor_hits else 0.0
    tone = min(1.0, 0.35 + 0.12 * len(positive_hits)) if anchor_hits else 0.0
    return {"anchor_hits": anchor_hits, "positive_hits": positive_hits, "news_hits": news_hits, "ad_hits": ad_hits, "trash_hits": trash_hits, "region_relevance_score": round(region,3), "newsiness_score": round(news,3), "ad_score": round(ad,3), "trash_score": round(trash,3), "text_value_score": round(value,3), "positive_or_useful_tone_score": round(tone,3)}


def image_scores_skipped(reason: str) -> dict[str, Any]:
    return {
        "technical_quality_score":0.0,"aesthetic_score":0.0,"postcardness_score":0.0,"region_visual_relevance_score":0.0,
        "publication_safety_score":0.0,"low_noise_score":0.0,"overall_media_score":0.0,
        "is_selected_for_publication":False,"image_publication_ready":"false","image_reviewable":"false","image_quality_bucket":"skipped",
        "recognized_visual_elements":"","model_short_explanation":"Image scoring skipped by text gate: " + reason,
        "failure_reason":"image_scoring_skipped_by_text_gate","model_id":"not_run_text_gate","model_version":"2026-07-05",
        "image_model_type":"not_run","image_model_runtime":"not_run","image_model_input_type":"none","image_scoring_mode": current_image_scoring_mode(),
        "image_model_device":"not_run","image_download_status":"not_needed_text_gate",
    }


def current_image_scoring_mode() -> str:
    return (os.getenv("REGION_TALK_IMAGE_SCORING_MODE") or "external_ydb_queue").strip().lower() or "external_ydb_queue"


def _external_image_queue_pending_scores(has_media: bool, *, reason_prefix: str = "") -> dict[str, Any]:
    reason = "queued_for_region_talk_image_diagnostic" if has_media else "no_media"
    prefix = (reason_prefix + "; ") if reason_prefix else ""
    return {
        "technical_quality_score":0.0,"aesthetic_score":0.0,"postcardness_score":0.0,"region_visual_relevance_score":0.0,
        "publication_safety_score":1.0 if has_media else 0.0,"low_noise_score":0.0,"overall_media_score":0.0,
        "is_selected_for_publication":False,"image_publication_ready":"false","image_reviewable":"false",
        "image_quality_bucket":"metadata_pending_actual_image" if has_media else "no_media",
        "recognized_visual_elements":"",
        "model_short_explanation":prefix + "CandidateReport does not run image models; actual-image scoring is delegated to RegionTalkImageDiagnostic via YDB image_queue_item.",
        "failure_reason":"needs_actual_image_fetch" if has_media else "no_media",
        "model_id":"external_region_talk_image_diagnostic","model_version":"2026-07-07",
        "image_model_type":"external_ydb_queue","image_model_runtime":"not_run_in_candidate_report",
        "image_model_input_type":"metadata_only" if has_media else "none",
        "image_scoring_mode": current_image_scoring_mode(),"image_model_device":"not_run",
        "image_download_status":reason,
    }


def _image_scores_from_ydb_queue(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not isinstance(row, dict):
        return None
    if str(row.get("image_queue_status") or "") != "actual_scored" and str(row.get("image_model_input_type") or "") != "actual_image":
        return None
    try:
        overall = float(row.get("overall_media_score") or row.get("final_visual_score") or 0)
    except Exception:
        overall = 0.0
    if overall <= 0:
        return None
    postcard = row.get("postcardness_score") or row.get("clip_postcardness_score") or row.get("cv_postcardness_score") or overall
    aesthetic = row.get("aesthetic_score") or row.get("laion_aesthetic_score") or row.get("cv_aesthetic_score") or overall
    technical = row.get("technical_quality_score") or row.get("cv_technical_quality_score") or overall
    selected = overall >= float(os.getenv("REGION_TALK_IMAGE_PUBLICATION_READY_MIN_SCORE", "0.72"))
    reviewable = selected or overall >= float(os.getenv("REGION_TALK_IMAGE_REVIEWABLE_MIN_SCORE", "0.55"))
    return {
        "technical_quality_score":technical,"aesthetic_score":aesthetic,"postcardness_score":postcard,
        "region_visual_relevance_score":row.get("region_visual_relevance_score") or postcard,
        "publication_safety_score":row.get("publication_safety_score") or 0.98,"low_noise_score":row.get("low_noise_score") or row.get("cv_low_noise_score") or "",
        "overall_media_score":round(overall,3),"is_selected_for_publication":selected,"image_publication_ready":str(selected).lower(),
        "image_reviewable":str(reviewable).lower(),"image_quality_bucket":"publication_ready" if selected else ("reviewable_image" if reviewable else "weak_image"),
        "recognized_visual_elements":"actual image scored by RegionTalkImageDiagnostic",
        "model_short_explanation":"Actual-image score read from YDB image_queue_item written by RegionTalkImageDiagnostic.",
        "failure_reason":"" if reviewable else "below_reviewable_image_threshold",
        "model_id":row.get("image_model_type") or "region_talk_image_diagnostic_consensus",
        "model_version":"2026-07-07","image_model_type":row.get("image_model_type") or "multi_model_visual_consensus",
        "image_model_runtime":"external_region_talk_image_diagnostic","image_model_input_type":"actual_image",
        "image_scoring_mode": current_image_scoring_mode(),"image_model_device":row.get("image_model_device") or "external",
        "image_download_status":"downloaded_actual_image",
        "media_acquisition_status":row.get("media_acquisition_status") or "actual_image_downloaded_and_scored",
        "images_scored_actual_count":row.get("images_scored_actual_count") or 1,
        "final_visual_score":row.get("final_visual_score") or overall,
        "model_disagreement_score":row.get("model_disagreement_score") or "",
    }


def _metadata_media_scores(has_media: bool, text_score: dict[str, Any], *, reason_prefix: str = "") -> dict[str, Any]:
    if not has_media:
        return {"technical_quality_score":0.0,"aesthetic_score":0.0,"postcardness_score":0.0,"region_visual_relevance_score":0.0,"publication_safety_score":1.0,"low_noise_score":0.0,"overall_media_score":0.0,"is_selected_for_publication":False,"image_publication_ready":"false","image_reviewable":"false","image_quality_bucket":"no_media","recognized_visual_elements":"","model_short_explanation":"No media detected; image gate failed.","failure_reason":"no_media","model_id":"cv_only_metadata_v1","model_version":"2026-07-05","image_model_type":"cv_only","image_model_runtime":"kaggle_local","image_model_input_type":"metadata_only","image_scoring_mode": current_image_scoring_mode(),"image_model_device":"cpu","image_download_status":"no_media"}
    anchors = text_score.get("anchor_hits") or []
    positives = text_score.get("positive_hits") or []
    technical = 0.72
    aesthetic = min(0.92, 0.68 + 0.04*len(positives))
    postcard = min(0.94, 0.66 + 0.06*len(anchors) + 0.03*len(positives))
    region_visual = min(0.95, 0.58 + 0.08*len(anchors))
    safety = 0.98 if not text_score.get("news_hits") and not text_score.get("trash_hits") else 0.78
    low_noise = 0.82
    overall = round((technical+aesthetic+postcard+region_visual+safety+low_noise)/6,3)
    elements = []
    low = " ".join(anchors + positives).lower()
    if "море" in low or "балтий" in low: elements.append("sea/coast")
    if "курш" in low or "дюн" in low: elements.append("dunes/nature")
    if "архитект" in low or "кёниг" in low or "калининград" in low: elements.append("city/architecture")
    if not elements: elements.append("travel visual candidate")
    prefix = (reason_prefix + "; ") if reason_prefix else ""
    return {"technical_quality_score":round(technical,3),"aesthetic_score":round(aesthetic,3),"postcardness_score":round(postcard,3),"region_visual_relevance_score":round(region_visual,3),"publication_safety_score":round(safety,3),"low_noise_score":round(low_noise,3),"overall_media_score":overall,"is_selected_for_publication":False,"image_publication_ready":"false","image_reviewable":"false","image_quality_bucket":"metadata_pending_actual_image","recognized_visual_elements":"; ".join(elements),"model_short_explanation":prefix + f"cv_only metadata fallback: media present; region anchors={len(anchors)}, travel/visual cues={len(positives)}; safety={'ok' if safety>=0.95 else 'blocked by news/trash cues'}. Actual image bytes are required before reviewable/publication-ready.","failure_reason":"needs_actual_image_fetch","model_id":"cv_only_metadata_v1","model_version":"2026-07-05","image_model_type":"cv_only","image_model_runtime":"kaggle_local","image_model_input_type":"metadata_only","image_scoring_mode": current_image_scoring_mode(),"image_model_device":"cpu","image_download_status":"actual_image_missing_or_fallback"}


def _local_image_stats(path: str) -> dict[str, float]:
    try:
        from PIL import Image, ImageFilter, ImageStat  # type: ignore
    except Exception:
        if getenv_bool("REGION_TALK_AUTO_INSTALL", True):
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "pillow"])
            from PIL import Image, ImageFilter, ImageStat  # type: ignore
        else:
            raise
    img = Image.open(path).convert("RGB")
    w, h = img.size
    small = img.resize((max(1, min(512, w)), max(1, int(h * min(512, w) / max(1, w)))), Image.Resampling.LANCZOS)
    gray = small.convert("L")
    stat = ImageStat.Stat(gray)
    mean = float(stat.mean[0]) / 255.0
    contrast = float(stat.stddev[0]) / 96.0
    edges = gray.filter(ImageFilter.FIND_EDGES)
    edge_mean = float(ImageStat.Stat(edges).mean[0]) / 255.0
    aspect_ok = 0.0 if min(w, h) < 420 else 1.0
    brightness_ok = max(0.0, 1.0 - abs(mean - 0.52) / 0.52)
    technical = max(0.0, min(1.0, 0.35*aspect_ok + 0.35*brightness_ok + 0.30*min(1.0, contrast)))
    low_noise = max(0.0, min(1.0, 1.0 - max(0.0, edge_mean - 0.24) * 2.2))
    return {"technical": technical, "contrast": min(1.0, contrast), "low_noise": low_noise, "width": float(w), "height": float(h)}


def _ensure_clip_model() -> tuple[Any, Any, str]:
    global _REGION_TALK_CLIP_MODEL, _REGION_TALK_CLIP_PROCESSOR, _REGION_TALK_CLIP_DEVICE
    if _REGION_TALK_CLIP_MODEL is not None and _REGION_TALK_CLIP_PROCESSOR is not None and _REGION_TALK_CLIP_DEVICE:
        return _REGION_TALK_CLIP_MODEL, _REGION_TALK_CLIP_PROCESSOR, _REGION_TALK_CLIP_DEVICE
    try:
        import torch  # type: ignore
        from transformers import CLIPModel, CLIPProcessor  # type: ignore
    except Exception:
        if getenv_bool("REGION_TALK_AUTO_INSTALL", True):
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "pillow", "transformers", "torch"])
            import torch  # type: ignore
            from transformers import CLIPModel, CLIPProcessor  # type: ignore
        else:
            raise
    model_id = os.getenv("REGION_TALK_CLIP_MODEL_ID") or "openai/clip-vit-base-patch32"
    device = "cuda" if getattr(torch, "cuda", None) and torch.cuda.is_available() else "cpu"
    _REGION_TALK_CLIP_PROCESSOR = CLIPProcessor.from_pretrained(model_id)
    _REGION_TALK_CLIP_MODEL = CLIPModel.from_pretrained(model_id).to(device)
    _REGION_TALK_CLIP_MODEL.eval()
    _REGION_TALK_CLIP_DEVICE = device
    return _REGION_TALK_CLIP_MODEL, _REGION_TALK_CLIP_PROCESSOR, device


def _clip_image_scores(image_path: str, text_score: dict[str, Any]) -> dict[str, Any]:
    from PIL import Image  # type: ignore
    import torch  # type: ignore
    model, processor, device = _ensure_clip_model()
    img = Image.open(image_path).convert("RGB")
    prompts = [
        "beautiful high quality travel landscape photo of the Baltic Sea coast dunes forest or historic city architecture",
        "useful authentic travel photo from Kaliningrad Oblast or Baltic resort town",
        "low quality screenshot meme poster flyer advertisement with lots of text",
        "news accident crime politics trash shocking image",
    ]
    inputs = processor(text=prompts, images=img, return_tensors="pt", padding=True).to(device)
    with torch.no_grad():
        outputs = model(**inputs)
        probs = outputs.logits_per_image.softmax(dim=1)[0].detach().cpu().tolist()
    stats = _local_image_stats(image_path)
    travel, region_like, ad_like, news_like = [float(x) for x in probs]
    anchors = text_score.get("anchor_hits") or []
    positives = text_score.get("positive_hits") or []
    technical = max(0.0, min(1.0, stats["technical"]))
    aesthetic = max(0.0, min(1.0, 0.55*travel + 0.25*stats["contrast"] + 0.20*technical))
    postcard = max(0.0, min(1.0, 0.62*travel + 0.18*region_like + 0.10*min(1.0, len(positives)/4) + 0.10*technical))
    region_visual = max(0.0, min(1.0, 0.58*region_like + 0.20*travel + 0.22*min(1.0, len(anchors)/4)))
    safety = max(0.0, min(1.0, 1.0 - max(ad_like, news_like)*0.85))
    low_noise = max(0.0, min(1.0, stats["low_noise"]))
    overall = round((technical+aesthetic+postcard+region_visual+safety+low_noise)/6,3)
    reviewable = technical>=0.50 and postcard>=0.55 and safety>=0.65 and overall>=0.60
    selected = technical>=0.62 and aesthetic>=0.66 and postcard>=0.68 and region_visual>=0.52 and safety>=0.82 and low_noise>=0.62 and overall>=0.72
    elements = []
    if travel >= 0.35: elements.append("travel/postcard visual")
    if region_like >= 0.25 or anchors: elements.append("region-compatible visual")
    if ad_like >= 0.35: elements.append("possible poster/ad screenshot")
    if news_like >= 0.35: elements.append("possible news/trash visual")
    if not elements: elements.append("visual candidate")
    model_id = os.getenv("REGION_TALK_CLIP_MODEL_ID") or "openai/clip-vit-base-patch32"
    return {
        "technical_quality_score":round(technical,3),"aesthetic_score":round(aesthetic,3),"postcardness_score":round(postcard,3),
        "region_visual_relevance_score":round(region_visual,3),"publication_safety_score":round(safety,3),"low_noise_score":round(low_noise,3),
        "overall_media_score":overall,"is_selected_for_publication":selected,"image_publication_ready":str(selected).lower(),
        "image_reviewable":str(reviewable).lower(),"image_quality_bucket":"publication_ready" if selected else ("reviewable_image" if reviewable else "weak_image"),
        "recognized_visual_elements":"; ".join(elements),
        "model_short_explanation":f"Kaggle-local CLIP actual-image scoring; probs travel={travel:.3f}, region_like={region_like:.3f}, ad_like={ad_like:.3f}, news_like={news_like:.3f}; device={device}.",
        "failure_reason":"" if selected else ("reviewable_below_publication_threshold" if reviewable else "below_reviewable_image_threshold"),
        "model_id":"clip_local_" + re.sub(r"[^A-Za-z0-9_.-]+", "_", model_id),"model_version":"2026-07-05",
        "image_model_type":"clip","image_model_runtime":"kaggle_local","image_model_input_type":"actual_image","image_scoring_mode": current_image_scoring_mode(),
        "image_model_device":device,"image_download_status":"downloaded_actual_image",
    }


def vk_wall_token() -> str:
    if getenv_bool("REGION_TALK_VK_READ_SERVICE_FIRST", True):
        return (os.getenv("VK_SERVICE_TOKEN") or os.getenv("VK_SERVICE_KEY") or os.getenv("VK_TOKEN") or os.getenv("VK_ACCESS_TOKEN") or "").strip()
    return (os.getenv("VK_ACCESS_TOKEN") or os.getenv("VK_SERVICE_TOKEN") or os.getenv("VK_SERVICE_KEY") or os.getenv("VK_TOKEN") or "").strip()


def vk_wall_token_kind() -> str:
    token = vk_wall_token()
    if not token:
        return "missing"
    for name in ["VK_SERVICE_TOKEN", "VK_SERVICE_KEY", "VK_TOKEN", "VK_ACCESS_TOKEN"]:
        if token == (os.getenv(name) or "").strip():
            return name
    return "configured"


def vk_domain_from_seed(seed: Seed) -> str:
    raw = (seed.url or seed.handle or "").strip()
    m = re.search(r"vk\.com/(?:club|public)?([A-Za-z0-9_.-]+)", raw, re.I)
    if m:
        domain = m.group(1)
        if domain.lower() in {"search", "video", "clips", "feed", "places", "market", "groups", "public"}:
            return canonical_handle(seed.handle).lstrip("@")
        return domain
    return canonical_handle(seed.handle).lstrip("@")


def fetch_vk_wall_for_seed(seed: Seed, output_dir: Path, max_posts: int) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    run_id = os.getenv("REGION_TALK_RUN_ID") or f"region-talk-{RUN_STARTED_AT.strftime('%Y%m%dT%H%M%SZ')}"
    token = vk_wall_token()
    src = source_status_row(seed, "skipped_vk_wall_not_configured" if not token else "ok", vk_wall_probe_status="not_configured" if not token else "ok", fetch_attempted=str(bool(token)).lower())
    if not token:
        src["source_probe_reason"] = "VK token is not configured; source retained in frontier/backlog"
        return src, []
    domain = vk_domain_from_seed(seed)
    if not domain:
        src.update({"fetch_status": "skipped_vk_wall_domain_missing", "vk_wall_probe_status": "domain_missing", "source_probe_reason": "Cannot extract VK wall domain from seed/catalog/search URL"})
        return src, []
    try:
        import urllib.parse, urllib.request
        params = urllib.parse.urlencode({"domain": domain, "count": max(1, min(max_posts, getenv_int("REGION_TALK_VK_MAX_WALL_POSTS_PER_SOURCE", max_posts))), "filter": "owner", "extended": 1, "access_token": token, "v": os.getenv("VK_API_VERSION") or "5.199"})
        with urllib.request.urlopen("https://api.vk.com/method/wall.get?" + params, timeout=getenv_int("REGION_TALK_VK_TIMEOUT_SECONDS", 20)) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        if payload.get("error"):
            err = payload["error"] or {}
            code = int(err.get("error_code") or 0)
            status = "skipped_vk_wall_access_denied" if code in {15, 18, 30} else "error_vk_wall_api"
            probe = "access_denied" if code in {15, 18, 30} else "error"
            src.update({"fetch_status": status, "vk_wall_probe_status": probe, "fetch_error_code": err.get("error_code"), "fetch_error_message": str(err.get("error_msg") or "")[:180], "vk_token_kind": vk_wall_token_kind()})
            return src, []
        response = payload.get("response") or {}
        items = response.get("items") or []
        groups = {str(g.get("id")): g for g in response.get("groups") or [] if isinstance(g, dict)}
        posts: list[dict[str, Any]] = []
        media_budget = getenv_int("REGION_TALK_VK_MAX_MEDIA_DOWNLOADS_PER_RUN", getenv_int("REGION_TALK_TG_MAX_MEDIA_DOWNLOADS_PER_RUN", 120))
        for item in items:
            if not isinstance(item, dict):
                continue
            mid = item.get("id")
            owner_id = item.get("owner_id") or item.get("from_id")
            text = str(item.get("text") or "").strip()
            if not mid or not owner_id or not text:
                continue
            dt = datetime.fromtimestamp(int(item.get("date") or 0), timezone.utc).isoformat() if item.get("date") else ""
            post_url = f"https://vk.com/wall{owner_id}_{mid}"
            attachments = item.get("attachments") or []
            photo_url = ""
            for a in attachments:
                if not isinstance(a, dict) or a.get("type") != "photo":
                    continue
                sizes = ((a.get("photo") or {}).get("sizes") or [])
                if sizes:
                    best = sorted([s for s in sizes if isinstance(s, dict) and s.get("url")], key=lambda s: int(s.get("width") or 0) * int(s.get("height") or 0), reverse=True)[0]
                    photo_url = str(best.get("url") or "")
                    break
            primary_media_path = ""
            if photo_url and media_budget > 0 and getenv_bool("REGION_TALK_DOWNLOAD_MEDIA_FOR_SCORING", True):
                try:
                    media_dir = output_dir / "media" / stable_hash("vk", domain)
                    media_dir.mkdir(parents=True, exist_ok=True)
                    ext = ".jpg" if ".jpg" in photo_url.lower() or ".jpeg" in photo_url.lower() else ".jpg"
                    target = media_dir / f"{mid}{ext}"
                    urllib.request.urlretrieve(photo_url, target)
                    primary_media_path = str(target)
                    media_budget -= 1
                except Exception:
                    primary_media_path = ""
            copy_history = item.get("copy_history") or []
            origin = copy_history[0] if copy_history and isinstance(copy_history[0], dict) else {}
            origin_owner = origin.get("owner_id") or ""
            origin_mid = origin.get("id") or ""
            forwarded_url = f"https://vk.com/wall{origin_owner}_{origin_mid}" if origin_owner and origin_mid else ""
            group = groups.get(str(abs(int(owner_id))) if str(owner_id).lstrip("-").isdigit() else "") or {}
            title = str(group.get("name") or seed.source_title or domain)
            post_row = {"post_id":"post_"+stable_hash("vk", owner_id, mid), "source_id": seed.source_id, "source_seed_id": seed.source_seed_id, "source_title": title, "platform":"vk", "handle": seed.handle or domain, "post_url": post_url, "platform_post_key": f"vk:{owner_id}:{mid}", "post_date": dt, "text": text, "text_excerpt": re.sub(r"\s+", " ", text)[:500], "has_media": bool(photo_url), "media_count": 1 if photo_url else 0, "primary_media_path": primary_media_path, "local_media_paths": primary_media_path, "rights_policy": seed.rights_policy, "source_kind": seed.source_kind, "source_type": seed.source_kind, "source_url": seed.canonical_url, "is_forwarded_or_repost": bool(forwarded_url), "forwarded_from_source_title": "", "forwarded_from_source_id": "src_" + stable_hash("vk_forward", forwarded_url) if forwarded_url else "", "forwarded_from_platform": "vk" if forwarded_url else "", "forwarded_from_handle": "", "forwarded_from_url": forwarded_url, "forwarded_from_post_url": forwarded_url, "forwarded_from_confidence": 0.75 if forwarded_url else 0.0, "original_source_candidate_id": "src_cand_" + stable_hash("vk", forwarded_url) if forwarded_url else ""}
            append_post_online(posts, post_row, run_id=run_id, stage="vk_wall_fetch")
        src.update({"fetch_status": "ok", "vk_wall_probe_status": "ok", "posts_scanned": len(posts), "history_fetch_mode": "vk_wall_primary_scan", "history_fetch_runtime_seconds": "", "last_seen_post_date": max([p.get("post_date") or "" for p in posts] or [""]), "source_probe_reason": "minimal VK wall.get fetch; text/photos/copy_history origins", "vk_token_kind": vk_wall_token_kind()})
        return src, posts
    except Exception as exc:
        src.update({"fetch_status": "error_vk_wall_fetch", "vk_wall_probe_status": "error", "fetch_error_code": type(exc).__name__, "fetch_error_message": str(exc)[:180]})
        return src, []


def media_scores(has_media: bool, text_score: dict[str, Any], post: dict[str, Any] | None = None) -> dict[str, Any]:
    mode = current_image_scoring_mode()
    allow_local_models = getenv_bool("REGION_TALK_CANDIDATE_REPORT_ALLOW_IMAGE_MODEL_SCORING", False)
    if not allow_local_models or mode in {"external_ydb_queue", "off", "queue_only"}:
        return _external_image_queue_pending_scores(has_media, reason_prefix="candidate_report_image_models_disabled")
    post = post or {}
    image_path = str(post.get("primary_media_path") or "").strip()
    if mode in {"cv_aesthetic_clip", "clip", "cv_aesthetic_clip_vlm"} and has_media and image_path and Path(image_path).exists():
        try:
            return _clip_image_scores(image_path, text_score)
        except Exception as exc:
            return _metadata_media_scores(has_media, text_score, reason_prefix=f"neural_image_scoring_unavailable:{type(exc).__name__}")
    return _metadata_media_scores(has_media, text_score, reason_prefix="actual_image_missing" if has_media and mode != "cv_only" else "")


def candidate_score(text_score: dict[str, Any], media: dict[str, Any], seed: Seed) -> float:
    source_quality = 0.72 if seed.priority == 1 else 0.60 if seed.priority == 2 else 0.50
    score = text_score["region_relevance_score"]*0.20 + source_quality*0.12 + 0.55*0.12 + text_score["text_value_score"]*0.16 + text_score["positive_or_useful_tone_score"]*0.12 + media["overall_media_score"]*0.20 + 0.2*0.04 + 0.2*0.04
    score -= text_score["newsiness_score"]*0.35 + text_score["trash_score"]*0.40 + text_score["ad_score"]*0.18
    if not media["is_selected_for_publication"]: score -= 0.25
    if seed.rights_policy == "unknown": score -= 0.03
    return round(max(0.0, min(1.0, score)), 3)


def parse_public_telegram_post_url(url: str) -> tuple[str, int] | None:
    m = re.match(r"^https?://t\.me/([A-Za-z0-9_]+)/([0-9]+)(?:\?.*)?$", str(url or "").strip())
    if not m:
        return None
    handle = m.group(1).strip()
    if not handle or handle == "c":
        return None
    return handle, int(m.group(2))


def ydb_candidate_link_rows_from_row_kv(limit: int) -> list[dict[str, Any]]:
    cfg = ydb_config_status()
    if cfg.get("missing"):
        return []
    try:
        ydb, driver, cfg = ydb_connect()
        table_path = ydb_kv_table_path(cfg)
        pool = ydb.SessionPool(driver)
        def op(session: Any) -> list[dict[str, Any]]:
            ensure_ydb_kv_table(ydb, session, table_path)
            rows: list[dict[str, Any]] = []
            for kind in ("candidate_memory_item", "image_queue_item"):
                items = ydb_select_kind_items(session, ydb, table_path, kind, limit=max(1, limit * 2))
                for pk, payload in items.items():
                    if not isinstance(payload, dict):
                        continue
                    url = str(payload.get("post_url") or "").strip()
                    if not url:
                        continue
                    rows.append({"_pk": pk, "_kind": kind, **payload})
            return rows
        raw_rows = pool.retry_operation_sync(op)
        driver.stop(timeout=5)
    except Exception:
        return []
    seen: set[str] = set()
    out: list[dict[str, Any]] = []
    def rank(row: dict[str, Any]) -> tuple[int, str]:
        status = str(row.get("image_queue_status") or row.get("current_lifecycle_status") or "")
        pri = 0 if status == "actual_scored" else 1 if "image" in status else 2
        return pri, str(row.get("post_date") or "")
    for row in sorted(raw_rows, key=rank):
        url = str(row.get("post_url") or "").strip()
        if not url or url in seen:
            continue
        seen.add(url)
        out.append(row)
        if len(out) >= limit:
            break
    return out


async def fetch_ydb_candidate_link_posts_with_telethon(client: Any, status: Any, output_dir: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    run_id = os.getenv("REGION_TALK_RUN_ID") or f"region-talk-{RUN_STARTED_AT.strftime('%Y%m%dT%H%M%SZ')}"
    limit = getenv_int("REGION_TALK_YDB_CANDIDATE_LINK_LIMIT", getenv_int("REGION_TALK_MAX_POSTS_TO_SCORE_PER_RUN", 180))
    rows = ydb_candidate_link_rows_from_row_kv(limit)
    status.event("ydb_candidate_links_loaded", phase="fetch", status="running", run_id=run_id, candidate_links=len(rows), posts_to_score=min(limit, len(rows)), progress_label=f"YDB candidate links {len(rows)}")
    posts: list[dict[str, Any]] = []
    source_stats: dict[str, dict[str, Any]] = {}
    entity_cache: dict[str, Any] = {}
    for idx, row in enumerate(rows, start=1):
        url = str(row.get("post_url") or "")
        parsed = parse_public_telegram_post_url(url)
        if not parsed:
            continue
        handle, mid = parsed
        stat = source_stats.setdefault(handle, {"handle": handle, "source_title": row.get("source_title") or handle, "attempted": 0, "ok": 0, "errors": 0, "last_seen_post_date": ""})
        stat["attempted"] += 1
        status.event("alive", phase="fetch", status="running", run_id=run_id, progress_label=f"YDB post links {idx}/{len(rows)} · {handle}/{mid}", sources_done=idx - 1, sources_total=len(rows), current_source_handle=handle, current_source_title=stat.get("source_title"), current_source_url=f"https://t.me/{handle}")
        try:
            entity = entity_cache.get(handle)
            if entity is None:
                entity = await client.get_entity(handle)
                entity_cache[handle] = entity
            msg = await client.get_messages(entity, ids=mid)
            text = str(getattr(msg, "message", None) or "").strip() if msg is not None else ""
            if not text:
                stat["errors"] += 1
                continue
            dt = getattr(msg, "date", None)
            has_media = bool(getattr(msg, "photo", None) or getattr(msg, "document", None) or getattr(msg, "media", None))
            title = str(getattr(entity, "title", None) or row.get("source_title") or handle)
            post_date = dt.isoformat() if dt else str(row.get("post_date") or "")
            post_row = {
                "post_id": "post_" + stable_hash("telegram_ydb_link", handle, mid),
                "source_id": "src_ydb_link_" + stable_hash("telegram", handle),
                "source_seed_id": "ydb_candidate_links",
                "source_title": title,
                "platform": "telegram",
                "handle": handle,
                "post_url": f"https://t.me/{handle}/{mid}",
                "platform_post_key": f"tg:{handle}:{mid}",
                "post_date": post_date,
                "text": text,
                "text_excerpt": re.sub(r"\s+", " ", text)[:500],
                "has_media": has_media,
                "media_count": 1 if has_media else 0,
                "primary_media_path": "",
                "local_media_paths": "",
                "rights_policy": "unknown",
                "source_kind": "ydb_candidate_link",
                "source_type": "ydb_candidate_link",
                "source_url": f"https://t.me/{handle}",
                "is_forwarded_or_repost": False,
                "forwarded_from_source_title": "",
                "forwarded_from_source_id": "",
                "forwarded_from_platform": "",
                "forwarded_from_handle": "",
                "forwarded_from_url": "",
                "forwarded_from_post_url": "",
                "forwarded_from_confidence": 0.0,
                "original_source_candidate_id": "",
                "ydb_candidate_link_kind": row.get("_kind") or "",
                "ydb_candidate_link_pk": row.get("_pk") or "",
            }
            append_post_online(posts, post_row, run_id=run_id, stage="ydb_candidate_link_fetch")
            stat["ok"] += 1
            stat["last_seen_post_date"] = max(str(stat.get("last_seen_post_date") or ""), post_date)
        except Exception as exc:
            stat["errors"] += 1
            stat["last_error_code"] = type(exc).__name__
            stat["last_error_message"] = str(exc)[:180]
    source_rows = []
    for handle, stat in source_stats.items():
        ok = int(stat.get("ok") or 0)
        errors = int(stat.get("errors") or 0)
        append_source_row_online(source_rows, {
            "source_id": "src_ydb_link_" + stable_hash("telegram", handle),
            "source_seed_id": "ydb_candidate_links",
            "source_title": stat.get("source_title") or handle,
            "platform": "telegram",
            "handle": handle,
            "canonical_url": f"https://t.me/{handle}",
            "fetch_status": "ok" if ok else "error_ydb_candidate_link_fetch",
            "fetch_attempted": "true",
            "posts_scanned": ok,
            "history_fetch_mode": "ydb_candidate_links_only_no_discovery",
            "source_probe_reason": "Refetched only post URLs already present in YDB candidate/image queues; no source discovery/history scan.",
            "last_seen_post_date": stat.get("last_seen_post_date") or "",
            "fetch_error_code": stat.get("last_error_code", "") if errors else "",
            "fetch_error_message": stat.get("last_error_message", "") if errors else "",
            "vk_wall_probe_status": "not_applicable",
        }, run_id=run_id, stage="ydb_candidate_link_fetch", sources_total=len(source_stats))
    status.event("ydb_candidate_posts_fetched", phase="fetch", status="running", run_id=run_id, candidate_links=len(rows), posts_fetched=len(posts), sources_scanned=len(source_rows), progress_label=f"YDB candidate posts fetched {len(posts)}/{len(rows)}")
    return source_rows, posts


async def fetch_telegram_posts(seeds: list[Seed], status: Status, output_dir: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    max_sources = getenv_int("REGION_TALK_MAX_SOURCES", 5)
    max_posts = getenv_int("REGION_TALK_MAX_POSTS_PER_SOURCE", 20)
    post_input_mode = (os.getenv("REGION_TALK_POST_INPUT_MODE") or os.getenv("REGION_TALK_FETCH_MODE") or "").strip().lower()
    fetch_enabled = getenv_bool("REGION_TALK_FETCH_TELEGRAM", True)
    discovery_mode = (os.getenv("REGION_TALK_DISCOVERY_MODE") or "mixed").strip().lower()
    history_scan_mode = (os.getenv("REGION_TALK_HISTORY_SCAN_MODE") or "primary_and_delta").strip().lower()
    run_id = os.getenv("REGION_TALK_RUN_ID") or f"region-talk-{RUN_STARTED_AT.strftime('%Y%m%dT%H%M%SZ')}"
    previous_state, _state_meta = load_region_talk_state(output_dir)
    governor = TelegramRequestGovernor(run_id, output_dir, previous_state)
    ydb_candidate_links_only = post_input_mode in {"ydb_candidate_links", "ydb_candidates", "candidate_links"}
    queue_dynamic = [] if ydb_candidate_links_only else unified_queue_dynamic_seeds(previous_state, getenv_int("REGION_TALK_MAX_SOURCES", max_sources))
    dynamic = [] if ydb_candidate_links_only else queue_dynamic + frontier_dynamic_seeds(previous_state, getenv_int("REGION_TALK_MAX_NEW_SOURCE_PROBES", 30))
    all_seed_candidates = list(seeds)
    seen_seed_urls = {s.canonical_url for s in all_seed_candidates}
    for s in dynamic:
        if s.canonical_url not in seen_seed_urls:
            all_seed_candidates.append(s)
            seen_seed_urls.add(s.canonical_url)
    selected = selected_sources_for_run(all_seed_candidates, max_sources)
    _REGION_TALK_TELEGRAM_RUNTIME["dynamic_frontier_seed_count"] = len(dynamic)
    _REGION_TALK_TELEGRAM_RUNTIME["unified_queue_dynamic_seed_count"] = len(queue_dynamic)
    _REGION_TALK_TELEGRAM_RUNTIME["history_sources_target"] = governor.max_history_sources
    monitored = [s for s in selected if s.platform == "telegram" and (s.handle or "t.me/" in s.url.lower())]
    source_rows: list[dict[str, Any]] = []
    posts: list[dict[str, Any]] = []
    entity_by_source: dict[str, Any] = {}
    for seed in selected:
        write_region_talk_online_source_item(
            source_status_row(seed, "selected_for_run", selected_for_planning="true", fetch_attempted="false"),
            run_id=run_id,
            stage="source_selected_for_run",
            status="selected_for_run",
        )
    write_region_talk_online_stats({
        "run_id": run_id,
        "phase": "source_selected_for_run",
        "status": "running",
        "sources_seen_this_run": len(selected),
        "sources_done": 0,
        "sources_total": len(selected),
        "progress_label": f"selected sources {len(selected)}",
    })
    for seed in selected:
        if seed.platform == "telegram":
            continue
        if seed.platform == "vk":
            if fetch_enabled:
                vk_row, vk_posts = fetch_vk_wall_for_seed(seed, output_dir, max_posts)
                append_source_row_online(source_rows, vk_row, run_id=run_id, stage="source_fetch", sources_total=len(selected))
                posts.extend(vk_posts)
            else:
                status_value = "skipped_vk_wall_not_configured" if not vk_wall_token() else "skipped_fetch_disabled"
                append_source_row_online(source_rows, source_status_row(seed, status_value, vk_wall_probe_status="fetch_disabled", source_probe_reason="VK wall fetch disabled by REGION_TALK_FETCH_TELEGRAM=0"), run_id=run_id, stage="source_fetch", sources_total=len(selected))
        elif seed.platform == "vkvideo":
            if fetch_enabled and getenv_bool("REGION_TALK_FETCH_VKVIDEO_WALL_FALLBACK", True) and "vk.com/" in seed.url.lower() and "/video" not in seed.url.lower():
                vk_row, vk_posts = fetch_vk_wall_for_seed(seed, output_dir, max_posts)
                vk_row["platform"] = "vk"
                vk_row["source_probe_reason"] = (vk_row.get("source_probe_reason") or "") + "; source row was marked vkvideo but has a public vk.com wall URL, fetched as VK wall fallback"
                append_source_row_online(source_rows, vk_row, run_id=run_id, stage="source_fetch", sources_total=len(selected))
                posts.extend(vk_posts)
            else:
                append_source_row_online(source_rows, source_status_row(seed, "skipped_vkvideo_auxiliary_not_implemented", vk_wall_probe_status="not_applicable_vkvideo_auxiliary", source_probe_reason="VK Video is auxiliary for discovery/media, not a wall fetch source in this MVP run."), run_id=run_id, stage="source_fetch", sources_total=len(selected))
        else:
            append_source_row_online(source_rows, source_status_row(seed, "skipped_unsupported_platform", vk_wall_probe_status="not_applicable", source_probe_reason="Non-Telegram source is tracked for coverage/discovery, not fetched in this MVP run."), run_id=run_id, stage="source_fetch", sources_total=len(selected))
    if not fetch_enabled:
        for s in monitored:
            append_source_row_online(source_rows, source_status_row(s, "skipped_fetch_disabled"), run_id=run_id, stage="source_fetch_disabled", sources_total=len(monitored))
        return source_rows, posts
    try:
        from telethon import TelegramClient  # type: ignore
        from telethon.sessions import StringSession  # type: ignore
        import telethon  # type: ignore
        _REGION_TALK_TELEGRAM_RUNTIME["telethon_version"] = str(getattr(telethon, "__version__", ""))
    except Exception:
        if getenv_bool("REGION_TALK_AUTO_INSTALL", True):
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "telethon"])
            from telethon import TelegramClient  # type: ignore
            from telethon.sessions import StringSession  # type: ignore
            import telethon  # type: ignore
            _REGION_TALK_TELEGRAM_RUNTIME["telethon_version"] = str(getattr(telethon, "__version__", ""))
        else:
            for s in monitored:
                append_source_row_online(source_rows, source_status_row(s, "skipped_telethon_not_installed"), run_id=run_id, stage="source_fetch", sources_total=len(monitored))
            return source_rows, posts
    bundle_env = (os.getenv("REGION_TALK_AUTH_BUNDLE_ENV") or "TELEGRAM_AUTH_BUNDLE_DISCOVERY").strip()
    raw = (os.getenv(bundle_env) or "").strip()
    api_id = int((os.getenv("TG_API_ID") or os.getenv("TELEGRAM_API_ID") or "0").strip() or 0)
    api_hash = (os.getenv("TG_API_HASH") or os.getenv("TELEGRAM_API_HASH") or "").strip()
    bundle: dict[str, Any] = {}
    if raw:
        bundle = json.loads(base64.urlsafe_b64decode(raw.encode("ascii")).decode("utf-8"))
    session = str(bundle.get("session") or os.getenv("TG_SESSION") or os.getenv("TELEGRAM_SESSION") or "").strip()
    if not session or not api_id or not api_hash:
        for s in monitored:
            append_source_row_online(source_rows, source_status_row(s, "skipped_missing_telethon_credentials"), run_id=run_id, stage="source_fetch", sources_total=len(monitored))
        return source_rows, posts
    client = TelegramClient(StringSession(session), api_id, api_hash, device_model=str(bundle.get("device_model") or "Region Talk Discovery"), system_version=str(bundle.get("system_version") or "Linux"), app_version=str(bundle.get("app_version") or "1.0"), lang_code=str(bundle.get("lang_code") or "ru"), system_lang_code=str(bundle.get("system_lang_code") or "ru"))
    try:
        client.flood_sleep_threshold = getenv_int("REGION_TALK_TG_FLOODWAIT_MAX_SLEEP_SECONDS", 60)
    except Exception:
        pass
    await client.connect()
    if not await client.is_user_authorized():
        await client.disconnect()
        raise RuntimeError("Telethon client is not authorized")
    try:
        if ydb_candidate_links_only:
            return await fetch_ydb_candidate_link_posts_with_telethon(client, status, output_dir)
        for idx, seed in enumerate(monitored, start=1):
            source_progress = {
                "phase": "fetch",
                "progress_label": f"источники {idx}/{len(monitored)} · {seed.source_title or seed.handle or seed.canonical_url}",
                "sources_done": idx - 1,
                "sources_total": len(monitored),
                "source_index": idx,
                "source_id": seed.source_id,
                "current_source_title": seed.source_title,
                "current_source_url": seed.canonical_url,
                "current_source_handle": seed.handle,
                "current_source_platform": seed.platform,
                "event_name": "alive",
            }
            status.event("alive", **source_progress)
            handle = seed.handle.lstrip("@")
            if not handle or "/" in handle or handle.startswith("http"):
                skipped = source_status_row(seed, "skipped_telegram_handle_not_configured", vk_wall_probe_status="not_applicable")
                append_source_row_online(source_rows, skipped, run_id=run_id, stage="source_fetch", sources_total=len(monitored))
                status.event("alive", **{**source_progress, "progress_label": f"источники {idx}/{len(monitored)} · skipped · {seed.source_title or seed.canonical_url}", "sources_done": idx, "fetch_status": skipped.get("fetch_status"), "posts_scanned": 0})
                continue
            src_row = source_status_row(seed, "ok", vk_wall_probe_status="not_applicable", selected_for_planning="true", fetch_attempted="false")
            entity, resolve_meta = await governor.resolve_entity(client, seed)
            src_row.update(resolve_meta)
            if entity is None:
                src_row.setdefault("fetch_status", resolve_meta.get("fetch_status") or "skipped_telegram_unresolved_deferred")
                append_source_row_online(source_rows, src_row, run_id=run_id, stage="source_fetch", sources_total=len(monitored))
                status.event("alive", **{**source_progress, "progress_label": f"источники {idx}/{len(monitored)} · unresolved · {seed.source_title or seed.canonical_url}", "sources_done": idx, "fetch_status": src_row.get("fetch_status"), "telegram_resolve_status": src_row.get("telegram_resolve_status"), "posts_scanned": 0, "fetch_error_code": src_row.get("last_resolve_error_code", ""), "fetch_error_message": src_row.get("last_resolve_error_message_short", "")})
                continue
            if history_scan_mode == "off" or discovery_mode in {"similar_only", "keyword_only"}:
                src_row.update({"fetch_status": "profile_resolved_history_disabled", "fetch_attempted": "false", "source_probe_reason": f"history scan disabled by discovery_mode={discovery_mode} history_scan_mode={history_scan_mode}", "history_fetch_mode": "history_disabled_discovery_only", "posts_scanned": 0})
                append_source_row_online(source_rows, src_row, run_id=run_id, stage="source_fetch", sources_total=len(monitored))
                entity_by_source[seed.source_id] = entity
                status.event("alive", **{**source_progress, "progress_label": f"источники {idx}/{len(monitored)} · history disabled · {src_row.get('resolved_title') or seed.source_title or seed.canonical_url}", "sources_done": idx, "fetch_status": src_row.get("fetch_status"), "posts_scanned": 0})
                continue
            if governor.history_sources_attempted >= governor.max_history_sources:
                src_row.update({"fetch_status": "skipped_telegram_budget_exhausted", "fetch_attempted": "false", "source_probe_reason": "history source budget exhausted"})
                append_source_row_online(source_rows, src_row, run_id=run_id, stage="source_fetch", sources_total=len(monitored))
                status.event("alive", **{**source_progress, "progress_label": f"источники {idx}/{len(monitored)} · budget exhausted · {seed.source_title or seed.canonical_url}", "sources_done": idx, "fetch_status": src_row.get("fetch_status"), "posts_scanned": 0})
                continue
            entity_by_source[seed.source_id] = entity
            governor.history_sources_attempted += 1
            src_row["fetch_attempted"] = "true"
            history_fetch_started = time.monotonic()
            src_row["history_fetch_mode"] = "delta_scan_active" if not seed.source_seed_id.startswith("frontier_dynamic_") else "first_probe_or_shallow_backfill"
            src_row["is_new_source_this_run"] = str(seed.source_seed_id.startswith("frontier_dynamic_")).lower()
            src_row["history_source_cached_entity"] = str(resolve_meta.get("telegram_resolve_status") == "resolved_from_private_cache").lower()
            src_row["history_source_network_resolved"] = str(resolve_meta.get("telegram_resolve_status") == "resolved_network").lower()
            try:
                title = str(resolve_meta.get("resolved_title") or getattr(entity, "title", None) or seed.source_title)
                src_row["resolved_title"] = title
                seen: set[int] = set()
                anchor_cap = getenv_int("REGION_TALK_TG_MAX_ANCHOR_QUERIES_PER_SOURCE", 3)
                queries = [None] + DEFAULT_ANCHORS[:max(0, anchor_cap)]
                for q in queries:
                    if not governor.has_total_request_budget("iter_messages", seed.source_id, seed.canonical_url):
                        src_row.update({"fetch_status": "skipped_telegram_total_request_budget_exhausted", "source_probe_reason": "total Telegram request budget exhausted"})
                        break
                    per_query = max(3, min(max_posts, getenv_int("REGION_TALK_TG_MAX_HISTORY_POSTS_PER_SOURCE", max_posts)) // len(queries) + 1)
                    governor.total_attempted += 1
                    governor.requests_by_method["iter_messages"] = governor.requests_by_method.get("iter_messages", 0) + 1
                    async for msg in client.iter_messages(entity, limit=per_query, search=q):
                        mid = int(getattr(msg, "id", 0) or 0)
                        if not mid or mid in seen:
                            continue
                        seen.add(mid)
                        text = str(getattr(msg, "message", None) or "").strip()
                        if not text:
                            continue
                        dt = getattr(msg, "date", None)
                        post_url = f"https://t.me/{handle}/{mid}"
                        has_media = bool(getattr(msg, "photo", None) or getattr(msg, "document", None) or getattr(msg, "media", None))
                        primary_media_path = ""
                        if has_media and getenv_bool("REGION_TALK_DOWNLOAD_MEDIA_FOR_SCORING", True) and governor.media_downloads_attempted < governor.max_media_downloads:
                            governor.media_downloads_attempted += 1
                            try:
                                media_dir = output_dir / "media" / stable_hash("telegram", handle)
                                media_dir.mkdir(parents=True, exist_ok=True)
                                downloaded = await client.download_media(msg, file=str(media_dir / f"{mid}"))
                                if downloaded:
                                    primary_media_path = str(downloaded)
                                    governor.media_downloads_ok += 1
                            except Exception:
                                primary_media_path = ""
                        fwd = getattr(msg, "fwd_from", None) or getattr(msg, "forward", None)
                        forwarded_from_handle = ""
                        forwarded_from_title = ""
                        forwarded_from_url = ""
                        forwarded_from_post_url = ""
                        forwarded_from_confidence = 0.0
                        if fwd:
                            from_name = str(getattr(fwd, "from_name", None) or "").strip()
                            forwarded_from_title = from_name
                            from_id = getattr(fwd, "from_id", None)
                            channel_id = str(getattr(from_id, "channel_id", "") or "")
                            if channel_id:
                                forwarded_from_handle = "channel_" + channel_id
                                forwarded_from_url = "https://t.me/c/" + channel_id
                                original_mid = getattr(fwd, "channel_post", None)
                                if original_mid:
                                    forwarded_from_post_url = forwarded_from_url + "/" + str(original_mid)
                                forwarded_from_confidence = 0.7
                            elif from_name:
                                forwarded_from_confidence = 0.45
                        post_row = {"post_id":"post_"+stable_hash("telegram", handle, mid), "source_id": seed.source_id, "source_seed_id": seed.source_seed_id, "source_title": title, "platform":"telegram", "handle": seed.handle, "post_url": post_url, "platform_post_key": f"tg:{handle}:{mid}", "post_date": dt.isoformat() if dt else "", "text": text, "text_excerpt": re.sub(r"\s+", " ", text)[:500], "has_media": has_media, "media_count": 1 if has_media else 0, "primary_media_path": primary_media_path, "local_media_paths": primary_media_path, "rights_policy": seed.rights_policy, "source_kind": seed.source_kind, "source_type": seed.source_kind, "source_url": seed.canonical_url, "is_forwarded_or_repost": bool(fwd), "forwarded_from_source_title": forwarded_from_title, "forwarded_from_source_id": "src_" + stable_hash("telegram_forward", forwarded_from_url or forwarded_from_title) if fwd else "", "forwarded_from_platform": "telegram" if fwd else "", "forwarded_from_handle": forwarded_from_handle, "forwarded_from_url": forwarded_from_url, "forwarded_from_post_url": forwarded_from_post_url, "forwarded_from_confidence": forwarded_from_confidence, "original_source_candidate_id": "src_cand_" + stable_hash("telegram", forwarded_from_url or forwarded_from_title) if fwd else ""}
                        append_post_online(posts, post_row, run_id=run_id, stage="telegram_history_fetch")
                        if len(seen) >= max_posts:
                            break
                    if len(seen) >= max_posts:
                        break
                governor.total_ok += 1
                governor.history_sources_ok += 1
                governor.history_posts_fetched += len(seen)
                src_row["posts_scanned"] = len(seen)
                post_dates = [str(p.get("post_date") or "") for p in posts if p.get("source_id") == seed.source_id]
                src_row["last_seen_post_date"] = max(post_dates or [""])
                src_row["history_fetch_runtime_seconds"] = round(time.monotonic() - history_fetch_started, 3)
                jitter = float(os.getenv("REGION_TALK_TG_HISTORY_SOURCE_JITTER_SECONDS") or "0.5")
                if jitter > 0:
                    await asyncio.sleep(random.uniform(0, jitter))
            except Exception as exc:
                seconds = int(getattr(exc, "seconds", 0) or 0)
                if type(exc).__name__ == "FloodWaitError" or seconds:
                    cooldown_until = governor.mark_floodwait("iter_messages", seed.source_id, seed.canonical_url, seconds)
                    src_row["fetch_status"] = "error_floodwait_history"
                    src_row["next_allowed_resolve_at"] = cooldown_until
                else:
                    src_row["fetch_status"] = "error_telegram_rpc"
                src_row["fetch_error_code"] = type(exc).__name__
                src_row["fetch_error_message"] = str(exc)[:180]
            append_source_row_online(source_rows, src_row, run_id=run_id, stage="source_fetch", sources_total=len(monitored))
            status.event("alive", **{
                **source_progress,
                "progress_label": f"источники {idx}/{len(monitored)} · done · {src_row.get('resolved_title') or seed.source_title or seed.canonical_url}",
                "sources_done": idx,
                "fetch_status": src_row.get("fetch_status"),
                "telegram_resolve_status": src_row.get("telegram_resolve_status"),
                "posts_scanned": src_row.get("posts_scanned", 0),
                "last_seen_post_date": src_row.get("last_seen_post_date", ""),
                "history_fetch_runtime_seconds": src_row.get("history_fetch_runtime_seconds", ""),
                "fetch_error_code": src_row.get("fetch_error_code", ""),
                "fetch_error_message": src_row.get("fetch_error_message", ""),
            })
        similar_rows, similar_edges = await discover_telegram_similar_channels(client, source_rows, entity_by_source, governor, run_id)
        keyword_rows, keyword_edges = await discover_telegram_keyword_sources(client, governor, run_id)
        _REGION_TALK_TELEGRAM_RUNTIME["similar_rows"] = similar_rows
        _REGION_TALK_TELEGRAM_RUNTIME["similar_edges"] = similar_edges
        _REGION_TALK_TELEGRAM_RUNTIME["keyword_rows"] = keyword_rows
        _REGION_TALK_TELEGRAM_RUNTIME["keyword_edges"] = keyword_edges
    finally:
        governor.write_ledger()
        _REGION_TALK_TELEGRAM_RUNTIME["entity_cache"] = governor.entity_cache
        _REGION_TALK_TELEGRAM_RUNTIME["cooldowns"] = governor.cooldowns
        _REGION_TALK_TELEGRAM_RUNTIME["observability"] = governor.observability_row(RUN_STARTED_AT.isoformat(), utc_now_iso())
        await client.disconnect()
    return source_rows, posts



def llm_text_gate_prompt(post: dict[str, Any], evidence: dict[str, Any]) -> str:
    text = str(post.get("text") or "")[:6000]
    payload = {
        "task": "Decide whether this Telegram/VK post is a reviewable Region Talk candidate.",
        "rules": [
            "Accept only if the main subject is Kaliningrad Oblast, not one item in a multi-region/country digest.",
            "Reject ads, promos, registrations, contests, paid/service CTAs, app-download promos, event announcements, and news digests.",
            "Prefer visit impressions, route/useful observations, emotional or memorable details about visiting the region.",
            "For evidence_hints_not_decisions.stage=publication_queue_final_verifier, accept only when supplied visual evidence says this is an actual-image candidate with strong postcard/aesthetic scores; otherwise return needs_review or reject.",
            "A single-location Kaliningrad Oblast card is NOT a multi-region roundup even if it belongs to a source rubric; classify it as single_location_photo_card or encyclopedic_card_candidate, not reject_multi_region_roundup.",
            "Encyclopedic/single-location cards may be research candidates but weaker than firsthand visit impressions; pure visual dumps or passing mentions are not publication candidates.",
            "Return JSON only.",
        ],
        "evidence_hints_not_decisions": evidence,
        "post": {
            "source_title": post.get("source_title"),
            "post_url": post.get("post_url"),
            "post_date": post.get("post_date"),
            "text": text,
        },
        "schema": {
            "decision": "accept|needs_review|reject",
            "whole_post_about_kaliningrad_oblast_score": "0..1",
            "kaliningrad_mention_role": "main_subject|one_item|passing_mention|footer|hashtag|link_only|unclear",
            "is_digest_or_roundup": "boolean",
            "is_multi_region_roundup": "boolean",
            "is_multi_topic_digest": "boolean",
            "is_single_location_card": "boolean",
            "is_author_visit_impression": "boolean",
            "is_official_route_material": "boolean",
            "is_photo_card_from_subscriber": "boolean",
            "is_ad_or_promo": "boolean",
            "is_news_or_trash": "boolean",
            "content_type": "visit_impression_candidate|route_useful_candidate|single_location_photo_card|encyclopedic_card_candidate|official_project_material|news_or_event|low_substance|reject",
            "visit_evidence_type": "firsthand_author_visit|subscriber_photo_report|route_guide|single_location_photo_card|official_project_material|news_or_event|unknown",
            "has_firsthand_visit_evidence": "boolean",
            "emotion_or_impression_evidence": "boolean",
            "review_or_opinion_evidence": "boolean",
            "memorable_detail_evidence": "boolean",
            "original_photo_evidence": "boolean",
            "reason": "short Russian explanation",
        },
    }
    return json.dumps(payload, ensure_ascii=False)


def parse_llm_json(text: str) -> dict[str, Any]:
    raw = (text or "").strip()
    raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw, flags=re.I | re.S).strip()
    try:
        data = json.loads(raw)
    except Exception:
        match = re.search(r"\{.*\}", raw, flags=re.S)
        data = json.loads(match.group(0)) if match else {}
    return data if isinstance(data, dict) else {}


class _SupabaseRestResult:
    def __init__(self, data: Any):
        self.data = data


class _SupabaseRestQuery:
    def __init__(self, client: "_SupabaseRestClient", table: str):
        self.client = client
        self.table = table
        self.params: dict[str, str] = {}
        self.order_parts: list[str] = []

    def select(self, columns: str = "*") -> "_SupabaseRestQuery":
        self.params["select"] = columns
        return self

    def eq(self, column: str, value: Any) -> "_SupabaseRestQuery":
        self.params[column] = "eq." + str(value).lower() if isinstance(value, bool) else "eq." + str(value)
        return self

    def in_(self, column: str, values: list[Any]) -> "_SupabaseRestQuery":
        encoded = ",".join(str(v) for v in values)
        self.params[column] = f"in.({encoded})"
        return self

    def order(self, column: str) -> "_SupabaseRestQuery":
        self.order_parts.append(column)
        return self

    def limit(self, n: int) -> "_SupabaseRestQuery":
        self.params["limit"] = str(int(n))
        return self

    def execute(self) -> _SupabaseRestResult:
        params = dict(self.params)
        if self.order_parts:
            params["order"] = ",".join(self.order_parts)
        data = self.client._request("GET", f"/rest/v1/{self.table}", params=params)
        return _SupabaseRestResult(data)


class _SupabaseRestRpc:
    def __init__(self, client: "_SupabaseRestClient", fn_name: str, payload: dict[str, Any]):
        self.client = client
        self.fn_name = fn_name
        self.payload = payload

    def execute(self) -> _SupabaseRestResult:
        data = self.client._request("POST", f"/rest/v1/rpc/{self.fn_name}", json_body=self.payload)
        return _SupabaseRestResult(data)


class _SupabaseRestClient:
    def __init__(self, url: str, key: str, *, schema: str = "public") -> None:
        self.url = url.rstrip("/")
        self.key = key
        self.schema = schema or "public"

    def table(self, table: str) -> _SupabaseRestQuery:
        return _SupabaseRestQuery(self, table)

    def rpc(self, fn_name: str, payload: dict[str, Any]) -> _SupabaseRestRpc:
        return _SupabaseRestRpc(self, fn_name, payload)

    def _request(self, method: str, path: str, *, params: dict[str, str] | None = None, json_body: Any = None) -> Any:
        try:
            import requests  # type: ignore
        except Exception:
            if getenv_bool("REGION_TALK_AUTO_INSTALL", True):
                subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "requests"])
                import requests  # type: ignore
            else:
                raise
        headers = {
            "apikey": self.key,
            "Authorization": "Bearer " + self.key,
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Accept-Profile": self.schema,
            "Content-Profile": self.schema,
        }
        resp = requests.request(method, self.url + path, headers=headers, params=params, json=json_body, timeout=30)
        if resp.status_code >= 400:
            raise RuntimeError(f"supabase_rest_{resp.status_code}: {resp.text[:500]}")
        if not resp.text.strip():
            return None
        return resp.json()


def build_region_talk_supabase_client() -> Any:
    global _REGION_TALK_SUPABASE_CLIENT
    if _REGION_TALK_SUPABASE_CLIENT is not None:
        return _REGION_TALK_SUPABASE_CLIENT
    if (os.getenv("SUPABASE_DISABLED") or "").strip() == "1":
        raise RuntimeError("supabase_disabled")
    base_url = (os.getenv("SUPABASE_URL") or "").strip().rstrip("/")
    key = (os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_KEY") or "").strip()
    if not base_url or not key:
        raise RuntimeError("missing_SUPABASE_URL_or_service_key")
    schema = (os.getenv("SUPABASE_SCHEMA") or "public").strip() or "public"
    _REGION_TALK_SUPABASE_CLIENT = _SupabaseRestClient(base_url, key, schema=schema)
    return _REGION_TALK_SUPABASE_CLIENT




def ensure_google_ai_import_path() -> None:
    candidates = [Path.cwd(), Path(__file__).resolve().parent, Path("/kaggle/working")]
    inp = Path("/kaggle/input")
    if inp.exists():
        candidates.extend(p.parent.parent for p in inp.rglob("google_ai/__init__.py"))
    for parent in candidates:
        if (parent / "google_ai" / "__init__.py").exists() and str(parent) not in sys.path:
            sys.path.insert(0, str(parent))

def get_region_talk_llm_gateway(default_env_var_name: str) -> Any:
    global _REGION_TALK_GOOGLE_CLIENT
    if _REGION_TALK_GOOGLE_CLIENT is not None:
        return _REGION_TALK_GOOGLE_CLIENT
    ensure_google_ai_import_path()
    try:
        from google_ai import GoogleAIClient, SecretsProvider  # type: ignore
    except Exception:
        raise RuntimeError("google_ai_gateway_package_missing")
    client = GoogleAIClient(
        supabase_client=build_region_talk_supabase_client(),
        secrets_provider=SecretsProvider(),
        consumer="region_talk_candidate_report",
        account_name=os.getenv("GOOGLE_API_LOCALNAME_REGION_TALK") or os.getenv("GOOGLE_API_LOCALNAME"),
        default_env_var_name=default_env_var_name or "GOOGLE_API_KEY",
    )
    client.allow_reserve_fallback = False
    client.allow_local_limiter_fallback = False
    client.allow_local_limiter_on_reserve_error = False
    _REGION_TALK_GOOGLE_CLIENT = client
    return client


def load_llm_limit_snapshot(model: str, default_env_var_name: str) -> dict[str, Any]:
    try:
        sb = build_region_talk_supabase_client()
        limits = sb.table("google_ai_model_limits").select("model,rpm,tpm,rpd,tpm_reserve_extra").eq("model", model).limit(1).execute().data or []
        keys = sb.table("google_ai_api_keys").select("id,env_var_name,key_alias,priority,is_active").eq("is_active", True).in_("env_var_name", [default_env_var_name]).order("priority").order("id").limit(5).execute().data or []
        return {
            "llm_limit_source": "supabase_google_ai",
            "llm_provider": "google_gemini",
            "llm_model": model,
            "llm_default_env_var_name": default_env_var_name,
            "supabase_limiter_model_found": str(bool(limits)).lower(),
            "supabase_scoped_key_found": str(bool(keys)).lower(),
            "supabase_model_rpm": (limits[0].get("rpm") if limits else ""),
            "supabase_model_tpm": (limits[0].get("tpm") if limits else ""),
            "supabase_model_rpd": (limits[0].get("rpd") if limits else ""),
            "llm_config_source": "supabase_google_ai_model_limits_and_google_ai_reserve",
        }
    except Exception as exc:
        return {
            "llm_limit_source": "supabase_required_unavailable",
            "llm_provider": "google_gemini",
            "llm_model": model,
            "llm_default_env_var_name": default_env_var_name,
            "supabase_limiter_model_found": "false",
            "supabase_scoped_key_found": "false",
            "llm_config_source": "supabase_required_but_failed",
            "llm_limit_error": f"{type(exc).__name__}: {str(exc)[:240]}",
        }


def call_region_talk_semantic_llm(post: dict[str, Any], evidence: dict[str, Any], *, model: str | None = None, default_env_var_name: str | None = None) -> dict[str, Any]:
    # Final semantic decision only. Budget/key selection must go through Supabase google_ai_reserve.
    model = (model or os.getenv("REGION_TALK_LLM_MODEL") or "gemini-3.1-flash-lite").strip()
    default_env_var_name = (default_env_var_name or os.getenv("REGION_TALK_LLM_DEFAULT_ENV_VAR_NAME") or "GOOGLE_API_KEY3").strip()
    try:
        from google_ai.exceptions import RateLimitError, ProviderError, ReservationError  # type: ignore
    except Exception:
        class RateLimitError(Exception): pass
        class ProviderError(Exception): pass
        class ReservationError(Exception): pass
    try:
        try:
            from google import genai as _genai  # noqa: F401
        except Exception:
            if getenv_bool("REGION_TALK_AUTO_INSTALL", True):
                subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "google-genai"])
        prompt = llm_text_gate_prompt(post, evidence)
        client = get_region_talk_llm_gateway(default_env_var_name)

        async def _call() -> tuple[str, Any]:
            return await client.generate_content_async(
                model=model,
                prompt=prompt,
                generation_config={"temperature": 0.1, "response_mime_type": "application/json"},
                max_output_tokens=700,
            )
        try:
            running_loop = asyncio.get_running_loop()
        except RuntimeError:
            running_loop = None
        if running_loop is not None:
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
                text, usage = pool.submit(lambda: asyncio.run(_call())).result()
        else:
            text, usage = asyncio.run(_call())
        data = parse_llm_json(text)
        decision = str(data.get("decision") or "needs_review").strip().lower()
        if decision not in {"accept", "needs_review", "reject"}:
            decision = "needs_review"
        return {
            "llm_gate_status": "ok",
            "llm_provider": "google_gemini",
            "llm_model": model,
            "llm_default_env_var_name": default_env_var_name,
            "llm_limit_source": "supabase_google_ai_reserve",
            "llm_decision": decision,
            "whole_post_about_kaliningrad_oblast_score": data.get("whole_post_about_kaliningrad_oblast_score", ""),
            "kaliningrad_mention_role": str(data.get("kaliningrad_mention_role") or "unclear"),
            "is_digest_or_roundup": str(bool(data.get("is_digest_or_roundup"))).lower(),
            "is_multi_region_roundup": str(bool(data.get("is_multi_region_roundup"))).lower(),
            "is_multi_topic_digest": str(bool(data.get("is_multi_topic_digest"))).lower(),
            "is_single_location_card": str(bool(data.get("is_single_location_card"))).lower(),
            "is_author_visit_impression": str(bool(data.get("is_author_visit_impression"))).lower(),
            "is_official_route_material": str(bool(data.get("is_official_route_material"))).lower(),
            "is_photo_card_from_subscriber": str(bool(data.get("is_photo_card_from_subscriber"))).lower(),
            "llm_is_ad_or_promo": str(bool(data.get("is_ad_or_promo"))).lower(),
            "llm_is_news_or_trash": str(bool(data.get("is_news_or_trash"))).lower(),
            "llm_content_type": str(data.get("content_type") or ""),
            "content_type": str(data.get("content_type") or ""),
            "visit_evidence_type": str(data.get("visit_evidence_type") or ""),
            "has_firsthand_visit_evidence": str(bool(data.get("has_firsthand_visit_evidence"))).lower(),
            "emotion_or_impression_evidence": str(bool(data.get("emotion_or_impression_evidence"))).lower(),
            "review_or_opinion_evidence": str(bool(data.get("review_or_opinion_evidence"))).lower(),
            "memorable_detail_evidence": str(bool(data.get("memorable_detail_evidence"))).lower(),
            "original_photo_evidence": str(bool(data.get("original_photo_evidence"))).lower(),
            "llm_reason": str(data.get("reason") or "")[:500],
            "llm_usage_input_tokens": getattr(usage, "input_tokens", ""),
            "llm_usage_output_tokens": getattr(usage, "output_tokens", ""),
            "llm_usage_total_tokens": getattr(usage, "total_tokens", ""),
        }
    except RateLimitError as exc:
        return {"llm_gate_status": "rate_limited", "llm_provider": "google_gemini", "llm_model": model, "llm_default_env_var_name": default_env_var_name, "llm_limit_source": "supabase_google_ai_reserve", "llm_reason": f"RateLimitError: {str(exc)[:240]}"}
    except (ProviderError, ReservationError) as exc:
        msg = str(exc)
        status = "rate_limited" if "429" in msg or "resource_exhausted" in msg.lower() or "rate limit" in msg.lower() else "error"
        return {"llm_gate_status": status, "llm_provider": "google_gemini", "llm_model": model, "llm_default_env_var_name": default_env_var_name, "llm_limit_source": "supabase_google_ai_reserve", "llm_reason": f"{type(exc).__name__}: {msg[:240]}"}
    except Exception as exc:
        return {"llm_gate_status": "error", "llm_provider": "google_gemini", "llm_model": model, "llm_default_env_var_name": default_env_var_name, "llm_limit_source": "supabase_google_ai_reserve", "llm_reason": f"{type(exc).__name__}: {str(exc)[:240]}"}


def should_send_to_llm(fresh: dict[str, Any], scope: dict[str, Any], post: dict[str, Any], *, ad_gate: dict[str, Any] | None = None, substance: dict[str, Any] | None = None, text_score: dict[str, Any] | None = None) -> bool:
    if not fresh.get("fresh_enough"):
        return False
    ad_gate = ad_gate or {}
    substance = substance or {}
    text_score = text_score or {}
    # Cost guard only: do not spend Supabase LLM quota on rows with no strong Kaliningrad-Oblast evidence.
    # The semantic decision for remaining rows is still owned by the LLM final gate.
    if not scope.get("kaliningrad_oblast_only_scope"):
        return False
    if not scope.get("matched_place_names") and not any(anchor.lower() in str(post.get("text") or "").lower() for anchor in DEFAULT_ANCHORS[:6]):
        return False
    if bool(ad_gate.get("is_ad_or_promo")) and float(substance.get("visit_impression_score") or 0) < 0.20:
        return False
    if float(substance.get("text_substance_score") or 0) < 0.18:
        return False
    if float(text_score.get("newsiness_score") or 0) >= 0.70 or float(text_score.get("trash_score") or 0) >= 0.70:
        return False
    return True


TEXT_EMBEDDING_MODELS = ["intfloat/multilingual-e5-base", "BAAI/bge-m3"]


def semantic_bank_v1() -> dict[str, list[str]]:
    """Prototype texts for vector-first Region Talk selection.

    The bank is intentionally semantic examples, not regex rules. Keyword/lexicon
    features remain evidence/fallback; when embeddings are enabled, acceptance and
    mass rejection come from distance to these positive/negative meaning classes.
    """
    return {
        "ko_visit_impression": [
            "Личный рассказ о поездке в Калининградскую область: что увидели, что понравилось, какие места запомнились.",
            "Автор делится впечатлениями от Зеленоградска, Светлогорска, Куршской косы, Балтийска или других городов Калининградской области.",
            "Фотоотчет или заметка путешественника о посещении Калининградской области с эмоциями и наблюдениями.",
        ],
        "ko_route_useful": [
            "Полезный маршрут по Калининградской области: как добраться, что посмотреть, где гулять, советы для поездки.",
            "Содержательная карточка о достопримечательности Калининградской области, истории места, природе, море, дюнах или архитектуре.",
            "Пост о нескольких городах или местах внутри Калининградской области без других регионов России.",
        ],
        "ko_visual_place_card": [
            "Красивое место Калининградской области с описанием вида, атмосферы, моря, пляжа, кирхи, форта, музея или природной локации.",
            "Один конкретный объект или локация в Калининградской области: чем интересен и почему стоит посмотреть.",
        ],
        "other_region_travel": [
            "Пост о Москве, московских парках, пляжах, маршрутах и прогулках, не связанный с Калининградской областью.",
            "Путешествие по Хайнаню, Турции, Беларуси, Европе, Кавказу, Байкалу, Сочи, Петербургу или другому региону, где Калининградская область не является основной темой.",
            "Рассказ о другом городе или стране, случайно содержащий слово, похожее на калининградский топоним.",
        ],
        "multi_region_roundup": [
            "Подборка разных регионов России: Калининград, Байкал, Дагестан, Сочи, Карелия, Алтай и другие направления одним списком.",
            "Дайджест куда поехать летом по России, где Калининградская область только один пункт среди многих регионов.",
            "Сравнение направлений или список городов из разных регионов и стран.",
        ],
        "news_report": [
            "Новость, официальное сообщение, заявление властей, происшествие, политика, суд, полиция, транспортные планы или исследовательская новость.",
            "Информационная заметка СМИ о факте, находке, решении, субсидиях, запуске парома или событии без личного опыта посещения региона.",
            "Локальная городская новость или федеральная новость, где место лишь контекст события.",
        ],
        "event_announcement": [
            "Анонс мероприятия, афиша, выставка, концерт, регистрация, билеты, расписание, программа, приглашаем прийти.",
            "Пост приглашает на событие или публикует календарь мероприятий, а не рассказывает о впечатлениях от региона.",
        ],
        "ad_or_promo": [
            "Реклама, промокод, скидка, конкурс, розыгрыш, тур, экскурсия, бронирование, покупка билетов, коммерческая услуга.",
            "Промо туристического сервиса или платной поездки, где основной смысл — продать или зарегистрировать.",
        ],
        "low_substance": [
            "Короткий пост без содержания: только фото, эмодзи, хэштег, поздравление или слабая подпись без полезной информации.",
            "Служебное объявление, репост, навигация по каналу, техническая новость или пустой визуальный дамп.",
        ],
    }


def _text_vector_mode() -> str:
    raw = (os.getenv("REGION_TALK_TEXT_VECTOR_MODE") or "").strip().lower()
    if raw:
        return raw
    if not getenv_bool("REGION_TALK_ENABLE_LOCAL_TEXT_EMBEDDINGS", True):
        return "prototype"
    # Avoid surprise model downloads in local unit tests; Kaggle/product runs opt in by environment or /kaggle.
    return "dual_embeddings" if Path("/kaggle/input").exists() else "prototype"


def _prefix_for_embedding_model(model_id: str, text: str, *, query: bool) -> str:
    if model_id.startswith("intfloat/multilingual-e5"):
        return ("query: " if query else "passage: ") + text
    return text


def _ensure_text_embedding_model(model_id: str) -> tuple[Any, Any, Any, str]:
    if model_id in _REGION_TALK_TEXT_MODELS:
        return _REGION_TALK_TEXT_MODELS[model_id]
    try:
        import torch  # type: ignore
        from transformers import AutoModel, AutoTokenizer  # type: ignore
    except Exception:
        if getenv_bool("REGION_TALK_AUTO_INSTALL", True):
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "transformers", "torch", "sentencepiece"])
            import torch  # type: ignore
            from transformers import AutoModel, AutoTokenizer  # type: ignore
        else:
            raise
    tokenizer = AutoTokenizer.from_pretrained(model_id)
    model = AutoModel.from_pretrained(model_id)
    device = "cuda" if getattr(torch, "cuda", None) and torch.cuda.is_available() else "cpu"
    model.to(device)
    model.eval()
    _REGION_TALK_TEXT_MODELS[model_id] = (tokenizer, model, torch, device)
    return _REGION_TALK_TEXT_MODELS[model_id]


def _mean_pool(last_hidden_state: Any, attention_mask: Any, torch: Any) -> Any:
    mask = attention_mask.unsqueeze(-1).expand(last_hidden_state.size()).float()
    return torch.sum(last_hidden_state * mask, 1) / torch.clamp(mask.sum(1), min=1e-9)


def embed_texts_for_model(model_id: str, texts: list[str], *, query: bool = False) -> Any:
    tokenizer, model, torch, device = _ensure_text_embedding_model(model_id)
    prefixed = [_prefix_for_embedding_model(model_id, t, query=query) for t in texts]
    encoded = tokenizer(prefixed, padding=True, truncation=True, max_length=512, return_tensors="pt")
    encoded = {k: v.to(device) for k, v in encoded.items()}
    with torch.no_grad():
        out = model(**encoded)
    pooled = _mean_pool(out.last_hidden_state, encoded["attention_mask"], torch)
    return torch.nn.functional.normalize(pooled, p=2, dim=1).detach().cpu()


def release_text_embedding_model(model_id: str) -> None:
    item = _REGION_TALK_TEXT_MODELS.pop(model_id, None)
    if not item:
        return
    try:
        _tokenizer, model, torch, _device = item
        del model
        if getattr(torch, "cuda", None) and torch.cuda.is_available():
            torch.cuda.empty_cache()
    except Exception:
        pass
    gc.collect()


def _fuse_dual_model_scores(per_model: dict[str, dict[str, float]], bank: dict[str, list[str]], semantic_bank_version: str, bank_hash: str, errors: list[str]) -> dict[str, Any]:
    labels = sorted(bank)
    positive_labels = {"ko_visit_impression", "ko_route_useful", "ko_visual_place_card"}
    negative_labels = set(labels) - positive_labels
    fused = {label: sum(scores.get(label, 0.0) for scores in per_model.values()) / max(1, len(per_model)) for label in labels}
    pos_label, pos_score = max(((l, fused.get(l, 0.0)) for l in positive_labels), key=lambda x: x[1])
    neg_label, neg_score = max(((l, fused.get(l, 0.0)) for l in negative_labels), key=lambda x: x[1])
    model_summaries: dict[str, Any] = {}
    top_labels = []
    for model_id, scores in per_model.items():
        mtop_label, mtop_score = max(scores.items(), key=lambda x: x[1])
        mneg_label, mneg_score = max(((l, scores.get(l, 0.0)) for l in negative_labels), key=lambda x: x[1])
        short = "e5" if model_id.startswith("intfloat/") else "bge_m3" if model_id == "BAAI/bge-m3" else stable_hash(model_id, length=8)
        model_summaries[f"{short}_top_class"] = mtop_label
        model_summaries[f"{short}_top_score"] = round(float(mtop_score), 3)
        model_summaries[f"{short}_negative_class"] = mneg_label
        model_summaries[f"{short}_negative_score"] = round(float(mneg_score), 3)
        top_labels.append(mtop_label)
    reason = "both_models" if len(set(top_labels)) == 1 and len(per_model) == 2 else ("model_disagreement" if len(per_model) == 2 else "single_model_fallback")
    return {
        **model_summaries,
        "text_embedding_runtime": "kaggle_local_dual_text_embeddings_sequential",
        "semantic_bank_version": semantic_bank_version,
        "semantic_bank_hash": bank_hash[:16],
        "text_embedding_model_id": "+".join(per_model.keys()),
        "vector_fusion_reason": reason,
        "vector_top_class": pos_label if pos_score >= neg_score else neg_label,
        "vector_top_score": round(max(pos_score, neg_score), 3),
        "vector_negative_class": neg_label,
        "vector_negative_score": round(float(neg_score), 3),
        "vector_positive_semantic_class": pos_label,
        "vector_positive_semantic_score": round(float(pos_score), 3),
        "vector_margin_positive_vs_negative": round(float(pos_score - neg_score), 3),
        "embedding_error": "; ".join(errors)[:500],
    }


def dual_model_semantic_scores_batch(texts: list[str], report_event: Any | None = None) -> list[dict[str, Any]]:
    mode = _text_vector_mode()
    if mode not in {"dual_embeddings", "embeddings", "real", "dual"}:
        return [{} for _ in texts]
    bank = semantic_bank_v1()
    semantic_bank_version, bank_hash = semantic_bank_version_and_hash(bank)
    flat_labels: list[str] = []
    flat_texts: list[str] = []
    for label, examples in bank.items():
        for ex in examples:
            flat_labels.append(label)
            flat_texts.append(ex)
    per_text: list[dict[str, dict[str, float]]] = [dict() for _ in texts]
    errors: list[str] = []
    require_dual = getenv_bool("REGION_TALK_REQUIRE_DUAL_TEXT_EMBEDDINGS", True)

    def emit(name: str, **payload: Any) -> None:
        if callable(report_event):
            try:
                report_event(name, phase="text_embedding", **payload)
            except Exception:
                pass

    for model_index, model_id in enumerate(TEXT_EMBEDDING_MODELS, start=1):
        started = time.monotonic()
        emit("text_embedding_model_pass_started", status="running", text_embedding_model_id=model_id, text_embedding_models_loaded=model_index - 1, text_embedding_models_required=len(TEXT_EMBEDDING_MODELS), texts_to_score=len(texts))
        try:
            cache_key = model_id + ":" + semantic_bank_version + ":" + bank_hash
            if cache_key not in _REGION_TALK_TEXT_PROTOTYPE_CACHE:
                cached_proto = ydb_load_semantic_bank_embedding(model_id, flat_labels, bank_hash)
                if cached_proto is not None:
                    _REGION_TALK_TEXT_PROTOTYPE_CACHE[cache_key] = cached_proto
                else:
                    _REGION_TALK_TEXT_PROTOTYPE_CACHE[cache_key] = embed_texts_for_model(model_id, flat_texts, query=False)
                    ydb_save_semantic_bank_embedding(model_id, flat_labels, flat_texts, bank_hash, _REGION_TALK_TEXT_PROTOTYPE_CACHE[cache_key])
            proto = _REGION_TALK_TEXT_PROTOTYPE_CACHE[cache_key]
            query_vecs = embed_texts_for_model(model_id, texts, query=True)
            sims_rows = (query_vecs @ proto.T).tolist()
            for i, sims in enumerate(sims_rows):
                scores: dict[str, float] = {}
                for label, sim in zip(flat_labels, sims):
                    scores[label] = max(scores.get(label, -1.0), float(sim))
                per_text[i][model_id] = scores
            emit("text_embedding_model_pass_done", status="running", text_embedding_model_id=model_id, text_embedding_models_loaded=model_index, text_embedding_models_required=len(TEXT_EMBEDDING_MODELS), texts_scored=len(texts), text_embedding_elapsed_seconds=round(time.monotonic() - started, 3))
        except Exception as exc:
            err = f"{model_id}:{type(exc).__name__}:{str(exc)[:180]}"
            errors.append(err)
            emit("text_embedding_model_pass_failed", status="error", text_embedding_model_id=model_id, text_embedding_models_loaded=model_index - 1, text_embedding_models_required=len(TEXT_EMBEDDING_MODELS), text_embedding_error=err, text_embedding_elapsed_seconds=round(time.monotonic() - started, 3))
        finally:
            release_text_embedding_model(model_id)
            emit("text_embedding_model_released", status="running", text_embedding_model_id=model_id)
    if require_dual and any(len(item) != len(TEXT_EMBEDDING_MODELS) for item in per_text):
        message = "; ".join(errors)[:700] or "not all dual text embedding models produced scores"
        emit("text_embedding_dual_requirement_failed", status="error", text_embedding_models_loaded=min((len(item) for item in per_text), default=0), text_embedding_models_required=len(TEXT_EMBEDDING_MODELS), text_embedding_error=message)
        raise RuntimeError("Region Talk dual text embeddings required but not fully available: " + message)
    out: list[dict[str, Any]] = []
    for item in per_text:
        if item:
            out.append(_fuse_dual_model_scores(item, bank, semantic_bank_version, bank_hash, errors))
        else:
            out.append({"embedding_error": "; ".join(errors)[:500], "text_embedding_runtime": "dual_embedding_failed_fallback"})
    return out


def semantic_bank_version_and_hash(bank: dict[str, list[str]]) -> tuple[str, str]:
    payload = json.dumps(bank, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return "semantic_bank_v1", hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _semantic_bank_cache_pk(model_id: str, bank_hash: str) -> str:
    return "semantic_bank_embedding:" + stable_hash(model_id, bank_hash, length=16)


def ydb_load_semantic_bank_embedding(model_id: str, labels: list[str], bank_hash: str) -> Any | None:
    if not getenv_bool("REGION_TALK_YDB_SEMANTIC_BANK_CACHE", True):
        return None
    if (os.getenv("REGION_TALK_STATE_BACKEND") or "").strip().lower() != "ydb" and not (os.getenv("REGION_TALK_YDB_ENDPOINT") or "").strip():
        return None
    try:
        ydb, driver, cfg = ydb_connect(); table_path = ydb_kv_table_path(cfg); pool = ydb.SessionPool(driver); pk = _semantic_bank_cache_pk(model_id, bank_hash)
        def op(session: Any) -> dict[str, Any]:
            ensure_ydb_kv_table(ydb, session, table_path)
            rs = session.transaction(ydb.StaleReadOnly()).execute(f"SELECT payload_json FROM `{table_path}` WHERE pk = '{pk}';", commit_tx=True)
            rows = rs[0].rows if rs else []
            if not rows:
                return {}
            payload = rows[0].payload_json
            return json.loads(payload) if isinstance(payload, str) else dict(payload or {})
        data = pool.retry_operation_sync(op); driver.stop(timeout=5)
        if not data or data.get("model_id") != model_id or data.get("bank_hash") != bank_hash or data.get("labels") != labels:
            return None
        import torch  # type: ignore
        return torch.tensor(data.get("vectors") or [], dtype=torch.float32)
    except Exception:
        return None


def ydb_save_semantic_bank_embedding(model_id: str, labels: list[str], texts: list[str], bank_hash: str, vectors: Any) -> None:
    if not getenv_bool("REGION_TALK_YDB_SEMANTIC_BANK_CACHE", True):
        return
    if (os.getenv("REGION_TALK_STATE_BACKEND") or "").strip().lower() != "ydb" and not (os.getenv("REGION_TALK_YDB_ENDPOINT") or "").strip():
        return
    try:
        ydb, driver, cfg = ydb_connect(); table_path = ydb_kv_table_path(cfg); pool = ydb.SessionPool(driver); now = utc_now_iso()
        version, _ = semantic_bank_version_and_hash(semantic_bank_v1())
        payload = {
            "semantic_bank_version": version, "bank_hash": bank_hash, "model_id": model_id,
            "labels": labels, "texts_hash": hashlib.sha256(json.dumps(texts, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest(),
            "embedding_dim": int(vectors.shape[1]) if hasattr(vectors, "shape") and len(vectors.shape) > 1 else 0,
            "vectors": vectors.tolist() if hasattr(vectors, "tolist") else vectors, "updated_at": now,
        }
        def op(session: Any) -> None:
            ensure_ydb_kv_table(ydb, session, table_path)
            ydb_upsert_json(session, ydb, table_path, _semantic_bank_cache_pk(model_id, bank_hash), "semantic_bank_embedding", payload, now)
        pool.retry_operation_sync(op); driver.stop(timeout=5)
    except Exception:
        return


def dual_model_semantic_scores(text: str) -> dict[str, Any] | None:
    rows = dual_model_semantic_scores_batch([text])
    return rows[0] if rows else None


def _prototype_vector_scores(text: str, ts: dict[str, Any], scope: dict[str, Any], ad_gate: dict[str, Any], substance: dict[str, Any]) -> dict[str, Any]:
    low = (text or "").lower()
    positive = max(float(substance.get("visit_impression_score") or 0), float(substance.get("useful_route_score") or 0), float(substance.get("emotion_observation_score") or 0))
    if scope.get("kaliningrad_oblast_only_scope"):
        positive += 0.25
    if scope.get("matched_place_names"):
        positive += 0.10
    news_event = float(ts.get("newsiness_score") or 0)
    event_terms = ["анонс", "мероприят", "конкурс", "диктант", "регистрац", "заявк", "афиша", "билет", "расписан", "состоится", "программа"]
    news_terms = [
        "уголов", "возбужден", "задерж", "следств", "силов", "прокурат", "суд", "побит рекорд", "температурный рекорд",
        "официально открыт", "мапп", "сообщили", "признан", "штраф", "проверка", "расследован", "помогает компаниям",
    ]
    event_like = any(w in low for w in event_terms)
    news_like = any(w in low for w in news_terms)
    if event_like or news_like:
        news_event += 0.65
    news_class = "event_announcement" if event_like and not news_like else "news_report"
    ad_promo = float(ts.get("ad_score") or 0)
    if bool(ad_gate.get("is_ad_or_promo")):
        ad_promo += 0.50
    roundup = 0.60 if scope.get("external_geo_mentions") else 0.0
    if any(w in low for w in ["подборка", "топ-", "топ ", "куда поехать", "мест россии", "регионов россии", "направлений"]):
        roundup += 0.25
    # Ambiguous lexicon hits such as Светлый/Пионерский/Сокольники must not create a strong KO vector positive by themselves.
    ambiguous_place_penalty = 0.18 if scope.get("matched_place_names") and not scope.get("matched_place_accepted_as_region_evidence") else 0.0
    low_substance = max(0.0, 0.65 - float(substance.get("text_substance_score") or 0) + ambiguous_place_penalty)
    other_region = 0.0
    if scope.get("external_geo_mentions"):
        other_region = max(other_region, 0.75)
    # Full-text external evidence catches cases where hashtags/boilerplate were removed from geo-main-content.
    full_norm = normalize_geo_text(text)
    full_external_regions = [term for term in EXTERNAL_REGION_TERMS if term_in_text(normalize_geo_text(term), full_norm)]
    full_external_countries = [term for term in EXTERNAL_COUNTRY_TERMS if term_in_text(normalize_geo_text(term), full_norm)]
    if full_external_regions or full_external_countries:
        other_region = max(other_region, 0.75)
    if not scope.get("kaliningrad_oblast_only_scope"):
        other_region = max(other_region, 0.55 if scope.get("matched_place_names") else 0.62)
    negative = max(news_event, ad_promo, roundup, low_substance, other_region)
    content_type = "visit_impression_candidate" if float(substance.get("visit_impression_score") or 0) >= 0.20 else ("route_useful_candidate" if float(substance.get("useful_route_score") or 0) >= 0.18 else "single_location_photo_card" if scope.get("matched_place_names") else "low_substance")
    return {
        "positive": positive,
        "news_event": news_event,
        "ad_promo": ad_promo,
        "roundup": roundup,
        "low_substance": low_substance,
        "other_region": other_region,
        "negative": negative,
        "news_class": news_class,
        "content_type": content_type,
    }


def text_vector_gate(text: str, ts: dict[str, Any], scope: dict[str, Any], ad_gate: dict[str, Any], substance: dict[str, Any], embedding_scores: dict[str, Any] | None = None) -> dict[str, Any]:
    """Vector-first text gate with deterministic evidence as fallback only."""
    proto = _prototype_vector_scores(text, ts, scope, ad_gate, substance)
    embed = embedding_scores if embedding_scores is not None else dual_model_semantic_scores(text)
    negative_class = ""
    if embed and embed.get("vector_negative_class"):
        positive = float(embed.get("vector_positive_semantic_score") or 0)
        negative = float(embed.get("vector_negative_score") or 0)
        negative_class = str(embed.get("vector_negative_class") or "")
        content_type = "visit_impression_candidate" if embed.get("vector_positive_semantic_class") == "ko_visit_impression" else ("route_useful_candidate" if embed.get("vector_positive_semantic_class") == "ko_route_useful" else "single_location_photo_card")
        runtime = str(embed.get("text_embedding_runtime") or "kaggle_local_dual_text_embeddings")
        model_id = str(embed.get("text_embedding_model_id") or "+".join(TEXT_EMBEDDING_MODELS))
    else:
        positive = float(proto["positive"])
        negative = float(proto["negative"])
        content_type = str(proto["content_type"])
        runtime = "kaggle_local_prototype_vector_gate"
        model_id = "dual_model_target:intfloat/multilingual-e5-base+BAAI/bge-m3;prototype_fallback_v2"
        if proto["other_region"] >= negative:
            negative_class = "other_region_travel"
        elif proto["roundup"] >= negative:
            negative_class = "multi_region_roundup"
        elif proto["news_event"] >= negative:
            negative_class = str(proto.get("news_class") or "news_report")
        elif proto["ad_promo"] >= negative:
            negative_class = "ad_or_promo"
        elif proto["low_substance"] >= negative:
            negative_class = "low_substance"
    # Deterministic evidence can raise negative confidence for obvious categories, but cannot be the sole positive signal.
    if proto["other_region"] >= 0.55 and negative < proto["other_region"]:
        negative, negative_class = float(proto["other_region"]), "other_region_travel"
    if proto["roundup"] >= 0.60 and negative < proto["roundup"]:
        negative, negative_class = float(proto["roundup"]), "multi_region_roundup"
    if proto["news_event"] >= 0.65 and negative < proto["news_event"]:
        negative, negative_class = float(proto["news_event"]), "event_announcement"
    if proto["ad_promo"] >= 0.65 and negative < proto["ad_promo"]:
        negative, negative_class = float(proto["ad_promo"]), "ad_or_promo"
    margin = round(max(0.0, min(1.0, positive)) - max(0.0, min(1.0, negative)), 3)
    status = "vector_ambiguous_keep_for_ranking"
    reason = ""
    if negative_class == "other_region_travel" and negative >= 0.55 and positive < 0.68:
        status, reason, content_type = "vector_reject_not_kaliningrad_oblast", "other-region semantic class dominates Kaliningrad fit", "other_region_travel"
    elif negative_class == "multi_region_roundup" and negative >= 0.55 and positive < 0.68:
        status, reason, content_type = "vector_reject_multi_region_roundup", "multi-region roundup semantic class dominates", "multi_region_roundup"
    elif negative_class in {"news_report", "event_announcement"} and negative >= 0.55 and positive < 0.68:
        status, reason, content_type = "vector_reject_news_event", "news/event/announcement semantic class dominates", "news_or_event"
    elif negative_class == "ad_or_promo" and negative >= 0.55 and positive < 0.70:
        status, reason, content_type = "vector_reject_ad_promo", "ad/promo semantic class dominates", "ad_or_promo"
    elif negative_class == "low_substance" and negative >= 0.62 and positive < 0.45:
        status, reason, content_type = "vector_reject_low_substance", "low-substance semantic class dominates", "low_substance"
    elif positive >= 0.55 and margin >= -0.12:
        status, reason = "vector_accept_candidate", "positive Kaliningrad visit/route/place semantic signal kept for local image ranking"
    out = {
        "text_embedding_model_id": model_id,
        "text_embedding_runtime": runtime,
        "vector_gate_status": status,
        "vector_content_type": content_type,
        "vector_positive_score": round(max(0.0, min(1.0, positive)), 3),
        "vector_news_event_score": round(max(0.0, min(1.0, max(float(proto.get("news_event") or 0), negative if negative_class in {"news_report", "event_announcement"} else 0))), 3),
        "vector_ad_promo_score": round(max(0.0, min(1.0, max(float(proto.get("ad_promo") or 0), negative if negative_class == "ad_or_promo" else 0))), 3),
        "vector_roundup_score": round(max(0.0, min(1.0, max(float(proto.get("roundup") or 0), negative if negative_class == "multi_region_roundup" else 0))), 3),
        "vector_low_substance_score": round(max(0.0, min(1.0, max(float(proto.get("low_substance") or 0), negative if negative_class == "low_substance" else 0))), 3),
        "vector_other_region_score": round(max(0.0, min(1.0, max(float(proto.get("other_region") or 0), negative if negative_class == "other_region_travel" else 0))), 3),
        "vector_region_scope_score": round(max(0.0, min(1.0, positive)), 3),
        "vector_not_other_region_score": round(max(0.0, min(1.0, 1.0 - max(float(proto.get("other_region") or 0), negative if negative_class == "other_region_travel" else 0))), 3),
        "vector_not_news_score": round(max(0.0, min(1.0, 1.0 - max(float(proto.get("news_event") or 0), negative if negative_class in {"news_report", "event_announcement"} else 0))), 3),
        "vector_not_ad_score": round(max(0.0, min(1.0, 1.0 - max(float(proto.get("ad_promo") or 0), negative if negative_class == "ad_or_promo" else 0))), 3),
        "vector_visit_impression_score": round(float(substance.get("visit_impression_score") or 0), 3),
        "vector_route_useful_score": round(float(substance.get("useful_route_score") or 0), 3),
        "vector_emotion_observation_score": round(float(substance.get("emotion_observation_score") or 0), 3),
        "vector_negative_class": negative_class,
        "vector_negative_score": round(max(0.0, min(1.0, negative)), 3),
        "vector_margin_positive_vs_negative": margin,
        "vector_rejection_reason": reason if status.startswith("vector_reject") else "",
        "vector_gate_confidence": round(abs(margin), 3),
        "needs_llm_final_verify": str(status in {"vector_accept_candidate", "vector_ambiguous_keep_for_ranking"}).lower(),
        "llm_not_called_reason": "wide_funnel_vector_gate_only",
        "llm_stage": "",
        "llm_status": "not_called_until_final_verifier" if status in {"vector_accept_candidate", "vector_ambiguous_keep_for_ranking"} else "not_called_vector_reject",
    }
    if embed:
        out.update(embed)
        out["vector_gate_status"] = status
        out["vector_content_type"] = content_type
        out["vector_positive_score"] = round(max(0.0, min(1.0, positive)), 3)
        out["vector_negative_class"] = negative_class
        out["vector_negative_score"] = round(max(0.0, min(1.0, negative)), 3)
        out["vector_margin_positive_vs_negative"] = margin
        out["needs_llm_final_verify"] = str(status in {"vector_accept_candidate", "vector_ambiguous_keep_for_ranking"}).lower()
        out["llm_status"] = "not_called_until_final_verifier" if status in {"vector_accept_candidate", "vector_ambiguous_keep_for_ranking"} else "not_called_vector_reject"
        out["vector_rejection_reason"] = reason if status.startswith("vector_reject") else ""
    return out


def build_similar_seed_queue(previous_state: dict[str, Any], source_rows: list[dict[str, Any]], candidate_memory_rows: list[dict[str, Any]], run_id: str, run_now: str) -> list[dict[str, Any]]:
    prev = previous_state.get("similar_seed_queue") if isinstance(previous_state.get("similar_seed_queue"), dict) else {}
    updates = _REGION_TALK_TELEGRAM_RUNTIME.get("similar_seed_updates") if isinstance(_REGION_TALK_TELEGRAM_RUNTIME.get("similar_seed_updates"), dict) else {}
    candidate_source_ids = {str(r.get("source_id") or "") for r in candidate_memory_rows if r.get("source_id")}
    out: dict[str, dict[str, Any]] = {str(k): dict(v) for k, v in (prev or {}).items() if isinstance(v, dict)}
    for s in source_rows:
        if str(s.get("platform") or "") != "telegram":
            continue
        if s.get("fetch_status") != "ok" and s.get("telegram_resolve_status") not in {"resolved_network", "resolved_from_private_cache"}:
            continue
        url = str(s.get("canonical_url") or "")
        if not url:
            continue
        key = "similar_seed_" + stable_hash(url)
        old = out.get(key) or {}
        priority = float(s.get("monitor_priority_score") or 0)
        if str(s.get("source_id") or "") in candidate_source_ids:
            priority += 0.35
        if str(s.get("source_kind") or "").lower().find("travel") >= 0 or str(s.get("source_type") or "").lower().find("travel") >= 0:
            priority += 0.10
        upd = updates.get(key) or updates.get(url) or {}
        out[key] = {
            **old,
            "similar_seed_id": key,
            "source_id": s.get("source_id"),
            "canonical_url": url,
            "source_title": s.get("source_title") or s.get("resolved_title"),
            "similar_seed_status": "ready",
            "similar_seed_first_seen_at": old.get("similar_seed_first_seen_at") or run_now,
            "similar_seed_last_seen_at": run_now,
            "similar_seed_last_used_at": upd.get("similar_last_used_at") or old.get("similar_seed_last_used_at", ""),
            "similar_seed_use_count": int(old.get("similar_seed_use_count") or 0) + int(upd.get("similar_use_count_increment") or 0),
            "similar_seed_error_count": int(old.get("similar_seed_error_count") or 0) + int(upd.get("similar_error_count_increment") or 0),
            "similar_seed_last_result_count": upd.get("similar_last_result_count", old.get("similar_seed_last_result_count", "")),
            "similar_seed_last_unique_count": upd.get("similar_last_unique_count", old.get("similar_seed_last_unique_count", "")),
            "similar_last_scanned_at": upd.get("similar_last_scanned_at") or old.get("similar_last_scanned_at", ""),
            "similar_seed_next_allowed_at": upd.get("similar_next_allowed_at") or old.get("similar_seed_next_allowed_at", ""),
            "similar_seed_priority": round(min(1.0, priority), 3),
            "last_run_id": run_id,
            "no_auto_join": "true",
        }
    return sorted(out.values(), key=lambda r: (-float(r.get("similar_seed_priority") or 0), str(r.get("similar_seed_last_used_at") or ""), str(r.get("canonical_url") or "")))


def build_source_delta_scan_sheet(source_rows: list[dict[str, Any]], previous_state: dict[str, Any], run_id: str) -> list[dict[str, Any]]:
    prev_cursors = previous_state.get("source_cursors") if isinstance(previous_state.get("source_cursors"), dict) else {}
    rows: list[dict[str, Any]] = []
    for s in source_rows:
        sid = str(s.get("source_id") or "")
        prev = prev_cursors.get(sid) if isinstance(prev_cursors, dict) else {}
        rows.append({
            "source_id": sid,
            "source_title": s.get("source_title") or s.get("resolved_title"),
            "platform": s.get("platform"),
            "fetch_status": s.get("fetch_status"),
            "history_fetch_mode": s.get("history_fetch_mode") or ("skipped" if str(s.get("fetch_status") or "").startswith("skipped") else ""),
            "is_new_source_this_run": s.get("is_new_source_this_run", "false"),
            "last_history_fetch_run_id_previous": (prev or {}).get("last_history_fetch_run_id", ""),
            "last_history_fetch_run_id": run_id if s.get("fetch_status") == "ok" else "",
            "last_history_fetch_at_previous": (prev or {}).get("last_history_fetch_at", ""),
            "last_seen_post_date_previous": (prev or {}).get("last_seen_post_date", ""),
            "last_seen_post_date": s.get("last_seen_post_date", ""),
            "posts_seen_total_previous": (prev or {}).get("posts_seen_total", ""),
            "posts_scanned_this_run": s.get("posts_scanned", 0),
            "kaliningrad_posts_seen_total_previous": (prev or {}).get("kaliningrad_posts_seen_total", ""),
            "candidate_posts_seen_total_previous": (prev or {}).get("candidate_posts_seen_total", ""),
            "last_candidate_seen_at_previous": (prev or {}).get("last_candidate_seen_at", ""),
            "source_delta_strategy": "since_last_scan_plus_anchor_overlap" if (prev or {}).get("last_history_fetch_at") else "first_probe_or_shallow_backfill",
        })
    return rows or [{"_sheet_note": "no source rows"}]


def build_source_yield_metrics(source_rows: list[dict[str, Any]], posts_by_source: dict[str, list[dict[str, Any]]], new_posts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    scanned = [s for s in source_rows if s.get("fetch_status") == "ok"]
    source_ids = [str(s.get("source_id") or "") for s in scanned]
    source_sets = {
        "sources_scanned_total_this_run": set(source_ids),
        "sources_scanned_new_this_run": {str(s.get("source_id") or "") for s in scanned if str(s.get("is_new_source_this_run") or "").lower() == "true"},
        "sources_with_any_ko_post": {sid for sid, rows in posts_by_source.items() if any(r.get("kaliningrad_oblast_only_scope") for r in rows)},
        "sources_with_fresh_ko_post": {sid for sid, rows in posts_by_source.items() if any(r.get("kaliningrad_oblast_only_scope") and r.get("fresh_enough") for r in rows)},
        "sources_with_non_ad_ko_post": {sid for sid, rows in posts_by_source.items() if any(r.get("kaliningrad_oblast_only_scope") and not r.get("is_ad_or_promo") for r in rows)},
        "sources_with_candidate_memory_post": {str(r.get("source_id") or "") for r in new_posts if r.get("current_stage") in CANDIDATE_MEMORY_STAGES},
        "sources_with_actual_image_candidate": {str(r.get("source_id") or "") for r in new_posts if r.get("image_model_input_type") == "actual_image" and r.get("current_stage") in CANDIDATE_MEMORY_STAGES},
        "sources_with_publication_ready_candidate": {str(r.get("source_id") or "") for r in new_posts if str(r.get("image_publication_ready")) == "true"},
    }
    denom = max(1, len(scanned))
    out = []
    for metric, vals in source_sets.items():
        count = len(vals)
        out.append({"metric": metric, "count": count, "per_1000_scanned_sources": round(count / denom * 1000, 2), "denominator_scanned_sources": len(scanned), "sample_bias_note": "biased by priority/frontier queue; not a random 1000-source sample"})
    return out


def build_all_time_metrics(state_to_write: dict[str, Any], source_frontier_unique: list[dict[str, Any]], candidate_memory_rows: list[dict[str, Any]]) -> dict[str, Any]:
    cursors = state_to_write.get("source_cursors") if isinstance(state_to_write.get("source_cursors"), dict) else {}
    posts_mem = state_to_write.get("posts") if isinstance(state_to_write.get("posts"), dict) else {}
    frontier_by_platform: dict[str, int] = {}
    for r in source_frontier_unique:
        p = str(r.get("platform") or r.get("platform_guess") or "unknown")
        frontier_by_platform[p] = frontier_by_platform.get(p, 0) + 1
    primary = [v for v in cursors.values() if isinstance(v, dict) and (v.get("primary_scan_completed_at") or int(v.get("posts_seen_total") or 0) > 0)]
    delta = [v for v in cursors.values() if isinstance(v, dict) and int(v.get("delta_scan_count_total") or 0) > 0]
    telegram_primary = [v for v in primary if str(v.get("platform") or v.get("source_id") or "").startswith("telegram") or str(v.get("source_id") or "").startswith("src_telegram")]
    vk_primary = [v for v in primary if str(v.get("platform") or "").startswith("vk")]
    scanned_keys = {str(v.get("source_id") or "") for v in primary}
    pending_primary = sum(1 for r in source_frontier_unique if str(r.get("frontier_source_id") or r.get("source_id") or "") not in scanned_keys and str(r.get("frontier_stage") or "") in {"history_due", "probe_due", "unresolved"})
    pending_similar = sum(1 for r in source_frontier_unique if not r.get("similar_last_scanned_at") and str(r.get("platform") or r.get("platform_guess") or "").startswith("telegram"))
    return {
        "sources_primary_scanned_total_all_time": len(primary),
        "telegram_sources_primary_scanned_total_all_time": len(telegram_primary),
        "vk_sources_primary_scanned_total_all_time": len(vk_primary),
        "sources_delta_scanned_total_all_time": len(delta),
        "sources_never_scanned_total": pending_primary,
        "frontier_total_by_platform": json.dumps(frontier_by_platform, ensure_ascii=False, sort_keys=True),
        "frontier_pending_primary_scan_total": pending_primary,
        "frontier_pending_similar_scan_total": pending_similar,
        "posts_memory_total": len(posts_mem),
        "candidates_memory_total": len(candidate_memory_rows),
        "publication_ready_total_all_time": sum(1 for r in candidate_memory_rows if str(r.get("image_publication_ready") or "") == "true"),
    }


def build_report(seeds: list[Seed], source_rows: list[dict[str, Any]], posts: list[dict[str, Any]], run_id: str, output_dir: Path, status: Any | None = None) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    run_now = utc_now_iso()
    lexicon_path = find_place_lexicon_file({})
    lexicon = load_place_lexicon(lexicon_path)
    source_seed_rows = [{**asdict(s), "source_id": s.source_id, "canonical_url": s.canonical_url} for s in seeds]
    candidates: list[dict[str, Any]] = []
    media_rows: list[dict[str, Any]] = []
    dropped: list[dict[str, Any]] = []
    new_posts: list[dict[str, Any]] = []
    increment: list[dict[str, Any]] = []
    discovered_rows: list[dict[str, Any]] = []
    graph_edges: list[dict[str, Any]] = []
    place_match_rows: list[dict[str, Any]] = []
    seed_by_id = {s.source_seed_id: s for s in seeds}
    pre_candidates: list[dict[str, Any]] = []
    llm_model = (os.getenv("REGION_TALK_LLM_MODEL") or "gemini-3.1-flash-lite").strip()
    llm_default_env_var_name = (os.getenv("REGION_TALK_LLM_DEFAULT_ENV_VAR_NAME") or "GOOGLE_API_KEY3").strip()
    llm_limit_snapshot = load_llm_limit_snapshot(llm_model, llm_default_env_var_name)
    llm_calls_used = 0
    llm_supabase_unavailable = llm_limit_snapshot.get("llm_limit_source") == "supabase_required_unavailable"
    previous_state, state_meta = load_region_talk_state(output_dir)
    previous_posts = previous_state.get("posts") if isinstance(previous_state.get("posts"), dict) else {}
    previous_image_queue_rows = _previous_rows_dict(previous_state.get("image_candidate_queue"))
    previous_image_scores_by_url = {str(r.get("post_url") or ""): r for r in previous_image_queue_rows if str(r.get("post_url") or "")}
    previous_discovered = previous_state.get("discovered_sources") if isinstance(previous_state.get("discovered_sources"), dict) else {}
    updated_posts_state: dict[str, Any] = dict(previous_posts)
    updated_discovered_state: dict[str, Any] = dict(previous_discovered)

    def report_event(name: str, **payload: Any) -> None:
        if status is not None and hasattr(status, "event"):
            try:
                status.event(name, run_id=run_id, **payload)
            except Exception:
                pass

    max_posts_to_score = getenv_int("REGION_TALK_MAX_POSTS_TO_SCORE_PER_RUN", 180)
    posts_for_scoring = posts[:max_posts_to_score] if max_posts_to_score > 0 else posts
    runtime_deferred_posts = posts[len(posts_for_scoring):]
    report_event("report_build_started", phase="report", status="running", posts_fetched=len(posts), posts_to_score=len(posts_for_scoring), posts_deferred=len(runtime_deferred_posts), ydb_read_status=state_meta.get("ydb_read_status"))

    semantic_gate_mode = (os.getenv("REGION_TALK_SEMANTIC_GATE_MODE") or "vector_first_final_llm").strip().lower()
    early_llm_enabled = getenv_bool("REGION_TALK_ENABLE_EARLY_LLM", False)
    vector_gates_enabled = getenv_bool("REGION_TALK_ENABLE_VECTOR_GATES", True)
    deterministic_override = getenv_bool("REGION_TALK_ALLOW_DETERMINISTIC_SEMANTIC_GATES", False)
    precomputed_embedding_scores: list[dict[str, Any]] = []
    if vector_gates_enabled and _text_vector_mode() in {"dual_embeddings", "embeddings", "real", "dual"}:
        precomputed_embedding_scores = dual_model_semantic_scores_batch([str(p.get("text") or "") for p in posts_for_scoring], report_event=report_event)

    for idx, p in enumerate(posts_for_scoring, start=1):
        if idx == 1 or idx % max(1, getenv_int("REGION_TALK_VECTOR_HEARTBEAT_EVERY_POSTS", 25)) == 0:
            report_event("vector_scoring_alive", phase="vector_scoring", status="running", progress_label=f"posts {idx}/{len(posts_for_scoring)}", posts_scored=idx, posts_to_score=len(posts_for_scoring), posts_deferred=len(runtime_deferred_posts))
        seed_for_post = seed_by_id.get(str(p.get('source_seed_id') or '')) or (seeds[0] if seeds else None)
        text = p.get("text") or ""
        ts = score_text(text)
        scope = kaliningrad_oblast_only_scope_gate(text, lexicon)
        fresh = freshness_gate(p.get("post_date"))
        ad_gate = ad_promo_gate(text)
        substance = score_substance(text)
        link_rows, edge_rows = discover_links_for_post(p, run_id)
        discovered_rows.extend(link_rows)
        graph_edges.extend(edge_rows)

        cid = "cand_" + stable_hash(p["post_id"], "region-talk-semantic-bank-v1", text[:120])
        mid = "media_" + stable_hash(p["post_id"], p.get("post_url"), "media0")
        drop_gate = ""
        rejection = ""
        visual_stage = "skipped_by_text_gate"
        visual_skip_reason = ""
        image_cost_saved = True
        current_stage = ""
        gate_trace: list[str] = []

        if not fresh["fresh_enough"]:
            drop_gate, rejection = "freshness_gate", "reject_stale_or_missing_date"
            visual_skip_reason = fresh["freshness_reason"]
            gate_trace.append("freshness_gate:reject")
        else:
            gate_trace.append("freshness_gate:pass")

        # LLM-first policy: regex/keyword/lexicon checks below are evidence and recall routing only.
        semantic_evidence_flags: list[str] = []
        if not scope["kaliningrad_oblast_only_scope"]:
            semantic_evidence_flags.append("deterministic_scope_evidence_not_oblast_only")
        if ad_gate["is_ad_or_promo"]:
            semantic_evidence_flags.append("deterministic_ad_promo_evidence")
        if substance["text_substance_score"] < 0.25:
            semantic_evidence_flags.append("deterministic_low_substance_evidence")
        if ts["newsiness_score"] >= 0.45:
            semantic_evidence_flags.append("deterministic_news_evidence")
        if ts["trash_score"] >= 0.35:
            semantic_evidence_flags.append("deterministic_trash_evidence")
        gate_trace.append("kaliningrad_oblast_only_scope_gate:evidence_only")
        gate_trace.append("ad_promo_announcement_gate:evidence_only")
        gate_trace.append("content_substance_visit_impression_gate:evidence_only")
        gate_trace.append("not_news_not_trash_gate:evidence_only")
        embedding_scores = precomputed_embedding_scores[idx - 1] if idx - 1 < len(precomputed_embedding_scores) else None
        vector_gate = text_vector_gate(text, ts, scope, ad_gate, substance, embedding_scores=embedding_scores) if vector_gates_enabled else {
            "text_embedding_model_id": "", "text_embedding_runtime": "disabled", "vector_gate_status": "disabled",
            "vector_content_type": "", "vector_positive_score": "", "vector_news_event_score": "", "vector_ad_promo_score": "",
            "vector_roundup_score": "", "vector_low_substance_score": "", "vector_visit_impression_score": "",
            "vector_route_useful_score": "", "vector_emotion_observation_score": "", "vector_margin_positive_vs_negative": "",
            "vector_rejection_reason": "", "vector_gate_confidence": "", "needs_llm_final_verify": "false",
            "llm_not_called_reason": "vector_gates_disabled", "llm_stage": "", "llm_status": "not_called_vector_disabled",
        }
        gate_trace.append("news_event_vector_gate:" + str(vector_gate.get("vector_gate_status") or "not_run"))

        llm_gate = {
            "llm_gate_status": "not_run", "llm_provider": "google_gemini", "llm_model": llm_model, "llm_default_env_var_name": llm_default_env_var_name, "llm_limit_source": llm_limit_snapshot.get("llm_limit_source", "supabase_google_ai"), "llm_decision": "",
            "whole_post_about_kaliningrad_oblast_score": round(0.65 if scope.get("kaliningrad_oblast_only_scope") else 0.15, 2),
            "kaliningrad_mention_role": "main_subject" if scope.get("kaliningrad_oblast_only_scope") else ("one_item" if scope.get("matched_place_names") else "unclear"),
            "is_digest_or_roundup": str(float(ts.get("newsiness_score") or 0) >= 0.45 or bool(scope.get("external_geo_mentions"))).lower(),
            "is_multi_region_roundup": str(bool(scope.get("external_geo_mentions"))).lower(),
            "is_multi_topic_digest": str(bool(scope.get("external_geo_mentions"))).lower(),
            "is_single_location_card": str(bool(scope.get("matched_place_names")) and not bool(scope.get("external_geo_mentions"))).lower(),
            "is_author_visit_impression": str(float(substance.get("visit_impression_score") or 0) >= 0.25).lower(),
            "is_official_route_material": "false", "is_photo_card_from_subscriber": "false",
            "llm_is_ad_or_promo": "", "llm_is_news_or_trash": "",
            "llm_content_type": str(vector_gate.get("vector_content_type") or ""),
            "content_type": str(vector_gate.get("vector_content_type") or ("single_location_photo_card" if scope.get("matched_place_names") else "low_substance")),
            "visit_evidence_type": "", "has_firsthand_visit_evidence": "", "emotion_or_impression_evidence": "",
            "review_or_opinion_evidence": "", "memorable_detail_evidence": "", "original_photo_evidence": "",
            "llm_reason": "",
            "llm_usage_input_tokens": "", "llm_usage_output_tokens": "", "llm_usage_total_tokens": "",
            "rejection_reason_primary": "", "rejection_reason_details": "",
        }
        if not rejection and str(vector_gate.get("vector_gate_status") or "").startswith("vector_reject"):
            drop_gate = "semantic_vector_gate"
            rejection = str(vector_gate.get("vector_gate_status") or "vector_reject")
            visual_skip_reason = str(vector_gate.get("vector_rejection_reason") or rejection)
            current_stage = "dropped_text_gate"
        elif not rejection and not scope.get("kaliningrad_oblast_only_scope") and str(vector_gate.get("vector_gate_status") or "") != "vector_accept_candidate":
            # Last-resort safety fallback for rows with no Kaliningrad evidence when vector embeddings are unavailable/ambiguous.
            # This is not a positive/negative semantic classifier; product mass decisions should be made by vector_gate above.
            drop_gate = "region_evidence_safety_gate"
            rejection = "vector_reject_not_kaliningrad_oblast"
            visual_skip_reason = str(scope.get("region_scope_reason") or "no Kaliningrad Oblast evidence for vector review")
            current_stage = "dropped_text_gate"
            vector_gate["needs_llm_final_verify"] = "false"
            vector_gate["llm_status"] = "not_called_region_evidence_safety"
            vector_gate["llm_not_called_reason"] = "region_evidence_safety_gate"
            gate_trace.append("region_evidence_safety_gate:reject_no_ko_evidence")
        llm_required = early_llm_enabled and semantic_gate_mode in {"llm_required", "llm", "semantic_required", "early_llm"} and not deterministic_override
        if not rejection and llm_required:
            if llm_supabase_unavailable:
                llm_gate["llm_gate_status"] = "not_run_supabase_limiter_unavailable"
                llm_gate["llm_reason"] = str(llm_limit_snapshot.get("llm_limit_error") or "Supabase limiter is required and unavailable")
                gate_trace.append("llm_semantic_gate:not_run_supabase_limiter_unavailable")
            elif should_send_to_llm(fresh, scope, p, ad_gate=ad_gate, substance=substance, text_score=ts):
                evidence = {
                    "matched_place_names": scope.get("matched_place_names"),
                    "external_geo_mentions": scope.get("external_geo_mentions"),
                    "ad_promo_hits": ad_gate.get("ad_promo_hits"),
                    "semantic_evidence_flags": semantic_evidence_flags,
                    "text_substance_score": substance.get("text_substance_score"),
                    "visit_impression_score": substance.get("visit_impression_score"),
                    "newsiness_score": ts.get("newsiness_score"),
                    "trash_score": ts.get("trash_score"),
                }
                llm_gate = {**llm_gate, **call_region_talk_semantic_llm(p, evidence, model=llm_model, default_env_var_name=llm_default_env_var_name)}
                if llm_gate.get("llm_gate_status") in {"ok", "error", "rate_limited"}:
                    llm_calls_used += 1
                gate_trace.append("llm_semantic_gate:" + str(llm_gate.get("llm_gate_status")))
            else:
                llm_gate["llm_gate_status"] = "not_run_pre_llm_cost_guard"
                gate_trace.append("llm_semantic_gate:not_run_pre_llm_cost_guard")

            if llm_gate.get("llm_gate_status") == "ok":
                decision = str(llm_gate.get("llm_decision") or "needs_review")
                # Guardrail for the observed z3b regression: a single-location Kaliningrad Oblast card is not a multi-region roundup.
                if decision == "reject" and scope.get("kaliningrad_oblast_only_scope") and not scope.get("external_geo_mentions"):
                    reason_l = str(llm_gate.get("llm_reason") or "").lower()
                    if "дайджест" in reason_l or "roundup" in reason_l or "подбор" in reason_l:
                        decision = "accept"
                        llm_gate["llm_decision"] = "accept"
                        llm_gate["content_type"] = llm_gate.get("content_type") or "single_location_photo_card"
                        llm_gate["llm_content_type"] = llm_gate.get("llm_content_type") or "single_location_photo_card"
                        llm_gate["is_single_location_card"] = "true"
                        llm_gate["is_multi_region_roundup"] = "false"
                        llm_gate["llm_reason"] = "Guardrail: single-location Kaliningrad Oblast card kept for research/manual review; not a multi-region roundup. " + str(llm_gate.get("llm_reason") or "")[:400]
                if decision == "reject":
                    drop_gate, rejection = "llm_semantic_gate", "llm_reject"
                    visual_skip_reason = str(llm_gate.get("llm_reason") or "LLM rejected semantic fit")
                    current_stage = "dropped_text_gate"
                elif str(llm_gate.get("llm_content_type") or llm_gate.get("content_type") or "") == "news_or_event" or str(llm_gate.get("llm_is_news_or_trash") or "").lower() == "true":
                    drop_gate, rejection = "llm_semantic_gate", "reject_news_or_event"
                    visual_skip_reason = str(llm_gate.get("llm_reason") or "LLM classified this as news/event rather than a visit/impression post")
                    current_stage = "dropped_text_gate"
                    llm_gate["llm_decision"] = "reject"
                    gate_trace.append("llm_semantic_gate:reject_news_or_event")
                elif decision == "needs_review":
                    drop_gate, rejection = "llm_semantic_gate", "llm_needs_review"
                    visual_skip_reason = str(llm_gate.get("llm_reason") or "LLM requested manual semantic review before image scoring")
                    current_stage = "pre_candidate_needs_llm"
                else:
                    gate_trace.append("llm_semantic_gate:accept")
            elif llm_gate.get("llm_gate_status") in {"rate_limited", "error"}:
                drop_gate, rejection = "llm_semantic_gate", "llm_retry_required"
                visual_skip_reason = str(llm_gate.get("llm_reason") or "LLM failed/rate-limited; retry required before image scoring")
                current_stage = "needs_llm_retry"
            elif llm_gate.get("llm_gate_status") == "not_run_pre_llm_cost_guard":
                precise_reason = classify_pre_llm_reject_reason(fresh, scope, ad_gate, substance, ts)
                drop_gate, rejection = "pre_llm_cost_guard", precise_reason
                visual_skip_reason = "Skipped before LLM to avoid spending Supabase quota on obvious non-region/ad/low-substance rows"
                current_stage = "debug_reject"
            else:
                drop_gate, rejection = "llm_semantic_gate", "semantic_gate_not_run"
                visual_skip_reason = "LLM semantic gate not run; row remains reviewable pre-candidate and image scoring is skipped"
                current_stage = "pre_candidate_needs_llm"
        elif not rejection and deterministic_override:
            current_stage = "pre_candidate_debug_deterministic"
            gate_trace.append("deterministic_semantic_override:pre_candidate_only")
        elif not rejection:
            llm_gate["llm_gate_status"] = "not_run_vector_first"
            llm_gate["llm_reason"] = "Early/broad LLM disabled; vector/local gates and image scoring run before optional final verifier"
            gate_trace.append("early_llm:disabled_vector_first")

        if rejection and current_stage in {"pre_candidate_needs_llm", "needs_llm_retry", "pre_candidate_debug_deterministic"}:
            ms = image_scores_skipped(visual_skip_reason or rejection)
            score = 0.0
        elif rejection:
            ms = image_scores_skipped(visual_skip_reason or rejection)
            score = 0.0
            current_stage = "dropped_text_gate"
        else:
            gate_trace.append("semantic_dual_model_enrichment:vector_gate_passed")
            visual_stage = "scored_after_vector_text_gates_before_final_llm"
            visual_skip_reason = ""
            image_cost_saved = False
            ms = _image_scores_from_ydb_queue(previous_image_scores_by_url.get(str(p.get("post_url") or ""))) or media_scores(bool(p.get("has_media")), ts, p)
            initial_image_status_fields = normalize_image_status(current_stage, ms)
            if initial_image_status_fields.get("image_status") == "needs_actual_image_fetch":
                ms = {**ms, **initial_image_status_fields}
            media_rows.append({"media_id": mid, "candidate_id": cid, "post_url": p.get("post_url"), "image_url_or_local_path": p.get("primary_media_path") or ((p.get("post_url") + "#media") if p.get("has_media") else ""), "thumbnail":"", **ms, **initial_image_status_fields})
            score = candidate_score(ts, ms, seed_for_post) if seed_for_post else 0.0
            image_status_fields = normalize_image_status(current_stage, ms)
            if image_status_fields.get("current_stage") == "image_fetch_retry_needed":
                current_stage, rejection, drop_gate = "image_fetch_retry_needed", "needs_actual_image_fetch", "image_fetch_gate"
                gate_trace.append("image_fetch_gate:metadata_only_pending")
            elif not ms["is_selected_for_publication"]:
                if str(ms.get("image_reviewable")) == "true":
                    current_stage, rejection, drop_gate = "needs_image_review", ms["failure_reason"], "image_postcardness_gate"
                    gate_trace.append("image_postcardness_gate:reviewable")
                else:
                    current_stage, rejection, drop_gate = "good_text_weak_media", ms["failure_reason"], "image_postcardness_gate"
                    gate_trace.append("image_postcardness_gate:weak")
            elif score >= 0.45:
                current_stage = "favorite" if score >= 0.62 else "semantic_candidate"
                gate_trace.append("image_postcardness_gate:pass")
            else:
                current_stage, rejection, drop_gate = "low_substance_but_region_relevant", "candidate_score_low", "candidate_score_gate"
                gate_trace.append("candidate_score_gate:review")

        for m in scope.get("place_matches") or []:
            place_match_rows.append({
                "post_id": p.get("post_id"), "post_url": p.get("post_url"),
                "matched_place_name": m.get("matched_place_name"), "canonical_name": m.get("canonical_name"),
                "place_type": m.get("place_type"), "municipality": m.get("municipality"),
                "priority_tier": m.get("priority_tier"), "alias_used": m.get("alias_used"),
                "match_context": m.get("match_context"), "requires_context": m.get("requires_context"),
                "accepted_as_region_evidence": m.get("accepted_as_region_evidence"),
                "match_context_short": m.get("match_context_short"),
            })

        semantic_visit_fields = infer_visit_semantic_fields(text, llm_gate, substance, p)
        image_status_fields = normalize_image_status(current_stage, ms)
        row = {
            **p, **ts, **scope, **fresh, **ad_gate, **substance, **ms, **semantic_visit_fields, **image_status_fields,
            "candidate_id": cid, "candidate_score": score, "current_stage": current_stage,
            "drop_gate": drop_gate, "rejection_reason": rejection,
            "short_summary": make_summary(text),
            "why_this_is_about_kaliningrad": scope["region_scope_reason"],
            "what_positive": ", ".join(ts.get("positive_hits") or []),
            "what_neutral_or_useful": "route/place/travel context" if substance["text_substance_score"] >= 0.25 else "",
            "what_concern": rejection or "text gates passed; image checked",
            "image_model_report_short": ms["model_short_explanation"],
            "risk_flags": "; ".join([rejection] if rejection else ([] if p.get("rights_policy") != "unknown" else ["rights_unknown"])),
            "suggested_action": (image_status_fields.get("next_action") or ("manual_review" if current_stage in {"favorite","semantic_candidate", "pre_candidate_needs_llm", "needs_llm_retry", "needs_image_review", "image_fetch_retry_needed", "low_substance_but_region_relevant", "pre_candidate_debug_deterministic"} else "reject")),
            "manual_decision":"", "reviewer_comment":"",
            "region_scope_gate": "pass" if scope["kaliningrad_oblast_only_scope"] else "fail",
            "visual_scoring_stage": visual_stage,
            "visual_scoring_skip_reason": visual_skip_reason,
            "image_scoring_cost_saved": str(image_cost_saved).lower(),
            "image_scoring_skipped": str(image_cost_saved).lower(),
            "discovery_edges_count": len(edge_rows),
            "gate_order_trace": " → ".join(gate_trace),
            "semantic_evidence_flags": "; ".join(semantic_evidence_flags),
            **vector_gate,
            **llm_gate,
            "rejection_reason_primary": rejection or "",
            "rejection_reason_details": visual_skip_reason or str(llm_gate.get("llm_reason") or ""),
            "semantic_gate_mode": semantic_gate_mode,
            "deterministic_semantic_gate_override": str(deterministic_override).lower(),
            "semantic_enrichment_stage": "final_llm_verifier_pending" if str(vector_gate.get("needs_llm_final_verify")) == "true" and current_stage in {"semantic_candidate", "favorite", "needs_image_review", "image_fetch_retry_needed", "good_text_weak_media", "low_substance_but_region_relevant"} else ("dual_model_vector_enrichment_pending" if current_stage in {"semantic_candidate", "favorite", "needs_image_review", "low_substance_but_region_relevant"} else "skipped_by_text_gate"),
        }
        row.pop("place_matches", None)
        new_posts.append(row)
        write_region_talk_online_post_item(row, run_id=run_id, stage="post_scored", status=current_stage)
        previous_post = previous_posts.get(str(p["post_id"])) if isinstance(previous_posts, dict) else None
        previous_stage = str((previous_post or {}).get("current_stage") or "")
        previous_score = (previous_post or {}).get("candidate_score_current", "")
        previous_media = (previous_post or {}).get("media_score_current", "")
        seen_count = int((previous_post or {}).get("seen_run_count") or 0) + 1
        first_seen_run = str((previous_post or {}).get("first_seen_run_id") or run_id)
        changed = (not previous_post) or previous_stage != current_stage or str(previous_score) != str(score) or str(previous_media) != str(ms["overall_media_score"])
        score_delta = ""
        media_delta = ""
        try:
            score_delta = round(float(score) - float(previous_score), 3) if previous_score != "" else ""
        except Exception:
            score_delta = ""
        try:
            media_delta = round(float(ms["overall_media_score"]) - float(previous_media), 3) if previous_media != "" else ""
        except Exception:
            media_delta = ""
        increment.append({"entity_type":"post", "entity_id":p["post_id"], "source_title":p.get("source_title"), "post_url":p.get("post_url"), "first_seen_run_id":first_seen_run, "previous_run_id":str((previous_post or {}).get("last_seen_run_id") or ""), "current_run_id":run_id, "first_seen_at":str((previous_post or {}).get("first_seen_at") or run_now), "last_seen_at":run_now, "seen_run_count":seen_count, "previous_stage":previous_stage, "current_stage":current_stage, "stage_transition":("new→"+current_stage if not previous_post else previous_stage+"→"+current_stage), "new_this_run":"yes" if not previous_post else "no", "changed_this_run":str(changed).lower(), "change_reason":rejection or ("first_seen" if not previous_post else ("stage_or_score_changed" if changed else "unchanged")), "candidate_score_previous":previous_score, "candidate_score_current":score, "candidate_score_delta":score_delta, "media_score_previous":previous_media, "media_score_current":ms["overall_media_score"], "media_score_delta":media_delta, "manual_review_status":"unreviewed", "next_action":row["suggested_action"]})
        updated_posts_state[str(p["post_id"])] = {
            "first_seen_run_id": first_seen_run,
            "first_seen_at": str((previous_post or {}).get("first_seen_at") or run_now),
            "last_seen_run_id": run_id,
            "last_seen_at": run_now,
            "seen_run_count": seen_count,
            "current_stage": current_stage,
            "candidate_score_current": score,
            "media_score_current": ms["overall_media_score"],
            "post_url": p.get("post_url"),
            "source_id": p.get("source_id"),
            "source_title": p.get("source_title"),
            "platform_post_key": p.get("platform_post_key"),
            "post_date": p.get("post_date"),
        }
        if current_stage in {"favorite", "semantic_candidate"}:
            candidates.append(row)
        elif current_stage in {"pre_candidate_needs_llm", "needs_llm_retry", "needs_image_review", "image_fetch_retry_needed", "low_substance_but_region_relevant", "pre_candidate_debug_deterministic"}:
            pre_candidates.append(row)
        else:
            dropped.append(row)

    similar_channel_rows = list(_REGION_TALK_TELEGRAM_RUNTIME.get("similar_rows") or [])
    similar_channel_edges = list(_REGION_TALK_TELEGRAM_RUNTIME.get("similar_edges") or [])
    keyword_discovery_rows = list(_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_rows") or [])
    keyword_discovery_edges = list(_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_edges") or [])
    public_blogger_rows = load_public_blogger_links({})
    discovered_rows.extend(similar_channel_rows)
    discovered_rows.extend(keyword_discovery_rows)
    discovered_rows.extend(public_blogger_rows)
    graph_edges.extend(similar_channel_edges)
    graph_edges.extend(keyword_discovery_edges)

    review_queue = sorted(candidates + pre_candidates, key=lambda r: (r.get("current_stage") not in {"favorite", "semantic_candidate"}, -float(r.get("candidate_score") or 0), int(r.get("post_age_days") or 9999)))
    for i, r in enumerate(review_queue, start=1):
        r["rank"] = i
        r["status_badge"] = "READY" if r["current_stage"] == "favorite" else "NEEDS_REVIEW"
        r["new_or_seen"] = "NEW"
    favorites = [r for r in review_queue if r.get("current_stage") == "favorite"]
    final_verifier_llm_calls = 0
    final_verifier_queue_rows = [
        r for r in review_queue
        if getenv_bool("REGION_TALK_ENABLE_FINAL_LLM_VERIFIER", False)
        and str(r.get("needs_llm_final_verify") or "").lower() == "true"
        and r.get("kaliningrad_oblast_only_scope")
        and str(r.get("kaliningrad_mention_role") or "main_subject") in {"", "main_subject"}
        and not r.get("is_ad_or_promo")
        and r.get("current_stage") in CANDIDATE_MEMORY_STAGES
    ]
    final_verifier_limit = getenv_int("REGION_TALK_MAX_LLM_FINAL_VERIFY", 10)
    for r in final_verifier_queue_rows[:max(0, final_verifier_limit)]:
        evidence = {
            "stage": "final_publication_verifier",
            "vector_gate_status": r.get("vector_gate_status"),
            "image_status": r.get("image_status"),
            "image_model_input_type": r.get("image_model_input_type"),
            "matched_place_names": r.get("matched_place_names"),
            "content_type": r.get("content_type") or r.get("vector_content_type"),
        }
        result = call_region_talk_semantic_llm(r, evidence, model=llm_model, default_env_var_name=llm_default_env_var_name)
        if result.get("llm_gate_status") in {"ok", "error", "rate_limited"}:
            final_verifier_llm_calls += 1
        r["llm_stage"] = "final_publication_verifier"
        r["final_verifier_status"] = result.get("llm_gate_status")
        r["final_verifier_decision"] = result.get("llm_decision", "")
        r["final_verifier_reason"] = result.get("llm_reason", "")
        r["final_verifier_model"] = result.get("llm_model", llm_model)
        r["llm_status"] = "final_verified" if result.get("llm_gate_status") == "ok" else "final_verifier_retry"
        if result.get("llm_gate_status") == "ok" and result.get("llm_decision") == "reject":
            r["current_stage"] = "dropped_final_verifier"
            r["drop_gate"] = "final_llm_verifier"
            r["rejection_reason"] = "llm_reject_final"
            r["llm_reject_final_reason"] = result.get("llm_reason", "")
            dropped.append(r)
    if final_verifier_llm_calls:
        candidates = [r for r in candidates if r.get("current_stage") in {"favorite", "semantic_candidate"}]
        pre_candidates = [r for r in pre_candidates if r.get("current_stage") in {"pre_candidate_needs_llm", "needs_llm_retry", "needs_image_review", "image_fetch_retry_needed", "low_substance_but_region_relevant", "pre_candidate_debug_deterministic", "good_text_weak_media"}]
        review_queue = sorted(candidates + pre_candidates, key=lambda r: (r.get("current_stage") not in {"favorite", "semantic_candidate"}, -float(r.get("candidate_score") or 0), int(r.get("post_age_days") or 9999)))
        for i, r in enumerate(review_queue, start=1):
            r["rank"] = i
            r["status_badge"] = "READY" if r["current_stage"] == "favorite" else "NEEDS_REVIEW"
            r["new_or_seen"] = "NEW"
        favorites = [r for r in review_queue if r.get("current_stage") == "favorite"]

    # Source profile probe MVP: derive probe metrics from fetched sample; sources are not automatically monitored from discovery.
    posts_by_source: dict[str, list[dict[str, Any]]] = {}
    for row in new_posts:
        posts_by_source.setdefault(str(row.get("source_id") or ""), []).append(row)
    enriched_source_rows: list[dict[str, Any]] = []
    for srow in source_rows:
        sid = str(srow.get("source_id") or "")
        sampled = posts_by_source.get(sid, [])
        n = len(sampled)
        kal_hits = sum(1 for r in sampled if r.get("kaliningrad_oblast_only_scope"))
        ad_hits = sum(1 for r in sampled if r.get("is_ad_or_promo"))
        news_hits = sum(1 for r in sampled if float(r.get("newsiness_score") or 0) >= 0.45)
        trash_hits = sum(1 for r in sampled if float(r.get("trash_score") or 0) >= 0.35)
        original_media = sum(1 for r in sampled if r.get("has_media"))
        link_richness = sum(int(r.get("discovery_edges_count") or 0) for r in sampled)
        monitor_score = round((kal_hits / max(1, n))*0.35 + (original_media/max(1,n))*0.20 + min(1, link_richness/10)*0.15 + max(0, 1-ad_hits/max(1,n))*0.15 + max(0, 1-news_hits/max(1,n))*0.10, 3) if n else 0.0
        status = "probed" if n else ("profile_resolved" if srow.get("fetch_status") == "ok" else "blocked")
        source_class = classify_source_profile(srow, sampled)
        enriched_source_rows.append({**srow,
            "source_probe_status": status, "sampled_post_count": n, "kaliningrad_hit_count": kal_hits,
            "russia_travel_score": round(kal_hits/max(1,n),3), "authorial_voice_score": 0.5,
            "original_media_score": round(original_media/max(1,n),3), "ad_ratio": round(ad_hits/max(1,n),3),
            "news_ratio": round(news_hits/max(1,n),3), "trash_ratio": round(trash_hits/max(1,n),3),
            "image_prevalence": round(original_media/max(1,n),3), "link_richness_score": round(min(1, link_richness/10),3),
            "forwarded_origin_richness_score": round(sum(1 for r in sampled if r.get("is_forwarded_or_repost"))/max(1,n),3),
            "source_graph_value_score": round(min(1, link_richness/10),3), "monitor_priority_score": monitor_score,
            **source_class,
            "source_probe_reason": "derived from recent fetched posts; no automatic monitoring without manual/probe acceptance",
        })
    source_rows = enriched_source_rows
    source_class_by_id = {str(s.get("source_id") or ""): {k: s.get(k) for k in ["source_geo_class", "source_topic_class", "ko_mention_ratio_recent", "travel_blogger_score", "personal_voice_score", "nonlocal_value_score", "source_priority_reason"]} for s in source_rows}
    for r in new_posts:
        r.update({k: v for k, v in source_class_by_id.get(str(r.get("source_id") or ""), {}).items() if v != "" and v is not None})

    for drow in discovered_rows:
        platform_norm = normalize_source_platform(str(drow.get("platform") or drow.get("platform_guess") or ""), str(drow.get("canonical_url") or drow.get("normalized_url") or drow.get("raw_url") or ""))
        canonical_url = canonical_source_url(platform_norm, str(drow.get("username_or_handle") or drow.get("recommended_username") or ""), str(drow.get("canonical_url") or drow.get("normalized_url") or ""))
        canonical_key = str(drow.get("canonical_source_key") or canonical_source_key(platform_norm, str(drow.get("username_or_handle") or drow.get("recommended_username") or ""), canonical_url))
        did = str(drow.get("source_candidate_id") or "src_cand_" + stable_hash(canonical_key or canonical_url))
        prev = updated_discovered_state.get(did) or {}
        updated_discovered_state[did] = {
            **prev,
            "source_candidate_id": did,
            "normalized_url": canonical_url or drow.get("normalized_url"),
            "canonical_url": canonical_url or drow.get("canonical_url") or drow.get("normalized_url"),
            "canonical_source_key": canonical_key,
            "platform_guess": platform_norm or drow.get("platform_guess"),
            "title_guess": drow.get("recommended_title") or drow.get("title_guess") or drow.get("to_source_title") or prev.get("title_guess", ""),
            "edge_types_all": "; ".join(sorted(set(str(x) for x in ((str(prev.get("edge_types_all") or "").split("; ") if prev.get("edge_types_all") else []) + [str(drow.get("edge_type") or "")]) if x))),
            "discovery_types": "; ".join(sorted(set(str(x) for x in ((str(prev.get("discovery_types") or "").split("; ") if prev.get("discovery_types") else []) + [str(drow.get("discovery_type") or "")]) if x))),
            "first_seen_run_id": prev.get("first_seen_run_id") or run_id,
            "last_seen_run_id": run_id,
            "seen_run_count": int(prev.get("seen_run_count") or 0) + 1,
            "candidate_source_status": drow.get("candidate_source_status") or prev.get("candidate_source_status") or "source_frontier",
            "frontier_priority": max(float(prev.get("frontier_priority") or 0), source_candidate_score(drow)),
        }
    previous_source_cursors = previous_state.get("source_cursors") if isinstance(previous_state.get("source_cursors"), dict) else {}
    source_cursors: dict[str, Any] = {str(k): dict(v) for k, v in (previous_source_cursors or {}).items() if isinstance(v, dict)}
    for srow in source_rows:
        sid = str(srow.get("source_id") or "")
        sampled = posts_by_source.get(sid, [])
        newest = max([str(r.get("post_date") or "") for r in sampled] or [""])
        prev_cursor = source_cursors.get(sid, {})
        scanned_ok = srow.get("fetch_status") == "ok"
        is_new_source = str(srow.get("is_new_source_this_run") or "").lower() == "true" or not prev_cursor.get("primary_scan_completed_at")
        source_cursors[sid] = {
            **prev_cursor,
            "source_id": sid,
            "source_title": srow.get("source_title") or srow.get("resolved_title"),
            "platform": srow.get("platform"),
            "canonical_url": srow.get("canonical_url"),
            "canonical_source_key": srow.get("canonical_source_key") or canonical_source_key(str(srow.get("platform") or ""), str(srow.get("handle") or ""), str(srow.get("canonical_url") or "")),
            "handle": srow.get("handle"),
            "fetch_status": srow.get("fetch_status"),
            "last_history_fetch_run_id": run_id if scanned_ok else prev_cursor.get("last_history_fetch_run_id", ""),
            "last_history_fetch_at": run_now if scanned_ok else prev_cursor.get("last_history_fetch_at", ""),
            "last_seen_post_key": max([str(r.get("platform_post_key") or "") for r in sampled] or [str(prev_cursor.get("last_seen_post_key") or "")]),
            "last_seen_post_date": newest or prev_cursor.get("last_seen_post_date", ""),
            "last_seen_post_published_at": newest or prev_cursor.get("last_seen_post_published_at", ""),
            "posts_seen_total": int(prev_cursor.get("posts_seen_total") or 0) + int(srow.get("posts_scanned") or 0),
            "kaliningrad_posts_seen_total": int(prev_cursor.get("kaliningrad_posts_seen_total") or 0) + sum(1 for r in sampled if r.get("kaliningrad_oblast_only_scope")),
            "candidate_posts_seen_total": int(prev_cursor.get("candidate_posts_seen_total") or 0) + sum(1 for r in sampled if r.get("current_stage") in CANDIDATE_MEMORY_STAGES),
            "last_candidate_seen_at": max([str(r.get("post_date") or "") for r in sampled if r.get("current_stage") in CANDIDATE_MEMORY_STAGES] or [str(prev_cursor.get("last_candidate_seen_at") or "")]),
            "primary_scan_completed_at": (prev_cursor.get("primary_scan_completed_at") or (run_now if scanned_ok and is_new_source else "")),
            "primary_scan_post_count": int(prev_cursor.get("primary_scan_post_count") or 0) or (int(srow.get("posts_scanned") or 0) if scanned_ok and is_new_source else 0),
            "last_successful_delta_scan_at": run_now if scanned_ok and not is_new_source else prev_cursor.get("last_successful_delta_scan_at", ""),
            "delta_scan_count_total": int(prev_cursor.get("delta_scan_count_total") or 0) + (1 if scanned_ok and not is_new_source else 0),
            "source_scan_tier": srow.get("source_scan_tier") or ("high_daily" if float(srow.get("monitor_priority_score") or 0) >= 0.5 else "medium_weekly" if sampled else "keyword_probe"),
            "next_history_scan_at": iso_after_seconds(86400 if float(srow.get("monitor_priority_score") or 0) >= 0.5 else 7 * 86400),
            "next_delta_scan_at": iso_after_seconds(86400 if float(srow.get("monitor_priority_score") or 0) >= 0.5 else 7 * 86400),
            "delta_scanned_this_run": str(bool(scanned_ok and not is_new_source)).lower(),
            "delta_scan_window_days": getenv_int("REGION_TALK_DELTA_SCAN_WINDOW_DAYS", 14),
            "delta_overlap_posts": getenv_int("REGION_TALK_DELTA_OVERLAP_POSTS", 10),
            "last_run_id": run_id,
            "cursor_strategy": "fresh_first_min_post_date_then_anchor_search",
        }
    source_frontier_unique = build_source_frontier_unique(discovered_rows, updated_discovered_state, run_id)
    active_frontier_tg_vk = [
        r for r in source_frontier_unique
        if normalize_source_platform(str(r.get("platform") or r.get("platform_guess") or ""), str(r.get("canonical_url") or r.get("normalized_url") or "")) in {"telegram", "vk"}
        and target_source_url_reason(
            normalize_source_platform(str(r.get("platform") or r.get("platform_guess") or ""), str(r.get("canonical_url") or r.get("normalized_url") or "")),
            str(r.get("canonical_url") or r.get("normalized_url") or ""),
        ) == "ok"
        and str(r.get("frontier_stage") or "") not in {"unsupported", "inactive_low_quality"}
    ]
    external_links_quarantine = [
        {
            **r,
            "quarantine_reason": r.get("external_quarantine_reason") or target_source_url_reason(
                normalize_source_platform(str(r.get("platform") or r.get("platform_guess") or ""), str(r.get("canonical_url") or r.get("normalized_url") or "")),
                str(r.get("canonical_url") or r.get("normalized_url") or ""),
            ),
            "active_scan_allowed": "false",
        }
        for r in source_frontier_unique
        if target_source_url_reason(
            normalize_source_platform(str(r.get("platform") or r.get("platform_guess") or ""), str(r.get("canonical_url") or r.get("normalized_url") or "")),
            str(r.get("canonical_url") or r.get("normalized_url") or ""),
        ) != "ok"
    ]
    candidate_memory_rows, previous_candidates_not_refetched, candidate_deltas = build_candidate_memory(previous_state, new_posts, source_rows, run_id, run_now)
    online_candidate_items_written = write_region_talk_online_candidate_items(candidate_memory_rows, run_id=run_id, stage="candidate_memory_built")
    source_frontier_queue_next = build_source_frontier_queue_next(source_frontier_unique, source_rows, candidate_memory_rows, run_id)
    candidate_memory_active_rows = [r for r in candidate_memory_rows if str(r.get("current_lifecycle_status") or "") in ACTIVE_CANDIDATE_MEMORY_STATUSES and str(r.get("manual_decision") or "") != "reject"]

    def memory_product_vector_ok(row: dict[str, Any]) -> bool:
        existing_status = str(row.get("vector_gate_status") or "")
        if existing_status.startswith("vector_reject"):
            row["memory_product_exclusion_reason"] = existing_status
            return False
        sample_text = str(row.get("short_summary") or row.get("text_excerpt") or row.get("post_url") or "")
        if not sample_text.strip():
            row["memory_product_exclusion_reason"] = "missing_text_for_memory_vector_recheck"
            return False
        scope2 = kaliningrad_oblast_only_scope_gate(sample_text, lexicon)
        gate2 = text_vector_gate(sample_text, score_text(sample_text), scope2, ad_promo_gate(sample_text), score_substance(sample_text))
        row["memory_vector_recheck_status"] = gate2.get("vector_gate_status")
        row["memory_vector_recheck_negative_class"] = gate2.get("vector_negative_class")
        row["memory_vector_recheck_reason"] = gate2.get("vector_rejection_reason")
        if str(gate2.get("vector_gate_status") or "").startswith("vector_reject"):
            row["memory_product_exclusion_reason"] = str(gate2.get("vector_gate_status") or "")
            return False
        topic = str(row.get("source_topic_class") or "")
        if topic in {"local_news", "federal_media", "ads_tours"} and str(gate2.get("vector_content_type") or "") not in {"visit_impression_candidate", "route_useful_candidate"}:
            row["memory_product_exclusion_reason"] = "low_priority_news_or_ads_source_without_positive_vector_fit"
            return False
        row["memory_product_exclusion_reason"] = ""
        return True

    candidate_memory_product_rows = [r for r in candidate_memory_active_rows if memory_product_vector_ok(r)]
    candidate_memory_top = sorted(candidate_memory_product_rows, key=lambda r: (
        str(r.get("not_refetched_this_run") or "") == "true",
        str(r.get("source_geo_class") or "") == "kaliningrad_local",
        str(r.get("source_topic_class") or "") in {"local_news", "federal_media", "ads_tours"},
        -float(r.get("publication_story_score") or 0),
        -float(r.get("nonlocal_value_score") or 0),
        -float(r.get("best_candidate_score_ever") or 0),
        -float(r.get("best_media_score_ever") or 0),
    ))[:100]
    similar_seed_queue = build_similar_seed_queue(previous_state, source_rows, candidate_memory_rows, run_id, run_now)
    source_delta_scan_sheet = build_source_delta_scan_sheet(source_rows, previous_state, run_id)
    source_yield_metrics_sheet = build_source_yield_metrics(source_rows, posts_by_source, new_posts)
    early_keyword_post_hit_rows = list(_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_post_hit_rows") or [])
    unified_source_queue_sheet, unified_source_queue_metrics = build_unified_source_queue(
        previous_state, seeds, source_rows, source_frontier_unique, public_blogger_rows,
        keyword_discovery_rows, early_keyword_post_hit_rows, posts_by_source, run_id, run_now,
    )
    online_source_queue_items_written = write_region_talk_online_queue_items(
        [r for r in unified_source_queue_sheet if isinstance(r, dict) and not r.get("_sheet_note")],
        kind="source_queue_item",
        id_fields=["canonical_source_key", "source_queue_id", "source_url", "canonical_url"],
        fields=SOURCE_QUEUE_STATE_FIELDS,
        run_id=run_id,
        stage="unified_source_queue_built",
    )
    write_region_talk_online_queue_items(
        [r for r in unified_source_queue_sheet if isinstance(r, dict) and not r.get("_sheet_note")],
        kind="source_status_item",
        id_fields=["canonical_source_key", "source_queue_id", "source_url", "canonical_url"],
        fields=SOURCE_QUEUE_STATE_FIELDS,
        run_id=run_id,
        stage="unified_source_queue_built",
    )
    image_candidate_queue_sheet, image_driven_top_sheet, image_queue_metrics = build_image_candidate_queue(
        previous_state, new_posts, candidate_memory_rows, media_rows, run_id, run_now,
    )
    online_image_queue_items_written = write_region_talk_online_queue_items(
        [r for r in image_candidate_queue_sheet if isinstance(r, dict) and not r.get("_sheet_note")],
        kind="image_queue_item",
        id_fields=["image_queue_id", "post_url"],
        fields=IMAGE_QUEUE_STATE_FIELDS,
        run_id=run_id,
        stage="image_candidate_queue_built",
    )
    previous_publication_rows = _previous_rows_dict(previous_state.get("publication_candidate_queue"))
    previous_publication_goal = previous_state.get("publication_goal") if isinstance(previous_state.get("publication_goal"), dict) else {}
    publication_candidate_rows, publication_goal = build_publication_candidate_queue(
        candidate_memory_product_rows,
        image_candidate_queue_sheet,
        previous_publication_rows,
        previous_publication_goal,
        run_id=run_id,
        run_now=run_now,
        llm_model=llm_model,
        llm_default_env_var_name=llm_default_env_var_name,
        current_run_preverified_llm_calls=final_verifier_llm_calls,
        report_event=report_event,
    )
    online_publication_candidate_items_written = write_region_talk_online_queue_items(
        [r for r in publication_candidate_rows if isinstance(r, dict) and not r.get("_sheet_note")],
        kind="publication_candidate_item",
        id_fields=["publication_candidate_id", "post_url"],
        fields=PUBLICATION_CANDIDATE_STATE_FIELDS,
        run_id=run_id,
        stage="publication_queue_built",
    )
    state_to_write = {
        "run_id": run_id,
        "state_schema_version": "region-talk-state-v2",
        "updated_at": run_now,
        "posts": updated_posts_state,
        "discovered_sources": updated_discovered_state,
        "source_cursors": source_cursors,
        "source_frontier_unique": {str(r.get("source_candidate_id")): r for r in source_frontier_unique if r.get("source_candidate_id")},
        "region_talk_sources": {str(r.get("canonical_source_key") or r.get("source_candidate_id")): r for r in source_frontier_unique if r.get("source_candidate_id")},
        "source_frontier_queue_next": {str(r.get("canonical_url") or r.get("queue_rank")): r for r in source_frontier_queue_next},
        "unified_source_queue": {str(r.get("canonical_source_key")): r for r in unified_source_queue_sheet if r.get("canonical_source_key")},
        "unified_source_queue_cursor_position": unified_source_queue_metrics.get("source_queue_cursor_position", 0),
        "unified_source_queue_cursor_key": unified_source_queue_metrics.get("source_queue_cursor_key", ""),
        "similar_seed_queue": {str(r.get("similar_seed_id") or r.get("canonical_url")): r for r in similar_seed_queue},
        "image_candidate_queue": {str(r.get("image_queue_id") or r.get("post_url")): r for r in image_candidate_queue_sheet if r.get("image_queue_id") or r.get("post_url")},
        "image_candidate_queue_cursor_position": image_queue_metrics.get("image_queue_cursor_position", 0),
        "image_candidate_queue_cursor_key": next((str(r.get("image_queue_id") or "") for r in image_candidate_queue_sheet if int(r.get("image_queue_order") or 0) == int(image_queue_metrics.get("image_queue_cursor_position") or 0)), ""),
        "candidate_memory": {str(r.get("candidate_memory_id")): r for r in candidate_memory_rows if r.get("candidate_memory_id")},
        "telegram_entity_cache": _REGION_TALK_TELEGRAM_RUNTIME.get("entity_cache") or previous_state.get("telegram_entity_cache") or {},
        "telegram_cooldowns": _REGION_TALK_TELEGRAM_RUNTIME.get("cooldowns") or previous_state.get("telegram_cooldowns") or {},
    }
    all_time_metrics = build_all_time_metrics(state_to_write, source_frontier_unique, candidate_memory_rows)
    state_to_write["all_time_metrics"] = all_time_metrics
    state_write_meta = save_region_talk_state(output_dir, state_to_write)

    stage_counts = {stage: sum(1 for r in new_posts if r.get("current_stage") == stage) for stage in sorted({str(r.get("current_stage") or "") for r in new_posts})}
    fresh_rows = [r for r in new_posts if r.get("fresh_enough")]
    fresh_with_place_evidence = [r for r in fresh_rows if r.get("matched_place_names")]
    llm_reviewed_rows = [r for r in new_posts if r.get("llm_gate_status") == "ok"]
    llm_accepted_rows = [r for r in llm_reviewed_rows if r.get("llm_decision") == "accept"]
    vector_scored_rows = [r for r in new_posts if r.get("vector_gate_status")]
    vector_rejected_rows = [r for r in vector_scored_rows if str(r.get("vector_gate_status") or "").startswith("vector_reject")]
    actual_image_scored_before_llm_count = sum(1 for r in media_rows if r.get("image_model_input_type") == "actual_image")
    final_verify_queue_count_estimate = len(final_verifier_queue_rows) if 'final_verifier_queue_rows' in locals() else sum(1 for r in new_posts if str(r.get("needs_llm_final_verify") or "") == "true" and r.get("current_stage") in {"favorite", "semantic_candidate", "needs_image_review", "image_fetch_retry_needed", "good_text_weak_media", "low_substance_but_region_relevant"})
    early_llm_enabled_summary = getenv_bool("REGION_TALK_ENABLE_EARLY_LLM", False)
    max_llm_final_verify = getenv_int("REGION_TALK_MAX_LLM_FINAL_VERIFY", 10)
    reviewable_rows = [r for r in review_queue if r.get("current_stage") in {"pre_candidate_needs_llm", "needs_llm_retry", "needs_image_review", "low_substance_but_region_relevant", "semantic_candidate", "favorite", "pre_candidate_debug_deterministic"}]
    git_info = git_provenance()
    tg_obs = _REGION_TALK_TELEGRAM_RUNTIME.get("observability") or {"telegram_phase_status": "not_configured"}
    tg_similar_metrics = {
        "telegram_similar_channels_enabled": str(getenv_bool("REGION_TALK_TG_SIMILAR_ENABLED", getenv_bool("REGION_TALK_DISCOVERY_ENABLE_TELEGRAM_SIMILAR", True))).lower(),
        "telegram_similar_channels_status": _REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_status", "not_run"),
        "telegram_similar_channels_seed_count": _REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_seed_count", 0),
        "telegram_similar_channels_raw_count": _REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_raw_count", 0),
        "telegram_similar_channels_unique_count": _REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_unique_count", 0),
        "telegram_similar_channels_added_to_frontier": _REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_added_to_frontier", 0),
        "telegram_similar_channels_fetch_errors": _REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_fetch_errors", 0),
        "telegram_similar_channels_floodwait_count": _REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_floodwait_count", 0),
    }
    summary_row = {
        "run_id":run_id,"started_at":RUN_STARTED_AT.isoformat(),"finished_at":run_now,"git_sha":git_info.get("git_sha", ""),"git_sha_short":git_info.get("git_sha_short", ""),"branch":git_info.get("branch", ""),
        "config_profile":"mvp1.x_llm_first_review_queue","dry_run":"1","ydb_namespace":os.getenv("REGION_TALK_YDB_NAMESPACE") or "dry-run-json",
        "REGION_TALK_DISCOVERY_MODE":os.getenv("REGION_TALK_DISCOVERY_MODE") or "mixed",
        "REGION_TALK_HISTORY_SCAN_MODE":os.getenv("REGION_TALK_HISTORY_SCAN_MODE") or "primary_and_delta",
        "REGION_TALK_MEDIA_SCORING_MODE":os.getenv("REGION_TALK_MEDIA_SCORING_MODE") or current_image_scoring_mode(),
        "REGION_TALK_ACTUAL_IMAGE_TARGET":getenv_int("REGION_TALK_ACTUAL_IMAGE_TARGET", 30),
        "seed_file_version":"v2" if any("v2" in str(s.get("source_seed_id")) for s in source_seed_rows) else "v1/v2-compatible",
        "place_lexicon_file":str(lexicon_path or ""),"place_lexicon_rows":len(lexicon),
        "source_count_seeded":len(seeds),"source_count_scanned":len(source_rows),"posts_fetched":len(posts),
        "source_count_selected":len(source_rows),"source_count_attempted":sum(1 for s in source_rows if str(s.get("fetch_attempted") or "").lower() == "true" or str(s.get("fetch_status") or "").startswith(("ok", "error"))),
        "source_count_ok":sum(1 for s in source_rows if s.get("fetch_status") == "ok"),
        "source_count_skipped":sum(1 for s in source_rows if str(s.get("fetch_status") or "").startswith("skipped")),
        "source_count_error":sum(1 for s in source_rows if str(s.get("fetch_status") or "").startswith("error")),
        "source_count_deferred_by_cooldown":sum(1 for s in source_rows if "cooldown" in str(s.get("fetch_status") or "")),
        "source_count_deferred_unresolved":sum(1 for s in source_rows if "unresolved" in str(s.get("fetch_status") or "")),
        "sources_selected_for_planning":len(source_rows),
        "sources_resolved_available":sum(1 for s in source_rows if s.get("telegram_resolve_status") in {"resolved_network", "resolved_from_private_cache"} or s.get("fetch_status") == "ok"),
        "sources_fetch_attempted":sum(1 for s in source_rows if str(s.get("fetch_attempted") or "").lower() == "true"),
        "sources_history_fetched_ok":sum(1 for s in source_rows if s.get("fetch_status") == "ok"),
        "history_sources_target":_REGION_TALK_TELEGRAM_RUNTIME.get("history_sources_target") or tg_obs.get("history_sources_target") or "",
        "history_sources_attempted":tg_obs.get("history_sources_attempted", sum(1 for s in source_rows if str(s.get("fetch_attempted") or "").lower() == "true")),
        "history_sources_ok":tg_obs.get("history_sources_ok", sum(1 for s in source_rows if s.get("fetch_status") == "ok")),
        "history_sources_new_to_system":sum(1 for s in source_rows if s.get("fetch_status") == "ok" and str(s.get("is_new_source_this_run") or "").lower() == "true"),
        "history_sources_cached_entity":sum(1 for s in source_rows if s.get("fetch_status") == "ok" and str(s.get("history_source_cached_entity") or "").lower() == "true"),
        "history_sources_network_resolved":sum(1 for s in source_rows if s.get("fetch_status") == "ok" and str(s.get("history_source_network_resolved") or "").lower() == "true"),
        "history_fetch_runtime_seconds":round(sum(float(s.get("history_fetch_runtime_seconds") or 0) for s in source_rows if str(s.get("history_fetch_runtime_seconds") or "").strip() not in {"", "None"}), 3),
        "posts_per_source_distribution":json.dumps([int(s.get("posts_scanned") or 0) for s in source_rows if s.get("fetch_status") == "ok"], ensure_ascii=False),
        "sources_skipped_due_resolve_budget":sum(1 for s in source_rows if "budget" in str(s.get("fetch_status") or "")),
        "sources_skipped_due_platform_not_configured":sum(1 for s in source_rows if "not_configured" in str(s.get("fetch_status") or "")),
        "sources_skipped_due_low_priority_defer":sum(1 for s in source_rows if "low_priority" in str(s.get("fetch_status") or "")),
        "vk_wall_probe_status":"ok" if any(s.get("platform") == "vk" and s.get("fetch_status") == "ok" for s in source_rows) else ("error" if any(s.get("platform") == "vk" and str(s.get("fetch_status") or "").startswith("error") for s in source_rows) else ("not_configured" if any(s.get("platform") == "vk" for s in source_rows) and not vk_wall_token() else ("configured_no_vk_selected" if vk_wall_token() else "not_selected"))),
        "posts_region_relevant":sum(1 for r in new_posts if r.get("kaliningrad_oblast_only_scope")),
        "fresh_posts":len(fresh_rows),"fresh_posts_with_place_evidence":len(fresh_with_place_evidence),
        "review_queue_rows":len(review_queue),"pre_candidates_created":len(pre_candidates),
        **llm_limit_snapshot,
        "llm_calls":llm_calls_used,"llm_max_calls":"supabase_reserve_controlled","llm_reviewed":len(llm_reviewed_rows),"llm_accepted":len(llm_accepted_rows),
        "llm_calls_attempted":llm_calls_used,"llm_calls_ok":len(llm_reviewed_rows),"llm_calls_error":sum(1 for r in new_posts if r.get("llm_gate_status") == "error"),"llm_quota_errors":sum(1 for r in new_posts if r.get("llm_gate_status") == "rate_limited"),"llm_retry_rows":sum(1 for r in new_posts if r.get("current_stage") == "needs_llm_retry"),
        "wide_funnel_llm_calls":llm_calls_used if early_llm_enabled_summary else 0,
        "final_verifier_llm_calls":final_verifier_llm_calls,
        "llm_calls_saved_by_vector_gate":len(vector_rejected_rows),
        "llm_calls_saved_by_deterministic_gate":sum(1 for r in new_posts if r.get("image_scoring_skipped") == "true" and not r.get("vector_gate_status")),
        "text_vector_model_id":"dual_model_target:intfloat/multilingual-e5-base+BAAI/bge-m3;prototype_fallback_v1",
        "text_vector_rows_scored":len(vector_scored_rows),
        "vector_rejected_news_event_count":sum(1 for r in vector_scored_rows if r.get("vector_gate_status") == "vector_reject_news_event"),
        "vector_rejected_ad_promo_count":sum(1 for r in vector_scored_rows if r.get("vector_gate_status") == "vector_reject_ad_promo"),
        "vector_rejected_roundup_count":sum(1 for r in vector_scored_rows if r.get("vector_gate_status") in {"vector_reject_roundup", "vector_reject_multi_region_roundup"}),
        "vector_rejected_not_ko_count":sum(1 for r in vector_scored_rows if r.get("vector_gate_status") == "vector_reject_not_kaliningrad_oblast"),
        "vector_rejected_low_substance_count":sum(1 for r in vector_scored_rows if r.get("vector_gate_status") == "vector_reject_low_substance"),
        "vector_ambiguous_kept_count":sum(1 for r in vector_scored_rows if r.get("vector_gate_status") == "vector_ambiguous_keep_for_ranking"),
        "actual_image_scored_before_llm_count":actual_image_scored_before_llm_count,
        "actual_image_target_met":str(actual_image_scored_before_llm_count >= getenv_int("REGION_TALK_ACTUAL_IMAGE_TARGET", 30)).lower(),
        "llm_final_verify_queue_count":final_verify_queue_count_estimate,
        "llm_budget_preserved_count":max(0, len(vector_scored_rows) - llm_calls_used),
        "REGION_TALK_ENABLE_EARLY_LLM":str(early_llm_enabled_summary).lower(),
        "REGION_TALK_ENABLE_VECTOR_GATES":str(getenv_bool("REGION_TALK_ENABLE_VECTOR_GATES", True)).lower(),
        "REGION_TALK_ENABLE_LOCAL_TEXT_EMBEDDINGS":str(getenv_bool("REGION_TALK_ENABLE_LOCAL_TEXT_EMBEDDINGS", True)).lower(),
        "REGION_TALK_TARGET_LLM_CALLS":getenv_int("REGION_TALK_TARGET_LLM_CALLS", 10),
        "REGION_TALK_MAX_LLM_FINAL_VERIFY":max_llm_final_verify,
        "llm_timeout_errors":sum(1 for r in new_posts if "timeout" in str(r.get("llm_reason") or "").lower()),
        "llm_malformed_json_errors":sum(1 for r in new_posts if "json" in str(r.get("llm_reason") or "").lower() and r.get("llm_gate_status") == "error"),
        "llm_cache_hits":0,"llm_budget_remaining":"supabase_controlled",
        "posts_with_strong_media":sum(1 for r in new_posts if r.get("is_selected_for_publication")),
        "candidates_created":len(candidates),"favorites_created":len(favorites),
        "dropped_news":sum(1 for r in dropped if r["rejection_reason"]=="newsiness"),"dropped_trash":sum(1 for r in dropped if r["rejection_reason"]=="trash"),
        "dropped_not_region":sum(1 for r in dropped if r["rejection_reason"]=="reject_not_kaliningrad_oblast_only"),
        "dropped_ad_or_promo":sum(1 for r in dropped if r["rejection_reason"]=="reject_ad_or_promo"),
        "dropped_stale":sum(1 for r in dropped if r["rejection_reason"]=="reject_stale_or_missing_date"),
        "dropped_low_substance":sum(1 for r in dropped if r["rejection_reason"]=="reject_low_substance"),
        "dropped_weak_media":sum(1 for r in dropped if "media" in str(r.get("rejection_reason"))),"dropped_duplicate":0,"dropped_rights":0,
        "image_model_calls":len(media_rows),"image_scoring_skipped_by_text_gate":sum(1 for r in new_posts if r.get("image_scoring_skipped") == "true"),
        "discovered_links":len(discovered_rows),"source_graph_edges":len(graph_edges),"errors_count":sum(1 for s in source_rows if str(s.get("fetch_status") or "").startswith("error")),
        "post_memory_total":len(updated_posts_state),
        "previous_seen_post_count":len(previous_posts),
        "current_fetched_post_count":len(new_posts),
        "previous_posts_not_refetched_this_run":len(set(str(k) for k in previous_posts.keys()) - {str(r.get("post_id")) for r in new_posts}),
        "candidate_memory_total":len(candidate_memory_rows),
        "candidate_memory_active":len(candidate_memory_active_rows),
        "candidate_memory_new_this_run":sum(1 for r in candidate_memory_rows if r.get("first_candidate_run_id") == run_id),
        "candidate_memory_retained_from_previous":sum(1 for r in candidate_memory_rows if r.get("first_candidate_run_id") != run_id),
        "candidate_memory_not_refetched_this_run":len(previous_candidates_not_refetched),
        "candidate_memory_upgraded_this_run":sum(1 for r in candidate_deltas if r.get("delta_bucket") == "stage_upgraded"),
        "candidate_memory_downgraded_this_run":sum(1 for r in candidate_deltas if r.get("delta_bucket") == "stage_downgraded"),
        "candidate_memory_expired_this_run":sum(1 for r in candidate_deltas if r.get("delta_bucket") == "expired_by_policy"),
        "source_frontier_unique_total":len(source_frontier_unique),
        "source_frontier_total":len(source_frontier_unique),
        "source_frontier_new_this_run":sum(1 for r in source_frontier_unique if r.get("discovery_first_run_id") == run_id),
        "source_frontier_promoted_this_run":len(source_frontier_queue_next),
        "source_frontier_ready_next_run":len(source_frontier_queue_next),
        "source_frontier_ready_to_probe":sum(1 for r in source_frontier_unique if str(r.get("frontier_stage") or "") in {"probe_due", "history_due", "unresolved"}),
        "source_frontier_deferred":sum(1 for r in source_frontier_unique if "defer" in str(r.get("frontier_status") or "")),
        "source_frontier_inactive_low_quality":sum(1 for r in source_frontier_unique if str(r.get("frontier_stage") or "") == "inactive_low_quality"),
        "similar_seed_queue_total":len(similar_seed_queue),
        "similar_seed_queue_ready":sum(1 for r in similar_seed_queue if str(r.get("similar_seed_status") or "") == "ready"),
        "similar_seed_queue_used_total":sum(1 for r in similar_seed_queue if int(r.get("similar_seed_use_count") or 0) > 0),
        "dynamic_frontier_seed_count":_REGION_TALK_TELEGRAM_RUNTIME.get("dynamic_frontier_seed_count", 0),
        "telegram_keyword_discovery_status":_REGION_TALK_TELEGRAM_RUNTIME.get("telegram_keyword_discovery_status", "not_run"),
        "keyword_search_queries_processed":_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_search_queries_processed", 0),
        "keyword_discovered_sources_unique":_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_discovered_sources_unique", 0),
        "keyword_discovery_errors":_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_discovery_errors", 0),
        "telegram_phase_status":tg_obs.get("telegram_phase_status"),
        "telegram_floodwait_count":tg_obs.get("resolve_network_floodwait", 0),
        "telegram_max_floodwait_seconds":tg_obs.get("max_floodwait_seconds", 0),
        "telegram_floodwait_method":tg_obs.get("floodwait_method", ""),
        "telegram_cooldown_until":tg_obs.get("floodwait_cooldown_until", ""),
        "telegram_resolve_cache_hits":tg_obs.get("resolve_cache_hits", 0),
        "telegram_resolve_network_attempts":tg_obs.get("resolve_network_attempts", 0),
        "telegram_resolve_network_budget":tg_obs.get("telegram_governor_decisions_json", ""),
        "entity_cache_loaded_from_path":tg_obs.get("entity_cache_loaded_from_path", ""),
        "entity_cache_write_path":tg_obs.get("entity_cache_write_path", ""),
        "entity_cache_hit_rate":tg_obs.get("entity_cache_hit_rate", ""),
        "resolved_sources_available_for_history_fetch":tg_obs.get("resolved_sources_available_for_history_fetch", ""),
        "resolved_sources_used_without_network_resolve":tg_obs.get("resolved_sources_used_without_network_resolve", ""),
        "telegram_sources_deferred_by_cooldown":sum(1 for s in source_rows if "cooldown" in str(s.get("fetch_status") or "")),
        "telegram_sources_deferred_unresolved":sum(1 for s in source_rows if "unresolved" in str(s.get("fetch_status") or "")),
        "telegram_similar_self_loop_count":_REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_self_loop_count", 0),
        "telegram_similar_duplicate_canonical_count":_REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_duplicate_canonical_count", 0),
        "similar_seed_cursor_advanced":_REGION_TALK_TELEGRAM_RUNTIME.get("similar_seed_cursor_advanced", "false"),
        **tg_similar_metrics,
        "new_posts_this_run":sum(1 for r in increment if r.get("new_this_run") == "yes"),
        "changed_posts_this_run":sum(1 for r in increment if r.get("changed_this_run") == "true"),
        "unchanged_posts_this_run":sum(1 for r in increment if r.get("changed_this_run") == "false"),
        "state_backend": state_write_meta.get("state_backend") or state_meta.get("state_backend") or region_talk_state_backend_requested(),
        "state_backend_requested": state_meta.get("state_backend_requested") or region_talk_state_backend_requested(),
        "state_fallback_used": state_write_meta.get("state_fallback_used") or state_meta.get("state_fallback_used") or "false",
        "state_fallback_reason": state_write_meta.get("state_fallback_reason") or state_meta.get("state_fallback_reason") or "",
        "ydb_read_status": state_meta.get("ydb_read_status", "not_requested"),
        "ydb_write_status": state_write_meta.get("ydb_write_status", "not_requested"),
        "ydb_tables_expected": state_write_meta.get("ydb_tables_expected") or state_meta.get("ydb_tables_expected") or "",
        "source_frontier_total_previous":len(previous_state.get("source_frontier_unique") or {}),
        "source_frontier_total_current":len(source_frontier_unique),
        "source_frontier_target_platform_total_current":len(active_frontier_tg_vk),
        "active_frontier_total":len(active_frontier_tg_vk),
        "external_links_quarantined_total":len(external_links_quarantine),
        "catalog_import_file_status":_REGION_TALK_TELEGRAM_RUNTIME.get("catalog_import_file_status", "not_checked"),
        "catalog_import_file_path":_REGION_TALK_TELEGRAM_RUNTIME.get("catalog_import_file_path", ""),
        "catalog_import_rows_total":_REGION_TALK_TELEGRAM_RUNTIME.get("catalog_import_rows_total", len(public_blogger_rows)),
        "catalog_import_telegram_unique":_REGION_TALK_TELEGRAM_RUNTIME.get("catalog_import_telegram_unique", len({r.get("canonical_source_key") for r in public_blogger_rows if r.get("platform") == "telegram"})),
        "catalog_import_vk_unique":_REGION_TALK_TELEGRAM_RUNTIME.get("catalog_import_vk_unique", len({r.get("canonical_source_key") for r in public_blogger_rows if r.get("platform") == "vk"})),
        "catalog_import_external_quarantine":_REGION_TALK_TELEGRAM_RUNTIME.get("catalog_import_external_quarantine", 0),
        "catalog_import_unique_sources":len({r.get("canonical_source_key") or r.get("canonical_url") or r.get("normalized_url") for r in public_blogger_rows}),
        "catalog_sources_in_authoritative_frontier":sum(1 for r in source_frontier_unique if "public_travel_blogger_catalog" in str(r.get("discovery_types") or "")),
        "catalog_sources_in_frontier":sum(1 for r in source_frontier_unique if "public_travel_blogger_catalog" in str(r.get("discovery_types") or "")),
        "frontier_duplicate_canonical_keys":len(source_frontier_unique) - len({r.get("canonical_source_key") or r.get("canonical_url") for r in source_frontier_unique}),
        "frontier_self_loops":0,
        "telegram_similar_seed_used":_REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_seed_count", 0),
        "telegram_similar_raw_count":_REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_raw_count", 0),
        "telegram_similar_unique_count":_REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_unique_count", 0),
        "keyword_queries_processed":_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_search_queries_processed", 0),
        "keyword_unique_sources":_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_discovered_sources_unique", 0),
        "keyword_post_hits_raw":_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_post_hits_raw", 0),
        "keyword_unique_channels":_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_unique_channels", _REGION_TALK_TELEGRAM_RUNTIME.get("keyword_discovered_sources_unique", 0)),
        "keyword_nonlocal_channels":0,
        "keyword_hit_posts_sent_to_pipeline":_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_hit_posts_sent_to_pipeline", 0),
        "sources_primary_scanned_this_run":sum(1 for s in source_rows if s.get("fetch_status") == "ok" and str(s.get("is_new_source_this_run") or "").lower() == "true"),
        "sources_delta_scanned_this_run":sum(1 for s in source_rows if s.get("fetch_status") == "ok" and str(s.get("history_fetch_mode") or "").startswith("delta")),
        "history_cursor_advanced":str(any(s.get("fetch_status") == "ok" and str(s.get("is_new_source_this_run") or "").lower() == "true" for s in source_rows)).lower(),
        "delta_cursor_advanced":str(any(s.get("fetch_status") == "ok" and str(s.get("history_fetch_mode") or "").startswith("delta") for s in source_rows)).lower(),
        "keyword_query_cursor_advanced":str(_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_search_queries_processed", 0) > 0).lower(),
        "posts_new_this_run":sum(1 for r in increment if r.get("new_this_run") == "yes"),
        "posts_delta_new_this_run":sum(1 for r in increment if r.get("new_this_run") == "yes" and not str(r.get("source_seed_id") or "").startswith("frontier_dynamic_")),
        "sample_bias_note":"priority-biased source sample; not a random conversion estimate",
        "sources_with_any_ko_post_per_1000":round(1000 * len({r.get("source_id") for r in new_posts if r.get("kaliningrad_oblast_only_scope")}) / max(1, len([s for s in source_rows if s.get("fetch_status") == "ok"])), 1),
        "sources_with_fresh_ko_post_per_1000":round(1000 * len({r.get("source_id") for r in new_posts if r.get("kaliningrad_oblast_only_scope") and r.get("fresh_enough")}) / max(1, len([s for s in source_rows if s.get("fetch_status") == "ok"])), 1),
        "sources_with_candidate_memory_post_per_1000":round(1000 * len({r.get("source_id") for r in candidate_memory_rows}) / max(1, len([s for s in source_rows if s.get("fetch_status") == "ok"])), 1),
        "sources_with_actual_image_candidate_per_1000":round(1000 * len({r.get("source_id") for r in new_posts if r.get("image_model_input_type") == "actual_image"}) / max(1, len([s for s in source_rows if s.get("fetch_status") == "ok"])), 1),
        "sources_with_publication_ready_candidate_per_1000":round(1000 * len({r.get("source_id") for r in candidate_memory_rows if str(r.get("image_publication_ready") or "") == "true"}) / max(1, len([s for s in source_rows if s.get("fetch_status") == "ok"])), 1),
        "current_run_reviewable_candidates":sum(1 for r in candidate_memory_rows if r.get("first_candidate_run_id") == run_id),
        "final_candidates":len(candidates),
        "favorites":len(favorites),
        "image_fetch_retry_needed":sum(1 for r in new_posts if r.get("current_stage") == "image_fetch_retry_needed"),
        "has_firsthand_visit_evidence_count":sum(1 for r in new_posts if str(r.get("has_firsthand_visit_evidence") or "").lower() == "true"),
        "visit_impression_candidate_count":sum(1 for r in new_posts if str(r.get("content_type") or r.get("vector_content_type") or "") == "visit_impression_candidate"),
        **state_meta,
        **state_write_meta,
        **all_time_metrics,
        **unified_source_queue_metrics,
        **image_queue_metrics,
        "publication_goal_id": publication_goal.get("publication_goal_id"),
        "publication_goal_status": publication_goal.get("goal_status"),
        "publication_goal_target": publication_goal.get("target_confirmed"),
        "publication_confirmed_count": publication_goal.get("confirmed_count"),
        "publication_sent_count": publication_goal.get("sent_count"),
        "publication_llm_calls_used_this_run": publication_goal.get("llm_calls_used_this_run"),
        "publication_llm_calls_used_total": publication_goal.get("llm_calls_used_total"),
        "publication_llm_budget_remaining": publication_goal.get("llm_budget_remaining"),
        "publication_candidate_rows": len(publication_candidate_rows),
        "publication_candidate_confirmed_rows": sum(1 for r in publication_candidate_rows if str(r.get("publication_candidate_status") or "") in PUBLICATION_CONFIRMED_STATUSES),
        "publication_candidate_ready_for_chat_rows": sum(1 for r in publication_candidate_rows if str(r.get("publication_candidate_status") or "") == "llm_confirmed" and str(r.get("sent_to_chat") or "").lower() != "true"),
        "online_candidate_items_written": online_candidate_items_written,
        "online_source_queue_items_written": online_source_queue_items_written,
        "online_image_queue_items_written": online_image_queue_items_written,
        "online_publication_candidate_items_written": online_publication_candidate_items_written,
        "stage_counts_json":json.dumps(stage_counts, ensure_ascii=False),"artifact_paths":""
    }
    run_summary = [summary_row]
    seq_fresh = [r for r in new_posts if r.get("fresh_enough")]
    seq_region = [r for r in seq_fresh if r.get("kaliningrad_oblast_only_scope")]
    seq_non_ad = [r for r in seq_region if not r.get("is_ad_or_promo")]
    seq_substantive = [r for r in seq_non_ad if float(r.get("text_substance_score") or 0) >= 0.18]
    seq_llm_reviewed = [r for r in seq_substantive if r.get("llm_gate_status") == "ok"]
    seq_llm_accepted = [r for r in seq_llm_reviewed if r.get("llm_decision") == "accept"]
    seq_image_actual = [r for r in media_rows if r.get("image_model_input_type") == "actual_image"]
    seq_reviewable_image = [r for r in new_posts if str(r.get("image_reviewable")) == "true"]
    seq_publication_ready = [r for r in new_posts if str(r.get("image_publication_ready")) == "true"]
    funnel = [
        {"stage":"post_fetched","current_run_count":len(new_posts),"notes":"all fetched posts"},
        {"stage":"fresh","current_run_count":len(seq_fresh),"notes":"sequential freshness gate"},
        {"stage":"kaliningrad_oblast_only","current_run_count":len(seq_region),"notes":"sequential region-scope evidence; final semantics still LLM-owned"},
        {"stage":"non_ad","current_run_count":len(seq_non_ad),"notes":"hard promo/ad excluded; possible promo may continue to LLM"},
        {"stage":"substantive","current_run_count":len(seq_substantive),"notes":"basic substance cost guard"},
        {"stage":"semantic_pre_candidate","current_run_count":sum(1 for r in seq_substantive if r.get("llm_gate_status") != "not_run_pre_llm_cost_guard"),"notes":"eligible for Supabase-limited final LLM semantic gate"},
        {"stage":"llm_reviewed","current_run_count":len(seq_llm_reviewed),"notes":"Google Gemini calls through Supabase google_ai_reserve; no direct env max/key2 bypass"},
        {"stage":"llm_accepted","current_run_count":len(seq_llm_accepted),"notes":"LLM semantic accept"},
        {"stage":"image_scored_actual","current_run_count":len(seq_image_actual),"notes":"Kaggle-local neural actual-image rows"},
        {"stage":"reviewable_image","current_run_count":len(seq_reviewable_image),"notes":"accepted text + image reviewable but not necessarily publication-ready"},
        {"stage":"publication_ready_image","current_run_count":len(seq_publication_ready),"notes":"strong image gate passed"},
        {"stage":"final_shortlist","current_run_count":len(final_shortlist) if 'final_shortlist' in locals() else 0,"notes":"filled after shortlist construction; see 04a"},
        {"stage":"favorite","current_run_count":len(favorites),"notes":"top auto-favorite rows"},
    ]
    independent_gate_counts = [
        {"gate":"fresh_enough","count":len(fresh_rows)},
        {"gate":"has_kaliningrad_place_evidence","count":sum(1 for r in new_posts if r.get("matched_place_names"))},
        {"gate":"deterministic_scope_evidence_ok","count":sum(1 for r in new_posts if r.get("kaliningrad_oblast_only_scope"))},
        {"gate":"no_ad_keyword_evidence","count":sum(1 for r in new_posts if not r.get("is_ad_or_promo"))},
        {"gate":"substance_score_ge_025","count":sum(1 for r in new_posts if float(r.get("text_substance_score") or 0) >= 0.25)},
        {"gate":"has_media","count":sum(1 for r in new_posts if r.get("has_media"))},
        {"gate":"semantic_review_required_or_precandidate","count":len(pre_candidates)},
        {"gate":"llm_retry_required","count":sum(1 for r in new_posts if r.get("current_stage") == "needs_llm_retry")},
        {"gate":"clip_actual_image_scored","count":sum(1 for r in media_rows if r.get("image_model_input_type") == "actual_image")},
    ]
    final_shortlist = sorted(
        [r for r in review_queue if r.get("current_stage") in {"favorite", "semantic_candidate", "needs_image_review", "needs_llm_retry", "pre_candidate_needs_llm"}
         and r.get("fresh_enough") and r.get("kaliningrad_oblast_only_scope") and not r.get("is_ad_or_promo")],
        key=lambda r: (
            0 if r.get("current_stage") in {"favorite", "semantic_candidate"} else 1 if r.get("current_stage") == "needs_image_review" else 2 if r.get("current_stage") == "needs_llm_retry" else 3,
            -float(r.get("publication_story_score") or 0),
            -float(r.get("nonlocal_value_score") or 0),
            -float(r.get("candidate_score") or 0),
            int(r.get("post_age_days") or 9999),
        ),
    )
    for i, r in enumerate(final_shortlist, start=1):
        r["human_shortlist_rank"] = i
        if r.get("current_stage") in {"favorite", "semantic_candidate"} and str(r.get("image_publication_ready")) == "true":
            r["decision_bucket"] = "publication_ready_candidate"
            r["next_action"] = "human_final_check"
        elif r.get("current_stage") == "needs_image_review" and str(r.get("image_reviewable")) == "true":
            r["decision_bucket"] = "reviewable_image"
            r["next_action"] = "human_image_review"
        elif r.get("current_stage") == "needs_llm_retry":
            r["decision_bucket"] = "needs_llm_retry"
            r["next_action"] = "retry_llm"
        else:
            r["decision_bucket"] = "needs_llm"
            r["next_action"] = "manual_review"
    for item in funnel:
        if item.get("stage") == "final_shortlist":
            item["current_run_count"] = len(final_shortlist)
    llm_error_rows = [r for r in new_posts if r.get("current_stage") == "needs_llm_retry" or r.get("llm_gate_status") in {"rate_limited", "error"}]
    debug_rejects = [r for r in dropped if r.get("current_stage") in {"debug_reject", "dropped_text_gate"} or str(r.get("rejection_reason") or "").startswith("debug_reject")]
    llm_error_sheet_rows = llm_error_rows or [{
        "post_url":"","current_stage":"","llm_gate_status":"","llm_provider":"","llm_model":"","llm_default_env_var_name":"",
        "llm_reason":"","retry_action":"","_sheet_note":"no LLM retry/error rows in this run",
    }]
    image_model_observability = []
    if media_rows:
        grouped: dict[tuple[str, str, str, str, str], int] = {}
        for mr in media_rows:
            key = (
                str(mr.get("image_scoring_mode") or ""),
                str(mr.get("model_id") or ""),
                str(mr.get("image_model_type") or ""),
                str(mr.get("image_model_runtime") or ""),
                str(mr.get("image_model_input_type") or ""),
            )
            grouped[key] = grouped.get(key, 0) + 1
        for (mode, model_id, model_type, runtime, input_type), count in sorted(grouped.items()):
            devices = sorted({str(r.get("image_model_device") or "") for r in media_rows if str(r.get("model_id") or "") == model_id})
            image_model_observability.append({
                "image_scoring_mode": mode,
                "image_model_id": model_id,
                "image_model_version": next((r.get("model_version") for r in media_rows if str(r.get("model_id") or "") == model_id), ""),
                "image_model_type": model_type,
                "image_model_runtime": runtime,
                "image_model_input_type": input_type,
                "image_model_device": "; ".join(d for d in devices if d),
                "rows": count,
                "actual_image_bytes_required": str(input_type == "actual_image").lower(),
                "fallback_note": "metadata fallback; not neural image understanding" if input_type == "metadata_only" else ("local CLIP actual-image scoring; CLIP is not a VLM" if model_type == "clip" else ""),
            })
    else:
        image_model_observability.append({"image_scoring_mode":current_image_scoring_mode(),"image_model_id":"","image_model_type":"","image_model_runtime":"","image_model_input_type":"","image_model_device":"","rows":0,"actual_image_bytes_required":"true","fallback_note":"no image rows reached scoring"})
    product_summary = [
        {"metric":"run_id","value":run_id},
        {"metric":"git_sha_short","value":git_info.get("git_sha_short")},
        {"metric":"branch","value":git_info.get("branch")},
        {"metric":"increment_status","value":"real increment" if state_meta.get("increment_state_loaded") == "true" else "baseline run, not real increment"},
        {"metric":"previous_run_id","value":state_meta.get("previous_run_id", "")},
        {"metric":"post_memory_total","value":summary_row.get("post_memory_total")},
        {"metric":"previous_posts_not_refetched_this_run","value":summary_row.get("previous_posts_not_refetched_this_run")},
        {"metric":"candidate_memory_total","value":summary_row.get("candidate_memory_total")},
        {"metric":"candidate_memory_active","value":summary_row.get("candidate_memory_active")},
        {"metric":"candidate_memory_new_this_run","value":summary_row.get("candidate_memory_new_this_run")},
        {"metric":"candidate_memory_not_refetched_this_run","value":summary_row.get("candidate_memory_not_refetched_this_run")},
        {"metric":"current_run_shortlist_count","value":len(final_shortlist)},
        {"metric":"publication_goal_status","value":summary_row.get("publication_goal_status")},
        {"metric":"publication_confirmed_count","value":summary_row.get("publication_confirmed_count")},
        {"metric":"publication_goal_target","value":summary_row.get("publication_goal_target")},
        {"metric":"publication_llm_calls_used_total","value":summary_row.get("publication_llm_calls_used_total")},
        {"metric":"publication_llm_budget_remaining","value":summary_row.get("publication_llm_budget_remaining")},
        {"metric":"publication_candidate_ready_for_chat_rows","value":summary_row.get("publication_candidate_ready_for_chat_rows")},
        {"metric":"cumulative_candidate_memory_count","value":len(candidate_memory_rows)},
        {"metric":"publication_ready_current_run_count","value":sum(1 for r in new_posts if str(r.get("image_publication_ready")) == "true")},
        {"metric":"publication_ready_cumulative_count","value":sum(1 for r in candidate_memory_rows if str(r.get("image_publication_ready")) == "true")},
        {"metric":"candidate_memory_note","value":("Current run shortlist is empty, but candidate_memory has %s active candidates; see 06b/21." % summary_row.get("candidate_memory_active")) if not final_shortlist and summary_row.get("candidate_memory_active") else ""},
        {"metric":"posts_fetched","value":len(posts)},
        {"metric":"fresh_posts","value":len(fresh_rows)},
        {"metric":"kaliningrad_oblast_only","value":sum(1 for r in fresh_rows if r.get("kaliningrad_oblast_only_scope"))},
        {"metric":"old_rejected","value":summary_row.get("dropped_stale")},
        {"metric":"ad_rejected","value":summary_row.get("dropped_ad_or_promo")},
        {"metric":"multi_region_or_not_region_rejected","value":summary_row.get("dropped_not_region")},
        {"metric":"news_trash_rejected","value":summary_row.get("dropped_news", 0) + summary_row.get("dropped_trash", 0)},
        {"metric":"low_substance_rejected","value":summary_row.get("dropped_low_substance")},
        {"metric":"fresh_posts_with_place_evidence","value":len(fresh_with_place_evidence)},
        {"metric":"llm_calls_supabase_reserved","value":llm_calls_used},
        {"metric":"wide_funnel_llm_calls","value":summary_row.get("wide_funnel_llm_calls")},
        {"metric":"final_verifier_llm_calls","value":summary_row.get("final_verifier_llm_calls")},
        {"metric":"llm_calls_saved_by_vector_gate","value":summary_row.get("llm_calls_saved_by_vector_gate")},
        {"metric":"llm_calls_saved_by_deterministic_gate","value":summary_row.get("llm_calls_saved_by_deterministic_gate")},
        {"metric":"text_vector_model_id","value":summary_row.get("text_vector_model_id")},
        {"metric":"text_vector_rows_scored","value":summary_row.get("text_vector_rows_scored")},
        {"metric":"vector_rejected_news_event_count","value":summary_row.get("vector_rejected_news_event_count")},
        {"metric":"vector_rejected_ad_promo_count","value":summary_row.get("vector_rejected_ad_promo_count")},
        {"metric":"vector_rejected_roundup_count","value":summary_row.get("vector_rejected_roundup_count")},
        {"metric":"vector_rejected_not_ko_count","value":summary_row.get("vector_rejected_not_ko_count")},
        {"metric":"vector_rejected_low_substance_count","value":summary_row.get("vector_rejected_low_substance_count")},
        {"metric":"vector_ambiguous_kept_count","value":summary_row.get("vector_ambiguous_kept_count")},
        {"metric":"actual_image_scored_before_llm_count","value":summary_row.get("actual_image_scored_before_llm_count")},
        {"metric":"llm_final_verify_queue_count","value":summary_row.get("llm_final_verify_queue_count")},
        {"metric":"llm_budget_preserved_count","value":summary_row.get("llm_budget_preserved_count")},
        {"metric":"llm_calls_ok","value":len(llm_reviewed_rows)},
        {"metric":"llm_retry_rows","value":len(llm_error_rows)},
        {"metric":"image_scored_rows","value":len(media_rows)},
        {"metric":"actual_image_neural_rows","value":sum(1 for r in media_rows if r.get("image_model_input_type") == "actual_image")},
        {"metric":"actual_image_target_met","value":summary_row.get("actual_image_target_met")},
        {"metric":"human_final_shortlist_rows","value":len(final_shortlist)},
        {"metric":"good_text_weak_media_rows","value":sum(1 for r in dropped if r.get("current_stage") == "good_text_weak_media")},
        {"metric":"publication_ready_rows","value":sum(1 for r in new_posts if str(r.get("image_publication_ready")) == "true")},
        {"metric":"final_candidates","value":len(candidates)},
        {"metric":"favorites","value":len(favorites)},
        {"metric":"why_zero_candidates_or_favorites","value":"Image gate did not produce publication_ready rows; review 04a_final_shortlist and 10_good_text_weak_media." if not candidates and not favorites else ""},
        {"metric":"source_coverage","value":f"selected={summary_row.get('source_count_selected')}, ok={summary_row.get('source_count_ok')}, skipped={summary_row.get('source_count_skipped')}, errors={summary_row.get('source_count_error')}"},
        {"metric":"history_sources_target","value":summary_row.get("history_sources_target")},
        {"metric":"history_sources_attempted","value":summary_row.get("history_sources_attempted")},
        {"metric":"history_sources_ok","value":summary_row.get("history_sources_ok")},
        {"metric":"history_sources_new_to_system","value":summary_row.get("history_sources_new_to_system")},
        {"metric":"history_sources_cached_entity","value":summary_row.get("history_sources_cached_entity")},
        {"metric":"history_sources_network_resolved","value":summary_row.get("history_sources_network_resolved")},
        {"metric":"history_fetch_runtime_seconds","value":summary_row.get("history_fetch_runtime_seconds")},
        {"metric":"posts_per_source_distribution","value":summary_row.get("posts_per_source_distribution")},
        {"metric":"vk_wall_probe_status","value":summary_row.get("vk_wall_probe_status")},
        {"metric":"telegram_phase_status","value":summary_row.get("telegram_phase_status")},
        {"metric":"telegram_floodwait_count","value":summary_row.get("telegram_floodwait_count")},
        {"metric":"telegram_max_floodwait_seconds","value":summary_row.get("telegram_max_floodwait_seconds")},
        {"metric":"telegram_floodwait_method","value":summary_row.get("telegram_floodwait_method")},
        {"metric":"telegram_cooldown_until","value":summary_row.get("telegram_cooldown_until")},
        {"metric":"telegram_resolve_cache_hits","value":summary_row.get("telegram_resolve_cache_hits")},
        {"metric":"telegram_resolve_network_attempts","value":summary_row.get("telegram_resolve_network_attempts")},
        {"metric":"entity_cache_loaded_from_path","value":summary_row.get("entity_cache_loaded_from_path")},
        {"metric":"entity_cache_hit_rate","value":summary_row.get("entity_cache_hit_rate")},
        {"metric":"resolved_sources_available_for_history_fetch","value":summary_row.get("resolved_sources_available_for_history_fetch")},
        {"metric":"telegram_sources_deferred_by_cooldown","value":summary_row.get("telegram_sources_deferred_by_cooldown")},
        {"metric":"telegram_sources_deferred_unresolved","value":summary_row.get("telegram_sources_deferred_unresolved")},
        {"metric":"telegram_similar_channels_status","value":summary_row.get("telegram_similar_channels_status")},
        {"metric":"telegram_similar_channels_seed_count","value":summary_row.get("telegram_similar_channels_seed_count")},
        {"metric":"telegram_similar_channels_raw_count","value":summary_row.get("telegram_similar_channels_raw_count")},
        {"metric":"telegram_similar_channels_unique_count","value":summary_row.get("telegram_similar_channels_unique_count")},
        {"metric":"telegram_similar_channels_added_to_frontier","value":summary_row.get("telegram_similar_channels_added_to_frontier")},
        {"metric":"similar_seed_queue_total","value":summary_row.get("similar_seed_queue_total")},
        {"metric":"similar_seed_queue_ready","value":summary_row.get("similar_seed_queue_ready")},
        {"metric":"similar_seed_queue_used_total","value":summary_row.get("similar_seed_queue_used_total")},
        {"metric":"similar_seed_cursor_advanced","value":summary_row.get("similar_seed_cursor_advanced")},
        {"metric":"telegram_keyword_discovery_status","value":summary_row.get("telegram_keyword_discovery_status")},
        {"metric":"keyword_search_queries_processed","value":summary_row.get("keyword_search_queries_processed")},
        {"metric":"keyword_discovered_sources_unique","value":summary_row.get("keyword_discovered_sources_unique")},
        {"metric":"dynamic_frontier_seed_count","value":summary_row.get("dynamic_frontier_seed_count")},
        {"metric":"source_frontier_unique_total","value":summary_row.get("source_frontier_unique_total")},
        {"metric":"source_frontier_new_this_run","value":summary_row.get("source_frontier_new_this_run")},
        {"metric":"catalog_import_rows_total","value":summary_row.get("catalog_import_rows_total")},
        {"metric":"catalog_sources_in_authoritative_frontier","value":summary_row.get("catalog_sources_in_authoritative_frontier")},
        {"metric":"frontier_duplicate_canonical_keys","value":summary_row.get("frontier_duplicate_canonical_keys")},
        {"metric":"source_frontier_ready_to_probe","value":summary_row.get("source_frontier_ready_to_probe")},
        {"metric":"source_frontier_promoted_this_run","value":summary_row.get("source_frontier_promoted_this_run")},
        {"metric":"source_frontier_ready_next_run","value":summary_row.get("source_frontier_ready_next_run")},
        {"metric":"source_frontier_deferred","value":summary_row.get("source_frontier_deferred")},
        {"metric":"sources_primary_scanned_total_all_time","value":summary_row.get("sources_primary_scanned_total_all_time")},
        {"metric":"telegram_sources_primary_scanned_total_all_time","value":summary_row.get("telegram_sources_primary_scanned_total_all_time")},
        {"metric":"vk_sources_primary_scanned_total_all_time","value":summary_row.get("vk_sources_primary_scanned_total_all_time")},
        {"metric":"sources_delta_scanned_this_run","value":summary_row.get("sources_delta_scanned_this_run")},
        {"metric":"sources_delta_scanned_total_all_time","value":summary_row.get("sources_delta_scanned_total_all_time")},
        {"metric":"sources_never_scanned_total","value":summary_row.get("sources_never_scanned_total")},
        {"metric":"frontier_total_by_platform","value":summary_row.get("frontier_total_by_platform")},
        {"metric":"frontier_pending_primary_scan_total","value":summary_row.get("frontier_pending_primary_scan_total")},
        {"metric":"frontier_pending_similar_scan_total","value":summary_row.get("frontier_pending_similar_scan_total")},
        {"metric":"posts_memory_total","value":summary_row.get("posts_memory_total")},
        {"metric":"candidates_memory_total","value":summary_row.get("candidates_memory_total")},
        {"metric":"publication_ready_total_all_time","value":summary_row.get("publication_ready_total_all_time")},
        {"metric":"llm_limit_source","value":llm_limit_snapshot.get("llm_limit_source")},
        {"metric":"llm_model","value":llm_model},
        {"metric":"llm_default_env_var_name","value":llm_default_env_var_name},
        {"metric":"image_scoring_mode","value":current_image_scoring_mode()},
    ]

    compact_current_run_shortlist = [compact_shortlist_row(r) for r in final_shortlist]
    compact_final_shortlist = [candidate_memory_shortlist_row(r, i) for i, r in enumerate(candidate_memory_top, start=1)] or compact_current_run_shortlist
    publication_queue_sheet = publication_candidate_rows or [{"_sheet_note":"no publication candidate rows yet; requires nonlocal text candidate + actual scored image + Gemini final verification"}]
    publication_confirmed_sheet = [r for r in publication_candidate_rows if str(r.get("publication_candidate_status") or "") in PUBLICATION_CONFIRMED_STATUSES] or [{"_sheet_note":"no Gemini-confirmed publication candidates yet"}]
    manual_review_queue = candidate_memory_top or compact_current_run_shortlist or [{"_sheet_note":"no active candidate memory/manual review rows yet"}]
    candidate_memory_sheet = candidate_memory_rows or [{"_sheet_note":"candidate memory is empty; no current or previous candidate rows reached memory criteria"}]
    candidate_memory_top_sheet = candidate_memory_top or [{"_sheet_note":"no active candidate memory rows"}]
    previous_not_refetched_sheet = previous_candidates_not_refetched or [{"_sheet_note":"no previous active candidates missed by this run"}]
    candidate_deltas_sheet = candidate_deltas or [{"_sheet_note":"no candidate deltas yet"}]
    source_frontier_queue_sheet = source_frontier_queue_next or [{"_sheet_note":"no source frontier queue rows"}]
    image_retry_queue = [r for r in new_posts if r.get("current_stage") == "image_fetch_retry_needed" or r.get("image_status") == "needs_actual_image_fetch"]
    seen_retry = {str(r.get("post_url") or r.get("candidate_memory_id") or "") for r in image_retry_queue}
    for r in candidate_memory_top:
        key = str(r.get("post_url") or r.get("candidate_memory_id") or "")
        if key not in seen_retry and (r.get("current_stage") == "image_fetch_retry_needed" or r.get("image_status") == "needs_actual_image_fetch"):
            image_retry_queue.append(r)
            seen_retry.add(key)
    image_retry_queue_sheet = image_retry_queue or [{"_sheet_note":"no metadata-only/image-fetch retry rows"}]
    llm_usage_by_stage = [
        {"stage":"broad_fetched_posts","rows_considered":len(new_posts),"llm_calls":summary_row.get("wide_funnel_llm_calls"),"allowed":"false","reason":"vector/deterministic only before compact shortlist","budget_cap":0},
        {"stage":"review_queue","rows_considered":len(review_queue),"llm_calls":0,"allowed":"false","reason":"wide review queue remains vector/local; no broad LLM classifier","budget_cap":0},
        {"stage":"candidate_memory_active","rows_considered":len(candidate_memory_active_rows),"llm_calls":0,"allowed":"true","reason":"eligible for future final verifier only; previous rows may lack text in MVP state","budget_cap":max_llm_final_verify},
        {"stage":"final_shortlist_top_n","rows_considered":len(compact_final_shortlist),"llm_calls":summary_row.get("final_verifier_llm_calls"),"allowed":"true","reason":"final verifier/tie-breaker only","budget_cap":max_llm_final_verify},
    ]
    vk_wall_setup_sheet = [{
        "status": summary_row.get("vk_wall_probe_status"),
        "needed_env": "VK_SERVICE_TOKEN or VK_ACCESS_TOKEN with wall read access",
        "publish_blocker": "VK publishing remains disabled in MVP; read-only wall probe is separate",
        "frontier_vk_rows": sum(1 for r in source_frontier_unique if str(r.get("platform") or "") == "vk"),
        "next_action": "configure read-only VK wall token or keep explicit blocker",
    }]
    favorites_sheet = [
        {"favorite_id": "fav_" + stable_hash(r.get("candidate_id"), r.get("post_url")), "why_selected": "favorite stage from vector/image ranking", "publication_readiness": r.get("image_publication_ready"), **r}
        for r in favorites
    ] or [{"_sheet_note":"no favorites in this run"}]
    candidates_sheet = candidates or [{"_sheet_note":"no final candidates in this run"}]
    favorites_non_placeholder = [r for r in favorites_sheet if r.get("post_url") and not r.get("_sheet_note")]
    candidates_non_placeholder = [r for r in candidates_sheet if r.get("post_url") and not r.get("_sheet_note")]
    consistency_errors = []
    if len(favorites) != len(favorites_non_placeholder):
        consistency_errors.append(f"favorites summary={len(favorites)} sheet={len(favorites_non_placeholder)}")
    if len(candidates) != len(candidates_non_placeholder):
        consistency_errors.append(f"candidates summary={len(candidates)} sheet={len(candidates_non_placeholder)}")
    summary_row["favorites_candidates_consistency_status"] = "ok" if not consistency_errors else "error: " + "; ".join(consistency_errors)
    product_summary.append({"metric":"favorites_candidates_consistency_status","value":summary_row["favorites_candidates_consistency_status"]})
    telegram_observability_rows = [tg_obs] if tg_obs else [{"run_id":run_id,"telegram_phase_status":"not_configured"}]
    similar_sheet_rows = similar_channel_rows or [{
        "run_id": run_id, "seed_source_id": "", "seed_channel_url": "", "seed_channel_title": "",
        "method_status": _REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_status", "not_run"),
        "method_error_code": _REGION_TALK_TELEGRAM_RUNTIME.get("telegram_similar_channels_error", ""),
        "method_error_message_short": "", "recommendation_rank": "", "recommended_title": "",
        "recommended_username": "", "recommended_canonical_url": "", "edge_type": "telegram_similar_channel",
        "frontier_action": "none", "frontier_reason": "no Telegram similar channel rows in this run",
    }]
    similar_seed_queue_sheet = similar_seed_queue or [{"_sheet_note":"no persistent similar-channel seeds yet"}]
    keyword_post_hit_rows = list(_REGION_TALK_TELEGRAM_RUNTIME.get("keyword_post_hit_rows") or [])
    for r in keyword_post_hit_rows:
        cls = classify_source_profile({
            "source_title": r.get("source_title") or r.get("username_or_handle"),
            "resolved_title": r.get("source_title") or r.get("username_or_handle"),
            "canonical_url": r.get("keyword_hit_source_url"),
            "source_kind": "telegram_keyword_search",
            "source_type": "telegram_keyword_search",
        }, [])
        r.update({k: cls.get(k, "") for k in ["source_geo_class", "source_topic_class", "ko_mention_ratio_recent", "travel_blogger_score", "personal_voice_score", "nonlocal_value_score", "source_priority_reason"]})
        r["product_priority_pool"] = "keyword_nonlocal_priority" if str(cls.get("source_geo_class")) != "kaliningrad_local" and cls.get("source_topic_class") in {"travel_blogger", "travel_media", "personal_blog", "unknown"} else "keyword_low_priority_local_or_unknown"
    summary_row["keyword_nonlocal_channels"] = len({r.get("canonical_source_key") for r in keyword_post_hit_rows if r.get("product_priority_pool") == "keyword_nonlocal_priority"})
    for metric in [
        "active_frontier_total", "external_links_quarantined_total",
        "catalog_import_file_status", "catalog_import_rows_total", "catalog_import_telegram_unique", "catalog_import_vk_unique",
        "keyword_post_hits_raw", "keyword_unique_channels", "keyword_nonlocal_channels", "keyword_hit_posts_sent_to_pipeline",
        "history_cursor_advanced", "delta_cursor_advanced", "keyword_query_cursor_advanced",
        "source_queue_total", "source_queue_pending_total", "source_queue_processed_total", "source_queue_retry_total",
        "source_queue_cursor_position", "source_queue_keyword_inserted_this_run", "source_queue_catalog_sources_total",
        "source_queue_telegram_total", "source_queue_vk_total", "source_queue_pending_telegram_total", "source_queue_pending_vk_total",
        "source_queue_non_target_skipped_this_run", "source_queue_low_image_quality_excluded_total",
        "source_queue_only_telegram_vk", "source_queue_only_target_source_urls", "image_queue_total", "image_queue_cursor_position",
        "image_queue_target_this_run", "image_queue_selected_next_batch", "image_queue_actual_scored_total",
        "image_queue_needs_actual_fetch_total",
        "publication_goal_status", "publication_confirmed_count", "publication_goal_target",
        "publication_candidate_rows", "publication_candidate_confirmed_rows", "publication_candidate_ready_for_chat_rows",
        "publication_llm_calls_used_this_run", "publication_llm_calls_used_total", "publication_llm_budget_remaining",
    ]:
        product_summary.append({"metric": metric, "value": summary_row.get(metric)})
    keyword_discovery_sheet = keyword_discovery_rows or [{"_sheet_note":"no Telegram keyword discovery rows in this run", "method_status": _REGION_TALK_TELEGRAM_RUNTIME.get("telegram_keyword_discovery_status", "not_run")}]
    keyword_post_hits_sheet = keyword_post_hit_rows or [{"_sheet_note":"no Telegram keyword post hits in this run", "method_status": _REGION_TALK_TELEGRAM_RUNTIME.get("telegram_keyword_discovery_status", "not_run")}]
    keyword_hit_candidates_sheet = [r for r in keyword_post_hit_rows if r.get("product_priority_pool") == "keyword_nonlocal_priority"] or [{"_sheet_note":"no keyword-hit candidates after source classification"}]
    source_classification_sheet = [
        {k: r.get(k, "") for k in ["source_id", "source_title", "canonical_url", "platform", "fetch_status", "source_geo_class", "source_topic_class", "ko_mention_ratio_recent", "travel_blogger_score", "personal_voice_score", "nonlocal_value_score", "source_priority_reason"]}
        for r in source_rows
    ] or [{"_sheet_note":"no source classification rows"}]
    actual_image_quality_rows = [r for r in media_rows if r.get("image_model_input_type") == "actual_image"]
    image_quality_debug_rows = [r for r in media_rows if r.get("image_model_input_type") != "actual_image"] or [{"_sheet_note":"no metadata/debug image rows"}]
    sheets = {
        "00_product_summary":product_summary,
        "00_readme":[{"field":"what","value":"Region Talk MVP-1.x Candidate Report Only; vector/local gates first, image scoring for non-ad Kaliningrad rows, optional final LLM verifier via Supabase google_ai limiter; no Telegram/VK publishing."},{"field":"run_id","value":run_id},{"field":"generated_at","value":run_now}],
        "01_run_summary":run_summary,
        "02_increment":increment,
        "02b_runtime_deferred_posts":[{"post_id": r.get("post_id"), "post_url": r.get("post_url"), "post_date": r.get("post_date"), "source_title": r.get("source_title"), "defer_reason":"runtime_vector_budget", "next_action":"score_in_next_bounded_run"} for r in runtime_deferred_posts[:500]],
        "03_funnel":funnel,
        "03b_gate_counts":independent_gate_counts,
        "04_review_queue":review_queue,
        "04a_final_shortlist":compact_final_shortlist,
        "04a_current_run_shortlist":compact_current_run_shortlist,
        "04a_final_shortlist_raw":candidate_memory_top or final_shortlist,
        "04p_publication_queue":publication_queue_sheet,
        "04q_publication_confirmed":publication_confirmed_sheet,
        "04b_needs_llm_retry":llm_error_sheet_rows,
        "04c_debug_rejects":debug_rejects,
        "05_favorites":favorites_sheet,
        "06_candidates_all":candidates_sheet,
        "06a_candidate_memory":candidate_memory_sheet,
        "06b_candidate_memory_top":candidate_memory_top_sheet,
        "07_new_posts_this_run":[r for r in new_posts if any(inc.get("entity_id") == r.get("post_id") and inc.get("new_this_run") == "yes" for inc in increment)],
        "07_current_run_posts":new_posts,
        "07b_prev_candidates_not_refetch":previous_not_refetched_sheet,
        "08_dropped_posts":dropped,
        "09_image_quality":actual_image_quality_rows or [{"_sheet_note":"no actual-image rows in this run; see 09c_image_quality_debug_fallback"}],
        "09a_image_candidate_queue":image_candidate_queue_sheet or [{"_sheet_note":"no image candidate queue rows"}],
        "09b_image_fetch_retry_queue":image_retry_queue_sheet,
        "09c_image_debug_fallback":image_quality_debug_rows,
        "09d_image_driven_top":image_driven_top_sheet or [{"_sheet_note":"no actual-image rows scored yet; image-driven ranking waits for media acquisition"}],
        "10_good_text_weak_media":[r for r in dropped if r.get("current_stage") == "good_text_weak_media"],
        "11_sources_seed":source_seed_rows,
        "12_source_queue":unified_source_queue_sheet or [{"_sheet_note":"no canonical Telegram/VK source queue rows"}],
        "12_sources_discovered":discovered_rows,
        "12a_source_frontier_unique":source_frontier_unique,
        "12a_active_tg_vk_frontier":active_frontier_tg_vk or [{"_sheet_note":"no active Telegram/VK frontier rows"}],
        "12b_telegram_similar_channels":similar_sheet_rows,
        "12c_source_frontier_queue_next":source_frontier_queue_sheet,
        "12d_similar_seed_queue":similar_seed_queue_sheet,
        "12e_telegram_keyword_discovery":keyword_discovery_sheet,
        "12e_keyword_posts":keyword_post_hits_sheet,
        "12f_source_classification":source_classification_sheet,
        "12g_external_links_quarantine":external_links_quarantine or [{"_sheet_note":"no external links quarantined"}],
        "04k_keyword_hit_candidates":keyword_hit_candidates_sheet,
        "13_sources_monitored":source_rows,
        "13b_source_delta_scan":source_delta_scan_sheet,
        "14_verifier_reports":[],
        "14b_pre_candidates_needing_llm":[r for r in pre_candidates if r.get("current_stage") == "pre_candidate_needs_llm"],
        "14c_llm_errors":llm_error_sheet_rows,
        "14d_llm_usage_by_stage":llm_usage_by_stage,
        "15_manual_decisions":[{"candidate_id":"","manual_decision":"favorite|reject|approve_for_preview|approve_for_queue|block_source","reviewer":"","reviewed_at":"","reviewer_comment":"","rights_override":"","source_status_override":""}],
        "16_publish_preview_future":[{"note":"Future only. REGION_TALK_DISABLE_PUBLISH=1; no real publishing in MVP-1.x."}],
        "17_source_graph_edges":graph_edges,
        "18_place_lexicon_matches":place_match_rows,
        "19_image_model_observability":image_model_observability,
        "20_telegram_rate_observability":telegram_observability_rows,
        "21_manual_review_queue":manual_review_queue,
        "22_candidate_deltas":candidate_deltas_sheet,
        "23_vk_wall_setup":vk_wall_setup_sheet,
        "24_source_yield_metrics":source_yield_metrics_sheet,
    }
    report_event("vector_scoring_done", phase="vector_scoring", status="running", posts_scored=len(posts_for_scoring), posts_deferred=len(runtime_deferred_posts), candidates_created=len(candidates), image_queue_total=len(image_candidate_queue_sheet))
    xlsx = output_dir / f"region-talk-candidates-{run_id}.xlsx"
    report_event("report_write_started", phase="report", status="running", xlsx=str(xlsx), candidates_created=len(candidates), image_queue_total=len(image_candidate_queue_sheet))
    write_xlsx(xlsx, sheets)
    candidate_event_rows = []
    def _candidate_event(stage: str, r: dict[str, Any]) -> dict[str, Any]:
        return {
            "event_type": stage,
            "run_id": run_id,
            "event_at": run_now,
            "source_id": r.get("source_id"),
            "source_title": r.get("source_title"),
            "source_url": r.get("source_url"),
            "post_id": r.get("post_id"),
            "post_url": r.get("post_url"),
            "post_date": r.get("post_date"),
            "matched_place_names": r.get("matched_place_names"),
            "content_type": r.get("content_type") or r.get("llm_content_type") or r.get("vector_content_type"),
            "candidate_score": r.get("candidate_score_current") or r.get("candidate_score") or r.get("best_candidate_score_ever"),
            "media_score": r.get("media_score_current") or r.get("overall_media_score") or r.get("best_media_score_ever"),
            "stage": r.get("current_stage") or r.get("decision_bucket"),
            "next_action": r.get("next_action") or r.get("suggested_action"),
            "short_summary": r.get("short_summary") or r.get("why_keep_in_memory") or r.get("why_this_is_about_kaliningrad"),
        }
    for r in source_rows:
        if r.get("fetch_status") == "ok" and r.get("source_geo_class") != "kaliningrad_local" and float(r.get("nonlocal_value_score") or 0) >= 0.25:
            candidate_event_rows.append({
                "event_type": "new_nonlocal_ko_channel_found",
                "run_id": run_id,
                "event_at": run_now,
                "source_id": r.get("source_id"),
                "source_title": r.get("source_title"),
                "source_url": r.get("canonical_url") or r.get("source_url"),
                "post_id": "",
                "post_url": "",
                "post_date": r.get("last_seen_post_date"),
                "matched_place_names": "",
                "content_type": r.get("source_topic_class"),
                "candidate_score": r.get("nonlocal_value_score"),
                "media_score": "",
                "stage": "source_discovery",
                "next_action": "prioritize_keyword_or_primary_scan",
                "short_summary": r.get("source_priority_reason"),
            })
    for r in keyword_post_hit_rows:
        if r.get("product_priority_pool") == "keyword_nonlocal_priority":
            candidate_event_rows.append({
                "event_type": "keyword_hit_candidate_found",
                "run_id": run_id,
                "event_at": run_now,
                "source_id": r.get("source_candidate_id"),
                "source_title": r.get("source_title") or r.get("username_or_handle"),
                "source_url": r.get("keyword_hit_source_url"),
                "post_id": "",
                "post_url": r.get("keyword_hit_post_url"),
                "post_date": "",
                "matched_place_names": r.get("matched_query"),
                "content_type": r.get("source_topic_class"),
                "candidate_score": r.get("nonlocal_value_score"),
                "media_score": "",
                "stage": "keyword_hit_source_context",
                "next_action": "fetch_hit_post_context_and_recent_history",
                "short_summary": r.get("keyword_hit_text_excerpt") or r.get("source_priority_reason"),
            })
    for r in new_posts:
        if r.get("kaliningrad_oblast_only_scope"):
            candidate_event_rows.append(_candidate_event("fresh_ko_post_found" if r.get("fresh_enough") else "new_source_with_ko_post", r))
        if r.get("current_stage") in {"pre_candidate_needs_llm", "semantic_candidate", "favorite", "needs_image_review", "image_fetch_retry_needed", "good_text_weak_media"}:
            candidate_event_rows.append(_candidate_event("pre_candidate_created", r))
        if r.get("current_stage") == "needs_image_review":
            candidate_event_rows.append(_candidate_event("reviewable_image_candidate_found", r))
        if r.get("current_stage") == "image_fetch_retry_needed":
            candidate_event_rows.append(_candidate_event("image_fetch_retry_needed", r))
        if str(r.get("image_publication_ready") or "").lower() == "true" and r.get("image_model_input_type") == "actual_image":
            candidate_event_rows.append(_candidate_event("publication_ready_candidate_found", r))
    if not candidate_event_rows:
        for r in (candidate_memory_top or final_shortlist)[:100]:
            candidate_event_rows.append(_candidate_event("candidate_found", r))
    (output_dir / "candidate_found.jsonl").write_text("\n".join(json.dumps(r, ensure_ascii=False) for r in candidate_event_rows) + ("\n" if candidate_event_rows else ""), encoding="utf-8")
    run_event_rows = [
        {"event_type":"run_started","run_id":run_id,"created_at":RUN_STARTED_AT.isoformat()},
        {"event_type":"report_written","run_id":run_id,"created_at":run_now,"xlsx_path":str(xlsx),"posts_fetched":len(posts),"candidate_events":len(candidate_event_rows)},
    ]
    (output_dir / "run_events.jsonl").write_text("\n".join(json.dumps(r, ensure_ascii=False) for r in run_event_rows) + "\n", encoding="utf-8")
    (output_dir / "stage_status.json").write_text(json.dumps({"run_id":run_id,"generated_at":run_now,"stage_counts":stage_counts,"summary":summary_row}, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(output_dir / f"region-talk-candidates-{run_id}.csv", review_queue)
    payload = {"ok": True, "run_id": run_id, "generated_at": run_now, "summary": run_summary[0], "sheets": sheets, "xlsx_path": str(xlsx)}
    (output_dir / f"region-talk-candidates-{run_id}.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    (output_dir / f"region-talk-candidates-{run_id}.md").write_text(render_md(payload), encoding="utf-8")
    (output_dir / f"region-talk-candidates-{run_id}.html").write_text(render_html(payload), encoding="utf-8")
    for p in [xlsx, output_dir / f"region-talk-candidates-{run_id}.json", output_dir / f"region-talk-candidates-{run_id}.md", output_dir / f"region-talk-candidates-{run_id}.html", output_dir / f"region-talk-candidates-{run_id}.csv", output_dir / "candidate_found.jsonl", output_dir / "run_events.jsonl", output_dir / "stage_status.json"]:
        target = Path.cwd() / p.name
        if target.resolve() != p.resolve():
            target.write_bytes(p.read_bytes())
    latest = Path.cwd() / "candidates-latest.xlsx"
    latest.write_bytes(xlsx.read_bytes())
    (Path.cwd() / "output.json").write_text(json.dumps({"ok": True, "run_id": run_id, "xlsx": str(latest), "summary": run_summary[0]}, ensure_ascii=False, indent=2), encoding="utf-8")
    payload["latest_xlsx"] = str(latest)
    return payload

def make_summary(text: str) -> str:
    clean = re.sub(r"\s+", " ", text or "").strip()
    if len(clean) <= 180:
        return clean
    return clean[:177].rsplit(" ",1)[0] + "…"


def rows_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for r in rows:
        for k in r.keys():
            if k not in headers:
                headers.append(k)
    return headers or ["empty"]


def col_name(n: int) -> str:
    s = ""
    while n:
        n, rem = divmod(n-1, 26)
        s = chr(65+rem) + s
    return s


def cell_xml(value: Any, row: int, col: int) -> str:
    ref = f"{col_name(col)}{row}"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f'<c r="{ref}"><v>{value}</v></c>'
    text = escape(str(value if value is not None else ""))
    return f'<c r="{ref}" t="inlineStr"><is><t>{text}</t></is></c>'


def sheet_xml(rows: list[dict[str, Any]]) -> str:
    headers = rows_headers(rows)
    all_rows = [dict(zip(headers, headers))] + rows
    body=[]
    for ri, r in enumerate(all_rows, start=1):
        cells = ''.join(cell_xml(r.get(h, ""), ri, ci) for ci, h in enumerate(headers, start=1))
        body.append(f'<row r="{ri}">{cells}</row>')
    last = f"{col_name(len(headers))}{max(1,len(all_rows))}"
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews><dimension ref="A1:{last}"/><sheetData>{''.join(body)}</sheetData><autoFilter ref="A1:{col_name(len(headers))}1"/></worksheet>'''



def excel_safe_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    if isinstance(value, (list, dict)):
        text = json.dumps(value, ensure_ascii=False)
    else:
        text = str(value)
    try:
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE  # type: ignore
        text = ILLEGAL_CHARACTERS_RE.sub("", text)
    except Exception:
        text = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", text)
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    if len(text) > 32767:
        text = text[:32700] + "… [truncated for Excel]"
    return text


def write_xlsx_openpyxl(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        if getenv_bool("REGION_TALK_AUTO_INSTALL", True):
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore
            from openpyxl.utils import get_column_letter  # type: ignore
        else:
            raise
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    header_border = Border(bottom=thin)
    row_fills = {
        "white_pending": PatternFill("solid", fgColor="FFFFFF"),
        "yellow_retry": PatternFill("solid", fgColor="FFF2CC"),
        "green_found_ko": PatternFill("solid", fgColor="E2F0D9"),
        "red_no_ko": PatternFill("solid", fgColor="F4CCCC"),
        "blue_cursor": PatternFill("solid", fgColor="D9EAF7"),
    }
    used_titles: set[str] = set()
    for sheet_name, raw_rows in sheets.items():
        title = re.sub(r"[:\\/?*\[\]]", "_", str(sheet_name))[:31] or "sheet"
        base_title = title
        i = 2
        while title in used_titles:
            suffix = f"_{i}"
            title = (base_title[:31-len(suffix)] + suffix)[:31]
            i += 1
        used_titles.add(title)
        rows = raw_rows if isinstance(raw_rows, list) else [{"value": raw_rows}]
        ws = wb.create_sheet(title=title)
        headers = rows_headers([r for r in rows if isinstance(r, dict)])
        headers = [h for h in headers if h != "place_matches"] or ["empty"]
        ws.append(headers)
        widths = {idx: len(str(header)) for idx, header in enumerate(headers, start=1)}
        for raw in rows:
            row = raw if isinstance(raw, dict) else {"value": raw}
            values = [excel_safe_value(row.get(header, "")) for header in headers]
            ws.append(values)
            fill_key = str(row.get("row_fill_color") or row.get("status_color_hint") or "").strip()
            if str(row.get("cursor_marker") or "") in {"cursor_here", "next_after_cursor"}:
                fill_key = "blue_cursor" if str(row.get("cursor_marker") or "") == "cursor_here" else fill_key
            fill = row_fills.get(fill_key)
            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = fill
            for idx, value in enumerate(values, start=1):
                widths[idx] = min(max(widths.get(idx, 0), len(str(value))), 80)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = max(10, min(width + 2, 60))
    wb.properties.creator = "events-bot-new / Region Talk"
    wb.properties.title = "Region Talk MVP-1.x candidate report"
    wb.properties.subject = "Excel-safe openpyxl workbook"
    wb.save(path)


def write_xlsx(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    try:
        write_xlsx_openpyxl(path, sheets)
    except Exception as exc:
        print(f"[region-talk] openpyxl_xlsx_writer_failed fallback=minimal error={type(exc).__name__}: {str(exc)[:160]}", flush=True)
        write_xlsx_minimal(path, sheets)

def write_xlsx_minimal(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    names = list(sheets.keys())
    sheet_overrides = ''.join(
        f'<Override PartName="/xl/worksheets/sheet{i}.xml" '
        f'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        for i in range(1, len(names) + 1)
    )
    sheet_relationships = ''.join(
        f'<Relationship Id="rId{i}" '
        f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        f'Target="worksheets/sheet{i}.xml"/>'
        for i in range(1, len(names) + 1)
    )
    styles_rid = len(names) + 1
    created_at = escape(RUN_STARTED_AT.replace(microsecond=0).isoformat().replace('+00:00', 'Z'))
    with zipfile.ZipFile(path, 'w', compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr('[Content_Types].xml', f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/><Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/><Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>{sheet_overrides}</Types>''')
        z.writestr('_rels/.rels', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/><Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/></Relationships>''')
        z.writestr('docProps/core.xml', f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"><dc:creator>events-bot-new</dc:creator><dc:title>Region Talk MVP-1 candidate report</dc:title><dcterms:created xsi:type="dcterms:W3CDTF">{created_at}</dcterms:created><dcterms:modified xsi:type="dcterms:W3CDTF">{created_at}</dcterms:modified></cp:coreProperties>''')
        z.writestr('docProps/app.xml', f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"><Application>events-bot-new</Application><DocSecurity>0</DocSecurity><ScaleCrop>false</ScaleCrop><HeadingPairs><vt:vector size="2" baseType="variant"><vt:variant><vt:lpstr>Worksheets</vt:lpstr></vt:variant><vt:variant><vt:i4>{len(names)}</vt:i4></vt:variant></vt:vector></HeadingPairs><TitlesOfParts><vt:vector size="{len(names)}" baseType="lpstr">{''.join(f'<vt:lpstr>{escape(name[:31])}</vt:lpstr>' for name in names)}</vt:vector></TitlesOfParts></Properties>''')
        z.writestr('xl/styles.xml', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><fonts count="1"><font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/><scheme val="minor"/></font></fonts><fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills><borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders><cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs><cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs><cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles><dxfs count="0"/><tableStyles count="0" defaultTableStyle="TableStyleMedium2" defaultPivotStyle="PivotStyleLight16"/></styleSheet>''')
        z.writestr('xl/workbook.xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets>' + ''.join(f'<sheet name="{escape(name[:31])}" sheetId="{i}" r:id="rId{i}"/>' for i,name in enumerate(names,1)) + '</sheets></workbook>')
        z.writestr('xl/_rels/workbook.xml.rels', f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{sheet_relationships}<Relationship Id="rId{styles_rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>''')
        for i, name in enumerate(names, 1):
            z.writestr(f'xl/worksheets/sheet{i}.xml', sheet_xml(sheets[name]))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    headers = rows_headers(rows)
    with path.open('w', encoding='utf-8', newline='') as f:
        w=csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
        w.writeheader(); w.writerows(rows)


def render_md(payload: dict[str, Any]) -> str:
    s = payload.get('summary', {})
    return f"# Region Talk Candidate Report\n\n- run_id: `{payload.get('run_id')}`\n- posts_fetched: **{s.get('posts_fetched',0)}**\n- candidates_created: **{s.get('candidates_created',0)}**\n- favorites_created: **{s.get('favorites_created',0)}**\n- xlsx: `{payload.get('latest_xlsx') or payload.get('xlsx_path')}`\n"


def render_html(payload: dict[str, Any]) -> str:
    return '<!doctype html><meta charset="utf-8"><pre>' + html.escape(render_md(payload)) + '</pre>'


async def amain() -> int:
    load_dotenv(Path('.env'))
    config = load_split_runtime_from_kaggle_input()
    for k, v in (config.get('env') or {}).items():
        if v is not None:
            # The Kaggle runtime process can retain or receive environment
            # variables from an older pushed session. The run config dataset is
            # the authoritative per-run contract; use it to override non-secret
            # REGION_TALK/GOOGLE_AI controls while secrets remain loaded through
            # the encrypted bundle above.
            os.environ[str(k)] = str(v)
    run_id = str(config.get('run_id') or os.getenv('REGION_TALK_RUN_ID') or f"region-talk-{RUN_STARTED_AT.strftime('%Y%m%dT%H%M%SZ')}")
    os.environ["REGION_TALK_RUN_ID"] = run_id
    if not getenv_bool('REGION_TALK_DRY_RUN', True):
        raise RuntimeError('REGION_TALK_DRY_RUN=1 is required for MVP-1')
    if not getenv_bool('REGION_TALK_DISABLE_PUBLISH', True):
        raise RuntimeError('REGION_TALK_DISABLE_PUBLISH=1 is required for MVP-1')
    status = Status()
    status.event('kernel_started', phase='start', status='running', run_id=run_id)
    seed_path = find_seed_file(config)
    seeds = load_seeds(seed_path)
    out_dir = Path(os.getenv('REGION_TALK_OUTPUT_DIR') or str(config.get('output_dir') or f"artifacts/region-talk/runs/{run_id}"))
    ydb_cfg = ydb_config_status()
    status.event('preflight_ok', phase='preflight', status='running', run_id=run_id, seeds=len(seeds), seed_file=str(seed_path), dry_run=True, disable_publish=True, state_backend=region_talk_state_backend_requested(), ydb_config_status="missing_config" if ydb_cfg.get("missing") else "configured", ydb_missing=ydb_cfg.get("missing"), vk_token_kind=vk_wall_token_kind())
    source_rows, posts = await fetch_telegram_posts(seeds, status, out_dir)
    status.event('posts_fetched', phase='fetch', status='running', sources_scanned=len(source_rows), posts_fetched=len(posts))
    payload = build_report(seeds, source_rows, posts, run_id, out_dir, status=status)
    summary = payload.get('summary', {})
    status.event('report_written', phase='report', status='done', run_id=run_id, posts_fetched=summary.get('posts_fetched'), candidates_created=summary.get('candidates_created'), favorites_created=summary.get('favorites_created'), xlsx=str(payload.get('latest_xlsx')), state_backend=summary.get('state_backend'), ydb_read_status=summary.get('ydb_read_status'), ydb_write_status=summary.get('ydb_write_status'), ydb_state_mode=summary.get('ydb_state_mode'), vk_wall_probe_status=summary.get('vk_wall_probe_status'), sources_history_fetched_ok=summary.get('sources_history_fetched_ok'), current_run_reviewable_candidates=summary.get('current_run_reviewable_candidates'))
    return 0


if __name__ == '__main__':
    raise SystemExit(asyncio.run(amain()))
