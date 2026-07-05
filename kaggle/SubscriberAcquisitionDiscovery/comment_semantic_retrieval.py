from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import platform
import random
import re
import subprocess
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable

STAGE_NAME = "acq_comment_semantic_retrieval.v1"
DEFAULT_MODELS = ["intfloat/multilingual-e5-base", "BAAI/bge-m3"]
REQUIRED_MODELS = list(DEFAULT_MODELS)
DEFAULT_GATE_MODEL = "intfloat/multilingual-e5-base"
DEFAULT_SCORING_METHOD = "positive_negative_margin"

INTENT_SETS: dict[str, list[str]] = {
    "route_poi_far_context": [
        "люди обсуждают поездки по Калининградской области",
        "люди обсуждают интересные места региона",
        "люди обсуждают туризм в Калининградской области",
        "люди делятся впечатлениями о местах, куда можно съездить",
        "люди обсуждают красивые места, города, побережье, замки, форты или музеи",
        "люди обсуждают, стоит ли посещать разные места области",
        "люди сравнивают туристические места",
        "люди спрашивают мнение о местах перед поездкой",
        "люди обсуждают отдых за пределами Калининграда",
        "люди обсуждают короткие поездки по области",
    ],
    "route_poi_medium_interest": [
        "человек спрашивает, куда съездить из Калининграда",
        "человек ищет место для поездки на один день",
        "человек спрашивает, что посмотреть за день",
        "человек спрашивает, что посмотреть в городе области",
        "человек ищет маршрут по Калининградской области",
        "человек спрашивает, куда поехать на выходных",
        "человек спрашивает, какие места стоит посетить",
        "человек ищет достопримечательности рядом",
        "человек спрашивает, как добраться до интересного места",
        "человек хочет понять, стоит ли место поездки",
        "человек спрашивает, что посмотреть кроме главной достопримечательности",
        "человек ищет маршрут для прогулки или поездки",
    ],
    "route_poi_close_actionable": [
        "посоветуйте маршрут на один день из Калининграда",
        "куда поехать из Калининграда на электричке",
        "что посмотреть в Светлогорске за один день",
        "что посмотреть в Зеленоградске за один день",
        "что посмотреть в Балтийске за один день",
        "что посмотреть в Янтарном за один день",
        "что посмотреть в Черняховске за один день",
        "стоит ли ехать в конкретное место",
        "стоит ли смотреть этот замок, форт, кирху, музей или парк",
        "как добраться до этого места без машины",
        "что совместить с этим местом в одной поездке",
        "что посмотреть рядом с этим местом",
        "куда съездить с детьми по области",
        "куда съездить на машине на один день",
        "куда съездить на выходные недалеко от Калининграда",
    ],
    "event_far_context": [
        "люди обсуждают мероприятия и афишу",
        "люди обсуждают концерты, выставки, лекции, спектакли или фестивали",
        "люди интересуются культурными событиями",
        "люди выбирают, на какое мероприятие пойти",
        "люди обсуждают городские мероприятия",
        "люди спрашивают о событиях на выходные",
    ],
    "event_close_question": [
        "человек спрашивает, во сколько начинается мероприятие",
        "человек спрашивает, сколько длится мероприятие",
        "человек спрашивает, где проходит мероприятие",
        "человек спрашивает, есть ли билеты",
        "человек спрашивает, сколько стоит билет",
        "человек спрашивает, нужна ли регистрация",
        "человек спрашивает, можно ли прийти с детьми",
        "человек спрашивает, не отменили ли событие",
        "человек спрашивает, перенесли ли мероприятие",
        "человек спрашивает, как попасть на мероприятие",
        "человек спрашивает программу мероприятия",
        "человек спрашивает, будет ли мероприятие сегодня, завтра или в выходные",
    ],
    "organizer_comment_fit": [
        "в комментариях под постом организатора задают практические вопросы о событии",
        "люди уточняют возрастные ограничения события",
        "люди уточняют вход, регистрацию или билеты",
        "люди спрашивают, где встреча или вход",
        "люди спрашивают, что будет при плохой погоде",
        "люди спрашивают, подходит ли событие детям",
        "люди спрашивают про доступность, коляски или собак",
        "люди спрашивают, остались ли места",
        "люди спрашивают, будет ли запись или трансляция",
    ],
    "organizer_event_post_context": [
        "организаторский пост анонсирует событие и подходит для уточняющего вопроса",
        "анонс события содержит дату, место или программу, но может не содержать важные детали",
        "пост организатора про выставку, концерт, лекцию, фестиваль или встречу",
        "можно спросить организатора о возрасте, длительности, регистрации, входе или программе",
        "можно уточнить у организатора условия посещения события",
    ],
    "event_site_search_or_listing": [
        "человек ищет сайт с афишей мероприятий",
        "человек спрашивает где посмотреть календарь событий",
        "человек ищет список выставок или популярных мероприятий",
        "человек спрашивает где найти поиск по событиям",
    ],
    "organizer_submission_or_partnership": [
        "человек спрашивает куда прислать анонс события",
        "организатор спрашивает как добавить мероприятие в афишу",
        "человек спрашивает про публикацию события или информационное партнёрство",
    ],
    "badge_filter_need": [
        "человек ищет события по Пушкинской карте",
        "человек ищет бесплатные мероприятия",
        "человек спрашивает мероприятия для детей или семьи",
        "человек ищет благотворительные события",
        "человек спрашивает есть ли запись или трансляция события",
    ],
    "negative_intents": [
        "политический спор без вопроса о поездке, месте или событии",
        "общий флуд",
        "оскорбление или эмоциональный комментарий без полезного вопроса",
        "благодарность без запроса информации",
        "комментарий круто без намерения посетить",
        "комментарий не нравится без вопроса",
        "реклама без вопроса пользователя",
        "обсуждение бытовой темы без туристического или событийного смысла",
        "вопрос про аренду квартиры, покупку жилья, ипотеку или риэлтора без связи с событием или маршрутом",
        "вопрос про врача, клинику, лечение, анализы или запись к врачу без связи с событием или маршрутом",
        "обсуждение транспорта без связи с поездкой к месту",
        "обсуждение погоды без связи с поездкой или мероприятием",
        "объявление о потерянном телефоне, найденных вещах или пропаже без связи с событием",
        "обсуждение цены бензина, дизеля, топлива или заправок без маршрута поездки",
        "вопрос где есть бензин, дизель или топливо на заправках",
        "вопрос о наличии бензина, дизеля или топлива на азс",
        "обсуждение выборов, политики или чиновников без вопроса о событии",
        "реклама товаров с маркетплейса, артикулом, wildberries или ссылкой на товар",
        "реклама помощи студентам, учебных работ, экзаменов, зачётов или услуг",
        "реклама психологических, эзотерических или системных расстановок без события",
        "жалоба на ремонт дороги, трамвай, остановку или городское благоустройство без события",
        "новостной пост о погоде, температуре моря или происшествии без вопроса пользователя",
        "вопрос не связан с местом, маршрутом, достопримечательностью или мероприятием",
    ],
}

REGION_POSITIVE_PHRASES: list[str] = [
    "площадка и комментарии относятся к Калининградской области",
    "события проходят в Калининграде или городах Калининградской области",
    "маршрут или поездка по Калининградской области",
    "вопрос о том, куда съездить из Калининграда на один день",
    "обсуждают Светлогорск, Зеленоградск, Балтийск, Янтарный или Куршскую косу",
    "обсуждают форты, кирхи, замки, музеи и побережье Калининградской области",
    "локальная афиша Калининграда, областные мероприятия и городские события",
    "пост или комментарий организатора события в Калининградской области",
]

REGION_NEGATIVE_PHRASES: list[str] = [
    "площадка относится к другому региону и не к Калининградской области",
    "маршрут, событие или вопрос про Москву, Санкт-Петербург, Беларусь или другой регион",
    "городская афиша другого города без связи с Калининградской областью",
    "поездка или туристический совет по другому региону",
    "пост о зарубежном или российском регионе вне Калининградской области",
]

POSITIVE_INTENT_SETS = [name for name in INTENT_SETS if name != "negative_intents"]
ROUTE_INTENT_SETS = {"route_poi_far_context", "route_poi_medium_interest", "route_poi_close_actionable"}
EVENT_INTENT_SETS = {"event_far_context", "event_close_question", "event_site_search_or_listing", "badge_filter_need"}
ORGANIZER_INTENT_SETS = {"organizer_comment_fit", "organizer_event_post_context", "organizer_submission_or_partnership"}
EVENT_FUTURE_REQUIRED_ACTIONS = {
    "event_recommendation_reply",
    "organizer_visibility_clarification",
    "event_site_search_or_listing",
    "badge_filter_need",
}
REGION_REQUIRED_ACTIONS = {
    "trip_route_poi_recommendation",
    "event_recommendation_reply",
    "organizer_visibility_clarification",
    "event_site_search_or_listing",
    "organizer_submission_or_partnership",
    "badge_filter_need",
}
REGION_ELIGIBLE_CONFIDENCES = {"confirmed", "probable"}

_TOKEN_RE = re.compile(r"[\wА-Яа-яЁё]+", re.U)
_HTML_RE = re.compile(r"<[^>]+>")
_CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_ROUTE_DEST_RE = re.compile(
    r"(?i)\b(светлогорск\w*|зеленоградск\w*|балтийск\w*|янтарн\w*|черняховск\w*|советск\w*|гусев\w*|куршск\w*\s+кос\w*|зам\w+|форт\w*|кирх\w*|побереж\w+)\b"
)
_TRANSPORT_RE = re.compile(r"(?i)\b(электричк\w*|поезд\w*|автобус\w*|машин\w*|без\s+машин\w*)\b")
_QUESTION_SIGNAL_RE = re.compile(
    r"(?i)(\?|"
    r"\b(?:куда|где|когда|как|что|сколько|какой|какая|какие|какое|зачем|почему)\b|"
    r"\b(?:подскажите|посоветуйте|посоветуй|ищу|ищем|интересует|нужен|нужна|нужно)\b|"
    r"\b(?:есть\s+ли|будет\s+ли|можно\s+ли|нужна\s+ли|нужно\s+ли|кто\s+знает)\b)"
)
_OFFER_NOISE_RE = re.compile(
    r"(?i)\b("
    r"сохраняйте|записывайтесь|приглашаем|предлагаем|предлагаю|приходите|жд[её]м\s+вас|"
    r"бронируйте|бронь|забронировать|купить\s+билет|стоимость|цена|скидк\w*|акци\w*|"
    r"в\s+наличии|прода[её]тся|продам|сдам|аренда|работа|ваканси\w*|резюме|"
    r"экскурси\w*\s+(?:по|на|в)|тур(?:ы|ов|ом)?|заезды|каждую\s+неделю|"
    r"пишите\s+(?:в\s+)?(?:лс|личку|директ)|whatsapp|ватсап|тел\.|телефон|подробности\s+по\s+ссылке"
    r")\b"
)
_URL_OR_MENTION_RE = re.compile(r"(?i)(https?://|t\.me/|vk\.com/|@[A-Za-z0-9_]{4,})")
_ROUTE_CONTEXT_RE = re.compile(
    r"(?i)\b(куда\s+(?:съездить|поехать|сходить)|что\s+посмотреть|маршрут|достопримечательн\w*|"
    r"посетить|светлогорск\w*|зеленоградск\w*|балтийск\w*|янтарн\w*|черняховск\w*|"
    r"куршск\w*\s+кос\w*|зам\w+|форт\w*|кирх\w*|электричк\w*|автобус\w*)\b"
)
_EVENT_CONTEXT_RE = re.compile(
    r"(?i)\b(мероприяти\w*|событи\w*|афиш\w*|концерт\w*|выставк\w*|спектакл\w*|"
    r"фестивал\w*|лекци\w*|мастер[- ]?класс\w*|билет\w*|регистрац\w*|начина\w*|"
    r"начало|длится|возрастн\w*|вход|программ\w*|запись|трансляц\w*|"
    r"куда\s+сходить|выходн\w*)\b"
)
_ORGANIZER_SUBMIT_CONTEXT_RE = re.compile(r"(?i)\b(афиш\w*|анонс\w*|мероприяти\w*|событи\w*|добавить|прислать|опубликовать|партн[её]рств\w*)\b")
_BADGE_CONTEXT_RE = re.compile(r"(?i)\b(пушкинск\w*|бесплатн\w*|дет\w*|семейн\w*|льгот\w*|инвалид\w*|доступн\w*|запись|трансляц\w*)\b")
_REAL_ESTATE_SCOPE_RE = re.compile(
    r"(?i)\b(недвижимост\w*|квартир\w*|комнат\w*|аренд\w*|снять\s+жиль[её]|сниму|сда[её]тся|"
    r"жк|ипотек\w*|риэлт\w*|новостройк\w*|застройщик\w*)\b"
)
_MEDICINE_SCOPE_RE = re.compile(
    r"(?i)\b(врач\w*|клиник\w*|стоматолог\w*|лор\b|педиатр\w*|терапевт\w*|поликлиник\w*|"
    r"больниц\w*|лечени\w*|анализ\w*|мрт\b|узи\b|при[её]м\s+врач\w*|запис\w*\s+к\s+врач)\b"
)
_GASOLINE_AVAILABILITY_SCOPE_RE = re.compile(
    r"(?i)("
    r"\b(?:где|есть\s+ли|подскажите|кто\s+знает|наличи\w*|появил\w*|остал\w*|ищу|найти|найд[её]тся)"
    r".{0,90}\b(?:бензин\w*|дизел\w*|топлив\w*|азс\b|заправк\w*)\b|"
    r"\b(?:бензин\w*|дизел\w*|топлив\w*|азс\b|заправк\w*)\b"
    r".{0,90}\b(?:где|есть|наличи\w*|появил\w*|остал\w*|законч\w*|найти|подскажите)\b"
    r")"
)
_SURFACE_MEDICINE_SCOPE_RE = re.compile(
    r"(?i)\b(врач\w*|клиник\w*|стоматолог\w*|педиатр\w*|здоровь\w*|симптом\w*|"
    r"грудн\w*\s+вскармливан\w*|гв\b|лечени\w*|психолог\w*|диет\w*|худе\w*)\b"
)
_RU_MONTHS = {
    "января": 1,
    "январь": 1,
    "февраля": 2,
    "февраль": 2,
    "марта": 3,
    "март": 3,
    "апреля": 4,
    "апрель": 4,
    "мая": 5,
    "май": 5,
    "июня": 6,
    "июнь": 6,
    "июля": 7,
    "июль": 7,
    "августа": 8,
    "август": 8,
    "сентября": 9,
    "сентябрь": 9,
    "октября": 10,
    "октябрь": 10,
    "ноября": 11,
    "ноябрь": 11,
    "декабря": 12,
    "декабрь": 12,
}
_NUMERIC_DATE_RE = re.compile(r"(?<!\d)(\d{1,2})[./](\d{1,2})(?:[./](\d{2,4}))?(?!\d)")
_TEXT_DATE_RE = re.compile(
    r"(?i)(?:(\d{1,2})\s*(?:-|–|—|по)\s*)?(\d{1,2})\s+"
    r"(января|январь|февраля|февраль|марта|март|апреля|апрель|мая|май|июня|июнь|июля|июль|"
    r"августа|август|сентября|сентябрь|октября|октябрь|ноября|ноябрь|декабря|декабрь)"
    r"(?:\s+(\d{4}))?"
)
_RELATIVE_TOMORROW_RE = re.compile(r"(?i)\b(?:завтра|послезавтра)\b")
_RELATIVE_TODAY_RE = re.compile(r"(?i)\bсегодня\b")
_PAST_EVENT_SIGNAL_RE = re.compile(
    r"(?i)\b("
    r"прош[её]л|прошла|прошли|состоял[ао]сь|состоялись|прошедш\w*|завершил[ао]сь|завершились|"
    r"стартовал\w*|начал[ао]?сь|начались|открыл\w*|"
    r"отч[её]т|фотоотч[её]т|итоги|после\s+мероприяти\w*|спасибо\s+организатор\w*|"
    r"поблагодар\w*|был[аио]?\s+на|были\s+на|кто\s+был\s+на|посетили|отметили|отпраздновали"
    r")\b"
)
_FUTURE_EVENT_SIGNAL_RE = re.compile(
    r"(?i)\b("
    r"состоится|пройд[её]т|будет|начн[её]тся|приглашаем|приходите|жд[её]м|регистрация\s+открыта|"
    r"анонс|афиша|запланирован\w*|откроется|стартует"
    r")\b"
)
_KGD_REGION_HINT_RE = re.compile(
    r"(?i)("
    r"калининград\w*|к[её]ниг\w*|kenig\w*|koenig\w*|\bkgd\b|\bkld\w*|(?<!\d)39(?!\d)|"
    r"светлогорск\w*|зеленоградск\w*|балтийск\w*|янтарн\w*|черняховск\w*|советск\w*|гусев\w*|"
    r"гурьевск\w*|гвардейск\w*|пионерск\w*|полесск\w*|правдинск\w*|багратионовск\w*|"
    r"мамоново|ладушкин\w*|неман\w*|славск\w*|краснознаменск\w*|оз[её]рск\w*|"
    r"куршск\w*\s+кос\w*|балтийск\w*\s+кос\w*|остров\s+канта|\bкант(?:а|у|е)?\b|"
    r"кафедральн\w*\s+собор|музе[йя]\s+янтар\w*|рыбн\w+\s+деревн\w*|"
    r"форт\s*(?:№|#)?\s*(?:1|2|3|4|5|11)|форт\s+д[её]нхофф|д[её]нхофф|"
    r"закхаймск\w*\s+ворот|фридландск\w*\s+ворот|железнодорожн\w*\s+ворот|"
    r"бранденбургск\w*\s+ворот|инстербург|тапиау|пиллау|раушен|кранц"
    r")"
)
_OUT_OF_REGION_HINT_RE = re.compile(
    r"(?i)("
    r"navahrudak|novogrud|новогруд\w*|минск\w*|беларус\w*|"
    r"москв\w*|санкт[-\s]?петербург\w*|\bспб\b|ленинградск\w+\s+област\w*|"
    r"нижн\w+\s+новгород\w*|казан\w*|соч\w*|краснодар\w*|ростов\w*|"
    r"екатеринбург\w*|новосибирск\w*|самар\w*|владивосток\w*|перм\w*|"
    r"турци\w*|грузи\w*"
    r")"
)


def truthy(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def semantic_retrieval_enabled() -> bool:
    return truthy(os.getenv("ACQ_ENABLE_COMMENT_SEMANTIC_RETRIEVAL"))


def _json_env(name: str, default: Any) -> Any:
    raw = (os.getenv(name) or "").strip()
    if not raw:
        return default
    try:
        return json.loads(raw)
    except Exception:
        return default


def _int_env(name: str, default: int, *, min_value: int | None = None) -> int:
    try:
        value = int(os.getenv(name) or default)
    except Exception:
        value = int(default)
    if min_value is not None:
        value = max(min_value, value)
    return value


def _float_env(name: str, default: float, *, min_value: float | None = None) -> float:
    try:
        value = float(os.getenv(name) or default)
    except Exception:
        value = float(default)
    if min_value is not None:
        value = max(min_value, value)
    return value


def _max_comment_age_days() -> int:
    return _int_env("ACQ_COMMENT_RETRIEVAL_MAX_COMMENT_AGE_DAYS", 365, min_value=1)


def _stale_activity_days() -> int:
    return _int_env("ACQ_COMMENT_RETRIEVAL_STALE_ACTIVITY_DAYS", 92, min_value=1)


def normalize_comment_text(text: str) -> str:
    cleaned = _CONTROL_RE.sub(" ", str(text or ""))
    cleaned = _HTML_RE.sub(" ", cleaned)
    return " ".join(cleaned.split())


def _prefix_text(model_name: str, text: str, *, is_query: bool) -> str:
    if "multilingual-e5" in model_name.lower():
        return ("query: " if is_query else "passage: ") + text
    return text


class HashingEmbeddingBackend:
    """Small deterministic backend for tests/offline smoke; not used for quality claims."""

    def __init__(self, dim: int = 96) -> None:
        self.dim = dim

    def encode(self, texts: list[str], *, model_name: str, is_query: bool, batch_size: int, max_length: int) -> list[list[float]]:
        out: list[list[float]] = []
        for text in texts:
            vec = [0.0] * self.dim
            for token in _TOKEN_RE.findall(text.lower()):
                digest = hashlib.blake2b(token.encode("utf-8"), digest_size=8).digest()
                idx = int.from_bytes(digest[:4], "little") % self.dim
                sign = -1.0 if digest[4] & 1 else 1.0
                vec[idx] += sign
            out.append(_normalize(vec))
        return out


class SentenceTransformerBackend:
    def __init__(self) -> None:
        self._models: dict[str, Any] = {}

    def _ensure_libs(self) -> None:
        try:
            import sentence_transformers  # noqa: F401
        except ImportError:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "sentence-transformers"])

    def _model(self, model_name: str) -> Any:
        self._ensure_libs()
        if model_name not in self._models:
            from sentence_transformers import SentenceTransformer

            self._models[model_name] = SentenceTransformer(model_name, device=os.getenv("ACQ_COMMENT_RETRIEVAL_DEVICE") or None)
        return self._models[model_name]

    def encode(self, texts: list[str], *, model_name: str, is_query: bool, batch_size: int, max_length: int) -> Any:
        model = self._model(model_name)
        prepared = [_prefix_text(model_name, text, is_query=is_query) for text in texts]
        # SentenceTransformer accepts max_seq_length as a mutable property.
        try:
            model.max_seq_length = int(max_length)
        except Exception:
            pass
        return model.encode(
            prepared,
            batch_size=batch_size,
            normalize_embeddings=True,
            convert_to_numpy=True,
            show_progress_bar=False,
        )


def _normalize(vec: list[float]) -> list[float]:
    norm = math.sqrt(sum(float(x) * float(x) for x in vec))
    if norm <= 0:
        return vec
    return [float(x) / norm for x in vec]


def _dot(a: Any, b: Any) -> float:
    return float(sum(float(x) * float(y) for x, y in zip(a, b)))


def _to_list_matrix(matrix: Any) -> list[list[float]]:
    if hasattr(matrix, "tolist"):
        return matrix.tolist()
    return [list(row) for row in matrix]


def _peak_ram_mb() -> float | None:
    try:
        import resource

        value = float(resource.getrusage(resource.RUSAGE_SELF).ru_maxrss)
        # Linux reports KB, macOS bytes. Kaggle/Linux path is KB.
        return value / 1024.0 if value > 10_000 else value / (1024.0 * 1024.0)
    except Exception:
        return None


def _percentile(values: list[float], q: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    if len(ordered) == 1:
        return float(ordered[0])
    pos = (len(ordered) - 1) * q
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return float(ordered[lo])
    return float(ordered[lo] * (hi - pos) + ordered[hi] * (pos - lo))


def _mean(values: list[float]) -> float:
    return float(sum(values) / len(values)) if values else 0.0


def _action_for_intent(intent_set: str) -> str:
    if intent_set in ROUTE_INTENT_SETS:
        return "trip_route_poi_recommendation"
    if intent_set in {"organizer_comment_fit", "organizer_event_post_context"}:
        return "organizer_visibility_clarification"
    if intent_set == "organizer_submission_or_partnership":
        return "organizer_submission_or_partnership"
    if intent_set == "event_site_search_or_listing":
        return "event_site_search_or_listing"
    if intent_set == "badge_filter_need":
        return "badge_filter_need"
    return "event_recommendation_reply"


def _route_target_hint(text: str, intent_set: str) -> dict[str, Any]:
    if intent_set not in ROUTE_INTENT_SETS:
        return {"route_target_status": "not_applicable", "destination_hint": "", "poi_hints": [], "transport_hint": "unknown", "event_ids": []}
    destinations = [m.group(0) for m in _ROUTE_DEST_RE.finditer(text or "")]
    transport = "unknown"
    transport_match = _TRANSPORT_RE.search(text or "")
    if transport_match:
        raw = transport_match.group(0).lower()
        if "электр" in raw or "поезд" in raw:
            transport = "train"
        elif "автоб" in raw:
            transport = "bus"
        elif "машин" in raw:
            transport = "car" if "без" not in raw else "unknown"
    return {
        "route_target_status": "route_needed",
        "destination_hint": destinations[0] if destinations else "",
        "poi_hints": destinations[:5],
        "transport_hint": transport,
        "event_ids": [],
    }


def _text_quality_features(text: str) -> dict[str, Any]:
    compact = normalize_comment_text(text)
    tokens = _TOKEN_RE.findall(compact)
    question_signal = bool(_QUESTION_SIGNAL_RE.search(compact))
    offer_signal = bool(_OFFER_NOISE_RE.search(compact))
    link_signal = bool(_URL_OR_MENTION_RE.search(compact))
    # URLs/mentions without a question are commonly ads, cross-posts or source
    # announcements. Keep them in artifacts, but do not spend LLM/reply budget.
    link_offer_signal = link_signal and not question_signal and len(tokens) >= 6
    too_short_statement = len(tokens) <= 2 and not question_signal
    if offer_signal:
        noise_type = "explicit_offer_or_ad"
    elif link_offer_signal:
        noise_type = "link_or_crosspost_without_question"
    elif too_short_statement:
        noise_type = "too_short_non_question"
    elif not question_signal:
        noise_type = "non_question_statement"
    else:
        noise_type = ""
    hard_noise = noise_type in {"explicit_offer_or_ad", "link_or_crosspost_without_question", "too_short_non_question"}
    return {
        "question_signal": question_signal,
        "offer_signal": offer_signal,
        "link_signal": link_signal,
        "noise_type": noise_type,
        "hard_noise": hard_noise,
        "token_count": len(tokens),
    }


def _out_of_scope_noise_type(text: str) -> str:
    compact = normalize_comment_text(text)
    if not compact:
        return ""
    if _GASOLINE_AVAILABILITY_SCOPE_RE.search(compact):
        return "out_of_scope_gasoline_availability"
    # Keep mixed questions when they also contain explicit route/event context:
    # "снимаем квартиру, куда съездить?" is a route question, but "где снять
    # квартиру?" is out of acquisition scope.
    has_our_context = bool(_ROUTE_CONTEXT_RE.search(compact) or _EVENT_CONTEXT_RE.search(compact) or _ORGANIZER_SUBMIT_CONTEXT_RE.search(compact))
    if has_our_context:
        return ""
    if _REAL_ESTATE_SCOPE_RE.search(compact):
        return "out_of_scope_real_estate"
    if _MEDICINE_SCOPE_RE.search(compact):
        return "out_of_scope_medicine"
    return ""


def _surface_out_of_scope_type(profile: dict[str, Any]) -> str:
    text = " ".join(str(profile.get(key) or "") for key in ["surface_title", "surface_url", "surface_key", "surface_type"])
    if not text:
        return ""
    has_our_context = bool(_ROUTE_CONTEXT_RE.search(text) or _EVENT_CONTEXT_RE.search(text) or _ORGANIZER_SUBMIT_CONTEXT_RE.search(text))
    if has_our_context:
        return ""
    if _REAL_ESTATE_SCOPE_RE.search(text):
        return "out_of_scope_real_estate_surface"
    if _SURFACE_MEDICINE_SCOPE_RE.search(text):
        return "out_of_scope_medicine_surface"
    return ""


def _first_match(pattern: re.Pattern[str], text: str) -> str:
    match = pattern.search(text or "")
    return match.group(0) if match else ""


def _retrieval_now() -> datetime:
    raw = (
        os.getenv("ACQ_COMMENT_RETRIEVAL_NOW_ISO")
        or os.getenv("KAGGLE_RUN_STARTED_AT")
        or os.getenv("ACQ_RUN_STARTED_AT")
        or ""
    ).strip()
    if raw:
        try:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except Exception:
            pass
    return datetime.now(timezone.utc)


def _normalize_event_year(year_raw: str | None, *, reference_year: int) -> int:
    if not year_raw:
        return reference_year
    year = int(year_raw)
    if year < 100:
        return 2000 + year
    return year


def _safe_date(year: int, month: int, day: int) -> datetime | None:
    try:
        return datetime(year, month, day, tzinfo=timezone.utc)
    except Exception:
        return None


def _extract_event_dates(text: str, *, reference: datetime) -> list[datetime]:
    compact = normalize_comment_text(text)
    dates: list[datetime] = []
    for match in _NUMERIC_DATE_RE.finditer(compact):
        day = int(match.group(1))
        month = int(match.group(2))
        year = _normalize_event_year(match.group(3), reference_year=reference.year)
        dt = _safe_date(year, month, day)
        if dt is not None:
            dates.append(dt)
    for match in _TEXT_DATE_RE.finditer(compact):
        start_day_raw, end_day_raw, month_raw, year_raw = match.groups()
        month = _RU_MONTHS.get(str(month_raw).casefold())
        if not month:
            continue
        year = _normalize_event_year(year_raw, reference_year=reference.year)
        for day_raw in [start_day_raw, end_day_raw]:
            if not day_raw:
                continue
            dt = _safe_date(year, month, int(day_raw))
            if dt is not None:
                dates.append(dt)
    base = reference.date()
    if _RELATIVE_TODAY_RE.search(compact):
        dates.append(datetime(base.year, base.month, base.day, tzinfo=timezone.utc))
    for match in _RELATIVE_TOMORROW_RE.finditer(compact):
        delta_days = 2 if match.group(0).casefold() == "послезавтра" else 1
        target = base.toordinal() + delta_days
        target_date = datetime.fromordinal(target)
        dates.append(target_date.replace(tzinfo=timezone.utc))
    # Deduplicate while preserving deterministic order.
    seen: set[str] = set()
    out: list[datetime] = []
    for dt in dates:
        key = dt.date().isoformat()
        if key in seen:
            continue
        seen.add(key)
        out.append(dt)
    return out


def _event_temporal_text(row: dict[str, Any]) -> str:
    return normalize_comment_text(" ".join(
        str(row.get(key) or "")
        for key in [
            "text",
            "text_snapshot",
            "source_post_text_snapshot",
            "source_post_text",
            "reply_parent_text_snapshot",
            "reply_parent_text",
            "analysis_context_snapshot",
            "analysis_text",
        ]
    ))


def _event_temporal_assessment(row: dict[str, Any], *, now: datetime | None = None) -> dict[str, Any]:
    action = str(row.get("candidate_action_type") or _action_for_intent(str(row.get("intent_set") or "")))
    if action not in EVENT_FUTURE_REQUIRED_ACTIONS:
        return {
            "event_temporal_status": "not_required",
            "event_temporal_evidence_ru": "для этой цели не требуется проверка будущей даты события",
            "event_latest_detected_date": "",
            "event_temporal_gate_passed": True,
        }
    now = (now or _retrieval_now()).astimezone(timezone.utc)
    reference = _parse_created_at(row.get("created_at")) or now
    text = _event_temporal_text(row)
    dates = _extract_event_dates(text, reference=reference)
    has_past = bool(_PAST_EVENT_SIGNAL_RE.search(text))
    has_future = bool(_FUTURE_EVENT_SIGNAL_RE.search(text))
    if dates:
        latest = max(dates)
        if latest.date() < now.date():
            return {
                "event_temporal_status": "past_event",
                "event_temporal_evidence_ru": f"найдена дата события {latest.strftime('%d.%m.%Y')}, она раньше даты запуска {now.strftime('%d.%m.%Y')}",
                "event_latest_detected_date": latest.date().isoformat(),
                "event_temporal_gate_passed": False,
            }
        if has_past and latest.date() <= now.date():
            return {
                "event_temporal_status": "past_event_signal",
                "event_temporal_evidence_ru": (
                    "в контексте есть формулировка прошедшего/уже стартовавшего события; "
                    "для acquisition берём только будущие анонсы и вопросы"
                ),
                "event_latest_detected_date": latest.date().isoformat(),
                "event_temporal_gate_passed": False,
            }
        return {
            "event_temporal_status": "future_or_today",
            "event_temporal_evidence_ru": f"найдена дата события {latest.strftime('%d.%m.%Y')}, она не раньше даты запуска {now.strftime('%d.%m.%Y')}",
            "event_latest_detected_date": latest.date().isoformat(),
            "event_temporal_gate_passed": True,
        }
    if has_past:
        return {
            "event_temporal_status": "past_event_signal",
            "event_temporal_evidence_ru": "по тексту похоже на отчёт/старт/итоги уже прошедшего или начавшегося события",
            "event_latest_detected_date": "",
            "event_temporal_gate_passed": False,
        }
    if has_future:
        return {
            "event_temporal_status": "future_signal_no_date",
            "event_temporal_evidence_ru": "есть смысловой сигнал будущего/анонсируемого события, но точная дата не извлечена",
            "event_latest_detected_date": "",
            "event_temporal_gate_passed": True,
        }
    return {
        "event_temporal_status": "unknown_event_time",
        "event_temporal_evidence_ru": "дата события не найдена; строка остаётся кандидатом только до финальной LLM-проверки",
        "event_latest_detected_date": "",
        "event_temporal_gate_passed": True,
    }


def _hard_semantic_rejected(row: dict[str, Any]) -> bool:
    if _truthy_value(row.get("semantic_candidate_rejected")):
        return True
    exclusion = str(row.get("semantic_exclusion_type") or row.get("candidate_noise_type") or "")
    return exclusion in {"past_event", "past_event_signal", "out_of_scope_gasoline_availability"}


def _region_positive_min_score() -> float:
    return _float_env("ACQ_COMMENT_RETRIEVAL_REGION_POSITIVE_MIN_SCORE", 0.12, min_value=0.0)


def _region_margin_min_score() -> float:
    return _float_env("ACQ_COMMENT_RETRIEVAL_REGION_MARGIN_MIN_SCORE", 0.06, min_value=0.0)


def _record_surface_region_text(record: dict[str, Any]) -> str:
    return normalize_comment_text(" ".join(
        str(record.get(key) or "")
        for key in [
            "surface_title",
            "title",
            "source_context_title",
            "surface_url",
            "surface_key",
            "surface_external_id",
            "surface_type",
            "platform",
            "source",
            "discovery_source_context",
        ]
    ))


def _record_content_region_text(record: dict[str, Any]) -> str:
    return normalize_comment_text(" ".join(
        str(record.get(key) or "")
        for key in [
            "text",
            "text_snapshot",
            "source_post_text_snapshot",
            "source_post_text",
            "reply_parent_text_snapshot",
            "reply_parent_text",
            "analysis_text",
        ]
    ))


def _record_region_text(record: dict[str, Any]) -> str:
    """Text embedded by the dedicated region gate.

    The region gate deliberately receives a wider context than the action-intent
    scorer: platform title/URL/handle + current text + source post + reply
    parent.  This lets a generic question like “куда съездить?” pass only when
    the surrounding surface/thread proves Kaliningrad Oblast, and keeps the
    action scorer itself free from a broad regex prefilter.
    """
    surface_text = _record_surface_region_text(record)
    content_text = _record_content_region_text(record)
    parts: list[str] = []
    if surface_text:
        parts.append(f"Площадка/метаданные: {surface_text[:500]}")
    if content_text:
        parts.append(f"Текст и контекст: {content_text[:1400]}")
    return "\n".join(parts)[:1900]


def _score_region_context(
    region_vec: Any,
    positive_vectors: list[Any],
    negative_vectors: list[Any],
) -> dict[str, Any]:
    positive_scores = [_dot(region_vec, vec) for vec in positive_vectors]
    negative_scores = [_dot(region_vec, vec) for vec in negative_vectors]
    positive_score = max(positive_scores) if positive_scores else 0.0
    negative_score = max(negative_scores) if negative_scores else 0.0
    positive_idx = positive_scores.index(positive_score) if positive_scores else -1
    negative_idx = negative_scores.index(negative_score) if negative_scores else -1
    return {
        "region_positive_score": float(positive_score),
        "region_negative_score": float(negative_score),
        "region_score": float(positive_score - negative_score),
        "top_region_phrase": REGION_POSITIVE_PHRASES[positive_idx] if 0 <= positive_idx < len(REGION_POSITIVE_PHRASES) else "",
        "top_negative_region_phrase": REGION_NEGATIVE_PHRASES[negative_idx] if 0 <= negative_idx < len(REGION_NEGATIVE_PHRASES) else "",
    }


def _assess_record_region(record: dict[str, Any], vector_scores: dict[str, Any] | None = None) -> dict[str, Any]:
    vector_scores = dict(vector_scores or {})
    surface_text = _record_surface_region_text(record)
    content_text = _record_content_region_text(record)
    all_text = normalize_comment_text(" ".join([surface_text, content_text]))
    positive_surface_hint = _first_match(_KGD_REGION_HINT_RE, surface_text)
    positive_content_hint = _first_match(_KGD_REGION_HINT_RE, content_text)
    if re.fullmatch(r"(?i)(39|\bkld\w*)", positive_content_hint or ""):
        # `39`/`kld*` are reliable in handles/URLs, but too noisy inside an
        # arbitrary comment body (age, bus number, random code).
        positive_content_hint = ""
    negative_hint = _first_match(_OUT_OF_REGION_HINT_RE, all_text)
    positive_hint = positive_surface_hint or positive_content_hint
    pos_score = float(vector_scores.get("region_positive_score") or 0.0)
    neg_score = float(vector_scores.get("region_negative_score") or 0.0)
    margin = pos_score - neg_score
    semantic_positive = pos_score >= _region_positive_min_score() and margin >= _region_margin_min_score()
    semantic_negative = neg_score >= _region_positive_min_score() and (neg_score - pos_score) >= _region_margin_min_score()

    if positive_hint:
        confidence = "confirmed"
        source = "surface_metadata_keyword" if positive_surface_hint else "comment_context_keyword"
        evidence = f"есть явная зацепка Калининградской области: «{positive_hint}»"
    elif negative_hint and not semantic_positive:
        confidence = "out_of_region"
        source = "out_of_region_keyword"
        evidence = f"есть зацепка другого региона без Калининградского контекста: «{negative_hint}»"
    elif semantic_positive:
        confidence = "probable"
        source = "semantic_region_vector"
        evidence = f"регион вероятен по embedding-сходству: {vector_scores.get('top_region_phrase') or 'Калининградская область'}"
    elif semantic_negative:
        confidence = "out_of_region"
        source = "semantic_out_of_region_vector"
        evidence = f"похоже на другой регион по embedding-сходству: {vector_scores.get('top_negative_region_phrase') or 'другой регион'}"
    else:
        confidence = "unknown"
        source = "no_region_signal"
        evidence = "нет достаточной зацепки, что вопрос/пост относится к Калининградской области"

    status = "region_ok" if confidence in REGION_ELIGIBLE_CONFIDENCES else f"region_{confidence}"
    return {
        **vector_scores,
        "region_confidence": confidence,
        "region_gate_status": status,
        "region_signal_source": source,
        "region_evidence_ru": evidence,
        "region_positive_hint": positive_hint,
        "region_out_of_region_hint": negative_hint if not positive_hint else "",
    }


def _candidate_region_required(row: dict[str, Any]) -> bool:
    action = str(row.get("candidate_action_type") or _action_for_intent(str(row.get("intent_set") or "")))
    return action in REGION_REQUIRED_ACTIONS


def _candidate_region_eligible(row: dict[str, Any]) -> bool:
    if not _candidate_region_required(row):
        return True
    return str(row.get("region_confidence") or "") in REGION_ELIGIBLE_CONFIDENCES


def _region_catalog_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for catalog_type, phrases in [("positive_kaliningrad_oblast", REGION_POSITIVE_PHRASES), ("negative_other_region", REGION_NEGATIVE_PHRASES)]:
        for idx, phrase in enumerate(phrases, start=1):
            rows.append({
                "region_catalog_type": catalog_type,
                "phrase_order": idx,
                "model_phrase": phrase,
                "is_negative": catalog_type.startswith("negative"),
                "note_ru": (
                    "Отдельный региональный embedding-gate: площадка/пост/комментарий должны быть про Калининградскую область; "
                    "unknown/out_of_region не попадают в monitoring/LLM-кандидаты."
                ),
            })
    return rows


def _intent_has_text_support(text: str, intent_set: str) -> bool:
    compact = normalize_comment_text(text)
    if intent_set in ROUTE_INTENT_SETS:
        return bool(_ROUTE_CONTEXT_RE.search(compact))
    if intent_set in {"event_far_context", "event_close_question", "event_site_search_or_listing"}:
        return bool(_EVENT_CONTEXT_RE.search(compact))
    if intent_set in {"organizer_comment_fit", "organizer_event_post_context"}:
        # Generic questions like “как зовут детей?” must not become organizer
        # clarification candidates without event/logistics context.
        return bool(_EVENT_CONTEXT_RE.search(compact))
    if intent_set == "organizer_submission_or_partnership":
        return bool(_ORGANIZER_SUBMIT_CONTEXT_RE.search(compact))
    if intent_set == "badge_filter_need":
        return bool(_BADGE_CONTEXT_RE.search(compact))
    return True


def _deterministic_prefilter_enabled() -> bool:
    # Default is deliberately false. In the vector-scan funnel every collected
    # comment/post context must be embedded and the LLM gate must receive the
    # top semantic rows, not a regex/question-signal shortlist.
    return truthy(os.getenv("ACQ_COMMENT_RETRIEVAL_DETERMINISTIC_PREFILTER"))


def _apply_text_quality_to_candidate(row: dict[str, Any], *, scoring_method: str) -> None:
    features = _text_quality_features(str(row.get("text") or row.get("text_snapshot") or ""))
    intent_supported = _intent_has_text_support(str(row.get("text") or row.get("text_snapshot") or ""), str(row.get("intent_set") or ""))
    out_of_scope = _out_of_scope_noise_type(str(row.get("text") or row.get("text_snapshot") or ""))
    context_out_of_scope = _out_of_scope_noise_type(_event_temporal_text(row))
    if context_out_of_scope and not out_of_scope:
        out_of_scope = context_out_of_scope
    temporal = _event_temporal_assessment(row)
    semantic_exclusion = ""
    if out_of_scope == "out_of_scope_gasoline_availability":
        semantic_exclusion = out_of_scope
    elif not temporal.get("event_temporal_gate_passed"):
        semantic_exclusion = str(temporal.get("event_temporal_status") or "past_event")
    raw_score = float(row.get(scoring_method) or row.get("score") or 0.0)
    relation = str(row.get("relation") or "").strip()
    source_post_relation = relation in {"vk_social_wall_post", "tg_channel_post_context"} or bool(row.get("is_post"))
    source_post_allowed = truthy(os.getenv("ACQ_ALLOW_SOURCE_POST_REPLY_CANDIDATES"))
    diagnostic_noise_type = features["noise_type"]
    if source_post_relation and not source_post_allowed:
        diagnostic_noise_type = "source_post_context"
    elif out_of_scope:
        diagnostic_noise_type = out_of_scope
    elif not intent_supported:
        diagnostic_noise_type = "intent_without_text_support"
    if semantic_exclusion:
        diagnostic_noise_type = semantic_exclusion
    row.update({
        "event_temporal_status": temporal.get("event_temporal_status"),
        "event_temporal_evidence_ru": temporal.get("event_temporal_evidence_ru"),
        "event_latest_detected_date": temporal.get("event_latest_detected_date"),
        "event_temporal_gate_passed": bool(temporal.get("event_temporal_gate_passed")),
        "semantic_exclusion_type": semantic_exclusion,
        "semantic_candidate_rejected": bool(semantic_exclusion),
    })
    if not _deterministic_prefilter_enabled():
        row["raw_score"] = raw_score
        row["question_boost"] = 0.0
        row["noise_penalty"] = 0.0
        row["score_for_rank"] = raw_score
        row["question_signal"] = bool(features["question_signal"])
        row["candidate_noise_type"] = diagnostic_noise_type
        row["intent_text_supported"] = bool(intent_supported)
        row["pre_llm_candidate_eligible"] = not bool(semantic_exclusion)
        row["llm_gate_selection_basis"] = "semantic_top_n_no_deterministic_prefilter"
        return
    if features["hard_noise"]:
        penalty = 0.35
    elif out_of_scope:
        penalty = 0.32
    elif source_post_relation and not source_post_allowed:
        penalty = 0.28
    elif not intent_supported:
        penalty = 0.18
    elif not features["question_signal"]:
        penalty = 0.08
    else:
        penalty = 0.0
    boost = 0.04 if features["question_signal"] else 0.0
    score_for_rank = raw_score + boost - penalty
    row["raw_score"] = raw_score
    row["question_boost"] = boost
    row["noise_penalty"] = penalty
    row["score_for_rank"] = score_for_rank
    row["question_signal"] = bool(features["question_signal"])
    row["candidate_noise_type"] = features["noise_type"]
    if source_post_relation and not source_post_allowed:
        row["candidate_noise_type"] = "source_post_not_comment"
    elif out_of_scope:
        row["candidate_noise_type"] = out_of_scope
    elif not intent_supported:
        row["candidate_noise_type"] = "intent_without_text_support"
    if semantic_exclusion:
        row["candidate_noise_type"] = semantic_exclusion
    row["intent_text_supported"] = bool(intent_supported)
    row["pre_llm_candidate_eligible"] = bool(
        features["question_signal"]
        and not features["hard_noise"]
        and not out_of_scope
        and intent_supported
        and (not source_post_relation or source_post_allowed)
        and not semantic_exclusion
    )
    row["llm_gate_selection_basis"] = "legacy_deterministic_prefilter"


def _record_analysis_text(record: dict[str, Any]) -> str:
    """Text embedded for retrieval: comment plus bounded parent/source context."""
    text = normalize_comment_text(str(record.get("text") or record.get("text_snapshot") or ""))
    parent = normalize_comment_text(str(record.get("reply_parent_text_snapshot") or record.get("reply_parent_text") or ""))
    source_post = normalize_comment_text(str(record.get("source_post_text_snapshot") or record.get("source_post_text") or ""))
    source_title = normalize_comment_text(str(record.get("source_context_title") or ""))
    context_parts: list[str] = []
    if source_post and source_post != text:
        context_parts.append(f"Исходный пост/тема: {source_post[:700]}")
    elif source_title:
        context_parts.append(f"Тема/контекст: {source_title[:300]}")
    if parent:
        context_parts.append(f"Родительский комментарий: {parent[:500]}")
    if not context_parts:
        return text
    return ("Текущий комментарий/пост: " + text + "\n\nКонтекст для понимания:\n" + "\n".join(context_parts))[:1800]


def _parse_created_at(value: Any) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except Exception:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _period_stats(records: list[dict[str, Any]]) -> dict[str, Any]:
    dates = [dt for dt in (_parse_created_at(r.get("created_at")) for r in records) if dt is not None]
    if not dates:
        return {
            "period_min_created_at": None,
            "period_max_created_at": None,
            "period_days": None,
            "period_label": "unknown_period",
        }
    start = min(dates)
    end = max(dates)
    seconds = max(0.0, (end - start).total_seconds())
    # A single-day/newest-N sample is still a sample period, not infinite rate.
    days = max(1.0, seconds / 86400.0)
    return {
        "period_min_created_at": start.isoformat(),
        "period_max_created_at": end.isoformat(),
        "period_days": round(days, 3),
        "period_label": f"{start.date().isoformat()}..{end.date().isoformat()} ({round(days, 1)}d)",
    }


def _latest_n_records(records: list[dict[str, Any]], *, limit: int = 100) -> list[dict[str, Any]]:
    dated: list[tuple[datetime, dict[str, Any]]] = []
    undated: list[dict[str, Any]] = []
    for record in records:
        dt = _parse_created_at(record.get("created_at"))
        if dt is None:
            undated.append(record)
        else:
            dated.append((dt, record))
    dated.sort(key=lambda item: item[0], reverse=True)
    ordered = [record for _dt, record in dated] + undated
    return ordered[: max(1, int(limit))]


def _days_since(value: Any, *, now: datetime | None = None) -> float | None:
    dt = _parse_created_at(value)
    if dt is None:
        return None
    now = (now or datetime.now(timezone.utc)).astimezone(timezone.utc)
    return max(0.0, (now - dt).total_seconds() / 86400.0)


def _latest_created_at(rows: Iterable[dict[str, Any]]) -> str | None:
    dates = [dt for dt in (_parse_created_at(row.get("created_at")) for row in rows) if dt is not None]
    if not dates:
        return None
    return max(dates).isoformat()


def _source_run_id_from_env() -> str:
    return (os.getenv("KAGGLE_RUN_ID") or os.getenv("ACQ_SOURCE_RUN_ID") or "").strip()


def _source_run_provenance() -> str:
    if (os.getenv("KAGGLE_RUN_ID") or "").strip():
        return "kaggle_run_id"
    if (os.getenv("ACQ_SOURCE_RUN_ID") or "").strip():
        return "explicit_source_run_id"
    return "missing_run_id_no_increment_claim"


def _record_within_analysis_window(record: dict[str, Any], *, now: datetime | None = None) -> bool:
    dt = _parse_created_at(record.get("created_at"))
    if dt is None:
        # Keep undated rows visible rather than silently losing potentially
        # current data; the report will show unknown dates.
        return True
    days = _days_since(dt, now=now)
    return days is None or days <= _max_comment_age_days()


def _record_usage_scope(record: dict[str, Any], *, now: datetime | None = None) -> str:
    if _record_within_analysis_window(record, now=now):
        return "monitoring_candidate"
    return "historical_calibration"


def _rate(count: int | float, period_days: Any, multiplier: float) -> float | None:
    try:
        days = float(period_days)
    except Exception:
        return None
    if days <= 0:
        return None
    return round(float(count) / days * multiplier, 3)


def _per_100(count: int | float, total: int | float) -> float | None:
    try:
        denominator = float(total)
    except Exception:
        return None
    if denominator <= 0:
        return None
    return round(float(count) / denominator * 100.0, 3)


def _truthy_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _question_pattern_label(text: str, *, action_type: str = "", intent_set: str = "") -> str:
    compact = normalize_comment_text(text).casefold()
    if not compact:
        return "other_question"
    if "пушкин" in compact:
        return "event_badge_pushkin_card"
    if re.search(r"\b(бесплатн|свободн|льгот|инвалид|доступн)\w*", compact):
        return "event_badge_access_or_free"
    if re.search(r"\b(куда\s+(?:съездить|поехать)|что\s+посмотреть|маршрут|на\s+один\s+день|выходн)", compact):
        if re.search(r"\b(дет|реб[её]н|семь|семейн)\w*", compact):
            return "route_with_children"
        if re.search(r"\b(электрич|поезд|автобус|без\s+машин|машин)\w*", compact):
            return "route_transport_or_car"
        return "route_where_to_go"
    if re.search(r"\b(как\s+добраться|доехать|ехать|транспорт)\b", compact):
        return "route_transport_or_car"
    if re.search(r"\b(куда|где).{0,20}(афиш|событ|мероприят|выставк|концерт)", compact):
        return "event_site_search"
    if re.search(r"\b(добавить|прислать|опубликовать|партн[её]рств)\w*.*\b(афиш|анонс|событ|мероприят)", compact):
        return "organizer_submission"
    if re.search(r"\b(билет|стоимост|цена|сколько\s+стоит|вход)\w*", compact):
        return "event_ticket_or_price"
    if re.search(r"\b(регистрац|запис|мест[ао]|остал)\w*", compact):
        return "event_registration_or_seats"
    if re.search(r"\b(возраст|лет|дет|реб[её]н)\w*", compact):
        return "event_age_or_children"
    if re.search(r"\b(во\s+сколько|когда|начал|длит|программ|расписан)\w*", compact):
        return "event_time_schedule_program"
    if re.search(r"\b(где|адрес|локац|место|вход|встреч)\w*", compact):
        return "event_location_or_entry"
    if re.search(r"\b(запись|трансляц|онлайн|стрим)\w*", compact):
        return "event_recording_or_stream"
    if action_type == "trip_route_poi_recommendation" or intent_set.startswith("route_poi"):
        return "route_other"
    if action_type == "organizer_submission_or_partnership":
        return "organizer_submission"
    if action_type == "badge_filter_need":
        return "event_badge_other"
    if action_type == "event_site_search_or_listing":
        return "event_site_search"
    if str(intent_set) in {"event_close_question", "organizer_comment_fit", "organizer_event_post_context"}:
        return "event_logistics_other"
    return "other_question"


_CANONICAL_QUESTIONS_RU = {
    "route_with_children": "Куда съездить с детьми на один день по Калининградской области?",
    "route_transport_or_car": "Куда поехать из Калининграда на электричке/автобусе/машине и что посмотреть по пути?",
    "route_where_to_go": "Куда съездить или что посмотреть в Калининградской области на один день/выходные?",
    "route_other": "Какой маршрут или место в Калининградской области посоветуете под этот запрос?",
    "event_site_search": "Где посмотреть актуальную афишу, выставки или календарь событий?",
    "organizer_submission": "Куда прислать анонс события или как добавить мероприятие в афишу?",
    "event_ticket_or_price": "Сколько стоит вход/билет и где его купить?",
    "event_registration_or_seats": "Нужна ли регистрация и остались ли места?",
    "event_age_or_children": "Есть ли возрастные ограничения и подходит ли событие детям?",
    "event_time_schedule_program": "Когда начало, сколько длится событие и какая программа?",
    "event_location_or_entry": "Где проходит событие, какой адрес/вход/место встречи?",
    "event_recording_or_stream": "Будет ли запись, трансляция или онлайн-доступ?",
    "event_badge_pushkin_card": "Можно ли посетить событие по Пушкинской карте?",
    "event_badge_access_or_free": "Есть ли бесплатный вход, льготы или условия доступности?",
    "event_badge_other": "Есть ли быстрые признаки события: бесплатно, детям, Пушкинская карта, доступность?",
    "event_logistics_other": "Какой важной практической информации о событии не хватает?",
    "other_question": "Какой полезный ответ можно дать на этот вопрос без рекламы?",
}


def _canonical_question_for_pattern(pattern: str) -> str:
    return _CANONICAL_QUESTIONS_RU.get(str(pattern or ""), _CANONICAL_QUESTIONS_RU["other_question"])


def _is_actual_question_text(text: str) -> bool:
    compact = normalize_comment_text(text)
    if not compact:
        return False
    if "?" in compact:
        return True
    return bool(re.match(
        r"(?i)^(?:а\s+)?(?:подскажите|посоветуйте|скажите|кто\s+знает|где|куда|когда|как|что|сколько|какой|какая|какие|есть\s+ли|будет\s+ли|можно\s+ли|нужна\s+ли)\b",
        compact,
    ))


_REPORT_REJECT_NOISE_TYPES = {
    "explicit_offer_or_ad",
    "link_or_crosspost_without_question",
    "too_short_non_question",
    "intent_without_text_support",
    "source_post_not_comment",
    "past_event",
    "past_event_signal",
    "out_of_scope_gasoline_availability",
}


def _is_source_post_context(row: dict[str, Any]) -> bool:
    relation = str(row.get("relation") or "")
    return relation in {"vk_social_wall_post", "tg_channel_post_context"} or _truthy_value(row.get("is_post"))


def _row_score(row: dict[str, Any]) -> float:
    for key in ["score", "score_for_rank", "positive_negative_margin"]:
        try:
            return float(row.get(key))
        except Exception:
            continue
    return 0.0


def _report_min_comment_score() -> float:
    return _float_env("ACQ_COMMENT_RETRIEVAL_REPORT_MIN_COMMENT_SCORE", 0.01)


def _report_min_source_post_score() -> float:
    return _float_env("ACQ_COMMENT_RETRIEVAL_REPORT_MIN_SOURCE_POST_SCORE", 0.01)


def _report_noise_rejected(row: dict[str, Any]) -> bool:
    noise = str(row.get("candidate_noise_type") or "").strip()
    return bool(noise in _REPORT_REJECT_NOISE_TYPES or noise.startswith("out_of_scope"))


def _report_candidate_eligible(row: dict[str, Any], *, allow_source_posts: bool = True, allow_historical: bool = False) -> bool:
    """Rows suitable for human-facing summaries/examples.

    This is intentionally *not* the LLM-gate selector.  The default discovery
    gate still receives top-N vector rows without a deterministic prefilter.
    The report/catalog path is stricter so garbage rows do not become
    “canonical questions” or selected-surface evidence.
    """
    scope = str(row.get("candidate_usage_scope") or "")
    if scope == "historical_calibration" and not allow_historical:
        return False
    if scope not in {"", "monitoring_candidate", "historical_calibration"}:
        return False
    if not _candidate_region_eligible(row):
        return False
    if _hard_semantic_rejected(row):
        return False
    text = str(row.get("text_snapshot") or row.get("text") or "")
    features = _text_quality_features(text)
    intent = str(row.get("intent_set") or "")
    action = str(row.get("candidate_action_type") or "")
    intent_supported = bool(_truthy_value(row.get("intent_text_supported")) and _intent_has_text_support(text, intent))
    if _is_source_post_context(row):
        if not allow_source_posts:
            return False
        if intent not in {"organizer_comment_fit", "organizer_event_post_context", "event_close_question"}:
            return False
        if not intent_supported:
            return False
        return _row_score(row) >= _report_min_source_post_score()
    if _report_noise_rejected(row):
        return False
    if features["hard_noise"] or _out_of_scope_noise_type(text):
        return False
    if not (_truthy_value(row.get("question_signal")) and features["question_signal"]):
        return False
    if not intent_supported:
        return False
    if not _is_actual_question_text(text):
        return False
    if action == "organizer_visibility_clarification" and intent not in {"organizer_comment_fit", "organizer_event_post_context", "event_close_question"}:
        return False
    return _row_score(row) >= _report_min_comment_score()


def _report_real_question_row(row: dict[str, Any], *, allow_historical: bool = True) -> bool:
    return (
        not _is_source_post_context(row)
        and _report_candidate_eligible(row, allow_source_posts=False, allow_historical=allow_historical)
        and _is_actual_question_text(str(row.get("text_snapshot") or row.get("text") or ""))
    )


def _surface_key(record: dict[str, Any]) -> str:
    return str(record.get("surface_key") or record.get("surface_external_id") or record.get("surface_url") or "unknown")


def _attach_surface_metadata_to_records(
    records: list[dict[str, Any]],
    surfaces_by_external: dict[str, dict[str, Any]] | None,
) -> None:
    if not surfaces_by_external:
        return
    for item in records:
        key = _surface_key(item)
        surface = surfaces_by_external.get(key)
        if not surface:
            continue
        item.setdefault("surface_key", key)
        if surface.get("title") and not item.get("surface_title"):
            item["surface_title"] = str(surface.get("title") or "")
        if surface.get("url") and not item.get("surface_url"):
            item["surface_url"] = str(surface.get("url") or "")
        if surface.get("source") and not item.get("source"):
            item["source"] = str(surface.get("source") or "")
        if surface.get("discovery_source_context") and not item.get("discovery_source_context"):
            item["discovery_source_context"] = str(surface.get("discovery_source_context") or "")
        if surface.get("topic_hint") and not item.get("topic_hint"):
            item["topic_hint"] = str(surface.get("topic_hint") or "")


def _dedupe_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str]] = set()
    out: list[dict[str, Any]] = []
    now = datetime.now(timezone.utc)
    for rec in records:
        text = normalize_comment_text(str(rec.get("text") or rec.get("text_snapshot") or ""))
        if not text:
            continue
        key = (_surface_key(rec), text[:500].casefold())
        if key in seen:
            continue
        seen.add(key)
        item = dict(rec)
        item["text"] = text
        item["text_snapshot"] = text[:500]
        age_days = _days_since(item.get("created_at"), now=now)
        usage_scope = _record_usage_scope(item, now=now)
        item["comment_age_days"] = round(age_days, 1) if age_days is not None else None
        item["is_within_monitoring_window"] = usage_scope == "monitoring_candidate"
        item["candidate_usage_scope"] = usage_scope
        analysis_text = _record_analysis_text(item)
        item["analysis_text"] = analysis_text
        item["analysis_context_snapshot"] = "\n".join(
            part for part in [
                normalize_comment_text(str(item.get("source_post_text_snapshot") or item.get("source_post_text") or item.get("source_context_title") or ""))[:700],
                normalize_comment_text(str(item.get("reply_parent_text_snapshot") or item.get("reply_parent_text") or ""))[:500],
            ]
            if part
        )[:1000]
        item.setdefault("surface_key", _surface_key(item))
        out.append(item)
    return out


def _score_comment(
    comment_vec: Any,
    intent_vectors: dict[str, list[Any]],
    negative_vectors: list[Any],
) -> list[dict[str, Any]]:
    negative_scores = [_dot(comment_vec, neg) for neg in negative_vectors]
    negative_score = max(negative_scores) if negative_scores else 0.0
    rows: list[dict[str, Any]] = []
    for intent_set, vectors in intent_vectors.items():
        scores = [_dot(comment_vec, vec) for vec in vectors]
        if not scores:
            continue
        max_score = max(scores)
        top3 = sorted(scores, reverse=True)[:3]
        centroid_score = _mean(scores)
        margin = max_score - negative_score
        top_idx = scores.index(max_score)
        rows.append({
            "intent_set": intent_set,
            "max_positive_similarity": float(max_score),
            "top3_positive_mean": float(_mean(top3)),
            "centroid_similarity": float(centroid_score),
            "positive_negative_margin": float(margin),
            "positive_score": float(max_score),
            "negative_score": float(negative_score),
            "top_intent_phrase": INTENT_SETS[intent_set][top_idx] if top_idx < len(INTENT_SETS[intent_set]) else "",
            "top_intent_score": float(max_score),
        })
    return rows


def _rank_candidates(rows: list[dict[str, Any]], *, scoring_method: str) -> list[dict[str, Any]]:
    rows = [r for r in rows if r.get("intent_set") != "negative_intents"]
    rows.sort(key=lambda r: float(r.get("score_for_rank") if r.get("score_for_rank") is not None else (r.get(scoring_method) or r.get("score") or 0.0)), reverse=True)
    for idx, row in enumerate(rows, start=1):
        row["rank_global"] = idx
        row["score"] = float(row.get("score_for_rank") if row.get("score_for_rank") is not None else (row.get(scoring_method) or 0.0))
    per_surface: dict[str, int] = defaultdict(int)
    for row in rows:
        key = str(row.get("surface_key") or "unknown")
        per_surface[key] += 1
        row["rank_within_surface"] = per_surface[key]
    total = max(1, len(rows))
    for row in rows:
        pct = 100.0 * int(row["rank_global"]) / total
        bucket = "top_10pct"
        if pct <= 0.5:
            bucket = "top_0_5pct"
        elif pct <= 1:
            bucket = "top_1pct"
        elif pct <= 3:
            bucket = "top_3pct"
        elif pct <= 5:
            bucket = "top_5pct"
        row["funnel_bucket"] = bucket
    return rows


def _surface_region_summary(surface_key: str, surface_records: list[dict[str, Any]], region_rows: list[dict[str, Any]]) -> dict[str, Any]:
    surface_text = normalize_comment_text(" ".join(
        str((surface_records[0] if surface_records else {}).get(key) or "")
        for key in ["surface_title", "source_context_title", "surface_url", "surface_key", "surface_external_id", "surface_type", "platform", "source"]
    ))
    positive_surface_hint = _first_match(_KGD_REGION_HINT_RE, surface_text)
    negative_surface_hint = _first_match(_OUT_OF_REGION_HINT_RE, surface_text)
    unique_by_context = _best_context_rows(region_rows)
    confidence_counts = Counter(str(row.get("region_confidence") or "unknown") for row in unique_by_context)
    scores = [float(row.get("region_score") or 0.0) for row in unique_by_context if row.get("region_score") is not None]
    positive_scores = [float(row.get("region_positive_score") or 0.0) for row in unique_by_context if row.get("region_positive_score") is not None]
    negative_scores = [float(row.get("region_negative_score") or 0.0) for row in unique_by_context if row.get("region_negative_score") is not None]
    evidence_row = next(
        (
            row for row in unique_by_context
            if str(row.get("region_confidence") or "") in {"confirmed", "probable", "out_of_region"}
        ),
        None,
    )
    if positive_surface_hint:
        confidence = "confirmed"
        source = "surface_metadata_keyword"
        evidence = f"площадка содержит явную зацепку Калининградской области: «{positive_surface_hint}»"
    elif negative_surface_hint and not (confidence_counts.get("confirmed", 0) or confidence_counts.get("probable", 0)):
        confidence = "out_of_region"
        source = "surface_out_of_region_keyword"
        evidence = f"площадка похожа на другой регион: «{negative_surface_hint}»"
    elif confidence_counts.get("confirmed", 0) > 0:
        confidence = "confirmed"
        source = "comment_context_keyword"
        evidence = str((evidence_row or {}).get("region_evidence_ru") or "в комментариях есть явная зацепка Калининградской области")
    elif confidence_counts.get("probable", 0) > 0:
        confidence = "probable"
        source = "semantic_region_vector"
        evidence = str((evidence_row or {}).get("region_evidence_ru") or "регион вероятен по embedding-сходству")
    elif confidence_counts.get("out_of_region", 0) > 0:
        confidence = "out_of_region"
        source = "semantic_or_keyword_out_of_region"
        evidence = str((evidence_row or {}).get("region_evidence_ru") or "есть признаки другого региона")
    else:
        confidence = "unknown"
        source = "no_region_signal"
        evidence = "нет достаточных доказательств, что площадка/комментарии относятся к Калининградской области"
    return {
        "region_confidence": confidence,
        "region_gate_status": "region_ok" if confidence in REGION_ELIGIBLE_CONFIDENCES else f"region_{confidence}",
        "region_signal_source": source,
        "region_evidence_ru": evidence,
        "region_confirmed_contexts": confidence_counts.get("confirmed", 0),
        "region_probable_contexts": confidence_counts.get("probable", 0),
        "region_unknown_contexts": confidence_counts.get("unknown", 0),
        "region_out_of_region_contexts": confidence_counts.get("out_of_region", 0),
        "region_score_max": max(scores) if scores else None,
        "region_score_p95": _percentile(scores, 0.95) if scores else None,
        "region_positive_score_max": max(positive_scores) if positive_scores else None,
        "region_negative_score_max": max(negative_scores) if negative_scores else None,
        "region_surface_hint": positive_surface_hint,
        "region_out_of_region_hint": negative_surface_hint if not positive_surface_hint else "",
    }


def _surface_profile(
    surface_key: str,
    surface_records: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    *,
    scoring_method: str,
    region_rows: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    by_intent: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_intent[str(row.get("intent_set"))].append(row)
    semantic_presence: dict[str, dict[str, Any]] = {}
    dominant: list[str] = []
    for intent_set in POSITIVE_INTENT_SETS:
        intent_rows = by_intent.get(intent_set, [])
        scores = [float(r.get("score") if r.get("score") is not None else (r.get(scoring_method) or 0.0)) for r in intent_rows]
        p95 = _percentile(scores, 0.95)
        candidate_count_top3pct = sum(1 for r in intent_rows if str(r.get("funnel_bucket")) in {"top_0_5pct", "top_1pct", "top_3pct"})
        if candidate_count_top3pct >= 5 or p95 >= 0.18:
            level = "high"
        elif candidate_count_top3pct >= 2 or p95 >= 0.10:
            level = "medium"
        elif scores and (candidate_count_top3pct >= 1 or p95 > 0.03):
            level = "low"
        else:
            level = "none"
        if level != "none":
            if intent_set in ROUTE_INTENT_SETS and "route_poi" not in dominant:
                dominant.append("route_poi")
            elif intent_set in EVENT_INTENT_SETS and "event_questions" not in dominant:
                dominant.append("event_questions")
            elif intent_set in ORGANIZER_INTENT_SETS and "organizer_fit" not in dominant:
                dominant.append("organizer_fit")
        examples = [r.get("comment_id") for r in sorted(intent_rows, key=lambda r: float(r.get("score") if r.get("score") is not None else (r.get(scoring_method) or 0.0)), reverse=True)[:3]]
        semantic_presence[intent_set] = {
            "present": level != "none",
            "level": level,
            "top_score_p95": round(p95, 6),
            "candidate_count_top3pct": int(candidate_count_top3pct),
            "example_comment_ids": [str(x) for x in examples if x is not None],
        }
    comments_total = len(surface_records)
    actionable = sum(1 for r in rows if str(r.get("funnel_bucket")) in {"top_0_5pct", "top_1pct", "top_3pct"})
    eligible_questions = sum(1 for r in rows if r.get("pre_llm_candidate_eligible"))
    if actionable >= 3 and dominant:
        decision = "monitor"
        reason = f"{actionable} top question-like semantic candidates; interests={', '.join(dominant)}"
    elif dominant and comments_total >= 5:
        decision = "sample_more"
        reason = f"semantic signal present but only {actionable} top candidates"
    elif dominant:
        decision = "low_priority"
        reason = "weak semantic signal in a small sample"
    else:
        decision = "reject"
        reason = "no acquisition-relevant semantic signal in embedded comments"
    period = _period_stats(surface_records)
    days_since_latest_activity = _days_since(period.get("period_max_created_at"))
    if days_since_latest_activity is not None and days_since_latest_activity > _stale_activity_days():
        decision = "reject_stale_inactive"
        reason = (
            f"latest observed comment is {days_since_latest_activity:.0f} days old; "
            f"current acquisition requires activity within {_stale_activity_days()} days. "
            "Keep only as low-frequency revival watch."
        )
    latest_100_records = _latest_n_records(surface_records, limit=100)
    latest_100_period = _period_stats(latest_100_records)
    relation_counts = Counter(str(r.get("relation") or "unknown") for r in surface_records)
    author_ids = {
        str(r.get("author_id") or "").strip()
        for r in surface_records
        if str(r.get("author_id") or "").strip()
    }
    region_summary = _surface_region_summary(surface_key, surface_records, region_rows or rows)
    return {
        "surface_key": surface_key,
        "platform": surface_records[0].get("platform") if surface_records else None,
        "surface_type": surface_records[0].get("surface_type") if surface_records else None,
        "surface_title": (surface_records[0].get("surface_title") or surface_records[0].get("source_context_title")) if surface_records else None,
        "surface_url": surface_records[0].get("surface_url") if surface_records else None,
        **region_summary,
        "comments_total": comments_total,
        "comments_embedded": comments_total,
        "eligible_question_candidates": eligible_questions,
        "period": {
            "min_created_at": period["period_min_created_at"],
            "max_created_at": period["period_max_created_at"],
        },
        "period_min_created_at": period["period_min_created_at"],
        "period_max_created_at": period["period_max_created_at"],
        "period_days": period["period_days"],
        "period_label": period["period_label"],
        "latest_comment_at": period["period_max_created_at"],
        "days_since_latest_activity": round(days_since_latest_activity, 1) if days_since_latest_activity is not None else None,
        "stale_activity_days": _stale_activity_days(),
        "analysis_max_comment_age_days": _max_comment_age_days(),
        "freshness_status": "stale_inactive" if days_since_latest_activity is not None and days_since_latest_activity > _stale_activity_days() else "active_or_unknown",
        "comments_per_day": _rate(comments_total, period["period_days"], 1),
        "comments_per_week": _rate(comments_total, period["period_days"], 7),
        "comments_per_30d": _rate(comments_total, period["period_days"], 30),
        "comments_per_90d": _rate(comments_total, period["period_days"], 90),
        "latest_100_comments": len(latest_100_records),
        "latest_100_min_created_at": latest_100_period["period_min_created_at"],
        "latest_100_max_created_at": latest_100_period["period_max_created_at"],
        "latest_100_period_days": latest_100_period["period_days"],
        "latest_100_period_label": latest_100_period["period_label"],
        "unique_commenters": len(author_ids) if author_ids else None,
        "unique_commenters_observed": bool(author_ids),
        "relation_counts": dict(relation_counts),
        "comment_records": sum(count for rel, count in relation_counts.items() if rel != "vk_social_wall_post"),
        "source_post_records": int(relation_counts.get("vk_social_wall_post") or 0),
        "semantic_presence": semantic_presence,
        "dominant_detected_interests": dominant,
        "monitoring_decision_hint": decision,
        "monitoring_reason": reason,
        "llm_budget_recommendation": {
            "send_top_comments_to_llm": min(actionable, _int_env("ACQ_COMMENT_RETRIEVAL_MAX_LLM_CANDIDATES", 24, min_value=1)),
            "reason": "top retrieval candidates only; no full-comment LLM pass",
        },
    }


def _best_context_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    best: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = str(row.get("context_url") or row.get("comment_id") or row.get("text_snapshot") or "")
        if not key:
            continue
        if key not in best or float(row.get("rank_global") or 999999) < float(best[key].get("rank_global") or 999999):
            best[key] = row
    return sorted(best.values(), key=lambda r: float(r.get("rank_global") or 999999))


def _examples_text(rows: list[dict[str, Any]], *, limit: int = 3) -> str:
    examples: list[str] = []
    for row in _best_context_rows(rows)[:limit]:
        text = str(row.get("text_snapshot") or "").replace("\n", " ")[:140]
        url = str(row.get("context_url") or "")
        if url:
            examples.append(f"{text} — {url}")
        else:
            examples.append(text)
    return "\n".join(examples)


def _selection_status_from_recommendation(recommendation: str, *, surface_status: str = "", scan_state: str = "") -> str:
    rec = str(recommendation or "").strip()
    status = str(surface_status or "").strip()
    state = str(scan_state or "").strip()
    if rec in {"both_monitor_replies_and_ask_clarifications", "monitor_for_reply_opportunities", "ask_organizer_clarification_questions"}:
        return "selected"
    if rec in {"sample_more", "low_priority"}:
        return "candidate"
    if rec in {"reject_stale_inactive", "reject_or_low_priority"} or rec.startswith("reject") or rec.startswith("out_of_scope"):
        return "rejected"
    if status.startswith("rejected") or state.startswith("checked_no") or state in {"resolved_no_comments", "checked_inaccessible"}:
        return "rejected"
    if rec == "not_profiled_or_no_comment_signal" and state in {"scanned", "comments_available", "resolved_no_comments", "checked_inaccessible"}:
        return "rejected"
    return "candidate"


def _surface_decision_summaries(
    profiles: list[dict[str, Any]],
    *,
    eligible_rows_by_surface: dict[str, list[dict[str, Any]]],
    all_gate_rows_by_surface: dict[str, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for profile in profiles:
        surface_key = str(profile.get("surface_key") or "")
        eligible = _best_context_rows(eligible_rows_by_surface.get(surface_key, []))
        all_rows = _best_context_rows(all_gate_rows_by_surface.get(surface_key, []))
        comment_eligible = [r for r in eligible if not _is_source_post_context(r)]
        route_rows = [r for r in comment_eligible if r.get("candidate_action_type") == "trip_route_poi_recommendation"]
        event_rows = [r for r in comment_eligible if r.get("candidate_action_type") == "event_recommendation_reply"]
        organizer_submit_rows = [r for r in comment_eligible if r.get("candidate_action_type") == "organizer_submission_or_partnership"]
        badge_rows = [r for r in comment_eligible if r.get("candidate_action_type") == "badge_filter_need"]
        site_rows = [r for r in comment_eligible if r.get("candidate_action_type") == "event_site_search_or_listing"]
        event_question_rows = [*event_rows, *site_rows, *badge_rows]
        ask_context_rows = [
            r for r in eligible
            if str(r.get("intent_set")) in {"event_close_question", "organizer_comment_fit", "organizer_event_post_context"}
        ]
        comment_rows = comment_eligible
        source_post_rows = [r for r in all_rows if _is_source_post_context(r)]
        noise_rows = [r for r in all_rows if str(r.get("candidate_noise_type") or "")]
        answerable_count = len(route_rows) + len(event_rows) + len(site_rows) + len(organizer_submit_rows) + len(badge_rows)
        ask_count = len(ask_context_rows)
        if answerable_count >= 3 and ask_count >= 2:
            recommendation = "both_monitor_replies_and_ask_clarifications"
        elif answerable_count >= 3:
            recommendation = "monitor_for_reply_opportunities"
        elif ask_count >= 2:
            recommendation = "ask_organizer_clarification_questions"
        elif answerable_count or ask_count:
            recommendation = "sample_more"
        else:
            recommendation = "reject_or_low_priority"
        surface_scope_noise = _surface_out_of_scope_type(profile)
        region_confidence = str(profile.get("region_confidence") or "unknown")
        if region_confidence == "out_of_region":
            recommendation = "reject_out_of_region"
        elif region_confidence not in REGION_ELIGIBLE_CONFIDENCES:
            recommendation = "reject_region_unknown"
        elif profile.get("monitoring_decision_hint") == "reject_stale_inactive":
            recommendation = "reject_stale_inactive"
        elif surface_scope_noise:
            recommendation = surface_scope_noise
        parts: list[str] = []
        if route_rows:
            parts.append(f"route/POI вопросов: {len(route_rows)}")
        if event_rows:
            parts.append(f"event-вопросов: {len(event_rows)}")
        if organizer_submit_rows:
            parts.append(f"organizer submission вопросов: {len(organizer_submit_rows)}")
        if site_rows:
            parts.append(f"поиск/афиша вопросов: {len(site_rows)}")
        if badge_rows:
            parts.append(f"badge/filter вопросов: {len(badge_rows)}")
        if ask_context_rows:
            own_post_count = len([r for r in ask_context_rows if _is_source_post_context(r)])
            if own_post_count:
                parts.append(f"контекстов постов для уточняющих вопросов организаторам: {own_post_count}")
            comment_ask_count = len(ask_context_rows) - own_post_count
            if comment_ask_count:
                parts.append(f"вопросов/комментариев для organizer-уточнений: {comment_ask_count}")
        if noise_rows:
            if _deterministic_prefilter_enabled():
                parts.append(f"отфильтровано рекламных/не-вопросных сигналов: {len(noise_rows)}")
            else:
                parts.append(f"диагностических шумовых/не-вопросных сигналов без prefilter-отсечения: {len(noise_rows)}")
        summary_ru = "; ".join(parts) if parts else "полезных вопросных сигналов не найдено"
        if recommendation == "reject_stale_inactive":
            summary_ru = (
                f"устаревшая активность: последний комментарий {profile.get('latest_comment_at') or 'неизвестно'}, "
                f"{profile.get('days_since_latest_activity') or '?'} дней назад; не выбирать сейчас, только revival-watch"
            )
        elif recommendation == "reject_out_of_region":
            summary_ru = f"не Калининградская область: {profile.get('region_evidence_ru') or 'региональная проверка не пройдена'}"
        elif recommendation == "reject_region_unknown":
            summary_ru = (
                "регион не доказан: для событий и маршрутов выбираем только площадки/стены с Калининградской областью; "
                f"{profile.get('region_evidence_ru') or 'нет региональной зацепки'}"
            )
        elif surface_scope_noise:
            summary_ru = f"площадка не по теме acquisition ({surface_scope_noise}); не выбирать для ответов/маршрутов/событий"
        increment_status = _surface_increment_status(profile, {
            "comments_embedded": profile.get("comments_embedded"),
        })
        out.append({
            "surface_key": surface_key,
            "platform": profile.get("platform"),
            "surface_type": profile.get("surface_type"),
            "surface_type_ru": _SURFACE_TYPE_RU.get(str(profile.get("surface_type") or ""), profile.get("surface_type")),
            "surface_title": profile.get("surface_title"),
            "surface_url": profile.get("surface_url"),
            "members_or_subscribers": profile.get("members_or_subscribers"),
            "period_min_created_at": profile.get("period_min_created_at"),
            "period_max_created_at": profile.get("period_max_created_at"),
            "period_days": profile.get("period_days"),
            "period_label": profile.get("period_label"),
            "latest_comment_at": profile.get("latest_comment_at"),
            "latest_event_question_at": _latest_created_at(event_question_rows),
            "latest_route_recommendation_at": _latest_created_at(route_rows),
            "increment_status": increment_status,
            "is_incremental_last_run": _is_counted_increment_status(increment_status),
            "last_analyzed_run_id": profile.get("last_analyzed_run_id"),
            "discovered_at": profile.get("discovered_at"),
            "discovered_in_run_id": profile.get("discovered_in_run_id"),
            "discovery_source_context": profile.get("discovery_source_context"),
            "days_since_latest_activity": profile.get("days_since_latest_activity"),
            "freshness_status": profile.get("freshness_status"),
            "region_confidence": profile.get("region_confidence"),
            "region_gate_status": profile.get("region_gate_status"),
            "region_signal_source": profile.get("region_signal_source"),
            "region_evidence_ru": profile.get("region_evidence_ru"),
            "region_confirmed_contexts": profile.get("region_confirmed_contexts"),
            "region_probable_contexts": profile.get("region_probable_contexts"),
            "region_unknown_contexts": profile.get("region_unknown_contexts"),
            "region_out_of_region_contexts": profile.get("region_out_of_region_contexts"),
            "region_score_max": profile.get("region_score_max"),
            "region_score_p95": profile.get("region_score_p95"),
            "region_positive_score_max": profile.get("region_positive_score_max"),
            "region_negative_score_max": profile.get("region_negative_score_max"),
            "analysis_max_comment_age_days": profile.get("analysis_max_comment_age_days"),
            "stale_activity_days": profile.get("stale_activity_days"),
            "unique_commenters": profile.get("unique_commenters"),
            "unique_commenters_note": (
                "уникальные авторы комментариев/сообщений за указанный период анализа площадки"
                if profile.get("unique_commenters_observed")
                else "не собрано в этом run"
            ),
            "comments_total": profile.get("comments_total"),
            "comments_embedded": profile.get("comments_embedded"),
            "comments_per_day": profile.get("comments_per_day"),
            "comments_per_week": profile.get("comments_per_week"),
            "comments_per_30d": profile.get("comments_per_30d"),
            "comments_per_90d": profile.get("comments_per_90d"),
            "latest_100_comments": profile.get("latest_100_comments"),
            "latest_100_min_created_at": profile.get("latest_100_min_created_at"),
            "latest_100_max_created_at": profile.get("latest_100_max_created_at"),
            "latest_100_period_days": profile.get("latest_100_period_days"),
            "latest_100_period_label": profile.get("latest_100_period_label"),
            "recommendation": recommendation,
            "selection_status": _selection_status_from_recommendation(recommendation),
            "summary_ru": summary_ru,
            "answerable_question_candidates": answerable_count,
            "ask_clarification_contexts": ask_count,
            "eligible_comment_contexts": len(comment_rows),
            "source_post_contexts": len(source_post_rows),
            "route_poi_questions": len(route_rows),
            "event_questions": len(event_rows),
            "event_site_search_questions": len(site_rows),
            "organizer_submission_questions": len(organizer_submit_rows),
            "badge_filter_questions": len(badge_rows),
            "answerable_questions_per_30d": _rate(answerable_count, profile.get("period_days"), 30),
            "answerable_questions_per_90d": _rate(answerable_count, profile.get("period_days"), 90),
            "answerable_questions_per_100_comments": _per_100(answerable_count, profile.get("comments_embedded") or 0),
            "ask_contexts_per_30d": _rate(ask_count, profile.get("period_days"), 30),
            "relation_counts_json": json.dumps(profile.get("relation_counts") or {}, ensure_ascii=False),
            "filtered_noise_contexts": len(noise_rows),
            "dominant_detected_interests": ",".join(profile.get("dominant_detected_interests") or []),
            "monitoring_decision_hint": profile.get("monitoring_decision_hint"),
            "monitoring_reason": profile.get("monitoring_reason"),
            "answerable_examples": _examples_text([*route_rows[:2], *event_rows[:2], *site_rows[:1], *organizer_submit_rows[:1], *badge_rows[:1]], limit=4),
            "ask_question_examples": _examples_text(ask_context_rows, limit=3),
        })
    return sorted(out, key=lambda r: (
        str(r.get("recommendation")) not in {"both_monitor_replies_and_ask_clarifications", "monitor_for_reply_opportunities", "ask_organizer_clarification_questions"},
        -int(r.get("answerable_question_candidates") or 0),
        -int(r.get("ask_clarification_contexts") or 0),
        str(r.get("surface_key") or ""),
    ))


def _build_question_patterns(rows: list[dict[str, Any]], *, limit_examples: int = 3) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in _best_context_rows([r for r in rows if _report_real_question_row(r)]):
        text = str(row.get("text_snapshot") or row.get("text") or "")
        pattern = _question_pattern_label(
            text,
            action_type=str(row.get("candidate_action_type") or ""),
            intent_set=str(row.get("intent_set") or ""),
        )
        surface_key = str(row.get("surface_key") or "")
        action = str(row.get("candidate_action_type") or "")
        key = (surface_key, pattern, action)
        item = grouped.setdefault(key, {
            "surface_key": surface_key,
            "platform": row.get("platform"),
            "surface_type": row.get("surface_type"),
            "pattern": pattern,
            "candidate_action_type": action,
            "intent_sets": set(),
            "models": set(),
            "count": 0,
            "canonical_question_ru": _canonical_question_for_pattern(pattern),
            "example_questions": [],
            "example_urls": [],
        })
        item["count"] += 1
        if row.get("intent_set"):
            item["intent_sets"].add(str(row.get("intent_set")))
        if row.get("model_name"):
            item["models"].add(str(row.get("model_name")))
        if len(item["example_questions"]) < limit_examples and text:
            item["example_questions"].append(text[:220])
        if len(item["example_urls"]) < limit_examples and row.get("context_url"):
            item["example_urls"].append(str(row.get("context_url")))
    out: list[dict[str, Any]] = []
    for item in grouped.values():
        out.append({
            **{k: v for k, v in item.items() if k not in {"intent_sets", "models", "example_questions", "example_urls"}},
            "intent_sets": ",".join(sorted(item["intent_sets"])),
            "models": ",".join(sorted(item["models"])),
            "example_questions": "\n".join(item["example_questions"]),
            "example_urls": "\n".join(item["example_urls"]),
        })
    return sorted(out, key=lambda r: (-int(r.get("count") or 0), str(r.get("surface_key") or ""), str(r.get("pattern") or "")))


def _canonical_question_catalog_rows(rows: list[dict[str, Any]], *, limit_examples: int = 5) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in _best_context_rows([r for r in rows if _report_real_question_row(r)]):
        text = str(row.get("text_snapshot") or row.get("text") or "")
        action = str(row.get("candidate_action_type") or "")
        pattern = _question_pattern_label(text, action_type=action, intent_set=str(row.get("intent_set") or ""))
        key = (pattern, action)
        item = grouped.setdefault(key, {
            "pattern": pattern,
            "candidate_action_type": action,
            "canonical_question_ru": _canonical_question_for_pattern(pattern),
            "real_question_examples_total": 0,
            "monitoring_candidate_examples": 0,
            "historical_calibration_examples": 0,
            "surfaces_count": set(),
            "example_questions": [],
            "example_urls": [],
            "source_note_ru": "Эталон собран из реальных вопросительных комментариев; historical используется для шаблонов/QA, fresh — для мониторинга.",
        })
        item["real_question_examples_total"] += 1
        scope = str(row.get("candidate_usage_scope") or "")
        if scope == "monitoring_candidate":
            item["monitoring_candidate_examples"] += 1
        elif scope == "historical_calibration":
            item["historical_calibration_examples"] += 1
        if row.get("surface_key"):
            item["surfaces_count"].add(str(row.get("surface_key")))
        if len(item["example_questions"]) < limit_examples:
            item["example_questions"].append(text[:220])
        if len(item["example_urls"]) < limit_examples and row.get("context_url"):
            item["example_urls"].append(str(row.get("context_url")))
    out: list[dict[str, Any]] = []
    for item in grouped.values():
        out.append({
            **{k: v for k, v in item.items() if k not in {"surfaces_count", "example_questions", "example_urls"}},
            "surfaces_count": len(item["surfaces_count"]),
            "example_questions": "\n".join(item["example_questions"]),
            "example_urls": "\n".join(item["example_urls"]),
        })
    return sorted(out, key=lambda r: (-int(r.get("monitoring_candidate_examples") or 0), -int(r.get("real_question_examples_total") or 0), str(r.get("pattern") or "")))


def _model_example_rows(candidates: list[dict[str, Any]], model_name: str, *, limit: int = 200) -> list[dict[str, Any]]:
    rows = [
        r for r in candidates
        if str(r.get("model_name") or "") == model_name
        and _report_real_question_row(r, allow_historical=False)
    ]
    return _best_context_rows(rows)[:limit]


def _model_ask_context_rows(candidates: list[dict[str, Any]], model_name: str, *, limit: int = 200) -> list[dict[str, Any]]:
    rows = [
        r for r in candidates
        if str(r.get("model_name") or "") == model_name
        and _is_source_post_context(r)
        and not _truthy_value(r.get("question_signal"))
        and _report_candidate_eligible(r)
    ]
    return _best_context_rows(rows)[:limit]


def _scope_rows(
    *,
    run_id: str,
    records: list[dict[str, Any]],
    profiles: list[dict[str, Any]],
    models: list[str],
    gate_model: str,
    scoring_method: str,
    summary_note: str | None = None,
) -> list[dict[str, Any]]:
    period = _period_stats(records)
    platforms = Counter(str(r.get("platform") or "unknown") for r in records)
    surface_types = Counter(str(r.get("surface_type") or "unknown") for r in records)
    relations = Counter(str(r.get("relation") or "unknown") for r in records)
    usage_scopes = Counter(str(r.get("candidate_usage_scope") or _record_usage_scope(r)) for r in records)
    profile_regions = Counter(str(p.get("region_confidence") or "unknown") for p in profiles)
    return [
        {"metric": "run_id", "value": run_id},
        {"metric": "report_generated_at", "value": datetime.now(timezone.utc).isoformat()},
        {"metric": "KAGGLE_RUN_ID", "value": os.getenv("KAGGLE_RUN_ID", "")},
        {"metric": "ACQ_SOURCE_RUN_ID", "value": os.getenv("ACQ_SOURCE_RUN_ID", "")},
        {"metric": "source_run_provenance", "value": _source_run_provenance()},
        {"metric": "stage", "value": STAGE_NAME},
        {"metric": "scope_note", "value": summary_note or "Limited to the Kaggle seed payload and configured per-run budgets; not a full historical DB scan unless the payload/budgets covered it."},
        {"metric": "models", "value": ", ".join(models)},
        {"metric": "gate_model_for_llm_budget", "value": gate_model},
        {"metric": "llm_gate_enabled", "value": os.getenv("ACQ_ENABLE_LLM_GATE", "")},
        {"metric": "llm_gate_model", "value": os.getenv("ACQ_LLM_MODEL", "")},
        {"metric": "scoring_method", "value": scoring_method},
        {"metric": "region_gate_required", "value": "true"},
        {"metric": "region_gate_rule", "value": "monitoring/LLM/goals require confirmed/probable Kaliningrad Oblast evidence from surface metadata, post/comment context, or dedicated region vector"},
        {"metric": "region_profile_confidence_json", "value": json.dumps(dict(profile_regions), ensure_ascii=False)},
        {"metric": "comments_after_filter", "value": len(records)},
        {"metric": "monitoring_window_comments", "value": usage_scopes.get("monitoring_candidate", 0)},
        {"metric": "historical_calibration_comments", "value": usage_scopes.get("historical_calibration", 0)},
        {"metric": "surfaces_profiled", "value": len(profiles)},
        {"metric": "period_min_created_at", "value": period["period_min_created_at"]},
        {"metric": "period_max_created_at", "value": period["period_max_created_at"]},
        {"metric": "period_days", "value": period["period_days"]},
        {"metric": "platform_comment_records_json", "value": json.dumps(dict(platforms), ensure_ascii=False)},
        {"metric": "surface_type_records_json", "value": json.dumps(dict(surface_types), ensure_ascii=False)},
        {"metric": "relation_records_json", "value": json.dumps(dict(relations), ensure_ascii=False)},
        {"metric": "ACQ_MAX_SURFACES_PER_RUN", "value": os.getenv("ACQ_MAX_SURFACES_PER_RUN", "")},
        {"metric": "ACQ_MAX_MESSAGES_PER_SURFACE", "value": os.getenv("ACQ_MAX_MESSAGES_PER_SURFACE", "")},
        {"metric": "ACQ_MAX_THREADS_PER_SURFACE", "value": os.getenv("ACQ_MAX_THREADS_PER_SURFACE", "")},
        {"metric": "ACQ_MAX_VK_SURFACES_PER_RUN", "value": os.getenv("ACQ_MAX_VK_SURFACES_PER_RUN", "")},
        {"metric": "ACQ_MAX_VK_POSTS_PER_SURFACE", "value": os.getenv("ACQ_MAX_VK_POSTS_PER_SURFACE", "")},
        {"metric": "ACQ_MAX_VK_COMMENTS_PER_POST", "value": os.getenv("ACQ_MAX_VK_COMMENTS_PER_POST", "")},
        {"metric": "ACQ_RUNTIME_DEADLINE_SECONDS", "value": os.getenv("ACQ_RUNTIME_DEADLINE_SECONDS", "")},
        {"metric": "ACQ_ENABLE_VK_MEMBER_PROFILE_DISCOVERY", "value": os.getenv("ACQ_ENABLE_VK_MEMBER_PROFILE_DISCOVERY", "")},
        {"metric": "ACQ_MAX_VK_MEMBER_PROFILES_DISCOVERED_PER_RUN", "value": os.getenv("ACQ_MAX_VK_MEMBER_PROFILES_DISCOVERED_PER_RUN", "")},
        {"metric": "ACQ_MAX_VK_AUTHOR_PROFILES_DISCOVERED_PER_RUN", "value": os.getenv("ACQ_MAX_VK_AUTHOR_PROFILES_DISCOVERED_PER_RUN", "")},
    ]


def _intent_catalog_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for intent_set, phrases in INTENT_SETS.items():
        action = "negative_filter" if intent_set == "negative_intents" else _action_for_intent(intent_set)
        for idx, phrase in enumerate(phrases, start=1):
            rows.append({
                "intent_set": intent_set,
                "intent_set_ru": _INTENT_SET_LABELS_RU.get(intent_set, intent_set),
                "phrase_order": idx,
                "model_phrase": phrase,
                "candidate_action_type": action,
                "is_negative": intent_set == "negative_intents",
                "note_ru": "Эта фраза эмбеддится как эталон смысла; top_intent_phrase — ближайшая такая фраза для строки-кандидата.",
            })
    return rows


_INTENT_SET_LABELS_RU = {
    "route_poi_far_context": "Маршруты/места: широкий контекст",
    "route_poi_medium_interest": "Маршруты/места: интерес",
    "route_poi_close_actionable": "Маршруты/места: можно отвечать",
    "event_far_context": "События: широкий контекст",
    "event_close_question": "События: практический вопрос",
    "organizer_comment_fit": "Организатор: вопросы в комментариях",
    "organizer_event_post_context": "Организатор: пост как контекст для вопроса",
    "event_site_search_or_listing": "Сайт/афиша/подборки",
    "organizer_submission_or_partnership": "Организатор: добавить событие/партнёрство",
    "badge_filter_need": "Фильтры/признаки события",
    "negative_intents": "Минус-смыслы/шум",
}


_SCAN_STATES_TOUCHED = {
    "scanned",
    "scanned_this_run",
    "comments_available",
    "commentability_resolved",
    "resolved_commentability",
    "resolved_no_comments",
    "rejected_after_resolve",
    "checked_inaccessible",
}


_COUNTED_INCREMENT_STATUSES = {
    "newly_discovered_this_run",
    "newly_discovered_and_analyzed_this_run",
    "analyzed_comments_this_run",
    "touched_no_eligible_comments_this_run",
}

_VISIBLE_DELTA_STATUSES = _COUNTED_INCREMENT_STATUSES | {
    "queued_discovered_backlog_this_run",
    "seed_backlog_visible_this_run",
}

_DISCOVERY_FRONTIER_SOURCES = {
    "discovered",
    "linked_discussion",
    "discovered_vk_author",
    "discovered_vk_member",
}

_STATIC_SEED_SOURCES = {
    "route_calibration",
    "telega_in",
    "tg_monitoring",
    "tg_monitoring_canonical",
    "vk_source",
    "smartik_kaliningrad_catalog",
    "vk_social_search",
    "allowlist",
    "seed",
}


_SURFACE_TYPE_RU = {
    "channel": "Telegram-канал",
    "group": "Telegram-группа/чат",
    "linked_discussion": "чат комментариев Telegram-канала",
    "community": "VK-сообщество",
    "profile": "VK-стена личного профиля",
}


def _is_counted_increment_status(status: Any) -> bool:
    return str(status or "") in _COUNTED_INCREMENT_STATUSES


def _is_visible_delta_status(status: Any) -> bool:
    return str(status or "") in _VISIBLE_DELTA_STATUSES


def _surface_increment_status(surface: dict[str, Any], summary: dict[str, Any]) -> str:
    source_run_id = _source_run_id_from_env()
    if not source_run_id:
        return "no_verified_run_id"
    discovered_run = str(surface.get("discovered_in_run_id") or "").strip()
    discovered_this_run = bool(discovered_run and discovered_run == source_run_id)
    source = str(surface.get("source") or "").strip().lower()
    touched_run = (
        surface.get("last_analyzed_run_id")
        or summary.get("last_analyzed_run_id")
        or surface.get("scan_run_id")
        or surface.get("run_id")
        or ""
    )
    touched_this_run = str(touched_run).strip() == source_run_id
    if int(summary.get("comments_embedded") or 0) > 0:
        if discovered_this_run:
            return "newly_discovered_and_analyzed_this_run"
        if touched_this_run:
            return "analyzed_comments_this_run"
        return "analyzed_comments_unverified_run"
    scan_state = str(surface.get("scan_state") or "").strip().lower()
    if discovered_this_run:
        if touched_this_run or scan_state in _SCAN_STATES_TOUCHED:
            return "newly_discovered_this_run"
        if source in _DISCOVERY_FRONTIER_SOURCES:
            return "queued_discovered_backlog_this_run"
        if source in _STATIC_SEED_SOURCES:
            return "seed_backlog_visible_this_run"
        return "queued_discovered_backlog_this_run"
    if touched_this_run and scan_state in _SCAN_STATES_TOUCHED:
        return "touched_no_eligible_comments_this_run"
    return "queued_or_existing"


def _surface_inventory_rows(
    surfaces_by_external: dict[str, dict[str, Any]] | None,
    surface_summaries: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    by_key = {str(row.get("surface_key") or ""): row for row in surface_summaries}
    rows: list[dict[str, Any]] = []
    keys = set(by_key)
    if surfaces_by_external:
        keys.update(str(k) for k in surfaces_by_external)
    for key in sorted(k for k in keys if k):
        surface = (surfaces_by_external or {}).get(key) or {}
        summary = by_key.get(key) or {}
        reach = surface.get("reach") if isinstance(surface.get("reach"), dict) else {}
        recommendation = summary.get("recommendation") or "not_profiled_or_no_comment_signal"
        metadata_region = _assess_record_region({
            "surface_key": key,
            "surface_external_id": key,
            "surface_title": surface.get("title"),
            "surface_url": surface.get("url"),
            "surface_type": surface.get("surface_type"),
            "platform": surface.get("platform"),
            "source": surface.get("source"),
        })
        selection_status = summary.get("selection_status") or _selection_status_from_recommendation(
            str(recommendation),
            surface_status=str(surface.get("status") or ""),
            scan_state=str(surface.get("scan_state") or ""),
        )
        increment_status = _surface_increment_status(surface, summary)
        rows.append({
            "surface_key": key,
            "platform": surface.get("platform") or summary.get("platform"),
            "surface_type": surface.get("surface_type") or summary.get("surface_type"),
            "surface_type_ru": _SURFACE_TYPE_RU.get(
                str(surface.get("surface_type") or summary.get("surface_type") or ""),
                surface.get("surface_type") or summary.get("surface_type"),
            ),
            "surface_title": surface.get("title") or summary.get("surface_title"),
            "surface_url": surface.get("url") or summary.get("surface_url"),
            "status": surface.get("status"),
            "scan_state": surface.get("scan_state"),
            "source": surface.get("source"),
            "members_or_subscribers": reach.get("members") or reach.get("members_count") or summary.get("members_or_subscribers"),
            "recommendation": recommendation,
            "selection_status": selection_status,
            "increment_status": increment_status,
            "is_incremental_last_run": _is_visible_delta_status(increment_status),
            "discovered_at": surface.get("discovered_at"),
            "discovered_in_run_id": surface.get("discovered_in_run_id"),
            "discovery_source_context": surface.get("discovery_source_context"),
            "latest_comment_at": summary.get("latest_comment_at"),
            "latest_event_question_at": summary.get("latest_event_question_at"),
            "latest_route_recommendation_at": summary.get("latest_route_recommendation_at"),
            "days_since_latest_activity": summary.get("days_since_latest_activity"),
            "freshness_status": summary.get("freshness_status") or ("unknown_no_profile" if not summary else ""),
            "region_confidence": summary.get("region_confidence") or metadata_region.get("region_confidence"),
            "region_gate_status": summary.get("region_gate_status") or metadata_region.get("region_gate_status"),
            "region_signal_source": summary.get("region_signal_source") or metadata_region.get("region_signal_source"),
            "region_evidence_ru": summary.get("region_evidence_ru") or metadata_region.get("region_evidence_ru"),
            "region_confirmed_contexts": summary.get("region_confirmed_contexts"),
            "region_probable_contexts": summary.get("region_probable_contexts"),
            "region_unknown_contexts": summary.get("region_unknown_contexts"),
            "region_out_of_region_contexts": summary.get("region_out_of_region_contexts"),
            "comments_embedded": summary.get("comments_embedded") or 0,
            "answerable_question_candidates": summary.get("answerable_question_candidates") or 0,
            "route_poi_questions": summary.get("route_poi_questions") or 0,
            "event_questions": summary.get("event_questions") or 0,
            "event_site_search_questions": summary.get("event_site_search_questions") or 0,
            "ask_clarification_contexts": summary.get("ask_clarification_contexts") or 0,
            "summary_ru": summary.get("summary_ru") or "В этом прогоне полезных сигналов не найдено или площадка не попала в comment retrieval.",
        })
    return sorted(rows, key=lambda r: (
        {"selected": 0, "candidate": 1, "rejected": 2}.get(str(r.get("selection_status")), 3),
        -int(float(r.get("answerable_question_candidates") or 0)),
        str(r.get("platform") or ""),
        str(r.get("surface_title") or r.get("surface_key") or ""),
    ))


def _decision_delta_label(row: dict[str, Any]) -> str:
    status = str(row.get("increment_status") or "")
    selection = str(row.get("selection_status") or "")
    status_ru = {
        "selected": "выбрано для мониторинга",
        "candidate": "кандидат / нужна досборка",
        "rejected": "пока отклонено",
    }.get(selection, selection or "без итогового статуса")
    if status == "newly_discovered_and_analyzed_this_run":
        return f"новая площадка, сразу проанализированы комментарии; решение: {status_ru}"
    if status == "newly_discovered_this_run":
        return f"новая площадка в этом запуске; комментарии ещё не дали профиля; решение: {status_ru}"
    if status == "queued_discovered_backlog_this_run":
        return f"новая ссылка найдена рантаймом и поставлена в очередь; ещё не сканировалась; решение: {status_ru}"
    if status == "seed_backlog_visible_this_run":
        return f"известный seed/backlog показан в карте отчёта; это не новый discovery-прирост; решение: {status_ru}"
    if status == "analyzed_comments_this_run":
        return f"в этом запуске заново обработаны комментарии; решение: {status_ru}"
    if status == "touched_no_eligible_comments_this_run":
        return f"в этом запуске проверена, но полезных комментариев не найдено; решение: {status_ru}"
    return f"не дельта последнего запуска; решение: {status_ru}"


def _decision_delta_rows(surface_inventory: list[dict[str, Any]], *, scope_rows: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    scope = {str(r.get("metric")): r.get("value") for r in (scope_rows or [])}
    verified_increment_source = _has_verified_increment_source(scope) if scope else True
    rows: list[dict[str, Any]] = []
    if not verified_increment_source:
        return [{
            "delta_type": "not_verified",
            "decision_change_ru": "KAGGLE_RUN_ID/ACQ_SOURCE_RUN_ID отсутствует: нельзя доказать, что строки относятся именно к последнему запуску.",
        }]
    for row in surface_inventory:
        if not _is_visible_delta_status(row.get("increment_status")):
            continue
        delta_type = str(row.get("increment_status") or "")
        rows.append({
            "delta_type": delta_type,
            "decision_change_ru": _decision_delta_label(row),
            "surface_key": row.get("surface_key"),
            "platform": row.get("platform"),
            "surface_type": row.get("surface_type"),
            "surface_type_ru": row.get("surface_type_ru"),
            "surface_title": row.get("surface_title"),
            "surface_url": row.get("surface_url"),
            "status": row.get("status"),
            "scan_state": row.get("scan_state"),
            "recommendation": row.get("recommendation"),
            "selection_status": row.get("selection_status"),
            "increment_status": row.get("increment_status"),
            "region_confidence": row.get("region_confidence"),
            "region_evidence_ru": row.get("region_evidence_ru"),
            "comments_embedded": row.get("comments_embedded"),
            "answerable_question_candidates": row.get("answerable_question_candidates"),
            "ask_clarification_contexts": row.get("ask_clarification_contexts"),
            "latest_comment_at": row.get("latest_comment_at"),
            "summary_ru": row.get("summary_ru"),
        })
    return sorted(rows, key=lambda r: (
        {
            "newly_discovered_and_analyzed_this_run": 0,
            "newly_discovered_this_run": 1,
            "analyzed_comments_this_run": 2,
            "touched_no_eligible_comments_this_run": 3,
            "queued_discovered_backlog_this_run": 4,
            "seed_backlog_visible_this_run": 5,
        }.get(str(r.get("delta_type")), 9),
        {"selected": 0, "candidate": 1, "rejected": 2}.get(str(r.get("selection_status")), 3),
        str(r.get("surface_title") or r.get("surface_key") or ""),
    ))


def _summary_count_rows(surface_inventory: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], Counter[str]] = defaultdict(Counter)
    for row in surface_inventory:
        key = (str(row.get("platform") or "unknown"), str(row.get("surface_type") or "unknown"))
        status = str(row.get("selection_status") or "candidate")
        grouped[key]["total"] += 1
        grouped[key][status] += 1
        increment_status = str(row.get("increment_status") or "")
        if increment_status in {"newly_discovered_this_run", "newly_discovered_and_analyzed_this_run"}:
            grouped[key]["newly_discovered_this_run"] += 1
        if increment_status == "queued_discovered_backlog_this_run":
            grouped[key]["queued_discovered_backlog_this_run"] += 1
        if increment_status == "seed_backlog_visible_this_run":
            grouped[key]["seed_backlog_visible_this_run"] += 1
        if _is_visible_delta_status(increment_status):
            grouped[key]["visible_delta_rows_this_run"] += 1
        if _is_counted_increment_status(increment_status):
            grouped[key]["increment_touched_this_run"] += 1
            grouped[key][f"{status}_in_delta_this_run"] += 1
            grouped[key]["comments_embedded_delta_this_run"] += int(float(row.get("comments_embedded") or 0))
            grouped[key]["answerable_questions_delta_this_run"] += int(float(row.get("answerable_question_candidates") or 0))
        if increment_status in {"analyzed_comments_this_run", "newly_discovered_and_analyzed_this_run"}:
            grouped[key]["analyzed_comments_this_run"] += 1
    rows: list[dict[str, Any]] = []
    total_counter: Counter[str] = Counter()
    for (platform, surface_type), counter in sorted(grouped.items()):
        total_counter.update(counter)
        rows.append({
            "platform": platform,
            "surface_type": surface_type,
            "total_surfaces": counter.get("total", 0),
            "selected_surfaces": counter.get("selected", 0),
            "candidate_surfaces": counter.get("candidate", 0),
            "rejected_surfaces": counter.get("rejected", 0),
            "newly_discovered_this_run": counter.get("newly_discovered_this_run", 0),
            "queued_discovered_backlog_this_run": counter.get("queued_discovered_backlog_this_run", 0),
            "seed_backlog_visible_this_run": counter.get("seed_backlog_visible_this_run", 0),
            "visible_delta_rows_this_run": counter.get("visible_delta_rows_this_run", 0),
            "increment_touched_this_run": counter.get("increment_touched_this_run", 0),
            "analyzed_comments_this_run": counter.get("analyzed_comments_this_run", 0),
            "removed_surfaces_this_run": counter.get("removed_surfaces_this_run", 0),
            "selected_in_delta_this_run": counter.get("selected_in_delta_this_run", 0),
            "candidate_in_delta_this_run": counter.get("candidate_in_delta_this_run", 0),
            "rejected_in_delta_this_run": counter.get("rejected_in_delta_this_run", 0),
            "comments_embedded_delta_this_run": counter.get("comments_embedded_delta_this_run", 0),
            "answerable_questions_delta_this_run": counter.get("answerable_questions_delta_this_run", 0),
        })
    rows.insert(0, {
        "platform": "ALL",
        "surface_type": "ALL",
        "total_surfaces": total_counter.get("total", 0),
        "selected_surfaces": total_counter.get("selected", 0),
        "candidate_surfaces": total_counter.get("candidate", 0),
        "rejected_surfaces": total_counter.get("rejected", 0),
        "newly_discovered_this_run": total_counter.get("newly_discovered_this_run", 0),
        "queued_discovered_backlog_this_run": total_counter.get("queued_discovered_backlog_this_run", 0),
        "seed_backlog_visible_this_run": total_counter.get("seed_backlog_visible_this_run", 0),
        "visible_delta_rows_this_run": total_counter.get("visible_delta_rows_this_run", 0),
        "increment_touched_this_run": total_counter.get("increment_touched_this_run", 0),
        "analyzed_comments_this_run": total_counter.get("analyzed_comments_this_run", 0),
        "removed_surfaces_this_run": total_counter.get("removed_surfaces_this_run", 0),
        "selected_in_delta_this_run": total_counter.get("selected_in_delta_this_run", 0),
        "candidate_in_delta_this_run": total_counter.get("candidate_in_delta_this_run", 0),
        "rejected_in_delta_this_run": total_counter.get("rejected_in_delta_this_run", 0),
        "comments_embedded_delta_this_run": total_counter.get("comments_embedded_delta_this_run", 0),
        "answerable_questions_delta_this_run": total_counter.get("answerable_questions_delta_this_run", 0),
    })
    return rows


def _has_verified_increment_source(scope: dict[str, Any]) -> bool:
    return str(scope.get("source_run_provenance") or "") in {"kaggle_run_id", "explicit_source_run_id"}


def _dashboard_summary_rows(
    *,
    summary: dict[str, Any] | None,
    surface_summaries: list[dict[str, Any]],
    scope_rows: list[dict[str, Any]],
    surface_inventory: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    summary = summary or {}
    scope = {str(r.get("metric")): r.get("value") for r in scope_rows}
    monitored = [r for r in surface_summaries if str(r.get("recommendation")) in {"both_monitor_replies_and_ask_clarifications", "monitor_for_reply_opportunities", "ask_organizer_clarification_questions", "sample_more"}]
    no_signal = [r for r in (surface_inventory or []) if int(float(r.get("answerable_question_candidates") or 0)) <= 0]
    counts = _summary_count_rows(surface_inventory or [])
    total = counts[0] if counts else {}
    verified_increment_source = _has_verified_increment_source(scope)
    run_status_ru = (
        "да, источник запуска подтверждён" if verified_increment_source
        else "нет доказанного нового запуска: KAGGLE_RUN_ID/ACQ_SOURCE_RUN_ID отсутствует, инкремент не засчитывается"
    )
    return [
        {"section": "Что это", "metric": "Назначение", "value": "Дашборд показывает, в каких TG/VK группах и обсуждениях есть вопросы для аккуратных ответов или места для уточняющих вопросов организаторам."},
        {"section": "Региональный gate", "metric": "Обязательный регион", "value": "События, маршруты и organizer-вопросы выбираются только если площадка/пост/комментарий подтверждает Калининградскую область."},
        {"section": "Региональный gate", "metric": "Как подтверждается", "value": "Отдельный region vector смотрит название/URL площадки, исходный пост, parent-комментарий и текущий комментарий; явные KGD-зацепки только усиливают evidence, но не заменяют embedding-оценку."},
        {"section": "Региональный gate", "metric": "Что отсекается", "value": "region_unknown и out_of_region не попадают в goal-листы, monitoring_targets и top-N LLM-gate до появления Калининградской зацепки."},
        {"section": "Как читать сначала", "metric": "decision_deltas", "value": "Первый рабочий лист после summary_ru: только площадки, изменившиеся/обработанные в последнем подтверждённом запуске. С него начинается просмотр дельт."},
        {"section": "Как читать сначала", "metric": "processed_comments_last_run", "value": "Показывает последние обработанные комментарии/пост-контексты и колонку criteria_status_ru: прошёл ли контекст критерии отчёта или почему отфильтрован."},
        {"section": "Дельта последнего запуска", "metric": "Был ли новый запуск", "value": run_status_ru},
        {"section": "Дельта последнего запуска", "metric": "Всего площадок в карте отчёта", "value": total.get("total_surfaces", len(surface_inventory or []))},
        {"section": "Дельта последнего запуска", "metric": "Новых площадок уже просканировано (+)", "value": total.get("newly_discovered_this_run", 0)},
        {"section": "Дельта последнего запуска", "metric": "Новых ссылок поставлено в очередь", "value": total.get("queued_discovered_backlog_this_run", 0)},
        {"section": "Дельта последнего запуска", "metric": "Seed/backlog показано, не новый прирост", "value": total.get("seed_backlog_visible_this_run", 0)},
        {"section": "Дельта последнего запуска", "metric": "Строк видно в дельте отчёта", "value": total.get("visible_delta_rows_this_run", 0)},
        {"section": "Дельта последнего запуска", "metric": "Площадок убрано (−)", "value": total.get("removed_surfaces_this_run", 0)},
        {"section": "Дельта последнего запуска", "metric": "Площадок реально затронуто/проверено", "value": total.get("increment_touched_this_run", 0)},
        {"section": "Дельта последнего запуска", "metric": "Из них выбрано для мониторинга", "value": total.get("selected_in_delta_this_run", 0)},
        {"section": "Дельта последнего запуска", "metric": "Из них кандидаты/дособрать", "value": total.get("candidate_in_delta_this_run", 0)},
        {"section": "Дельта последнего запуска", "metric": "Из них отклонено", "value": total.get("rejected_in_delta_this_run", 0)},
        {"section": "Дельта последнего запуска", "metric": "Комментариев в дельте", "value": total.get("comments_embedded_delta_this_run", 0)},
        {"section": "Дельта последнего запуска", "metric": "Вопросов для ответа в дельте", "value": total.get("answerable_questions_delta_this_run", 0)},
        {"section": "Охват", "metric": "Комментариев обработано", "value": summary.get("comments_embedded") or scope.get("comments_after_filter")},
        {"section": "Охват", "metric": "Свежих для мониторинга", "value": scope.get("monitoring_window_comments")},
        {"section": "Охват", "metric": "Исторических для эталонов/контроля", "value": scope.get("historical_calibration_comments")},
        {"section": "Охват", "metric": "Площадок с анализом комментариев", "value": summary.get("surface_profiles_count") or scope.get("surfaces_profiled")},
        {"section": "Охват", "metric": "Все площадки в списке", "value": len(surface_inventory or [])},
        {"section": "Итог", "metric": "Выбрано", "value": total.get("selected_surfaces", 0)},
        {"section": "Итог", "metric": "Кандидаты", "value": total.get("candidate_surfaces", 0)},
        {"section": "Итог", "metric": "Отклонено", "value": total.get("rejected_surfaces", 0)},
        {"section": "Период", "metric": "С даты", "value": scope.get("period_min_created_at")},
        {"section": "Период", "metric": "По дату", "value": scope.get("period_max_created_at")},
        {"section": "Модели", "metric": "Модели смысла", "value": summary.get("models") or scope.get("models")},
        {"section": "Модели", "metric": "Embedding-модель для top-N LLM-очереди", "value": f"{summary.get('recommended_model') or scope.get('gate_model_for_llm_budget')} — только ранжирует строки для LLM-бюджета; обе модели смысла считаются и сравниваются в отчёте."},
        {"section": "Модели", "metric": "LLM-модель финальной проверки", "value": scope.get("llm_gate_model") or os.getenv("ACQ_LLM_MODEL") or "Gemma gate может быть выключен в этом run"},
        {"section": "Итог", "metric": "Где есть хотя бы слабый смысл", "value": len(monitored)},
        {"section": "Итог", "metric": "Где ничего не найдено/не попало", "value": len(no_signal)},
        {"section": "Как читать", "metric": "surface_summary", "value": "Главный лист: где мониторить, сколько вопросов и как часто они встречаются."},
        {"section": "Как читать", "metric": "surface_backlog", "value": "Отдельный лист подтверждённых, кандидатов и ожидающих проверки комментариев/обсуждений площадок; именно там видны event-source publics, которые ещё только ждут resolve/скан."},
        {"section": "Как читать", "metric": "full_surface_list", "value": "Полный список площадок из payload/run: видно, где ничего не нашлось или площадка не анализировалась."},
        {"section": "Как читать", "metric": "summary_counts", "value": "Сводка по каждому типу: всего / выбрано / кандидаты / отклонено и явные +/−/затронуто по последнему запуску."},
        {"section": "Как читать", "metric": "intent_catalog", "value": "Список модельных смыслов, по которым ищем. top_intent_phrase в примерах берётся именно отсюда."},
        {"section": "Ограничения", "metric": "Контекст", "value": "Новые прогоны сохраняют исходный пост/родительский комментарий; в старых артефактах контекст восстановить нельзя."},
        {"section": "Ограничения", "metric": "Не наши темы", "value": "Недвижимость и медицина отсекаются как out_of_scope, если нет явной связи с маршрутом/событием."},
        {"section": "Freshness", "metric": "Активность", "value": f"Для включения в мониторинг учитываются свежие комментарии до {_max_comment_age_days()} дней; площадки без активности больше {_stale_activity_days()} дней отклоняются сейчас, но исторические строки остаются для эталонных вопросов и контроля смыслов."},
    ]


def _write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


_RU_HEADERS = {
    "label": "Метка",
    "action_class": "Класс действия",
    "is_actionable_reply_opportunity": "Можно отвечать?",
    "false_positive_type": "Тип ошибки",
    "model_disagreement_bucket": "Согласие моделей",
    "model_name": "Модель",
    "intent_set": "Группа смысла",
    "intent_set_ru": "Смысл по-русски",
    "model_phrase": "Модельная фраза",
    "phrase_order": "№",
    "is_negative": "Минус-смысл?",
    "note_ru": "Пояснение",
    "score": "Оценка",
    "rank_global": "Ранг общий",
    "rank_within_surface": "Ранг в площадке",
    "surface_key": "ID площадки",
    "platform": "Платформа",
    "surface_type": "Тип (техн.)",
    "surface_type_ru": "Тип площадки",
    "surface_title": "Название",
    "surface_url": "Ссылка на площадку/чат",
    "status": "Статус",
    "selection_status": "Итоговый статус",
    "scan_state": "Сканирование",
    "source": "Источник",
    "increment_status": "Инкремент",
    "is_incremental_last_run": "Инкремент последнего запуска?",
    "discovered_at": "Добавлено/найдено",
    "discovered_in_run_id": "Найдено в run_id",
    "discovery_source_context": "Где найдено",
    "relation": "Тип контекста",
    "is_post": "Это пост?",
    "author_id": "Автор",
    "created_at": "Дата",
    "comment_age_days": "Возраст, дней",
    "candidate_usage_scope": "Роль строки",
    "is_within_monitoring_window": "В окне мониторинга?",
    "context_url": "Ссылка на конкретный комментарий/пост",
    "text_snapshot": "Текущий текст (комментарий или пост)",
    "analysis_context_snapshot": "Общий контекст для модели",
    "current_comment_text": "Текущий комментарий",
    "current_post_text": "Текущий пост / анонс",
    "source_post_text": "Пост, под которым написан комментарий",
    "source_post_text_snapshot": "Пост, под которым написан комментарий",
    "reply_parent_comment_text": "Комментарий-родитель (если это reply)",
    "reply_parent_text_snapshot": "Комментарий-родитель (если это reply)",
    "evidence_type_ru": "Что это за текст",
    "future_goal_ru": "Будущая цель",
    "top_intent_phrase": "Ближайшая модельная фраза",
    "positive_score": "Похожесть +",
    "negative_score": "Похожесть −",
    "funnel_bucket": "Воронка",
    "candidate_action_type": "Что можно делать",
    "destination_hint": "Место/направление",
    "transport_hint": "Транспорт",
    "question_signal": "Есть вопрос?",
    "candidate_noise_type": "Диагностика шума",
    "semantic_exclusion_type": "Смысловой стоп-фильтр",
    "semantic_candidate_rejected": "Отклонено стоп-фильтром?",
    "event_temporal_status": "Дата события: статус",
    "event_temporal_evidence_ru": "Дата события: доказательство",
    "event_latest_detected_date": "Дата события: найдено",
    "event_temporal_gate_passed": "Дата события: прошла gate?",
    "intent_text_supported": "Смысл подтверждён текстом",
    "pre_llm_candidate_eligible": "В LLM-gate?",
    "llm_gate_selection_basis": "Основа отбора в LLM",
    "recommendation": "Рекомендация",
    "members_or_subscribers": "Участники/подписчики (если собрано)",
    "period_min_created_at": "Период анализа с",
    "period_max_created_at": "Период анализа по",
    "period_days": "Дней",
    "period_label": "Период",
    "latest_comment_at": "Последний комментарий",
    "latest_event_question_at": "Последний event-вопрос",
    "latest_route_recommendation_at": "Последняя route-рекомендация",
    "days_since_latest_activity": "Дней без активности",
    "freshness_status": "Свежесть",
    "region_confidence": "Регион: уверенность",
    "region_gate_status": "Региональный gate",
    "region_signal_source": "Регион: источник сигнала",
    "region_evidence_ru": "Регион: доказательство",
    "region_score": "Регион score",
    "region_positive_score": "Регион +",
    "region_negative_score": "Регион −",
    "top_region_phrase": "Региональная модельная фраза",
    "top_negative_region_phrase": "Минус-регион фраза",
    "region_positive_hint": "Регион + зацепка",
    "region_out_of_region_hint": "Регион − зацепка",
    "region_confirmed_contexts": "Регион confirmed",
    "region_probable_contexts": "Регион probable",
    "region_unknown_contexts": "Регион unknown",
    "region_out_of_region_contexts": "Регион out",
    "region_score_max": "Регион score max",
    "region_score_p95": "Регион score p95",
    "region_positive_score_max": "Регион + max",
    "region_negative_score_max": "Регион − max",
    "region_catalog_type": "Тип регионального эталона",
    "analysis_max_comment_age_days": "Окно анализа, дней",
    "stale_activity_days": "Порог неактивности, дней",
    "unique_commenters": "Уникальные авторы комментариев за период",
    "unique_commenters_note": "Авторы: примечание",
    "comments_total": "Всего комментариев",
    "comments_embedded": "В анализе",
    "comments_per_day": "Комм./день",
    "comments_per_week": "Комм./нед.",
    "comments_per_30d": "Комм./30д",
    "comments_per_90d": "Комм./90д",
    "latest_100_comments": "Последние N",
    "latest_100_min_created_at": "Latest-100 с",
    "latest_100_max_created_at": "Latest-100 по",
    "latest_100_period_days": "Latest-100 дней",
    "latest_100_period_label": "Latest-100 период",
    "summary_ru": "Вывод",
    "answerable_question_candidates": "Вопросы для ответа",
    "answerable_questions_per_30d": "Вопросы/30д",
    "answerable_questions_per_90d": "Вопросы/90д",
    "answerable_questions_per_100_comments": "Вопросы/100 комм.",
    "ask_clarification_contexts": "Где спрашивать самим",
    "ask_contexts_per_30d": "Уточнения/30д",
    "eligible_comment_contexts": "Контексты-комментарии",
    "source_post_contexts": "Контексты-посты",
    "route_poi_questions": "Маршруты",
    "event_questions": "События",
    "event_site_search_questions": "Поиск/афиша",
    "organizer_submission_questions": "Организаторам",
    "badge_filter_questions": "Фильтры",
    "filtered_noise_contexts": "Шум/диагн.",
    "relation_counts_json": "Типы контекстов",
    "dominant_detected_interests": "Темы",
    "monitoring_decision_hint": "Решение",
    "monitoring_reason": "Причина",
    "answerable_examples": "Примеры ответов",
    "ask_question_examples": "Примеры уточнений",
    "pattern": "Паттерн",
    "canonical_question_ru": "Эталонный вопрос",
    "count": "Количество",
    "intent_sets": "Группы смыслов",
    "models": "Модели",
    "llm_gate_model": "LLM-модель",
    "llm_gate_enabled": "LLM gate включён?",
    "example_questions": "Примеры вопросов",
    "example_urls": "Ссылки примеров",
    "metric": "Показатель",
    "value": "Значение",
    "section": "Раздел",
    "total_surfaces": "Всего",
    "selected_surfaces": "Выбрано",
    "candidate_surfaces": "Кандидаты",
    "rejected_surfaces": "Отклонено",
    "newly_discovered_this_run": "Новых просканировано",
    "queued_discovered_backlog_this_run": "Новых в очередь",
    "seed_backlog_visible_this_run": "Seed/backlog не +",
    "visible_delta_rows_this_run": "Строк в дельте",
    "increment_touched_this_run": "Реально проверено",
    "analyzed_comments_this_run": "С анализом комм.",
    "real_question_examples_total": "Реальных вопросов",
    "monitoring_candidate_examples": "Свежих примеров",
    "historical_calibration_examples": "Исторических примеров",
    "surfaces_count": "Площадок",
    "source_note_ru": "Пояснение",
    "delta_type": "Тип дельты",
    "decision_change_ru": "Что изменилось / как читать",
    "criteria_status": "Критерии (код)",
    "criteria_status_ru": "Прошёл критерии?",
    "analysis_kind": "Что анализировали",
    "removed_surfaces_this_run": "Удалено (−)",
    "selected_in_delta_this_run": "Выбрано в дельте",
    "candidate_in_delta_this_run": "Кандидатов в дельте",
    "rejected_in_delta_this_run": "Отклонено в дельте",
    "comments_embedded_delta_this_run": "Комм. в дельте",
    "answerable_questions_delta_this_run": "Вопросов в дельте",
    "processed_in_run_id": "Обработано в run_id",
    "last_processed_at": "Последняя обработка",
    "models_matched": "Модели, где найдено",
    "model_count": "Моделей",
    "best_score": "Лучшая оценка",
    "llm_final_check_ru": "LLM-проверка",
    "llm_queue_status_ru": "Очередь LLM",
    "is_in_llm_gate_queue": "В top-N LLM queue?",
    "goal_sheet": "Лист цели",
    "backlog_bucket": "Backlog (код)",
    "backlog_bucket_ru": "Backlog: как читать",
}


def _format_date_ru(value: Any) -> Any:
    dt = _parse_created_at(value)
    if dt is None:
        return value
    return dt.strftime("%d.%m.%Y")


def _format_datetime_ru(value: Any) -> Any:
    dt = _parse_created_at(value)
    if dt is None:
        return value
    return dt.strftime("%d.%m.%Y %H:%M")


def _xlsx_value(header: str, value: Any) -> Any:
    if header == "members_or_subscribers" and (value is None or value == ""):
        return "не собрано"
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    if header in {"discovered_at", "report_generated_at", "generated_at", "last_processed_at"}:
        return _format_datetime_ru(value)
    if header.endswith("_created_at") or header == "created_at" or header.endswith("_at") or header.endswith("_date") or header in {"period_min_created_at", "period_max_created_at"}:
        return _format_date_ru(value)
    if isinstance(value, float):
        return round(value, 1)
    if isinstance(value, str):
        stripped = value.strip()
        if header.endswith("_days") or "_per_" in header or header.endswith("_score") or header == "score":
            try:
                return round(float(stripped), 1)
            except Exception:
                return value
    return value


def _append_grouped_header(ws: Any, headers: list[str], groups: dict[str, str] | None = None) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    groups = groups or {}
    group_values = [groups.get(h, "") for h in headers]
    ws.append(group_values)
    ws.append([_RU_HEADERS.get(h, h) for h in headers])
    for row_idx, color in [(1, "B7DEE8"), (2, "D9EAF7")]:
        for cell in ws[row_idx]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    start = 1
    while start <= len(headers):
        group = group_values[start - 1]
        end = start
        while end < len(headers) and group_values[end] == group:
            end += 1
        if group and end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        start = end + 1


_FUTURE_GOAL_RU = {
    "trip_route_poi_recommendation": "отвечать на вопросы рекомендацией маршрута",
    "event_recommendation_reply": "отвечать на вопросы о событиях",
    "organizer_visibility_clarification": "задавать уточняющие вопросы о событиях",
    "event_site_search_or_listing": "отвечать ссылкой на афишу/поиск/подборку событий",
    "organizer_submission_or_partnership": "подсказывать организаторам, куда добавить событие / инфопартнёрство",
    "badge_filter_need": "отвечать подборкой или фильтром событий: детям, бесплатно, Пушкинская карта и т.п.",
}


def _run_id_datetime(run_id: Any) -> str:
    raw = str(run_id or "").strip()
    match = re.match(r"^(\d{8})-(\d{6})$", raw)
    if not match:
        return ""
    try:
        dt = datetime.strptime("".join(match.groups()), "%Y%m%d%H%M%S").replace(tzinfo=timezone.utc)
        return dt.isoformat()
    except Exception:
        return ""


def _display_row(row: dict[str, Any]) -> dict[str, Any]:
    enriched = dict(row)
    text = str(enriched.get("text_snapshot") or enriched.get("text") or "")
    source_post = str(enriched.get("source_post_text_snapshot") or "")
    parent = str(enriched.get("reply_parent_text_snapshot") or "")
    is_source_post = _is_source_post_context(enriched)
    action_type = str(enriched.get("candidate_action_type") or "")
    if not action_type:
        action_type = _action_for_intent(str(enriched.get("intent_set") or ""))
    if is_source_post:
        enriched["evidence_type_ru"] = enriched.get("evidence_type_ru") or "исходный пост/анонс, не комментарий"
        enriched["current_comment_text"] = ""
        enriched["current_post_text"] = text
        # There is no “post under the comment” for a source-post row; repeating
        # the same text in both columns made the report look like a comment was
        # duplicated. Keep the source-post text only in current_post_text.
        enriched["source_post_text"] = ""
    else:
        enriched["evidence_type_ru"] = enriched.get("evidence_type_ru") or "пользовательский комментарий/сообщение"
        enriched["current_comment_text"] = text
        enriched["current_post_text"] = ""
        enriched["source_post_text"] = source_post
    enriched["reply_parent_comment_text"] = parent
    if not enriched.get("processed_in_run_id") and enriched.get("run_id"):
        enriched["processed_in_run_id"] = enriched.get("run_id")
    if not enriched.get("last_processed_at"):
        enriched["last_processed_at"] = _run_id_datetime(enriched.get("processed_in_run_id") or enriched.get("run_id"))
    criteria = str(enriched.get("criteria_status") or "")
    if criteria.startswith("rejected"):
        enriched.setdefault("future_goal_ru", "")
    else:
        enriched.setdefault("future_goal_ru", _FUTURE_GOAL_RU.get(action_type, "диагностика/не выбрано для будущего действия"))
    return enriched


def _append_data_rows(
    ws: Any,
    rows: list[dict[str, Any]],
    headers: list[str],
    *,
    hyperlink_field: str | None = None,
    style: str | None = None,
) -> None:
    from openpyxl.styles import PatternFill

    selected_fill = PatternFill("solid", fgColor="C6EFCE")
    increment_fill = PatternFill("solid", fgColor="FFF2CC")
    goal_fill = PatternFill("solid", fgColor="E2F0D9")
    rejected_fill = PatternFill("solid", fgColor="F4CCCC")
    neutral_fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        display_row = _display_row(row)
        ws.append([_xlsx_value(h, display_row.get(h)) for h in headers])
        if style == "surface":
            fill = None
            if str(display_row.get("selection_status") or "") == "selected":
                fill = selected_fill
            elif _is_visible_delta_status(display_row.get("increment_status")):
                fill = increment_fill
            if fill is not None:
                for cell in ws[ws.max_row]:
                    cell.fill = fill
        elif style == "summary_counts":
            delta_headers = [
                "newly_discovered_this_run", "queued_discovered_backlog_this_run", "seed_backlog_visible_this_run",
                "visible_delta_rows_this_run", "removed_surfaces_this_run", "increment_touched_this_run", "analyzed_comments_this_run",
                "selected_in_delta_this_run", "candidate_in_delta_this_run", "rejected_in_delta_this_run",
                "comments_embedded_delta_this_run", "answerable_questions_delta_this_run",
            ]
            if any(int(float(row.get(h) or 0)) > 0 for h in delta_headers if h in headers):
                for header in delta_headers:
                    if header in headers:
                        ws.cell(ws.max_row, headers.index(header) + 1).fill = increment_fill
        elif style == "goal":
            for cell in ws[ws.max_row]:
                cell.fill = goal_fill
        elif style == "processed":
            criteria = str(display_row.get("criteria_status") or "")
            if criteria.startswith("accepted"):
                fill = selected_fill
            elif criteria.startswith("rejected"):
                fill = rejected_fill
            else:
                fill = neutral_fill
            for cell in ws[ws.max_row]:
                cell.fill = fill
        if hyperlink_field and display_row.get(hyperlink_field):
            url_cell = ws.cell(ws.max_row, headers.index(hyperlink_field) + 1)
            url_cell.hyperlink = str(display_row.get(hyperlink_field))
            url_cell.style = "Hyperlink"


def _criteria_status(row: dict[str, Any]) -> tuple[str, str]:
    if _report_real_question_row(row, allow_historical=False):
        return "accepted_reply_candidate", "соответствует: реальный вопрос, можно рассматривать ответ"
    if _is_source_post_context(row) and _report_candidate_eligible(row, allow_source_posts=True, allow_historical=False):
        return "accepted_ask_context", "соответствует: это пост/контекст, где можно задать уточняющий вопрос организатору"
    scope = str(row.get("candidate_usage_scope") or "")
    if scope == "historical_calibration":
        return "rejected_historical", "не fresh-дельта: исторический пример только для эталонов/QA"
    if scope and scope != "monitoring_candidate":
        return "rejected_wrong_scope", f"не в свежем окне мониторинга: {scope}"
    if _candidate_region_required(row) and not _candidate_region_eligible(row):
        confidence = str(row.get("region_confidence") or "unknown")
        evidence = str(row.get("region_evidence_ru") or "нет доказательства Калининградской области")
        return f"rejected_region:{confidence}", f"не прошёл региональный gate Калининградской области: {evidence}"
    noise = str(row.get("candidate_noise_type") or "").strip()
    if _hard_semantic_rejected(row):
        exclusion = str(row.get("semantic_exclusion_type") or noise or "")
        if exclusion == "out_of_scope_gasoline_availability":
            return "rejected_noise:out_of_scope_gasoline_availability", "отфильтровано: вопросы наличия бензина/топлива/АЗС не наша acquisition-тема"
        if exclusion in {"past_event", "past_event_signal"}:
            evidence = str(row.get("event_temporal_evidence_ru") or "событие выглядит прошедшим")
            return f"rejected_temporal:{exclusion}", f"отфильтровано: работаем с будущими событиями; {evidence}"
    if _is_source_post_context(row):
        return "rejected_source_post_context", "это исходный пост/контекст, не пользовательский комментарий для ответа"
    if _report_noise_rejected(row) or noise.startswith("out_of_scope"):
        return f"rejected_noise:{noise or 'noise'}", f"отфильтровано как шум/не наша тема: {noise or 'noise'}"
    if not _truthy_value(row.get("question_signal")):
        return "rejected_no_question", "нет вопроса от пользователя"
    if not _truthy_value(row.get("intent_text_supported")):
        return "rejected_intent_not_supported", "векторная близость есть, но текст не подтверждает нужный смысл"
    if _row_score(row) < _report_min_comment_score():
        return "rejected_low_score", "ниже порога человекочитаемого отчёта"
    return "rejected_report_gate", "не прошло строгие критерии отчёта; векторный кандидат оставлен только для диагностики"


def _processed_comment_rows(candidates: list[dict[str, Any]], *, gate_model: str | None = None, limit: int = 500) -> list[dict[str, Any]]:
    monitoring_rows = [r for r in candidates if str(r.get("candidate_usage_scope") or "") == "monitoring_candidate"]
    by_context: dict[str, dict[str, Any]] = {}
    for row in monitoring_rows:
        key = str(row.get("context_url") or row.get("comment_id") or row.get("text_snapshot") or "")
        if not key:
            continue
        bucket = by_context.setdefault(key, {"rows": [], "models": set()})
        bucket["rows"].append(row)
        if row.get("model_name"):
            bucket["models"].add(str(row.get("model_name")))
    best: list[dict[str, Any]] = []
    for bucket in by_context.values():
        rows = bucket["rows"]
        if gate_model:
            gate_rows = [r for r in rows if str(r.get("model_name") or "") == gate_model]
            rows = gate_rows or rows
        row = sorted(rows, key=lambda r: float(r.get("rank_global") or 999999))[0]
        enriched = dict(row)
        models = sorted(bucket["models"])
        enriched["models_matched"] = ", ".join(models)
        enriched["model_count"] = len(models)
        enriched["best_score"] = max(_row_score(r) for r in bucket["rows"])
        best.append(enriched)
    out: list[dict[str, Any]] = []
    for row in sorted(best, key=lambda r: str(r.get("created_at") or ""), reverse=True)[:limit]:
        status, status_ru = _criteria_status(row)
        enriched = dict(row)
        enriched["criteria_status"] = status
        enriched["criteria_status_ru"] = status_ru
        enriched["analysis_kind"] = "post_context" if _is_source_post_context(row) else "user_comment"
        out.append(enriched)
    return out


def _rejected_noise_rows(processed_rows: list[dict[str, Any]], *, limit: int = 300) -> list[dict[str, Any]]:
    return [
        row for row in processed_rows
        if str(row.get("criteria_status") or "").startswith("rejected")
    ][:limit]


_GOAL_SHEET_SPECS: dict[str, dict[str, Any]] = {
    "goal_ask_event_details": {
        "title": "уточняющие вопросы организаторам",
        "actions": {"organizer_visibility_clarification"},
        "source_posts": True,
    },
    "goal_reply_events": {
        "title": "ответы на вопросы о событиях",
        "actions": {"event_recommendation_reply", "event_site_search_or_listing", "badge_filter_need"},
        "source_posts": False,
    },
    "goal_reply_routes": {
        "title": "ответы рекомендацией маршрута",
        "actions": {"trip_route_poi_recommendation"},
        "source_posts": False,
    },
    "goal_other_acq": {
        "title": "прочие acquisition-кандидаты",
        "actions": {"organizer_submission_or_partnership"},
        "source_posts": False,
    },
}


def _goal_candidate_rows(
    candidates: list[dict[str, Any]],
    *,
    actions: set[str],
    source_posts: bool | None,
    limit: int = 200,
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in candidates:
        action = str(row.get("candidate_action_type") or _action_for_intent(str(row.get("intent_set") or "")))
        if action not in actions:
            continue
        is_source_post = _is_source_post_context(row)
        if source_posts is True and not is_source_post:
            continue
        if source_posts is True and _truthy_value(row.get("question_signal")):
            continue
        if source_posts is False and is_source_post:
            continue
        if not _report_candidate_eligible(row, allow_source_posts=True, allow_historical=False):
            continue
        key = (str(row.get("context_url") or row.get("comment_id") or row.get("text_snapshot") or ""), action)
        if not key[0]:
            continue
        bucket = grouped.setdefault(key, {"best": row, "models": set(), "intent_sets": set(), "phrases": set()})
        bucket["models"].add(str(row.get("model_name") or "unknown"))
        bucket["intent_sets"].add(str(row.get("intent_set") or ""))
        if row.get("top_intent_phrase"):
            bucket["phrases"].add(str(row.get("top_intent_phrase")))
        if _row_score(row) > _row_score(bucket["best"]):
            bucket["best"] = row
    out: list[dict[str, Any]] = []
    for bucket in grouped.values():
        best = dict(bucket["best"])
        models = sorted(m for m in bucket["models"] if m)
        best["models_matched"] = ", ".join(models)
        best["model_count"] = len(models)
        best["best_score"] = _row_score(best)
        best["future_goal_ru"] = _FUTURE_GOAL_RU.get(str(best.get("candidate_action_type") or ""), "")
        best["llm_final_check_ru"] = "нужна финальная LLM-проверка перед реальным ответом/вопросом"
        best["llm_queue_status_ru"] = "в top-N LLM queue" if _truthy_value(best.get("is_in_llm_gate_queue")) else "векторный кандидат; ждёт LLM gate"
        out.append(best)
    return sorted(out, key=lambda r: (-int(r.get("model_count") or 0), -float(r.get("best_score") or 0), str(r.get("created_at") or "")))[:limit]


def _monitoring_target_rows(surface_summaries: list[dict[str, Any]], *, limit: int = 500) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in surface_summaries:
        answerable = int(float(row.get("answerable_question_candidates") or 0))
        ask = int(float(row.get("ask_clarification_contexts") or 0))
        status = str(row.get("selection_status") or "")
        if status == "rejected" and answerable <= 0 and ask <= 0:
            continue
        if answerable <= 0 and ask <= 0 and status != "selected":
            continue
        enriched = dict(row)
        enriched["last_processed_at"] = _run_id_datetime(row.get("last_analyzed_run_id") or row.get("discovered_in_run_id"))
        rows.append(enriched)
    return sorted(rows, key=lambda r: (
        {"selected": 0, "candidate": 1, "rejected": 2}.get(str(r.get("selection_status") or ""), 9),
        -int(float(r.get("answerable_question_candidates") or 0)),
        -int(float(r.get("ask_clarification_contexts") or 0)),
        str(r.get("surface_title") or r.get("surface_key") or ""),
    ))[:limit]


def _surface_backlog_rows(surface_inventory: list[dict[str, Any]], *, limit: int = 1000) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in surface_inventory:
        status = str(row.get("status") or "").strip().lower()
        selection = str(row.get("selection_status") or "").strip().lower()
        scan_state = str(row.get("scan_state") or "").strip().lower()
        source = str(row.get("source") or "").strip().lower()
        if selection == "selected":
            bucket = "selected_for_monitoring"
            bucket_ru = "подтверждено/выбрано для постоянного мониторинга"
        elif selection == "candidate" and int(float(row.get("answerable_question_candidates") or 0)) + int(float(row.get("ask_clarification_contexts") or 0)) > 0:
            bucket = "candidate_has_signal"
            bucket_ru = "кандидат: есть сигнал, нужно больше данных/LLM/ручной отсмотр"
        elif status in {"needs_comment_resolve", "seed", "candidate", "approved"} or scan_state in {"queued_waiting_replyable_budget", "queued_waiting_scan_budget", "queued", "waiting_scan"}:
            bucket = "waiting_scan_or_commentability"
            bucket_ru = "ожидает проверки комментариев/обсуждения или скана"
        elif source in {"tg_monitoring", "tg_monitoring_canonical", "vk_source", "telega_in", "smartik_kaliningrad_catalog", "vk_social_search"} and selection != "rejected":
            bucket = "known_source_backlog"
            bucket_ru = "известный источник в backlog, ещё нет полезного профиля"
        else:
            continue
        enriched = dict(row)
        enriched["backlog_bucket"] = bucket
        enriched["backlog_bucket_ru"] = bucket_ru
        rows.append(enriched)
    return sorted(rows, key=lambda r: (
        {
            "selected_for_monitoring": 0,
            "candidate_has_signal": 1,
            "waiting_scan_or_commentability": 2,
            "known_source_backlog": 3,
        }.get(str(r.get("backlog_bucket") or ""), 9),
        str(r.get("source") or ""),
        str(r.get("surface_title") or r.get("surface_key") or ""),
    ))[:limit]


def _write_xlsx(
    path: Path,
    rows: list[dict[str, Any]],
    *,
    surface_summaries: list[dict[str, Any]] | None = None,
    model_examples: dict[str, list[dict[str, Any]]] | None = None,
    model_ask_contexts: dict[str, list[dict[str, Any]]] | None = None,
    question_patterns: list[dict[str, Any]] | None = None,
    canonical_questions: list[dict[str, Any]] | None = None,
    scope_rows: list[dict[str, Any]] | None = None,
    surface_inventory: list[dict[str, Any]] | None = None,
    summary_counts: list[dict[str, Any]] | None = None,
    dashboard_rows: list[dict[str, Any]] | None = None,
    intent_catalog_rows: list[dict[str, Any]] | None = None,
    decision_delta_rows: list[dict[str, Any]] | None = None,
    processed_comment_rows: list[dict[str, Any]] | None = None,
    rejected_noise_rows: list[dict[str, Any]] | None = None,
    goal_candidate_rows: dict[str, list[dict[str, Any]]] | None = None,
    monitoring_target_rows: list[dict[str, Any]] | None = None,
    surface_backlog_rows: list[dict[str, Any]] | None = None,
    region_catalog_rows: list[dict[str, Any]] | None = None,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return
    if surface_backlog_rows is None and surface_inventory is not None:
        surface_backlog_rows = _surface_backlog_rows(surface_inventory)
    wb = Workbook()
    ws = wb.active
    ws.title = "manual_review"
    headers = [
        "label", "action_class", "is_actionable_reply_opportunity", "false_positive_type", "model_disagreement_bucket",
        "model_name", "intent_set", "score", "rank_global", "rank_within_surface", "surface_key", "platform", "surface_type",
        "relation", "is_post", "evidence_type_ru", "author_id", "created_at", "comment_age_days", "candidate_usage_scope", "context_url",
        "current_comment_text", "reply_parent_comment_text", "source_post_text", "current_post_text", "future_goal_ru",
        "region_confidence", "region_gate_status", "region_evidence_ru", "region_signal_source", "region_score",
        "top_intent_phrase", "positive_score", "negative_score", "funnel_bucket",
        "destination_hint", "transport_hint", "question_signal", "candidate_noise_type", "semantic_exclusion_type",
        "event_temporal_status", "event_temporal_evidence_ru", "intent_text_supported",
        "is_in_llm_gate_queue", "llm_queue_status_ru", "llm_gate_selection_basis",
    ]
    manual_groups = {h: "Разметка" for h in headers[:5]} | {h: "Скоринг" for h in headers[5:10]} | {h: "Площадка" for h in headers[10:20]} | {h: "Текст и контекст" for h in headers[20:25]} | {h: "Регион" for h in headers[25:30]} | {h: "Решение" for h in headers[30:]}
    _append_grouped_header(ws, headers, manual_groups)
    _append_data_rows(ws, rows, headers, hyperlink_field="context_url")
    for idx, width in enumerate([12, 28, 28, 24, 28, 34, 28, 12, 12, 16, 28, 10, 18, 18, 10, 28, 16, 18, 14, 28, 48, 80, 70, 90, 90, 54, 55, 14, 14, 14, 22, 16, 14, 24, 20, 24], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A3"

    if dashboard_rows:
        dash = wb.create_sheet("summary_ru", 0)
        dash_headers = ["section", "metric", "value"]
        _append_grouped_header(dash, dash_headers, {h: "Краткое объяснение" for h in dash_headers})
        _append_data_rows(dash, dashboard_rows, dash_headers)
        dash.freeze_panes = "A3"
        for idx, width in enumerate([18, 32, 120], start=1):
            dash.column_dimensions[get_column_letter(idx)].width = width

    if decision_delta_rows is not None:
        delta = wb.create_sheet("decision_deltas", 1 if dashboard_rows else 0)
        delta_headers = [
            "delta_type", "decision_change_ru", "surface_key", "platform", "surface_type_ru", "surface_type",
            "surface_title", "surface_url", "status", "scan_state", "recommendation", "selection_status",
            "increment_status", "region_confidence", "region_evidence_ru", "comments_embedded", "answerable_question_candidates", "ask_clarification_contexts",
            "latest_comment_at", "summary_ru",
        ]
        delta_groups = {h: "Что изменилось" for h in delta_headers[:2]} | {h: "Площадка" for h in delta_headers[2:8]} | {h: "Решение" for h in delta_headers[8:13]} | {h: "Сигналы" for h in delta_headers[13:]}
        _append_grouped_header(delta, delta_headers, delta_groups)
        if not decision_delta_rows:
            delta.append(["empty", "В этом артефакте нет подтверждённых дельт последнего запуска. Проверьте KAGGLE_RUN_ID/ACQ_SOURCE_RUN_ID и лист full_surface_list.", *[""] * (len(delta_headers) - 2)])
        _append_data_rows(delta, decision_delta_rows, delta_headers, hyperlink_field="surface_url", style="surface")
        delta.freeze_panes = "A3"
        for idx, width in enumerate([30, 70, 30, 10, 30, 18, 44, 46, 24, 24, 36, 18, 30, 14, 16, 16, 20, 90], start=1):
            delta.column_dimensions[get_column_letter(idx)].width = width

        run_delta = wb.create_sheet("run_delta_sources")
        run_delta_headers = [
            "decision_change_ru", "surface_title", "surface_url", "platform", "surface_type_ru",
            "selection_status", "increment_status", "region_confidence", "region_evidence_ru", "comments_embedded", "answerable_question_candidates",
            "ask_clarification_contexts", "latest_comment_at", "summary_ru",
        ]
        run_delta_groups = {h: "Что изменилось в последнем запуске" for h in run_delta_headers[:3]} | {h: "Площадка" for h in run_delta_headers[3:5]} | {h: "Вердикт" for h in run_delta_headers[5:]}
        _append_grouped_header(run_delta, run_delta_headers, run_delta_groups)
        if not decision_delta_rows:
            run_delta.append(["В этом запуске нет подтверждённых изменений источников/площадок.", *[""] * (len(run_delta_headers) - 1)])
        _append_data_rows(run_delta, decision_delta_rows, run_delta_headers, hyperlink_field="surface_url", style="surface")
        run_delta.freeze_panes = "A3"
        for idx, width in enumerate([74, 44, 48, 10, 30, 18, 30, 14, 16, 16, 18, 90], start=1):
            run_delta.column_dimensions[get_column_letter(idx)].width = width

    if monitoring_target_rows is not None:
        targets = wb.create_sheet("monitoring_targets")
        target_headers = [
            "selection_status", "recommendation", "surface_title", "surface_url", "platform", "surface_type_ru",
            "last_analyzed_run_id", "last_processed_at", "discovered_at", "discovered_in_run_id",
            "latest_comment_at", "latest_event_question_at", "latest_route_recommendation_at",
            "region_confidence", "region_evidence_ru",
            "answerable_question_candidates", "route_poi_questions", "event_questions", "event_site_search_questions",
            "ask_clarification_contexts", "comments_embedded", "unique_commenters", "members_or_subscribers", "summary_ru",
        ]
        target_groups = {h: "Вердикт для постоянного мониторинга" for h in target_headers[:4]} | {h: "Площадка" for h in target_headers[4:10]} | {h: "Накопительный результат" for h in target_headers[10:]}
        _append_grouped_header(targets, target_headers, target_groups)
        if not monitoring_target_rows:
            targets.append(["empty", "Пока нет площадок с подтверждённым полезным сигналом для постоянного мониторинга.", *[""] * (len(target_headers) - 2)])
        _append_data_rows(targets, monitoring_target_rows, target_headers, hyperlink_field="surface_url", style="surface")
        targets.freeze_panes = "A3"
        for idx, width in enumerate([18, 36, 44, 48, 10, 30, 20, 20, 20, 20, 18, 18, 18, 16, 14, 14, 14, 16, 14, 14, 20, 90], start=1):
            targets.column_dimensions[get_column_letter(idx)].width = width

    if surface_backlog_rows is not None:
        backlog = wb.create_sheet("surface_backlog")
        backlog_headers = [
            "backlog_bucket_ru", "backlog_bucket", "selection_status", "recommendation",
            "surface_title", "surface_url", "platform", "surface_type_ru", "surface_type",
            "status", "scan_state", "source", "increment_status", "is_incremental_last_run",
            "discovered_at", "discovered_in_run_id", "latest_comment_at", "latest_event_question_at", "latest_route_recommendation_at",
            "region_confidence", "region_evidence_ru", "comments_embedded", "answerable_question_candidates",
            "route_poi_questions", "event_questions", "event_site_search_questions", "ask_clarification_contexts",
            "members_or_subscribers", "summary_ru",
        ]
        backlog_groups = {h: "Куда смотреть дальше" for h in backlog_headers[:4]} | {h: "Площадка" for h in backlog_headers[4:12]} | {h: "Прогресс" for h in backlog_headers[12:19]} | {h: "Потенциал" for h in backlog_headers[19:]}
        _append_grouped_header(backlog, backlog_headers, backlog_groups)
        if not surface_backlog_rows:
            backlog.append(["Нет подтверждённых/кандидатных/ожидающих площадок в текущем payload.", *[""] * (len(backlog_headers) - 1)])
        _append_data_rows(backlog, surface_backlog_rows, backlog_headers, hyperlink_field="surface_url", style="surface")
        backlog.freeze_panes = "A3"
        for idx, width in enumerate([62, 30, 18, 36, 44, 48, 10, 30, 18, 24, 30, 24, 30, 14, 20, 20, 18, 18, 18, 16, 90, 14, 14, 14, 14, 16, 14, 20, 90], start=1):
            backlog.column_dimensions[get_column_letter(idx)].width = width

    if goal_candidate_rows is not None:
        goal_headers = [
            "future_goal_ru", "llm_final_check_ru", "llm_queue_status_ru", "models_matched", "model_count", "best_score",
            "surface_title", "surface_key", "platform", "surface_type", "context_url", "created_at", "comment_age_days",
            "region_confidence", "region_evidence_ru",
            "evidence_type_ru", "current_comment_text", "reply_parent_comment_text", "source_post_text", "current_post_text",
            "top_intent_phrase", "intent_set", "candidate_action_type", "event_temporal_status", "event_temporal_evidence_ru",
            "processed_in_run_id", "last_processed_at",
        ]
        for sheet_name, spec in _GOAL_SHEET_SPECS.items():
            sheet_rows = goal_candidate_rows.get(sheet_name, [])
            goal = wb.create_sheet(sheet_name)
            goal_groups = {h: "Кандидат по цели" for h in goal_headers[:6]} | {h: "Где найдено" for h in goal_headers[6:13]} | {h: "Регион" for h in goal_headers[13:15]} | {h: "Текст и контекст" for h in goal_headers[15:20]} | {h: "Почему подходит" for h in goal_headers[20:]}
            _append_grouped_header(goal, goal_headers, goal_groups)
            if not sheet_rows:
                goal.append([f"Нет кандидатов: {spec.get('title')}", *[""] * (len(goal_headers) - 1)])
            _append_data_rows(goal, sheet_rows, goal_headers, hyperlink_field="context_url", style="goal")
            goal.freeze_panes = "A3"
            for idx, width in enumerate([54, 54, 28, 42, 10, 12, 44, 30, 10, 18, 48, 18, 14, 28, 90, 70, 90, 90, 60, 28, 34, 20, 20], start=1):
                goal.column_dimensions[get_column_letter(idx)].width = width

    if processed_comment_rows is not None:
        processed_index = 2 if (dashboard_rows and decision_delta_rows is not None) else (1 if (dashboard_rows or decision_delta_rows is not None) else 0)
        processed = wb.create_sheet("processed_comments_last_run", processed_index)
        processed_headers = [
        "criteria_status_ru", "criteria_status", "analysis_kind", "model_name", "models_matched", "model_count", "score", "best_score", "rank_global",
        "surface_key", "platform", "surface_type", "relation", "is_post", "created_at", "comment_age_days",
        "context_url", "evidence_type_ru", "current_comment_text", "reply_parent_comment_text", "source_post_text", "current_post_text",
        "future_goal_ru", "region_confidence", "region_evidence_ru", "top_intent_phrase", "intent_set",
        "candidate_action_type", "question_signal", "candidate_noise_type", "semantic_exclusion_type",
        "event_temporal_status", "event_temporal_evidence_ru", "intent_text_supported",
        "is_in_llm_gate_queue", "llm_queue_status_ru",
    ]
        processed_groups = {h: "Критерии" for h in processed_headers[:3]} | {h: "Скоринг" for h in processed_headers[3:9]} | {h: "Где найдено" for h in processed_headers[9:16]} | {h: "Текст и контекст" for h in processed_headers[16:23]} | {h: "Регион" for h in processed_headers[23:25]} | {h: "Почему принято/отклонено" for h in processed_headers[25:]}
        _append_grouped_header(processed, processed_headers, processed_groups)
        _append_data_rows(processed, processed_comment_rows, processed_headers, hyperlink_field="context_url", style="processed")
        processed.freeze_panes = "A3"
        for idx, width in enumerate([62, 30, 16, 34, 42, 10, 12, 12, 12, 30, 10, 18, 22, 10, 18, 14, 48, 28, 90, 70, 90, 90, 54, 60, 28, 34, 14, 30, 18, 18, 28], start=1):
            processed.column_dimensions[get_column_letter(idx)].width = width

    if rejected_noise_rows is not None:
        noise_index = None
        if processed_comment_rows is not None:
            noise_index = 3 if dashboard_rows and decision_delta_rows is not None else 2
        noise = wb.create_sheet("rejected_noise_examples", noise_index)
        noise_headers = [
            "criteria_status_ru", "criteria_status", "analysis_kind", "surface_key", "platform", "surface_type",
            "relation", "created_at", "context_url", "evidence_type_ru", "current_comment_text", "reply_parent_comment_text", "source_post_text", "current_post_text",
            "region_confidence", "region_evidence_ru", "top_intent_phrase", "score", "positive_score", "negative_score",
            "candidate_noise_type", "semantic_exclusion_type", "event_temporal_status", "event_temporal_evidence_ru",
        ]
        noise_groups = {h: "Почему это не кандидат" for h in noise_headers[:3]} | {h: "Где найдено" for h in noise_headers[3:9]} | {h: "Текст и контекст" for h in noise_headers[9:14]} | {h: "Регион" for h in noise_headers[14:16]} | {h: "Шум/диагностика" for h in noise_headers[16:]}
        _append_grouped_header(noise, noise_headers, noise_groups)
        _append_data_rows(noise, rejected_noise_rows, noise_headers, hyperlink_field="context_url", style="processed")
        noise.freeze_panes = "A3"
        for idx, width in enumerate([62, 30, 16, 30, 10, 18, 22, 18, 48, 28, 90, 70, 90, 90, 60, 12, 12, 12, 30], start=1):
            noise.column_dimensions[get_column_letter(idx)].width = width

    if surface_summaries:
        surf = wb.create_sheet("surface_summary")
        surf_headers = [
            "recommendation", "selection_status", "surface_key", "platform", "surface_type_ru", "surface_type", "surface_title", "surface_url",
            "members_or_subscribers", "period_min_created_at", "period_max_created_at", "period_days",
            "latest_comment_at", "latest_event_question_at", "latest_route_recommendation_at", "days_since_latest_activity", "freshness_status", "unique_commenters",
            "region_confidence", "region_gate_status", "region_evidence_ru", "region_confirmed_contexts", "region_probable_contexts", "region_unknown_contexts", "region_out_of_region_contexts",
            "comments_total", "comments_embedded", "comments_per_day", "comments_per_week", "comments_per_30d", "comments_per_90d",
            "latest_100_comments", "latest_100_min_created_at", "latest_100_max_created_at", "latest_100_period_days",
            "increment_status", "is_incremental_last_run", "discovered_at", "discovered_in_run_id", "discovery_source_context",
            "summary_ru", "answerable_question_candidates", "answerable_questions_per_30d", "answerable_questions_per_90d",
            "answerable_questions_per_100_comments", "ask_clarification_contexts", "ask_contexts_per_30d", "eligible_comment_contexts", "source_post_contexts",
            "route_poi_questions", "event_questions", "event_site_search_questions", "organizer_submission_questions", "badge_filter_questions",
            "filtered_noise_contexts", "relation_counts_json", "unique_commenters_note", "answerable_examples", "ask_question_examples",
        ]
        surf_groups = {
            **{h: "Площадка" for h in surf_headers[:8]},
            **{h: "Период, свежесть и объём" for h in surf_headers[8:18]},
            **{h: "Регион" for h in surf_headers[18:25]},
            **{h: "Потенциал" for h in surf_headers[25:46]},
            **{h: "Пояснения" for h in surf_headers[42:]},
        }
        _append_grouped_header(surf, surf_headers, surf_groups)
        _append_data_rows(surf, surface_summaries, surf_headers, hyperlink_field="surface_url", style="surface")
        surf.freeze_panes = "A3"
        for idx, width in enumerate([34, 30, 10, 18, 30, 42, 16, 22, 22, 12, 18, 14, 14, 14, 14, 14, 14, 14, 22, 22, 14, 70, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 60, 24, 90, 90], start=1):
            surf.column_dimensions[get_column_letter(idx)].width = width

    if surface_inventory:
        inv = wb.create_sheet("full_surface_list")
        inv_headers = [
            "surface_key", "platform", "surface_type_ru", "surface_type", "surface_title", "surface_url", "status", "scan_state", "source",
            "members_or_subscribers", "recommendation", "selection_status", "increment_status", "is_incremental_last_run",
            "discovered_at", "discovered_in_run_id", "latest_comment_at", "latest_event_question_at", "latest_route_recommendation_at", "days_since_latest_activity", "freshness_status",
            "region_confidence", "region_gate_status", "region_evidence_ru",
            "comments_embedded", "answerable_question_candidates",
            "route_poi_questions", "event_questions", "event_site_search_questions", "ask_clarification_contexts", "summary_ru",
        ]
        inv_groups = {h: "Площадка" for h in inv_headers[:9]} | {h: "Результат анализа" for h in inv_headers[9:]}
        _append_grouped_header(inv, inv_headers, inv_groups)
        _append_data_rows(inv, surface_inventory, inv_headers, hyperlink_field="surface_url", style="surface")
        inv.freeze_panes = "A3"
        for idx, width in enumerate([30, 10, 18, 34, 44, 20, 24, 22, 14, 36, 16, 16, 14, 14, 14, 14, 14, 14, 14, 90], start=1):
            inv.column_dimensions[get_column_letter(idx)].width = width

    if summary_counts:
        counts = wb.create_sheet("summary_counts")
        count_headers = [
            "platform", "surface_type", "total_surfaces", "selected_surfaces", "candidate_surfaces", "rejected_surfaces",
            "newly_discovered_this_run", "queued_discovered_backlog_this_run", "seed_backlog_visible_this_run",
            "visible_delta_rows_this_run", "removed_surfaces_this_run", "increment_touched_this_run", "analyzed_comments_this_run",
            "selected_in_delta_this_run", "candidate_in_delta_this_run", "rejected_in_delta_this_run",
            "comments_embedded_delta_this_run", "answerable_questions_delta_this_run",
        ]
        count_groups = {h: "Тип площадки" for h in count_headers[:2]} | {h: "Итог" for h in count_headers[2:6]} | {h: "Дельта последнего запуска" for h in count_headers[6:]}
        _append_grouped_header(counts, count_headers, count_groups)
        _append_data_rows(counts, summary_counts, count_headers, style="summary_counts")
        counts.freeze_panes = "A3"
        for idx, width in enumerate([14, 24, 12, 12, 12, 12], start=1):
            counts.column_dimensions[get_column_letter(idx)].width = width

    for model_name, example_rows in (model_examples or {}).items():
        model_key = str(model_name).lower()
        if "multilingual-e5" in model_key:
            sheet_name = "answerable_e5_base"
        elif "bge-m3" in model_key:
            sheet_name = "answerable_bge_m3"
        else:
            sheet_name = ("answerable_" + re.sub(r"[^A-Za-z0-9]+", "_", model_name).strip("_"))[:31] or "answerable_model"
        ex = wb.create_sheet(sheet_name)
        ex_headers = [
            "model_name", "score", "rank_global", "rank_within_surface", "surface_key", "platform", "surface_type",
            "relation", "is_post", "created_at", "comment_age_days", "candidate_usage_scope", "intent_set", "candidate_action_type",
            "candidate_noise_type", "semantic_exclusion_type", "event_temporal_status", "event_temporal_evidence_ru",
            "llm_gate_selection_basis", "region_confidence", "region_evidence_ru", "context_url", "evidence_type_ru", "current_comment_text", "reply_parent_comment_text", "source_post_text", "current_post_text", "future_goal_ru",
            "top_intent_phrase", "destination_hint", "transport_hint",
        ]
        ex_groups = {h: "Скоринг" for h in ex_headers[:4]} | {h: "Площадка" for h in ex_headers[4:12]} | {h: "Решение" for h in ex_headers[12:16]} | {h: "Регион" for h in ex_headers[16:18]} | {h: "Текст и контекст" for h in ex_headers[18:25]} | {h: "Действие" for h in ex_headers[25:]}
        _append_grouped_header(ex, ex_headers, ex_groups)
        if not example_rows:
            ex.append(["Нет успешных вопросов/сообщений, по которым можно доказать потенциальный ответ, в этом run."])
        _append_data_rows(ex, example_rows, ex_headers, hyperlink_field="context_url")
        ex.freeze_panes = "A3"
        for idx, width in enumerate([34, 12, 12, 16, 30, 10, 18, 18, 10, 12, 28, 30, 24, 34, 30, 45, 48, 28, 90, 70, 90, 90, 54, 55, 22, 18], start=1):
            ex.column_dimensions[get_column_letter(idx)].width = width

    for model_name, context_rows in (model_ask_contexts or {}).items():
        model_key = str(model_name).lower()
        if "multilingual-e5" in model_key:
            sheet_name = "ask_contexts_e5_base"
        elif "bge-m3" in model_key:
            sheet_name = "ask_contexts_bge_m3"
        else:
            sheet_name = ("ask_contexts_" + re.sub(r"[^A-Za-z0-9]+", "_", model_name).strip("_"))[:31] or "ask_contexts_model"
        ctx = wb.create_sheet(sheet_name)
        ctx_headers = [
            "model_name", "score", "rank_global", "rank_within_surface", "surface_key", "platform", "surface_type",
            "relation", "is_post", "created_at", "comment_age_days", "candidate_usage_scope", "intent_set", "candidate_action_type",
            "candidate_noise_type", "semantic_exclusion_type", "event_temporal_status", "event_temporal_evidence_ru",
            "llm_gate_selection_basis", "region_confidence", "region_evidence_ru", "context_url", "evidence_type_ru", "current_comment_text", "reply_parent_comment_text", "source_post_text", "current_post_text", "future_goal_ru",
            "top_intent_phrase", "destination_hint", "transport_hint",
        ]
        ctx_groups = {h: "Скоринг" for h in ctx_headers[:4]} | {h: "Площадка" for h in ctx_headers[4:12]} | {h: "Решение" for h in ctx_headers[12:16]} | {h: "Регион" for h in ctx_headers[16:18]} | {h: "Это пост/контекст для вопроса, не user-comment reply" for h in ctx_headers[18:25]} | {h: "Действие" for h in ctx_headers[25:]}
        _append_grouped_header(ctx, ctx_headers, ctx_groups)
        if not context_rows:
            ctx.append(["Нет контекстов постов, по которым можно было бы самим задать уточняющий вопрос организатору, в этом run."])
        _append_data_rows(ctx, context_rows, ctx_headers, hyperlink_field="context_url")
        ctx.freeze_panes = "A3"
        for idx, width in enumerate([34, 12, 12, 16, 30, 10, 18, 18, 10, 12, 28, 30, 24, 34, 30, 45, 48, 28, 90, 70, 90, 90, 54, 55, 22, 18], start=1):
            ctx.column_dimensions[get_column_letter(idx)].width = width

    if question_patterns is not None:
        qp = wb.create_sheet("question_patterns")
        qp_headers = [
            "surface_key", "platform", "surface_type", "pattern", "candidate_action_type",
            "canonical_question_ru", "count", "intent_sets", "models", "example_questions", "example_urls",
        ]
        qp_groups = {h: "Где найдено" for h in qp_headers[:3]} | {h: "Эталон вопроса" for h in qp_headers[3:7]} | {h: "Модели и примеры" for h in qp_headers[7:]}
        _append_grouped_header(qp, qp_headers, qp_groups)
        _append_data_rows(qp, question_patterns, qp_headers)
        qp.freeze_panes = "A3"
        for idx, width in enumerate([30, 10, 18, 28, 34, 70, 10, 45, 45, 90, 70], start=1):
            qp.column_dimensions[get_column_letter(idx)].width = width

    if canonical_questions is not None:
        cq = wb.create_sheet("canonical_questions")
        cq_headers = [
            "pattern", "candidate_action_type", "canonical_question_ru", "real_question_examples_total",
            "monitoring_candidate_examples", "historical_calibration_examples", "surfaces_count",
            "example_questions", "example_urls", "source_note_ru",
        ]
        cq_groups = {h: "Эталон вопроса" for h in cq_headers[:7]} | {h: "Реальные примеры" for h in cq_headers[7:9]} | {"source_note_ru": "Пояснение"}
        _append_grouped_header(cq, cq_headers, cq_groups)
        _append_data_rows(cq, canonical_questions, cq_headers)
        cq.freeze_panes = "A3"
        for idx, width in enumerate([28, 34, 76, 14, 14, 14, 14, 100, 80, 100], start=1):
            cq.column_dimensions[get_column_letter(idx)].width = width

    if intent_catalog_rows:
        cat = wb.create_sheet("intent_catalog")
        cat_headers = ["intent_set", "intent_set_ru", "phrase_order", "model_phrase", "candidate_action_type", "is_negative", "note_ru"]
        cat_groups = {h: "Группа смысла" for h in cat_headers[:3]} | {h: "Модельная фраза" for h in cat_headers[3:]}
        _append_grouped_header(cat, cat_headers, cat_groups)
        _append_data_rows(cat, intent_catalog_rows, cat_headers)
        cat.freeze_panes = "A3"
        for idx, width in enumerate([32, 38, 8, 80, 34, 12, 90], start=1):
            cat.column_dimensions[get_column_letter(idx)].width = width

    if region_catalog_rows:
        reg = wb.create_sheet("region_catalog")
        reg_headers = ["region_catalog_type", "phrase_order", "model_phrase", "is_negative", "note_ru"]
        reg_groups = {h: "Региональный embedding-gate" for h in reg_headers}
        _append_grouped_header(reg, reg_headers, reg_groups)
        _append_data_rows(reg, region_catalog_rows, reg_headers)
        reg.freeze_panes = "A3"
        for idx, width in enumerate([34, 8, 90, 12, 110], start=1):
            reg.column_dimensions[get_column_letter(idx)].width = width

    if scope_rows:
        scope = wb.create_sheet("scope")
        _append_grouped_header(scope, ["metric", "value"], {"metric": "Охват запуска", "value": "Охват запуска"})
        _append_data_rows(scope, scope_rows, ["metric", "value"])
        scope.freeze_panes = "A3"
        scope.column_dimensions["A"].width = 36
        scope.column_dimensions["B"].width = 120

    summary = wb.create_sheet("summary")
    _append_grouped_header(summary, ["metric", "value"], {"metric": "Файл", "value": "Файл"})
    _append_data_rows(summary, [
        {"metric": "rows", "value": len(rows)},
        {"metric": "generated_at", "value": datetime.now(timezone.utc).isoformat()},
    ], ["metric", "value"])

    preferred_order = [
        "summary_ru",
        "run_delta_sources",
        "monitoring_targets",
        "surface_backlog",
        "goal_ask_event_details",
        "goal_reply_events",
        "goal_reply_routes",
        "goal_other_acq",
        "surface_summary",
        "full_surface_list",
        "summary_counts",
        "decision_deltas",
        "processed_comments_last_run",
        "rejected_noise_examples",
        "manual_review",
        "answerable_e5_base",
        "answerable_bge_m3",
        "ask_contexts_e5_base",
        "ask_contexts_bge_m3",
        "question_patterns",
        "canonical_questions",
        "intent_catalog",
        "region_catalog",
        "scope",
        "summary",
    ]
    insert_at = 0
    for sheet_name in preferred_order:
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]
        wb._sheets.remove(sheet)
        wb._sheets.insert(insert_at, sheet)
        insert_at += 1
    wb.save(path)


def _select_manual_rows(candidates: list[dict[str, Any]], *, max_rows: int) -> list[dict[str, Any]]:
    rows = candidates[: max_rows // 2]
    # Include near-threshold and lower-ranked rows for calibration, deterministic sample.
    if len(candidates) > len(rows):
        step = max(1, len(candidates) // max(1, max_rows - len(rows)))
        rows.extend(candidates[len(rows)::step][: max_rows - len(rows)])
    out: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row in rows[:max_rows]:
        key = (str(row.get("context_url")), str(row.get("model_name")))
        if key in seen:
            continue
        seen.add(key)
        enriched = dict(row)
        enriched.setdefault("label", "")
        enriched.setdefault("is_actionable_reply_opportunity", "")
        enriched.setdefault("false_positive_type", "")
        out.append(enriched)
    return out


def _report_md(
    summary: dict[str, Any],
    profiles: list[dict[str, Any]],
    candidates: list[dict[str, Any]],
    speed_rows: list[dict[str, Any]],
    surface_summaries: list[dict[str, Any]] | None = None,
) -> str:
    top_route = [c for c in candidates if c.get("candidate_action_type") == "trip_route_poi_recommendation"][:10]
    lines = [
        "# Subscriber Acquisition Comment Semantic Retrieval Dry Run",
        "",
        "## Executive summary",
        f"- Stage: `{STAGE_NAME}`",
        f"- Comments embedded: {summary.get('comments_embedded', 0)} / {summary.get('comments_total', 0)}",
        f"- Surfaces profiled: {summary.get('surface_profiles_count', 0)}",
        f"- Candidate rows: {summary.get('candidate_count', 0)}",
        f"- Recommended gate model: `{summary.get('recommended_model')}`",
        "- No external Telegram/VK sends and no full-comment LLM pass were used.",
        "",
        "## Speed benchmark",
        "| model | batch | max_length | comments/sec | total sec | peak RAM MB |",
        "| --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for row in speed_rows:
        lines.append(f"| {row.get('model_name')} | {row.get('batch_size')} | {row.get('max_length')} | {row.get('comments_per_sec', 0):.2f} | {row.get('total_sec', 0):.2f} | {row.get('peak_ram_mb') or ''} |")
    lines.extend(["", "## Top monitoring surfaces", "| decision | surface | comments | interests | reason |", "| --- | --- | ---: | --- | --- |"])
    for profile in sorted(profiles, key=lambda p: (p.get("monitoring_decision_hint") != "monitor", -(p.get("comments_embedded") or 0)))[:20]:
        lines.append(f"| {profile.get('monitoring_decision_hint')} | {profile.get('surface_key')} | {profile.get('comments_embedded')} | {', '.join(profile.get('dominant_detected_interests') or [])} | {profile.get('monitoring_reason')} |")
    if surface_summaries:
        lines.extend(["", "## Surface decision summary", "| recommendation | surface | comments | summary |", "| --- | --- | ---: | --- |"])
        for row in surface_summaries[:20]:
            lines.append(f"| {row.get('recommendation')} | {row.get('surface_key')} | {row.get('comments_embedded')} | {str(row.get('summary_ru') or '').replace('|', ' ')} |")
    lines.extend(["", "## Best route/POI examples", "| score | surface | url | text |", "| ---: | --- | --- | --- |"])
    for row in top_route:
        text = str(row.get("text_snapshot") or "").replace("|", " ")[:160]
        lines.append(f"| {row.get('score', 0):.4f} | {row.get('surface_key')} | {row.get('context_url')} | {text} |")
    return "\n".join(lines) + "\n"


def _load_models_from_env() -> list[str]:
    models = _json_env("ACQ_COMMENT_RETRIEVAL_MODELS_JSON", DEFAULT_MODELS)
    if isinstance(models, str):
        models = [models]
    out: list[str] = []
    for model in models:
        clean = str(model).strip()
        if clean and clean not in out:
            out.append(clean)
    # Product contract: every collected comment/post-context is embedded by
    # both supported semantic models. Operators may add extra models, but cannot
    # accidentally downgrade a live discovery run to a one-model smoke.
    for required in REQUIRED_MODELS:
        if required not in out:
            out.append(required)
    return out or list(REQUIRED_MODELS)


def _batch_for_model(model_name: str) -> int:
    key = "ACQ_COMMENT_RETRIEVAL_BGE_BATCH_SIZE" if "bge" in model_name.lower() else "ACQ_COMMENT_RETRIEVAL_E5_BATCH_SIZE"
    default = 8 if "bge" in model_name.lower() else 32
    return _int_env(key, default, min_value=1)


def run_comment_semantic_retrieval(
    comment_records: list[dict[str, Any]],
    *,
    surfaces_by_external: dict[str, dict[str, Any]] | None = None,
    output_dir: str | Path | None = None,
    backend: Any | None = None,
    progress_callback: Callable[[str, dict[str, Any]], None] | None = None,
) -> dict[str, Any]:
    started = time.perf_counter()
    output_path = Path(output_dir or os.getenv("ACQ_OUTPUT_DIR") or "/kaggle/working")
    output_path.mkdir(parents=True, exist_ok=True)
    scoring_method = os.getenv("ACQ_COMMENT_RETRIEVAL_SCORING_METHOD") or DEFAULT_SCORING_METHOD
    max_length = _int_env("ACQ_COMMENT_RETRIEVAL_MAX_LENGTH", 128, min_value=16)
    max_llm_candidates = _int_env("ACQ_COMMENT_RETRIEVAL_MAX_LLM_CANDIDATES", 24, min_value=1)
    max_manual_rows = _int_env("ACQ_COMMENT_RETRIEVAL_MANUAL_SAMPLE_ROWS", 800, min_value=20)
    models = _load_models_from_env()
    source_run_id = _source_run_id_from_env()
    records = _dedupe_records(comment_records)
    _attach_surface_metadata_to_records(records, surfaces_by_external)
    backend = backend or SentenceTransformerBackend()
    progress_callback = progress_callback or (lambda _phase, _payload: None)
    progress_callback("loading_comments", {"comments_total": len(comment_records), "comments_after_filter": len(records), "progress_percent": 5})

    positive_phrases = {name: phrases for name, phrases in INTENT_SETS.items() if name != "negative_intents"}
    negative_phrases = INTENT_SETS["negative_intents"]
    all_candidates: list[dict[str, Any]] = []
    speed_rows: list[dict[str, Any]] = []
    distributions: list[dict[str, Any]] = []

    for model_index, model_name in enumerate(models, start=1):
        model_started = time.perf_counter()
        batch_size = _batch_for_model(model_name)
        progress_callback("embedding_intents", {"model_name": model_name, "progress_percent": 5 + int(10 * model_index / max(1, len(models)))})
        intent_vectors: dict[str, list[Any]] = {}
        intent_embed_sec = 0.0
        for intent_set, phrases in positive_phrases.items():
            t0 = time.perf_counter()
            intent_vectors[intent_set] = _to_list_matrix(backend.encode(phrases, model_name=model_name, is_query=True, batch_size=batch_size, max_length=max_length))
            intent_embed_sec += time.perf_counter() - t0
        t0 = time.perf_counter()
        negative_vectors = _to_list_matrix(backend.encode(negative_phrases, model_name=model_name, is_query=True, batch_size=batch_size, max_length=max_length))
        intent_embed_sec += time.perf_counter() - t0
        t0 = time.perf_counter()
        region_positive_vectors = _to_list_matrix(backend.encode(REGION_POSITIVE_PHRASES, model_name=model_name, is_query=True, batch_size=batch_size, max_length=max_length))
        region_negative_vectors = _to_list_matrix(backend.encode(REGION_NEGATIVE_PHRASES, model_name=model_name, is_query=True, batch_size=batch_size, max_length=max_length))
        intent_embed_sec += time.perf_counter() - t0

        comments = [str(r.get("analysis_text") or r.get("text") or "") for r in records]
        region_contexts = [_record_region_text(r) for r in records]
        progress_callback("embedding_comments", {"model_name": model_name, "comments_total": len(comments), "comments_processed": 0, "progress_percent": 20})
        t0 = time.perf_counter()
        comment_vectors = _to_list_matrix(backend.encode(comments, model_name=model_name, is_query=False, batch_size=batch_size, max_length=max_length)) if comments else []
        comment_embedding_sec = time.perf_counter() - t0
        t0 = time.perf_counter()
        region_vectors = _to_list_matrix(backend.encode(region_contexts, model_name=model_name, is_query=False, batch_size=batch_size, max_length=max_length)) if region_contexts else []
        region_embedding_sec = time.perf_counter() - t0
        progress_callback("scoring", {"model_name": model_name, "comments_processed": len(comments), "progress_percent": 60})
        t0 = time.perf_counter()
        model_rows: list[dict[str, Any]] = []
        for rec, vec, region_vec in zip(records, comment_vectors, region_vectors):
            region_scores = _score_region_context(region_vec, region_positive_vectors, region_negative_vectors)
            region_assessment = _assess_record_region(rec, region_scores)
            scored = _score_comment(vec, intent_vectors, negative_vectors)
            for row in scored:
                enriched = dict(rec)
                enriched.update(row)
                enriched.update(region_assessment)
                enriched["model_name"] = model_name
                enriched["max_length"] = max_length
                enriched["batch_size"] = batch_size
                enriched["scoring_method"] = scoring_method
                enriched["candidate_action_type"] = _action_for_intent(str(row.get("intent_set")))
                enriched["target_hint"] = _route_target_hint(str(rec.get("analysis_text") or rec.get("text") or ""), str(row.get("intent_set")))
                enriched["destination_hint"] = enriched["target_hint"].get("destination_hint")
                enriched["transport_hint"] = enriched["target_hint"].get("transport_hint")
                _apply_text_quality_to_candidate(enriched, scoring_method=scoring_method)
                model_rows.append(enriched)
        model_rows = _rank_candidates(model_rows, scoring_method=scoring_method)
        scoring_sec = time.perf_counter() - t0
        all_candidates.extend(model_rows)
        scores = [float(r.get("score") or 0.0) for r in model_rows]
        distributions.append({
            "model_name": model_name,
            "scoring_method": scoring_method,
            "count": len(scores),
            "p50": _percentile(scores, 0.50),
            "p90": _percentile(scores, 0.90),
            "p95": _percentile(scores, 0.95),
            "p99": _percentile(scores, 0.99),
            "max": max(scores) if scores else 0.0,
        })
        total_sec = time.perf_counter() - model_started
        comments_per_sec = len(records) / total_sec if total_sec > 0 else 0.0
        speed_rows.append({
            "model_name": model_name,
            "device": os.getenv("ACQ_COMMENT_RETRIEVAL_DEVICE") or "auto",
            "batch_size": batch_size,
            "max_length": max_length,
            "comments_total": len(records),
            "intent_embedding_sec": round(intent_embed_sec, 4),
            "comment_embedding_sec": round(comment_embedding_sec, 4),
            "region_embedding_sec": round(region_embedding_sec, 4),
            "scoring_sec": round(scoring_sec, 4),
            "total_sec": round(total_sec, 4),
            "comments_per_sec": round(comments_per_sec, 4),
            "comments_per_hour": round(comments_per_sec * 3600, 2),
            "peak_ram_mb": round(_peak_ram_mb() or 0.0, 2),
            "cpu": platform.processor() or platform.machine(),
        })

    # Cross-model ranking and disagreement labels.
    all_candidates = _rank_candidates(all_candidates, scoring_method=scoring_method)
    by_context_model = {(str(c.get("context_url")), str(c.get("model_name")), str(c.get("intent_set"))): c for c in all_candidates}
    models_set = {str(m) for m in models}
    for c in all_candidates:
        context = str(c.get("context_url"))
        intent = str(c.get("intent_set"))
        present = {m for m in models_set if (context, m, intent) in by_context_model and int(by_context_model[(context, m, intent)].get("rank_global") or 999999) <= max(1000, max_llm_candidates * 10)}
        if len(present) == len(models_set):
            bucket = "both_models"
        elif str(c.get("model_name")) in present:
            bucket = f"{c.get('model_name')}_only"
        else:
            bucket = "near_threshold"
        c["model_disagreement_bucket"] = bucket

    gate_model = os.getenv("ACQ_COMMENT_RETRIEVAL_GATE_MODEL") or (DEFAULT_GATE_MODEL if DEFAULT_GATE_MODEL in models else models[0])
    gate_candidates = [
        c for c in all_candidates
        if c.get("model_name") == gate_model
        and c.get("candidate_usage_scope") == "monitoring_candidate"
        and _candidate_region_eligible(c)
        and not _hard_semantic_rejected(c)
    ]
    gate_candidates = _rank_candidates(gate_candidates, scoring_method=scoring_method)[:max_llm_candidates]
    gate_candidate_keys = {
        (str(c.get("context_url") or c.get("comment_id") or c.get("text_snapshot") or ""), str(c.get("model_name") or ""))
        for c in gate_candidates
    }
    for c in all_candidates:
        c["is_in_llm_gate_queue"] = (
            str(c.get("context_url") or c.get("comment_id") or c.get("text_snapshot") or ""),
            str(c.get("model_name") or ""),
        ) in gate_candidate_keys
        c["llm_queue_status_ru"] = "в top-N LLM queue" if c["is_in_llm_gate_queue"] else "нет"

    progress_callback("surface_profile", {"surfaces": len({r.get('surface_key') for r in records}), "progress_percent": 78})
    profiles: list[dict[str, Any]] = []
    rows_by_surface: dict[str, list[dict[str, Any]]] = defaultdict(list)
    all_gate_rows_by_surface: dict[str, list[dict[str, Any]]] = defaultdict(list)
    records_by_surface: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in all_candidates:
        # Profiles are based on selected gate model to avoid double-counting two benchmark models.
        if row.get("model_name") == gate_model:
            all_gate_rows_by_surface[str(row.get("surface_key") or "unknown")].append(row)
            if row.get("candidate_usage_scope") == "monitoring_candidate" and _report_candidate_eligible(row):
                rows_by_surface[str(row.get("surface_key") or "unknown")].append(row)
    for rec in records:
        records_by_surface[str(rec.get("surface_key") or "unknown")].append(rec)
    for surface_key, surface_records in records_by_surface.items():
        profile = _surface_profile(
            surface_key,
            surface_records,
            rows_by_surface.get(surface_key, []),
            scoring_method=scoring_method,
            region_rows=all_gate_rows_by_surface.get(surface_key, []),
        )
        if source_run_id:
            profile["last_analyzed_run_id"] = source_run_id
        if surfaces_by_external and surface_key in surfaces_by_external:
            s = surfaces_by_external[surface_key]
            profile["surface_title"] = s.get("title")
            profile["surface_url"] = s.get("url")
            profile["status"] = s.get("status")
            profile["scan_state"] = s.get("scan_state")
            profile["source"] = s.get("source")
            profile["discovered_at"] = s.get("discovered_at")
            profile["discovered_in_run_id"] = s.get("discovered_in_run_id")
            profile["discovery_source_context"] = s.get("discovery_source_context")
            reach = s.get("reach") if isinstance(s.get("reach"), dict) else {}
            profile["members_or_subscribers"] = (
                reach.get("members")
                or reach.get("members_count")
                or reach.get("subscribers")
                or reach.get("participants")
                or None
            )
        profiles.append(profile)
    surface_summaries = _surface_decision_summaries(
        profiles,
        eligible_rows_by_surface=rows_by_surface,
        all_gate_rows_by_surface=all_gate_rows_by_surface,
    )
    # Canonical question/pattern catalogs may use historical calibration rows
    # after the strict report-quality gate, but monitoring/surface decisions
    # above stay limited to fresh monitoring candidates.
    question_patterns = _build_question_patterns(all_candidates)
    canonical_questions = _canonical_question_catalog_rows(all_candidates)
    model_examples = {model_name: _model_example_rows(all_candidates, model_name) for model_name in models}
    model_ask_contexts = {model_name: _model_ask_context_rows(all_candidates, model_name) for model_name in models}
    surface_inventory = _surface_inventory_rows(surfaces_by_external, surface_summaries)
    summary_counts = _summary_count_rows(surface_inventory)
    processed_last_run = _processed_comment_rows(
        all_candidates,
        gate_model=gate_model,
        limit=_int_env("ACQ_COMMENT_RETRIEVAL_PROCESSED_COMMENTS_XLSX_ROWS", 500, min_value=20),
    )
    rejected_noise_examples = _rejected_noise_rows(
        processed_last_run,
        limit=_int_env("ACQ_COMMENT_RETRIEVAL_REJECTED_NOISE_XLSX_ROWS", 300, min_value=20),
    )
    intent_catalog = _intent_catalog_rows()
    region_catalog = _region_catalog_rows()
    run_id = source_run_id or "unknown_source_run"
    scope_rows = _scope_rows(
        run_id=run_id,
        records=records,
        profiles=profiles,
        models=models,
        gate_model=gate_model,
        scoring_method=scoring_method,
    )
    decision_deltas = _decision_delta_rows(surface_inventory, scope_rows=scope_rows)
    goal_candidate_sheets = {
        sheet_name: _goal_candidate_rows(
            all_candidates,
            actions=set(spec["actions"]),
            source_posts=bool(spec["source_posts"]),
        )
        for sheet_name, spec in _GOAL_SHEET_SPECS.items()
    }
    monitoring_targets = _monitoring_target_rows(surface_summaries)
    surface_backlog = _surface_backlog_rows(surface_inventory)
    dashboard_rows = _dashboard_summary_rows(
        summary={
            "comments_embedded": len(records),
            "surface_profiles_count": len(profiles),
            "models": ", ".join(models),
            "recommended_model": gate_model,
        },
        surface_summaries=surface_summaries,
        scope_rows=scope_rows,
        surface_inventory=surface_inventory,
    )

    artifact_prefix = os.getenv("ACQ_COMMENT_RETRIEVAL_ARTIFACT_PREFIX") or "comment_retrieval"
    candidates_csv = output_path / f"{artifact_prefix}_candidates.csv"
    profiles_csv = output_path / f"{artifact_prefix}_surface_profiles.csv"
    surface_summary_csv = output_path / f"{artifact_prefix}_surface_decision_summary.csv"
    question_patterns_csv = output_path / f"{artifact_prefix}_question_patterns.csv"
    canonical_questions_csv = output_path / f"{artifact_prefix}_canonical_questions.csv"
    distributions_csv = output_path / f"{artifact_prefix}_score_distributions.csv"
    speed_csv = output_path / f"{artifact_prefix}_speed_metrics.csv"
    manual_xlsx = output_path / f"{artifact_prefix}_manual_review_sample.xlsx"
    report_md = output_path / f"{artifact_prefix}_report.md"
    summary_json = output_path / "acq_comment_retrieval_run_summary.json"

    candidate_fields = [
        "run_id", "surface_key", "platform", "surface_type", "relation", "is_post", "author_id", "retrieval", "search_query",
        "context_url", "comment_id", "post_id", "topic_id", "thread_id",
        "created_at", "comment_age_days", "candidate_usage_scope", "is_within_monitoring_window",
        "text_snapshot", "analysis_context_snapshot", "source_post_text_snapshot", "reply_parent_text_snapshot",
        "region_confidence", "region_gate_status", "region_signal_source", "region_evidence_ru", "region_score",
        "region_positive_score", "region_negative_score", "top_region_phrase", "top_negative_region_phrase",
        "model_name", "max_length", "batch_size", "intent_set", "score", "positive_score", "negative_score",
        "scoring_method", "raw_score", "question_boost", "noise_penalty", "top_intent_phrase", "top_intent_score", "rank_global", "rank_within_surface",
        "funnel_bucket", "candidate_action_type", "destination_hint", "transport_hint", "question_signal", "candidate_noise_type",
        "semantic_exclusion_type", "semantic_candidate_rejected", "event_temporal_status", "event_temporal_evidence_ru",
        "event_latest_detected_date", "event_temporal_gate_passed", "intent_text_supported", "pre_llm_candidate_eligible",
        "is_in_llm_gate_queue", "llm_queue_status_ru", "llm_gate_selection_basis", "model_disagreement_bucket",
    ]
    _write_csv(candidates_csv, all_candidates, candidate_fields)
    _write_csv(profiles_csv, [{**p, "semantic_presence": json.dumps(p.get("semantic_presence"), ensure_ascii=False), "dominant_detected_interests": ",".join(p.get("dominant_detected_interests") or [])} for p in profiles], [
        "surface_key", "platform", "surface_type", "surface_title", "surface_url", "members_or_subscribers",
        "period_min_created_at", "period_max_created_at", "period_days", "period_label", "latest_comment_at",
        "latest_event_question_at", "latest_route_recommendation_at", "increment_status", "is_incremental_last_run", "discovered_at", "discovered_in_run_id",
        "days_since_latest_activity", "freshness_status", "analysis_max_comment_age_days", "stale_activity_days", "unique_commenters",
        "region_confidence", "region_gate_status", "region_signal_source", "region_evidence_ru",
        "region_confirmed_contexts", "region_probable_contexts", "region_unknown_contexts", "region_out_of_region_contexts",
        "region_score_max", "region_score_p95", "region_positive_score_max", "region_negative_score_max",
        "comments_total", "comments_embedded", "comment_records", "source_post_records",
        "comments_per_day", "comments_per_week", "comments_per_30d", "comments_per_90d",
        "latest_100_comments", "latest_100_min_created_at", "latest_100_max_created_at", "latest_100_period_days", "latest_100_period_label",
        "eligible_question_candidates", "monitoring_decision_hint", "monitoring_reason", "dominant_detected_interests", "semantic_presence",
    ])
    _write_csv(surface_summary_csv, surface_summaries, [
        "surface_key", "platform", "surface_type", "surface_type_ru", "surface_title", "surface_url", "members_or_subscribers", "selection_status",
        "period_min_created_at", "period_max_created_at", "period_days", "period_label", "unique_commenters", "unique_commenters_note",
        "latest_comment_at", "latest_event_question_at", "latest_route_recommendation_at", "increment_status", "is_incremental_last_run", "discovered_at", "discovered_in_run_id",
        "days_since_latest_activity", "freshness_status", "analysis_max_comment_age_days", "stale_activity_days",
        "region_confidence", "region_gate_status", "region_signal_source", "region_evidence_ru",
        "region_confirmed_contexts", "region_probable_contexts", "region_unknown_contexts", "region_out_of_region_contexts",
        "region_score_max", "region_score_p95", "region_positive_score_max", "region_negative_score_max",
        "comments_total", "comments_embedded", "comments_per_day", "comments_per_week", "comments_per_30d", "comments_per_90d",
        "latest_100_comments", "latest_100_min_created_at", "latest_100_max_created_at", "latest_100_period_days", "latest_100_period_label",
        "recommendation", "summary_ru", "answerable_question_candidates", "answerable_questions_per_30d",
        "answerable_questions_per_90d", "answerable_questions_per_100_comments", "ask_clarification_contexts",
        "ask_contexts_per_30d", "eligible_comment_contexts", "source_post_contexts",
        "route_poi_questions", "event_questions", "event_site_search_questions", "organizer_submission_questions", "badge_filter_questions",
        "filtered_noise_contexts", "relation_counts_json", "dominant_detected_interests",
        "monitoring_decision_hint", "monitoring_reason", "answerable_examples", "ask_question_examples",
    ])
    _write_csv(question_patterns_csv, question_patterns, [
        "surface_key", "platform", "surface_type", "pattern", "candidate_action_type", "canonical_question_ru", "count", "intent_sets", "models", "example_questions", "example_urls",
    ])
    _write_csv(canonical_questions_csv, canonical_questions, [
        "pattern", "candidate_action_type", "canonical_question_ru", "real_question_examples_total",
        "monitoring_candidate_examples", "historical_calibration_examples", "surfaces_count",
        "example_questions", "example_urls", "source_note_ru",
    ])
    _write_csv(distributions_csv, distributions, ["model_name", "scoring_method", "count", "p50", "p90", "p95", "p99", "max"])
    _write_csv(speed_csv, speed_rows, ["model_name", "device", "batch_size", "max_length", "comments_total", "intent_embedding_sec", "comment_embedding_sec", "region_embedding_sec", "scoring_sec", "total_sec", "comments_per_sec", "comments_per_hour", "peak_ram_mb", "cpu"])
    manual_rows = _select_manual_rows(all_candidates, max_rows=max_manual_rows)
    _write_xlsx(
        manual_xlsx,
        manual_rows,
        surface_summaries=surface_summaries,
        model_examples=model_examples,
        model_ask_contexts=model_ask_contexts,
        question_patterns=question_patterns,
        canonical_questions=canonical_questions,
        scope_rows=scope_rows,
        surface_inventory=surface_inventory,
        summary_counts=summary_counts,
        dashboard_rows=dashboard_rows,
        intent_catalog_rows=intent_catalog,
        region_catalog_rows=region_catalog,
        decision_delta_rows=decision_deltas,
        processed_comment_rows=processed_last_run,
        rejected_noise_rows=rejected_noise_examples,
        goal_candidate_rows=goal_candidate_sheets,
        monitoring_target_rows=monitoring_targets,
        surface_backlog_rows=surface_backlog,
    )

    summary = {
        "stage": STAGE_NAME,
        "run_id": run_id,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "models": models,
        "recommended_model": gate_model,
        "scoring_method": scoring_method,
        "region_gate_required": True,
        "region_gate_confidence_counts": dict(Counter(str(row.get("region_confidence") or "unknown") for row in surface_summaries)),
        "comments_total": len(comment_records),
        "comments_after_filter": len(records),
        "comments_embedded": len(records),
        "surface_profiles_count": len(profiles),
        "candidate_count": len(all_candidates),
        "llm_gate_candidate_count": len(gate_candidates),
        "estimated_llm_reduction_vs_all_comments": round(1.0 - (len(gate_candidates) / max(1, len(records))), 6),
        "speed_metrics": speed_rows,
        "score_distributions": distributions,
        "artifacts": {
            "summary_json": str(summary_json),
            "candidates_csv": str(candidates_csv),
            "surface_profiles_csv": str(profiles_csv),
            "surface_decision_summary_csv": str(surface_summary_csv),
            "question_patterns_csv": str(question_patterns_csv),
            "canonical_questions_csv": str(canonical_questions_csv),
            "score_distributions_csv": str(distributions_csv),
            "speed_metrics_csv": str(speed_csv),
            "manual_review_xlsx": str(manual_xlsx),
            "report_md": str(report_md),
        },
        "total_sec": round(time.perf_counter() - started, 4),
    }
    report_md.write_text(_report_md(summary, profiles, gate_candidates, speed_rows, surface_summaries), encoding="utf-8")
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    progress_callback("complete", {"comments_embedded": len(records), "candidate_count": len(all_candidates), "progress_percent": 100})
    return {
        "summary": summary,
        "surface_profiles": profiles,
        "surface_decision_summaries": surface_summaries,
        "canonical_questions": canonical_questions,
        "candidates": all_candidates,
        "llm_gate_candidates": gate_candidates,
        "artifacts": summary["artifacts"],
    }
