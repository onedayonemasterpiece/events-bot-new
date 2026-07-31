#!/usr/bin/env python3
from __future__ import annotations
import argparse, json, os, shutil, tempfile, time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in os.sys.path: os.sys.path.insert(0, str(PROJECT_ROOT))
import importlib.util
_candidate_executor_path = PROJECT_ROOT / "kaggle" / "execute_region_talk_candidate_report.py"
_spec = importlib.util.spec_from_file_location("region_talk_candidate_executor", _candidate_executor_path)
if _spec is None or _spec.loader is None:
    raise RuntimeError(f"cannot load {_candidate_executor_path}")
_candidate_executor = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_candidate_executor)
DirectKaggleClient = _candidate_executor.DirectKaggleClient
KaggleClient = _candidate_executor.KaggleClient
load_env_file = _candidate_executor.load_env_file
create_or_replace_dataset = _candidate_executor.create_or_replace_dataset
wait_dataset_ready = _candidate_executor.wait_dataset_ready
build_input_datasets = _candidate_executor.build_input_datasets
poll_kernel = _candidate_executor.poll_kernel
preflight_ydb_access = _candidate_executor.preflight_ydb_access
assert_region_talk_kaggle_slots_free = _candidate_executor.assert_region_talk_kaggle_slots_free
cleanup_stale_region_talk_input_datasets = _candidate_executor.cleanup_stale_region_talk_input_datasets

KERNEL_PATH = PROJECT_ROOT / "kaggle" / "RegionTalkImageDiagnostic"
OUT_ROOT = PROJECT_ROOT / "artifacts" / "codex" / "kaggle" / "region-talk-image-diagnostic"
DEFAULT_CLIP_KAGGLE_MODEL_SOURCE = "yujkaggle/openaiclip-vit-base-patch32/PyTorch/default/1"

def _rows_from_xlsx(path: Path, top_n: int) -> tuple[list[dict[str, Any]], int]:
    from openpyxl import load_workbook

    wb=load_workbook(path, read_only=True, data_only=True); ws=wb["09a_image_candidate_queue"]
    it=ws.iter_rows(values_only=True); headers=[str(v) if v is not None else "" for v in next(it)]
    rows=[]
    for vals in it:
        d={headers[i] if i<len(headers) and headers[i] else f"col{i}": vals[i] if i<len(vals) else None for i in range(max(len(headers),len(vals)))}
        if any(v not in (None,"") for v in vals) and not d.get("_sheet_note"): rows.append(d)
    wb.close()
    def pri(d):
        selected=str(d.get("selected_for_next_image_batch") or "").lower() in {"1","true","yes"}
        actual=d.get("image_model_input_type")=="actual_image" or str(d.get("media_acquisition_status") or "").startswith("actual_image")
        return (0 if selected else 1, 0 if actual else 1, int(d.get("image_queue_order") or 10**9))
    out=[]
    for d in sorted(rows,key=pri)[:top_n]:
        out.append({
            "image_queue_id":d.get("image_queue_id") or f"imgq_{len(out)+1}", "post_url":d.get("post_url") or "", "source_url":d.get("source_url") or "", "source_title":d.get("source_title") or "",
            "image_url_or_local_path":d.get("image_url_or_local_path") or "", "media_ref":d.get("platform_post_key") or "", "image_queue_status":d.get("image_queue_status") or "",
            "media_acquisition_status":d.get("media_acquisition_status") or "", "image_queue_order":d.get("image_queue_order"), "source_post_id":d.get("post_id") or "", "post_date":d.get("post_date") or "",
            "candidate_score":d.get("candidate_score"), "previous_overall_media_score":d.get("overall_media_score"), "previous_postcardness_score":d.get("postcardness_score"),
            "previous_aesthetic_score":d.get("aesthetic_score"), "previous_technical_quality_score":d.get("technical_quality_score"), "previous_publication_safety_score":d.get("publication_safety_score"),
            "previous_image_model_input_type":d.get("image_model_input_type"), "previous_image_model_type":d.get("image_model_type"), "text_excerpt":"",
        })
    return out, len(rows)

def latest_queue_xlsx() -> Path:
    candidates=[]
    for root in [PROJECT_ROOT / "artifacts" / "codex" / "kaggle" / "region-talk-candidate-report", PROJECT_ROOT / "artifacts" / "region-talk"]:
        if root.exists(): candidates.extend(root.rglob("*.xlsx"))
    candidates=[p for p in candidates if p.is_file() and ("z10" in str(p) or p.name=="candidates-latest.xlsx")]
    if not candidates: raise FileNotFoundError("No Region Talk queue workbook found")
    return sorted(candidates, key=lambda p:p.stat().st_mtime, reverse=True)[0]

def stage_kernel(run_id: str, kernel_slug: str) -> Path:
    tmp=Path(tempfile.mkdtemp(prefix="region-talk-image-diag-kernel-")); dst=tmp/KERNEL_PATH.name
    shutil.copytree(KERNEL_PATH,dst,ignore=shutil.ignore_patterns("__pycache__","*.pyc"))
    meta=json.loads((dst/"kernel-metadata.json").read_text())
    username=(os.getenv("KAGGLE_USERNAME") or "").strip()
    if username: meta["id"]=f"{username}/{kernel_slug}"
    meta["slug"]=kernel_slug; meta["title"]="Region Talk Image Diagnostic"; meta["enable_gpu"]=False; meta["enable_internet"]=True
    model_source=str(os.getenv("REGION_TALK_CLIP_KAGGLE_MODEL_SOURCE") or DEFAULT_CLIP_KAGGLE_MODEL_SOURCE).strip()
    model_sources=[str(item).strip() for item in (meta.get("model_sources") or []) if str(item).strip()]
    if model_source and model_source not in model_sources: model_sources.append(model_source)
    meta["model_sources"]=model_sources
    (dst/"kernel-metadata.json").write_text(json.dumps(meta,ensure_ascii=False,indent=2),encoding="utf-8")
    return dst

def main() -> int:
    ap=argparse.ArgumentParser(); ap.add_argument("--env-file",type=Path,default=PROJECT_ROOT/".env"); ap.add_argument("--run-id",default=""); ap.add_argument("--queue-xlsx",type=Path,default=None); ap.add_argument("--source",choices=["ydb","xlsx"],default="ydb"); ap.add_argument("--top-n",type=int,default=50); ap.add_argument("--max-items-per-run",type=int,default=50); ap.add_argument("--batch-size",type=int,default=30); ap.add_argument("--wait-initial-seconds",type=int,default=600); ap.add_argument("--wait-after-drain-seconds",type=int,default=600); ap.add_argument("--image-poll-interval-seconds",type=int,default=60); ap.add_argument("--timeout-minutes",type=int,default=60); ap.add_argument("--poll-interval-seconds",type=int,default=20); ap.add_argument("--kernel-slug",default="region-talk-image-diagnostic"); ap.add_argument("--keep-input-datasets",action="store_true"); ap.add_argument("--allow-active-region-talk-kernel",action="store_true"); ap.add_argument("--no-wait", action="store_true")
    args=ap.parse_args(); load_env_file(args.env_file)
    run_id=args.run_id or "region-talk-image-diagnostic-"+time.strftime("%Y%m%dT%H%M%SZ",time.gmtime())
    os.environ["REGION_TALK_RUN_ID"]=run_id; os.environ.setdefault("REGION_TALK_DRY_RUN","1"); os.environ.setdefault("REGION_TALK_DISABLE_PUBLISH","1"); os.environ.setdefault("REGION_TALK_AUTH_BUNDLE_ENV","TELEGRAM_AUTH_BUNDLE_DISCOVERY2")
    os.environ["REGION_TALK_IMAGE_DIAG_SOURCE"]=args.source
    os.environ.setdefault("REGION_TALK_IMAGE_DIAG_QUEUE_SCAN_LIMIT", "5000")
    os.environ.setdefault("REGION_TALK_IMAGE_DIAG_PUBLIC_TG_HTML_FALLBACK", "0")
    os.environ.setdefault("REGION_TALK_STATE_BACKEND", "ydb" if args.source == "ydb" else os.environ.get("REGION_TALK_STATE_BACKEND", "local"))
    qxlsx=None; rows=[]; total=0
    if args.source == "xlsx":
        qxlsx=args.queue_xlsx or latest_queue_xlsx(); rows,total=_rows_from_xlsx(qxlsx,args.top_n)
        if not rows: print("image queue empty, diagnostic image scoring skipped", flush=True); return 0
    client=KaggleClient() if KaggleClient is not None else DirectKaggleClient(); username=(os.getenv("KAGGLE_USERNAME") or "").strip()
    if not username: raise RuntimeError("KAGGLE_USERNAME is required")
    cleanup_stale_region_talk_input_datasets(client)
    if args.source == "ydb":
        os.environ.setdefault("REGION_TALK_REQUIRE_YDB_STATE", "1")
        preflight_ydb_access()
    def write_input(folder: Path):
        (folder/"image_diag_input.json").write_text(json.dumps({"run_id":run_id,"source":args.source,"top_n":args.top_n,"max_items_per_run":args.max_items_per_run,"batch_size":args.batch_size,"wait_initial_seconds":args.wait_initial_seconds,"wait_after_drain_seconds":args.wait_after_drain_seconds,"poll_interval_seconds":args.image_poll_interval_seconds,"queue_xlsx":str(qxlsx or ""),"queue_rows_total":total,"rows":rows},ensure_ascii=False,indent=2),encoding="utf-8")
        safe_env_keys=[
            "REGION_TALK_RUN_ID", "REGION_TALK_IMAGE_DIAG_SOURCE", "REGION_TALK_IMAGE_DIAG_TOP_N",
            "REGION_TALK_IMAGE_DIAG_MAX_ITEMS_PER_RUN", "REGION_TALK_IMAGE_DIAG_BATCH_SIZE",
            "REGION_TALK_IMAGE_DIAG_WAIT_INITIAL_SECONDS", "REGION_TALK_IMAGE_DIAG_WAIT_AFTER_DRAIN_SECONDS",
            "REGION_TALK_IMAGE_DIAG_POLL_INTERVAL_SECONDS",
            "REGION_TALK_IMAGE_DIAG_QUEUE_SCAN_LIMIT",
            "REGION_TALK_IMAGE_DIAG_PUBLIC_TG_HTML_FALLBACK",
            "REGION_TALK_IMAGE_DIAG_STALE_LEASE_SECONDS",
            "REGION_TALK_IMAGE_MAX_IMAGES_PER_POST",
            "REGION_TALK_IMAGE_VLM_ENABLED", "REGION_TALK_IMAGE_VLM_MAX_CALLS_PER_RUN",
            "REGION_TALK_IMAGE_VLM_MODEL", "REGION_TALK_IMAGE_VLM_DEFAULT_ENV_VAR_NAME",
            "REGION_TALK_IMAGE_VLM_TIMEOUT_SECONDS", "REGION_TALK_IMAGE_VLM_MAX_SIDE",
            "REGION_TALK_LLM_MODEL", "REGION_TALK_LLM_DEFAULT_ENV_VAR_NAME",
            "REGION_TALK_LLM_BUDGET_ID", "REGION_TALK_LLM_BUDGET_MAX",
            "REGION_TALK_CLIP_MODEL_LOCAL_PATH", "REGION_TALK_CLIP_REQUIRE_LOCAL_MODEL",
            "REGION_TALK_CLIP_KAGGLE_MODEL_SOURCE",
            "REGION_TALK_STATE_BACKEND", "REGION_TALK_REQUIRE_YDB_STATE",
            "REGION_TALK_REQUIRE_NONINTERACTIVE_YDB_CREDENTIAL", "REGION_TALK_ALLOW_KAGGLE_YDB_SECRET",
            "REGION_TALK_KAGGLE_SECRET_NAMES",
            "REGION_TALK_YDB_ENDPOINT", "REGION_TALK_YDB_DATABASE", "REGION_TALK_YDB_NAMESPACE",
            "REGION_TALK_YDB_MAX_SOURCE_ROWS", "REGION_TALK_YDB_MAX_CANDIDATE_ROWS",
            "REGION_TALK_SOURCE_IMAGE_MIN_ACTUAL_SCORED", "REGION_TALK_SOURCE_IMAGE_MIN_AVG_SCORE",
            "REGION_TALK_PUBLICATION_ELIGIBILITY_GATE_VERSION",
            "REGION_TALK_IMAGE_DIAG_EXPECTED_ELIGIBILITY_GATE_VERSION",
            "REGION_TALK_AUTH_BUNDLE_ENV",
        ]
        runtime_env={k:os.environ.get(k, "") for k in safe_env_keys if os.environ.get(k, "") != ""}
        runtime_env["REGION_TALK_IMAGE_DIAG_TOP_N"]=str(args.top_n)
        runtime_env["REGION_TALK_IMAGE_DIAG_MAX_ITEMS_PER_RUN"]=str(args.max_items_per_run)
        runtime_env["REGION_TALK_IMAGE_DIAG_BATCH_SIZE"]=str(args.batch_size)
        runtime_env["REGION_TALK_IMAGE_DIAG_WAIT_INITIAL_SECONDS"]=str(args.wait_initial_seconds)
        runtime_env["REGION_TALK_IMAGE_DIAG_WAIT_AFTER_DRAIN_SECONDS"]=str(args.wait_after_drain_seconds)
        runtime_env["REGION_TALK_IMAGE_DIAG_POLL_INTERVAL_SECONDS"]=str(args.image_poll_interval_seconds)
        (folder/"region_talk_run_config.json").write_text(json.dumps({"run_id":run_id,"source":args.source,"env":runtime_env},ensure_ascii=False,indent=2),encoding="utf-8")
        shutil.copy2(PROJECT_ROOT / "region_talk_llm_runtime.py", folder / "region_talk_llm_runtime.py")
    diag_ref=create_or_replace_dataset(client, username, f"rt-img-diag-{run_id[-18:].lower().replace(':','').replace('_','-')}", "RT image diag input", write_input)
    wait_dataset_ready(client, diag_ref, expected_files=["image_diag_input.json", "region_talk_run_config.json", "region_talk_llm_runtime.py"])
    dataset_sources=build_input_datasets(client, run_id=run_id, username=username)+[diag_ref]
    kernel_path=stage_kernel(run_id,args.kernel_slug); kernel_ref=f"{username}/{args.kernel_slug}"
    candidate_kernel_ref=f"{username}/region-talk-candidate-report"
    assert_region_talk_kaggle_slots_free(
        client,
        [kernel_ref],
        optional_kernel_refs=[candidate_kernel_ref],
        optional_kernel_auth_bundle_envs={candidate_kernel_ref: "TELEGRAM_AUTH_BUNDLE_DISCOVERY1"},
        allow_active=bool(args.allow_active_region_talk_kernel),
        auth_bundle_env=os.environ.get("REGION_TALK_AUTH_BUNDLE_ENV", "TELEGRAM_AUTH_BUNDLE_DISCOVERY2"),
    )
    print(f"[region-talk-image-diagnostic] pushing {kernel_ref} run_id={run_id} source={args.source} queue_rows={total} top_n={args.top_n if args.source=='ydb' else len(rows)} max_items={args.max_items_per_run} batch_size={args.batch_size}", flush=True)
    client.push_kernel(kernel_path=kernel_path, dataset_sources=dataset_sources)
    if args.no_wait:
        print(f"[region-talk-image-diagnostic] pushed {kernel_ref}; not waiting; input datasets retained: {dataset_sources}", flush=True)
        return 0
    completed=False
    try:
        poll_kernel(client,kernel_ref,timeout_minutes=args.timeout_minutes,poll_interval_seconds=args.poll_interval_seconds)
        completed=True
        out_dir=OUT_ROOT/run_id; out_dir.mkdir(parents=True,exist_ok=True)
        files=client.download_kernel_output(kernel_ref,path=out_dir,force=True,file_pattern=r".*(xlsx|html|md|json|jsonl|jpg|png)$")
        print(f"[region-talk-image-diagnostic] downloaded {len(files)} files to {out_dir}", flush=True)
    finally:
        if not args.keep_input_datasets and completed:
            for ref in dataset_sources:
                try: client.delete_dataset(ref); print(f"[region-talk-image-diagnostic] cleaned input dataset {ref}", flush=True)
                except Exception as exc: print(f"[region-talk-image-diagnostic] WARNING cleanup failed {ref}: {type(exc).__name__}: {exc}", flush=True)
        elif not completed:
            print(f"[region-talk-image-diagnostic] keeping input datasets because run did not complete locally: {dataset_sources}", flush=True)
    return 0
if __name__=="__main__": raise SystemExit(main())
