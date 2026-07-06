from __future__ import annotations
import asyncio, base64, csv, datetime as dt, gzip, hashlib, html, json, math, os, random, re, subprocess, sys, tarfile, time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

RUN_SCHEMA_VERSION = "event-comment-feedback-kaggle-dual-prototype-v2"
PHRASE_BANK_VERSION = "event-comment-feedback-phrase-bank-v1-json"
READ_MODE = "api_read_paced_v1"
REQUIRED_MODELS = ["intfloat/multilingual-e5-base", "BAAI/bge-m3"]
GATE_MODEL = "intfloat/multilingual-e5-base"
SCRIPT_DIR = Path(globals().get('__file__', Path.cwd() / 'event_comment_feedback_discovery.py')).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

WORK = Path('/kaggle/working') if Path('/kaggle/working').exists() else Path.cwd()
STATUS_PATH = WORK / 'event_comment_feedback_status.jsonl'
STATUS_CLIENT = None
STATUS_RESOURCES: list[str] = []
STATUS_PROGRESS: dict[str, Any] = {
    'phase': 'bootstrap',
    'progress_percent': 0,
    'progress_label': 'подготовка',
}

EVENT_PHASES = {
    'kernel_started': 'preflight',
    'pip_install_start': 'preflight',
    'pip_install_done': 'preflight',
    'secrets_loaded': 'preflight',
    'preflight_ok': 'preflight',
    'fetch_tg_progress': 'fetch',
    'fetch_vk_progress': 'fetch',
    'fetch_done': 'fetch',
    'fetch_error_summary_written': 'fetch',
    'model_load_start': 'embed',
    'model_encode_start': 'embed',
    'model_encode_done': 'embed',
    'report_written': 'report',
    'kernel_done': 'report',
    'kernel_failed': 'failed',
}

def _find_status_client_loader():
    try:
        from kaggle_status_client import load_status_client  # type: ignore
        return load_status_client
    except Exception:
        pass
    for root in [SCRIPT_DIR, Path.cwd(), WORK, Path('/kaggle/input')]:
        if not root.exists():
            continue
        try:
            matches = sorted(root.rglob('kaggle_status_client.py'))
        except Exception:
            matches = []
        for path in matches:
            try:
                import importlib.util
                spec = importlib.util.spec_from_file_location('kaggle_status_client_dynamic', path)
                if not spec or not spec.loader:
                    continue
                mod = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(mod)
                return getattr(mod, 'load_status_client', None)
            except Exception:
                continue
    return None

def init_status() -> None:
    global STATUS_CLIENT
    loader = _find_status_client_loader()
    if loader is None:
        return
    try:
        STATUS_CLIENT = loader(output_dir=WORK, log=lambda message: print(message, flush=True))
    except Exception as exc:
        print(f'[event-comment-feedback] status init failed: {type(exc).__name__}: {exc}', flush=True)
        STATUS_CLIENT = None

def _status_enabled() -> bool:
    return bool(STATUS_CLIENT is not None and getattr(STATUS_CLIENT, 'enabled', False))

def _status_event(event: str, *, payload: dict[str, Any], status: str | None = None, message: str | None = None) -> None:
    if not _status_enabled():
        return
    phase = str(payload.get('phase') or EVENT_PHASES.get(event) or event)
    for key in ('progress_percent', 'progress_label', 'done', 'total', 'events', 'source_posts', 'selected', 'tg', 'vk', 'comments', 'errors', 'model', 'count'):
        if key in payload:
            STATUS_PROGRESS[key] = payload[key]
    STATUS_PROGRESS['phase'] = phase
    try:
        STATUS_CLIENT.event(
            event,
            phase=phase,
            status=status or ('done' if event in {'report_written', 'kernel_done'} else 'failed' if event == 'kernel_failed' else 'running'),
            progress=dict(STATUS_PROGRESS),
            message=message,
        )
    except Exception as exc:
        print(f'[event-comment-feedback] status event failed {event}: {type(exc).__name__}: {exc}', flush=True)

def acquire_status_resources() -> None:
    if not _status_enabled():
        return
    for key in STATUS_CLIENT.config.get('resource_leases') or []:
        if not STATUS_CLIENT.acquire_resource(str(key), ttl_seconds=3 * 60 * 60):
            raise RuntimeError(f'Required Kaggle resource is busy: {key}')
        STATUS_RESOURCES.append(str(key))
    STATUS_CLIENT.start_alive(interval_seconds=60, progress_provider=lambda: dict(STATUS_PROGRESS))

def finish_status() -> None:
    if STATUS_CLIENT is None:
        return
    for key in list(STATUS_RESOURCES):
        try:
            STATUS_CLIENT.release_resource(key)
            STATUS_RESOURCES.remove(key)
        except Exception as exc:
            print(f'[event-comment-feedback] status resource release failed {key}: {exc}', flush=True)
    try:
        STATUS_CLIENT.stop_alive()
    except Exception:
        pass

def emit(event: str, **payload: Any) -> None:
    row = {"ts": dt.datetime.now(dt.timezone.utc).isoformat(), "event": event, **payload}
    print(json.dumps(row, ensure_ascii=False), flush=True)
    with STATUS_PATH.open('a', encoding='utf-8') as f:
        f.write(json.dumps(row, ensure_ascii=False)+'\n')
    _status_event(event, payload=payload, message=payload.get('message') if isinstance(payload.get('message'), str) else None)


def human_delay(base_seconds: float, *, index: int = 0, platform: str = '') -> float:
    base = max(0.2, float(base_seconds or 0.45))
    # API paced read: non-uniform pauses and longer breath between batches.
    delay = random.uniform(base * 0.75, base * 1.85)
    if index and index % 17 == 0:
        delay += random.uniform(2.0, 5.0)
    if index and index % 53 == 0:
        delay += random.uniform(8.0, 18.0)
    return min(delay, 35.0)

async def async_human_pause(base_seconds: float, *, index: int = 0, platform: str = '') -> None:
    await asyncio.sleep(human_delay(base_seconds, index=index, platform=platform))

def sync_human_pause(base_seconds: float, *, index: int = 0, platform: str = '') -> None:
    time.sleep(human_delay(base_seconds, index=index, platform=platform))

def ensure_deps() -> None:
    missing=[]
    for mod,pkg in [('telethon','telethon'),('openpyxl','openpyxl'),('sentence_transformers','sentence-transformers'),('cryptography','cryptography'),('requests','requests')]:
        try: __import__(mod)
        except Exception: missing.append(pkg)
    if missing:
        emit('pip_install_start', packages=missing)
        subprocess.check_call([sys.executable,'-m','pip','install','-q',*missing])
        emit('pip_install_done', packages=missing)

PAYLOAD_EXTRACTED = False

def extract_payload_tarball_once() -> None:
    global PAYLOAD_EXTRACTED
    if PAYLOAD_EXTRACTED:
        return
    PAYLOAD_EXTRACTED = True
    for root in [Path('/kaggle/input'), Path.cwd(), WORK]:
        if not root.exists():
            continue
        for archive in root.rglob('event_comment_feedback_payload.tarball'):
            target = WORK / 'event_comment_feedback_payload'
            target.mkdir(parents=True, exist_ok=True)
            with tarfile.open(archive, 'r:*') as tf:
                tf.extractall(target)
            emit('payload_extracted', archive=str(archive), target=str(target), progress_percent=3, progress_label='payload распакован')
            return

def find_input_file(name: str) -> Path:
    for attempt in range(2):
        for root in [Path('/kaggle/input'), Path.cwd(), WORK]:
            if not root.exists(): continue
            for p in root.rglob(name): return p
        if attempt == 0:
            extract_payload_tarball_once()
    raise FileNotFoundError(name)

def load_json_any(path: Path) -> Any:
    data = path.read_bytes()
    if path.suffix == '.gz': data = gzip.decompress(data)
    return json.loads(data.decode('utf-8'))

def load_secrets() -> dict[str, Any]:
    from cryptography.fernet import Fernet
    enc = find_input_file('secrets.enc').read_bytes(); key = find_input_file('fernet.key').read_bytes().strip()
    secrets = json.loads(Fernet(key).decrypt(enc).decode('utf-8'))
    for k,v in secrets.items():
        if v is not None and str(v).strip(): os.environ.setdefault(k, str(v))
    emit('secrets_loaded', keys=sorted(secrets.keys()))
    return secrets

def decode_bundle(name: str):
    raw=(os.getenv(name) or '').strip()
    if not raw: raise RuntimeError(f'{name} missing')
    b=json.loads(base64.urlsafe_b64decode(raw.encode('ascii')).decode('utf-8'))
    kwargs={}
    for key in ('device_model','system_version','app_version','lang_code','system_lang_code'):
        if b.get(key): kwargs[key]=str(b[key])
    return b.get('session') or '', kwargs

def normalize_text(text: str) -> str:
    text=html.unescape(str(text or ''))
    text=re.sub(r'https?://\S+',' ',text); text=re.sub(r'[\uFE0F\u200D]','',text)
    return re.sub(r'\s+',' ',text).strip()
def low(text: str) -> str: return normalize_text(text).lower().replace('ё','е')
def short_hash(value: str, n:int=16) -> str: return hashlib.sha256(value.encode('utf-8', errors='ignore')).hexdigest()[:n]
def is_link_or_emoji_only(text: str) -> bool: return len(re.findall(r'[A-Za-zА-Яа-яЁё]', re.sub(r'https?://\S+','',text or ''))) < 2
def is_probable_source_copy(text: str) -> bool:
    t=low(text)
    if len(t)>760: return True
    formal=sum(1 for w in ['приглашаем','состоится','выставка','концерт','программа','регистрация','билеты','подробнее','начало','стоимость','адрес','организатор'] if w in t)
    user=any(w in t for w in ['жду','ждём','ждем','пойду','идём','идем','хочу','ура','класс','спасибо','?','жаль','дорого','подскажите','можно','интересно','люблю'])
    return len(t)>320 and formal>=3 and not user


def strip_vk_mention_prefix(text: str) -> str:
    # VK replies often start with an addressed mention, e.g.
    # [id123|Имя], здравствуйте! ...  Keep the content, remove only the address.
    return re.sub(r'^\s*\[(?:id|club|public)\d+\|[^\]]+\]\s*,?\s*', '', normalize_text(text), flags=re.I)

def classify_comment_role(text: str) -> tuple[str, str | None]:
    raw = normalize_text(text)
    t = low(raw)
    stripped = strip_vk_mention_prefix(raw)
    st = low(stripped)
    letters=len(re.findall(r'[a-zа-яё]', t, re.I))
    if not t or letters < 2: return 'empty_or_emoji', 'too_few_letters'
    if is_probable_private_ticket_request(t): return 'ticket_resale_or_private_ticket_request', 'private_ticket_request'
    if re.search(r'(?i)(подработк|заработ|оплат[аы]\s+от|\b[0-9]{3,6}\s*(?:₽|руб|р\b).*день|график\s+[0-9-]+\s*час)', t):
        return 'job_spam_or_earnings_ad', 'job_or_earnings_ad'
    if re.search(r'(?i)(участвую\s+в\s+розыгрыш|розыгрыш|хочу\s+выиграть|победител|отмеч[ау].*друз|репост)', t):
        return 'giveaway_participation', 'giveaway_or_contest'
    if len(t) <= 48 and re.fullmatch(r'(?i)(такой\s+же\s+вопрос|и\s+мне|мне\s+тоже|а\s+где\??|подскажите\s+тоже|спасибо|благодарю|благодарю\s+за\s+подсказку.*|\+1|тоже\s+интересно)', t):
        return 'contextless_short_reply', 'needs_parent_context'
    if len(t) <= 80 and re.fullmatch(r'[«»"\'\sа-яa-z0-9:;,.!?—–-]+', t, re.I) and len(re.findall(r'[?!]', t)) == 0 and len(t.split()) <= 8 and re.search(r'[«"]', t):
        return 'title_only_or_entity_only', 'title_or_entity_without_predicate'
    if re.search(r'(?i)(администрац|дорог[аиу]|тротуар|грунт|ремонтир|ремонт\s+дорог)', t) and not re.search(r'(?i)(парковк|где\s+парков|оставить\s+машин|как\s+добраться|общественн\w*\s+транспорт|автобус)', t):
        return 'municipal_road_complaint_offtopic', 'road_or_municipal_complaint'
    if re.search(r'(?i)(афиш|постер|анонс|картинк|фото\s+анонс)', t) and not re.search(r'(?i)(постановк|спектакл|сцен|мюзикл|выставк|представлен|концерт)', t):
        return 'poster_or_announcement_reaction', 'poster_not_event_experience'
    if re.search(r'(?i)(питух|бойцовск\w*\s+птиц|ахах|лол\b|кек\b|мем)', t):
        return 'offtopic_meme_or_noise', 'meme_or_noise'
    official_greeting = re.search(r'(?i)^(здравствуйте|добрый\s+день|добрый\s+вечер)[!.:,\s]+', st)
    user_question_like = has_direct_question_marker(st) or re.search(r'(?i)\b(как\s+можно|где|подскажите|можно\s+ли|будет\s+ли|есть\s+ли|когда|во\s+сколько|сколько)\b', st)
    explicit_official_marker = re.search(r'(?i)(держим\s+вопрос\s+на\s+контроле|информация\s+будет\s+опубликована|информация\s+размещена|обратитесь\s+по\s+ссылке|подробности\s+у\s+организатора|регистрация\s+закрыта|запись\s+закрыта|приносим\s+извинения|благодарим\s+за\s+обращение|следите(?:,?\s+пожалуйста,?)?\s+за\s+новостями|можете\s+ознакомиться)', st)
    operational_official_context = re.search(r'(?i)(при\s+прохождении\s+входного\s+контроля\s+сотрудники|территори[яи]\s+стадиона|предусмотрены\s+точки\s+питания|вход[^.!?]{0,80}запрещен|будут\s+опубликованы\s+еще\s+билеты|в\s+последнем\s+посте[^.!?]{0,80}ссылк)', st)
    if ((official_greeting and explicit_official_marker) or operational_official_context) and not has_explicit_problem_report(st):
        return 'official_reply', 'official_or_admin_answer'
    if is_probable_source_copy(t): return 'source_copy_or_announcement', 'source_like_text'
    return 'user_feedback', None

def is_probable_private_ticket_request(text: str) -> bool:
    t=low(text)
    return bool(RESALE_RE.search(t)) if 'RESALE_RE' in globals() else bool(re.search(r'(?i)(приму\s+в\s+дар|дайте\s+.*билет|проходк|лишн\w*\s+билет|у\s+кого.*билет)', t))

def source_comment_cap(post: dict[str, Any], global_cap: int) -> int:
    ceiling = max(1, int(global_cap or 300))
    metric = int(post.get('metric_comments') or 0)
    if metric <= 0:
        wanted = min(60, ceiling)
    elif metric <= 20:
        wanted = 30
    elif metric <= 100:
        wanted = 100
    elif metric <= 500:
        wanted = 200
    else:
        wanted = 300
    return max(1, min(ceiling, wanted))

def redact_error_example(err: dict[str, Any]) -> dict[str, Any]:
    allow = {'platform','platform_post_key','status','error_type','code','message','seconds','cap','fetched','total_available','skip_reason'}
    return {k: v for k, v in err.items() if k in allow and v is not None}

def write_fetch_error_summary(errors: list[dict[str, Any]], skipped: list[dict[str, Any]], source_diagnostics: list[dict[str, Any]]) -> dict[str, Any]:
    buckets: dict[str, dict[str, Any]] = {}
    def add(kind: str, item: dict[str, Any]) -> None:
        key = '|'.join([kind, str(item.get('platform') or ''), str(item.get('status') or item.get('skip_reason') or ''), str(item.get('code') or item.get('error_type') or '')])
        b = buckets.setdefault(key, {'kind': kind, 'platform': item.get('platform'), 'status': item.get('status') or item.get('skip_reason'), 'code': item.get('code'), 'error_type': item.get('error_type'), 'count': 0, 'examples': []})
        b['count'] += 1
        if len(b['examples']) < 5:
            b['examples'].append(redact_error_example(item))
    for err in errors:
        add('fetch_error', err)
    for sk in skipped:
        add('source_skipped', sk)
    summary = {
        'schema_version': 'event-comment-feedback-fetch-error-summary-v1',
        'generated_at': dt.datetime.now(dt.timezone.utc).isoformat(),
        'errors_total': len(errors),
        'skipped_total': len(skipped),
        'source_diagnostics_total': len(source_diagnostics),
        'buckets': sorted(buckets.values(), key=lambda b: (-int(b['count']), str(b.get('platform') or ''), str(b.get('status') or ''))),
        'source_diagnostics_samples': source_diagnostics[:200],
    }
    (WORK/'fetch_error_summary.json').write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    emit('fetch_error_summary_written', errors=len(errors), skipped=len(skipped), buckets=len(summary['buckets']), progress_label=f"ошибки источников: {len(errors)} / skipped {len(skipped)}")
    return summary


def derive_comments_capability(item: dict[str, Any]) -> tuple[str, str | None]:
    status=str(item.get('status') or item.get('skip_reason') or '')
    code=str(item.get('code') or item.get('error_type') or '')
    if status in {'capability_no_comments', 'no_comments'}: return 'no_comments', 'ttl_7d'
    if status in {'capability_no_discussion_or_deleted', 'no_discussion_or_deleted'}: return 'no_discussion_or_deleted', 'ttl_14d'
    if status in {'capability_entity_resolution_failed', 'telegram_missing_username_or_chat_id','vk_missing_owner_or_post_id'}: return 'entity_resolution_failed', 'ttl_30d'
    if status in {'capability_forbidden_or_deleted', 'not_accessible_or_deleted'}: return 'forbidden_or_deleted', 'ttl_30d'
    if status in {'capability_rate_limited', 'flood_wait'}: return 'rate_limited', 'ttl_floodwait'
    if code in {'MsgIdInvalidError','BadRequestError'}: return 'no_discussion_or_deleted', 'ttl_14d'
    if code == 'ValueError' or re.search(r'(?i)(no\s+username|could\s+not\s+find\s+entity|cannot\s+find\s+any\s+entity)', str(item.get('message') or '')):
        return 'entity_resolution_failed', 'ttl_30d'
    if status in {'missing_token'}: return 'unknown_retryable', None
    if int(item.get('fetched') or 0) > 0: return 'available', None
    if item.get('total_available') == 0 or (item.get('scanned') == 0 and item.get('fetched') == 0): return 'no_comments', 'ttl_7d'
    return 'unknown_retryable', None

def write_source_capability_cache(errors: list[dict[str, Any]], skipped: list[dict[str, Any]], source_diagnostics: list[dict[str, Any]]) -> dict[str, Any]:
    now=dt.datetime.now(dt.timezone.utc)
    rows=[]
    for collection, kind in [(errors,'error'), (skipped,'skipped'), (source_diagnostics,'diagnostic')]:
        for item in collection:
            capability, ttl = derive_comments_capability(item)
            next_after=None
            if ttl == 'ttl_7d': next_after=(now+dt.timedelta(days=7)).isoformat()
            elif ttl == 'ttl_14d': next_after=(now+dt.timedelta(days=14)).isoformat()
            elif ttl == 'ttl_30d': next_after=(now+dt.timedelta(days=30)).isoformat()
            elif ttl == 'ttl_floodwait': next_after=(now+dt.timedelta(seconds=int(item.get('seconds') or 3600)+1800)).isoformat()
            rows.append({'platform_post_key': item.get('platform_post_key'), 'platform': item.get('platform'), 'kind': kind, 'last_status': item.get('status') or item.get('skip_reason') or ('ok' if capability == 'available' else None), 'last_error_code': item.get('code') or item.get('error_type'), 'comments_capability': capability, 'last_checked_at': now.isoformat(), 'next_check_after': next_after})
    payload={'schema_version':'event-comment-feedback-source-capability-cache-v1','generated_at':now.isoformat(),'sources':rows}
    (WORK/'source_capability_cache.json').write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    return payload

def classify_comment_type(text: str) -> tuple[str, str | None]:
    return classify_comment_role(text)

def record_filter_sample(summary: dict[str, Any], comment: dict[str, Any], comment_type: str, reason: str | None) -> None:
    bucket=summary.setdefault(comment_type, {'count':0,'reason':reason,'examples':[]})
    bucket['count']+=1
    if len(bucket['examples'])<8:
        bucket['examples'].append({'comment_key':comment.get('comment_key'),'platform':comment.get('platform'),'platform_post_key':comment.get('platform_post_key'),'reason':reason,'text_snippet':str(comment.get('text') or '')[:220]})

def write_comment_filter_summary(filter_summary: dict[str, Any]) -> dict[str, Any]:
    payload={'schema_version':'event-comment-feedback-comment-filter-summary-v1','generated_at':dt.datetime.now(dt.timezone.utc).isoformat(),'buckets':filter_summary}
    (WORK/'comment_filter_summary.json').write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    return payload

def phrase_min_sparse(p: dict[str, Any]) -> float:
    if p.get('singular_safe'): return 0.08
    if p.get('tone') == 'positive': return 0.12
    return 0.10

def phrase_margin_min(p: dict[str, Any]) -> float:
    if p.get('singular_safe') or p.get('risk') == 'low': return 0.03
    if p.get('tone') == 'positive': return 0.05
    if p.get('tone') == 'concern' or p.get('risk') in {'medium','high'}: return 0.07
    return 0.05

def public_candidate_gate(p: dict[str, Any], cand: dict[str, Any]) -> bool:
    if not cand.get('model_agreement'): return False
    if int(cand.get('e5_rank') or 999) > 2: return False
    if int(cand.get('bge_rank') or 999) > (3 if p.get('singular_safe') else 2): return False
    if float(cand.get('sparse_score') or 0.0) < phrase_min_sparse(p): return False
    if float(cand.get('positive_margin') or 0.0) < phrase_margin_min(p): return False
    return True

def parse_event_date(value: Any) -> dt.date | None:
    if not value: return None
    try:
        return dt.date.fromisoformat(str(value)[:10])
    except Exception:
        return None

def event_site_flags(event: dict[str, Any], generated_at: dt.datetime | None = None) -> dict[str, Any]:
    now = generated_at or dt.datetime.now(dt.timezone.utc)
    run_date = now.date()
    ed = parse_event_date(event.get('date') or event.get('start_date') or event.get('event_date'))
    is_past = bool(ed and ed < run_date)
    return {'run_date': run_date.isoformat(), 'run_datetime': now.isoformat(), 'is_past_event': is_past, 'eligible_for_site_export': not is_past}

def load_source_capability_cache_optional() -> dict[str, Any]:
    try:
        return load_json_any(find_input_file('source_capability_cache.json'))
    except FileNotFoundError:
        return {'schema_version':'event-comment-feedback-source-capability-cache-v1','sources':[]}

def load_previous_state_optional() -> dict[str, Any]:
    try:
        return load_json_any(find_input_file('event_comment_feedback_state.json'))
    except FileNotFoundError:
        return {'schema_version':'event-comment-feedback-state-v1','comments':{},'source_capabilities':{}}

def build_state_payload(comments: list[dict[str, Any]], source_capability_cache: dict[str, Any], previous_state: dict[str, Any] | None, state_mode: str) -> tuple[dict[str, Any], dict[str, Any]]:
    now=dt.datetime.now(dt.timezone.utc).isoformat()
    prev_comments=dict((previous_state or {}).get('comments') or {})
    before=len(prev_comments)
    reused=0; new_count=0
    for c in comments:
        key=str(c.get('comment_key') or '')
        if not key: continue
        event_ids=sorted({int(l['event_id']) for l in (c.get('links') or []) if l.get('event_id') is not None})
        text_hash=short_hash(low(c.get('text') or ''), 24)
        if key in prev_comments:
            reused += 1
            first_seen=prev_comments[key].get('first_seen_at') or now
        else:
            new_count += 1
            first_seen=now
        prev_comments[key]={'text_hash':text_hash,'event_ids':event_ids,'platform_post_key':c.get('platform_post_key'),'first_seen_at':first_seen,'last_seen_at':now}
    source_caps={}
    for row in (source_capability_cache or {}).get('sources') or []:
        k=row.get('platform_post_key')
        if not k: continue
        source_caps[str(k)]={'comments_capability':row.get('comments_capability') or 'unknown_retryable','last_checked_at':row.get('last_checked_at'),'next_check_after':row.get('next_check_after')}
    payload={'schema_version':'event-comment-feedback-state-v1','generated_at':now,'state_mode':state_mode,'comments':prev_comments,'source_capabilities':source_caps}
    (WORK/'event_comment_feedback_state.json').write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    stats={'state_mode':state_mode,'comments_known_before':before,'comments_known_after':len(prev_comments),'new_comments_this_run':new_count,'comments_reused_from_cache':reused,'source_capabilities_loaded':len((previous_state or {}).get('source_capabilities') or {}),'source_capabilities_updated':len(source_caps)}
    return payload, stats

def capability_skip_reason(row: dict[str, Any], now: dt.datetime | None = None) -> str | None:
    cap=str(row.get('comments_capability') or '')
    if cap not in {'forbidden_or_deleted','entity_resolution_failed','no_discussion_or_deleted','no_comments','rate_limited'}:
        return None
    next_after=row.get('next_check_after')
    if next_after:
        try:
            if dt.datetime.fromisoformat(str(next_after)) <= (now or dt.datetime.now(dt.timezone.utc)):
                return None
        except Exception:
            pass
    return f'capability_{cap}'

def build_fetch_posts(manifest: dict[str,Any], source_capability_cache: dict[str, Any] | None = None):
    events=manifest['events']; grouped={}
    cap_rows={r.get('platform_post_key'): r for r in ((source_capability_cache or {}).get('sources') or []) if r.get('platform_post_key')}
    for link in sorted(manifest['source_links'], key=lambda s:(-int(s.get('metric_comments') or 0), events.get(str(s['event_id']),{}).get('date') or '', int(s['event_id']), s['platform_post_key'])):
        parsed=link.get('parsed') or {}; platform=link.get('platform')
        if platform not in {'telegram','vk'}: continue
        key=link['platform_post_key']; item=grouped.setdefault(key, {'platform':platform,'platform_post_key':key,'parsed':parsed,'links':[],'metric_comments':0,'source_urls':[]})
        item['links'].append({'event_id':link['event_id'],'event_source_id':link.get('event_source_id'),'source_url':link.get('source_url')})
        item['metric_comments']=max(int(item['metric_comments']), int(link.get('metric_comments') or 0))
        if link.get('source_url') and link.get('source_url') not in item['source_urls']: item['source_urls'].append(link.get('source_url'))
    selected=[]; skipped=[]
    for p in sorted(grouped.values(), key=lambda p:(-int(p.get('metric_comments') or 0), p['platform_post_key'])):
        parsed=p.get('parsed') or {}
        cap_reason=capability_skip_reason(cap_rows.get(p['platform_post_key']) or {})
        if cap_reason:
            skipped.append({**p,'skip_reason':cap_reason})
        elif p['platform']=='telegram' and not (parsed.get('username') or parsed.get('chat_id')):
            skipped.append({**p,'skip_reason':'telegram_missing_username_or_chat_id'})
        elif p['platform']=='vk' and (parsed.get('owner_id') is None or parsed.get('post_id') is None):
            skipped.append({**p,'skip_reason':'vk_missing_owner_or_post_id'})
        else:
            selected.append(p)
    return selected, skipped

async def fetch_tg(posts, max_comments:int, sleep_s:float):
    from telethon import TelegramClient
    from telethon.sessions import StringSession
    from telethon.errors import FloodWaitError
    api_id=int(os.getenv('TG_API_ID') or os.getenv('TELEGRAM_API_ID') or '0'); api_hash=os.getenv('TG_API_HASH') or os.getenv('TELEGRAM_API_HASH') or ''
    session, kwargs=decode_bundle('TELEGRAM_AUTH_BUNDLE_DISCOVERY')
    client=TelegramClient(StringSession(session), api_id, api_hash, **kwargs)
    comments=[]; errors=[]; diagnostics=[]; account={}
    await client.connect()
    try:
        if not await client.is_user_authorized(): raise RuntimeError('TELEGRAM_AUTH_BUNDLE_DISCOVERY unauthorized')
        me=await client.get_me(); account={'id':int(me.id),'username':getattr(me,'username',None),'bundle':'TELEGRAM_AUTH_BUNDLE_DISCOVERY'}
        for idx,p in enumerate(posts,1):
            parsed=p['parsed']; username=parsed.get('username'); chat_id=parsed.get('chat_id'); mid=int(parsed.get('message_id') or 0); cap=source_comment_cap(p,max_comments)
            entity_ref = username or chat_id
            got=0; scanned=0
            try:
                entity=await client.get_entity(entity_ref)
                async for msg in client.iter_messages(entity, reply_to=mid, limit=cap):
                    scanned+=1
                    txt=normalize_text(getattr(msg,'message',None) or '')
                    if not txt or getattr(msg,'post',False) or getattr(msg,'fwd_from',None) or is_link_or_emoji_only(txt) or is_probable_source_copy(txt) or is_probable_private_ticket_request(txt): continue
                    sender_id=getattr(msg,'sender_id',None)
                    source_key = f'tg:{username}:{mid}' if username else f'tgid:{chat_id}:{mid}'
                    comments.append({'platform':'telegram','platform_post_key':p['platform_post_key'],'source_urls':p['source_urls'],'comment_key':f'{source_key}:{msg.id}','created_at':getattr(msg,'date',None).isoformat() if getattr(msg,'date',None) else None,'author_hash':short_hash(f'tg:{sender_id}' if sender_id else f'tgmsg:{msg.id}'),'text':txt,'links':p['links']}); got+=1
                diagnostics.append({'platform':'telegram','platform_post_key':p['platform_post_key'],'cap':cap,'fetched':got,'scanned':scanned,'metric_comments':p.get('metric_comments'),'entity_mode':'username' if username else 'chat_id'})
                if idx%25==0: emit('fetch_tg_progress', done=idx,total=len(posts),comments=len(comments),errors=len(errors), progress_label=f'TG источники {idx}/{len(posts)} · комментарии {len(comments)}')
                await async_human_pause(sleep_s, index=idx, platform='telegram')
            except FloodWaitError as exc:
                errors.append({'platform':'telegram','platform_post_key':p['platform_post_key'],'status':'flood_wait','seconds':int(exc.seconds),'cap':cap})
                if exc.seconds<=45: await asyncio.sleep(exc.seconds)
                else: break
            except Exception as exc:
                errors.append({'platform':'telegram','platform_post_key':p['platform_post_key'],'status':'error','error_type':type(exc).__name__,'message':str(exc)[:240],'cap':cap})
    finally: await client.disconnect()
    return comments, errors, diagnostics, account

def _vk_error_status(code: Any) -> str:
    try: code=int(code)
    except Exception: return 'api_error'
    if code in {15, 18, 30, 100}: return 'not_accessible_or_deleted'
    if code in {6, 9, 10, 29}: return 'retryable_api_error'
    return 'api_error'

def fetch_vk(posts, max_comments:int, sleep_s:float):
    import requests
    token=(os.getenv('VK_SERVICE_TOKEN') or os.getenv('VK_SERVICE_KEY') or os.getenv('VK_ACCESS_TOKEN') or '').strip(); comments=[]; errors=[]; diagnostics=[]
    if not token: return comments, [{'platform':'vk','status':'missing_token'}], diagnostics
    session=requests.Session()
    for idx,p in enumerate(posts,1):
        parsed=p['parsed']; owner=parsed.get('owner_id'); post=parsed.get('post_id'); cap=source_comment_cap(p,max_comments)
        fetched=0; scanned=0; total_available=None; offset=0; seen=set(); pages=0
        try:
            while fetched < cap:
                count=max(1,min(100,cap-fetched)); pages+=1
                params={'owner_id':int(owner),'post_id':int(post),'count':count,'offset':offset,'need_likes':0,'thread_items_count':10,'v':'5.199','access_token':token,'sort':'desc'}
                data=None
                for attempt in range(1,4):
                    try:
                        data=session.get('https://api.vk.com/method/wall.getComments', params=params, timeout=25).json(); break
                    except Exception as exc:
                        if attempt>=3: raise
                        time.sleep(1.5*attempt)
                if 'error' in data:
                    err=data['error']; code=err.get('error_code'); status=_vk_error_status(code)
                    errors.append({'platform':'vk','platform_post_key':p['platform_post_key'],'status':status,'code':code,'message':err.get('error_msg'),'cap':cap})
                    if status=='retryable_api_error' and pages==1:
                        time.sleep(2.5); continue
                    break
                resp=data.get('response') or {}; total_available=resp.get('count', total_available); items=resp.get('items') or []
                if not items: break
                for item in items:
                    scanned+=1
                    key=f"vk:{owner}:{post}:{item.get('id')}"
                    if key in seen: continue
                    seen.add(key)
                    txt=normalize_text(item.get('text') or '')
                    if txt and not is_link_or_emoji_only(txt) and not is_probable_source_copy(txt) and not is_probable_private_ticket_request(txt):
                        comments.append({'platform':'vk','platform_post_key':p['platform_post_key'],'source_urls':p['source_urls'],'comment_key':key,'created_at':dt.datetime.fromtimestamp(int(item.get('date') or 0), dt.timezone.utc).isoformat() if item.get('date') else None,'author_hash':short_hash(f"vk:{item.get('from_id')}"),'text':txt,'links':p['links']}); fetched+=1
                    thread=(item.get('thread') or {}).get('items') or []
                    for reply in thread:
                        rkey=f"vk:{owner}:{post}:{item.get('id')}:{reply.get('id')}"
                        if rkey in seen: continue
                        seen.add(rkey); rtxt=normalize_text(reply.get('text') or '')
                        if rtxt and not is_link_or_emoji_only(rtxt) and not is_probable_source_copy(rtxt) and not is_probable_private_ticket_request(rtxt):
                            comments.append({'platform':'vk','platform_post_key':p['platform_post_key'],'source_urls':p['source_urls'],'comment_key':rkey,'parent_comment_key':key,'created_at':dt.datetime.fromtimestamp(int(reply.get('date') or 0), dt.timezone.utc).isoformat() if reply.get('date') else None,'author_hash':short_hash(f"vk:{reply.get('from_id')}"),'text':rtxt,'links':p['links']}); fetched+=1
                    if fetched>=cap: break
                offset += len(items)
                if offset >= int(total_available or 0): break
                sync_human_pause(max(0.05, sleep_s/2), index=offset, platform='vk_page')
            diagnostics.append({'platform':'vk','platform_post_key':p['platform_post_key'],'cap':cap,'fetched':fetched,'scanned':scanned,'total_available':total_available,'pages':pages,'metric_comments':p.get('metric_comments')})
            if idx%50==0: emit('fetch_vk_progress', done=idx,total=len(posts),comments=len(comments),errors=len(errors), progress_label=f'VK источники {idx}/{len(posts)} · комментарии {len(comments)}')
            sync_human_pause(sleep_s, index=idx, platform='vk')
        except Exception as exc:
            errors.append({'platform':'vk','platform_post_key':p['platform_post_key'],'status':'error','error_type':type(exc).__name__,'message':str(exc)[:240],'cap':cap})
    return comments, errors, diagnostics

def quoted_values(line: str): return [normalize_text(v) for v in re.findall(r'[“"]([^”"]+)[”"]', line) if normalize_text(v)]
def parse_phrase_bank_markdown(path: Path):
    text=path.read_text(encoding='utf-8'); parts=re.split(r'(?m)^###\s+(\d+)\.\s+`([^`]+)`\s*$', text); out=[]
    for idx in range(1,len(parts),3):
        p={'num':int(parts[idx]),'id':parts[idx+1],'category':'','signal_type':'','tone':'internal','icon':'','risk':'internal','vector_only_allowed':False,'requires_llm_verification':True,'publishable':True,'public_sentence':None,'card_title':None,'card_text':None,'family':None,'singular_safe':False,'min_evidence_count':2,'min_unique_authors':2,'positive_prototypes':[],'hard_negatives':[]}
        for raw in parts[idx+2].splitlines():
            line=raw.strip()
            if line.startswith('- **Category:**'): p['category']=normalize_text(line.split(':**',1)[1])
            elif line.startswith('- **signal_type:**'):
                m=re.search(r'`([^`]+)`',line); p['signal_type']=m.group(1) if m else ''
            elif line.startswith('- **tone/icon/risk:**'):
                vals=re.findall(r'`([^`]+)`',line)
                if len(vals)>=3: p['tone'],p['icon'],p['risk']=vals[:3]
            elif line.startswith('- **Policy:**'): p['vector_only_allowed']='vector_only_allowed=true' in line; p['requires_llm_verification']='requires_llm_verification=true' in line
            elif line.startswith('- **public_sentence:**'):
                if '`null`' in line or re.search(r'\bnull\b',line,re.I): p['public_sentence']=None; p['publishable']=False
                else:
                    vals=quoted_values(line); p['public_sentence']=vals[0] if vals else None
            elif line.startswith('- **min evidence:**'):
                m1=re.search(r'min_evidence_count=(\d+)',line); m2=re.search(r'min_unique_authors=(\d+)',line)
                if m1: p['min_evidence_count']=int(m1.group(1))
                if m2: p['min_unique_authors']=int(m2.group(1))
            elif line.startswith('- **Positive prototypes:**'): p['positive_prototypes']=quoted_values(line)
            elif line.startswith('- **Hard negatives:**'): p['hard_negatives']=quoted_values(line)
        if not p['public_sentence']: p['publishable']=False
        p['card_title']=p.get('card_title') or p.get('public_sentence') or p['id']
        p['card_text']=p.get('card_text') or p.get('public_sentence') or p['card_title']
        p['family']=p.get('family') or p.get('category') or p['id']
        out.append(p)
    return out

def parse_phrase_bank(path: Path):
    if path.suffix == '.json':
        data=json.loads(path.read_text(encoding='utf-8'))
        phrases=data.get('phrases') if isinstance(data,dict) else data
        if not isinstance(phrases,list): raise RuntimeError('phrase bank JSON must contain phrases[]')
        for p in phrases:
            if p.get('public_sentence') in {'null','None',''}: p['public_sentence']=None
            p['publishable']=bool(p.get('publishable') and p.get('public_sentence'))
            p.setdefault('card_title', p.get('public_sentence') or p.get('id'))
            p.setdefault('card_text', p.get('public_sentence') or p.get('card_title'))
            p.setdefault('family', p.get('category') or p.get('id'))
            p.setdefault('positive_prototypes', [])
            p.setdefault('hard_negatives', [])
            p.setdefault('singular_safe', False)
        return phrases
    return parse_phrase_bank_markdown(path)

def phrase_docs(p):
    docs=[]
    for value in [p.get('card_title'), p.get('card_text'), p.get('public_sentence'), *(p.get('positive_prototypes') or [])]:
        v=normalize_text(value or '')
        if v and v not in docs: docs.append(v)
    return docs or [str(p.get('id'))]

def neg_docs(p): return [normalize_text(v) for v in (p.get('hard_negatives') or []) if normalize_text(v)]
def phrase_sparse_text(p): return ' | '.join([str(p.get('card_title') or ''), str(p.get('card_text') or ''), str(p.get('public_sentence') or ''), '; '.join(p.get('positive_prototypes') or [])])
def prefix(model,text,is_query): return (('query: ' if is_query else 'passage: ') + text) if 'multilingual-e5' in model.lower() else text

def sparse_counter(text):
    c=Counter();
    for w in re.findall(r'[a-zа-я0-9]+', low(text)):
        if len(w)>2:
            c[f'w:{w}']+=2; pad=f' {w} '
            for n in (3,4):
                for i in range(max(0,len(pad)-n+1)): c[f'c:{pad[i:i+n]}']+=1
        else: c[f'w:{w}']+=1
    return c
def cos_counter(a,b):
    if not a or not b: return 0.0
    if len(a)>len(b): a,b=b,a
    dot=sum(v*b.get(k,0) for k,v in a.items());
    if dot<=0: return 0.0
    return float(dot/(math.sqrt(sum(v*v for v in a.values()))*math.sqrt(sum(v*v for v in b.values()))))
QUESTION_IDS={'ticket_availability_question','time_questions','duration_questions','children_questions','location_questions','pushkin_card_questions','accessibility_questions','parking_questions','payment_questions','age_limit_questions','online_recording_questions','registration_interest','extra_places_question','extra_date_request','refund_exchange_questions','barcode_or_eticket_question','late_entry_question','admission_rules_question','food_drinks_question','security_rules_question','sector_or_seat_question','stroller_question','wheelchair_accessibility_question'}
RESALE_RE=re.compile(r'(?i)(\b(?:продам|продаю|куплю|ищу|нужен|нужны)\b[^\n]{0,50}\b(?:билет|проходк)|\b(?:билет|проходк)[^\n]{0,50}\b(?:продам|продаю|куплю|ищу|нужен|нужны)|есть\s+у\s+кого|у\s+кого(?:-то)?\s+есть|лишн\w*\s+билет|приму\s+в\s+дар[^\n]{0,40}\bбилет|дайте[^\n]{0,40}\bбилет|напишите\s+(?:пожалуйста\s+)?(?:в\s+)?(?:личк|лс))')
QUESTION_MARKER_RE=re.compile(r'(?i)(\?|подскаж\w*|уточните|можно\s+ли|есть\s+ли|будет\s+ли|остал\w*|где\b|когда\b|во\s+сколько|сколько\b|как\b|нужн\w*\s+ли|пустят)')
EXPLICIT_PROBLEM_RE=re.compile(r'(?i)(не\s+открывается|не\s+работает|не\s+могу\s+зарегистрироваться|не\s+получается\s+купить|ошибк|не\s+приш[её]л\s+билет|не\s+грузится|ссылка\s+не)')
def has_direct_question_marker(text: str) -> bool: return bool(QUESTION_MARKER_RE.search(low(text)))
def has_explicit_problem_report(text: str) -> bool: return bool(EXPLICIT_PROBLEM_RE.search(low(text)))
TICKET_TOPIC=[r'билет',r'билеты',r'места?',r'регистрац',r'попасть',r'проходк']
GUARDS={
    'children_questions':[r'дет',r'реб[её]н',r'школь',r'возраст'],
    'accessibility_questions':[r'пандус',r'маломобиль',r'инвалид',r'коляск',r'лифт'],
    'weather_concern':[r'дожд',r'погод',r'ливн',r'гр[оа]з',r'ветер'],
    'ticket_availability_question':TICKET_TOPIC,
    'ticket_interest_high':TICKET_TOPIC,
    'sold_out_discussion':[r'билет',r'мест',r'регистрац',r'законч',r'разобрал',r'нет\s+мест',r'нет\s+билет'],
    'sold_out_disappointment':[r'билет',r'мест',r'регистрац',r'законч',r'разобрал',r'не\s+успел',r'нет\s+мест',r'нет\s+билет'],
    'high_demand_from_ticket_friction':[r'билет',r'мест',r'регистрац',r'попасть',r'тираж',r'успева',r'очеред'],
    'ticket_purchase_technical_problem':[r'ссылк',r'куп',r'оплат',r'не\s+работ',r'ошибк',r'vpn',r'сайт',r'не\s+получ'],
    'ticket_link_broken':[r'ссылк',r'не\s+откры',r'не\s+работ',r'ошибк',r'сайт',r'не\s+груз'],
    'registration_closed':[r'регистрац',r'мест',r'билет',r'законч',r'закрыт'],
    'barcode_or_eticket_question':[r'штрих',r'qr',r'электронн\w*\s+билет',r'номер\s+билет',r'почт',r'не\s+приш[её]л\s+билет'],
    'late_entry_question':[r'после\s+начала',r'опозда',r'прийти\s+позже',r'пустят'],
    'admission_rules_question':[r'вход',r'проход',r'пустят',r'паспорт',r'правил',r'можно\s+взять|можно\s+с'],
    'parking_questions':[r'парковк',r'где\s+парков',r'оставить\s+машин',r'как\s+добраться',r'общественн\w*\s+транспорт',r'автобус'],
    'location_questions':[r'адрес',r'где\s+это',r'как\s+добраться',r'откуда\s+вход',r'как\s+пройти'],
    'time_questions':[r'во сколько',r'время',r'начал',r'вход',r'длит'],
    'visual_quality_positive':[r'постановк',r'спектакл',r'сцен',r'мюзикл',r'выставк',r'представлен',r'концерт',r'атмосфер'],
    'recommendation_from_past_visit':[r'посетил',r'сходил',r'понрав',r'рекоменд',r'совет',r'стоит\s+сход'],
    'performance_praised':[r'постановк',r'исполнен',r'артист',r'спектакл',r'игра',r'сцен'],
    'stroller_question':[r'детск.*коляск|коляск.*детск|с\s+коляск'],
    'wheelchair_accessibility_question':[r'инвалид|маломобиль|пандус|лифт|коляск.*инвалид'],
    'food_drinks_question':[r'вод|напит|ед|снек|перекус|трибун|пронос'],
    'food_at_venue_question':[r'снек|ед|перекус|трибун|купить\s+.*(?:ед|снек|напит)'],
    'security_rules_question':[r'досмотр|нельзя|сумк|безопас|пронос'],
    'sector_or_seat_question':[r'сектор|мест|ряд|танцпол|сидяч'],
    'crowd_management_concern':[r'толп|очеред|много\s+люд|вход'],
}
def guard(text,pid):
    t=low(text); reasons=[]
    if pid in QUESTION_IDS:
        if RESALE_RE.search(t): reasons.append('ticket_resale_or_private_ticket_request')
        elif not (has_direct_question_marker(t) or has_explicit_problem_report(t)):
            reasons.append('question_phrase_without_direct_question_or_problem')
    if pid in {'sold_out_discussion','sold_out_disappointment'} and not any(re.search(p,t,re.I) for p in TICKET_TOPIC):
        reasons.append('sold_out_without_ticket_topic')
    if pid in {'accessibility_questions','accessibility_concern'} and any(re.search(p,t,re.I) for p in [r'ссылк',r'куп',r'билет',r'оплат',r'vpn']):
        reasons.append('accessibility_confused_with_ticket_purchase')
    pats=GUARDS.get(pid)
    if pats and not any(re.search(p,t,re.I) for p in pats): reasons.append('phrase_lexical_guard')
    return reasons

def encode_with_loaded_model(model_name, model, texts, is_query):
    batch=32 if 'multilingual-e5' in model_name.lower() else 8
    emit('model_encode_start', model=model_name, count=len(texts), is_query=is_query, batch=batch, progress_label=f'embeddings {model_name}: {len(texts)} текстов')
    arr=model.encode([prefix(model_name,t,is_query) for t in texts], batch_size=batch, normalize_embeddings=True, convert_to_numpy=True, show_progress_bar=True)
    emit('model_encode_done', model=model_name, count=len(texts), is_query=is_query, progress_label=f'embeddings готовы {model_name}')
    return arr

def model_phrase_scores(comment_vec, proto_vecs, proto_meta, neg_vecs, neg_meta):
    phrase_scores=defaultdict(lambda:{'score':-9.0,'prototype':None,'negative_score':0.0})
    vals=comment_vec @ proto_vecs.T
    for i,score_v in enumerate(vals):
        pid, proto = proto_meta[i]
        s=float(score_v)
        if s > phrase_scores[pid]['score']:
            phrase_scores[pid]={'score':s,'prototype':proto,'negative_score':0.0}
    if len(neg_meta):
        nvals=comment_vec @ neg_vecs.T
        for i,score_v in enumerate(nvals):
            pid, _neg = neg_meta[i]
            phrase_scores[pid]['negative_score']=max(float(score_v), float(phrase_scores[pid].get('negative_score') or 0.0))
    ranked=sorted([(v['score'], pid, v) for pid,v in phrase_scores.items()], reverse=True)
    ranks={pid:rank for rank,(_s,pid,_v) in enumerate(ranked,1)}
    return phrase_scores, ranked, ranks

def score(comments, manifest, phrases):
    runtime=[p for p in phrases if p.get('runtime_enabled', True)]
    unique={}; event_comments=defaultdict(list); filter_summary={}; context_by_event_source=defaultdict(list)
    source_post_event_link_count={c.get('platform_post_key'): len({int(l['event_id']) for l in (c.get('links') or []) if l.get('event_id') is not None}) for c in comments}
    for c in comments:
        ctype, reason = classify_comment_type(str(c.get('text') or ''))
        role = 'user_feedback' if ctype == 'user_feedback' else ctype
        if ctype != 'user_feedback':
            record_filter_sample(filter_summary, c, ctype, reason)
            if ctype == 'official_reply':
                for link in c.get('links') or []:
                    context_by_event_source[(int(link['event_id']), c.get('platform_post_key'))].append({**c, 'comment_role': role, 'filter_reason': reason})
            continue
        for link in c.get('links') or []:
            item=dict(c); item['event_id']=int(link['event_id']); item['comment_type']=ctype; item['comment_role']=role; item['filter_reason']=reason; item['parent_text_available']=False; item['thread_context_used']=False; item['source_post_event_link_count']=source_post_event_link_count.get(c.get('platform_post_key'), 1); event_comments[item['event_id']].append(item)
            th=short_hash(low(c['text']),24); unique.setdefault(th, {'text_hash':th,'text':c['text']})
    filter_payload=write_comment_filter_summary(filter_summary)
    ulist=list(unique.values()); texts=[u['text'] for u in ulist]
    if not texts: return {}, [], filter_payload
    proto_meta=[]; proto_docs=[]; neg_meta=[]; neg_texts=[]
    for p in runtime:
        for doc in phrase_docs(p):
            proto_meta.append((p['id'], doc)); proto_docs.append(doc)
        for doc in neg_docs(p):
            neg_meta.append((p['id'], doc)); neg_texts.append(doc)
    model_scores={}
    from sentence_transformers import SentenceTransformer
    for model_name in REQUIRED_MODELS:
        emit('model_load_start', model=model_name, progress_percent=45, progress_label=f'загрузка модели {model_name}')
        model=SentenceTransformer(model_name, device=os.getenv('ACQ_COMMENT_RETRIEVAL_DEVICE') or None)
        try: model.max_seq_length=128
        except Exception: pass
        pos_vec=encode_with_loaded_model(model_name, model, proto_docs, True)
        neg_vec=encode_with_loaded_model(model_name, model, neg_texts, True) if neg_texts else []
        cvec=encode_with_loaded_model(model_name, model, texts, False)
        rows={}
        for ci,u in enumerate(ulist):
            ps, ranked, ranks = model_phrase_scores(cvec[ci], pos_vec, proto_meta, neg_vec, neg_meta)
            rows[u['text_hash']]={'phrase_scores':ps,'ranked':ranked[:12],'ranks':ranks,'top_phrase_id':ranked[0][1] if ranked else None,'top_score':float(ranked[0][0]) if ranked else 0.0,'top5':{pid:rank for rank,(_s,pid,_v) in enumerate(ranked[:5],1)},'scores12':{pid:s for s,pid,_v in ranked[:12]}}
        model_scores[model_name]=rows
        del model
    by_phrase={p['id']:p for p in runtime}; sparse={p['id']:sparse_counter(phrase_sparse_text(p)) for p in runtime}
    event_results={}; evidence=[]
    for eid, rows in event_comments.items():
        seen=set(); groups=defaultdict(lambda:{'phrase':None,'comments':[],'authors':set(),'sources':set(),'scores':[]}); suppressed_internal=[]
        for c in rows:
            th=short_hash(low(c['text']),24); ah=c.get('author_hash') or ''; k=(th,ah)
            if k in seen: continue
            seen.add(k)
            e5=model_scores[REQUIRED_MODELS[0]].get(th); bge=model_scores[REQUIRED_MODELS[1]].get(th)
            if not e5 or not bge: continue
            candidate_ids=set(pid for _s,pid,_v in e5['ranked'][:8]) | set(pid for _s,pid,_v in bge['ranked'][:8])
            best=None
            for pid in candidate_ids:
                p=by_phrase.get(pid)
                if not p: continue
                e5_rank=e5['ranks'].get(pid,999); bge_rank=bge['ranks'].get(pid,999)
                e5_row=e5['phrase_scores'].get(pid,{}); bge_row=bge['phrase_scores'].get(pid,{})
                e5_score=float(e5_row.get('score') or 0.0); bge_score=float(bge_row.get('score') or 0.0)
                neg=max(float(e5_row.get('negative_score') or 0.0), float(bge_row.get('negative_score') or 0.0))
                ss=cos_counter(sparse_counter(c['text']), sparse.get(pid, Counter()))
                reasons=guard(c['text'],pid)
                if ss < 0.006: reasons.append('sparse_support_low')
                positive_margin=max(e5_score,bge_score)-neg
                if positive_margin < -0.01: reasons.append('negative_margin_low')
                if not ((e5_rank<=5 and bge_rank<=5) or (e5_rank==1 and bge_rank<=10) or (bge_rank==1 and e5_rank<=10)):
                    reasons.append('dual_rank_weak')
                ensemble=0.40*e5_score+0.40*bge_score+0.10*(1.0/max(e5_rank,1)+1.0/max(bge_rank,1))+0.10*ss-0.15*max(0.0,neg-max(e5_score,bge_score))
                cand={'phrase_id':pid,'e5_score':round(e5_score,4),'bge_score':round(bge_score,4),'e5_rank':e5_rank,'bge_rank':bge_rank,'sparse_score':round(ss,4),'negative_score':round(neg,4),'positive_margin':round(positive_margin,4),'positive_margin_min':phrase_margin_min(p),'ensemble_score':round(ensemble,4),'e5_prototype':e5_row.get('prototype'),'bge_prototype':bge_row.get('prototype'),'model_agreement':e5.get('top_phrase_id')==bge.get('top_phrase_id')}
                cand['public_gate_pass']=public_candidate_gate(p, cand)
                if reasons: continue
                if best is None or ensemble > best[0]: best=(ensemble,pid,p,cand)
            if best is None: continue
            _ensemble,pid,p,cand=best
            if not p.get('publishable') or p.get('risk')=='internal' or p.get('tone')=='internal':
                suppressed_internal.append({'phrase_id':pid,'comment_key':c.get('comment_key'),'text_snippet':c['text'][:220],**cand})
                continue
            g=groups[pid]; g['phrase']=p; item=dict(c); item['candidate']=cand; item['guard_reasons']=';'.join(reasons); item['is_user_evidence']=True; item['is_official_context']=False; g['comments'].append(item); g['authors'].add(ah); g['sources'].add(c.get('platform_post_key')); g['scores'].append(cand)
        accepted=[]; other=[]
        for pid,g in groups.items():
            p=g['phrase']; ev=len(g['comments']); au=len(g['authors']); min_ev=int(p.get('min_evidence_count') or 2); min_au=int(p.get('min_unique_authors') or 2)
            official_context=[ctx for src in g['sources'] for ctx in context_by_event_source.get((eid, src), [])]
            public_comments=[c for c in g['comments'] if c.get('candidate',{}).get('public_gate_pass') and c.get('comment_role') == 'user_feedback']
            public_authors={c.get('author_hash') or '' for c in public_comments if c.get('comment_role') == 'user_feedback'}
            pev=len(public_comments); pau=len(public_authors)
            if pev>=min_ev and pau>=min_au and p.get('vector_only_allowed') and not p.get('requires_llm_verification') and p.get('risk')=='low':
                status='semantic_public_ready_dual_kaggle'; semantic_status='semantic_public_gate_pass'
            elif p.get('singular_safe') and pev>=1 and pau>=1 and p.get('vector_only_allowed') and not p.get('requires_llm_verification') and p.get('risk') in {'low','medium'}:
                status='semantic_public_ready_singular_dual_kaggle'; semantic_status='semantic_public_gate_pass'
            elif ev>=min_ev and au>=min_au:
                status='needs_review_dual_kaggle'; semantic_status='semantic_needs_review'
            else:
                status='suppressed_weak_dual_kaggle'; semantic_status='semantic_suppressed_weak'
            all_cs=sorted(g['comments'], key=lambda x:x['candidate']['ensemble_score'], reverse=True)
            public_status=semantic_status == 'semantic_public_gate_pass'
            cs=sorted(public_comments, key=lambda x:x['candidate']['ensemble_score'], reverse=True) if public_status else all_cs
            display_ev=pev if public_status else ev
            display_au=pau if public_status else au
            feedback_scope='event_series' if max([int(c.get('source_post_event_link_count') or 1) for c in g['comments']] or [1]) > 1 else 'event_instance'
            rec={'phrase_id':pid,'tone':p.get('tone'),'risk_class':p.get('risk'),'public_sentence':p.get('public_sentence'),'card_title':p.get('card_title'),'card_text':p.get('card_text'),'singular_safe':bool(p.get('singular_safe')),'public_gate_evidence_count':pev,'public_gate_unique_authors_count':pau,'user_evidence_count':pev if public_status else ev,'user_unique_authors_count':pau if public_status else au,'official_context_count':len(official_context),'evidence_count':display_ev,'unique_authors_count':display_au,'raw_group_evidence_count':ev+len(official_context),'raw_group_unique_authors_count':au,'sources_count':len(g['sources']),'feedback_scope':feedback_scope,'source_post_event_link_count':max([int(c.get('source_post_event_link_count') or 1) for c in g['comments']] or [1]),'status':status,'semantic_status':semantic_status,'site_export_status':None,'evidence_snippets':[c['text'][:220] for c in cs[:4]],'score_samples':[{**c['candidate'],'comment_key':c['comment_key'],'comment_role':c.get('comment_role'),'filter_reason':c.get('filter_reason'),'guard_reasons':c.get('guard_reasons'),'is_user_evidence':c.get('is_user_evidence',False),'is_official_context':False,'parent_comment_key':c.get('parent_comment_key'),'parent_text_available':c.get('parent_text_available',False),'thread_context_used':c.get('thread_context_used',False),'source_post_event_link_count':c.get('source_post_event_link_count'),'feedback_scope':feedback_scope,'shared_source_evidence':feedback_scope=='event_series','risk_class':p.get('risk'),'tone':p.get('tone'),'text_snippet':c['text'][:220]} for c in cs[:4]]}
            for c in cs[:4]: evidence.append({'event_id':eid,'phrase_id':pid,'status':status,'semantic_status':semantic_status,'site_export_status':None,**c['candidate'],'comment_key':c['comment_key'],'comment_type':c.get('comment_type'),'comment_role':c.get('comment_role'),'filter_reason':c.get('filter_reason'),'guard_reasons':c.get('guard_reasons'),'is_user_evidence':c.get('is_user_evidence',False),'is_official_context':False,'parent_comment_key':c.get('parent_comment_key'),'parent_text_available':c.get('parent_text_available',False),'thread_context_used':c.get('thread_context_used',False),'source_post_event_link_count':c.get('source_post_event_link_count'),'feedback_scope':feedback_scope,'shared_source_evidence':feedback_scope=='event_series','risk_class':p.get('risk'),'tone':p.get('tone'),'text_snippet':c['text'][:220]})
            (accepted if semantic_status == 'semantic_public_gate_pass' else other).append(rec)
        event_results[str(eid)]={'event':manifest['events'].get(str(eid),{'id':eid}),'comments_seen_count':len(rows),'dedup_comments_count':len(seen),'accepted_items':accepted,'review_or_suppressed_items':other,'official_context_count':sum(len(v) for (event_id,_src),v in context_by_event_source.items() if event_id==eid),'suppressed_internal_samples':suppressed_internal[:20]}
    return event_results,evidence,filter_payload

def site_export_status_for(semantic_status: str, flags: dict[str, Any]) -> tuple[str, str]:
    if semantic_status == 'semantic_public_gate_pass':
        if flags.get('eligible_for_site_export') and not flags.get('is_past_event'):
            return 'site_public_ready', 'site_public_ready_dual_kaggle'
        if flags.get('is_past_event'):
            return 'site_ineligible_past_event', 'site_ineligible_past_gate_pass'
        return 'site_ineligible_not_site_event', 'site_ineligible_not_site_event'
    if semantic_status == 'semantic_needs_review': return 'site_ineligible_review_required', 'needs_review_dual_kaggle'
    if semantic_status == 'semantic_filtered_internal': return 'site_ineligible_weak', 'filtered_internal'
    return 'site_ineligible_weak', 'suppressed_weak_dual_kaggle'

def write_reports(manifest, comments, errors, event_results, evidence, fetch_error_summary=None, comment_filter_summary=None, source_capability_cache=None, state_stats=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    generated_at=dt.datetime.now(dt.timezone.utc)
    for eid,res in event_results.items():
        flags=event_site_flags(res.get('event') or {}, generated_at)
        res.update(flags)
        for it in (res.get('accepted_items') or []) + (res.get('review_or_suppressed_items') or []):
            it['is_site_eligible_event']=flags['eligible_for_site_export']
            it['run_date']=flags['run_date']; it['run_datetime']=flags['run_datetime']; it['is_past_event']=flags['is_past_event']
            site_status, public_status = site_export_status_for(str(it.get('semantic_status') or ''), flags)
            it['site_export_status']=site_status; it['status']=public_status
    for e in evidence:
        flags=event_site_flags((manifest.get('events') or {}).get(str(e.get('event_id')), {}), generated_at)
        site_status, public_status = site_export_status_for(str(e.get('semantic_status') or ''), flags)
        e.update({'run_date':flags['run_date'],'run_datetime':flags['run_datetime'],'is_past_event':flags['is_past_event'],'eligible_for_site_export':flags['eligible_for_site_export'],'is_site_eligible_event':flags['eligible_for_site_export'],'site_export_status':site_status,'status':public_status})
    filtered_total=sum(int(v.get('count') or 0) for v in ((comment_filter_summary or {}).get('buckets') or {}).values()) if isinstance(comment_filter_summary, dict) else 0
    candidate_rows_total=len(evidence)
    semantic_public_gate_pass_rows_total=sum(1 for e in evidence if e.get('public_gate_pass'))
    semantic_public_ready=[e for e in evidence if e.get('semantic_status') == 'semantic_public_gate_pass']
    site_public_ready=[e for e in evidence if e.get('site_export_status') == 'site_public_ready' and e.get('is_user_evidence') and not e.get('is_official_context') and e.get('model_agreement')]
    non_public_gate_pass=[e for e in evidence if e.get('public_gate_pass') and e.get('site_export_status') != 'site_public_ready']
    past_public_gate=[e for e in evidence if e.get('public_gate_pass') and e.get('is_past_event')]
    low_risk_suppressed=[e for e in evidence if e.get('eligible_for_site_export') and not e.get('is_past_event') and e.get('model_agreement') and int(e.get('e5_rank') or 999)==1 and int(e.get('bge_rank') or 999)==1 and not e.get('is_official_context') and e.get('risk_class')=='low' and e.get('site_export_status')!='site_public_ready']
    official_context_rows=sum(int(r.get('official_context_count') or 0) for r in event_results.values())
    user_evidence_rows=sum(1 for e in evidence if e.get('is_user_evidence'))
    state_stats=state_stats or {'state_mode':'one_off_non_cumulative','comments_known_before':0,'comments_known_after':len(comments),'new_comments_this_run':len(comments),'comments_reused_from_cache':0,'source_capabilities_loaded':0,'source_capabilities_updated':len((source_capability_cache or {}).get('sources') or [])}
    summary={'schema_version':RUN_SCHEMA_VERSION,'generated_at':generated_at.isoformat(),'embedding_api_allowed':False,'provider_calls_total_this_process':0,'read_mode':READ_MODE,'models':REQUIRED_MODELS,'events_in_manifest':manifest.get('event_count'),'source_links':manifest.get('source_link_count'),'source_posts':manifest.get('source_post_count'),'comments_fetched':len(comments),'comments_fetched_this_run':len(comments),'comments_known_total':state_stats.get('comments_known_after'),'new_comments_this_run':state_stats.get('new_comments_this_run'),'comments_reused_from_cache':state_stats.get('comments_reused_from_cache'),'source_posts_known_total':manifest.get('source_post_count'),'source_posts_checked_this_run':manifest.get('source_post_count'),'source_posts_skipped_by_capability':sum(1 for r in ((source_capability_cache or {}).get('sources') or []) if r.get('kind')=='skipped' and str(r.get('last_status') or '').startswith('capability_')),'events_with_comments':len(event_results),'fetch_errors':len(errors),'fetch_error_buckets':len((fetch_error_summary or {}).get('buckets') or []),'events_with_public_ready':len({e.get('event_id') for e in site_public_ready}),'events_with_review_candidates':sum(1 for r in event_results.values() if any(i['site_export_status']=='site_ineligible_review_required' for i in r.get('review_or_suppressed_items',[]))),'candidate_rows_total':candidate_rows_total,'candidate_rows_public_gate_pass':semantic_public_gate_pass_rows_total,'public_ready_evidence_rows':len(site_public_ready),'public_ready_events':len({e.get('event_id') for e in site_public_ready}),'semantic_public_gate_pass_rows_total':semantic_public_gate_pass_rows_total,'semantic_public_ready_rows_all_events':len(semantic_public_ready),'site_export_public_ready_rows':len(site_public_ready),'site_export_public_ready_events':len({e.get('event_id') for e in site_public_ready}),'past_public_gate_pass_rows':len(past_public_gate),'site_ineligible_public_gate_pass_rows':len(non_public_gate_pass),'non_public_rows_with_public_gate_pass':len(non_public_gate_pass),'low_risk_site_eligible_suppressed_rows':len(low_risk_suppressed),'filtered_comments_total':filtered_total,'scored_comments_total':len(comments)-filtered_total,'official_context_rows':official_context_rows,'user_evidence_rows':user_evidence_rows,'source_capability_rows':len((source_capability_cache or {}).get('sources') or []),**state_stats}
    (WORK/'event_comment_feedback_summary.json').write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    (WORK/'event_comment_feedback_probe.json').write_text(json.dumps({**summary,'events':event_results}, ensure_ascii=False, indent=2), encoding='utf-8')
    rows=[]
    for eid,res in sorted(event_results.items(), key=lambda kv:((kv[1]['event'].get('date') or ''), int(kv[0]))):
        ev=res['event']; items=res['accepted_items']+res['review_or_suppressed_items']
        def join(site_status,tone=None):
            vals=[]
            for it in items:
                if it.get('site_export_status') != site_status: continue
                if tone and it.get('tone')!=tone: continue
                vals.append(f"{it.get('card_title') or it.get('public_sentence')} — {it.get('card_text') or ''} ({it['evidence_count']}/{it['unique_authors_count']}; {it.get('feedback_scope','event_instance')})")
            return '\n'.join(vals[:6])
        rows.append({'event_id':int(eid),'date':ev.get('date'),'time':ev.get('time'),'run_date':res.get('run_date'),'run_datetime':res.get('run_datetime'),'is_past_event':res.get('is_past_event'),'eligible_for_site_export':res.get('eligible_for_site_export'),'title':ev.get('title'),'venue':ev.get('location_name'),'comments_seen':res['comments_seen_count'],'dedup_comments':res['dedup_comments_count'],'official_context_count':res.get('official_context_count',0),'public_positive':join('site_public_ready','positive'),'public_neutral':join('site_public_ready','neutral'),'public_negative':join('site_public_ready','concern'),'review_positive':join('site_ineligible_review_required','positive'),'review_neutral':join('site_ineligible_review_required','neutral'),'review_negative':join('site_ineligible_review_required','concern'),'suppressed_weak_top':join('site_ineligible_weak'),'evidence_snippets':'\n---\n'.join([sn for it in items[:6] for sn in it.get('evidence_snippets',[])][:6]),'note':'Kaggle dual local prototype ensemble: intfloat/multilingual-e5-base + BAAI/bge-m3; embedding API disabled; status is site-aware; official replies are context, not public evidence'})
    headers=list(rows[0].keys()) if rows else ['note']
    with (WORK/'event_comment_feedback_full_table.csv').open('w',encoding='utf-8-sig',newline='') as f:
        w=csv.DictWriter(f,fieldnames=headers); w.writeheader(); w.writerows(rows or [{'note':'No rows'}])
    wb=Workbook(); ws=wb.active; ws.title='summary'
    for k,v in summary.items(): ws.append([k,json.dumps(v,ensure_ascii=False) if isinstance(v,(list,dict,bool)) else v])
    ws.column_dimensions['A'].width=42; ws.column_dimensions['B'].width=95
    evs=wb.create_sheet('events'); evs.append(headers)
    for cell in evs[1]: cell.font=Font(bold=True); cell.fill=PatternFill('solid', fgColor='D9EAD3')
    for r in rows: evs.append([r.get(h,'') for h in headers])
    evs.freeze_panes='A2'
    for col,w in {'A':10,'B':12,'C':12,'D':28,'H':46,'I':30,'M':52,'N':52,'O':52,'P':52,'Q':52,'R':52,'S':52,'T':70,'U':60}.items(): evs.column_dimensions[col].width=w
    def excel_value(value):
        if isinstance(value, (dict, list, tuple, bool)):
            return json.dumps(value, ensure_ascii=False)
        return value
    def add_sheet(name, rows_list):
        sh=wb.create_sheet(name); hh=sorted({k for row in rows_list for k in row.keys()}) if rows_list else ['note']; sh.append(hh)
        for row in rows_list or [{'note':'No rows'}]: sh.append([excel_value(row.get(h,'')) for h in hh])
        for cell in sh[1]: cell.font=Font(bold=True); cell.fill=PatternFill('solid', fgColor='D9EAD3')
        sh.freeze_panes='A2'
        return sh
    add_sheet('evidence_samples', evidence[:2000])
    add_sheet('public_ready_evidence_only', site_public_ready[:2000])
    add_sheet('public_gate_pass_not_published', non_public_gate_pass[:2000])
    add_sheet('low_risk_site_eligible_suppressed', low_risk_suppressed[:2000])
    if fetch_error_summary:
        add_sheet('fetch_error_buckets', (fetch_error_summary.get('buckets') or [])[:1000])
    if comment_filter_summary:
        filter_rows=[]
        for k,v in (comment_filter_summary.get('buckets') or {}).items():
            filter_rows.append({'comment_type':k,'count':v.get('count'),'reason':v.get('reason'),'examples_json':json.dumps(v.get('examples') or [],ensure_ascii=False)})
        add_sheet('comment_filter_summary', filter_rows)
    if source_capability_cache:
        add_sheet('source_capability_cache', (source_capability_cache.get('sources') or [])[:2000])
    for sh in wb.worksheets:
        for row in sh.iter_rows():
            for cell in row: cell.alignment=Alignment(wrap_text=True, vertical='top')
    wb.save(WORK/'event_comment_feedback_full_table.xlsx')
    emit('report_written', **summary, progress_percent=100, progress_label='таблица готова')

def main():
    init_status()
    try:
        emit('kernel_started', schema=RUN_SCHEMA_VERSION, progress_percent=1, progress_label='запуск Kaggle')
        acquire_status_resources()
        ensure_deps(); load_secrets()
        config_path=find_input_file('run_config.json')
        manifest_path=find_input_file('prod_source_manifest_full.json.gz')

        try:
            phrase_path=find_input_file('phrase-bank-v1.json')
        except FileNotFoundError:
            phrase_path=find_input_file('phrase-bank-v1.md')
        emit('input_files_found', run_config=str(config_path), manifest=str(manifest_path), phrase_bank=str(phrase_path), progress_percent=4, progress_label='payload найден')
        config=load_json_any(config_path); manifest=load_json_any(manifest_path); phrases=parse_phrase_bank(phrase_path); prior_source_capability_cache=load_source_capability_cache_optional(); previous_state=load_previous_state_optional(); state_mode=str(config.get('state_mode') or 'one_off_non_cumulative')
        max_comments=int(config.get('max_comments_per_source') or 60); sleep_s=float(config.get('request_sleep') or 0.18)
        selected, skipped=build_fetch_posts(manifest, prior_source_capability_cache); tg=[p for p in selected if p['platform']=='telegram']; vk=[p for p in selected if p['platform']=='vk']
        emit('preflight_ok', events=manifest.get('event_count'), source_posts=manifest.get('source_post_count'), selected=len(selected), tg=len(tg), vk=len(vk), skipped=len(skipped), models=REQUIRED_MODELS, embedding_api_allowed=False, read_mode=READ_MODE, request_sleep=sleep_s, progress_percent=8, progress_label=f'источники: TG {len(tg)} / VK {len(vk)}')
        tg_comments,tg_errors,tg_diag,tg_account=asyncio.run(fetch_tg(tg,max_comments,sleep_s)) if tg else ([],[],[],{})
        vk_comments,vk_errors,vk_diag=fetch_vk(vk,max_comments,sleep_s) if vk else ([],[],[])
        comments=tg_comments+vk_comments; errors=tg_errors+vk_errors; source_diagnostics=tg_diag+vk_diag
        fetch_summary=write_fetch_error_summary(errors, skipped, source_diagnostics)
        source_capability_cache=write_source_capability_cache(errors, skipped, source_diagnostics)
        _state_payload,state_stats=build_state_payload(comments, source_capability_cache, previous_state, state_mode)
        (WORK/'comments_raw_redacted.json').write_text(json.dumps({'comments':comments,'errors':errors,'source_diagnostics':source_diagnostics,'telegram_account':tg_account}, ensure_ascii=False, indent=2), encoding='utf-8')
        emit('fetch_done', comments=len(comments), errors=len(errors), tg_comments=len(tg_comments), vk_comments=len(vk_comments), progress_percent=40, progress_label=f'комментарии: {len(comments)}')
        event_results,evidence,filter_summary=score(comments, manifest, phrases) if comments else ({}, [], {'buckets': {}})
        write_reports(manifest, comments, errors, event_results, evidence, fetch_summary, filter_summary, source_capability_cache, state_stats)
        emit('kernel_done', progress_percent=100, progress_label='готово')
    except Exception as exc:
        emit('kernel_failed', error_type=type(exc).__name__, message=str(exc)[:1000], progress_label='ошибка')
        # Mark terminal state in the shared status ledger: terminal events are report_written/render_done.
        _status_event('report_written', payload={'phase': 'failed', 'progress_label': 'ошибка'}, status='failed', message=f'{type(exc).__name__}: {exc}')
        raise
    finally:
        finish_status()
if __name__=='__main__': main()
